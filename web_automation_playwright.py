from playwright.sync_api import sync_playwright, Page, Browser
from excel_state_manager import ExcelStateManager
try:
    from playwright.async_api import async_playwright
    # AsyncPage และ AsyncBrowser คือ type aliases ของ Page และ Browser ใน async context
    # ไม่ต้อง import โดยตรง
    ASYNC_PLAYWRIGHT_AVAILABLE = True
except ImportError:
    ASYNC_PLAYWRIGHT_AVAILABLE = False
    async_playwright = None
import time
import os
import re
import json
import asyncio
from typing import Dict, Optional, List, Callable
from pathlib import Path
from config import Config
from report_manager import get_global_report_manager, line_notify, line_oa_push
from file_manager import FileManager
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    load_workbook = None

class AsyncPageWrapper:
    """Wrapper class สำหรับ AsyncPage เพื่อให้สามารถใช้แบบ sync ได้"""
    def __init__(self, async_page, loop):
        self._async_page = async_page
        self._loop = loop
    
    def _run_async(self, coro):
        """Helper method สำหรับรัน async function ใน sync context"""
        try:
            loop = self._loop or asyncio.get_event_loop()
            if loop.is_running():
                import concurrent.futures
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    future = executor.submit(asyncio.run, coro)
                    return future.result(timeout=30)
            else:
                return loop.run_until_complete(coro)
        except RuntimeError:
            return asyncio.run(coro)
        except Exception:
            return asyncio.run(coro)
    
    def locator(self, selector):
        """Wrapper สำหรับ locator"""
        # locator() ใน Async API ไม่ใช่ async method แต่ return Locator object โดยตรง
        loc = self._async_page.locator(selector)
        return AsyncLocatorWrapper(loc, self._loop)
    
    def goto(self, url, **kwargs):
        """Wrapper สำหรับ goto"""
        return self._run_async(self._async_page.goto(url, **kwargs))
    
    def fill(self, selector, value):
        """Wrapper สำหรับ fill"""
        return self._run_async(self._async_page.fill(selector, value))
    
    def click(self, selector):
        """Wrapper สำหรับ click"""
        return self._run_async(self._async_page.click(selector))
    
    def wait_for_selector(self, selector, timeout=None):
        """Wrapper สำหรับ wait_for_selector"""
        return self._run_async(self._async_page.wait_for_selector(selector, timeout=timeout))
    
    def wait_for_load_state(self, state, timeout=None):
        """Wrapper สำหรับ wait_for_load_state"""
        return self._run_async(self._async_page.wait_for_load_state(state, timeout=timeout))
    
    def set_input_files(self, selector, files):
        """Wrapper สำหรับ set_input_files"""
        return self._run_async(self._async_page.set_input_files(selector, files))
    
    def title(self):
        """Wrapper สำหรับ title"""
        return self._run_async(self._async_page.title())
    
    def content(self):
        """Wrapper สำหรับ content"""
        return self._run_async(self._async_page.content())
    
    def close(self):
        """Wrapper สำหรับ close"""
        return self._run_async(self._async_page.close())
    
    def set_default_timeout(self, timeout):
        """Wrapper สำหรับ set_default_timeout"""
        self._async_page.set_default_timeout(timeout)
    
    def set_viewport_size(self, size):
        """Wrapper สำหรับ set_viewport_size"""
        return self._run_async(self._async_page.set_viewport_size(size))
    
    def evaluate(self, expression):
        """Wrapper สำหรับ evaluate"""
        return self._run_async(self._async_page.evaluate(expression))
    
    @property
    def url(self):
        """Wrapper สำหรับ url property"""
        return self._async_page.url
    
    def screenshot(self, **kwargs):
        """Wrapper สำหรับ screenshot"""
        return self._run_async(self._async_page.screenshot(**kwargs))
    
    def reload(self, **kwargs):
        """Wrapper สำหรับ reload"""
        return self._run_async(self._async_page.reload(**kwargs))
    
    def go_back(self, **kwargs):
        """Wrapper สำหรับ go_back"""
        return self._run_async(self._async_page.go_back(**kwargs))
    
    def go_forward(self, **kwargs):
        """Wrapper สำหรับ go_forward"""
        return self._run_async(self._async_page.go_forward(**kwargs))

class AsyncLocatorWrapper:
    """Wrapper class สำหรับ AsyncLocator เพื่อให้สามารถใช้แบบ sync ได้"""
    def __init__(self, async_locator, loop):
        self._async_locator = async_locator
        self._loop = loop
    
    def _run_async(self, coro):
        """Helper method สำหรับรัน async function ใน sync context"""
        try:
            loop = self._loop or asyncio.get_event_loop()
            if loop.is_running():
                import concurrent.futures
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    future = executor.submit(asyncio.run, coro)
                    return future.result(timeout=30)
            else:
                return loop.run_until_complete(coro)
        except RuntimeError:
            return asyncio.run(coro)
        except Exception:
            return asyncio.run(coro)
    
    def count(self):
        """Wrapper สำหรับ count"""
        return self._run_async(self._async_locator.count())
    
    def fill(self, value):
        """Wrapper สำหรับ fill"""
        return self._run_async(self._async_locator.fill(value))
    
    def click(self):
        """Wrapper สำหรับ click"""
        return self._run_async(self._async_locator.click())
    
    @property
    def first(self):
        """Wrapper สำหรับ first (property)"""
        # first เป็น property ที่ return AsyncLocator ใหม่
        new_locator = self._async_locator.first
        return AsyncLocatorWrapper(new_locator, self._loop)
    
    def wait_for(self, **kwargs):
        """Wrapper สำหรับ wait_for"""
        return self._run_async(self._async_locator.wait_for(**kwargs))
    
    def input_value(self):
        """Wrapper สำหรับ input_value"""
        return self._run_async(self._async_locator.input_value())
    
    def text_content(self):
        """Wrapper สำหรับ text_content"""
        return self._run_async(self._async_locator.text_content())
    
    def set_input_files(self, files):
        """Wrapper สำหรับ set_input_files"""
        return self._run_async(self._async_locator.set_input_files(files))
    
    def press(self, key):
        """Wrapper สำหรับ press"""
        return self._run_async(self._async_locator.press(key))
    
    def is_visible(self):
        """Wrapper สำหรับ is_visible"""
        return self._run_async(self._async_locator.is_visible())

class WebAutomationPlaywright:
    def __init__(
        self,
        progress_callback: Optional[Callable[..., None]] = None,
        status_callback: Optional[Callable[..., None]] = None,
        log_callback: Optional[Callable[..., None]] = None,
        force_sync_mode: bool = False
    ):
        self.playwright = None
        self.browser = None
        self.page = None
        self.is_logged_in = False
        self.credentials = {}
        self.company_link = ""
        self.express_link = ""
        self.progress_callback = progress_callback
        self.status_callback = status_callback
        self.log_callback = log_callback
        self.file_manager = FileManager()  # สร้าง FileManager instance
        self.last_error_message = None  # เก็บ error message ล่าสุดสำหรับแสดงในหน้าเว็บ
        self.latest_iptnumber_text = None  # เก็บค่า iptnumber ล่าสุดที่อ่านได้ระหว่างกรอกฟอร์ม
        self.latest_created_file_path = None  # เก็บพาธไฟล์ที่สร้างล่าสุด เพื่อนำไปย้ายเมื่อพบแจ้งเตือนเอกสารซ้ำ
        self._current_pdf_data_for_retry = None  # เก็บ pdf_data ปัจจุบันเพื่อใช้กรอกใหม่เมื่อมีแจ้งเตือนบังคับกรอก
        self._current_excel_rows_for_retry = None  # เก็บ Excel rows (ทั้งหมดของลำดับ) เพื่อใช้กรอกใหม่เมื่อมีแจ้งเตือนบังคับกรอก
        self._current_excel_sequence_info_for_retry = None  # เก็บข้อมูล sequence (sequence, is_vat_sheet, excel_path) สำหรับ retry
        self._playwright_error = None  # เก็บ error message ของ Playwright
        self._refill_attempt_count = 0  # นับจำนวนครั้งที่กรอกใหม่จากแจ้งเตือนเพื่อกันวนไม่สิ้นสุด
        self._async_playwright = None  # สำหรับ async API ใน Streamlit
        self._async_browser = None
        self._async_page = None
        self._is_async_mode = False  # ตรวจสอบว่าใช้ async mode หรือไม่
        self._loop = None  # เก็บ event loop สำหรับ async mode
        self._force_sync_mode = force_sync_mode

    def set_progress_callback(self, callback: Optional[Callable[..., None]]):
        self.progress_callback = callback

    def set_status_callback(self, callback: Optional[Callable[..., None]]):
        self.status_callback = callback

    def set_log_callback(self, callback: Optional[Callable[..., None]]):
        self.log_callback = callback

    def _convert_date_format(self, date_value: str) -> str:
        """แปลงวันที่จากรูปแบบต่างๆ เป็น dd/mm/yyyy
        
        รองรับ:
        - yyyy-mm-dd HH:MM:SS (เช่น 2025-11-01 00:00:00)
        - yyyy-mm-dd (เช่น 2025-11-01)
        - dd/mm/yyyy (เช่น 01/11/2025) - คืนค่าเดิม
        - datetime object (ถ้าเป็น string representation)
        
        Args:
            date_value: วันที่ในรูปแบบต่างๆ
            
        Returns:
            วันที่ในรูปแบบ dd/mm/yyyy
        """
        if not date_value:
            return date_value
        
        date_str = str(date_value).strip()
        
        # ถ้าเป็นรูปแบบ dd/mm/yyyy อยู่แล้ว ให้คืนค่าเดิม
        if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', date_str):
            # แปลงให้เป็น dd/mm/yyyy (เติม 0 ถ้าจำเป็น)
            parts = date_str.split('/')
            if len(parts) == 3:
                day = parts[0].zfill(2)
                month = parts[1].zfill(2)
                year = parts[2]
                return f"{day}/{month}/{year}"
            return date_str
        
        # ถ้าเป็นรูปแบบ yyyy-mm-dd HH:MM:SS หรือ yyyy-mm-dd
        # Pattern: 2025-11-01 00:00:00 หรือ 2025-11-01
        date_match = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})(?:\s+\d{2}:\d{2}:\d{2})?', date_str)
        if date_match:
            year = date_match.group(1)
            month = date_match.group(2).zfill(2)
            day = date_match.group(3).zfill(2)
            return f"{day}/{month}/{year}"
        
        # ถ้าไม่ตรงกับรูปแบบใดๆ ให้คืนค่าเดิม
        return date_str
    
    def _clean_remark_text(self, remark: str) -> str:
        """ทำความสะอาดข้อความหมายเหตุ: ลบ .pdf, None_vat_, VAT_, WHT_, _VAT, _WHT"""
        import re
        cleaned = remark.strip()
        
        # ลบ .pdf ที่ท้ายข้อความ
        cleaned = re.sub(r'\.pdf$', '', cleaned, flags=re.IGNORECASE)
        
        # ลบ None_vat_ ที่ต้นข้อความ
        cleaned = re.sub(r'^None_vat_', '', cleaned, flags=re.IGNORECASE)
        
        # ลบ VAT_ ที่ต้นข้อความ
        cleaned = re.sub(r'^VAT_', '', cleaned, flags=re.IGNORECASE)
        
        # ลบ WHT_ ที่ต้นข้อความ
        cleaned = re.sub(r'^WHT_', '', cleaned, flags=re.IGNORECASE)
        
        # ลบ _VAT ที่ท้ายข้อความ (หลังจากลบ .pdf แล้ว)
        cleaned = re.sub(r'_VAT$', '', cleaned, flags=re.IGNORECASE)
        
        # ลบ _WHT ที่ท้ายข้อความ (หลังจากลบ .pdf แล้ว)
        cleaned = re.sub(r'_WHT$', '', cleaned, flags=re.IGNORECASE)
        
        return cleaned.strip()
    
    def _handle_file_rename_and_upload(self, row_data: Dict, document_number: str, is_vat_sheet: bool, excel_path: str) -> Optional[str]:
        """เปลี่ยนชื่อไฟล์และอัปโหลดไฟล์หลังจากบันทึกสำเร็จ
        
        Args:
            row_data: ข้อมูลจาก Excel row
            document_number: เลขที่เอกสารที่อ่านได้จากระบบ (เช่น EXP-20251000008)
            is_vat_sheet: True ถ้าเป็นชีต "มีภาษีมูลค่าเพิ่ม", False ถ้าเป็น "ไม่มีภาษีมูลค่าเพิ่ม"
            excel_path: Path ของไฟล์ Excel
        
        Returns:
            new_filename ถ้าสำเร็จ, None ถ้าไม่สำเร็จ
        """
        try:
            import json
            import re  # ต้อง import re ใน scope นี้เพื่อใช้ในฟังก์ชัน
            from datetime import datetime
            
            # 1. อ่านชื่อไฟล์เก่าจาก Excel
            old_filename = row_data.get('ชื่อไฟล์เก่า', '').strip()
            if not old_filename:
                self._log(f"⚠️ ไม่พบชื่อไฟล์เก่าใน Excel", level="warning")
                return None
            
            # 2. หาไฟล์ PDF ในโฟลเดอร์เดียวกับไฟล์ Excel
            excel_dir = os.path.dirname(excel_path)
            old_file_path = os.path.join(excel_dir, old_filename)
            
            if not os.path.exists(old_file_path):
                self._log(f"⚠️ ไม่พบไฟล์: {old_file_path}", level="warning")
                return None
            
            self._log(f"📄 พบไฟล์เก่า: {old_file_path}")
            
            # 3. อ่าน folder_settings.json และหา folder_code (เหมือนระบบอ่าน PDF)
            folder_settings = self.file_manager.read_folder_settings()
            folder_code = self.file_manager.get_folder_code_from_path(Path(excel_dir)) if excel_dir else None
            folder_group = None
            
            if folder_code:
                self._log(f"📊 พบ folder_code: {folder_code}")
                if folder_code in folder_settings:
                    folder_info = folder_settings[folder_code]
                    folder_group = folder_info.get('group', 'unknown')
                    self._log(f"📊 พบ folder_group: {folder_group}")
                else:
                    self._log(f"⚠️ ไม่พบ folder_code ใน folder_settings", level="warning")
            else:
                self._log(f"⚠️ ไม่พบ folder_code จาก path: {excel_dir}", level="warning")
            
            # 4. อ่านชื่อไฟล์ใหม่จาก Excel
            new_filename_base = row_data.get('ชื่อไฟล์ใหม่', '').strip()
            if not new_filename_base:
                self._log(f"⚠️ ไม่พบชื่อไฟล์ใหม่ใน Excel", level="warning")
                return None
            
            # ลบ .pdf ออกจากชื่อไฟล์ใหม่ (ถ้ามี)
            if new_filename_base.lower().endswith('.pdf'):
                new_filename_base = new_filename_base[:-4]
            
            # ลบ _VAT, _WHT, VAT, WHT ออกจากชื่อไฟล์ใหม่ (ถ้ามี) เพื่อป้องกันการซ้ำ
            new_filename_base = re.sub(r'(_VAT|_WHT|VAT|WHT)$', '', new_filename_base, flags=re.IGNORECASE)
            new_filename_base = new_filename_base.strip()
            
            # 5. ตรวจสอบเปอร์เซ็นต์หัก ณ ที่จ่าย (ถ้ามีข้อมูลในคอลัมน์)
            withholding_str = row_data.get('เปอร์เซ็นต์หัก ณ ที่จ่าย', '').strip()
            has_withholding = bool(withholding_str)  # ตรวจสอบว่ามีข้อมูลหรือไม่ (ไม่ว่าง)
            
            if has_withholding:
                try:
                    # ทำความสะอาดค่า (ลบอักขระที่ไม่ใช่ตัวเลขและจุดทศนิยม)
                    # ใช้ re ที่ import ไว้แล้วที่บรรทัด 13
                    withholding_clean = re.sub(r'[^\d.]', '', str(withholding_str))
                    if withholding_clean:
                        withholding_percent = float(withholding_clean)
                        self._log(f"📊 พบเปอร์เซ็นต์หัก ณ ที่จ่าย: {withholding_percent}%")
                except Exception as e:
                    self._log(f"⚠️ ไม่สามารถอ่านเปอร์เซ็นต์หัก ณ ที่จ่ายได้: {e}", level="warning")
            
            # 6. สร้างชื่อไฟล์ใหม่ตามเงื่อนไข
            new_filename = None
            
            if folder_group == 'regular':
                if is_vat_sheet:
                    # สำหรับ regular + มีภาษีมูลค่าเพิ่ม → ต้องมีวันที่เสมอ
                    date_str = row_data.get('วันที่', '').strip()
                    
                    if not date_str:
                        self._log(f"❌ ไม่พบวันที่ใน Excel สำหรับ regular + มีภาษีมูลค่าเพิ่ม - ต้องมีวันที่เสมอ", level="error")
                        return None
                    
                    # แปลงวันที่จาก DD/MM/YYYY เป็น DD.MM.YYYY
                    try:
                        date_parts = date_str.split('/')
                        if len(date_parts) == 3:
                            # ใช้ zfill(2) เพื่อให้เป็น 2 หลัก (เช่น 01 แทน 1)
                            formatted_date = f"{date_parts[0].zfill(2)}.{date_parts[1].zfill(2)}.{date_parts[2]}"
                        else:
                            self._log(f"❌ รูปแบบวันที่ไม่ถูกต้อง: {date_str} - ต้องเป็น DD/MM/YYYY", level="error")
                            return None
                    except Exception as e:
                        self._log(f"❌ ไม่สามารถแปลงวันที่ได้: {e}", level="error")
                        return None
                    
                    # สร้างชื่อไฟล์ (มีวันที่เสมอ)
                    if has_withholding:
                        # regular + มีภาษีมูลค่าเพิ่ม + หัก ณ ที่จ่าย: {DD.MM.YYYY}_WHT_{เลขที่เอกสาร}_{ชื่อไฟล์ใหม่}_VAT.pdf
                        new_filename = f"{formatted_date}_WHT_{document_number}_{new_filename_base}_VAT.pdf"
                    else:
                        # regular + มีภาษีมูลค่าเพิ่ม: {DD.MM.YYYY}_{เลขที่เอกสาร}_{ชื่อไฟล์ใหม่}_VAT.pdf
                        new_filename = f"{formatted_date}_{document_number}_{new_filename_base}_VAT.pdf"
                else:
                    # regular + ไม่มีภาษีมูลค่าเพิ่ม → ไม่ใช้วันที่
                    if has_withholding:
                        # regular + ไม่มีภาษีมูลค่าเพิ่ม + หัก ณ ที่จ่าย: WHT_{เลขที่เอกสาร}_{ชื่อไฟล์ใหม่}.pdf
                        new_filename = f"WHT_{document_number}_{new_filename_base}.pdf"
                    else:
                        # regular + ไม่มีภาษีมูลค่าเพิ่ม: {เลขที่เอกสาร}_{ชื่อไฟล์ใหม่}.pdf
                        new_filename = f"{document_number}_{new_filename_base}.pdf"
            elif folder_group == 'special':
                # special → ไม่ใช้วันที่เสมอ (ไม่ว่าจะเป็นชีตไหน)
                if has_withholding:
                    # special + หัก ณ ที่จ่าย: WHT_{เลขที่เอกสาร}_{ชื่อไฟล์ใหม่}.pdf
                    new_filename = f"WHT_{document_number}_{new_filename_base}.pdf"
                else:
                    # special + ทั้งสองชีต: {เลขที่เอกสาร}_{ชื่อไฟล์ใหม่}.pdf
                    new_filename = f"{document_number}_{new_filename_base}.pdf"
            else:
                # ถ้าไม่พบ folder_group ให้ใช้รูปแบบ default (ไม่ใช้วันที่)
                if has_withholding:
                    new_filename = f"WHT_{document_number}_{new_filename_base}.pdf"
                else:
                    self._log(f"⚠️ ไม่พบ folder_group ใช้รูปแบบ default", level="warning")
                    new_filename = f"{document_number}_{new_filename_base}.pdf"
            
            if not new_filename:
                self._log(f"⚠️ ไม่สามารถสร้างชื่อไฟล์ใหม่ได้", level="warning")
                return None
            
            # 6. สร้างไฟล์ใหม่ (คัดลอกไฟล์ต้นฉบับ) แทนการเปลี่ยนชื่อ (เหมือนระบบอ่าน PDF)
            new_file_path = os.path.join(excel_dir, new_filename)
            original_file_path = old_file_path  # เก็บ path ไฟล์ต้นฉบับไว้
            
            try:
                # คัดลอกไฟล์แทนการ rename เพื่อเก็บไฟล์ต้นฉบับไว้
                if self.file_manager._copy_with_retry(old_file_path, new_file_path):
                    self._log(f"✅ สร้างไฟล์ใหม่สำเร็จ: {new_filename}")
                else:
                    self._log(f"⚠️ ไม่สามารถสร้างไฟล์ใหม่ได้", level="warning")
                    return None
            except Exception as e:
                self._log(f"⚠️ เกิดข้อผิดพลาดในการสร้างไฟล์ใหม่: {e}", level="warning")
                return None
            
            # 7. อัปโหลดไฟล์เข้าเว็บ
            self._log(f"📤 กำลังอัปโหลดไฟล์: {new_filename}")
            upload_success = False
            if self.upload_file_directly(new_file_path):
                self._log(f"✅ อัปโหลดไฟล์สำเร็จ: {new_filename}")
                upload_success = True
            else:
                self._log(f"⚠️ อัปโหลดไฟล์ไม่สำเร็จ: {new_filename}", level="warning")
                return None
            
            # 8. ย้ายไฟล์ไปยังโฟลเดอร์ต่างๆ (เหมือนการอ่าน PDF)
            if upload_success:
                try:
                    # ถ้า folder_group = 'special' → ย้ายไป "เอกสาร NoneVat" เสมอ
                    if folder_group == 'special':
                        move_group = 'special'
                        self._log(f"📋 folder_group='special' → ย้ายไป 'เอกสาร NoneVat' เสมอ")
                    else:
                        # สำหรับ regular → ตรวจสอบว่าชื่อไฟล์มีรูปแบบวันที่หรือไม่ (DD.MM.YYYY)
                        # ถ้ามีรูปแบบวันที่ → ย้ายไป "เอกสาร Vat" (move_group='regular')
                        # ถ้าไม่มีรูปแบบวันที่ → ย้ายไป "เอกสาร NoneVat" (move_group='special')
                        import re
                        has_date_pattern = re.search(r'\d{2}\.\d{2}\.\d{4}', new_filename)
                        
                        if has_date_pattern:
                            # มีรูปแบบวันที่ → ย้ายไป "เอกสาร Vat"
                            move_group = 'regular'
                            self._log(f"📅 พบรูปแบบวันที่ในชื่อไฟล์ → ย้ายไป 'เอกสาร Vat'")
                        else:
                            # ไม่มีรูปแบบวันที่ → ย้ายไป "เอกสาร NoneVat"
                            move_group = 'special'
                            self._log(f"📄 ไม่พบรูปแบบวันที่ในชื่อไฟล์ → ย้ายไป 'เอกสาร NoneVat'")
                    
                    self._log(f"📁 กำลังย้ายไฟล์: original='{original_file_path}', processed='{new_file_path}', group='{move_group}'")
                    # ใช้ file_manager.move_original_and_processed() เพื่อย้ายไฟล์
                    # ไฟล์ต้นฉบับ (original_file_path) → เอกสารต้นฉบับ
                    # ไฟล์ที่เปลี่ยนชื่อแล้ว (new_file_path) → เอกสาร Vat หรือ เอกสาร NoneVat
                    # หมายเหตุ: original_file_path อาจไม่มีอยู่แล้วเพราะเปลี่ยนชื่อแล้ว แต่ file_manager จะตรวจสอบเอง
                    if self.file_manager.move_original_and_processed(original_file_path, new_file_path, move_group):
                        self._log(f"✅ ย้ายไฟล์สำเร็จ")
                        return new_filename
                    else:
                        self._log(f"⚠️ ย้ายไฟล์ไม่สำเร็จ", level="warning")
                        return new_filename  # ยังคืนค่า new_filename เพราะอัปโหลดสำเร็จแล้ว
                except Exception as e:
                    self._log(f"⚠️ เกิดข้อผิดพลาดในการย้ายไฟล์: {e}", level="warning")
                    return new_filename  # ยังคืนค่า new_filename เพราะอัปโหลดสำเร็จแล้ว
            
            return new_filename if upload_success else None
                
        except Exception as e:
            self._log(f"⚠️ เกิดข้อผิดพลาดในการเปลี่ยนชื่อไฟล์และอัปโหลด: {e}", level="warning")
            import traceback
            self._log(f"📋 Traceback: {traceback.format_exc()}", level="error")
            return False

    def _log(self, message: str, level: str = "info"):
        try:
            print(message)
        except Exception:
            pass
        if self.log_callback:
            try:
                self.log_callback(message, level)
            except Exception:
                pass

    def _status_update(self, *, folder=None, file=None, step=None):
        if self.status_callback:
            try:
                self.status_callback(folder=folder, file=file, step=step)
            except Exception:
                pass

    def _notify_progress(self, *, total_delta=0, success_delta=0, failure_delta=0, duplicate_delta=0, reset=False):
        if self.progress_callback:
            try:
                self.progress_callback(
                    total_delta=total_delta,
                    success_delta=success_delta,
                    failure_delta=failure_delta,
                    duplicate_delta=duplicate_delta,
                    reset=reset
                )
            except Exception as e:
                print(f"⚠️ ไม่สามารถอัพเดตความคืบหน้า (automation): {e}")
        
    def read_config_from_txt(self, txt_file_path: str) -> bool:
        """อ่านข้อมูลการตั้งค่าจากไฟล์ Build000.txt"""
        try:
            self._log(f"📖 กำลังอ่านข้อมูลการตั้งค่าจาก: {txt_file_path}")
            
            with open(txt_file_path, 'r', encoding='utf-8') as file:
                content = file.read()
                
            # แยกข้อมูลตามบรรทัด
            lines = content.strip().split('\n')
            self._log(f"🔍 เนื้อหาที่อ่านได้ ({len(lines)} บรรทัด):")
            for i, line in enumerate(lines):
                self._log(f"   บรรทัด {i+1}: '{line}'")
            
            # Reset credentials และ links
            self.credentials = {}
            self.company_link = ""
            self.express_link = ""
            
            for line in lines:
                line = line.strip()
                if not line:  # ข้ามบรรทัดว่าง
                    continue
                    
                self._log(f"🔍 ตรวจสอบบรรทัด: '{line}'")
                
                if line.startswith('Username :'):
                    username = line.split(':', 1)[-1].strip()
                    self.credentials['Username'] = username
                    self._log(f"✅ อ่าน Username: {username}")
                    
                elif line.startswith('Password :'):
                    password = line.split(':', 1)[-1].strip()
                    self.credentials['Password'] = password
                    self._log(f"✅ อ่าน Password: {'*' * len(password)}")  # ไม่แสดง password จริง
                    
                elif line.startswith('Link company :'):
                    company_link = line.split(':', 1)[-1].strip()
                    self.company_link = company_link
                    self._log(f"✅ อ่าน Company Link: {company_link}")
                    
                elif line.startswith('Link Express :'):
                    express_link = line.split(':', 1)[-1].strip()
                    self.express_link = express_link
                    self._log(f"✅ อ่าน Express Link: {express_link}")
                else:
                    self._log(f"⚠️ ไม่ตรงกับรูปแบบที่กำหนด: '{line}'", level="warning")
            
            # แสดงข้อมูลที่อ่านได้
            self._log(f"📊 สรุปข้อมูลที่อ่านได้:")
            self._log(f"   Username: '{self.credentials.get('Username', 'ไม่พบ')}'")
            self._log(f"   Password: {'พบ' if self.credentials.get('Password') else 'ไม่พบ'}")
            self._log(f"   Company Link: '{self.company_link}'")
            self._log(f"   Express Link: '{self.express_link}'")
            
            # ตรวจสอบว่าอ่านข้อมูลครบหรือไม่
            if (self.credentials.get('Username') and 
                self.credentials.get('Password') and 
                self.company_link and 
                self.express_link):
                self._log(f"✅ อ่านข้อมูลการตั้งค่าเสร็จสิ้น")
                return True
            else:
                missing = []
                if not self.credentials.get('Username'):
                    missing.append('Username')
                if not self.credentials.get('Password'):
                    missing.append('Password')
                if not self.company_link:
                    missing.append('Company Link')
                if not self.express_link:
                    missing.append('Express Link')
                self._log(f"❌ อ่านข้อมูลการตั้งค่าไม่ครบ: ขาด {', '.join(missing)}", level="error")
                return False
                
        except Exception as e:
            self._log(f"❌ เกิดข้อผิดพลาดในการอ่านไฟล์การตั้งค่า: {e}", level="error")
            import traceback
            self._log(f"📋 Traceback: {traceback.format_exc()}", level="error")
            return False
        
    def setup_driver(self):
        """ตั้งค่า Playwright Browser"""
        try:
            import sys
            import asyncio
            
            # ตรวจสอบว่าอยู่ใน Streamlit environment หรือไม่
            in_streamlit = 'streamlit' in sys.modules
            
            if in_streamlit and not self._force_sync_mode:
                # ใน Streamlit ต้องใช้ Async API แทน Sync API
                if not ASYNC_PLAYWRIGHT_AVAILABLE or async_playwright is None:
                    error_msg = "❌ Async Playwright API ไม่พร้อมใช้งาน"
                    print(error_msg)
                    self._playwright_error = error_msg
                    return False
                
                print("🔧 กำลังรัน Playwright ด้วย Async API (Streamlit mode)...")
                
                async def setup_async_playwright():
                    """Setup Playwright ด้วย Async API"""
                    try:
                        self._async_playwright = await async_playwright().start()
                        
                        # เลือก browser (chromium, firefox, webkit)
                        self._async_browser = await self._async_playwright.chromium.launch(
                            headless=False,  # แสดงหน้าต่าง browser
                            args=[
                                '--no-sandbox',
                                '--disable-dev-shm-usage',
                                '--disable-gpu',
                                '--start-maximized',  # เปิดเต็มจอตั้งแต่แรก
                                '--window-size=1920,1080'
                            ]
                        )
                        
                        # สร้าง page ใหม่
                        self._async_page = await self._async_browser.new_page()
                        
                        # ตั้งค่า viewport ให้เต็มจอ
                        await self._async_page.set_viewport_size({"width": 1920, "height": 1080})
                        
                        # ตั้งค่า timeout
                        self._async_page.set_default_timeout(Config.SELENIUM_TIMEOUT * 1000)  # แปลงเป็น milliseconds
                        
                        # เก็บ async objects ไว้และตั้งค่า flag
                        self._is_async_mode = True
                        # เก็บ event loop สำหรับใช้ในภายหลัง
                        try:
                            self._loop = asyncio.get_event_loop()
                        except RuntimeError:
                            self._loop = asyncio.new_event_loop()
                            asyncio.set_event_loop(self._loop)
                        
                        self.playwright = self._async_playwright
                        self.browser = self._async_browser
                        # ใช้ wrapper class เพื่อให้สามารถใช้แบบ sync ได้
                        self.page = AsyncPageWrapper(self._async_page, self._loop)
                        
                        print("✅ Playwright setup completed successfully (Async API - Streamlit mode)")
                        return True
                    except Exception as e:
                        print(f"❌ Error setting up async Playwright: {e}")
                        import traceback
                        traceback.print_exc()
                        return False
                
                # รัน async function ใน event loop
                try:
                    # ตรวจสอบว่ามี event loop อยู่หรือไม่
                    try:
                        loop = asyncio.get_event_loop()
                        if loop.is_running():
                            # ถ้า loop กำลังรันอยู่ ใช้ create_task
                            import concurrent.futures
                            with concurrent.futures.ThreadPoolExecutor() as executor:
                                future = executor.submit(asyncio.run, setup_async_playwright())
                                result = future.result(timeout=30)
                        else:
                            # ถ้า loop ไม่ได้รัน ใช้ run
                            result = loop.run_until_complete(setup_async_playwright())
                    except RuntimeError:
                        # ถ้าไม่มี loop ให้สร้างใหม่
                        result = asyncio.run(setup_async_playwright())
                    
                    if result:
                        return True
                    else:
                        error_msg = "❌ ไม่สามารถตั้งค่า Playwright ด้วย Async API ได้"
                        print(error_msg)
                        self._playwright_error = error_msg
                        return False
                        
                except Exception as e:
                    error_msg = (
                        f"❌ Playwright ไม่สามารถทำงานได้: {e}\n"
                        "💡 **แนะนำ**: ใช้ Tkinter GUI แทน Streamlit\n"
                        "   รันคำสั่ง: python bot_gui_tkinter.py"
                    )
                    print(error_msg)
                    self._playwright_error = error_msg
                    return False
            else:
                # ไม่ใช่ Streamlit ให้ใช้วิธีปกติ
                self.playwright = sync_playwright().start()
                
                # เลือก browser (chromium, firefox, webkit)
                self.browser = self.playwright.chromium.launch(
                    headless=False,  # แสดงหน้าต่าง browser
                    args=[
                        '--no-sandbox',
                        '--disable-dev-shm-usage',
                        '--disable-gpu',
                        '--start-maximized',  # เปิดเต็มจอตั้งแต่แรก
                        '--window-size=1920,1080'
                    ]
                )
                
                # สร้าง page ใหม่
                self.page = self.browser.new_page()
                
                # ตั้งค่า viewport ให้เต็มจอ
                self.page.set_viewport_size({"width": 1920, "height": 1080})
                
                # ตั้งค่า timeout
                self.page.set_default_timeout(Config.SELENIUM_TIMEOUT * 1000)  # แปลงเป็น milliseconds
                
                print("✅ Playwright setup completed successfully")
                return True
            
        except NotImplementedError as e:
            error_msg = (
                f"❌ ไม่สามารถเปิด Playwright ได้: {e}\n"
                "💡 หมายเหตุ: ปัญหานี้มักเกิดใน Streamlit environment บน Windows\n"
                "💡 แนะนำให้:\n"
                "   1. ใช้ Tkinter GUI แทน Streamlit (แนะนำ): python bot_gui_tkinter.py\n"
                "   2. หรือติดตั้ง nest-asyncio: pip install nest-asyncio\n"
                "   3. หรือใช้ Docker/Linux environment"
            )
            print(error_msg)
            # เก็บ error message ไว้ในตัวแปรเพื่อให้ Streamlit แสดงได้
            self._playwright_error = error_msg
            return False
        except Exception as e:
            print(f"❌ Error setting up Playwright: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def login_to_peak_engine(self, credentials: Dict) -> bool:
        """ล็อกอินเข้า Peak Engine และคลิกปุ่มกลับไปยัง Peak เก่า"""
        try:
            if not self.page:
                self._log("Playwright page not initialized", level="error")
                return False
            
            # ไปยังหน้า login
            self._log(f"🌐 กำลังไปยังหน้า Login: {Config.PEAK_ENGINE_URL}")
            try:
                # ใช้ load แทน networkidle เพื่อลด timeout
                self.page.goto(Config.PEAK_ENGINE_URL, wait_until='load', timeout=60000)
                self._log(f"✅ โหลดหน้า Login สำเร็จ")
            except Exception as e:
                self._log(f"⚠️ เกิดข้อผิดพลาดในการโหลดหน้า Login: {e}", level="warning")
                # ลองต่อแม้มี error
                self._log(f"🔄 กำลังลองต่อ...")
            
            # รอให้หน้าโหลดเสร็จ
            try:
                self.page.wait_for_load_state('domcontentloaded', timeout=30000)
                self._log(f"✅ DOM โหลดเสร็จ")
            except Exception as e:
                self._log(f"⚠️ รอ DOM timeout: {e}", level="warning")
            
            # รอให้หน้าเว็บโหลดเสร็จสมบูรณ์ (ลดเป็น 1 วินาที)
            self._log(f"⏳ รอให้หน้าเว็บโหลดเสร็จสมบูรณ์ (1 วินาที)...")
            time.sleep(1)  # รอ 1 วินาทีเพื่อให้หน้าเว็บและ JavaScript โหลดเสร็จ
            
            # ลองหา username field ด้วย selector หลายแบบ (ใช้ selector เดียวกับ peak_engine_workflow)
            username_selectors = [
                'input[placeholder="อีเมล"]',
                'input[type="email"]',
                'input[name="email"]',
                'input[name="username"]',
                'input[type="text"][name="username"]',
                'input#username',
                'input.email',
                'input[placeholder*="email" i]',
                'input[placeholder*="username" i]',
                'input[placeholder*="อีเมล" i]'
            ]
            
            username_field = None
            for selector in username_selectors:
                try:
                    locator = self.page.locator(selector)
                    count = locator.count()
                    if count > 0:
                        self._log(f"✅ พบ username field ด้วย selector: {selector} (พบ {count} ตัว)")
                        username_field = locator.first
                        break
                    else:
                        self._log(f"🔍 ลอง selector: {selector} - ไม่พบ", level="debug")
                except Exception as e:
                    self._log(f"⚠️ Error กับ selector {selector}: {e}", level="debug")
                    continue
            
            if not username_field:
                self._log("❌ ไม่พบ username field ด้วย selector ใดๆ", level="error")
                # ลองหา input ทั้งหมดเพื่อ debug
                try:
                    all_inputs = self.page.locator('input').all()
                    self._log(f"🔍 พบ input ทั้งหมด {len(all_inputs)} ตัว:")
                    for i, inp in enumerate(all_inputs[:10]):  # แสดงแค่ 10 ตัวแรก
                        try:
                            input_type = inp.get_attribute('type') or 'text'
                            input_name = inp.get_attribute('name') or ''
                            input_id = inp.get_attribute('id') or ''
                            input_placeholder = inp.get_attribute('placeholder') or ''
                            self._log(f"   Input {i+1}: type={input_type}, name={input_name}, id={input_id}, placeholder={input_placeholder}")
                        except:
                            pass
                except Exception as e:
                    self._log(f"⚠️ ไม่สามารถอ่าน input ทั้งหมดได้: {e}", level="warning")
                
                # ลอง screenshot เพื่อ debug
                try:
                    screenshot_path = "debug_login_username_not_found.png"
                    self.page.screenshot(path=screenshot_path, full_page=True)
                    self._log(f"📸 บันทึก screenshot: {screenshot_path}")
                except:
                    pass
                return False
            
            # รอให้ field พร้อมใช้งาน
            try:
                username_field.wait_for(state='visible', timeout=15000)
            except Exception as e:
                self._log(f"⚠️ รอ username field timeout: {e}", level="warning")
                # ลองต่อแม้ timeout
            
            username_value = credentials.get('Username', '')
            if not username_value:
                self._log("❌ ไม่พบ Username ใน credentials", level="error")
                return False
            
            try:
                username_field.fill(username_value)
                self._log(f"✅ กรอก username สำเร็จ: {username_value}")
            except Exception as e:
                self._log(f"❌ ไม่สามารถกรอก username ได้: {e}", level="error")
                return False
            time.sleep(0.3)  # ลด delay
            
            # ลองหา password field ด้วย selector หลายแบบ
            password_selectors = [
                'input[name="password"]',
                'input[type="password"]',
                'input#password',
                'input.password'
            ]
            
            password_field = None
            for selector in password_selectors:
                try:
                    locator = self.page.locator(selector)
                    if locator.count() > 0:
                        self._log(f"✅ พบ password field ด้วย selector: {selector}")
                        password_field = locator.first
                        break
                except:
                    continue
            
            if not password_field:
                self._log("❌ ไม่พบ password field", level="error")
                return False
            
            # รอให้ field พร้อมใช้งาน
            try:
                password_field.wait_for(state='visible', timeout=15000)
            except Exception as e:
                self._log(f"⚠️ รอ password field timeout: {e}", level="warning")
                # ลองต่อแม้ timeout
            
            password_value = credentials.get('Password', '')
            if not password_value:
                self._log("❌ ไม่พบ Password ใน credentials", level="error")
                return False
            
            try:
                password_field.fill(password_value)
                self._log(f"✅ กรอก password สำเร็จ")
            except Exception as e:
                self._log(f"❌ ไม่สามารถกรอก password ได้: {e}", level="error")
                return False
            time.sleep(0.3)  # ลด delay
            
            # หาและกดปุ่ม login (ใช้ selector เดียวกับ peak_engine_workflow)
            login_selectors = [
                '#loginbtn',
                'button#loginbtn',
                'input#loginbtn',
                'button:has-text("เข้าใช้งาน")',
                'button:has-text("Login")',
                'button[type="submit"]',
                'button.teal',
                'button[class*="teal"]',
                'button:has-text("เข้าสู่ระบบ")',
                'input[type="submit"]',
                'button.btn-primary',
                'button.login'
            ]
            
            login_button = None
            for selector in login_selectors:
                try:
                    locator = self.page.locator(selector)
                    count = locator.count()
                    if count > 0:
                        self._log(f"✅ พบ login button ด้วย selector: {selector} (พบ {count} ตัว)")
                        login_button = locator.first
                        break
                    else:
                        self._log(f"🔍 ลอง selector: {selector} - ไม่พบ", level="debug")
                except Exception as e:
                    self._log(f"⚠️ Error กับ selector {selector}: {e}", level="debug")
                    continue
            
            if not login_button:
                self._log("❌ ไม่พบ login button ด้วย selector ใดๆ", level="error")
                # ลองหา button ทั้งหมดเพื่อ debug
                try:
                    all_buttons = self.page.locator('button, input[type="submit"]').all()
                    self._log(f"🔍 พบ button ทั้งหมด {len(all_buttons)} ตัว:")
                    for i, btn in enumerate(all_buttons[:10]):  # แสดงแค่ 10 ตัวแรก
                        try:
                            btn_text = btn.text_content() or ''
                            btn_id = btn.get_attribute('id') or ''
                            btn_type = btn.get_attribute('type') or ''
                            btn_class = btn.get_attribute('class') or ''
                            self._log(f"   Button {i+1}: text='{btn_text[:30]}', id={btn_id}, type={btn_type}, class={btn_class[:50]}")
                        except:
                            pass
                except Exception as e:
                    self._log(f"⚠️ ไม่สามารถอ่าน button ทั้งหมดได้: {e}", level="warning")
                
                # ลอง screenshot เพื่อ debug
                try:
                    screenshot_path = "debug_login_button_not_found.png"
                    self.page.screenshot(path=screenshot_path, full_page=True)
                    self._log(f"📸 บันทึก screenshot: {screenshot_path}")
                except:
                    pass
                return False
            
            # รอให้ button พร้อมใช้งาน
            try:
                login_button.wait_for(state='visible', timeout=15000)
            except Exception as e:
                self._log(f"⚠️ รอ login button timeout: {e}", level="warning")
                # ลองต่อแม้ timeout
            
            try:
                login_button.click()
                self._log(f"✅ คลิกปุ่ม Login สำเร็จ")
                # รอสักครู่หลังคลิก (ลดเป็น 0.5 วินาที)
                time.sleep(0.5)
            except Exception as e:
                self._log(f"❌ ไม่สามารถคลิกปุ่ม Login ได้: {e}", level="error")
                import traceback
                self._log(f"📋 Traceback: {traceback.format_exc()}", level="error")
                return False
            
            # รอให้ล็อกอินสำเร็จ (รอให้ URL เปลี่ยน)
            self._log(f"⏳ กำลังรอผลลัพธ์การ Login...")
            try:
                # รอให้หน้าโหลดเสร็จก่อน (ลด timeout)
                self.page.wait_for_load_state('load', timeout=8000)
                self._log(f"✅ หน้าโหลดเสร็จ")
                
                # รอให้ URL เปลี่ยน (ไม่ใช่หน้า login) - ลด timeout
                self.page.wait_for_function(
                    "() => !window.location.href.includes('/Login') && !window.location.href.includes('/login')",
                    timeout=20000  # ลดจาก 30 เป็น 20 วินาที
                )
                self._log(f"✅ รอหน้าโหลดหลัง Login สำเร็จ")
                # ไม่ต้องรอเพิ่ม - ใช้เวลาจาก wait_for_function แล้ว
            except Exception as e:
                self._log(f"⚠️ รอ URL เปลี่ยน timeout: {e}", level="warning")
                # ลองตรวจสอบ URL ปัจจุบัน
                try:
                    current_url = self.page.url
                    self._log(f"📍 URL ปัจจุบัน: {current_url}")
                    if "login" not in current_url.lower():
                        self._log(f"✅ URL ไม่มี 'login' ถือว่าล็อกอินสำเร็จ")
                    else:
                        # รอเพิ่มอีก 3 วินาที
                        self._log(f"⏳ รอเพิ่มอีก 3 วินาที...")
                        time.sleep(3)
                        current_url = self.page.url
                        self._log(f"📍 URL หลังรอ: {current_url}")
                        if "login" in current_url.lower():
                            self._log(f"❌ ยังอยู่ที่หน้า login", level="error")
                            return False
                except Exception as url_error:
                    self._log(f"⚠️ ไม่สามารถตรวจสอบ URL ได้: {url_error}", level="warning")
                    # ลองต่อแม้มี error
                    time.sleep(2)
            
            # ตรวจสอบว่าล็อกอินสำเร็จหรือไม่ (ใช้วิธีเดียวกับ peak_engine_workflow)
            try:
                current_url = self.page.url
                self._log(f"📍 URL ปัจจุบัน: {current_url}")
            except Exception as e:
                self._log(f"⚠️ ไม่สามารถอ่าน URL ได้: {e}", level="warning")
                current_url = ""
            
            # ตรวจสอบเงื่อนไขการล็อกอิน (เหมือน workflow ปกติ)
            self._log(f"🔍 ตรวจสอบเงื่อนไขการล็อกอิน...")
            login_success = False
            
            # เงื่อนไขที่ 1: URL ไม่มี "home/login" (ไม่ใช่แค่ "login" เพราะอาจมีใน query string)
            if current_url:
                has_home_login = "home/login" in current_url.lower()
                has_login_only = "login" in current_url.lower() and not has_home_login
                is_login_page = current_url.lower() == "https://secure.peakengine.com/home/login"
                
                self._log(f"   URL ปัจจุบัน: {current_url}")
                self._log(f"   มี 'home/login' ใน URL: {has_home_login}")
                self._log(f"   มี 'login' ใน URL (แต่ไม่ใช่ home/login): {has_login_only}")
                self._log(f"   เท่ากับ URL เริ่มต้น: {is_login_page}")
                
                if not has_home_login and not is_login_page:
                    login_success = True
                    self._log(f"✅ ล็อกอินสำเร็จ (เงื่อนไขที่ 1): URL ไม่ใช่หน้า login")
            
            # เงื่อนไขที่ 2: ไม่มีฟิลด์ login อยู่แล้ว (สำคัญที่สุด!)
            if not login_success:
                try:
                    login_fields = self.page.locator('input[type="email"], input[type="password"]')
                    login_field_count = login_fields.count()
                    self._log(f"   พบฟิลด์ login: {login_field_count} ตัว")
                    if login_field_count == 0:
                        login_success = True
                        self._log(f"✅ ล็อกอินสำเร็จ (เงื่อนไขที่ 2): ไม่มีฟิลด์ login อยู่แล้ว")
                except Exception as e:
                    self._log(f"⚠️ ไม่สามารถตรวจสอบฟิลด์ login ได้: {e}", level="warning")
            
            # เงื่อนไขที่ 3: มีข้อความแสดงว่าล็อกอินสำเร็จ
            if not login_success:
                try:
                    success_messages = self.page.locator('text=ยินดีด้วย, text=พร้อมเริ่มต้น, text=PEAK')
                    if success_messages.count() > 0:
                        login_success = True
                        self._log(f"✅ ล็อกอินสำเร็จ (เงื่อนไขที่ 3): พบข้อความแสดงความสำเร็จ")
                except:
                    pass
            
            if login_success:
                self.is_logged_in = True
                self._log("✅ Login สำเร็จ!")
                
                # หลังจากล็อกอินสำเร็จ ให้คลิกปุ่ม "กลับไปยัง Peak เก่า"
                self._log(f"🔄 กำลังคลิกปุ่มกลับไปยัง Peak เก่า...")
                try:
                    # รอให้ปุ่มปรากฏ (เพิ่ม timeout และ error handling)
                    try:
                        self.page.wait_for_selector('#btnBackToOldPeak', timeout=15000)
                    except Exception as e:
                        self._log(f"⚠️ ไม่พบปุ่มกลับไปยัง Peak เก่า: {e}", level="warning")
                        # ลองตรวจสอบว่ามีปุ่มนี้หรือไม่
                        back_button_check = self.page.locator('#btnBackToOldPeak')
                        if back_button_check.count() == 0:
                            self._log(f"ℹ️ ไม่พบปุ่มกลับไปยัง Peak เก่า - อาจจะไม่จำเป็นต้องคลิก", level="info")
                            return True  # ถือว่าล็อกอินสำเร็จแล้ว
                        else:
                            raise
                    
                    # คลิกปุ่มด้วยวิธีที่ถูกต้อง
                    back_button = self.page.locator('#btnBackToOldPeak')
                    if back_button.count() > 0:
                        # รอให้ปุ่มพร้อมใช้งาน
                        back_button.first.wait_for(state='visible')
                        
                        # คลิกปุ่ม
                        back_button.first.click()
                        self._log(f"✅ คลิกปุ่มกลับไปยัง Peak เก่าสำเร็จ")
                        
                        # รอให้หน้าโหลดเสร็จ (ใช้ load แทน networkidle เพื่อความเร็ว)
                        self.page.wait_for_load_state('load', timeout=15000)
                        time.sleep(0.5)  # ลด delay
                        
                        # ตรวจสอบ URL หลังคลิกปุ่ม
                        current_url = self.page.url
                        self._log(f"📍 URL หลังคลิกปุ่มกลับ: {current_url}")
                        
                    else:
                        self._log(f"⚠️ ไม่พบปุ่มกลับไปยัง Peak เก่า", level="warning")
                        
                except Exception as e:
                    self._log(f"⚠️ เกิดข้อผิดพลาดในการคลิกปุ่มกลับ: {e}", level="warning")
                    # ลองใช้วิธีอื่น
                    try:
                        self._log(f"🔍 ลองใช้วิธีอื่นในการคลิกปุ่ม...")
                        self.page.click('#btnBackToOldPeak')
                        self._log(f"✅ คลิกปุ่มสำเร็จด้วยวิธีอื่น")
                        
                        # รอให้หน้าโหลดเสร็จ (ใช้ load แทน networkidle เพื่อความเร็ว)
                        self.page.wait_for_load_state('load', timeout=15000)
                        time.sleep(0.5)  # ลด delay
                        
                        # ตรวจสอบ URL หลังคลิกปุ่ม
                        current_url = self.page.url
                        self._log(f"📍 URL หลังคลิกปุ่มกลับ: {current_url}")
                        
                    except Exception as e2:
                        self._log(f"❌ ไม่สามารถคลิกปุ่มได้ด้วยวิธีใดเลย: {e2}", level="warning")
                
                return True
            else:
                self._log("❌ Login ไม่สำเร็จ - ยังอยู่ที่หน้า login", level="error")
                return False
                
        except Exception as e:
            self._log(f"❌ เกิดข้อผิดพลาดระหว่าง Login: {e}", level="error")
            import traceback
            self._log(f"📋 Traceback: {traceback.format_exc()}", level="error")
            self._playwright_error = str(e)
            # ลอง screenshot เพื่อ debug
            try:
                if self.page:
                    screenshot_path = "debug_login_error.png"
                    self.page.screenshot(path=screenshot_path, full_page=True)
                    self._log(f"📸 บันทึก screenshot: {screenshot_path}")
            except Exception as screenshot_error:
                self._log(f"⚠️ ไม่สามารถบันทึก screenshot ได้: {screenshot_error}", level="warning")
            
            # ตรวจสอบว่า page ยังเปิดอยู่หรือไม่
            try:
                if self.page:
                    current_url = self.page.url
                    self._log(f"📍 URL หลัง error: {current_url}")
            except:
                self._log(f"⚠️ ไม่สามารถเข้าถึง page ได้ - อาจจะปิดไปแล้ว", level="warning")
            
            return False
    
    def navigate_to_company_link(self, company_link: str) -> bool:
        """ไปยัง Link Company"""
        try:
            if not self.is_logged_in:
                print("Not logged in")
                return False
            
            print(f"Navigating to company link: {company_link}")
            self.page.goto(company_link)
            
            
            return True
            
        except Exception as e:
            print(f"Error navigating to company link: {e}")
            return False
    
    def navigate_to_express_link(self, express_link: str) -> bool:
        """ไปยัง Link Express"""
        try:
            if not self.is_logged_in:
                print("Not logged in")
                return False
            
            print(f"Navigating to express link: {express_link}")
            self.page.goto(express_link)
            
            
            return True
            
        except Exception as e:
            print(f"Error navigating to express link: {e}")
            return False
    
    def fill_form_data(self, pdf_data: Dict) -> bool:
        """กรอกข้อมูลในฟอร์มด้วยข้อมูลจาก PDF"""
        try:
            if not self.is_logged_in:
                print("Not logged in")
                return False
            
            print(f"📝 กรอกข้อมูลในฟอร์ม: {pdf_data.get('filename', 'ไม่ทราบชื่อ')}")
            print(f"📊 ข้อมูลที่จะกรอก:")
            print(f"   Company: {pdf_data.get('company_name', '')}")
            print(f"   Group: {pdf_data.get('group', 'unknown')}")
            print(f"   Service name: {pdf_data.get('service_name', '')}")
            print(f"   Customer ID: {pdf_data.get('customer_id', '')}")
            print(f"   Account Code: {pdf_data.get('account_code', '')}")
            print(f"   เลขที่เอกสาร: {pdf_data.get('document_number', '')}")
            print(f"   วันที่เอกสาร: {pdf_data.get('document_date', '')}")
            print(f"   ยอดก่อนภาษีมูลค่าเพิ่ม: {pdf_data.get('total_ex_vat', '')}")
            print(f"   ยอดก่อนภาษีมูลค่าเพิ่ม (NoneVat): {pdf_data.get('total_ex_vat_none', '')}")
            print(f"   ยอดภาษีมูลค่าเพิ่ม: {pdf_data.get('vat_value', '')}")
            print(f"   ยอดหลังบวกภาษีมูลค่าเพิ่ม: {pdf_data.get('total_in_vat', '')}")
            
            # รอให้ฟอร์มโหลดเสร็จ
            time.sleep(0.5)
            
            # กรอกข้อมูลในฟอร์มตามข้อมูลจาก PDF
            try:
                # 1. กรอก Customer ID - //*[@id="iptcontactname"] >> เลื่อน Arrow down >> enter
                print(f"🔍 กำลังกรอก Customer ID...")
                try:
                    # รอให้ element iptnumber โหลดก่อนเพื่อป้องกันการกรอกเร็วเกินไป
                    try:
                        print("⏳ รอให้ iptnumber พร้อมใช้งาน...")
                        self.page.wait_for_selector('//*[@id="iptnumber"]', timeout=5000)
                        print("✅ iptnumber พร้อมสำหรับการทำงาน")
                    except Exception as wait_err:
                        print(f"⚠️ รอ iptnumber ไม่สำเร็จ: {wait_err}")

                    customer_field = self.page.locator('//*[@id="iptcontactname"]')
                    selected_from_dropdown = False
                    if customer_field.count() > 0:
                        customer_field.first.fill(pdf_data.get('customer_id', ''))
                        print(f"✅ กรอก Customer ID: {pdf_data.get('customer_id', '')}")
                    
                    # รอให้ดรอปดาวน์ปรากฏ และตรวจสอบรายการ
                    print(f"⏳ รอรายการลูกค้า 1 วินาที...")
                    time.sleep(1)
                    
                    try:
                        dropdown_items = self.page.locator('css=ul.ui-autocomplete:visible li')
                        dropdown_count = dropdown_items.count()
                        print(f"📋 จำนวนรายการที่แสดงในดรอปดาวน์: {dropdown_count}")
                        if dropdown_count > 1:
                            for item_index in range(dropdown_count):
                                candidate = self.page.locator(f'css=ul.ui-autocomplete:visible li:nth-child({item_index + 1})')
                                if candidate.count() == 0:
                                    continue
                                item_text = (candidate.first.text_content() or '').strip()
                                if not item_text or '+ เพิ่มผู้ติดต่อ' in item_text:
                                    continue
                                candidate.first.click()
                                selected_from_dropdown = True
                                print(f"✅ เลือกรายการลูกค้าโดยอัตโนมัติ: '{item_text}'")
                                break
                    except Exception as dropdown_err:
                        print(f"⚠️ ตรวจสอบดรอปดาวน์ไม่สำเร็จ: {dropdown_err}")
                    
                    if not selected_from_dropdown:
                        # ลองหาตัวเลือกรายการที่ตรงกับ Customer ID
                        customer_id = pdf_data.get('customer_id', '')
                        print(f"🔍 กำลังหาตัวเลือกที่มี '{customer_id}'...")
                        
                        # วิธีที่ 1: หาด้วย text content ที่มี Customer ID
                        customer_option = self.page.locator(f'text={customer_id}')
                        if customer_option.count() > 0:
                            print(f"✅ พบตัวเลือกที่มี '{customer_id}'")
                            customer_option.first.click()
                            print(f"✅ คลิกตัวเลือกสำเร็จ")
                        else:
                            # วิธีที่ 2: กดลง 1 ครั้งเพื่อเลือกรายการที่ 2 (ข้าม "+ เพิ่มผู้ติดต่อ")
                            print(f"🔽 ลองใช้วิธีกดลง 1 ครั้ง...")
                            customer_field.first.press('ArrowDown')
                            time.sleep(0.5)
                            
                            # ตรวจสอบว่าตัวเลือกที่เลือกถูกต้องหรือไม่
                            try:
                                selected_text = customer_field.first.input_value()
                                print(f"📋 ตัวเลือกที่เลือก: '{selected_text}'")
                                
                                if customer_id in str(selected_text):
                                    print(f"✅ เลือกรายการที่ถูกต้อง: {selected_text}")
                                else:
                                    # วิธีที่ 3: กดลงอีก 1 ครั้ง (รวมเป็น 2 ครั้ง)
                                    print(f"🔽 กดลงครั้งที่ 2...")
                                    customer_field.first.press('ArrowDown')
                                    time.sleep(0.5)
                                    
                                    # ตรวจสอบอีกครั้ง
                                    selected_text_2 = customer_field.first.input_value()
                                    print(f"📋 ตัวเลือกหลังกดลง 2 ครั้ง: '{selected_text_2}'")
                                    
                                    if customer_id in str(selected_text_2):
                                        print(f"✅ เลือกรายการที่ถูกต้องหลังกดลง 2 ครั้ง: {selected_text_2}")
                                    else:
                                        print(f"⚠️ ไม่สามารถเลือกรายการที่ถูกต้องได้")
                            except Exception as e:
                                print(f"⚠️ ไม่สามารถตรวจสอบตัวเลือกที่เลือกได้: {e}")
                        
                            # กด Enter เพื่อยืนยันการเลือก
                            print(f"⏎ กด Enter เพื่อยืนยัน...")
                            customer_field.first.press('Enter')
                            time.sleep(0.5)
                            print(f"✅ กด Enter สำเร็จ")
                except Exception as e:
                    print(f"⚠️ เกิดข้อผิดพลาดในการกรอก Customer ID: {e}")
                    # ลองใช้วิธี fallback
                    try:
                        customer_field = self.page.locator('#iptcontactname')
                        if customer_field.count() > 0:
                            customer_field.first.fill(pdf_data.get('customer_id', ''))
                            customer_field.first.press("ArrowDown")
                            time.sleep(0.5)
                            customer_field.first.press("Enter")
                            print(f"✅ ใช้วิธี fallback สำเร็จ")
                    except Exception as e2:
                        print(f"❌ ไม่สามารถกรอก Customer ID ได้: {e2}")
                        return False
                    
                # รอ 1 วินาทีหลังกรอก Customer ID
                print(f"⏳ รอ 1 วินาทีหลังกรอก Customer ID...")
                time.sleep(0.5)
                
                # ตรวจสอบที่อยู่ //*[@id="lbcontactaddress"] ต้องไม่เป็น "-"
                print(f"🔍 ตรวจสอบที่อยู่...")
                address_field = self.page.locator('//*[@id="lbcontactaddress"]')
                if address_field.count() > 0:
                    address_text = address_field.first.text_content()
                    print(f"📍 ที่อยู่ปัจจุบัน: '{address_text}'")
                    
                    if address_text and address_text.strip() != "-" and address_text.strip() != "":
                        print(f"✅ ที่อยู่ถูกต้อง: {address_text}")
                    else:
                        print(f"⚠️ ที่อยู่ยังไม่ถูกต้อง (เป็น '-' หรือว่างเปล่า)")
                        # รอเพิ่มเติมให้ที่อยู่โหลดเสร็จ
                        print(f"⏳ รอให้ที่อยู่โหลดเสร็จ...")
                        time.sleep(1)
                        
                        # ตรวจสอบอีกครั้ง
                        address_text_retry = address_field.first.text_content()
                        print(f"📍 ที่อยู่หลังรอ: '{address_text_retry}'")
                        
                        if address_text_retry and address_text_retry.strip() != "-" and address_text_retry.strip() != "":
                            print(f"✅ ที่อยู่ถูกต้องหลังรอ: {address_text_retry}")
                        else:
                            print(f"❌ ที่อยู่ยังไม่ถูกต้อง แม้จะรอแล้ว")
                else:
                    print(f"⚠️ ไม่พบฟิลด์ที่อยู่ (lbcontactaddress)")
                
                # 2. กรอก Account Code - //*[@id="iptaccountcode1"] >> enter
                print(f"🔍 กำลังกรอก Account Code...")
                try:
                    account_field = self.page.locator('//*[@id="iptaccountcode1"]')
                    if account_field.count() > 0:
                        account_field.first.fill(pdf_data.get('account_code', ''))
                        print(f"✅ กรอก Account Code: {pdf_data.get('account_code', '')}")
                        account_field.first.press('Enter')
                        print(f"✅ กด Enter สำเร็จ")
                    else:
                        print(f"⚠️ ไม่พบฟิลด์ Account Code (iptaccountcode1)")
                except Exception as e:
                    print(f"⚠️ เกิดข้อผิดพลาดในการกรอก Account Code: {e}")
                
                # 3. กรอกวันที่เอกสาร - //*[@id="iptdate"] (ย้ายขึ้นมาก่อน)
                print(f"🔍 กำลังกรอกวันที่เอกสาร...")
                try:
                    date_field = self.page.locator('//*[@id="iptdate"]')
                    if date_field.count() > 0:
                        date_field.first.fill(pdf_data.get('document_date', ''))
                        print(f"✅ กรอกวันที่เอกสาร: {pdf_data.get('document_date', '')}")
                        # ถ้าวันที่ยังว่าง ให้กรอก Customer ID ใหม่ทันที (ไม่ต้องรอ)
                        try:
                            current_date_val = date_field.first.input_value()
                        except Exception:
                            current_date_val = ''
                        if not current_date_val:
                            print("⚠️ ไม่พบค่าวันที่หลังกรอก จะกรอก Customer ID ใหม่ทันที...")
                            self.fill_customer_id_again(pdf_data)
                    else:
                        print(f"⚠️ ไม่พบฟิลด์วันที่เอกสาร (iptdate)")
                except Exception as e:
                    print(f"⚠️ เกิดข้อผิดพลาดในการกรอกวันที่เอกสาร: {e}")
                
                # 4. กรอกเลขที่เอกสาร (เฉพาะบริษัท VAT เท่านั้น)
                try:
                    company_name = pdf_data.get('company_name', '')
                    company_vat_status = Config.COMPANY_VAT_STATUS.get(company_name, 'VAT')
                    folder_group = pdf_data.get('group') or 'unknown'
                    
                    self._log(f"🔍 ตรวจสอบเงื่อนไข: folder_group='{folder_group}', company_vat_status='{company_vat_status}'")
                    
                    # ตรวจสอบจาก folder_group ก่อน (special = NoneVat, regular = VAT)
                    if folder_group == 'special':
                        print(f"ℹ️ โฟลเดอร์ special (NoneVat): ข้ามขั้นตอนกรอกเลขที่เอกสาร")
                        skip_invoice = True
                    elif company_vat_status == 'NoneVat':
                        print(f"ℹ️ บริษัท {company_name}: ทำงานแบบ NoneVat - ข้ามขั้นตอนกรอกเลขที่เอกสาร")
                        skip_invoice = True
                    else:
                        print(f"ℹ️ บริษัท {company_name}: ทำงานแบบ VAT - กรอกเลขที่เอกสาร")
                        skip_invoice = False
                except Exception as e:
                    print(f"⚠️ เกิดข้อผิดพลาดในการตรวจสอบเงื่อนไข: {e}")
                    skip_invoice = False
                
                # 5. กรอกเลขที่เอกสาร (เฉพาะบริษัท VAT เท่านั้น)
                if not skip_invoice:
                    print(f"🔍 กำลังกรอกเลขที่เอกสาร...")
                    try:
                        invoice_button = self.page.locator('//*[@id="receivedTaxInvoiceAddButton"]')
                        if invoice_button.count() > 0:
                            invoice_button.first.click()
                            print(f"✅ คลิกปุ่ม receivedTaxInvoiceAddButton สำเร็จ")
                            
                            # รอป๊อปอัป
                            time.sleep(1)
                            
                            # กรอกเลขที่เอกสารในป๊อปอัป
                            invoice_input = self.page.locator('//*[@id="inputModalReceivedTaxInvoiceNumber"]')
                            if invoice_input.count() > 0:
                                invoice_input.first.fill(pdf_data.get('document_number', ''))
                                print(f"✅ กรอกเลขที่เอกสาร: {pdf_data.get('document_number', '')}")
                                
                                # คลิกปุ่มในป๊อปอัป
                                popup_button = self.page.locator('//*[@id="receivedTaxInvoiceAddModalSize"]/div/div[3]/div[2]')
                                if popup_button.count() > 0:
                                    popup_button.first.click()
                                    print(f"✅ คลิกปุ่มในป๊อปอัปสำเร็จ")
                                else:
                                    print(f"⚠️ ไม่พบปุ่มในป๊อปอัป")
                            else:
                                print(f"⚠️ ไม่พบฟิลด์เลขที่เอกสารในป๊อปอัป")
                        else:
                            print(f"⚠️ ไม่พบปุ่ม receivedTaxInvoiceAddButton")
                    except Exception as e:
                        print(f"⚠️ เกิดข้อผิดพลาดในการกรอกเลขที่เอกสาร: {e}")
                else:
                    print(f"⏭️ ข้ามการกรอกเลขที่เอกสาร (NoneVat)")
                
                # 5. กรอกยอดหลังบวกภาษีมูลค่าเพิ่ม - //*[@id="iptprice1"]
                print(f"🔍 กำลังกรอกยอดหลังบวกภาษีมูลค่าเพิ่ม...")
                price_field = self.page.locator('//*[@id="iptprice1"]')
                if price_field.count() > 0:
                    price_field.first.fill(str(pdf_data.get('total_in_vat', '')))
                    print(f"✅ กรอกยอดหลังบวกภาษีมูลค่าเพิ่ม: {pdf_data.get('total_in_vat', '')}")
                else:
                    print(f"⚠️ ไม่พบฟิลด์ยอดหลังบวกภาษีมูลค่าเพิ่ม (iptprice1)")
                
                # 6. เลือกดรอปดาวน์ภาษี (แตกต่างกันตาม folder_group และ company_vat_status)
                # ใช้ตัวแปรที่คำนวณไว้ด้านบน (company_vat_status, folder_group)
                # ให้ความสำคัญกับ folder_group ก่อน
                if folder_group == 'special' or company_vat_status == 'NoneVat':
                    # สำหรับ NoneVat: เลือก "ไม่มี" (value="1") ใน ddlvattypeid1
                    print(f"🔍 [NoneVat] กำลังเลือกดรอปดาวน์ประเภทภาษี (ddlvattypeid1) เป็น 'ไม่มี'...")
                    vat_type_dropdown = self.page.locator('//*[@id="ddlvattypeid1"]')
                    if vat_type_dropdown.count() > 0:
                        # วิธีที่ 1: เลือกด้วย value="1"
                        try:
                            self.page.select_option('//*[@id="ddlvattypeid1"]', value='1')
                            print(f"✅ เลือก 'ไม่มี' สำเร็จด้วยวิธีที่ 1 (value='1')")
                        except:
                            # วิธีที่ 2: ใช้ JavaScript
                            try:
                                self.page.evaluate("""
                                    const dropdown = document.querySelector('#ddlvattypeid1');
                                    if (dropdown) {
                                        dropdown.value = '1';
                                        dropdown.dispatchEvent(new Event('change', {bubbles: true}));
                                    }
                                """)
                                print(f"✅ เลือก 'ไม่มี' สำเร็จด้วยวิธีที่ 2 (JavaScript)")
                            except:
                                # วิธีที่ 3: คลิกแล้วเลือกด้วย text
                                try:
                                    vat_type_dropdown.first.click()
                                    time.sleep(0.5)
                                    no_vat_option = self.page.locator('text=ไม่มี')
                                    if no_vat_option.count() > 0:
                                        no_vat_option.first.click()
                                        print(f"✅ เลือก 'ไม่มี' สำเร็จด้วยวิธีที่ 3 (text)")
                                    else:
                                        # วิธีที่ 4: เลือกด้วย index
                                        self.page.select_option('//*[@id="ddlvattypeid1"]', index=1)
                                        print(f"✅ เลือก 'ไม่มี' สำเร็จด้วยวิธีที่ 4 (index=1)")
                                except Exception as e:
                                    print(f"⚠️ ไม่สามารถเลือก 'ไม่มี' ได้: {e}")
                    else:
                        print(f"⚠️ ไม่พบดรอปดาวน์ประเภทภาษี (ddlvattypeid1)")
                    
                    # ตรวจสอบว่าการเลือกสำเร็จหรือไม่ (สำหรับ NoneVat)
                    try:
                        selected_value = vat_type_dropdown.first.input_value()
                        print(f"📋 ค่าประเภทภาษีที่เลือกได้: '{selected_value}'")
                        if selected_value == '1':
                            print(f"✅ เลือก 'ไม่มี' สำเร็จ!")
                        else:
                            print(f"⚠️ ค่าที่เลือกไม่ตรงกับที่ต้องการ: '{selected_value}'")
                    except Exception as e:
                        print(f"📋 ไม่สามารถอ่านค่าที่เลือกได้: {e}")

                else:
                    # สำหรับ VAT: เลือก "รวมภาษี" ใน ddltaxstatus
                    print(f"🔍 [VAT] กำลังเลือกดรอปดาวน์รวมภาษี (ddltaxstatus)...")
                    tax_status_dropdown = self.page.locator('//*[@id="ddltaxstatus"]')
                    if tax_status_dropdown.count() > 0:
                        # คลิกดรอปดาวน์เพื่อเปิดตัวเลือก
                        tax_status_dropdown.first.click()
                        print(f"✅ เปิดดรอปดาวน์รวมภาษีสำเร็จ")
                        
                        # รอให้ดรอปดาวน์เปิดและรายการปรากฏ
                        print(f"⏳ รอให้รายการดรอปดาวน์ปรากฏ...")
                        time.sleep(1)  # รอให้ดรอปดาวน์เปิดสมบูรณ์
                        
                        # ลองหาตัวเลือก "รวมภาษี" ด้วยหลายวิธี
                        tax_included_option = None
                        
                        # วิธีที่ 1: หาด้วย text content
                        tax_included_option = self.page.locator('text=รวมภาษี, text=Include Tax')
                        if tax_included_option.count() > 0:
                            tax_included_option.first.click()
                            print(f"✅ เลือกรวมภาษีสำเร็จด้วยวิธีที่ 1")
                        else:
                            # วิธีที่ 2: เลือกด้วย index
                            try:
                                self.page.select_option('//*[@id="ddltaxstatus"]', index=1)
                                print(f"✅ เลือกรวมภาษีสำเร็จด้วยวิธีที่ 2 (index=1)")
                            except:
                                # วิธีที่ 3: ใช้ JavaScript
                                try:
                                    self.page.evaluate("""
                                        const dropdown = document.querySelector('#ddltaxstatus');
                                        if (dropdown) {
                                            for (let i = 0; i < dropdown.options.length; i++) {
                                                if (dropdown.options[i].text.includes('รวมภาษี') || 
                                                    dropdown.options[i].text.includes('Include Tax')) {
                                                    dropdown.selectedIndex = i;
                                                    dropdown.dispatchEvent(new Event('change'));
                                                    break;
                                                }
                                            }
                                        }
                                    """)
                                    print(f"✅ เลือกรวมภาษีสำเร็จด้วยวิธีที่ 3 (JavaScript)")
                                except:
                                    # วิธีที่ 4: กดลง 2 ครั้ง แล้วกด Enter
                                    print(f"🔽 ลองใช้วิธีกดลง 2 ครั้ง...")
                                    tax_status_dropdown.first.press('ArrowDown')
                                    time.sleep(0.3)
                                    
                                    print(f"🔽 กดลงครั้งที่ 2...")
                                    tax_status_dropdown.first.press('ArrowDown')
                                    time.sleep(0.3)
                                    
                                    print(f"⏎ กด Enter...")
                                    tax_status_dropdown.first.press('Enter')
                                    time.sleep(0.3)
                                    
                                    print(f"✅ เลือกรวมภาษีสำเร็จด้วยวิธีที่ 4 (กดลง 2 ครั้ง + Enter)")
                    else:
                        print(f"⚠️ ไม่พบดรอปดาวน์รวมภาษี (ddltaxstatus)")
                        

                
                # 6.1 อ่านเลขเอกสารจาก iptnumber ล่วงหน้า ก่อนคลิกปุ่มสุดท้าย
                try:
                    ipt_loc = self.page.locator('//*[@id="iptnumber"]')
                    if ipt_loc.count() > 0:
                        try:
                            txt = (ipt_loc.first.input_value() or '').strip()
                        except Exception:
                            txt = (ipt_loc.first.text_content() or '').strip()
                        self.latest_iptnumber_text = txt if txt else None
                        print(f"📄 บันทึก iptnumber ล่วงหน้า: '{self.latest_iptnumber_text or ''}'")
                    else:
                        print("⚠️ ไม่พบ element iptnumber สำหรับบันทึกล่วงหน้า")
                except Exception as _:
                    print("⚠️ อ่าน iptnumber ล่วงหน้าไม่สำเร็จ")
                except Exception as e:
                    print(f"⚠️ เกิดข้อผิดพลาดในการกรอกเลขที่เอกสาร: {e}")

                # 7. คลิกปุ่ม hidePaymentModal - //*[@id="hidePaymentModal"]
                print(f"🔍 กำลังคลิกปุ่ม hidePaymentModal...")
                try:
                    hide_payment_button = self.page.locator('//*[@id="hidePaymentModal"]')
                    if hide_payment_button.count() > 0:
                        hide_payment_button.first.click()
                        print(f"✅ คลิกปุ่ม hidePaymentModal สำเร็จ")
                    else:
                        print(f"⚠️ ไม่พบปุ่ม hidePaymentModal")
                except Exception as e:
                    print(f"⚠️ เกิดข้อผิดพลาดในการคลิกปุ่ม hidePaymentModal: {e}")
                
                # 8. คลิกปุ่มสุดท้าย - //*[@id="content"]/div[6]/div[6]/div/div[2]/div[2]/div[1]
                print(f"🔍 กำลังกดปุ่มสุดท้าย...")
                # รอ 1.5 วินาทีก่อนกดปุ่มบันทึก
                time.sleep(1.5)
                try:
                    final_button = self.page.locator('//*[@id="content"]/div[6]/div[6]/div/div[2]/div[2]/div[1]')
                    if final_button.count() > 0:
                        final_button.first.click()
                        print(f"✅ คลิกปุ่มสุดท้ายสำเร็จ")
                    else:
                        print(f"⚠️ ไม่พบปุ่มสุดท้าย")
                except Exception as e:
                    print(f"⚠️ เกิดข้อผิดพลาดในการคลิกปุ่มสุดท้าย: {e}")

                print(f"✅ กรอกข้อมูลในฟอร์มสำเร็จ!")
                return True
                
            except Exception as e:
                print(f"⚠️ เกิดข้อผิดพลาดในการกรอกฟิลด์เฉพาะ: {e}")
                print(f"🔍 ลองใช้วิธีกรอกฟอร์มทั่วไป...")
                # ถ้ากรอกฟิลด์เฉพาะไม่ได้ ให้ลองกรอกฟอร์มทั่วไป
                try:
                    return self.fill_form_generic(pdf_data)
                except Exception as e2:
                    print(f"❌ เกิดข้อผิดพลาดในการกรอกฟอร์มทั่วไป: {e2}")
                    return False
            
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการกรอกข้อมูล: {e}")
            return False
    
    def fill_form_generic(self, pdf_data: Dict) -> bool:
        """กรอกฟอร์มแบบทั่วไป (fallback)"""
        try:
            print(f"🔍 ใช้วิธีกรอกฟอร์มทั่วไป...")
            
            # หา input fields ทั้งหมด
            input_fields = self.page.locator('input[type="text"], input[type="number"], textarea')
            print(f"📊 พบ input fields: {input_fields.count()} ฟิลด์")
            
            # กรอกข้อมูลตามลำดับจาก PDF (ตรงกับ pdf_reader.py)
            data_to_fill = [
                pdf_data.get('customer_id', ''),         # Customer ID
                pdf_data.get('account_code', ''),        # Account Code
                pdf_data.get('account_code2', ''),       # Account Code2
                pdf_data.get('document_number', ''),     # เลขที่เอกสาร
                pdf_data.get('document_date', ''),       # วันที่เอกสาร
                str(pdf_data.get('total_ex_vat', '')),   # ยอดก่อนภาษีมูลค่าเพิ่ม
                str(pdf_data.get('total_ex_vat_none', '')), # ยอดก่อนภาษีมูลค่าเพิ่ม (NoneVat)
                str(pdf_data.get('vat_value', '')),      # ยอดภาษีมูลค่าเพิ่ม
                str(pdf_data.get('total_in_vat', '')),   # ยอดหลังบวกภาษีมูลค่าเพิ่ม
                pdf_data.get('company_name', '')         # ชื่อบริษัท
            ]
            
            # แสดงข้อมูลที่จะกรอกในรูปแบบที่กำหนด
            print(f"📝 ข้อมูลที่จะกรอก:")
            print(f"   Customer ID: {pdf_data.get('customer_id', '')}")
            print(f"   Account Code: {pdf_data.get('account_code', '')}")
            print(f"   เลขที่เอกสาร: {pdf_data.get('document_number', '')}")
            print(f"   วันที่เอกสาร: {pdf_data.get('document_date', '')}")
            print(f"   ยอดก่อนภาษีมูลค่าเพิ่ม: {pdf_data.get('total_ex_vat', '')}")
            print(f"   ยอดก่อนภาษีมูลค่าเพิ่ม (NoneVat): {pdf_data.get('total_ex_vat_none', '')}")
            print(f"   ยอดภาษีมูลค่าเพิ่ม: {pdf_data.get('vat_value', '')}")
            print(f"   ยอดหลังบวกภาษีมูลค่าเพิ่ม: {pdf_data.get('total_in_vat', '')}")
            print(f"   ชื่อบริษัท: {pdf_data.get('company_name', '')}")
            
            print(f"📝 ข้อมูลที่จะกรอก: {data_to_fill}")
            
            filled_count = 0
            for i, data in enumerate(data_to_fill):
                if i < input_fields.count() and data:
                    try:
                        input_fields.nth(i).fill(data)
                        print(f"✅ กรอกฟิลด์ที่ {i+1}: {data}")
                        filled_count += 1
                        time.sleep(0.5)
                    except Exception as e:
                        print(f"⚠️ ไม่สามารถกรอกฟิลด์ที่ {i+1}: {e}")
            
            print(f"✅ กรอกข้อมูลสำเร็จ {filled_count}/{len([d for d in data_to_fill if d])} ฟิลด์")
            return filled_count > 0
            
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการกรอกฟอร์มทั่วไป: {e}")
            return False
    
    def upload_file(self, file_path: str) -> bool:
        """อัปโหลดไฟล์"""
        try:
            if not self.is_logged_in:
                print("Not logged in")
                return False
            
            print(f"Uploading file: {file_path}")
            
            # 1. เปิดเต็มจอ
            print(f"🖥️ กำลังเปิดเต็มจอ...")
            self.page.bring_to_front()
            self.page.evaluate("document.documentElement.requestFullscreen()")
            time.sleep(1)
            
            # 2. รอให้หน้าเว็บโหลดเสร็จ (ลด timeout และใช้วิธีอื่น)
            print(f"⏳ รอให้หน้าเว็บโหลดเสร็จ...")
            try:
                # ลองรอ networkidle แต่ลด timeout
                self.page.wait_for_load_state('networkidle', timeout=5000)
                print(f"✅ หน้าเว็บโหลดเสร็จ (networkidle)")
            except:
                # ถ้า timeout ให้รอแค่ DOM content loaded
                try:
                    self.page.wait_for_load_state('domcontentloaded', timeout=3000)
                    print(f"✅ หน้าเว็บโหลดเสร็จ (domcontentloaded)")
                except:
                    print(f"⚠️ รอหน้าเว็บโหลด timeout ใช้การรอแบบธรรมดา")
                
            time.sleep(2)  # รอเพิ่มเติมให้แน่ใจว่าโหลดเสร็จแล้ว
            
            # 3. เลื่อนลงไปหาปุ่ม "เพิ่มไฟล์" ที่อยู่ด้านล่าง
            print(f"🔍 กำลังเลื่อนลงไปหาปุ่มเพิ่มไฟล์...")
            
            # วิธีที่ 1: เลื่อนลงไปด้านล่างของหน้า
            print(f"📜 วิธีที่ 1: เลื่อนลงไปด้านล่างของหน้า...")
            self.page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(1)
            
            # วิธีที่ 2: เลื่อนทีละส่วน
            print(f"📜 วิธีที่ 2: เลื่อนทีละส่วน...")
            self.page.evaluate("window.scrollBy(0, 500)")
            time.sleep(0.5)
            self.page.evaluate("window.scrollBy(0, 500)")
            time.sleep(0.5)
            self.page.evaluate("window.scrollBy(0, 500)")
            time.sleep(0.5)
            
            # วิธีที่ 3: เลื่อนไปยัง element ที่มีคำว่า "เพิ่มไฟล์"
            print(f"📜 วิธีที่ 3: เลื่อนไปยัง element ที่มีคำว่า 'เพิ่มไฟล์'...")
            try:
                add_file_elements = self.page.locator('text=เพิ่มไฟล์, text=Add file, text=Upload, text=อัปโหลด')
                if add_file_elements.count() > 0:
                    # เลื่อนไปยัง element แรกที่พบ
                    add_file_elements.first.scroll_into_view_if_needed()
                    print(f"✅ เลื่อนไปยัง element 'เพิ่มไฟล์' สำเร็จ")
                    time.sleep(1)
                else:
                    print(f"⚠️ ไม่พบ element ที่มีคำว่า 'เพิ่มไฟล์'")
            except Exception as scroll_error:
                print(f"⚠️ ไม่สามารถเลื่อนไปยัง element ได้: {scroll_error}")
            
            # วิธีที่ 4: เลื่อนไปยังปุ่มชำระเงิน (ถ้ามี)
            print(f"📜 วิธีที่ 4: เลื่อนไปยังปุ่มชำระเงิน...")
            try:
                pay_button = self.page.locator('button:has-text("ชำระเงิน"), button:has-text("Pay"), button:has-text("Payment")')
                if pay_button.count() > 0:
                    pay_button.first.scroll_into_view_if_needed()
                    print(f"✅ เลื่อนไปยังปุ่มชำระเงินสำเร็จ")
                    time.sleep(1)
                else:
                    print(f"⚠️ ไม่พบปุ่มชำระเงิน")
            except Exception as scroll_error:
                print(f"⚠️ ไม่สามารถเลื่อนไปยังปุ่มชำระเงินได้: {scroll_error}")
            
            # 2. หาปุ่ม "เพิ่มไฟล์" โดยใช้ XPath ที่ระบุ
            print(f"🔍 กำลังหาปุ่มเพิ่มไฟล์...")
            
            # ใช้ XPath ที่ผู้ใช้ระบุ
            specific_xpath = '//*[@id="content"]/div[6]/div[1]/div[4]/div[7]/div[3]/div[2]/div[3]/div'
            
                        # หาปุ่มด้วย XPath ที่ระบุ
            add_file_button = self.page.locator(f'xpath={specific_xpath}')
            
            if add_file_button.count() > 0:
                print(f"✅ พบปุ่มเพิ่มไฟล์ที่ XPath: {specific_xpath}")
                
                # เลื่อนไปยังปุ่มนั้น
                print(f"📜 เลื่อนไปยังปุ่มเพิ่มไฟล์...")
                add_file_button.first.scroll_into_view_if_needed()
                time.sleep(1)
                
                # คลิกปุ่มเพิ่มไฟล์
                print(f"🖱️ กำลังคลิกปุ่มเพิ่มไฟล์...")
                add_file_button.first.click()
                time.sleep(1)  # รอให้ dialog เปิด
                print(f"✅ คลิกปุ่มเพิ่มไฟล์สำเร็จ")
                
            else:
                print(f"❌ ไม่พบปุ่มเพิ่มไฟล์ที่ XPath: {specific_xpath}")
                
                # ลองหาปุ่มเพิ่มไฟล์ด้วยวิธีเดิม (เป็น fallback)
                print(f"🔍 ลองหาปุ่มเพิ่มไฟล์ด้วยวิธีอื่น...")
                add_file_button = self.page.locator('button:has-text("เพิ่มไฟล์"), button:has-text("Add file"), button:has-text("Upload"), button:has-text("อัปโหลด")')
                
                if add_file_button.count() > 0:
                    print(f"✅ พบปุ่มเพิ่มไฟล์ (fallback): {add_file_button.count()} ปุ่ม")
                    
                    # คลิกปุ่มเพิ่มไฟล์
                    print(f"🖱️ กำลังคลิกปุ่มเพิ่มไฟล์ (fallback)...")
                    add_file_button.first.click()
                    time.sleep(1)  # รอให้ dialog เปิด
                    print(f"✅ คลิกปุ่มเพิ่มไฟล์สำเร็จ")
                else:
                    print(f"❌ ไม่พบปุ่มเพิ่มไฟล์")
                    return False
            
            # 3. หา file input field หลังจากคลิกปุ่ม
            print(f"🔍 กำลังหาฟิลด์อัปโหลดหลังจากคลิกปุ่ม...")
            file_input = None
            
            # วิธีที่ 1: หาด้วย type="file"
            file_input = self.page.locator('input[type="file"]')
            if file_input.count() > 0:
                print(f"✅ พบฟิลด์อัปโหลดด้วยวิธีที่ 1: input[type='file']")
            else:
                # วิธีที่ 2: หาด้วย id ที่มีคำว่า file, upload, attachment
                file_input = self.page.locator('input[id*="file"], input[id*="upload"], input[id*="attachment"]')
                if file_input.count() > 0:
                    print(f"✅ พบฟิลด์อัปโหลดด้วยวิธีที่ 2: id ที่มีคำว่า file/upload/attachment")
                else:
                    # วิธีที่ 3: หาด้วย name ที่มีคำว่า file, upload, attachment
                    file_input = self.page.locator('input[name*="file"], input[name*="upload"], input[name*="attachment"]')
                    if file_input.count() > 0:
                        print(f"✅ พบฟิลด์อัปโหลดด้วยวิธีที่ 3: name ที่มีคำว่า file/upload/attachment")
                    else:
                        # วิธีที่ 4: หาด้วย class ที่มีคำว่า file, upload, attachment
                        file_input = self.page.locator('input[class*="file"], input[name*="upload"], input[name*="attachment"]')
                        if file_input.count() > 0:
                            print(f"✅ พบฟิลด์อัปโหลดด้วยวิธีที่ 4: class ที่มีคำว่า file/upload/attachment")
                        else:
                            # วิธีที่ 5: หาด้วย placeholder ที่มีคำว่า file, upload, attachment
                            file_input = self.page.locator('input[placeholder*="file"], input[placeholder*="upload"], input[placeholder*="attachment"]')
                            if file_input.count() > 0:
                                print(f"✅ พบฟิลด์อัปโหลดด้วยวิธีที่ 5: placeholder ที่มีคำว่า file/upload/attachment")
                            else:
                                # วิธีที่ 6: หาด้วย label ที่มีคำว่า file, upload, attachment
                                file_labels = self.page.locator('label:has-text("file"), label:has-text("upload"), label:has-text("attachment"), label:has-text("ไฟล์"), label:has-text("อัปโหลด")')
                                if file_labels.count() > 0:
                                    # หา input ที่เกี่ยวข้องกับ label
                                    for i in range(file_labels.count()):
                                        label_text = file_labels.nth(i).text_content()
                                        print(f"🔍 พบ label: '{label_text}'")
                                        # หา input ที่เกี่ยวข้อง
                                        related_input = self.page.locator(f'input[aria-label*="{label_text}"], input[title*="{label_text}"]')
                                        if related_input.count() > 0:
                                            file_input = related_input
                                            print(f"✅ พบฟิลด์อัปโหลดด้วยวิธีที่ 6: label ที่เกี่ยวข้อง")
                                            break
                                else:
                                    # วิธีที่ 7: หาด้วย XPath ที่อาจเป็นฟิลด์อัปโหลด
                                    file_input = self.page.locator('//input[contains(@id, "file") or contains(@name, "file") or contains(@class, "file")]')
                                    if file_input.count() > 0:
                                        print(f"✅ พบฟิลด์อัปโหลดด้วยวิธีที่ 7: XPath")
                                    else:
                                        print(f"⚠️ ไม่พบฟิลด์อัปโหลดด้วยวิธีใดเลย")
                                        return False
            
            if file_input and file_input.count() > 0:
                print(f"📁 พบฟิลด์อัปโหลด: {file_input.count()} ฟิลด์")
                
                # แสดงข้อมูลของฟิลด์ที่พบ
                for i in range(min(file_input.count(), 3)):  # แสดงแค่ 3 ฟิลด์แรก
                    try:
                        input_id = file_input.nth(i).get_attribute('id')
                        input_name = file_input.nth(i).get_attribute('name')
                        input_class = file_input.nth(i).get_attribute('class')
                        print(f"   ฟิลด์ที่ {i+1}: id='{input_id}', name='{input_name}', class='{input_class}'")
                    except:
                        pass
                
                # อัปโหลดไฟล์
                print(f"📤 กำลังอัปโหลดไฟล์: {file_path}")
                file_input.first.set_input_files(file_path)
                
                # รอให้อัปโหลดเสร็จ
                time.sleep(1)
                
                print("✅ File uploaded successfully")
                return True
            else:
                print("❌ No file input field found")
                return False
            
        except Exception as e:
            print(f"❌ Error uploading file: {e}")
            return False
    
    def wait_for_processing(self, timeout: int = 60) -> bool:
        """รอให้การประมวลผลเสร็จสิ้น"""
        try:
            start_time = time.time()
            
            while time.time() - start_time < timeout:
                # ตรวจสอบแจ้งเตือนสำเร็จ (dvgreenalert) ถ้ามีถือว่าเสร็จงานทันที
                try:
                    success_alert = self.page.locator('//*[@id="dvgreenalert"]')
                    if success_alert.count() > 0 and success_alert.first.is_visible():
                        try:
                            alert_label = self.page.locator('//*[@id="lbgreenalert"]')
                            alert_text = (alert_label.first.text_content() or '').strip() if alert_label.count() > 0 else (success_alert.first.text_content() or '').strip()
                        except Exception:
                            alert_text = (success_alert.first.text_content() or '').strip()
                        # เคสอัปโหลดสำเร็จปกติ
                        if any(keyword in alert_text for keyword in ['เพิ่มไฟล์สำเร็จ', 'อัปโหลดสำเร็จ', 'อัปไฟล์สำเร็จ', 'Upload', 'Uploaded']):
                            print("✅ พบการแจ้งเตือนสำเร็จ: เพิ่มไฟล์สำเร็จ → จบการทำงาน")
                            return True
                        # เคสเอกสารซ้ำ ให้ย้ายไฟล์ที่สร้างไปไว้ในโฟลเดอร์ 'เอกสารซ้ำรอตรวจ'
                        if 'เลขที่ใบกำกับภาษีอ้างอิงนี้สำหรับผู้ติดต่อนี้เคยถูกใช้งานแล้ว' in alert_text or 'ผู้ติดต่อนี้เคยถูกใช้งานแล้ว' in alert_text:
                            print('⚠️ พบแจ้งเตือนเอกสารซ้ำ → ย้ายไฟล์ไปยังโฟลเดอร์ เอกสารซ้ำรอตรวจ')
                            try:
                                if getattr(self, 'latest_created_file_path', None):
                                    moved_to = self.file_manager.move_file_to_duplicate_folder(self.latest_created_file_path)
                                    if moved_to:
                                        print(f"✅ ย้ายไฟล์ไปยังโฟลเดอร์เอกสารซ้ำรอตรวจแล้ว: {moved_to}")
                                    else:
                                        print("❌ ย้ายไฟล์ไปยังโฟลเดอร์เอกสารซ้ำรอตรวจไม่สำเร็จ")
                                else:
                                    print("⚠️ ไม่มีพาธไฟล์ล่าสุดให้ย้าย (latest_created_file_path ว่าง)")
                            except Exception as move_err:
                                print(f"⚠️ ย้ายไฟล์ไปโฟลเดอร์เอกสารซ้ำรอตรวจล้มเหลว: {move_err}")
                            return True
                except Exception:
                    pass

                # ตรวจสอบแจ้งเตือน error สีแดง (dvredalert) สำหรับเอกสารซ้ำ
                try:
                    error_alert = self.page.locator('//*[@id="dvredalert"]')
                    if error_alert.count() > 0 and error_alert.first.is_visible():
                        try:
                            alert_label = self.page.locator('//*[@id="lbredalert"]')
                            alert_text = (alert_label.first.text_content() or '').strip() if alert_label.count() > 0 else (error_alert.first.text_content() or '').strip()
                        except Exception:
                            alert_text = (error_alert.first.text_content() or '').strip()
                        if 'เลขที่ใบกำกับภาษีอ้างอิงนี้สำหรับผู้ติดต่อนี้เคยถูกใช้งานแล้ว' in alert_text or 'ผู้ติดต่อนี้เคยถูกใช้งานแล้ว' in alert_text:
                            print('⚠️ พบแจ้งเตือนเอกสารซ้ำ (แดง) → ย้ายไฟล์ไปยังโฟลเดอร์ เอกสารซ้ำรอตรวจ')
                            try:
                                if getattr(self, 'latest_created_file_path', None):
                                    moved_to = self.file_manager.move_file_to_duplicate_folder(self.latest_created_file_path)
                                    if moved_to:
                                        print(f"✅ ย้ายไฟล์ไปยังโฟลเดอร์เอกสารซ้ำรอตรวจแล้ว: {moved_to}")
                                    else:
                                        print("❌ ย้ายไฟล์ไปยังโฟลเดอร์เอกสารซ้ำรอตรวจไม่สำเร็จ")
                                else:
                                    print("⚠️ ไม่มีพาธไฟล์ล่าสุดให้ย้าย (latest_created_file_path ว่าง)")
                            except Exception as move_err:
                                print(f"⚠️ ย้ายไฟล์ไปโฟลเดอร์เอกสารซ้ำรอตรวจล้มเหลว: {move_err}")
                            return True
                except Exception:
                    pass
                
                # ตรวจสอบสถานะการประมวลผล
                try:
                    # หา element ที่แสดงสถานะ
                    status_element = self.page.locator('.status, .progress, [data-status]')
                    
                    if status_element.count() > 0:
                        status_text = status_element.first.text_content()
                        
                        if 'complete' in status_text.lower() or 'เสร็จสิ้น' in status_text:
                            print("Processing completed")
                            return True
                        elif 'error' in status_text.lower() or 'ผิดพลาด' in status_text:
                            print("Processing error detected")
                            return False
                    
                    # ตรวจสอบ URL ว่ามีการเปลี่ยนหรือไม่
                    current_url = self.page.url
                    if 'success' in current_url.lower() or 'complete' in current_url.lower():
                        print("Processing completed (URL change detected)")
                        return True
                    
                except Exception as e:
                    print(f"Error checking status: {e}")
                
                time.sleep(0.5)
            
            print("Processing timeout")
            return False
            
        except Exception as e:
            print(f"Error waiting for processing: {e}")
            return False
    
    def create_new_pdf_and_rename(self, pdf_data: Dict) -> bool:
        """สร้างไฟล์ PDF ใหม่และเปลี่ยนชื่อไฟล์ตามรูปแบบที่กำหนด"""
        try:
            print(f"📄 กำลังสร้างไฟล์ PDF ใหม่...")
            
            # 1. ดึงข้อมูลจาก h3 element
            print(f"🔍 กำลังดึงข้อมูลจาก h3 element...")
            h3_element = self.page.locator('//*[@id="content"]/div[6]/div[1]/div[3]/div[1]/h3/text()')
            
            if h3_element.count() > 0:
                h3_text = h3_element.first.text_content()
                print(f"📋 ข้อมูลจาก h3: '{h3_text}'")
                
                # 2. ใช้ FileManager สร้างไฟล์ PDF ใหม่
                new_pdf_filename = self.file_manager.create_pdf_from_form_data(pdf_data, h3_text)
                
                if new_pdf_filename:
                    print(f"✅ สร้างไฟล์ PDF ใหม่สำเร็จ: {new_pdf_filename}")
                    
                    # 3. อัปโหลดไฟล์ใหม่
                    print(f"📤 กำลังอัปโหลดไฟล์ใหม่...")
                    if self.upload_file(new_pdf_filename):
                        print(f"✅ อัปโหลดไฟล์ใหม่สำเร็จ: {new_pdf_filename}")
                        
                        # 4. ลบไฟล์ชั่วคราว
                        self.file_manager.cleanup_temp_files()
                        
                        return True
                    else:
                        print(f"❌ ไม่สามารถอัปโหลดไฟล์ใหม่ได้")
                        return False
                else:
                    print(f"❌ ไม่สามารถสร้างไฟล์ PDF ใหม่ได้")
                    return False
                    
            else:
                print(f"⚠️ ไม่พบ h3 element")
                return False
                
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการสร้างไฟล์ PDF ใหม่: {e}")
            return False
    
    def create_and_upload_pdf(self, pdf_data: Dict) -> bool:
        """สร้างไฟล์ PDF ใหม่และอัปโหลด"""
        try:
            print(f"📄 เริ่มกระบวนการสร้างและอัปโหลดไฟล์ PDF...")
            
            # 1. สร้างไฟล์ PDF ใหม่
            if self.create_new_pdf_and_rename(pdf_data):
                print(f"✅ สร้างไฟล์ PDF ใหม่สำเร็จ")
                return True
            else:
                print(f"❌ ไม่สามารถสร้างไฟล์ PDF ใหม่ได้")
                return False
                
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการสร้างและอัปโหลดไฟล์ PDF: {e}")
            return False
    
    def _run_async(self, coro):
        """Helper method สำหรับรัน async function ใน sync context"""
        if not self._is_async_mode:
            return coro
        
        try:
            loop = self._loop or asyncio.get_event_loop()
            if loop.is_running():
                # ถ้า loop กำลังรันอยู่ ใช้ ThreadPoolExecutor
                import concurrent.futures
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    future = executor.submit(asyncio.run, coro)
                    return future.result(timeout=30)
            else:
                return loop.run_until_complete(coro)
        except RuntimeError:
            return asyncio.run(coro)
    
    def close_driver(self):
        """ปิด Playwright"""
        try:
            if self._is_async_mode:
                # ใช้ async mode
                async def close_async():
                    try:
                        # Cancel pending tasks ก่อน
                        try:
                            loop = asyncio.get_event_loop()
                            # หา pending tasks ที่เกี่ยวข้องกับ Playwright
                            pending_tasks = [t for t in asyncio.all_tasks(loop) if not t.done()]
                            for task in pending_tasks:
                                task_name = getattr(task, 'get_name', lambda: '')()
                                if 'Connection.run' in str(task) or 'playwright' in str(task).lower():
                                    try:
                                        task.cancel()
                                    except:
                                        pass
                        except Exception as e:
                            print(f"⚠️ Error canceling tasks: {e}")
                        
                        # ปิด page ก่อน (ถ้ามี)
                        if self._async_page:
                            try:
                                # ปิด page context แทน page โดยตรง
                                context = getattr(self._async_page, 'context', None)
                                if context:
                                    await context.close()
                                else:
                                    await self._async_page.close()
                            except Exception as e:
                                print(f"⚠️ Error closing page: {e}")
                        
                        # รอให้ pending operations เสร็จ
                        await asyncio.sleep(0.3)
                        
                        # ปิด browser (จะปิด contexts ทั้งหมดด้วย)
                        if self._async_browser:
                            try:
                                # ปิด browser contexts ทั้งหมดก่อน
                                contexts = self._async_browser.contexts
                                for context in contexts:
                                    try:
                                        await context.close()
                                    except:
                                        pass
                                
                                await asyncio.sleep(0.2)
                                
                                # ปิด browser
                                await self._async_browser.close()
                            except Exception as e:
                                print(f"⚠️ Error closing browser: {e}")
                        
                        # รอให้ pending operations เสร็จ
                        await asyncio.sleep(0.3)
                        
                        # ปิด playwright สุดท้าย
                        if self._async_playwright:
                            try:
                                await self._async_playwright.stop()
                            except Exception as e:
                                print(f"⚠️ Error stopping playwright: {e}")
                        
                        # รอให้ cleanup เสร็จ
                        await asyncio.sleep(0.2)
                    except Exception as e:
                        print(f"⚠️ Error in close_async: {e}")
                
                try:
                    # ใช้ timeout เพื่อป้องกันการค้าง
                    import concurrent.futures
                    with concurrent.futures.ThreadPoolExecutor() as executor:
                        future = executor.submit(asyncio.run, close_async())
                        try:
                            future.result(timeout=5)  # timeout 5 วินาที
                        except concurrent.futures.TimeoutError:
                            print("⚠️ Timeout closing Playwright, forcing cleanup...")
                            # Force cleanup
                            self._async_page = None
                            self._async_browser = None
                            self._async_playwright = None
                        except asyncio.CancelledError:
                            print("⚠️ close_async cancelled because event loop is closing")
                        except Exception as e:
                            print(f"⚠️ close_async terminated with error: {e}")
                except Exception as e:
                    print(f"⚠️ Error running close_async: {e}")
                    # ลองปิดแบบ force ถ้า async ไม่สำเร็จ
                    try:
                        self._async_page = None
                        self._async_browser = None
                        self._async_playwright = None
                    except:
                        pass
            else:
                # ใช้ sync mode
                if self.page:
                    try:
                        self.page.close()
                    except Exception as e:
                        print(f"⚠️ Error closing page: {e}")
                if self.browser:
                    try:
                        self.browser.close()
                    except Exception as e:
                        print(f"⚠️ Error closing browser: {e}")
                if self.playwright:
                    try:
                        self.playwright.stop()
                    except Exception as e:
                        print(f"⚠️ Error stopping playwright: {e}")
            
            # Reset flags
            self._is_async_mode = False
            self._async_page = None
            self._async_browser = None
            self._async_playwright = None
            self.page = None
            self.browser = None
            self.playwright = None
            
            print("Playwright closed")
        except Exception as e:
            print(f"Error closing Playwright: {e}")
    
    def is_page_loaded(self) -> bool:
        """ตรวจสอบว่าเพจโหลดเสร็จแล้วหรือไม่"""
        try:
            return self.page.evaluate("document.readyState") == "complete"
        except:
            return False
    
    def take_screenshot(self, filename: str = "screenshot.png"):
        """ถ่ายภาพหน้าจอ"""
        try:
            if self.page:
                self.page.screenshot(path=filename)
                print(f"Screenshot saved: {filename}")
                return True
        except Exception as e:
            print(f"Error taking screenshot: {e}")
        return False
    
    def wait_for_element(self, selector: str, timeout: int = 30) -> bool:
        """รอให้ element ปรากฏ"""
        try:
            self.page.wait_for_selector(selector, timeout=timeout * 1000)
            return True
        except Exception as e:
            print(f"Element {selector} not found within {timeout} seconds")
            return False
    
    def click_element(self, selector: str) -> bool:
        """คลิก element"""
        try:
            element = self.page.locator(selector)
            if element.count() > 0:
                element.first.click()
                return True
            return False
        except Exception as e:
            print(f"Error clicking element {selector}: {e}")
            return False
    
    def get_page_content(self) -> str:
        """ดึงเนื้อหาของหน้าเว็บ"""
        try:
            return self.page.content()
        except Exception as e:
            print(f"Error getting page content: {e}")
            return ""
    
    def execute_script(self, script: str):
        """รัน JavaScript"""
        try:
            return self.page.evaluate(script)
        except Exception as e:
            print(f"Error executing script: {e}")
            return None
    
    def peak_engine_workflow(self, pdf_data_list: list, start_index: int = 0):
        """ทำงานกับ PeakEngine ตามข้อมูลจาก PDF"""
        try:
            if not self.page:
                print("❌ Playwright page not initialized")
                return False
            
            print(f"🚀 เริ่มต้นการทำงานกับ PeakEngine...")
            print(f"📊 ข้อมูล PDF ที่จะประมวลผล: {len(pdf_data_list)} ไฟล์")
            
            # 1. เข้าหน้า Login
            print(f"🔐 เข้าหน้า Login...")
            self.page.goto("https://secure.peakengine.com/Home/Login")
            
            
            # 2. ล็อกอิน (ปรับให้ตรงกับโครงสร้างเว็บไซต์ PeakEngine)
            print(f"🔑 กำลังล็อกอิน...")
            try:
                # หาและกรอกอีเมล (Username)
                print(f"📧 กำลังกรอกอีเมล...")
                email_field = self.page.locator('input[placeholder="อีเมล"], input[type="email"], input[name="email"]')
                if email_field.count() > 0:
                    email_field.first.fill(self.credentials.get('Username', ''))
                    print(f"✅ กรอกอีเมลสำเร็จ: {self.credentials.get('Username', '')}")
                else:
                    print(f"⚠️ ไม่พบฟิลด์อีเมล")
                    return False
                
                # หาและกรอกรหัสผ่าน (Password)
                print(f"🔒 กำลังกรอกรหัสผ่าน...")
                password_field = self.page.locator('input[placeholder="รหัสผ่าน"], input[type="password"], input[name="password"]')
                if password_field.count() > 0:
                    password_field.first.fill(self.credentials.get('Password', ''))
                    print(f"✅ กรอกรหัสผ่านสำเร็จ")
                else:
                    print(f"⚠️ ไม่พบฟิลด์รหัสผ่าน")
                    return False
                
                # กดปุ่ม "เข้าใช้งาน" (Login) - ใช้ ID ที่ระบุ
                print(f"🚀 กำลังกดปุ่มเข้าใช้งาน...")
                # หาปุ่มด้วย ID loginbtn ที่ระบุ
                login_button = self.page.locator('#loginbtn, button#loginbtn, input#loginbtn')
                if login_button.count() > 0:
                    login_button.first.click()
                    print(f"✅ กดปุ่มเข้าใช้งานสำเร็จ (ID: loginbtn)")
                else:
                    # ถ้าไม่พบ ID ให้ลองหาด้วยวิธีอื่น
                    print(f"🔍 ไม่พบ ID loginbtn ลองหาด้วยวิธีอื่น...")
                    login_button = self.page.locator('button:has-text("เข้าใช้งาน"), button:has-text("Login"), button[type="submit"], button.teal, button[class*="teal"], button:has-text("เข้าสู่ระบบ")')
                    if login_button.count() > 0:
                        login_button.first.click()
                        print(f"✅ กดปุ่มเข้าใช้งานสำเร็จ")
                    else:
                        # ถ้าไม่พบปุ่ม ให้ลองหาด้วย text content
                        print(f"🔍 ลองหาปุ่มด้วยวิธีอื่น...")
                        all_buttons = self.page.locator('button')
                        for i in range(all_buttons.count()):
                            button_text = all_buttons.nth(i).text_content()
                            print(f"   ปุ่มที่ {i+1}: '{button_text}'")
                            if "เข้าใช้งาน" in button_text or "Login" in button_text or "เข้าสู่ระบบ" in button_text:
                                all_buttons.nth(i).click()
                                print(f"✅ กดปุ่มสำเร็จ: '{button_text}'")
                                break
                        else:
                            print(f"❌ ไม่พบปุ่มเข้าใช้งาน")
                            return False
                                    # รอให้ล็อกอินเสร็จ
                print(f"⏳ รอการล็อกอิน...")
                
                try:
                    # รอให้ network idle และ redirect เสร็จสิ้น
                    self.page.wait_for_load_state('networkidle', timeout=10000)
                    
                    
                    # รอให้ redirect เสร็จสิ้น (เพิ่มการรอ)
                    print(f"🔄 รอการ redirect...")
                    max_wait = 10  # รอสูงสุด 10 วินาที
                    wait_count = 0
                    
                    while wait_count < max_wait:
                        current_url = self.page.url
                        if current_url != "https://secure.peakengine.com/Home/Login":
                            print(f"✅ Redirect เสร็จสิ้น: {current_url}")
                            break
                        time.sleep(0.5)
                        wait_count += 1
                        print(f"   รอ redirect... {wait_count}/{max_wait}")
                    
                    # ตรวจสอบว่าล็อกอินสำเร็จหรือไม่
                    current_url = self.page.url
                    print(f"📍 URL ปัจจุบัน: {current_url}")
                    
                    # ตรวจสอบหลายเงื่อนไข
                    print(f"🔍 ตรวจสอบเงื่อนไขการล็อกอิน...")
                    print(f"   URL ปัจจุบัน: {current_url}")
                    print(f"   มี 'login' ใน URL: {'login' in current_url.lower()}")
                    print(f"   มี 'home/login' ใน URL: {'home/login' in current_url.lower()}")
                    print(f"   เท่ากับ URL เริ่มต้น: {current_url == 'https://secure.peakengine.com/Home/Login'}")
                    
                    # ตรวจสอบว่าล็อกอินสำเร็จหรือไม่ (ปรับเงื่อนไขให้ยืดหยุ่นขึ้น)
                    login_success = False
                    
                    # เงื่อนไขที่ 1: URL เปลี่ยนไปแล้ว
                    if ("login" not in current_url.lower() and 
                        "home/login" not in current_url.lower() and
                        current_url != "https://secure.peakengine.com/Home/Login"):
                        login_success = True
                        print(f"✅ ล็อกอินสำเร็จ (เงื่อนไขที่ 1): URL เปลี่ยนไปแล้ว")
                    
                    # เงื่อนไขที่ 2: ไม่มีฟิลด์ login อยู่แล้ว
                    try:
                        login_fields = self.page.locator('input[type="email"], input[type="password"]')
                        if login_fields.count() == 0:
                            login_success = True
                            print(f"✅ ล็อกอินสำเร็จ (เงื่อนไขที่ 2): ไม่มีฟิลด์ login อยู่แล้ว")
                    except:
                        pass
                    
                    # เงื่อนไขที่ 3: มีข้อความแสดงว่าล็อกอินสำเร็จ
                    try:
                        success_messages = self.page.locator('text=ยินดีด้วย, text=พร้อมเริ่มต้น, text=PEAK')
                        if success_messages.count() > 0:
                            login_success = True
                            print(f"✅ ล็อกอินสำเร็จ (เงื่อนไขที่ 3): พบข้อความแสดงความสำเร็จ")
                    except:
                        pass
                    
                    if login_success:
                        self.is_logged_in = True
                        print(f"✅ ล็อกอินสำเร็จ! ย้ายไปหน้า: {current_url}")
                        
                        # หลังจากล็อกอินสำเร็จ ให้ไปที่ลิงค์ Peak เก่าโดยตรง
                        print(f"🔄 กำลังไปยัง Peak เก่า...")
                        try:
                            # ไปที่ลิงค์ Peak เก่าโดยตรง
                            old_peak_url = "https://secure.peakengine.com/selectlist"
                            print(f"📍 ไปยังลิงค์: {old_peak_url}")
                            
                            self.page.goto(old_peak_url, timeout=10000)  # ลด timeout เป็น 10 วินาที
                            time.sleep(0.5)  # รอแค่ 0.5 วินาที
                            
                            
                            # ตรวจสอบ URL หลังไปยังลิงค์
                            current_url = self.page.url
                            print(f"📍 URL หลังไปยัง Peak เก่า: {current_url}")
                            
                            if "selectlist" in current_url.lower():
                                print(f"✅ ไปยัง Peak เก่าสำเร็จ!")
                                
                                # ต่อไปให้ไปที่ Link Company
                                if self.company_link:
                                    print(f"🏢 ไปยังหน้า Company: {self.company_link}")
                                    self.page.goto(self.company_link)
                                    
                                    
                                    print(f"✅ ไปยังหน้า Company สำเร็จ!")
                                    
                                    # ต่อไปให้ไปที่ Link Express
                                    if self.express_link:
                                        print(f"📝 ไปยังหน้า Express: {self.express_link}")
                                        self.page.goto(self.express_link)
                                        
                                        
                                        print(f"✅ ไปยังหน้า Express สำเร็จ!")
                                        
                                        # ตรวจสอบ URL สุดท้าย
                                        final_url = self.page.url
                                        print(f"📍 URL สุดท้าย: {final_url}")
                                        print(f"🎯 ระบบพร้อมทำงานกับข้อมูล PDF แล้ว!")
                                        
                                        # เริ่มทำงานกับข้อมูล PDF ทันที
                                        print(f"🚀 เริ่มทำงานกับข้อมูล PDF...")
                                        # ไม่ return ออกมา ให้ทำงานต่อใน workflow
                                        pdf_processing_result = self.process_pdf_data(pdf_data_list, start_index=start_index)
                                        print(f"📊 ผลการประมวลผล PDF: {'สำเร็จ' if pdf_processing_result else 'ไม่สำเร็จ'}")
                                        return True
                                    else:
                                        print(f"❌ ไม่พบ Express Link")
                                else:
                                    print(f"❌ ไม่พบ Company Link")
                            else:
                                print(f"⚠️ ไปยัง Peak เก่าไม่สำเร็จ")
                                
                        except Exception as e:
                            print(f"⚠️ เกิดข้อผิดพลาดในการไปยัง Peak เก่า: {e}")
                            print(f"🔍 ลองใช้วิธีอื่น...")
                            try:
                                # ลองใช้ JavaScript navigate
                                self.page.evaluate("window.location.href = 'https://secure.peakengine.com/selectlist'")
                                print(f"✅ ไปยัง Peak เก่าสำเร็จด้วย JavaScript")
                                
                                # รอให้หน้าโหลดเสร็จ
                                
                                
                                
                                # ตรวจสอบ URL หลังไปยังลิงค์
                                current_url = self.page.url
                                print(f"📍 URL หลังไปยัง Peak เก่า: {current_url}")
                                
                                # ต่อไปให้ไปที่ Link Company
                                if self.company_link:
                                    print(f"🏢 ไปยังหน้า Company: {self.company_link}")
                                    self.page.goto(self.company_link)
                                    
                                    
                                    print(f"✅ ไปยังหน้า Company สำเร็จ!")
                                    
                                    # ต่อไปให้ไปที่ Link Express
                                    if self.express_link:
                                        print(f"📝 ไปยังหน้า Express: {self.express_link}")
                                        self.page.goto(self.express_link)
                                        
                                        
                                        print(f"✅ ไปยังหน้า Express สำเร็จ!")
                                        
                                        # ตรวจสอบ URL สุดท้าย
                                        final_url = self.page.url
                                        print(f"📍 URL สุดท้าย: {final_url}")
                                        print(f"🎯 ระบบพร้อมทำงานกับข้อมูล PDF แล้ว!")
                                    else:
                                        print(f"❌ ไม่พบ Express Link")
                                else:
                                    print(f"❌ ไม่พบ Company Link")
                                
                            except Exception as e2:
                                print(f"❌ ไม่สามารถไปยัง Peak เก่าได้ด้วยวิธีใดเลย: {e2}")
                        
                        return True
                    else:
                        print(f"⚠️ เงื่อนไขการล็อกอินไม่ผ่าน - ยังอยู่ที่หน้า login")
                        # ลองตรวจสอบ error message หรือ validation
                        print(f"🔍 ตรวจสอบข้อผิดพลาด...")
                        try:
                            error_messages = self.page.locator('.error, .alert, .message, [class*="error"], [class*="alert"]')
                            if error_messages.count() > 0:
                                error_text = error_messages.first.text_content()
                                print(f"⚠️ พบข้อผิดพลาด: {error_text}")
                        except:
                            print(f"⚠️ ไม่สามารถตรวจสอบ error message ได้")
                        
                        # ตรวจสอบว่ายังมีฟิลด์ login อยู่หรือไม่
                        try:
                            login_fields = self.page.locator('input[type="email"], input[type="password"]')
                            if login_fields.count() > 0:
                                print(f"⚠️ ยังมีฟิลด์ login อยู่ - ล็อกอินไม่สำเร็จ")
                            else:
                                print(f"✅ ไม่มีฟิลด์ login - ล็อกอินอาจสำเร็จ")
                                self.is_logged_in = True
                                return True
                        except:
                            print(f"⚠️ ไม่สามารถตรวจสอบฟิลด์ login ได้")
                        
                        print(f"⚠️ ล็อกอินไม่สำเร็จ - ยังอยู่ที่หน้า login")
                        return False
                        
                except Exception as e:
                    print(f"⚠️ เกิดข้อผิดพลาดในการตรวจสอบล็อกอิน: {e}")
                    # ลองตรวจสอบ URL อีกครั้ง
                    try:
                        current_url = self.page.url
                        print(f"📍 URL หลัง error: {current_url}")
                        if "login" not in current_url.lower():
                            self.is_logged_in = True
                            print(f"✅ ล็อกอินสำเร็จ! (ตรวจสอบจาก URL)")
                            return True
                    except:
                        pass
                    
                    return False
                    
            except Exception as e:
                print(f"❌ เกิดข้อผิดพลาดในการล็อกอิน: {e}")
                return False
            
            print(f"✅ ล็อกอินสำเร็จ!")
            
            # 4. ข้ามขั้นตอนการไปยัง Company และ Express (ทำไปแล้วในขั้นตอนการล็อกอิน)
            print(f"✅ ข้ามขั้นตอนการไปยัง Company และ Express (ทำไปแล้ว)")
            
            # 5. ข้อมูล PDF จะถูกประมวลผลใน process_pdf_data method
            print(f"✅ ระบบพร้อมทำงานกับข้อมูล PDF แล้ว!")
            return True
            
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการทำงานกับ PeakEngine: {e}")
            try:
                self._playwright_error = str(e)
            except Exception:
                self._playwright_error = None
            return False
    
    def get_peak_engine_status(self):
        """ตรวจสอบสถานะการทำงานของ PeakEngine"""
        try:
            if not self.page:
                return "ไม่มีการเชื่อมต่อ"
            
            current_url = self.page.url
            page_title = self.page.title()
            
            return {
                "current_url": current_url,
                "page_title": page_title,
                "is_logged_in": self.is_logged_in,
                "browser_status": "เปิดใช้งาน" if self.browser else "ปิด"
            }
            
        except Exception as e:
            return f"เกิดข้อผิดพลาด: {e}"
    
    def process_pdf_data(self, pdf_data_list: list, start_index: int = 0) -> bool:
        """ประมวลผลข้อมูล PDF ทันทีหลังจากไปยัง Express Link สำเร็จ"""
        try:
            # จำกัดจำนวนการประมวลผลสูงสุด (สามารถปรับได้จาก config)
            max_items = getattr(Config, 'MAX_PROCESSING_ITEMS', 60)
            total_items = len(pdf_data_list)
            
            if total_items > max_items:
                self._log(f"⚠️ จำนวนรายการ ({total_items}) เกินกำหนดสูงสุด ({max_items})", level="warning")
                self._log(f"📊 จะประมวลผลเฉพาะ {max_items} รายการแรก")
                pdf_data_list = pdf_data_list[:max_items]
            
            self._log(f"📊 เริ่มประมวลผลข้อมูล PDF: {len(pdf_data_list)} ไฟล์ (สูงสุด {max_items} รายการ)")
            self._status_update(step="กำลังประมวลผล PDF", file='-', folder=None)
            
            # ประมวลผลข้อมูล PDF แต่ละไฟล์
            for i, pdf_data in enumerate(pdf_data_list, 1):
                if i < max(1, int(start_index)):
                    continue
                # แสดงความคืบหน้าการประมวลผล
                progress_percent = (i / len(pdf_data_list)) * 100
                current_filename = pdf_data.get('filename', 'ไม่ทราบชื่อ')
                self._log(f"📊 ความคืบหน้า: {i}/{len(pdf_data_list)} ({progress_percent:.1f}%) - {current_filename}")
                self._status_update(
                    step=f"กำลังประมวลผล ({i}/{len(pdf_data_list)})",
                    file=current_filename,
                    folder=None
                )
                
                # แจ้งเตือนเมื่อใกล้ครบ 60 รายการ
                if i >= max_items - 5:
                    remaining = max_items - i + 1
                    self._log(f"⚠️ เหลืออีก {remaining} รายการ จะสิ้นสุดการประมวลผล", level="warning")
                
                # รีหน้า Express ทุกครั้งก่อนเริ่มไฟล์ใหม่ เพื่อความสะอาดของสถานะ
                self.refresh_express_page()
                self._log(f"\n ประมวลผลไฟล์ที่ {i}: {current_filename}")
                
                # บันทึกข้อมูลปัจจุบันสำหรับ retry กรณีเจอแจ้งเตือนบังคับกรอก
                self._current_pdf_data_for_retry = pdf_data
                self._refill_attempt_count = 0
                
                # 0. ระบุ group จาก folder_settings + folder_code และบันทึกลง pdf_data
                try:
                    folder_settings = self.file_manager.read_folder_settings()
                    source_folder = pdf_data.get('source_folder') or os.path.dirname(pdf_data.get('file_path', '') or pdf_data.get('pdf_path', '') or '')
                    folder_code = self.file_manager.get_folder_code_from_path(Path(source_folder)) if source_folder else None
                    if folder_code and 'folder_code' not in pdf_data:
                        pdf_data['folder_code'] = folder_code
                    group = 'unknown'
                    if folder_code and folder_code in folder_settings:
                        group = folder_settings[folder_code].get('group', 'unknown')
                    pdf_data['group'] = group
                    # คำนวณชื่อบริการจาก mapping เพื่อใช้แสดงผล
                    try:
                        comp = pdf_data.get('company_name', '')
                        service_name = self.file_manager._get_service_name(comp, folder_settings, pdf_data)
                    except Exception:
                        service_name = ''
                    if service_name:
                        pdf_data['service_name'] = service_name
                    self._log(f"🗂️ จัดกลุ่มไฟล์: folder_code='{folder_code or ''}', group='{group}'")
                except Exception as _:
                    group = pdf_data.get('group') or 'unknown'

                # 1. กรอกข้อมูลในฟอร์ม (กรณี NoneVat จะไม่กรอกเลขที่เอกสาร ตาม pdf_data['group'])
                if self.fill_form_data(pdf_data):
                    self._log(f"✅ กรอกข้อมูลสำเร็จ")
                    
                    # 2. รอผลบันทึก/ยืนยัน (อนุมัติ หรือ เอกสารซ้ำ)
                    self._log(f"⏳ รอผลบันทึก/ยืนยัน...")
                    result = self.wait_for_save_result(timeout=20)
                    self._log(f"📊 ผลการบันทึก: {result}")
                    
                    if result == 'duplicate':
                        # อัปเดตตัวนับในรายงาน: เอกสารซ้ำ
                        try:
                            rm = get_global_report_manager()
                            if rm:
                                rm.add_duplicate(1)
                        except Exception:
                            pass
                        self._notify_progress(duplicate_delta=1, failure_delta=1)
                        # ย้ายไฟล์ต้นฉบับไปโฟลเดอร์ เอกสารซ้ำรอตรวจ แล้วไปไฟล์ถัดไป
                        original_path = pdf_data.get('file_path') or pdf_data.get('pdf_path')
                        if original_path:
                            moved = self.file_manager.move_file_to_duplicate_folder(original_path)
                            if moved:
                                self._log(f"✅ ย้ายไฟล์ต้นฉบับไปยัง 'เอกสารซ้ำรอตรวจ': {moved}")
                            else:
                                self._log("❌ ย้ายไฟล์ต้นฉบับไปยังโฟลเดอร์เอกสารซ้ำรอตรวจไม่สำเร็จ", level="warning")
                        else:
                            self._log("⚠️ ไม่พบพาธไฟล์ต้นฉบับใน pdf_data", level="warning")
                        time.sleep(0.2)
                        self._status_update(step="เอกสารซ้ำ ข้ามไฟล์", file=current_filename, folder=None)
                        continue
                    elif result == 'timeout':
                        self._log("⚠️ ไม่พบผลบันทึกภายในเวลา 20 วินาที", level="warning")
                        self._log("🔍 ตรวจสอบสถานะหน้าเว็บ...")
                        
                        # ตรวจสอบว่าหน้าเว็บยังทำงานอยู่หรือไม่
                        try:
                            # ตรวจสอบว่ามี error message หรือไม่
                            error_elements = [
                                '//*[@id="dvredalert"]',
                                '//*[@id="dvgreenalert"]',
                                '[class*="error"]',
                                '[class*="alert"]'
                            ]
                            
                            found_error = False
                            for selector in error_elements:
                                try:
                                    element = self.page.locator(selector)
                                    if element.count() > 0 and element.first.is_visible():
                                        text = element.first.text_content() or ''
                                        if text.strip():
                                            self._log(f"🔍 พบข้อความ: {text.strip()}")
                                            found_error = True
                                except Exception:
                                    pass
                            
                            if not found_error:
                                self._log("ℹ️ ไม่พบ alert box หรือ error message")
                                self._log("🔄 ลองรีเฟรชหน้าเว็บ...")
                                self.refresh_express_page()
                                time.sleep(2)
                        except Exception as e:
                            self._log(f"⚠️ ตรวจสอบสถานะหน้าเว็บไม่สำเร็จ: {e}", level="warning")
                        
                        self._log("⏭️ ข้ามไฟล์นี้ไปไฟล์ถัดไป", level="warning")
                        time.sleep(0.2)
                        self._notify_progress(failure_delta=1)
                        self._status_update(step="เกิดปัญหา ข้ามไฟล์", file=current_filename, folder=None)
                        continue
                    # result == 'approved' → ค่อยสร้างไฟล์ใหม่และอัปโหลด
                    try:
                        rm = get_global_report_manager()
                        if rm:
                            rm.add_processed_success(1)
                    except Exception:
                        pass
                    
                    # 2.x สำหรับบริษัท VAT เท่านั้นจึงตรวจ/แก้ VAT
                    company_name = pdf_data.get('company_name', '')
                    company_vat_status = Config.COMPANY_VAT_STATUS.get(company_name, 'VAT')
                    folder_group = pdf_data.get('group') or group or 'unknown'
                    
                    print(f"🔍 ตรวจสอบเงื่อนไข: folder_group='{folder_group}', company_vat_status='{company_vat_status}'")
                    
                    # ให้ความสำคัญกับ folder_group ก่อน
                    if folder_group == 'special' or company_vat_status == 'NoneVat':
                        self._log(f"ℹ️ โฟลเดอร์ special (NoneVat): ข้ามขั้นตอนตรวจ/แก้ไข VAT")
                    elif company_vat_status == 'VAT':
                        # 2.1 หน่วง 1 วิ และเลื่อนลงก่อนตรวจ VAT (หลังอนุมัติสำเร็จ)
                        self._log(f"ℹ️ บริษัท {company_name}: ทำงานแบบ VAT - ตรวจ/แก้ไข VAT")
                        time.sleep(1)
                        try:
                            self.page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                        except Exception:
                            pass
                        # 2.2 ตรวจและแก้ไข VAT ถ้าจำเป็น (ก่อนสร้างไฟล์/อัปโหลด)
                        try:
                            vat_ok = self.check_and_fix_vat_value(str(pdf_data.get('vat_value', '')))
                            if not vat_ok:
                                self._log('❌ แก้ไข VAT ไม่สำเร็จ ข้ามไฟล์นี้เพื่อป้องกันข้อมูลผิดพลาด', level="warning")
                                time.sleep(0.2)
                                self._notify_progress(failure_delta=1)
                                self._status_update(step="แก้ไข VAT ไม่สำเร็จ", file=current_filename, folder=None)
                                continue
                        except Exception as _:
                            self._log('⚠️ ข้ามขั้นตอนตรวจ/แก้ไข VAT เนื่องจากเกิดข้อยกเว้น', level="warning")
                            time.sleep(0.2)
                            self._notify_progress(failure_delta=1)
                            self._status_update(step="ข้ามขั้นตอนตรวจ VAT", file=current_filename, folder=None)
                            continue
                    else:
                        self._log(f"ℹ️ บริษัท {company_name}: ทำงานแบบ NoneVat - ข้ามขั้นตอนตรวจ/แก้ไข VAT")
                    
                    # 3. ตรวจสอบ iptnumber จากระบบก่อนสร้างไฟล์ (พร้อม retry และรีเฟรชหน้าเว็บเมื่อค้าง)
                    max_retries = 3
                    upload_success = False
                    
                    for retry_attempt in range(max_retries):
                        if retry_attempt > 0:
                            self._log(f"🔄 ลองใหม่ครั้งที่ {retry_attempt + 1}/{max_retries}...")
                            # ตรวจสอบว่าหน้าเว็บค้างหรือไม่
                            if self.check_page_stuck(timeout=3):
                                self._log(f"⚠️ ตรวจพบว่าหน้าเว็บค้าง กำลังรีเฟรชหน้าเว็บ...")
                                self.refresh_page()
                                time.sleep(2)
                        
                        self._log(f"🔍 กำลังตรวจสอบ iptnumber จากระบบก่อนสร้างไฟล์...")
                        verified_iptnumber = self.verify_and_get_iptnumber()
                        
                        if not verified_iptnumber:
                            self._log(f"❌ ไม่สามารถตรวจสอบ iptnumber จากระบบได้ หรือ iptnumber ยังไม่พร้อม", level="warning")
                            if retry_attempt < max_retries - 1:
                                self._log(f"⏳ รอสักครู่แล้วลองใหม่...")
                                time.sleep(2)
                                continue
                            else:
                                self._log(f"❌ ไม่สามารถตรวจสอบ iptnumber ได้หลังจากลอง {max_retries} ครั้ง ข้ามไฟล์นี้", level="warning")
                                break
                        
                        self._log(f"✅ ตรวจสอบ iptnumber สำเร็จ: '{verified_iptnumber}' พร้อมสร้างไฟล์")
                        
                        # 4. สร้างไฟล์ PDF ใหม่จากข้อมูลที่กรอกในฟอร์ม (ใช้ iptnumber ที่ตรวจสอบแล้ว)
                        self._log(f" เริ่มสร้างไฟล์ใหม่จากข้อมูลที่กรอก...")
                        new_pdf_filename = self.create_new_pdf_from_form_data(pdf_data, verified_iptnumber)
                        
                        if not new_pdf_filename or not os.path.exists(new_pdf_filename):
                            self._log(f"❌ ไม่สามารถสร้างไฟล์ใหม่ได้", level="warning")
                            if retry_attempt < max_retries - 1:
                                self._log(f"⏳ รอสักครู่แล้วลองใหม่...")
                                time.sleep(2)
                                continue
                            else:
                                self._log(f"❌ ไม่สามารถสร้างไฟล์ใหม่ได้หลังจากลอง {max_retries} ครั้ง ข้ามไฟล์นี้", level="warning")
                                break
                        
                        self._log(f"✅ สร้างไฟล์ใหม่สำเร็จ: {new_pdf_filename}")
                        
                        # 5. โยนไฟล์เข้าเว็บโดยตรง (ไม่ต้องกดปุ่มอัปโหลด)
                        self._log(f"📤 โยนไฟล์เข้าเว็บโดยตรง...")
                        upload_result = self.upload_file_directly(new_pdf_filename, pdf_data)
                        
                        if not upload_result:
                            self._log(f"❌ โยนไฟล์เข้าเว็บไม่สำเร็จ", level="warning")
                            if retry_attempt < max_retries - 1:
                                self._log(f"⏳ รอสักครู่แล้วลองใหม่...")
                                # ตรวจสอบว่าหน้าเว็บค้างหรือไม่
                                if self.check_page_stuck(timeout=3):
                                    self._log(f"⚠️ ตรวจพบว่าหน้าเว็บค้าง กำลังรีเฟรชหน้าเว็บ...")
                                    self.refresh_page()
                                time.sleep(2)
                                continue
                            else:
                                self._log(f"❌ โยนไฟล์เข้าเว็บไม่สำเร็จหลังจากลอง {max_retries} ครั้ง", level="error")
                                break
                        
                        self._log(f"✅ โยนไฟล์เข้าเว็บสำเร็จ: {new_pdf_filename}")
                        
                        # 6. รอการประมวลผลหลังจากโยนไฟล์
                        self._log(f"⏳ รอการประมวลผลหลังจากโยนไฟล์...")
                        processing_result = self.wait_for_processing()
                        
                        if not processing_result:
                            self._log(f"⚠️ การประมวลผลไม่เสร็จสิ้นหรือเกิดข้อผิดพลาด", level="warning")
                            if retry_attempt < max_retries - 1:
                                self._log(f"⏳ รอสักครู่แล้วลองใหม่...")
                                # ตรวจสอบว่าหน้าเว็บค้างหรือไม่
                                if self.check_page_stuck(timeout=3):
                                    self._log(f"⚠️ ตรวจพบว่าหน้าเว็บค้าง กำลังรีเฟรชหน้าเว็บ...")
                                    self.refresh_page()
                                time.sleep(2)
                                continue
                            else:
                                self._log(f"❌ การประมวลผลไม่เสร็จสิ้นหลังจากลอง {max_retries} ครั้ง", level="error")
                                break
                        
                        self._log(f"✅ ประมวลผลเสร็จสิ้น")
                        
                        # 7. ย้ายไฟล์ก่อน: ต้นฉบับ → เอกสารต้นฉบับ, ไฟล์ที่สร้าง/อัปโหลด → เอกสาร Vat/NoneVat
                        try:
                            original_path = pdf_data.get('file_path') or pdf_data.get('pdf_path')
                            processed_path = new_pdf_filename  # new_pdf_filename คือพาธเต็มจาก create_new_pdf_from_form_data
                            
                            # ใช้ folder_group และ company_vat_status ในการกำหนดโฟลเดอร์ปลายทาง
                            # กรณีพิเศษ: folder_group = 'regular' แต่ company_vat_status = 'NoneVat'
                            # → ทำงานแบบ NoneVat แต่ย้ายไฟล์ไป "เอกสาร Vat" (ตามโฟลเดอร์)
                            
                            # ตรวจสอบว่าต้องทำงานแบบ NoneVat หรือไม่
                            is_nonevat_workflow = (folder_group == 'special' or company_vat_status == 'NoneVat')
                            
                            if is_nonevat_workflow:
                                # ถ้าทำงานแบบ NoneVat ให้ย้ายไฟล์ตาม company_vat_status
                                if company_vat_status == 'NoneVat':
                                    move_group = 'special'  # ย้ายไป "เอกสาร NoneVat"
                                else:
                                    move_group = 'regular'  # ย้ายไป "เอกสาร Vat"
                            else:
                                # ถ้าทำงานแบบ VAT ให้ย้ายไฟล์ตาม folder_group
                                if folder_group in ['regular', 'special']:
                                    move_group = folder_group
                                else:
                                    move_group = 'regular'  # fallback
                            
                            self._log(f"🔍 ย้ายไฟล์: folder_group='{folder_group}', company_vat_status='{company_vat_status}', move_group='{move_group}'")
                            self.file_manager.move_original_and_processed(original_path, processed_path, move_group)
                        except Exception as _:
                            self._log('⚠️ ย้ายไฟล์หลังอัปโหลดไม่สำเร็จ', level="warning")
                        
                        # สำเร็จแล้ว - ออกจาก retry loop
                        upload_success = True
                        self._log(f"✅ รายการนี้เสร็จสมบูรณ์: สร้างไฟล์และอัปโหลดสำเร็จ")
                        self._status_update(step="สำเร็จ", file=current_filename, folder=None)
                        self._notify_progress(success_delta=1)
                        break
                    
                    if not upload_success:
                        self._log(f"❌ รายการนี้ไม่สำเร็จ: ไม่สามารถสร้างไฟล์หรืออัปโหลดได้หลังจากลอง {max_retries} ครั้ง", level="error")
                        self._notify_progress(failure_delta=1)
                        # ตรวจสอบว่าหน้าเว็บค้างหรือไม่ และรีเฟรช
                        if self.check_page_stuck(timeout=3):
                            self._log(f"⚠️ ตรวจพบว่าหน้าเว็บค้าง กำลังรีเฟรชหน้าเว็บ...")
                            self.refresh_page()
                    
                    # รอสักครู่ก่อนไฟล์ถัดไป
                    time.sleep(0.5)
                else:
                    self._log(f"❌ กรอกข้อมูลไม่สำเร็จ", level="warning")
                    self._notify_progress(failure_delta=1)
                    self._status_update(step="กรอกข้อมูลไม่สำเร็จ", file=current_filename, folder=None)
            
            self._log(f"\n🎯 สรุป: ประมวลผลข้อมูล PDF {len(pdf_data_list)} ไฟล์ เสร็จสิ้น!")
            self._status_update(step="ประมวลผลครบถ้วน", file='-', folder=None)
            
            # แจ้งเตือนถ้าจำนวนรายการเกิน 60
            if total_items > max_items:
                remaining_items = total_items - max_items
                self._log(f"⚠️ หมายเหตุ: ยังมีรายการที่เหลืออีก {remaining_items} รายการ", level="warning")
                self._log(f"💡 แนะนำ: รันระบบใหม่เพื่อประมวลผลรายการที่เหลือ")
            
            return True
            
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการประมวลผลข้อมูล PDF: {e}")
            return False
        
    def verify_and_get_iptnumber(self) -> Optional[str]:
        """ตรวจสอบ iptnumber จากระบบและคืนค่า iptnumber ที่พร้อมใช้ (ลบคำว่า 'พ้นกำหนดรับชำระ' ออกแล้ว)
        
        Returns:
            iptnumber ที่พร้อมใช้ หรือ None ถ้าไม่สามารถตรวจสอบได้
        """
        try:
            # 1. อ่าน iptnumber จากระบบ
            iptnumber_text = getattr(self, 'latest_iptnumber_text', None)
            if not iptnumber_text:
                print(f"🔍 กำลังดึงข้อมูลจาก iptnumber element...")
                iptnumber_element = self.page.locator('//*[@id="iptnumber"]')
                count = iptnumber_element.count()
                if count > 0:
                    print(f"✅ พบ iptnumber element: {count} ตัว")
                    try:
                        iptnumber_text = (iptnumber_element.first.input_value() or '').strip()
                    except Exception as e1:
                        print(f"⚠️ ไม่สามารถอ่าน input_value ได้: {e1} ลอง text_content...")
                        try:
                            iptnumber_text = (iptnumber_element.first.text_content() or '').strip()
                        except Exception as e2:
                            print(f"❌ ไม่สามารถอ่าน text_content ได้: {e2}")
                            iptnumber_text = ''
                else:
                    print(f"❌ ไม่พบ iptnumber element")
                    return None
            
            print(f"📄 ข้อมูลจาก iptnumber (ล่าสุด/DOM): '{iptnumber_text or ''}'")
            
            # 2. ถ้า iptnumber ว่างหรือเป็น '-' ให้ลองอ่านใหม่
            if not iptnumber_text or iptnumber_text == '-':
                print(f"⚠️ ไม่พบข้อมูลใน iptnumber (เป็น '-' หรือว่างเปล่า) กำลังลองอ่านใหม่...")
                try:
                    iptnumber_element = self.page.locator('//*[@id="iptnumber"]')
                    if iptnumber_element.count() > 0:
                        try:
                            iptnumber_text = (iptnumber_element.first.input_value() or '').strip()
                        except Exception:
                            iptnumber_text = (iptnumber_element.first.text_content() or '').strip()
                        print(f"📄 ข้อมูลจาก iptnumber (อ่านใหม่): '{iptnumber_text or ''}'")
                except Exception as e:
                    print(f"⚠️ ไม่สามารถอ่าน iptnumber ใหม่ได้: {e}")
                
                if not iptnumber_text or iptnumber_text == '-':
                    print(f"⚠️ ไม่พบข้อมูลใน iptnumber หลังจากลองอ่านใหม่")
                    return None
            
            # 3. ตรวจสอบความตรงกันระหว่าง iptnumber กับเลขที่เอกสารจาก DOM (h3 element)
            print(f"🔍 กำลังตรวจสอบความตรงกันของเลขที่เอกสาร...")
            try:
                h3_element = self.page.locator('//*[@id="content"]/div[6]/div[1]/div[4]/div[1]/h3')
                if h3_element.count() > 0:
                    h3_text = (h3_element.first.text_content() or '').strip()
                    print(f"📄 ข้อความจาก h3 element: '{h3_text}'")
                    
                    # แยกส่วนหลัง # จากข้อความ
                    document_number_from_h3 = None
                    if '#' in h3_text:
                        parts = h3_text.split('#')
                        if len(parts) > 1:
                            document_number_from_h3 = parts[1].strip()
                            print(f"📄 เลขที่เอกสารจาก h3 (หลัง #): '{document_number_from_h3}'")
                    
                    # เปรียบเทียบกับ iptnumber_text
                    if document_number_from_h3:
                        iptnumber_clean = iptnumber_text.replace(' ', '').upper()
                        h3_clean = document_number_from_h3.replace(' ', '').upper()
                        
                        if iptnumber_clean != h3_clean:
                            print(f"⚠️ เลขที่เอกสารไม่ตรงกัน!")
                            print(f"   iptnumber: '{iptnumber_text}' → '{iptnumber_clean}'")
                            print(f"   h3 element: '{document_number_from_h3}' → '{h3_clean}'")
                            print(f"✅ ใช้เลขที่เอกสารจาก h3 element แทน: '{document_number_from_h3}'")
                            iptnumber_text = document_number_from_h3
                        else:
                            print(f"✅ เลขที่เอกสารตรงกัน: '{iptnumber_text}' == '{document_number_from_h3}'")
                    else:
                        print(f"⚠️ ไม่สามารถแยกเลขที่เอกสารจาก h3 element ได้ (ไม่พบ #)")
                else:
                    print(f"⚠️ ไม่พบ h3 element สำหรับตรวจสอบเลขที่เอกสาร ใช้ iptnumber เดิม")
            except Exception as e:
                print(f"⚠️ เกิดข้อผิดพลาดในการตรวจสอบเลขที่เอกสารจาก h3: {e} ใช้ iptnumber เดิม")
                # ยังคงใช้ iptnumber เดิมต่อไป
            
            # 4. ตรวจสอบว่า iptnumber มีค่าหรือไม่
            if not iptnumber_text or not iptnumber_text.strip():
                print(f"❌ iptnumber ว่างเปล่า ไม่สามารถใช้ได้")
                return None
            
            # 5. ลบคำว่า "พ้นกำหนดชำระ" และ "รอชำระ" ออกจาก iptnumber
            words_to_remove = ['พ้นกำหนดชำระ', 'พ้นกำหนดรับชำระ', 'รอชำระ', 'รอรับชำระ']
            cleaned_iptnumber = iptnumber_text
            for word in words_to_remove:
                cleaned_iptnumber = cleaned_iptnumber.replace(word, '').strip()
            
            # ลบช่องว่างที่ซ้ำกัน
            cleaned_iptnumber = ' '.join(cleaned_iptnumber.split())
            
            # ตรวจสอบอีกครั้งว่าหลังลบคำแล้วยังมีค่าหรือไม่
            if not cleaned_iptnumber or not cleaned_iptnumber.strip():
                print(f"❌ iptnumber ว่างเปล่าหลังลบคำ ไม่สามารถใช้ได้")
                return None
            
            if cleaned_iptnumber != iptnumber_text:
                print(f"🧹 ลบคำออกจาก iptnumber: '{iptnumber_text}' → '{cleaned_iptnumber}'")
            
            print(f"✅ iptnumber ที่พร้อมใช้: '{cleaned_iptnumber}'")
            return cleaned_iptnumber
            
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการตรวจสอบ iptnumber: {e}")
            import traceback
            print(f"📋 รายละเอียดข้อผิดพลาด:\n{traceback.format_exc()}")
            return None
    
    def create_new_pdf_from_form_data(self, pdf_data: Dict, verified_iptnumber: str = None) -> Optional[str]:
        """สร้างไฟล์ PDF ใหม่จากข้อมูลที่กรอกในฟอร์ม
        
        Args:
            pdf_data: ข้อมูล PDF
            verified_iptnumber: iptnumber ที่ตรวจสอบแล้ว (ถ้าไม่ระบุจะอ่านจาก DOM)
        """
        try:
            print(f" เริ่มสร้างไฟล์ PDF ใหม่...")
            
            # ใช้ iptnumber ที่ตรวจสอบแล้ว หรืออ่านจาก DOM
            if verified_iptnumber:
                iptnumber_text = verified_iptnumber
                print(f"📄 ใช้ iptnumber ที่ตรวจสอบแล้ว: '{iptnumber_text}'")
            else:
                # ใช้ค่า iptnumber ที่บันทึกล่วงหน้าถ้ามี มิฉะนั้นอ่านจาก DOM
                iptnumber_text = getattr(self, 'latest_iptnumber_text', None)
                if not iptnumber_text:
                    print(f"🔍 กำลังดึงข้อมูลจาก iptnumber element...")
                    iptnumber_element = self.page.locator('//*[@id="iptnumber"]')
                    count = iptnumber_element.count()
                    if count > 0:
                        print(f"✅ พบ iptnumber element: {count} ตัว")
                        try:
                            iptnumber_text = (iptnumber_element.first.input_value() or '').strip()
                        except Exception as e1:
                            print(f"⚠️ ไม่สามารถอ่าน input_value ได้: {e1} ลอง text_content...")
                            try:
                                iptnumber_text = (iptnumber_element.first.text_content() or '').strip()
                            except Exception as e2:
                                print(f"❌ ไม่สามารถอ่าน text_content ได้: {e2}")
                                iptnumber_text = ''
                    else:
                        print(f"❌ ไม่พบ iptnumber element")
                        return None
                    
                    print(f"📄 ข้อมูลจาก iptnumber (ล่าสุด/DOM): '{iptnumber_text or ''}'")
                    
                    # ถ้า iptnumber ว่างหรือเป็น '-' ให้ลองอ่านใหม่
                    if not iptnumber_text or iptnumber_text == '-':
                        print(f"⚠️ ไม่พบข้อมูลใน iptnumber (เป็น '-' หรือว่างเปล่า) กำลังลองอ่านใหม่...")
                        try:
                            iptnumber_element = self.page.locator('//*[@id="iptnumber"]')
                            if iptnumber_element.count() > 0:
                                try:
                                    iptnumber_text = (iptnumber_element.first.input_value() or '').strip()
                                except Exception:
                                    iptnumber_text = (iptnumber_element.first.text_content() or '').strip()
                                print(f"📄 ข้อมูลจาก iptnumber (อ่านใหม่): '{iptnumber_text or ''}'")
                        except Exception as e:
                            print(f"⚠️ ไม่สามารถอ่าน iptnumber ใหม่ได้: {e}")
                        
                        if not iptnumber_text or iptnumber_text == '-':
                            print(f"⚠️ ไม่พบข้อมูลใน iptnumber หลังจากลองอ่านใหม่ ยกเลิกการสร้างไฟล์สำหรับรายการนี้")
                            return None
                    
                    # ลบคำว่า "พ้นกำหนดชำระ" และ "รอชำระ" ออกจาก iptnumber
                    words_to_remove = ['พ้นกำหนดชำระ', 'พ้นกำหนดรับชำระ', 'รอชำระ', 'รอรับชำระ']
                    for word in words_to_remove:
                        iptnumber_text = iptnumber_text.replace(word, '').strip()
                    
                    # ลบช่องว่างที่ซ้ำกัน
                    iptnumber_text = ' '.join(iptnumber_text.split())

            # อ่านวันที่จากเว็บเพื่อให้ตรงกับข้อมูลที่กรอก
            print(f"🔍 กำลังอ่านวันที่จากเว็บ...")
            date_from_web = None
            try:
                date_field = self.page.locator('//*[@id="iptdate"]')
                date_count = date_field.count()
                if date_count > 0:
                    date_from_web = (date_field.first.input_value() or '').strip()
                    if date_from_web:
                        print(f"📅 อ่านวันที่จากเว็บสำเร็จ: {date_from_web}")
                        # อัปเดตวันที่ใน pdf_data ให้ตรงกับที่กรอกในเว็บ
                        pdf_data['document_date'] = date_from_web
                    else:
                        print(f"⚠️ วันที่บนเว็บว่างเปล่า ใช้วันที่เดิมจาก PDF")
                else:
                    print(f"⚠️ ไม่พบ element วันที่บนเว็บ (#iptdate) ใช้วันที่เดิมจาก PDF")
            except Exception as e:
                print(f"⚠️ อ่านวันที่จากเว็บไม่ได้: {e} ใช้วันที่เดิมจาก PDF")

            # มีค่า iptnumber → สร้างไฟล์
            print(f"✅ ข้อมูล iptnumber มีเนื้อหา: '{iptnumber_text}'")
            print(f"📤 ส่งข้อมูลไปยัง FileManager...")
            source_folder = pdf_data.get('source_folder')
            if not source_folder:
                src_path = pdf_data.get('file_path', '')
                source_folder = os.path.dirname(src_path) if src_path else ''
            # อ่าน folder_settings เพื่อส่งไปยัง create_pdf_from_form_data
            folder_settings = self.file_manager.read_folder_settings()
            
            # เพิ่ม folder_code ลงใน pdf_data ถ้ายังไม่มี
            if 'folder_code' not in pdf_data:
                # พยายามดึง folder_code จาก source_folder
                if source_folder:
                    folder_code = self.file_manager.get_folder_code_from_path(Path(source_folder))
                    if folder_code:
                        pdf_data['folder_code'] = folder_code
                        print(f"🔍 เพิ่ม folder_code: '{folder_code}'")
            
            new_pdf_path = self.file_manager.create_pdf_from_form_data(pdf_data, iptnumber_text, source_folder, folder_settings)
            if new_pdf_path:
                # ตรวจสอบว่าไฟล์ถูกสร้างจริงหรือไม่
                if os.path.exists(new_pdf_path):
                    file_size = os.path.getsize(new_pdf_path)
                    print(f"✅ สร้างไฟล์ PDF ใหม่สำเร็จ: {new_pdf_path} (ขนาด: {file_size} bytes)")
                    self.latest_created_file_path = new_pdf_path
                    return new_pdf_path
                else:
                    print(f"❌ ไฟล์ถูกสร้างแต่ไม่พบในระบบไฟล์: {new_pdf_path}")
                    return None
            print(f"❌ ไม่สามารถสร้างไฟล์ PDF ใหม่ได้ (FileManager ส่งกลับ None)")
            return None
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการสร้างไฟล์ PDF ใหม่: {e}")
            return None

    def fill_customer_id_again(self, pdf_data: Dict) -> bool:
        """กรอก Customer ID ใหม่"""
        try:
            print(f"🔄 กรอก Customer ID ใหม่: {pdf_data.get('customer_id', '')}")
            
            # หา input field สำหรับ Customer ID โดยลองหลาย selector แบบแยกกัน (CSS/XPath)
            selector_candidates = [
                '#iptcontactname',
                'input#iptcontactname',
                'xpath=//*[@id="iptcontactname"]',
                'input[name="customer_id"]',
                'input#customer_id',
                '#customer_id'
            ]

            customer_locator = None
            for sel in selector_candidates:
                loc = self.page.locator(sel)
                if loc.count() > 0:
                    customer_locator = loc.first
                    break

            if not customer_locator:
                print(f"❌ ไม่พบ input field สำหรับ Customer ID")
                return False

            # ลบข้อมูลเก่าและกรอกใหม่
            try:
                customer_locator.fill('')
            except Exception:
                pass
            customer_locator.fill(pdf_data.get('customer_id', ''))

            # ยืนยันการกรอก
            try:
                customer_locator.press('Enter')
            except Exception:
                try:
                    customer_locator.press('Tab')
                except Exception:
                    pass

            print(f"✅ กรอก Customer ID ใหม่สำเร็จ")
            return True
                
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการกรอก Customer ID ใหม่: {e}")
            return False
        
    def upload_file_directly(self, file_path: str, pdf_data: Dict = None) -> bool:
        """โยนไฟล์เข้าเว็บโดยตรงโดยไม่ต้องกดปุ่มอัปโหลด
        
        Args:
            file_path: พาธไฟล์ที่จะอัปโหลด
            pdf_data: ข้อมูล PDF สำหรับตรวจสอบเลขที่เอกสาร (optional)
        """
        try:
            if not self.is_logged_in:
                print("Not logged in")
                return False
            
            print(f"�� โยนไฟล์เข้าเว็บโดยตรง: {file_path}")
            
            # 1. หา file input field ที่ซ่อนอยู่
            print(f"�� กำลังหาฟิลด์อัปโหลดที่ซ่อนอยู่...")
            file_input = self.page.locator('input[type="file"]')
            
            count = file_input.count()
            if count > 0:
                print(f"✅ พบฟิลด์อัปโหลด: {count} ฟิลด์")
                
                # โยนไฟล์เข้าไปโดยตรง
                print(f"�� โยนไฟล์เข้าเว็บ...")
                file_input.first.set_input_files(file_path)
                
                # รอให้อัปโหลดเสร็จ
                time.sleep(1)
                
                print(f"✅ โยนไฟล์เข้าเว็บสำเร็จ: {file_path}")
                return True
            else:
                print(f"❌ ไม่พบฟิลด์อัปโหลด")
                return False
                
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการโยนไฟล์: {e}")
            import traceback
            print(f"📋 รายละเอียดข้อผิดพลาด:\n{traceback.format_exc()}")
            return False
    
    def extract_document_number_from_filename(self, file_path: str) -> Optional[str]:
        """ดึงเลขที่เอกสารจากชื่อไฟล์"""
        try:
            filename = os.path.basename(file_path)
            filename_no_ext = os.path.splitext(filename)[0]  # ลบ .pdf
            
            print(f"🔍 กำลังดึงเลขที่เอกสารจากชื่อไฟล์: '{filename_no_ext}'")
            
            # ลองหาจากรูปแบบที่อาจมี:
            # 1. ชื่อไฟล์อาจมีรูปแบบ: "DD.MM.YYYY เลขที่เอกสาร ชื่อบริการ" หรือ "เลขที่เอกสาร ชื่อบริการ"
            # 2. ให้ลองดึงเลขที่เอกสารจาก pdf_data ก่อน ถ้าไม่มีจึงดึงจากชื่อไฟล์
            
            # วิธีที่ 1: ถ้ามี pdf_data ให้ใช้ document_number จาก pdf_data
            # (จะต้องส่งผ่าน parameter แทน)
            
            # วิธีที่ 2: ดึงจากชื่อไฟล์โดยการลบส่วนวันที่และชื่อบริการออก
            # แต่เนื่องจากชื่อบริการอาจแตกต่างกัน ให้ใช้วิธีอื่น
            
            # วิธีที่ 3: คืนค่า None เพื่อให้ใช้ pdf_data แทน
            return None
            
        except Exception as e:
            print(f"⚠️ เกิดข้อผิดพลาดในการดึงเลขที่เอกสารจากชื่อไฟล์: {e}")
            return None
    
    def verify_filename_matches_iptnumber(self, file_path: str) -> bool:
        """ตรวจสอบว่าเลขที่เอกสารในชื่อไฟล์ตรงกับ iptnumber ที่บันทึกในระบบหรือไม่
        
        Args:
            file_path: พาธไฟล์ที่จะตรวจสอบ
        
        Returns:
            True ถ้าเลขที่เอกสารในชื่อไฟล์ตรงกับ iptnumber, False ถ้าไม่ตรงกัน
        """
        try:
            # 1. ดึงเลขที่เอกสารจากชื่อไฟล์
            filename = os.path.basename(file_path)
            filename_no_ext = os.path.splitext(filename)[0]  # ลบ .pdf
            
            print(f"🔍 กำลังตรวจสอบชื่อไฟล์: '{filename_no_ext}'")
            
            # 2. อ่าน iptnumber จากระบบ
            iptnumber_text = getattr(self, 'latest_iptnumber_text', None)
            if not iptnumber_text:
                # ลองอ่านจาก DOM
                try:
                    iptnumber_element = self.page.locator('//*[@id="iptnumber"]')
                    if iptnumber_element.count() > 0:
                        try:
                            iptnumber_text = (iptnumber_element.first.input_value() or '').strip()
                        except Exception:
                            iptnumber_text = (iptnumber_element.first.text_content() or '').strip()
                except Exception as e:
                    print(f"⚠️ ไม่สามารถอ่าน iptnumber ได้: {e}")
                    return False
            
            if not iptnumber_text or iptnumber_text == '-':
                print(f"⚠️ ไม่พบข้อมูล iptnumber ในระบบ")
                return False
            
            # ทำความสะอาดข้อมูล
            iptnumber_clean = iptnumber_text.replace(' ', '').upper().strip()
            filename_clean = filename_no_ext.replace(' ', '').upper()
            
            print(f"📄 iptnumber จากระบบ: '{iptnumber_text}' → '{iptnumber_clean}'")
            print(f"📄 ชื่อไฟล์: '{filename_no_ext}' → '{filename_clean}'")
            
            # ตรวจสอบว่าเลขที่เอกสารในชื่อไฟล์มี iptnumber หรือไม่
            if iptnumber_clean in filename_clean:
                print(f"✅ เลขที่เอกสารในชื่อไฟล์ตรงกับ iptnumber ที่บันทึกในระบบ")
                return True
            else:
                print(f"❌ เลขที่เอกสารในชื่อไฟล์ไม่ตรงกับ iptnumber ที่บันทึกในระบบ")
                print(f"   iptnumber: '{iptnumber_clean}'")
                print(f"   ชื่อไฟล์: '{filename_clean}'")
                return False
                
        except Exception as e:
            print(f"⚠️ เกิดข้อผิดพลาดในการตรวจสอบชื่อไฟล์: {e}")
            import traceback
            print(f"📋 รายละเอียดข้อผิดพลาด:\n{traceback.format_exc()}")
            return False
    
    def verify_document_number_match(self, file_path: str, pdf_data: Dict, max_wait: int = 5) -> bool:
        """เปรียบเทียบเลขที่เอกสารจากไฟล์กับเลขที่เอกสารบนเว็บ
        
        Args:
            file_path: พาธไฟล์ที่จะอัปโหลด
            pdf_data: ข้อมูล PDF ที่มี document_number
            max_wait: จำนวนครั้งที่รอให้เลขที่เอกสารตรงกัน (ครั้งละ 0.5 วินาที)
        
        Returns:
            True ถ้าเลขที่เอกสารตรงกัน, False ถ้าไม่ตรงกัน
        """
        try:
            print(f"🔍 กำลังตรวจสอบเลขที่เอกสารก่อนอัปโหลด...")
            
            # 1. ดึงเลขที่เอกสารจากไฟล์ (จาก pdf_data หรือชื่อไฟล์)
            document_number_from_file = pdf_data.get('document_number', '')
            
            # ถ้าไม่มีใน pdf_data ให้ลองดึงจากชื่อไฟล์
            if not document_number_from_file:
                doc_from_filename = self.extract_document_number_from_filename(file_path)
                if doc_from_filename:
                    document_number_from_file = doc_from_filename
                else:
                    # ถ้าไม่มีเลย ให้ใช้ชื่อไฟล์ทั้งหมด (ลบ .pdf)
                    filename = os.path.basename(file_path)
                    document_number_from_file = os.path.splitext(filename)[0]
            
            # ทำความสะอาดเลขที่เอกสาร
            document_number_from_file = ' '.join(document_number_from_file.split()).strip()
            print(f"📄 เลขที่เอกสารจากไฟล์: '{document_number_from_file}'")
            
            if not document_number_from_file:
                print(f"⚠️ ไม่พบเลขที่เอกสารจากไฟล์ ข้ามการตรวจสอบ")
                return True  # ถ้าไม่มีเลขที่เอกสาร ให้อัปโหลดได้ (เพื่อไม่ให้ขัดจังหวะ)
            
            # 2. อ่านเลขที่เอกสารจากเว็บ (iptnumber)
            for attempt in range(max_wait):
                try:
                    iptnumber_element = self.page.locator('//*[@id="iptnumber"]')
                    if iptnumber_element.count() > 0:
                        try:
                            iptnumber_text = (iptnumber_element.first.input_value() or '').strip()
                        except Exception:
                            iptnumber_text = (iptnumber_element.first.text_content() or '').strip()
                        
                        # ทำความสะอาดเลขที่เอกสารจากเว็บ
                        iptnumber_text = ' '.join(iptnumber_text.split()).strip()
                        print(f"🌐 เลขที่เอกสารจากเว็บ (ครั้งที่ {attempt+1}): '{iptnumber_text}'")
                        
                        # เปรียบเทียบ (ไม่สนใจตัวพิมพ์เล็ก/ใหญ่)
                        if not iptnumber_text or iptnumber_text == '-':
                            print(f"⏳ เลขที่เอกสารบนเว็บยังว่าง รออีกครั้ง... ({attempt+1}/{max_wait})")
                            time.sleep(0.5)
                            continue
                        
                        # เปรียบเทียบ (ไม่สนใจช่องว่างและตัวพิมพ์เล็ก/ใหญ่)
                        file_doc_clean = document_number_from_file.replace(' ', '').upper()
                        web_doc_clean = iptnumber_text.replace(' ', '').upper()
                        
                        if file_doc_clean == web_doc_clean:
                            print(f"✅ เลขที่เอกสารตรงกัน: '{document_number_from_file}' == '{iptnumber_text}'")
                            return True
                        else:
                            # ตรวจสอบว่ามีส่วนที่ตรงกันหรือไม่ (แบบยืดหยุ่น)
                            if file_doc_clean in web_doc_clean or web_doc_clean in file_doc_clean:
                                print(f"✅ เลขที่เอกสารตรงกัน (บางส่วน): '{document_number_from_file}' ~= '{iptnumber_text}'")
                                return True
                            else:
                                print(f"⚠️ เลขที่เอกสารไม่ตรงกัน: '{document_number_from_file}' != '{iptnumber_text}'")
                                if attempt < max_wait - 1:
                                    print(f"⏳ รอให้เลขที่เอกสารตรงกันอีกครั้ง... ({attempt+1}/{max_wait})")
                                    time.sleep(0.5)
                                    continue
                                else:
                                    print(f"❌ เลขที่เอกสารไม่ตรงกันหลังจากรอ {max_wait} ครั้ง")
                                    return False
                    else:
                        print(f"⚠️ ไม่พบ element iptnumber บนเว็บ")
                        if attempt < max_wait - 1:
                            time.sleep(0.5)
                            continue
                        else:
                            print(f"⚠️ ไม่สามารถอ่านเลขที่เอกสารจากเว็บได้ ข้ามการตรวจสอบ")
                            return True  # ถ้าอ่านไม่ได้ ให้อัปโหลดได้
                except Exception as e:
                    print(f"⚠️ เกิดข้อผิดพลาดในการอ่านเลขที่เอกสารจากเว็บ (ครั้งที่ {attempt+1}): {e}")
                    if attempt < max_wait - 1:
                        time.sleep(0.5)
                        continue
                    else:
                        print(f"⚠️ ไม่สามารถตรวจสอบเลขที่เอกสารได้ ข้ามการตรวจสอบ")
                        return True  # ถ้ามี error ให้อัปโหลดได้
            
            return False
            
        except Exception as e:
            print(f"⚠️ เกิดข้อผิดพลาดในการตรวจสอบเลขที่เอกสาร: {e}")
            return True  # ถ้ามี error ให้อัปโหลดได้ (เพื่อไม่ให้ขัดจังหวะ)
        
    def execute_peak_engine_workflow(self, pdf_data_list: list, main_folder: str) -> bool:
        """จัดการการทำงานทั้งหมดของ PeakEngine (รวมการเปิด/ปิด Playwright)"""
        try:
            self._log(f"🚀 เริ่มต้นการทำงานด้วย Playwright...")
            self._status_update(step="กำลังเตรียม Playwright", folder=main_folder, file='-')
            success = self.setup_driver()
            
            if not success:
                self._log(f"❌ ไม่สามารถเปิด Playwright ได้", level="error")
                return False
            
            self._log(f"✅ เปิด Playwright สำเร็จ")
            self._log(f"📊 รวบรวมข้อมูล PDF: {len(pdf_data_list)} ไฟล์")
            self._status_update(step="กำลังเชื่อมต่อ PeakEngine", folder=main_folder, file='-')
            
            # อ่านข้อมูลการตั้งค่าจากไฟล์ txt ที่ตรงกับชื่อโฟลเดอร์ Build*
            # หาโฟลเดอร์ Build* จาก main_folder
            main_folder_path = Path(main_folder)
            build_folder = None
            current_path = main_folder_path
            while current_path != current_path.parent:
                if current_path.name.startswith('Build'):
                    build_folder = current_path
                    break
                current_path = current_path.parent
            
            if build_folder:
                # ใช้แค่หมายเลข BuildXXX (เช่น Build000 ทดสอบระบบ → Build000)
                folder_name = build_folder.name
                build_number = folder_name.split()[0]  # แยกเอาแค่ "Build000"
                # ตำแหน่งหลักภายในโฟลเดอร์งาน
                local_txt = Path(build_folder) / "รหัส" / f"{build_number}.txt"
                # ตำแหน่งสำรองศูนย์รวมในไดรฟ์ V: ภายใต้ A./AA./AAA.โฟรเดอร์หลัก/Build000 ทดสอบระบบ/รหัส
                drive_root = Path(f"{Config.BASE_FOLDER}:/")
                candidates = [local_txt]
                for main_name in getattr(Config, 'MAIN_FOLDERS', ["A.โฟร์เดอร์หลัก", "AA.โฟรเดอร์หลัก", "AAA.โฟรเดอร์หลัก"]):
                    candidates.append(
                        drive_root / main_name / f"Build{Config.TEST_SYSTEM_FOLDER}" / "รหัส" / f"{build_number}.txt"
                    )
                # เผื่อมีโครงสร้าง V:/Build000 ทดสอบระบบ/รหัส ด้วย
                candidates.append(drive_root / f"Build{Config.TEST_SYSTEM_FOLDER}" / "รหัส" / f"{build_number}.txt")

                # เลือกไฟล์แรกที่มีอยู่จริง
                chosen = None
                for c in candidates:
                    if c.exists():
                        chosen = c
                        break
                # debug แสดง path ที่ลองหา
                try:
                    print("🔎 ค้นหา TXT จากตำแหน่งต่อไปนี้ (ลำดับความสำคัญ):")
                    for i, c in enumerate(candidates, start=1):
                        print(f"  {i}) {c}")
                    print(f"➡️ ใช้งาน: {chosen if chosen else 'ไม่พบไฟล์ที่ตรงกัน'}")
                except Exception:
                    pass

                config_file_path = str(chosen) if chosen else str(local_txt)
            else:
                self._log(f"❌ ไม่พบโฟลเดอร์ Build* จาก: {main_folder}", level="error")
                self.close_driver()
                return False
            if not self.read_config_from_txt(config_file_path):
                self._log(f"❌ ไม่สามารถอ่านข้อมูลการตั้งค่าได้", level="error")
                self.close_driver()
                return False
            
            self._status_update(step="กำลังประมวลผล PDF", folder=main_folder, file='-')
            # เรียกใช้ PeakEngine Workflow
            # รองรับการเริ่มต้นจากไฟล์ลำดับที่กำหนดผ่าน pdf_data_list meta: pdf_data_list.__start_index__
            start_index = 0
            try:
                start_index = int(getattr(pdf_data_list, '__start_index__', 0))
            except Exception:
                start_index = 0
            workflow_success = False
            max_attempts = 2
            for attempt in range(max_attempts):
                workflow_success = self.peak_engine_workflow(pdf_data_list, start_index=start_index)
                if workflow_success:
                    break
                
                connection_lost = (
                    isinstance(self._playwright_error, str)
                    and "'NoneType' object has no attribute 'send'" in self._playwright_error
                )
                if not connection_lost or attempt == max_attempts - 1:
                    break
                
                print("⚠️ ตรวจพบว่าการเชื่อมต่อ Playwright หลุด ('NoneType.send') กำลังรีสตาร์ทและลองใหม่...")
                self.close_driver()
                time.sleep(1)
                if not self.setup_driver():
                    print("❌ ไม่สามารถเปิด Playwright ใหม่ได้สำหรับการลองซ้ำ")
                    break
                if not self.read_config_from_txt(config_file_path):
                    print("❌ ไม่สามารถอ่านข้อมูลการตั้งค่าหลังรีสตาร์ท Playwright ได้")
                    self.close_driver()
                    return False
                self._playwright_error = None
            
            leftover_pdf_files: List[Path] = []
            try:
                leftover_pdf_files = [
                    p for p in self.file_manager.get_pdf_files(main_folder_path)
                    if p.suffix.lower() == '.pdf'
                ]
            except Exception as leftover_err:
                print(f"⚠️ ตรวจสอบไฟล์ PDF ค้างไม่สำเร็จ: {leftover_err}")
                leftover_pdf_files = []

            if leftover_pdf_files:
                print("⚠️ พบไฟล์ PDF ที่ยังค้างอยู่หลังจบงาน:")
                for pending_file in leftover_pdf_files:
                    try:
                        print(f"   - {pending_file}")
                    except Exception:
                        pass
                workflow_success = False

            if workflow_success:
                self._log(f"✅ PeakEngine Workflow สำเร็จ!")
                self._status_update(step="สำเร็จ", folder=main_folder, file='-')
                
                # ตรวจสอบสถานะ
                status = self.get_peak_engine_status()
                self._log(f"📊 สถานะ PeakEngine: {status}")
                
            else:
                self._log(f"❌ PeakEngine Workflow ไม่สำเร็จ", level="warning")
                self._status_update(step="มีข้อผิดพลาด", folder=main_folder, file='-')
            
            # แจ้งเตือนจบงานผ่าน LINE (ครบถ้วน/บางส่วน) หากมี ReportManager จากขั้นตอนอ่าน
            try:
                rm = get_global_report_manager()
            except Exception:
                rm = None
            if rm is not None:
                if leftover_pdf_files:
                    try:
                        rm.pending_action_count += len(leftover_pdf_files)
                    except Exception:
                        pass
                # ประเมินความสำเร็จ: ถ้าทำงานผ่านเว็บสำเร็จให้ถือว่าครบถ้วน, ไม่เช่นนั้นบางส่วน
                end_message = rm.end_success() if workflow_success else rm.end_partial()
                sent_end = False
                try:
                    if getattr(Config, 'LINE_OA_CHANNEL_ACCESS_TOKEN', '') and getattr(Config, 'LINE_OA_DEFAULT_TO', ''):
                        sent_end = line_oa_push(end_message)
                    elif getattr(Config, 'LINE_NOTIFY_TOKEN', ''):
                        sent_end = line_notify(end_message)
                except Exception:
                    sent_end = False
                if not sent_end:
                    print("ℹ️ ไม่ได้ส่ง LINE end (ไม่มี token OA/Notify หรือล้มเหลว)")

                # สร้างรายงานฉบับสุดท้าย (รวมผลอัปโหลด/ซ้ำ/รอดำเนินการ/อ่านไม่ได้/ฐานข้อมูล)
                try:
                    report_path = rm.write_txt_report()
                    if report_path:
                        print(f"📝 สร้างไฟล์รายงานการทำงาน: {report_path}")
                    else:
                        print("⚠️ สร้างไฟล์รายงานการทำงานไม่สำเร็จ")
                except Exception:
                    print("⚠️ เกิดข้อผิดพลาดระหว่างสร้างไฟล์รายงานฉบับสุดท้าย")

            # ปิด Playwright
            self.close_driver()
            self._log(f"🔒 ปิด Playwright แล้ว")
            
            return workflow_success
            
        except Exception as e:
            self._log(f"❌ เกิดข้อผิดพลาดในการทำงานกับ PeakEngine: {e}", level="error")
            try:
                self.close_driver()
            except:
                pass
            self._status_update(step="เกิดข้อผิดพลาด", folder=main_folder, file='-')
            return False

    def _check_and_handle_duplicate_before_upload(self) -> bool:
        """ตรวจจับแจ้งเตือน 'เลขที่ใบกำกับภาษีอ้างอิงนี้สำหรับผู้ติดต่อนี้เคยถูกใช้งานแล้ว' ก่อนอัปโหลด ถ้าพบให้ย้ายไฟล์ไปโฟลเดอร์ 'เอกสารซ้ำรอตรวจ' และคืนค่า True"""
        try:
            duplicate_texts = [
                'เลขที่ใบกำกับภาษีอ้างอิงนี้สำหรับผู้ติดต่อนี้เคยถูกใช้งานแล้ว',
                'ผู้ติดต่อนี้เคยถูกใช้งานแล้ว'
            ]
            # ตรวจกรอบแดงก่อน
            red = self.page.locator('//*[@id="dvredalert"]')
            if red.count() > 0 and red.first.is_visible():
                try:
                    lb = self.page.locator('//*[@id="lbredalert"]')
                    msg = (lb.first.text_content() or '').strip() if lb.count() > 0 else (red.first.text_content() or '').strip()
                except Exception:
                    msg = (red.first.text_content() or '').strip()
                if any(t in msg for t in duplicate_texts):
                    print('⚠️ พบแจ้งเตือนเอกสารซ้ำ (แดง) ก่อนอัปโหลด')
                    if getattr(self, 'latest_created_file_path', None):
                        moved_to = self.file_manager.move_file_to_duplicate_folder(self.latest_created_file_path)
                        if moved_to:
                            print(f"✅ ย้ายไฟล์ไปยัง 'เอกสารซ้ำรอตรวจ': {moved_to}")
                        else:
                            print('❌ ย้ายไฟล์ไปยังโฟลเดอร์เอกสารซ้ำรอตรวจไม่สำเร็จ')
                    else:
                        print('⚠️ ไม่มีพาธไฟล์ล่าสุดให้ย้าย')
                    return True
            # เผื่อบางระบบแจ้งในกรอบเขียวด้วยข้อความเดียวกัน
            green = self.page.locator('//*[@id="dvgreenalert"]')
            if green.count() > 0 and green.first.is_visible():
                try:
                    lb = self.page.locator('//*[@id="lbgreenalert"]')
                    msg = (lb.first.text_content() or '').strip() if lb.count() > 0 else (green.first.text_content() or '').strip()
                except Exception:
                    msg = (green.first.text_content() or '').strip()
                if any(t in msg for t in duplicate_texts):
                    print('⚠️ พบแจ้งเตือนเอกสารซ้ำ (เขียว) ก่อนอัปโหลด')
                    if getattr(self, 'latest_created_file_path', None):
                        moved_to = self.file_manager.move_file_to_duplicate_folder(self.latest_created_file_path)
                        if moved_to:
                            print(f"✅ ย้ายไฟล์ไปยัง 'เอกสารซ้ำรอตรวจ': {moved_to}")
                        else:
                            print('❌ ย้ายไฟล์ไปยังโฟลเดอร์เอกสารซ้ำรอตรวจไม่สำเร็จ')
                    else:
                        print('⚠️ ไม่มีพาธไฟล์ล่าสุดให้ย้าย')
                    return True
        except Exception as e:
            print(f"⚠️ ตรวจแจ้งเตือนเอกสารซ้ำก่อนอัปโหลดล้มเหลว: {e}")
        return False

    def wait_for_save_result(self, timeout: int = 20) -> str:
        """รอผลการกดบันทึก/ยืนยัน: คืนค่า 'approved' | 'duplicate' | 'timeout'"""
        try:
            start_time = time.time()
            print(f"⏳ เริ่มรอผลบันทึก (timeout: {timeout} วินาที)...")
            while time.time() - start_time < timeout:
                # เช็คกรอบแดง: เอกสารซ้ำ
                try:
                    red = self.page.locator('//*[@id="dvredalert"]')
                    if red.count() > 0 and red.first.is_visible():
                        try:
                            lb = self.page.locator('//*[@id="lbredalert"]')
                            msg = (lb.first.text_content() or '').strip() if lb.count() > 0 else (red.first.text_content() or '').strip()
                        except Exception:
                            msg = (red.first.text_content() or '').strip()
                        if 'เลขที่ใบกำกับภาษีอ้างอิงนี้สำหรับผู้ติดต่อนี้เคยถูกใช้งานแล้ว' in msg or 'ผู้ติดต่อนี้เคยถูกใช้งานแล้ว' in msg:
                            print('⚠️ ผลบันทึก: เอกสารซ้ำ (แดง)')
                            return 'duplicate'
                        if 'โปรดกรอกข้อมูลในช่อง' in msg:
                            print('⚠️ แจ้งเตือน: โปรดกรอกข้อมูลในช่อง → กรอกข้อมูลทั้งหมดใหม่อีกครั้ง')
                            # ตรวจสอบว่ามีข้อมูล PDF หรือ Excel สำหรับ retry
                            if getattr(self, '_current_pdf_data_for_retry', None) and self._refill_attempt_count < 2:
                                try:
                                    self._refill_attempt_count += 1
                                    # รีหน้าแล้วกรอกใหม่ทั้งชุด (PDF)
                                    self.refresh_express_page()
                                    self.fill_form_data(self._current_pdf_data_for_retry)
                                    # หลังกรอกใหม่ จะยังคงวนในลูปรอผลต่อ
                                    continue
                                except Exception as _:
                                    print('⚠️ รีฟิลไม่สำเร็จ จะรอตรวจผลต่อ')
                            elif getattr(self, '_current_excel_rows_for_retry', None) and self._refill_attempt_count < 2:
                                try:
                                    self._refill_attempt_count += 1
                                    print(f'🔄 เริ่มรีเซ็ตและกรอกข้อมูลใหม่ทั้งหมด (ครั้งที่ {self._refill_attempt_count})')
                                    # รีหน้าแล้วกรอกใหม่ทั้งชุด (Excel - ทั้งหมดของลำดับ)
                                    self.refresh_express_page()
                                    time.sleep(1)  # รอให้หน้าเว็บโหลดเสร็จ
                                    
                                    sequence_info = getattr(self, '_current_excel_sequence_info_for_retry', {})
                                    is_vat_sheet = sequence_info.get('is_vat_sheet', False)
                                    excel_path = sequence_info.get('excel_path', '')
                                    rows = self._current_excel_rows_for_retry
                                    sequence = sequence_info.get('sequence', '')
                                    
                                    print(f'📝 กำลังกรอกข้อมูลใหม่ทั้งหมด: ลำดับ {sequence}, {len(rows)} แถว')
                                    
                                    # กรอกข้อมูลแถวแรก
                                    if len(rows) > 0:
                                        first_row = rows[0]
                                        effective_is_vat_sheet = is_vat_sheet if getattr(self, 'current_folder_group', 'regular') != 'special' else False
                                        if self.fill_form_from_excel_row(first_row, effective_is_vat_sheet, row_index=1, is_first_row=True, folder_group=getattr(self, 'current_folder_group', 'regular')):
                                            print(f'✅ กรอกข้อมูลแถวแรกสำเร็จ')
                                            
                                            # กรอกแถวที่ 2, 3, ... (ถ้ามี)
                                            if len(rows) > 1:
                                                for row_idx in range(1, len(rows)):
                                                    additional_row = rows[row_idx]
                                                    row_number = row_idx + 1
                                                    
                                                    # คลิกปุ่ม "+ เพิ่มรายการ"
                                                    try:
                                                        add_item_button = self.page.locator('div.button-main.button-green.different-border.float-left:has-text("+ เพิ่มรายการ")')
                                                        if add_item_button.count() == 0:
                                                            add_item_button = self.page.locator('//div[contains(text(), "+ เพิ่มรายการ")]')
                                                        if add_item_button.count() > 0:
                                                            add_item_button.first.click()
                                                            print(f'✅ คลิกปุ่ม "+ เพิ่มรายการ" สำหรับแถวที่ {row_number}')
                                                            time.sleep(0.5)
                                                    except Exception as e:
                                                        print(f'⚠️ ไม่สามารถคลิกปุ่ม "+ เพิ่มรายการ" ได้: {e}')
                                                    
                                                    # กรอกข้อมูลแถวที่ row_number
                                                    self.fill_form_from_excel_row(
                                                        additional_row,
                                                        is_vat_sheet,
                                                        row_index=row_number,
                                                        is_first_row=False
                                                    )
                                                    print(f'✅ กรอกข้อมูลแถวที่ {row_number} สำเร็จ')
                                            
                                            # ทำขั้นตอนสุดท้าย (หมายเหตุ, เลือกรวมภาษี, คลิกปุ่ม)
                                            # 5.5. กรอกหมายเหตุ (ใช้ข้อมูลจากแถวแรก)
                                            remark = first_row.get('หมายเหตุ', '').strip()
                                            if remark:
                                                remark_cleaned = self._clean_remark_text(remark)
                                                print(f'📝 กำลังกรอกหมายเหตุ: {remark_cleaned}')
                                                try:
                                                    remark_field = self.page.locator('#tarremark')
                                                    if remark_field.count() > 0:
                                                        remark_field.first.fill(remark_cleaned)
                                                        print(f'✅ กรอกหมายเหตุสำเร็จ')
                                                        time.sleep(0.5)
                                                except Exception as e:
                                                    print(f'⚠️ เกิดข้อผิดพลาดในการกรอกหมายเหตุ: {e}')
                                            
                                            # 7. เลือกดรอปดาวน์ภาษี (เฉพาะ is_vat_sheet)
                                            if is_vat_sheet:
                                                print(f'🔍 [VAT] กำลังเลือกดรอปดาวน์รวมภาษี...')
                                                try:
                                                    tax_status_dropdown = self.page.locator('//*[@id="ddltaxstatus"]')
                                                    if tax_status_dropdown.count() > 0:
                                                        tax_status_dropdown.first.click()
                                                        time.sleep(1)
                                                        tax_included_option = self.page.locator('text=รวมภาษี, text=Include Tax')
                                                        if tax_included_option.count() > 0:
                                                            tax_included_option.first.click()
                                                            print(f'✅ เลือกรวมภาษีสำเร็จ')
                                                        else:
                                                            try:
                                                                self.page.select_option('//*[@id="ddltaxstatus"]', index=1)
                                                                print(f'✅ เลือกรวมภาษีสำเร็จด้วยวิธี fallback')
                                                            except:
                                                                print(f'⚠️ ไม่สามารถเลือกรวมภาษีได้')
                                                except Exception as e:
                                                    print(f'⚠️ เกิดข้อผิดพลาดในการเลือกรวมภาษี: {e}')
                                            
                                            # 8. คลิกปุ่ม hidePaymentModal
                                            print(f'🔍 กำลังคลิกปุ่ม hidePaymentModal...')
                                            try:
                                                hide_payment_button = self.page.locator('//*[@id="hidePaymentModal"]')
                                                if hide_payment_button.count() > 0:
                                                    hide_payment_button.first.click()
                                                    print(f'✅ คลิกปุ่ม hidePaymentModal สำเร็จ')
                                            except Exception as e:
                                                print(f'⚠️ เกิดข้อผิดพลาดในการคลิกปุ่ม hidePaymentModal: {e}')
                                            
                                            # 9. คลิกปุ่มสุดท้าย
                                            print(f'🔍 กำลังกดปุ่มสุดท้าย...')
                                            time.sleep(1.5)
                                            try:
                                                final_button = self.page.locator('//*[@id="content"]/div[6]/div[6]/div/div[2]/div[2]/div[1]')
                                                if final_button.count() > 0:
                                                    final_button.first.click()
                                                    print(f'✅ คลิกปุ่มสุดท้ายสำเร็จ')
                                                    time.sleep(2)  # รอให้หน้าเว็บโหลด
                                            except Exception as e:
                                                print(f'⚠️ เกิดข้อผิดพลาดในการคลิกปุ่มสุดท้าย: {e}')
                                    
                                    # หลังกรอกใหม่ จะยังคงวนในลูปรอผลต่อ
                                    continue
                                except Exception as e:
                                    print(f'⚠️ รีฟิล Excel ไม่สำเร็จ: {e} จะรอตรวจผลต่อ')
                            else:
                                print('⚠️ ข้ามการรีฟิล (ไม่มีข้อมูล หรือพยายามเกินกำหนด)')
                except Exception:
                    pass

                # เช็คกรอบเขียว: อนุมัติบันทึกรายจ่าย ... สำเร็จ
                try:
                    green = self.page.locator('//*[@id="dvgreenalert"]')
                    if green.count() > 0 and green.first.is_visible():
                        try:
                            lb = self.page.locator('//*[@id="lbgreenalert"]')
                            msg = (lb.first.text_content() or '').strip() if lb.count() > 0 else (green.first.text_content() or '').strip()
                        except Exception:
                            msg = (green.first.text_content() or '').strip()
                        if ('อนุมัติบันทึกรายจ่าย' in msg and 'สำเร็จ' in msg) or ('Approved expense' in msg):
                            print('✅ ผลบันทึก: อนุมัติสำเร็จ')
                            return 'approved'
                except Exception:
                    pass

                # ตรวจสอบว่าหน้าเว็บยังโหลดอยู่หรือไม่
                try:
                    # ตรวจสอบว่ามี loading indicator หรือไม่
                    loading = self.page.locator('[class*="loading"], [class*="spinner"], [id*="loading"]')
                    if loading.count() > 0 and loading.first.is_visible():
                        print(f"⏳ หน้าเว็บกำลังโหลด...")
                    
                    # ตรวจสอบว่าหน้าเว็บยังตอบสนองหรือไม่
                    try:
                        # ลองหาองค์ประกอบพื้นฐานของหน้าเว็บ
                        basic_elements = [
                            '//*[@id="iptcontactname"]',
                            '//*[@id="iptaccountcode1"]',
                            '//*[@id="iptdate"]'
                        ]
                        
                        found_basic = False
                        for selector in basic_elements:
                            try:
                                element = self.page.locator(selector)
                                if element.count() > 0:
                                    found_basic = True
                                    break
                            except Exception:
                                pass
                        
                        if not found_basic:
                            print(f"⚠️ หน้าเว็บอาจไม่ตอบสนอง - ไม่พบองค์ประกอบพื้นฐาน")
                    except Exception:
                        pass
                except Exception:
                    pass

                # แสดงความคืบหน้า
                elapsed = int(time.time() - start_time)
                if elapsed % 3 == 0:  # แสดงทุก 3 วินาที
                    print(f"⏳ รอผลบันทึก... ({elapsed}/{timeout} วินาที)")
                
                time.sleep(0.3)
            
            print('⌛ ผลบันทึก: timeout (ไม่พบ alert box)')
            return 'timeout'
        except Exception as e:
            print(f"⚠️ ตรวจผลบันทึกล้มเหลว: {e}")
            return 'timeout'

    def check_and_fix_vat_value(self, expected_vat_text: str) -> bool:
        """ตรวจสอบยอดภาษีมูลค่าเพิ่มบนหน้ารวม และแก้ไขถ้าไม่ตรงกับค่าที่คาดหวัง
        expected_vat_text: ข้อความตัวเลขจาก pdf_data['vat_value'] เช่น '324.29'
        คืนค่า True ถ้าถูกต้องแล้วหรือแก้ไขสำเร็จ
        """
        try:
            # ดึงตัวเลขจาก expected
            try:
                expected_val = float(re.sub(r"[^0-9\.]+", "", expected_vat_text).replace(",", "")) if expected_vat_text else None
            except Exception:
                expected_val = None
            if expected_val is None:
                print("⚠️ ไม่มีค่า VAT ที่คาดหวัง ไม่ทำการตรวจสอบ")
                return True

            # 0) เลื่อนลงไปด้านล่างเพื่อดูส่วนสรุป VAT ให้แน่ใจว่า element อยู่ในวิวนักพัฒนาก่อนอ่านค่า
            try:
                self.page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                time.sleep(0.2)
            except Exception:
                pass

            # อ่าน VAT ที่แสดงบนหน้า (ใช้ selector หลายแบบเพื่อความทนทาน)
            def _select_vat_locator():
                candidates = [
                    'p[data-cypress="transactionVatAmount"]',
                    '//*[@data-cypress="transactionVatAmount"]',
                    'xpath=//p[@data-cypress="transactionVatAmount"]',
                    'css=[data-cypress="transactionVatAmount"]',
                    'xpath=//p[contains(normalize-space(.), "ภาษีมูลค่าเพิ่ม")]/following-sibling::p[1]',
                    'xpath=//div[@name="fromCurrencyBox"]//div[contains(@class, "subtotal")]/p[2]',
                    '//*[@id="content"]/div[6]/div[1]/div[3]/div[7]/div[2]/div[2]/div[3]/p[2]',
                ]
                for sel in candidates:
                    try:
                        loc = self.page.locator(sel)
                        if loc.count() > 0:
                            return loc
                    except Exception:
                        pass
                return None

            display_loc = _select_vat_locator()
            if not display_loc or display_loc.count() == 0:
                print("⚠️ ไม่พบตำแหน่งแสดง VAT บนหน้า (ลองใช้ data-cypress แล้ว)")
                return True
            display_text = (display_loc.first.text_content() or '').strip()
            try:
                display_val = float(re.sub(r"[^0-9\.]+", "", display_text).replace(",", "")) if display_text else None
            except Exception:
                display_val = None
            print(f"📊 ตรวจ VAT: คาดหวัง {expected_val} | หน้าจอ {display_val}")

            if display_val is not None and abs(display_val - expected_val) < 0.001:
                print("✅ VAT บนหน้าตรงกับข้อมูล ไม่ต้องแก้ไข")
                return True

            # รีเช็คหน้าเว็บสองรอบก่อนแก้ไข VAT ถ้ายอดไม่ตรง
            print("⚠️ VAT ไม่ตรง กำลังรีเช็คหน้าเว็บสองรอบ...")
            for check_round in range(2):
                print(f"🔍 รีเช็คหน้าเว็บรอบที่ {check_round + 1}/2...")
                time.sleep(1)
                
                # รีเฟรชหน้าเว็บ
                try:
                    self.page.reload(timeout=10000)
                    self.page.wait_for_load_state('domcontentloaded', timeout=5000)
                    time.sleep(0.2)
                except Exception:
                    pass
                
                # อ่าน VAT อีกครั้ง
                try:
                    self.page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                    time.sleep(0.2)
                except Exception:
                    pass
                
                display_loc_check = _select_vat_locator()
                if display_loc_check and display_loc_check.count() > 0:
                    display_text_check = (display_loc_check.first.text_content() or '').strip()
                    try:
                        display_val_check = float(re.sub(r"[^0-9\.]+", "", display_text_check).replace(",", "")) if display_text_check else None
                    except Exception:
                        display_val_check = None
                    
                    print(f"📊 รีเช็ครอบที่ {check_round + 1}: คาดหวัง {expected_val} | หน้าจอ {display_val_check}")
                    
                    if display_val_check is not None and abs(display_val_check - expected_val) < 0.001:
                        print(f"✅ VAT ตรงกันแล้วหลังรีเช็ครอบที่ {check_round + 1} ไม่ต้องแก้ไข")
                        return True
                else:
                    print(f"⚠️ ไม่พบตำแหน่งแสดง VAT ในรอบที่ {check_round + 1}")
            
            print("⚠️ VAT ยังไม่ตรงหลังรีเช็คสองรอบ เริ่มกระบวนการแก้ไข")

            # ลองแก้ไขสูงสุด 2 รอบ
            for attempt in range(2):
                print(f"🛠️ แก้ไข VAT รอบที่ {attempt + 1}/2: เตรียมกรอกค่าใหม่ 2 รอบ แล้วบันทึกครั้งเดียว")

                # 1) เลื่อนกลับขึ้นบนสุดเพื่อเข้าถึงปุ่ม 'ตัวเลือก' และหน่วง 0.8 วิ (ให้ element เสถียร)
                try:
                    self.page.evaluate('window.scrollTo(0, 0)')
                except Exception:
                    pass
                time.sleep(0.8)

                # 2) คลิกปุ่ม 'ตัวเลือก' แล้วเปิดเมนู
                options_btn_xpath = '//*[@id="content"]/div[6]/div[1]/div[3]/div[1]/div/div[3]'
                options_menu_xpath = '//*[@id="content"]/div[6]/div[1]/div[3]/div[1]/div/div[3]/div'
                try:
                    candidates_btn = [
                        options_btn_xpath,
                        "css=div.ui.button.dropdown.option-button.button-blue",
                        "css=div.option-button.button-blue",
                        "css=div.option-button",
                        "css=div.ui.button.dropdown:has-text('ตัวเลือก')",
                        "text=ตัวเลือก"
                    ]
                    options_btn = None
                    for sel in candidates_btn:
                        try:
                            loc = self.page.locator(sel)
                            if loc.count() > 0:
                                options_btn = loc.first
                                break
                        except Exception:
                            continue
                    if not options_btn:
                        print("⚠️ ไม่พบปุ่มตัวเลือก (options)")
                        return False
                    try:
                        options_btn.scroll_into_view_if_needed()
                        time.sleep(0.2)
                    except Exception:
                        pass
                    options_btn.click()
                    menu_visible = False
                    try:
                        self.page.locator("css=div.menu.transition:not(.hidden)").first.wait_for(state='visible', timeout=3000)
                        menu_visible = True
                    except Exception:
                        try:
                            self.page.wait_for_selector(options_menu_xpath, timeout=3000)
                            menu_visible = True
                        except Exception:
                            time.sleep(0.5)
                    if not menu_visible:
                        try:
                            options_btn.click()
                            time.sleep(0.5)
                        except Exception:
                            pass
                except Exception:
                    print("⚠️ คลิกตัวเลือกไม่สำเร็จ")
                    return False

                # 3) คลิก 'แก้ไข'
                try:
                    edit_item = self.page.locator('text=แก้ไข')
                    if edit_item.count() == 0:
                        edit_item = self.page.locator('//*[@id="content"]/div[6]/div[1]/div[3]/div[1]/div/div[3]/div/div[2]')
                    if edit_item.count() > 0:
                        edit_item.first.click()
                    else:
                        print("⚠️ ไม่พบเมนู 'แก้ไข'")
                        return False
                except Exception:
                    print("⚠️ คลิก 'แก้ไข' ไม่สำเร็จ")
                    return False

                # 4) รอโหลดหน้าแก้ไข
                try:
                    self.page.wait_for_load_state('domcontentloaded', timeout=7000)
                except Exception:
                    pass
                time.sleep(0.2)

                # 5) เลื่อนลงไปยังช่อง VAT
                try:
                    self.page.evaluate('window.scrollBy(0, 800)')
                except Exception:
                    pass
                try:
                    self.page.wait_for_selector('//*[@id="iptTransactionCESummaryVat"]', timeout=5000)
                except Exception:
                    time.sleep(0.2)

                # 6) กรอก VAT ใหม่
                vat_input = None
                for sel in [
                    '//*[@id="iptTransactionCESummaryVat"]',
                    'css=#iptTransactionCESummaryVat',
                    'css=input#iptTransactionCESummaryVat',
                    'css=input[name="iptTransactionCESummaryVat"]',
                    'xpath=//input[contains(@id, "TransactionCESummaryVat") or contains(@name, "TransactionCESummaryVat")]'
                ]:
                    try:
                        loc = self.page.locator(sel)
                        if loc.count() > 0:
                            vat_input = loc
                            break
                    except Exception:
                        continue
                if vat_input and vat_input.count() > 0:
                    for fill_round in range(2):
                        try:
                            vat_input.first.fill("")
                            vat_input.first.fill(f"{expected_val:.2f}")
                        except Exception:
                            try:
                                vat_input.first.click()
                                vat_input.first.fill("")
                                vat_input.first.fill(f"{expected_val:.2f}")
                            except Exception:
                                pass
                        print(f"✏️ อัปเดต VAT รอบที่ {fill_round+1}/2 เป็น {expected_val:.2f}")
                        try:
                            self.page.evaluate(
                                "(sel)=>{ const el = document.querySelector(sel); if(!el) return; el.dispatchEvent(new Event('input',{bubbles:true})); el.dispatchEvent(new Event('change',{bubbles:true})); el.blur && el.blur(); }",
                                '#iptTransactionCESummaryVat'
                            )
                        except Exception:
                            pass
                        try:
                            self.page.evaluate("()=>{ if (document.activeElement && document.activeElement.blur) { document.activeElement.blur(); } }")
                        except Exception:
                            pass
                        time.sleep(1)
                        try:
                            loading = self.page.locator('//*[@id="tagTransactionLoading"], #tagTransactionLoading')
                            if loading.count() > 0:
                                try:
                                    loading.first.wait_for(state='visible', timeout=1500)
                                except Exception:
                                    pass
                                try:
                                    loading.first.wait_for(state='hidden', timeout=6000)
                                except Exception:
                                    pass
                        except Exception:
                            pass
                else:
                    print("⚠️ ไม่พบช่องกรอก VAT (iptTransactionCESummaryVat/fallback)")
                    return False

                # 7) คลิกบันทึก/แก้ไข ตาม XPaths ที่ระบุ
                clicked_save = False
                click_source = None
                try:
                    # อัปเดตตำแหน่งตามที่ให้มา และระบุเป็น xpath= เพื่อหลีกเลี่ยงการตีความเป็น CSS
                    container_xpath = 'xpath=//*[@id="content"]/div[6]/div[1]/div[6]/div/div[2]'
                    abs_save_xpath = 'xpath=//*[@id="content"]/div[6]/div[1]/div[6]/div/div[2]/div[1]'
                    try:
                        self.page.evaluate('document.querySelector("body").scrollTop = document.body.scrollHeight;')
                    except Exception:
                        try:
                            self.page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                        except Exception:
                            pass
                    try:
                        self.page.wait_for_selector(container_xpath, timeout=5000)
                    except Exception:
                        pass
                    try:
                        container_loc = self.page.locator(container_xpath)
                        if container_loc.count() > 0:
                            try:
                                container_loc.first.scroll_into_view_if_needed()
                                container_loc.first.wait_for(state='visible', timeout=2000)
                            except Exception:
                                pass
                    except Exception:
                        pass
                    try:
                        self.page.wait_for_selector(abs_save_xpath, timeout=4000)
                    except Exception:
                        pass
                    abs_btn = self.page.locator(abs_save_xpath)
                    if abs_btn.count() > 0:
                        print("🔘 พบปุ่มแก้ไขแบบ absolute กำลังกด…")
                        try:
                            abs_btn.first.scroll_into_view_if_needed()
                        except Exception:
                            pass
                        # 1) ปกติ
                        try:
                            abs_btn.first.click()
                            clicked_save = True
                            click_source = 'absolute_xpath.click()'
                        except Exception as e1:
                            print(f"   ↪︎ ปุ่ม absolute.click() ล้มเหลว: {e1}")
                            # 2) force click
                            try:
                                abs_btn.first.click(force=True)
                                clicked_save = True
                                click_source = 'absolute_xpath.click(force)'
                            except Exception as e2:
                                print(f"   ↪︎ ปุ่ม absolute.click(force) ล้มเหลว: {e2}")
                                # 3) JS click
                                try:
                                    self.page.evaluate("el=>el.click()", abs_btn.first)
                                    clicked_save = True
                                    click_source = 'absolute_xpath.js_click'
                                except Exception as e3:
                                    print(f"   ↪︎ ปุ่ม absolute.js_click ล้มเหลว: {e3}")
                                    # 4) mouse click by bounding box
                                    try:
                                        box = abs_btn.first.bounding_box()
                                        if box:
                                            self.page.mouse.click(box['x'] + box['width']/2, box['y'] + box['height']/2)
                                            clicked_save = True
                                            click_source = 'absolute_xpath.mouse_click'
                                    except Exception as e4:
                                        print(f"   ↪︎ ปุ่ม absolute.mouse_click ล้มเหลว: {e4}")
                                        # 5) focus + Enter
                                        try:
                                            abs_btn.first.focus()
                                            self.page.keyboard.press('Enter')
                                            clicked_save = True
                                            click_source = 'absolute_xpath.focus+Enter'
                                        except Exception as e5:
                                            print(f"   ↪︎ ปุ่ม absolute.focus+Enter ล้มเหลว: {e5}")
                                            # 6) submit form
                                            try:
                                                self.page.evaluate("el=>{const f=el.closest('form'); if(f){(f.requestSubmit||f.submit).call(f)} }", abs_btn.first)
                                                clicked_save = True
                                                click_source = 'absolute_xpath.form_submit'
                                            except Exception as e6:
                                                print(f"   ↪︎ ปุ่ม absolute.form_submit ล้มเหลว: {e6}")
                    if not clicked_save:
                        # ให้ความสำคัญกับปุ่ม 'แก้ไข' ที่มี onclick=editAndApproveInvoice
                        special_targets = [
                            'xpath=//div[contains(@class, "ui button button-blue") and contains(@onclick, "editAndApproveInvoice(")]',
                            'css=div.ui.button.button-blue[onclick*="editAndApproveInvoice("]',
                            'text=แก้ไข'
                        ]
                        for sel in special_targets:
                            sp = self.page.locator(sel)
                            if sp.count() > 0:
                                print(f"🔘 พบปุ่มแก้ไขแบบ onclick ({sel}) กำลังกด…")
                                try:
                                    sp.first.scroll_into_view_if_needed()
                                    sp.first.wait_for(state='visible', timeout=3000)
                                except Exception:
                                    pass
                                try:
                                    sp.first.click()
                                    clicked_save = True
                                    click_source = f'{sel}.click()'
                                    break
                                except Exception as e1:
                                    print(f"   ↪︎ {sel}.click() ล้มเหลว: {e1}")
                                    try:
                                        sp.first.click(force=True)
                                        clicked_save = True
                                        click_source = f'{sel}.click(force)'
                                        break
                                    except Exception as e2:
                                        print(f"   ↪︎ {sel}.click(force) ล้มเหลว: {e2}")
                                        try:
                                            self.page.evaluate("() => { if (typeof window.editAndApproveInvoice === 'function') { window.editAndApproveInvoice(1, 3, true); } }")
                                            clicked_save = True
                                            click_source = 'call_editAndApproveInvoice()'
                                            break
                                        except Exception as e3:
                                            print(f"   ↪︎ call_editAndApproveInvoice() ล้มเหลว: {e3}")
                        if not clicked_save:
                            fallback_sels = [
                                'text=บันทึก',
                                'button:has-text("บันทึก")',
                                'css=button[type="submit"]',
                                'text=แก้ไข',
                                'css=div.ui.button.button-blue',
                                'xpath=//div[contains(@class, "ui button button-blue") and contains(normalize-space(.), "แก้ไข")]'
                            ]
                            for sel in fallback_sels:
                                btn = self.page.locator(sel)
                                if btn.count() > 0:
                                    print(f"🔘 พบปุ่ม ({sel}) กำลังกด…")
                                    try:
                                        btn.first.scroll_into_view_if_needed()
                                        time.sleep(0.2)
                                    except Exception:
                                        pass
                                    try:
                                        btn.first.click()
                                        clicked_save = True
                                        click_source = f'{sel}.click()'
                                        break
                                    except Exception as e1:
                                        print(f"   ↪︎ {sel}.click() ล้มเหลว: {e1}")
                                        try:
                                            btn.first.click(force=True)
                                            clicked_save = True
                                            click_source = f'{sel}.click(force)'
                                            break
                                        except Exception as e2:
                                            print(f"   ↪︎ {sel}.click(force) ล้มเหลว: {e2}")
                                            try:
                                                self.page.evaluate("el=>el.click()", btn.first)
                                                clicked_save = True
                                                click_source = f'{sel}.js_click'
                                                break
                                            except Exception as e3:
                                                print(f"   ↪︎ {sel}.js_click ล้มเหลว: {e3}")
                                                try:
                                                    box = btn.first.bounding_box()
                                                    if box:
                                                        self.page.mouse.click(box['x'] + box['width']/2, box['y'] + box['height']/2)
                                                        clicked_save = True
                                                        click_source = f'{sel}.mouse_click'
                                                        break
                                                except Exception as e4:
                                                    print(f"   ↪︎ {sel}.mouse_click ล้มเหลว: {e4}")
                    if clicked_save:
                        print(f"💾 สั่งคลิกปุ่มแก้ไข/บันทึกแล้ว (source={click_source})")
                    else:
                        print("🚫 ไม่พบปุ่มแก้ไข/บันทึกให้คลิก (ตรวจทั้ง absolute, onclick และ fallback แล้ว)")
                        return False
                except Exception as e:
                    print(f"⚠️ เกิดข้อผิดพลาดระหว่างคลิกปุ่มแก้ไข/บันทึก: {e}")
                    return False

                # 8) รอให้หน้าโหลดเสร็จ
                try:
                    self.page.wait_for_load_state('networkidle', timeout=8000)
                except Exception:
                    try:
                        self.page.wait_for_load_state('domcontentloaded', timeout=5000)
                    except Exception:
                        pass
                time.sleep(0.6)

                # 9) ตรวจสอบหลังกดบันทึก (2 รอบ)
                print("🔍 กำลังตรวจสอบ VAT หลังกดบันทึก...")
                for verify_round in range(2):
                    print(f"🔍 ตรวจสอบรอบที่ {verify_round + 1}/2...")
                    
                    # รอให้หน้าโหลดเสร็จ
                    try:
                        self.page.wait_for_load_state('networkidle', timeout=5000)
                    except Exception:
                        try:
                            self.page.wait_for_load_state('domcontentloaded', timeout=3000)
                        except Exception:
                            pass
                    time.sleep(1)
                    
                    # เลื่อนลงไปยังส่วนสรุปอีกครั้งและตรวจซ้ำ
                    try:
                        self.page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                        time.sleep(0.3)
                    except Exception:
                        pass
                    
                    display_loc2 = _select_vat_locator()
                    if not display_loc2 or display_loc2.count() == 0:
                        print(f"⚠️ ไม่พบตำแหน่งแสดง VAT ในรอบที่ {verify_round + 1}")
                        if verify_round < 1:
                            # รีเฟรชหน้าเว็บแล้วลองใหม่
                            try:
                                self.page.reload(timeout=10000)
                                self.page.wait_for_load_state('domcontentloaded', timeout=5000)
                                time.sleep(1)
                            except Exception:
                                pass
                            continue
                        else:
                            break
                    
                    display_text2 = (display_loc2.first.text_content() or '').strip() if display_loc2 and display_loc2.count() > 0 else ''
                    try:
                        display_val2 = float(re.sub(r"[^0-9\.]+", "", display_text2).replace(",", "")) if display_text2 else None
                    except Exception:
                        display_val2 = None
                    
                    print(f"📊 ตรวจสอบรอบที่ {verify_round + 1}: คาดหวัง {expected_val} | หน้าจอ {display_val2}")
                    
                    if display_val2 is not None and abs(display_val2 - expected_val) < 0.001:
                        print(f"✅ VAT หลังแก้ไขตรงตามคาด (ตรวจสอบรอบที่ {verify_round + 1})")
                        return True
                    
                    if verify_round < 1:
                        print(f"⚠️ VAT ยังไม่ตรงในรอบที่ {verify_round + 1} กำลังรีเฟรชหน้าเว็บแล้วตรวจสอบอีกครั้ง...")
                        # รีเฟรชหน้าเว็บแล้วลองใหม่
                        try:
                            self.page.reload(timeout=10000)
                            self.page.wait_for_load_state('domcontentloaded', timeout=5000)
                            time.sleep(1)
                        except Exception:
                            pass
                
                # ถ้าตรวจสอบรอบสองแล้วไม่ตรง ให้แก้ไขใหม่อีกรอบ
                print("⚠️ VAT หลังแก้ไขยังไม่ตรงหลังตรวจสอบสองรอบ กำลังแก้ไขใหม่อีกรอบ...")
                # กลับไปแก้ไขใหม่ (จะวน loop ใหม่)
                continue

            # ถ้าลองครบแล้วยังไม่ตรง
            print("❌ VAT ยังไม่ตรงหลังจากลองแก้ไขหลายรอบ")
            return False
        except Exception as e:
            print(f"❌ ตรวจ/แก้ไข VAT ล้มเหลว: {e}")
            return False

    def refresh_express_page(self):
        """รีหน้า Express เพื่อเริ่มกรอกใหม่ให้สะอาด"""
        try:
            if self.express_link:
                print(f"🔄 รีหน้า Express: {self.express_link}")
                self.page.goto(self.express_link, timeout=10000)
                try:
                    self.page.wait_for_load_state('domcontentloaded', timeout=5000)
                except Exception:
                    pass
                time.sleep(0.3)
                return True
            print("⚠️ ไม่มี express_link สำหรับรีหน้า")
            return False
        except Exception as e:
            print(f"⚠️ รีหน้า Express ล้มเหลว: {e}")
            return False
    
    def check_page_stuck(self, timeout: int = 5) -> bool:
        """ตรวจสอบว่าหน้าเว็บค้างหรือไม่ โดยตรวจสอบว่า element ตอบสนองหรือไม่
        
        Args:
            timeout: เวลาที่รอให้ element ตอบสนอง (วินาที)
        
        Returns:
            True ถ้าหน้าเว็บค้าง, False ถ้าไม่ค้าง
        """
        try:
            # ตรวจสอบว่า iptnumber element ตอบสนองหรือไม่
            iptnumber_element = self.page.locator('//*[@id="iptnumber"]')
            try:
                # ลองอ่านค่า element ภายใน timeout
                iptnumber_element.first.wait_for(state='attached', timeout=timeout * 1000)
                return False  # ไม่ค้าง
            except Exception:
                print(f"⚠️ ตรวจพบว่าหน้าเว็บอาจค้าง (element ไม่ตอบสนองภายใน {timeout} วินาที)")
                return True  # ค้าง
        except Exception as e:
            print(f"⚠️ ตรวจสอบหน้าเว็บค้างล้มเหลว: {e}")
            return True  # ถ้าเกิด error ให้ถือว่าค้าง
    
    def refresh_page(self):
        """รีเฟรชหน้าเว็บ (กด F5 หรือ reload)"""
        try:
            print(f"🔄 กำลังรีเฟรชหน้าเว็บ (F5/reload)...")
            
            # วิธีที่ 1: ใช้ reload
            try:
                self.page.reload(timeout=15000)
                self.page.wait_for_load_state('domcontentloaded', timeout=8000)
                time.sleep(1)
                print(f"✅ รีเฟรชหน้าเว็บสำเร็จ (reload)")
                return True
            except Exception as e1:
                print(f"⚠️ reload ไม่สำเร็จ: {e1} ลองใช้ F5...")
            
            # วิธีที่ 2: ใช้ keyboard shortcut F5
            try:
                self.page.keyboard.press('F5')
                time.sleep(3)
                self.page.wait_for_load_state('domcontentloaded', timeout=8000)
                time.sleep(1)
                print(f"✅ รีเฟรชหน้าเว็บสำเร็จ (F5)")
                return True
            except Exception as e2:
                print(f"⚠️ F5 ไม่สำเร็จ: {e2} ลองใช้ refresh_express_page...")
            
            # วิธีที่ 3: ใช้ refresh_express_page
            return self.refresh_express_page()
            
        except Exception as e:
            print(f"⚠️ รีเฟรชหน้าเว็บล้มเหลว: {e}")
            # ลองใช้ refresh_express_page แทน
            return self.refresh_express_page()
    
    def _calculate_excel_file_hash(self, excel_path: str) -> str:
        """คำนวณ hash ของไฟล์ Excel เพื่อตรวจสอบว่าไฟล์เปลี่ยนหรือไม่"""
        try:
            if not os.path.exists(excel_path):
                return ""
            
            import hashlib
            hash_md5 = hashlib.md5()
            with open(excel_path, "rb") as f:
                # อ่านเฉพาะส่วนแรกของไฟล์ (1024 bytes) เพื่อความเร็ว
                chunk = f.read(1024)
                hash_md5.update(chunk)
            return hash_md5.hexdigest()
        except Exception:
            return ""
    
    def _get_excel_file_mtime(self, excel_path: str) -> float:
        """อ่าน last modified time ของไฟล์ Excel"""
        try:
            if os.path.exists(excel_path):
                return os.path.getmtime(excel_path)
            return 0.0
        except Exception:
            return 0.0
    
    def _validate_status_json(self, excel_path: str, json_path: str) -> bool:
        """ตรวจสอบว่า JSON status file ยังใช้ได้กับไฟล์ Excel ปัจจุบันหรือไม่
        
        Returns:
            True ถ้ายังใช้ได้, False ถ้าไฟล์ Excel เปลี่ยนแล้ว
        """
        try:
            if not os.path.exists(json_path):
                return False
            
            # อ่าน JSON
            with open(json_path, 'r', encoding='utf-8') as f:
                status_data = json.load(f)
            
            # ตรวจสอบว่าเป็นไฟล์ Excel เดียวกันหรือไม่
            stored_excel_path = status_data.get('excel_path', '')
            if stored_excel_path != excel_path:
                return False
            
            # ตรวจสอบ hash (ถ้ามี)
            stored_hash = status_data.get('excel_hash', '')
            if stored_hash:
                current_hash = self._calculate_excel_file_hash(excel_path)
                if current_hash and current_hash != stored_hash:
                    return False
            
            # ตรวจสอบ last modified time (ถ้ามี)
            stored_mtime = status_data.get('excel_mtime', 0)
            if stored_mtime:
                current_mtime = self._get_excel_file_mtime(excel_path)
                if current_mtime and abs(current_mtime - stored_mtime) > 1.0:  # ต่างกันมากกว่า 1 วินาที
                    return False
            
            return True
            
        except Exception:
            return False
    
    def _reset_status_json(self, json_path: str) -> bool:
        """รีเซ็ต JSON status file"""
        try:
            if os.path.exists(json_path):
                os.remove(json_path)
                return True
            return True
        except Exception as e:
            print(f"⚠️ ไม่สามารถลบ JSON status file: {e}")
            return False
    
    def update_excel_status(self, excel_path: str, sheet_name: str, row_index: int, status: str) -> bool:
        """อัปเดตสถานะใน JSON file (แทนการเขียนลง Excel เพื่อหลีกเลี่ยงปัญหา file lock)
        
        Args:
            excel_path: Path ของไฟล์ Excel
            sheet_name: ชื่อชีต
            row_index: หมายเลขแถว (1-based, แถวแรกคือ header)
            status: สถานะ (เช่น "สำเร็จ", "ล้มเหลว", "ข้าม (ไฟล์ไม่พบ)", "เอกสารซ้ำ", "กำลังประมวลผล")
        
        Returns:
            True ถ้าสำเร็จ, False ถ้าไม่สำเร็จ
        """
        try:
            from datetime import datetime
            
            # สร้าง JSON file path (อยู่ในโฟลเดอร์เดียวกับ Excel)
            excel_dir = os.path.dirname(excel_path)
            excel_filename = os.path.basename(excel_path)
            json_filename = f"{excel_filename}.status.json"
            json_path = os.path.join(excel_dir, json_filename)
            
            # อ่านสถานะปัจจุบันจาก JSON (ถ้ามี)
            status_data = {}
            if os.path.exists(json_path):
                try:
                    with open(json_path, 'r', encoding='utf-8') as f:
                        status_data = json.load(f)
                except:
                    status_data = {}
            
            # สร้าง key สำหรับแถวนี้ (sheet_name_row_index)
            status_key = f"{sheet_name}_{row_index}"
            
            # อัปเดตสถานะ
            if 'statuses' not in status_data:
                status_data['statuses'] = {}
            
            status_data['statuses'][status_key] = {
                'status': status,
                'sheet_name': sheet_name,
                'row_index': row_index,
                'updated_at': datetime.now().isoformat()
            }
            
            # บันทึกข้อมูลเพิ่มเติม (รวม hash และ mtime เพื่อตรวจสอบว่าไฟล์เปลี่ยนหรือไม่)
            status_data['excel_path'] = excel_path
            status_data['excel_hash'] = self._calculate_excel_file_hash(excel_path)
            status_data['excel_mtime'] = self._get_excel_file_mtime(excel_path)
            status_data['last_updated'] = datetime.now().isoformat()
            
            # บันทึกลง JSON file
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(status_data, f, ensure_ascii=False, indent=2)
            
            return True
            
        except Exception as e:
            self._log(f"⚠️ ไม่สามารถอัปเดตสถานะใน JSON ได้: {e}", level="warning")
            return False
    
    def _write_renamed_filenames_to_excel(self, excel_path: str) -> bool:
        """เขียนชื่อไฟล์ที่เปลี่ยนแล้วลง Excel โดยใช้ Save As สร้างไฟล์ใหม่
        
        Args:
            excel_path: Path ของไฟล์ Excel เดิม
        
        Returns:
            True ถ้าสำเร็จ, False ถ้าไม่สำเร็จ
        """
        try:
            if not OPENPYXL_AVAILABLE:
                self._log(f"⚠️ openpyxl ไม่พร้อมใช้งาน → ข้ามการเขียนชื่อไฟล์ที่เปลี่ยนแล้ว", level="warning")
                return False
            
            if not hasattr(self, 'renamed_files_mapping') or not self.renamed_files_mapping:
                self._log(f"ℹ️ ไม่มีชื่อไฟล์ที่เปลี่ยนแล้ว → ข้ามการเขียน", level="info")
                return True
            
            self._log(f"📝 กำลังเขียนชื่อไฟล์ที่เปลี่ยนแล้วลง Excel...")
            
            # สร้างชื่อไฟล์ใหม่ (Save As) โดยใช้รูปแบบ OCR_{ชื่อไฟล์เดิม}
            excel_dir = os.path.dirname(excel_path)
            excel_filename = os.path.basename(excel_path)
            excel_name_without_ext = os.path.splitext(excel_filename)[0]
            excel_ext = os.path.splitext(excel_filename)[1]
            new_excel_filename = f"OCR_{excel_filename}"
            new_excel_path = os.path.join(excel_dir, new_excel_filename)
            
            self._log(f"📄 ไฟล์เดิม: {excel_filename}")
            self._log(f"📄 ไฟล์ใหม่: {new_excel_filename}")
            
            # ตรวจสอบว่าไฟล์เดิมถูกเปิดอยู่หรือไม่
            temp_lock_file = os.path.join(excel_dir, f"~${excel_filename}")
            
            if os.path.exists(temp_lock_file):
                self._log(f"⚠️ ไฟล์ Excel ถูกเปิดอยู่ → จะเขียนข้อมูลเมื่อปิดไฟล์แล้ว", level="warning")
                # เก็บข้อมูลไว้ใน JSON เพื่อเขียนทีหลัง
                self._save_renamed_filenames_to_json(excel_path)
                return False
            
            # โหลด Excel จากไฟล์เดิม
            try:
                workbook = load_workbook(excel_path)
            except Exception as e:
                self._log(f"⚠️ ไม่สามารถเปิดไฟล์ Excel ได้: {e} → จะเขียนข้อมูลเมื่อปิดไฟล์แล้ว", level="warning")
                self._save_renamed_filenames_to_json(excel_path)
                return False
            
            column_name = "ชื่อไฟล์ที่เปลี่ยนแล้ว"
            updated_count = 0
            
            # Log ข้อมูล mapping ที่จะเขียน
            self._log(f"📋 พบข้อมูลที่จะเขียน: {len(self.renamed_files_mapping)} รายการ")
            for (sheet_name, row_index), new_filename in self.renamed_files_mapping.items():
                self._log(f"   - ชีต '{sheet_name}' แถวที่ {row_index}: {new_filename}")
            
            try:
                # ประมวลผลแต่ละชีต
                for (sheet_name, row_index), new_filename in self.renamed_files_mapping.items():
                    if sheet_name not in workbook.sheetnames:
                        self._log(f"⚠️ ไม่พบชีต '{sheet_name}' ในไฟล์ Excel", level="warning")
                        continue
                    
                    sheet = workbook[sheet_name]
                    self._log(f"📝 กำลังเขียนข้อมูลลงชีต '{sheet_name}' แถวที่ {row_index}...")
                    
                    # หาคอลัมน์ "ชื่อไฟล์ที่เปลี่ยนแล้ว" (หรือสร้างใหม่ถ้ายังไม่มี)
                    header_row = 1
                    column_index = None
                    
                    # ค้นหาคอลัมน์ที่มีอยู่แล้ว
                    for col_idx, cell in enumerate(sheet[header_row], start=1):
                        if cell.value == column_name:
                            column_index = col_idx
                            self._log(f"✅ พบคอลัมน์ '{column_name}' ที่คอลัมน์ {column_index}")
                            break
                    
                    # ถ้ายังไม่มีคอลัมน์ ให้เพิ่มใหม่
                    if column_index is None:
                        # หาคอลัมน์สุดท้าย
                        max_col = sheet.max_column
                        column_index = max_col + 1
                        # เขียน header
                        header_cell = sheet.cell(row=header_row, column=column_index, value=column_name)
                        
                        # Format header ให้เหมือนกับคอลัมน์อื่นๆ
                        try:
                            # ใช้ format เหมือนกับ invoice_excel_writer.py
                            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                            header_font = Font(name='TH Sarabun New', size=14, bold=True, color="FFFFFF")
                            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            
                            header_cell.fill = header_fill
                            header_cell.font = header_font
                            header_cell.alignment = header_alignment
                            header_cell.border = border
                            
                            # ตั้งความกว้างคอลัมน์ (ประมาณ 40 สำหรับชื่อไฟล์)
                            col_letter = get_column_letter(column_index)
                            sheet.column_dimensions[col_letter].width = 40
                            
                            self._log(f"✅ สร้างคอลัมน์ใหม่ '{column_name}' ที่คอลัมน์ {column_index} พร้อม format header")
                        except Exception as format_error:
                            self._log(f"⚠️ ไม่สามารถ format header ได้: {format_error} → ข้าม format", level="warning")
                            self._log(f"✅ สร้างคอลัมน์ใหม่ '{column_name}' ที่คอลัมน์ {column_index} (ไม่มี format)")
                    
                    # ตรวจสอบว่าแถวที่ row_index มีอยู่จริงหรือไม่
                    if row_index > sheet.max_row:
                        self._log(f"⚠️ แถวที่ {row_index} ไม่มีในชีต '{sheet_name}' (มีแค่ {sheet.max_row} แถว) → ข้าม", level="warning")
                        continue
                    
                    # เขียนค่า new_filename ลงแถวที่ row_index
                    data_cell = sheet.cell(row=row_index, column=column_index, value=new_filename)
                    
                    # Format เซลล์ข้อมูลให้เหมือนกับคอลัมน์อื่นๆ (มีเส้นตาราง)
                    try:
                        # สร้าง border, font, alignment เหมือนกับ invoice_excel_writer.py
                        data_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        data_font = Font(name='TH Sarabun New', size=13)
                        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                        
                        # สลับสีพื้นหลังตามแถว (เหมือนคอลัมน์อื่นๆ)
                        if row_index % 2 == 0:
                            data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        else:
                            data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        
                        data_cell.border = data_border
                        data_cell.font = data_font
                        data_cell.alignment = data_alignment
                        data_cell.fill = data_fill
                        
                        self._log(f"✅ เขียนค่า '{new_filename}' ลงแถวที่ {row_index} คอลัมน์ {column_index} พร้อม format เส้นตาราง")
                    except Exception as format_error:
                        self._log(f"⚠️ ไม่สามารถ format เซลล์ข้อมูลได้: {format_error} → ข้าม format", level="warning")
                        self._log(f"✅ เขียนค่า '{new_filename}' ลงแถวที่ {row_index} คอลัมน์ {column_index} (ไม่มี format)")
                    
                    updated_count += 1
                
                if updated_count == 0:
                    self._log(f"⚠️ ไม่มีข้อมูลที่เขียนได้ (อาจเป็นเพราะแถวไม่พบ)", level="warning")
                    workbook.close()
                    return False
                
                # บันทึกไฟล์ใหม่ (Save As)
                self._log(f"💾 กำลังบันทึกไฟล์ Excel ใหม่: {new_excel_path}")
                try:
                    workbook.save(new_excel_path)
                    self._log(f"✅ บันทึกไฟล์ Excel ใหม่สำเร็จ: {new_excel_filename}")
                except Exception as save_error:
                    self._log(f"❌ ไม่สามารถบันทึกไฟล์ Excel ใหม่ได้: {save_error}", level="error")
                    workbook.close()
                    raise save_error
                
                workbook.close()
                
                # ตรวจสอบว่าไฟล์ถูกบันทึกจริงหรือไม่
                if os.path.exists(new_excel_path):
                    file_size = os.path.getsize(new_excel_path)
                    self._log(f"✅ ยืนยันว่าไฟล์ถูกบันทึกแล้ว: {new_excel_filename} (ขนาด: {file_size} bytes)")
                else:
                    self._log(f"❌ ไม่พบไฟล์ Excel หลังจากบันทึก!", level="error")
                
                self._log(f"✅ เขียนชื่อไฟล์ที่เปลี่ยนแล้วลง Excel สำเร็จ: {updated_count} รายการ")
                self._log(f"📁 ไฟล์เดิม: {excel_filename}")
                self._log(f"📁 ไฟล์ใหม่: {new_excel_filename}")
                return True
                
            except Exception as e:
                workbook.close()
                raise e
                
        except Exception as e:
            self._log(f"⚠️ ไม่สามารถเขียนชื่อไฟล์ที่เปลี่ยนแล้วลง Excel ได้: {e}", level="warning")
            # เก็บข้อมูลไว้ใน JSON เพื่อเขียนทีหลัง
            self._save_renamed_filenames_to_json(excel_path)
            return False
    
    def _save_renamed_filenames_to_json(self, excel_path: str) -> bool:
        """เก็บข้อมูลชื่อไฟล์ที่เปลี่ยนแล้วไว้ใน JSON เพื่อเขียนทีหลัง"""
        try:
            excel_dir = os.path.dirname(excel_path)
            excel_filename = os.path.basename(excel_path)
            json_filename = f"{excel_filename}.renamed_files.json"
            json_path = os.path.join(excel_dir, json_filename)
            
            # อ่านข้อมูลเดิม (ถ้ามี)
            existing_data = {}
            if os.path.exists(json_path):
                try:
                    with open(json_path, 'r', encoding='utf-8') as f:
                        existing_data = json.load(f)
                except:
                    existing_data = {}
            
            # อัปเดตข้อมูล
            if 'renamed_files' not in existing_data:
                existing_data['renamed_files'] = {}
            
            for (sheet_name, row_index), new_filename in self.renamed_files_mapping.items():
                key = f"{sheet_name}_{row_index}"
                existing_data['renamed_files'][key] = {
                    'sheet_name': sheet_name,
                    'row_index': row_index,
                    'new_filename': new_filename
                }
            
            # บันทึก
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(existing_data, f, ensure_ascii=False, indent=2)
            
            self._log(f"💾 เก็บข้อมูลชื่อไฟล์ที่เปลี่ยนแล้วไว้ใน JSON: {json_path}")
            return True
            
        except Exception as e:
            self._log(f"⚠️ ไม่สามารถเก็บข้อมูลชื่อไฟล์ที่เปลี่ยนแล้วได้: {e}", level="warning")
            return False
    
    def read_excel_status(self, excel_path: str, sheet_name: str, row_index: int) -> Optional[str]:
        """อ่านสถานะจาก JSON file (แทนการอ่านจาก Excel เพื่อหลีกเลี่ยงปัญหา file lock)
        
        Args:
            excel_path: Path ของไฟล์ Excel
            sheet_name: ชื่อชีต
            row_index: หมายเลขแถว (1-based, แถวแรกคือ header)
        
        Returns:
            สถานะ หรือ None ถ้าไม่พบ
        """
        try:
            # สร้าง JSON file path (อยู่ในโฟลเดอร์เดียวกับ Excel)
            excel_dir = os.path.dirname(excel_path)
            excel_filename = os.path.basename(excel_path)
            json_filename = f"{excel_filename}.status.json"
            json_path = os.path.join(excel_dir, json_filename)
            
            # อ่านสถานะจาก JSON (ถ้ามี)
            if not os.path.exists(json_path):
                return None
            
            with open(json_path, 'r', encoding='utf-8') as f:
                status_data = json.load(f)
            
            # สร้าง key สำหรับแถวนี้ (sheet_name_row_index)
            status_key = f"{sheet_name}_{row_index}"
            
            if 'statuses' in status_data and status_key in status_data['statuses']:
                return status_data['statuses'][status_key].get('status')
            
            return None
            
        except Exception as e:
            # ไม่ log error เพื่อไม่ให้รบกวน
            return None
    
    def read_excel_data(self, excel_path: str) -> Dict[str, List[Dict]]:
        """อ่านข้อมูลจากไฟล์ Excel และแยกตามชื่อชีต
        
        Args:
            excel_path: Path ของไฟล์ Excel
            
        Returns:
            Dictionary ที่มี key เป็นชื่อชีต และ value เป็น list ของ dictionaries (แต่ละแถว)
        """
        try:
            if not OPENPYXL_AVAILABLE:
                print("❌ openpyxl ไม่พร้อมใช้งาน กรุณาติดตั้ง: pip install openpyxl")
                return {}
            
            print(f"📖 กำลังอ่านข้อมูลจาก Excel: {excel_path}")
            
            # ตรวจสอบว่าไฟล์มีอยู่จริงและสามารถอ่านได้
            if not os.path.exists(excel_path):
                print(f"❌ ไม่พบไฟล์ Excel: {excel_path}")
                return {}
            
            # ตรวจสอบว่ามีไฟล์ชั่วคราวของ Excel หรือไม่ (~$filename.xlsx)
            import tempfile
            excel_dir = os.path.dirname(excel_path)
            excel_filename = os.path.basename(excel_path)
            temp_lock_file = os.path.join(excel_dir, f"~${excel_filename}")
            
            # ถ้ามีไฟล์ lock แสดงว่าไฟล์ถูกเปิดอยู่ใน Excel
            if os.path.exists(temp_lock_file):
                print(f"⚠️ ไฟล์ Excel ถูกเปิดอยู่ใน Excel application (พบไฟล์ lock: {temp_lock_file})")
                print(f"💡 กำลังคัดลอกไฟล์ไปยัง temp directory เพื่ออ่าน...")
                
                # คัดลอกไฟล์ไปยัง temp directory เพื่ออ่าน
                try:
                    temp_dir = tempfile.gettempdir()
                    temp_excel_path = os.path.join(temp_dir, f"temp_{excel_filename}")
                    
                    # คัดลอกไฟล์
                    import shutil
                    shutil.copy2(excel_path, temp_excel_path)
                    print(f"✅ คัดลอกไฟล์สำเร็จ: {temp_excel_path}")
                    
                    # อ่านจากไฟล์ที่คัดลอก
                    try:
                        workbook = load_workbook(temp_excel_path, read_only=True, data_only=True, keep_links=False)
                    except:
                        workbook = load_workbook(temp_excel_path, data_only=True, keep_links=False)
                    
                    # ลบไฟล์ชั่วคราวหลังจากอ่านเสร็จ
                    try:
                        os.remove(temp_excel_path)
                    except:
                        pass
                    
                except Exception as copy_error:
                    print(f"⚠️ ไม่สามารถคัดลอกไฟล์ได้: {copy_error}")
                    print(f"💡 กำลังลองอ่านไฟล์โดยตรง...")
                    # ลองอ่านไฟล์โดยตรง
                    try:
                        workbook = load_workbook(excel_path, read_only=True, data_only=True, keep_links=False)
                    except:
                        workbook = load_workbook(excel_path, data_only=True, keep_links=False)
            else:
                # ไม่มีไฟล์ lock → อ่านไฟล์โดยตรง
                try:
                    workbook = load_workbook(excel_path, read_only=True, data_only=True, keep_links=False)
                except Exception as e1:
                    # ถ้าเปิด read-only ไม่ได้ ให้ลองเปิดแบบปกติ
                    try:
                        workbook = load_workbook(excel_path, data_only=True, keep_links=False)
                    except Exception as e2:
                        # ถ้าเป็น BadZipFile แสดงว่าไฟล์อาจเสียหายหรือถูกเปิดอยู่ → คัดลอกไป temp
                        if "BadZipFile" in str(type(e2).__name__) or "not a zip file" in str(e2).lower():
                            print(f"⚠️ ไฟล์ Excel อาจถูกเปิดอยู่หรือเสียหาย → กำลังคัดลอกไฟล์ไปยัง temp directory...")
                            try:
                                import tempfile
                                import shutil
                                temp_dir = tempfile.gettempdir()
                                temp_excel_path = os.path.join(temp_dir, f"temp_{excel_filename}")
                                shutil.copy2(excel_path, temp_excel_path)
                                print(f"✅ คัดลอกไฟล์สำเร็จ: {temp_excel_path}")
                                
                                try:
                                    workbook = load_workbook(temp_excel_path, read_only=True, data_only=True, keep_links=False)
                                except:
                                    workbook = load_workbook(temp_excel_path, data_only=True, keep_links=False)
                                
                                # ลบไฟล์ชั่วคราวหลังจากอ่านเสร็จ
                                try:
                                    os.remove(temp_excel_path)
                                except:
                                    pass
                            except Exception as copy_error:
                                print(f"❌ ไม่สามารถคัดลอกไฟล์ได้: {copy_error}")
                                print(f"💡 คำแนะนำ: กรุณาปิดไฟล์ Excel และลองใหม่อีกครั้ง")
                                return {}
                        else:
                            raise e2
            
            result = {}
            
            # ชื่อชีตที่ต้องการ (ไม่รวม "ที่อยู่แต่ละบริษัท" เพราะใช้เฉพาะสำหรับกรอกข้อมูลในส่วนของ "+เพิ่มผู้ติดต่อ")
            target_sheets = ["มีภาษีมูลค่าเพิ่ม", "ไม่มีภาษีมูลค่าเพิ่ม"]
            
            for sheet_name in target_sheets:
                if sheet_name not in workbook.sheetnames:
                    print(f"⚠️ ไม่พบชีต: {sheet_name}")
                    continue
                
                sheet = workbook[sheet_name]
                rows_data = []
                
                # อ่าน header (แถวแรก)
                headers = []
                for cell in sheet[1]:
                    headers.append(cell.value if cell.value else "")
                
                print(f"📋 Headers ในชีต '{sheet_name}': {headers}")
                
                # อ่านข้อมูล (เริ่มจากแถวที่ 2)
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                    row_data = {}
                    has_data = False
                    
                    for col_idx, cell in enumerate(row):
                        if col_idx < len(headers):
                            header = headers[col_idx]
                            value = cell.value
                            
                            # แปลงค่า None เป็น empty string
                            if value is None:
                                value = ""
                            # แปลงตัวเลข float ที่เป็น .0 เป็น int string
                            elif isinstance(value, float) and value.is_integer():
                                value = str(int(value))
                            else:
                                value = str(value).strip()
                            
                            row_data[header] = value
                            
                            if value:
                                has_data = True
                    
                    # เพิ่มเฉพาะแถวที่มีข้อมูล
                    if has_data:
                        # อ่านสถานะจาก JSON file (ถ้ามี)
                        status = self.read_excel_status(excel_path, sheet_name, row_idx)
                        if status:
                            row_data['สถานะ'] = status
                        else:
                            row_data['สถานะ'] = 'รอประมวลผล'  # Default status
                        row_data['_row_index'] = row_idx  # เก็บ row_index สำหรับอัปเดตสถานะ
                        rows_data.append(row_data)
                
                if rows_data:
                    result[sheet_name] = rows_data
                    print(f"✅ อ่านข้อมูลจากชีต '{sheet_name}' สำเร็จ: {len(rows_data)} แถว")
                else:
                    print(f"⚠️ ไม่พบข้อมูลในชีต '{sheet_name}'")
            
            workbook.close()
            return result
            
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการอ่าน Excel: {e}")
            import traceback
            traceback.print_exc()
            return {}
    
    def read_address_from_excel(self, excel_path: str, tax_id: str) -> Optional[Dict[str, str]]:
        """อ่านข้อมูลที่อยู่จาก Excel sheet "ที่อยู่แต่ละบริษัท" โดยใช้ tax_id เป็น key
        
        Args:
            excel_path: Path ของไฟล์ Excel
            tax_id: เลขประจำตัวผู้เสียภาษี (13 หลัก)
            
        Returns:
            Dictionary ที่มีข้อมูลที่อยู่ หรือ None ถ้าไม่พบ
        """
        try:
            if not OPENPYXL_AVAILABLE:
                self._log(f"❌ openpyxl ไม่พร้อมใช้งาน", level="error")
                return None
            
            self._log(f"📖 กำลังอ่านข้อมูลที่อยู่จาก Excel: {excel_path} (Tax ID: {tax_id})")
            
            # เปิดไฟล์แบบ read-only เพื่อไม่ให้ล็อกไฟล์
            try:
                workbook = load_workbook(excel_path, read_only=True, data_only=True, keep_links=False)
            except:
                # ถ้าเปิด read-only ไม่ได้ ให้ลองเปิดแบบปกติ
                workbook = load_workbook(excel_path, data_only=True, keep_links=False)
            
            # ตรวจสอบว่ามีชีต "ที่อยู่แต่ละบริษัท" หรือไม่
            if "ที่อยู่แต่ละบริษัท" not in workbook.sheetnames:
                self._log(f"⚠️ ไม่พบชีต 'ที่อยู่แต่ละบริษัท'", level="warning")
                workbook.close()
                return None
            
            sheet = workbook["ที่อยู่แต่ละบริษัท"]
            
            # อ่าน header (แถวแรก)
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value if cell.value else "")
            
            self._log(f"📋 Headers ในชีต 'ที่อยู่แต่ละบริษัท': {headers}")
            
            # หา index ของคอลัมน์ "เลขประจำตัวผู้เสียภาษี"
            tax_id_col_idx = None
            for idx, header in enumerate(headers):
                if header and "เลขประจำตัวผู้เสียภาษี" in str(header):
                    tax_id_col_idx = idx
                    break
            
            if tax_id_col_idx is None:
                self._log(f"❌ ไม่พบคอลัมน์ 'เลขประจำตัวผู้เสียภาษี'", level="error")
                workbook.close()
                return None
            
            # อ่านข้อมูล (เริ่มจากแถวที่ 2) และหา tax_id ที่ตรงกัน
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                if tax_id_col_idx < len(row):
                    cell_value = row[tax_id_col_idx].value
                    # แปลงเป็น string และลบช่องว่าง
                    if cell_value:
                        cell_value_str = str(cell_value).strip().replace('-', '').replace(' ', '')
                        tax_id_clean = tax_id.replace('-', '').replace(' ', '')
                        
                        if cell_value_str == tax_id_clean:
                            # พบข้อมูลที่ตรงกัน - อ่านข้อมูลทั้งหมด
                            self._log(f"✅ พบข้อมูลที่ตรงกันที่แถว {row_idx}")
                            
                            address_data = {}
                            # อ่านข้อมูลจากแต่ละคอลัมน์
                            for col_idx, header in enumerate(headers):
                                if col_idx < len(row):
                                    value = row[col_idx].value
                                    if value is None:
                                        value = ""
                                    elif isinstance(value, float) and value.is_integer():
                                        value = str(int(value))
                                    else:
                                        value = str(value).strip()
                                    
                                    address_data[header] = value
                            
                            workbook.close()
                            self._log(f"📋 ข้อมูลที่อ่านได้: {address_data}")
                            return address_data
            
            workbook.close()
            self._log(f"⚠️ ไม่พบข้อมูลที่ตรงกับ Tax ID: {tax_id}", level="warning")
            return None
            
        except Exception as e:
            self._log(f"❌ เกิดข้อผิดพลาดในการอ่านข้อมูลที่อยู่จาก Excel: {e}", level="error")
            import traceback
            self._log(f"📋 Traceback: {traceback.format_exc()}", level="error")
            return None
    
    def fill_address_in_modal(self, address_data: Dict[str, str]) -> bool:
        """กรอกข้อมูลที่อยู่ใน modal "เพิ่มผู้ขาย/ผู้รับเงิน"
        
        Args:
            address_data: Dictionary ที่มีข้อมูลที่อยู่จาก Excel
            
        Returns:
            True ถ้ากรอกสำเร็จ, False ถ้าไม่สำเร็จ
        """
        try:
            self._log(f"📝 กำลังกรอกข้อมูลที่อยู่ใน modal...")
            
            # 1. กรอกชื่อบริษัท
            company_name = address_data.get('ชื่อบริษัท', '').strip()
            if company_name:
                try:
                    company_field = self.page.locator('#contactmerchantname')
                    if company_field.count() > 0:
                        company_field.first.fill(company_name)
                        self._log(f"✅ กรอกชื่อบริษัท: {company_name}")
                        time.sleep(0.3)
                    else:
                        self._log(f"⚠️ ไม่พบฟิลด์ชื่อบริษัท (#contactmerchantname)", level="warning")
                except Exception as e:
                    self._log(f"⚠️ ไม่สามารถกรอกชื่อบริษัทได้: {e}", level="warning")
            
            # 2. กรอกที่อยู่ (อื่นๆ + เลขที่ + ถนน)
            other_info = address_data.get('อื่นๆ', '').strip()
            building_number = address_data.get('เลขที่', '').strip()
            road = address_data.get('ถนน', '').strip()
            
            # รวมกัน: "อื่นๆ เลขที่ ถนน" (เช่น "MSC Building 571 Sukhumvit 71 Rd.")
            address_parts = []
            if other_info:
                address_parts.append(other_info)
            if building_number:
                address_parts.append(building_number)
            if road:
                address_parts.append(road)
            
            address_full = ' '.join(address_parts).strip()
            if address_full:
                try:
                    address_field = self.page.locator('#customerThAddress')
                    if address_field.count() > 0:
                        address_field.first.fill(address_full)
                        self._log(f"✅ กรอกที่อยู่: {address_full}")
                        time.sleep(0.3)
                    else:
                        self._log(f"⚠️ ไม่พบฟิลด์ที่อยู่ (#customerThAddress)", level="warning")
                except Exception as e:
                    self._log(f"⚠️ ไม่สามารถกรอกที่อยู่ได้: {e}", level="warning")
            
            # 3. กรอกแขวง/ตำบล
            subdistrict = address_data.get('แขวง/ตำบล', '').strip()
            if subdistrict:
                try:
                    subdistrict_field = self.page.locator('#customerThDistrict1')
                    if subdistrict_field.count() > 0:
                        subdistrict_field.first.fill(subdistrict)
                        self._log(f"✅ กรอกแขวง/ตำบล: {subdistrict}")
                        time.sleep(0.3)
                    else:
                        self._log(f"⚠️ ไม่พบฟิลด์แขวง/ตำบล (#customerThDistrict1)", level="warning")
                except Exception as e:
                    self._log(f"⚠️ ไม่สามารถกรอกแขวง/ตำบลได้: {e}", level="warning")
            
            # 4. กรอกเขต/อำเภอ
            district = address_data.get('เขต/อำเภอ', '').strip()
            if district:
                try:
                    district_field = self.page.locator('#customerThDistrict2')
                    if district_field.count() > 0:
                        district_field.first.fill(district)
                        self._log(f"✅ กรอกเขต/อำเภอ: {district}")
                        time.sleep(0.3)
                    else:
                        self._log(f"⚠️ ไม่พบฟิลด์เขต/อำเภอ (#customerThDistrict2)", level="warning")
                except Exception as e:
                    self._log(f"⚠️ ไม่สามารถกรอกเขต/อำเภอได้: {e}", level="warning")
            
            # 5. กรอกจังหวัด
            province = address_data.get('จังหวัด', '').strip()
            if province:
                try:
                    province_field = self.page.locator('#customerThProvince')
                    if province_field.count() > 0:
                        province_field.first.fill(province)
                        self._log(f"✅ กรอกจังหวัด: {province}")
                        time.sleep(0.3)
                    else:
                        self._log(f"⚠️ ไม่พบฟิลด์จังหวัด (#customerThProvince)", level="warning")
                except Exception as e:
                    self._log(f"⚠️ ไม่สามารถกรอกจังหวัดได้: {e}", level="warning")
            
            # 6. กรอกรหัสไปรษณีย์ (ทีละตัว)
            postal_code = address_data.get('รหัสไปรษณีย์', '').strip()
            if postal_code:
                # ลบช่องว่างและตัวอักษรที่ไม่ใช่ตัวเลข
                postal_code_clean = re.sub(r'[^\d]', '', str(postal_code))
                
                if len(postal_code_clean) == 5:
                    try:
                        # กรอกทีละตัว (5 ตัว)
                        for i in range(1, 6):
                            input_id = f"customerThPostCode{i}"
                            digit = postal_code_clean[i-1] if i-1 < len(postal_code_clean) else ""
                            
                            if digit:
                                postcode_field = self.page.locator(f'#{input_id}')
                                if postcode_field.count() > 0:
                                    postcode_field.first.fill(digit)
                                    self._log(f"✅ กรอกตัวเลขรหัสไปรษณีย์ที่ {i}: {digit}")
                                    time.sleep(0.1)
                                else:
                                    self._log(f"⚠️ ไม่พบฟิลด์รหัสไปรษณีย์ที่ {i} (#{input_id})", level="warning")
                    except Exception as e:
                        self._log(f"⚠️ ไม่สามารถกรอกรหัสไปรษณีย์ได้: {e}", level="warning")
                else:
                    self._log(f"⚠️ รหัสไปรษณีย์ไม่ถูกต้อง (ต้อง 5 หลัก): {postal_code_clean}", level="warning")
            
            self._log(f"✅ กรอกข้อมูลที่อยู่เสร็จสิ้น")
            return True
            
        except Exception as e:
            self._log(f"❌ เกิดข้อผิดพลาดในการกรอกข้อมูลที่อยู่: {e}", level="error")
            import traceback
            self._log(f"📋 Traceback: {traceback.format_exc()}", level="error")
            return False
    
    def click_run_with_excel_button(self) -> bool:
        """กดปุ่ม 'รันด้วยไฟล์ Excel'
        
        Returns:
            True ถ้ากดสำเร็จ, False ถ้าไม่สำเร็จ
        """
        try:
            print(f"🔍 กำลังหาปุ่ม 'รันด้วยไฟล์ Excel'...")
            
            # ลองหาปุ่มหลายวิธี
            button_selectors = [
                'text=รันด้วยไฟล์ Excel',
                'button:has-text("รันด้วยไฟล์ Excel")',
                '//button[contains(text(), "รันด้วยไฟล์ Excel")]',
                '//*[contains(text(), "รันด้วยไฟล์ Excel")]',
                'css=button:has-text("รันด้วยไฟล์ Excel")',
            ]
            
            button_found = False
            for selector in button_selectors:
                try:
                    button = self.page.locator(selector)
                    if button.count() > 0:
                        print(f"✅ พบปุ่ม 'รันด้วยไฟล์ Excel' ด้วย selector: {selector}")
                        button.first.scroll_into_view_if_needed()
                        time.sleep(0.3)
                        button.first.click()
                        button_found = True
                        print(f"✅ กดปุ่ม 'รันด้วยไฟล์ Excel' สำเร็จ")
                        time.sleep(1)  # รอให้หน้าโหลด
                        break
                except Exception as e:
                    print(f"⚠️ ไม่พบปุ่มด้วย selector {selector}: {e}")
                    continue
            
            if not button_found:
                print(f"❌ ไม่พบปุ่ม 'รันด้วยไฟล์ Excel'")
                return False
            
            return True
            
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการกดปุ่ม 'รันด้วยไฟล์ Excel': {e}")
            return False
    
    def fill_tax_id_in_modal(self, tax_id: str, excel_path: Optional[str] = None) -> bool:
        """กรอกเลขประจำตัวผู้เสียภาษีใน modal "เพิ่มผู้ขาย/ผู้รับเงิน"
        
        Args:
            tax_id: เลขประจำตัวผู้เสียภาษี (13 หลัก)
            excel_path: Path ของไฟล์ Excel (ถ้ามี) สำหรับอ่านข้อมูลที่อยู่
            
        Returns:
            True ถ้ากรอกสำเร็จ, False ถ้าไม่สำเร็จ
        """
        try:
            self._log(f"🔍 กำลังกรอกเลขประจำตัวผู้เสียภาษีใน modal: {tax_id}")
            
            # รอให้ element #contactthtaxid แสดง (นี่คือ element ที่จะกรอกเลขประจำตัวผู้เสียภาษี)
            self._log(f"⏳ รอให้ element #contactthtaxid แสดง...")
            try:
                # รอให้ element #contactthtaxid แสดงและพร้อมใช้งาน
                self.page.wait_for_selector('#contactthtaxid', timeout=15000, state='visible')
                self._log(f"✅ พบ element #contactthtaxid")
                
                # รออีกนิดเพื่อให้แน่ใจว่า input fields พร้อม
                time.sleep(0.5)
            except Exception as e:
                self._log(f"⚠️ ไม่พบ element #contactthtaxid ภายใน 15 วินาที: {e}", level="warning")
                # ลองรอด้วยวิธีอื่น
                try:
                    self.page.wait_for_selector('div.modal-header', timeout=5000, state='visible')
                    self._log(f"✅ พบ modal header แล้ว ลองหา #contactthtaxid...")
                    time.sleep(1)
                except:
                    pass
                
                # ตรวจสอบอีกครั้งว่า #contactthtaxid มีอยู่หรือไม่
                contact_taxid_container = self.page.locator('#contactthtaxid')
                if contact_taxid_container.count() == 0:
                    self._log(f"❌ ไม่พบ element #contactthtaxid", level="error")
                    return False
            
            # ลบช่องว่างและตัวอักษรที่ไม่ใช่ตัวเลข
            tax_id_clean = re.sub(r'[^\d]', '', str(tax_id))
            
            if len(tax_id_clean) != 13:
                self._log(f"⚠️ เลขประจำตัวผู้เสียภาษีไม่ถูกต้อง (ต้อง 13 หลัก): {tax_id_clean}", level="warning")
                return False
            
            self._log(f"📝 กำลังกรอกเลขประจำตัวผู้เสียภาษี: {tax_id_clean} (13 หลัก)")
            
            # ตรวจสอบว่า element contactthtaxid มีอยู่
            contact_taxid_container = self.page.locator('#contactthtaxid')
            if contact_taxid_container.count() == 0:
                self._log(f"⚠️ ไม่พบ element #contactthtaxid", level="warning")
                return False
            
            self._log(f"✅ พบ element #contactthtaxid")
            
            # กรอกเลขประจำตัวผู้เสียภาษีทีละตัว (13 ตัว) ใน element #contactthtaxid
            for i in range(1, 14):
                input_id = f"mdccipttaxid{i}"
                digit = tax_id_clean[i-1] if i-1 < len(tax_id_clean) else ""
                
                try:
                    # ใช้ selector ที่ชัดเจน: input ภายใน #contactthtaxid
                    input_field = self.page.locator(f'#contactthtaxid input#{input_id}')
                    if input_field.count() > 0:
                        input_field.first.fill(digit)
                        self._log(f"✅ กรอกตัวเลขที่ {i}: {digit} (ใน {input_id})")
                        time.sleep(0.1)  # รอเล็กน้อยระหว่างกรอก
                    else:
                        # Fallback: ลองหาโดยใช้ id โดยตรง
                        input_field_fallback = self.page.locator(f'#{input_id}')
                        if input_field_fallback.count() > 0:
                            input_field_fallback.first.fill(digit)
                            self._log(f"✅ กรอกตัวเลขที่ {i}: {digit} (ใน {input_id} - fallback)")
                            time.sleep(0.1)
                        else:
                            self._log(f"⚠️ ไม่พบ input field: {input_id}", level="warning")
                            return False
                except Exception as e:
                    self._log(f"⚠️ ไม่สามารถกรอกตัวเลขที่ {i} ได้: {e}", level="warning")
                    return False
            
            self._log(f"✅ กรอกเลขประจำตัวผู้เสียภาษีเสร็จสิ้น: {tax_id_clean}")
            
            # คลิกปุ่ม "ค้นหา" เพื่อดึงข้อมูลจากเลขประจำตัวผู้เสียภาษี
            try:
                search_button = self.page.locator('#contactgetinfobtn')
                if search_button.count() > 0:
                    self._log(f"🔍 กำลังคลิกปุ่ม 'ค้นหา' เพื่อดึงข้อมูลจากเลขประจำตัวผู้เสียภาษี...")
                    search_button.first.click()
                    self._log(f"✅ คลิกปุ่ม 'ค้นหา' สำเร็จ")
                    
                    # รอให้ element #mdccperrmsg แสดง (รอจนกว่าจะแสดง ไม่จำกัดเวลา)
                    self._log(f"⏳ รอให้ element #mdccperrmsg แสดง...")
                    try:
                        # รอจนกว่าจะแสดง โดยไม่จำกัดเวลา (timeout=0 หมายถึงรอไม่จำกัดเวลา)
                        self.page.wait_for_selector('#mdccperrmsg', timeout=0, state='visible')
                        self._log(f"✅ พบ element #mdccperrmsg")
                        time.sleep(0.5)  # รอให้ข้อความแสดงเสร็จ
                    except Exception as e:
                        self._log(f"⚠️ เกิดข้อผิดพลาดในการรอ element #mdccperrmsg: {e}", level="warning")
                        return False
                    
                    # ตรวจสอบข้อความและสีของ #mdccperrmsg
                    try:
                        error_msg_element = self.page.locator('#mdccperrmsg')
                        if error_msg_element.count() > 0:
                            error_text = error_msg_element.first.text_content()
                            error_style = error_msg_element.first.evaluate("el => window.getComputedStyle(el).color")
                            
                            self._log(f"📋 ข้อความที่พบ: '{error_text}'")
                            self._log(f"🎨 สีที่พบ: {error_style}")
                            
                            # ตรวจสอบว่าข้อความเป็น "ค้นหาสำเร็จ" (สีเขียว) หรือ "*ไม่พบข้อมูลลูกค้า" (สีแดง)
                            if "ค้นหาสำเร็จ" in error_text or "rgb(39, 190, 149)" in error_style or "rgb(39,190,149)" in error_style:
                                # พบข้อมูลสำเร็จ - คลิกปุ่ม "เพิ่มผู้ขาย/ผู้รับเงิน"
                                self._log(f"✅ พบข้อมูลสำเร็จ - กำลังคลิกปุ่ม 'เพิ่มผู้ขาย/ผู้รับเงิน'...")
                                try:
                                    create_button = self.page.locator('#contactcreatebtn')
                                    if create_button.count() > 0:
                                        create_button.first.click()
                                        self._log(f"✅ คลิกปุ่ม 'เพิ่มผู้ขาย/ผู้รับเงิน' สำเร็จ")
                                        time.sleep(1)
                                    else:
                                        self._log(f"⚠️ ไม่พบปุ่ม 'เพิ่มผู้ขาย/ผู้รับเงิน' (#contactcreatebtn)", level="warning")
                                except Exception as e:
                                    self._log(f"⚠️ ไม่สามารถคลิกปุ่ม 'เพิ่มผู้ขาย/ผู้รับเงิน' ได้: {e}", level="warning")
                                
                            elif "*ไม่พบข้อมูลลูกค้า" in error_text or "rgb(217, 92, 92)" in error_style or "rgb(217,92,92)" in error_style:
                                # ไม่พบข้อมูล - ต้องกรอกข้อมูลจาก Excel
                                self._log(f"⚠️ ไม่พบข้อมูลลูกค้า - กำลังอ่านข้อมูลจาก Excel...")
                                
                                # หา excel path
                                excel_path_to_use = excel_path
                                if not excel_path_to_use:
                                    # หา excel path จาก folder path ที่ใช้อยู่
                                    if hasattr(self, 'current_folder_path') and self.current_folder_path:
                                        excel_path_to_use = os.path.join(self.current_folder_path, 'Invoice_Data.xlsx')
                                    elif hasattr(self, 'excel_path') and self.excel_path:
                                        excel_path_to_use = self.excel_path
                                
                                if not excel_path_to_use or not os.path.exists(excel_path_to_use):
                                    self._log(f"❌ ไม่พบไฟล์ Excel: {excel_path_to_use}", level="error")
                                    return False
                                
                                # อ่านข้อมูลจาก Excel
                                address_data = self.read_address_from_excel(excel_path_to_use, tax_id_clean)
                                
                                if not address_data:
                                    self._log(f"❌ ไม่พบข้อมูลที่อยู่สำหรับ Tax ID: {tax_id_clean}", level="error")
                                    return False
                                
                                # กรอกข้อมูลที่อยู่
                                if self.fill_address_in_modal(address_data):
                                    self._log(f"✅ กรอกข้อมูลที่อยู่สำเร็จ")
                                    time.sleep(1)
                                    
                                    # คลิกปุ่ม "เพิ่มผู้ขาย/ผู้รับเงิน" หลังจากกรอกข้อมูลที่อยู่สำเร็จ
                                    try:
                                        create_button = self.page.locator('#contactcreatebtn')
                                        if create_button.count() > 0:
                                            create_button.first.click()
                                            self._log(f"✅ คลิกปุ่ม 'เพิ่มผู้ขาย/ผู้รับเงิน' สำเร็จ")
                                            time.sleep(1.5)  # รอให้ระบบบันทึกข้อมูล
                                        else:
                                            self._log(f"⚠️ ไม่พบปุ่ม 'เพิ่มผู้ขาย/ผู้รับเงิน' (#contactcreatebtn)", level="warning")
                                    except Exception as e:
                                        self._log(f"⚠️ ไม่สามารถคลิกปุ่ม 'เพิ่มผู้ขาย/ผู้รับเงิน' ได้: {e}", level="warning")
                                else:
                                    self._log(f"❌ ไม่สามารถกรอกข้อมูลที่อยู่ได้", level="error")
                                    return False
                            else:
                                self._log(f"⚠️ พบข้อความที่ไม่ทราบ: '{error_text}'", level="warning")
                        else:
                            self._log(f"⚠️ ไม่พบ element #mdccperrmsg", level="warning")
                    except Exception as e:
                        self._log(f"⚠️ เกิดข้อผิดพลาดในการตรวจสอบข้อความ: {e}", level="warning")
                        import traceback
                        self._log(f"📋 Traceback: {traceback.format_exc()}", level="error")
                else:
                    self._log(f"⚠️ ไม่พบปุ่ม 'ค้นหา' (#contactgetinfobtn)", level="warning")
            except Exception as e:
                self._log(f"⚠️ ไม่สามารถคลิกปุ่ม 'ค้นหา' ได้: {e}", level="warning")
                import traceback
                self._log(f"📋 Traceback: {traceback.format_exc()}", level="error")
            
            return True
            
        except Exception as e:
            self._log(f"❌ เกิดข้อผิดพลาดในการกรอกเลขประจำตัวผู้เสียภาษีใน modal: {e}", level="error")
            import traceback
            self._log(f"📋 Traceback: {traceback.format_exc()}", level="error")
            return False
    
    def check_and_add_new_contact(self, company_name: str, tax_id: str, customer_field) -> bool:
        """ตรวจสอบว่าต้องเพิ่มผู้ติดต่อใหม่หรือไม่ และเพิ่มถ้าจำเป็น
        
        Args:
            company_name: ชื่อบริษัท
            tax_id: เลขประจำตัวผู้เสียภาษี
            customer_field: Locator ของฟิลด์ชื่อบริษัท (เพื่อใช้ลบและกรอกใหม่)
            
        Returns:
            True ถ้าสำเร็จ (ไม่ต้องเพิ่ม หรือเพิ่มสำเร็จ), False ถ้าไม่สำเร็จ
        """
        try:
            self._log(f"🔍 ตรวจสอบว่าต้องเพิ่มผู้ติดต่อใหม่หรือไม่...")
            
            # รอให้ dropdown แสดง
            time.sleep(1)
            
            # ตรวจสอบรายการใน dropdown
            dropdown_items = self.page.locator('css=ul.ui-autocomplete:visible li')
            dropdown_count = dropdown_items.count()
            
            self._log(f"📋 จำนวนรายการใน dropdown: {dropdown_count}")
            
            if dropdown_count == 0:
                self._log(f"⚠️ ไม่พบรายการใน dropdown", level="warning")
                return False
            
            # ตรวจสอบข้อความในรายการทั้งหมด
            items_text = []
            for i in range(dropdown_count):
                try:
                    item = dropdown_items.nth(i)
                    item_text = (item.text_content() or '').strip()
                    items_text.append(item_text)
                    self._log(f"📋 รายการที่ {i+1}: '{item_text}'")
                except:
                    pass
            
            # ตรวจสอบว่ามีรายการอื่นที่ไม่ใช่ "+ เพิ่มผู้ติดต่อ" หรือไม่
            has_other_items = False
            other_item_index = None
            for i, item_text in enumerate(items_text):
                if "+ เพิ่มผู้ติดต่อ" not in item_text and "เพิ่มผู้ติดต่อ" not in item_text:
                    has_other_items = True
                    other_item_index = i
                    self._log(f"✅ พบรายการอื่นที่ไม่ใช่ '+ เพิ่มผู้ติดต่อ': '{item_text}' (index: {i})")
                    break
            
            # ถ้ามีรายการอื่น ให้เลือกรายการนั้นเลย
            if has_other_items and other_item_index is not None:
                try:
                    item = dropdown_items.nth(other_item_index)
                    item.click()
                    self._log(f"✅ เลือกรายการ: '{items_text[other_item_index]}'")
                    time.sleep(0.5)
                    return True
                except Exception as e:
                    self._log(f"⚠️ ไม่สามารถเลือกรายการได้: {e}", level="warning")
                    # ถ้าเลือกไม่ได้ ให้ดำเนินการต่อเหมือนไม่มีรายการอื่น
            
            # ตรวจสอบว่ามีแค่ "+ เพิ่มผู้ติดต่อ" หรือไม่
            has_only_add_contact = (
                dropdown_count == 1 and 
                ("+ เพิ่มผู้ติดต่อ" in items_text[0] or "เพิ่มผู้ติดต่อ" in items_text[0])
            ) or (
                dropdown_count > 1 and 
                not has_other_items  # มีหลายรายการแต่ทั้งหมดเป็น "+ เพิ่มผู้ติดต่อ"
            )
            
            if has_only_add_contact:
                # ถ้ามีแค่ "+ เพิ่มผู้ติดต่อ" ให้ลบชื่อบริษัทและกรอกเลขประจำตัวผู้เสียภาษีแทน
                self._log(f"✅ พบแค่ '+ เพิ่มผู้ติดต่อ' - จะลบชื่อบริษัทและกรอกเลขประจำตัวผู้เสียภาษีแทน")
                
                # ลบชื่อบริษัท
                try:
                    customer_field.first.clear()
                    self._log(f"✅ ลบชื่อบริษัทสำเร็จ")
                    time.sleep(0.5)
                except Exception as e:
                    self._log(f"⚠️ ไม่สามารถลบชื่อบริษัทได้: {e}", level="warning")
                
                # กรอกเลขประจำตัวผู้เสียภาษีแทน
                if tax_id:
                    try:
                        customer_field.first.fill(tax_id)
                        self._log(f"✅ กรอกเลขประจำตัวผู้เสียภาษีแทน: {tax_id}")
                        time.sleep(1)  # รอให้ dropdown แสดง
                        
                        # ตรวจสอบว่ามีรายการใน dropdown หรือไม่
                        new_dropdown_items = self.page.locator('css=ul.ui-autocomplete:visible li')
                        new_dropdown_count = new_dropdown_items.count()
                        
                        if new_dropdown_count > 0:
                            # ตรวจสอบรายการทั้งหมด
                            new_items_text = []
                            for i in range(new_dropdown_count):
                                try:
                                    item = new_dropdown_items.nth(i)
                                    item_text = (item.text_content() or '').strip()
                                    new_items_text.append(item_text)
                                    self._log(f"📋 รายการใหม่ที่ {i+1}: '{item_text}'")
                                except:
                                    pass
                            
                            # ตรวจสอบว่ามีรายการอื่นที่ไม่ใช่ "+ เพิ่มผู้ติดต่อ" หรือไม่
                            has_other_new_items = False
                            other_new_item_index = None
                            for i, item_text in enumerate(new_items_text):
                                if "+ เพิ่มผู้ติดต่อ" not in item_text and "เพิ่มผู้ติดต่อ" not in item_text:
                                    has_other_new_items = True
                                    other_new_item_index = i
                                    self._log(f"✅ พบรายการอื่นที่ไม่ใช่ '+ เพิ่มผู้ติดต่อ': '{item_text}' (index: {i})")
                                    break
                            
                            # ถ้ามีรายการอื่น ให้เลือกรายการนั้นเลย
                            if has_other_new_items and other_new_item_index is not None:
                                try:
                                    item = new_dropdown_items.nth(other_new_item_index)
                                    item.click()
                                    self._log(f"✅ เลือกรายการ: '{new_items_text[other_new_item_index]}'")
                                    time.sleep(0.5)
                                    return True
                                except Exception as e:
                                    self._log(f"⚠️ ไม่สามารถเลือกรายการได้: {e}", level="warning")
                                    # ถ้าเลือกไม่ได้ ให้ดำเนินการต่อเหมือนไม่มีรายการอื่น
                            
                            # ตรวจสอบว่ายังมีแค่ "+ เพิ่มผู้ติดต่อ" หรือไม่
                            first_new_item = new_dropdown_items.first
                            first_new_item_text = (first_new_item.text_content() or '').strip()
                            
                            if "+ เพิ่มผู้ติดต่อ" in first_new_item_text or "เพิ่มผู้ติดต่อ" in first_new_item_text:
                                # ถ้ายังมีแค่ "+ เพิ่มผู้ติดต่อ" ให้คลิกที่ "+ เพิ่มผู้ติดต่อ"
                                self._log(f"✅ ยังพบแค่ '+ เพิ่มผู้ติดต่อ' - กำลังคลิก...")
                                first_new_item.click()
                                time.sleep(1.5)  # รอให้ modal เปิด
                                
                                # รอให้ modal แสดง (รอให้ element #contactthtaxid แสดง)
                                try:
                                    # รอให้ element #contactthtaxid แสดง (นี่คือ element ที่จะกรอกเลขประจำตัวผู้เสียภาษี)
                                    self.page.wait_for_selector('#contactthtaxid', timeout=15000, state='visible')
                                    self._log(f"✅ พบ element #contactthtaxid (modal แสดงแล้ว)")
                                    
                                    # ตรวจสอบว่าเป็น modal "เพิ่มผู้ขาย/ผู้รับเงิน" (optional check)
                                    modal_title = self.page.locator('h4.modal-title#contentModalHeader')
                                    if modal_title.count() > 0:
                                        title_text = (modal_title.first.text_content() or '').strip()
                                        if "เพิ่มผู้ขาย/ผู้รับเงิน" in title_text:
                                            self._log(f"✅ ยืนยันว่าเป็น modal 'เพิ่มผู้ขาย/ผู้รับเงิน'")
                                        else:
                                            self._log(f"⚠️ Modal title ไม่ตรง: '{title_text}' แต่จะดำเนินการต่อ", level="warning")
                                    else:
                                        self._log(f"⚠️ ไม่พบ modal title แต่จะดำเนินการต่อ", level="warning")
                                    
                                    # กรอกเลขประจำตัวผู้เสียภาษีใน modal (ใน element #contactthtaxid)
                                    # หา excel_path ถ้ามี
                                    excel_path_for_modal = None
                                    if hasattr(self, 'excel_path') and self.excel_path:
                                        excel_path_for_modal = self.excel_path
                                    elif hasattr(self, 'current_folder_path') and self.current_folder_path:
                                        excel_path_for_modal = os.path.join(self.current_folder_path, 'Invoice_Data.xlsx')
                                    
                                    if self.fill_tax_id_in_modal(tax_id, excel_path_for_modal):
                                        self._log(f"✅ กรอกเลขประจำตัวผู้เสียภาษีใน modal สำเร็จ")
                                        
                                        # รอให้ระบบตรวจสอบ VAT
                                        time.sleep(1.5)
                                        
                                        # คลิกปุ่มบันทึกใน modal (ถ้ามี)
                                        try:
                                            # ลองหาปุ่มบันทึกหลายแบบ
                                            save_selectors = [
                                                'button:has-text("บันทึก")',
                                                'button:has-text("Save")',
                                                'button[type="submit"]',
                                                '.modal-footer button.btn-primary',
                                                '.modal-footer button:has-text("บันทึก")'
                                            ]
                                            
                                            save_clicked = False
                                            for selector in save_selectors:
                                                try:
                                                    save_button = self.page.locator(selector)
                                                    if save_button.count() > 0:
                                                        save_button.first.click()
                                                        self._log(f"✅ คลิกปุ่มบันทึกใน modal สำเร็จ (ใช้ selector: {selector})")
                                                        time.sleep(1.5)
                                                        save_clicked = True
                                                        break
                                                except:
                                                    continue
                                            
                                            if not save_clicked:
                                                self._log(f"⚠️ ไม่พบปุ่มบันทึกใน modal แต่ถือว่าสำเร็จแล้ว", level="warning")
                                            
                                            return True
                                        except Exception as e:
                                            self._log(f"⚠️ ไม่สามารถคลิกปุ่มบันทึกได้: {e} แต่ถือว่าสำเร็จแล้ว", level="warning")
                                            return True
                                    else:
                                        self._log(f"❌ ไม่สามารถกรอกเลขประจำตัวผู้เสียภาษีใน modal ได้", level="error")
                                        return False
                                except Exception as e:
                                    self._log(f"⚠️ ไม่พบ element #contactthtaxid ภายใน 15 วินาที: {e}", level="warning")
                                    # ลองรออีกครั้งด้วยวิธีอื่น
                                    try:
                                        self.page.wait_for_selector('div.modal-header', timeout=5000, state='visible')
                                        self._log(f"✅ พบ modal header แล้ว ลองกรอกเลขประจำตัวผู้เสียภาษี...")
                                        # หา excel_path ถ้ามี
                                        excel_path_for_modal = None
                                        if hasattr(self, 'excel_path') and self.excel_path:
                                            excel_path_for_modal = self.excel_path
                                        elif hasattr(self, 'current_folder_path') and self.current_folder_path:
                                            excel_path_for_modal = os.path.join(self.current_folder_path, 'Invoice_Data.xlsx')
                                        
                                        if self.fill_tax_id_in_modal(tax_id, excel_path_for_modal):
                                            return True
                                    except:
                                        pass
                                    return False
                            else:
                                # ถ้ามีรายการอื่นที่ไม่ใช่ "+ เพิ่มผู้ติดต่อ" ให้เลือกรายการนั้น
                                for i in range(new_dropdown_count):
                                    try:
                                        item = new_dropdown_items.nth(i)
                                        item_text = (item.text_content() or '').strip()
                                        if "+ เพิ่มผู้ติดต่อ" not in item_text and "เพิ่มผู้ติดต่อ" not in item_text:
                                            item.click()
                                            self._log(f"✅ เลือกรายการ: '{item_text}'")
                                            time.sleep(0.5)
                                            return True
                                    except:
                                        continue
                                
                                # ถ้าไม่พบรายการอื่น ให้กด Enter
                                customer_field.first.press('Enter')
                                self._log(f"✅ กด Enter เพื่อยืนยัน")
                                time.sleep(0.5)
                                return True
                        else:
                            # ถ้าไม่มี dropdown ให้กด Enter
                            customer_field.first.press('Enter')
                            self._log(f"✅ กด Enter เพื่อยืนยัน (ไม่มี dropdown)")
                            time.sleep(0.5)
                            return True
                    except Exception as e:
                        self._log(f"❌ ไม่สามารถกรอกเลขประจำตัวผู้เสียภาษีได้: {e}", level="error")
                        return False
                else:
                    self._log(f"⚠️ ไม่มีเลขประจำตัวผู้เสียภาษีให้กรอก", level="warning")
                    return False
            else:
                # ถ้ามีมากกว่า "+ เพิ่มผู้ติดต่อ" ให้เลือกรายการที่ไม่ใช่ "+ เพิ่มผู้ติดต่อ"
                self._log(f"✅ พบหลายรายการ - จะเลือกรายการที่ไม่ใช่ '+ เพิ่มผู้ติดต่อ'")
                
                for i in range(dropdown_count):
                    try:
                        item = dropdown_items.nth(i)
                        item_text = (item.text_content() or '').strip()
                        
                        # ข้าม "+ เพิ่มผู้ติดต่อ"
                        if "+ เพิ่มผู้ติดต่อ" in item_text or "เพิ่มผู้ติดต่อ" in item_text:
                            continue
                        
                        # เลือกรายการแรกที่ไม่ใช่ "+ เพิ่มผู้ติดต่อ"
                        item.click()
                        self._log(f"✅ เลือกรายการ: '{item_text}'")
                        time.sleep(0.5)
                        return True
                    except Exception as e:
                        self._log(f"⚠️ ไม่สามารถคลิกรายการที่ {i+1} ได้: {e}", level="warning")
                        continue
                
                # ถ้าไม่พบรายการอื่น ให้กด Enter
                self._log(f"⚠️ ไม่พบรายการอื่น - จะกด Enter", level="warning")
                customer_field.first.press('Enter')
                time.sleep(0.5)
                return True
                
        except Exception as e:
            self._log(f"❌ เกิดข้อผิดพลาดในการตรวจสอบ/เพิ่มผู้ติดต่อ: {e}", level="error")
            import traceback
            self._log(f"📋 Traceback: {traceback.format_exc()}", level="error")
            return False
    
    def process_excel_data_and_fill_form(self, excel_path: str) -> bool:
        """ประมวลผลข้อมูลจาก Excel และกรอกข้อมูลบนเว็บ
        
        Args:
            excel_path: Path ของไฟล์ Excel
            
        Returns:
            True ถ้าสำเร็จ, False ถ้าไม่สำเร็จ
        """
        try:
            self._log(f"🚀 เริ่มประมวลผลข้อมูลจาก Excel: {excel_path}")
            self._status_update(step="กำลังอ่านข้อมูลจาก Excel", file='-')
            
            # เพิ่มตัวแปรเก็บ mapping ของชื่อไฟล์ที่เปลี่ยนแล้ว
            self.renamed_files_mapping = {}  # {(sheet_name, row_index): new_filename}
            
            # ตรวจสอบประเภทของบริษัทจาก folder_settings.json (สำคัญมาก!)
            excel_dir = os.path.dirname(excel_path) if os.path.exists(excel_path) else None
            folder_code = None
            folder_group = None
            
            if excel_dir:
                try:
                    # หา folder_code จาก excel_dir
                    folder_code = self.file_manager.get_folder_code_from_path(Path(excel_dir))
                    
                    if folder_code:
                        self._log(f"📊 พบ folder_code: {folder_code}")
                        
                        # อ่าน folder_settings.json
                        folder_settings_path = r"V:\A.โฟร์เดอร์หลัก\Build000 ทดสอบระบบ\folder_settings\folder_settings.json"
                        if os.path.exists(folder_settings_path):
                            try:
                                with open(folder_settings_path, 'r', encoding='utf-8') as f:
                                    folder_settings = json.load(f)
                                
                                if folder_code in folder_settings:
                                    folder_info = folder_settings[folder_code]
                                    folder_group = folder_info.get('group', 'unknown')
                                    folder_message = folder_info.get('message', '')
                                    self._log(f"✅ พบ folder_code '{folder_code}' ใน folder_settings.json")
                                    self._log(f"📋 ประเภทบริษัท: {folder_group} ({folder_message})")
                                    
                                    # เก็บ folder_group ไว้ใช้ในภายหลัง
                                    self.current_folder_group = folder_group
                                else:
                                    self._log(f"⚠️ ไม่พบ folder_code '{folder_code}' ใน folder_settings.json", level="warning")
                                    self._log(f"⚠️ ระบบจะใช้ค่า default (regular)", level="warning")
                                    self.current_folder_group = 'regular'  # Default
                            except Exception as e:
                                self._log(f"⚠️ เกิดข้อผิดพลาดในการอ่าน folder_settings.json: {e}", level="warning")
                                self.current_folder_group = 'regular'  # Default
                        else:
                            self._log(f"⚠️ ไม่พบไฟล์ folder_settings.json ที่: {folder_settings_path}", level="warning")
                            self.current_folder_group = 'regular'  # Default
                    else:
                        self._log(f"⚠️ ไม่พบ folder_code จากพาธ Excel: {excel_dir}", level="warning")
                        self._log(f"⚠️ ระบบจะใช้ค่า default (regular)", level="warning")
                        self.current_folder_group = 'regular'  # Default
                except Exception as e:
                    self._log(f"⚠️ เกิดข้อผิดพลาดในการตรวจสอบ folder_code: {e}", level="warning")
                    self.current_folder_group = 'regular'  # Default
            else:
                self._log(f"⚠️ ไม่พบ excel_dir", level="warning")
                self.current_folder_group = 'regular'  # Default
            
            if not folder_group:
                self._log(f"⚠️ ไม่สามารถระบุประเภทบริษัทได้ - ใช้ค่า default (regular)", level="warning")
                folder_group = 'regular'
            
            # รีหน้า Express ก่อนเริ่มประมวลผล (เหมือนระบบอ่าน PDF)
            self.refresh_express_page()
            
            # ตั้งค่า excel_path และ current_folder_path สำหรับใช้ใน fill_tax_id_in_modal
            self.excel_path = excel_path
            # หา folder path จาก excel path
            if os.path.exists(excel_path):
                self.current_folder_path = os.path.dirname(excel_path)
            
            # 0.5. ตรวจสอบและรีเซ็ต JSON status file ถ้าไฟล์ Excel เปลี่ยน
            excel_dir = os.path.dirname(excel_path) if os.path.exists(excel_path) else None
            if excel_dir:
                excel_filename = os.path.basename(excel_path)
                json_filename = f"{excel_filename}.status.json"
                json_path = os.path.join(excel_dir, json_filename)
                
                # ตรวจสอบว่า JSON status file ยังใช้ได้กับไฟล์ Excel ปัจจุบันหรือไม่
                if os.path.exists(json_path):
                    if not self._validate_status_json(excel_path, json_path):
                        self._log(f"⚠️ ตรวจพบว่าไฟล์ Excel เปลี่ยนแล้ว → รีเซ็ต JSON status file")
                        if self._reset_status_json(json_path):
                            self._log(f"✅ รีเซ็ต JSON status file สำเร็จ (จะเริ่มทำงานใหม่ทั้งหมด)")
                        else:
                            self._log(f"⚠️ ไม่สามารถรีเซ็ต JSON status file ได้", level="warning")
                    else:
                        self._log(f"✅ JSON status file ยังใช้ได้กับไฟล์ Excel ปัจจุบัน")
            
            # 1. อ่านข้อมูลจาก Excel
            excel_data = self.read_excel_data(excel_path)
            
            if not excel_data:
                self._log(f"❌ ไม่พบข้อมูลใน Excel", level="error")
                return False
            
            # 1.5. ตรวจสอบคอลัมน์ที่จำเป็นก่อนเริ่มการทำงาน
            required_columns = [
                'ลำดับ',
                'ชื่อบริษัท',
                'เลขประจำตัวผู้เสียภาษี',
                'วันที่',
                'ชื่อบัญชี / โค้ดบัญชี',
                'ยอดก่อนภาษีมูลค่าเพิ่ม',
                'ยอดภาษีมูลค่าเพิ่ม',
                'ยอดหลังบวกภาษีมูลค่าเพิ่ม',
                'ชื่อไฟล์ใหม่',
                'ชื่อไฟล์เก่า'
            ]
            
            validation_errors = []
            
            # ตรวจสอบเฉพาะชีต "มีภาษีมูลค่าเพิ่ม" และ "ไม่มีภาษีมูลค่าเพิ่ม" ที่มีข้อมูลเท่านั้น
            sheets_to_check = ["มีภาษีมูลค่าเพิ่ม", "ไม่มีภาษีมูลค่าเพิ่ม"]
            
            for sheet_name in sheets_to_check:
                # ตรวจสอบว่าเป็นชีตที่มีข้อมูลใน excel_data หรือไม่
                if sheet_name not in excel_data:
                    # ชีตไม่มีข้อมูล → ข้ามการตรวจสอบ
                    self._log(f"ℹ️ ข้ามการตรวจสอบชีต '{sheet_name}' เพราะไม่มีข้อมูล", level="info")
                    continue
                
                rows_data = excel_data[sheet_name]
                
                # ตรวจสอบว่ามีข้อมูลจริงหรือไม่ (ไม่ใช่ list ว่าง)
                if not rows_data or len(rows_data) == 0:
                    # ชีตมี key แต่ไม่มีข้อมูล → ข้ามการตรวจสอบ
                    self._log(f"ℹ️ ข้ามการตรวจสอบชีต '{sheet_name}' เพราะไม่มีข้อมูล", level="info")
                    continue
                
                # มีข้อมูล → ตรวจสอบความครบถ้วน
                for row_idx, row_data in enumerate(rows_data, start=2):  # เริ่มจากแถวที่ 2 (เพราะแถวแรกเป็น header)
                    missing_columns = []
                    
                    for col_name in required_columns:
                        value = row_data.get(col_name, '').strip()
                        if not value:
                            missing_columns.append(col_name)
                    
                    if missing_columns:
                        validation_errors.append({
                            'sheet': sheet_name,
                            'row': row_idx,
                            'missing_columns': missing_columns
                        })
            
            # ถ้ามีข้อผิดพลาด ให้แจ้งเตือนและหยุดการทำงาน
            if validation_errors:
                error_message = "❌ พบข้อมูลไม่ครบใน Excel:\n\n"
                for error in validation_errors:
                    error_message += f"📋 ชีต '{error['sheet']}' แถวที่ {error['row']}:\n"
                    error_message += f"   ⚠️ ขาดคอลัมน์: {', '.join(error['missing_columns'])}\n\n"
                
                error_message += "กรุณาตรวจสอบและกรอกข้อมูลให้ครบถ้วนก่อนดำเนินการต่อ"
                
                # เก็บ error message ไว้ใน instance variable เพื่อให้ web_app.py เรียกใช้ได้
                self.last_error_message = error_message
                
                self._log(error_message, level="error")
                self._status_update(step="พบข้อมูลไม่ครบ - กรุณาตรวจสอบ Excel", file='-')
                
                # แจ้งเตือนผ่าน status update (เพื่อแสดงในหน้าเว็บ)
                if hasattr(self, '_status_callback') and self._status_callback:
                    try:
                        self._status_callback({
                            'step': 'error',
                            'message': error_message,
                            'file': '-',
                            'folder': None
                        })
                    except:
                        pass
                
                return False
            
            # นับจำนวนลำดับที่ไม่ซ้ำกัน (ไม่นับแถวทั้งหมด) เพื่อให้ผู้ใช้เห็นจำนวนที่ถูกต้อง
            total_unique_sequences = 0
            for sheet_name, rows_data in excel_data.items():
                # จัดกลุ่มข้อมูลตามลำดับเพื่อนับจำนวนลำดับที่ไม่ซ้ำกัน
                sequences_in_sheet = set()
                for row in rows_data:
                    sequence = row.get('ลำดับ', '').strip()
                    if sequence:
                        sequences_in_sheet.add(sequence)
                total_unique_sequences += len(sequences_in_sheet)
            
            total_rows = sum(len(rows) for rows in excel_data.values())
            self._log(f"📊 พบข้อมูลใน Excel: {len(excel_data)} ชีต, {total_rows} แถว, {total_unique_sequences} ลำดับ")
            self._log(f"✅ ตรวจสอบคอลัมน์ที่จำเป็นเสร็จสิ้น - ข้อมูลครบถ้วน")
            # ใช้จำนวนลำดับที่ไม่ซ้ำกันแทนจำนวนแถวทั้งหมด
            self._notify_progress(total_delta=total_unique_sequences, reset=True)
            
            # 1.6. ตรวจสอบสถานะจาก Excel (Resume mode)
            completed_count = 0
            for sheet_name, rows_data in excel_data.items():
                for row in rows_data:
                    status = row.get('สถานะ', 'รอประมวลผล')
                    if status == 'สำเร็จ':
                        completed_count += 1
            
            if completed_count > 0:
                self._log(f"📋 พบรายการที่เสร็จแล้ว: {completed_count} รายการ → จะข้ามรายการเหล่านี้")
            
            # 2. ประมวลผลแต่ละชีต (ไม่ต้องกดปุ่ม "รันด้วยไฟล์ Excel" เพราะไม่มีปุ่มนี้บนหน้า Peak Engine)
            processed_count = 0
            for sheet_name, rows_data in excel_data.items():
                self._log(f"\n📊 เริ่มประมวลผลชีต: {sheet_name}")
                self._log(f"📋 จำนวนแถว: {len(rows_data)}")
                self._status_update(step=f"กำลังประมวลผลชีต: {sheet_name}", file='-')
                
                # ตรวจสอบว่าเป็นชีต "มีภาษีมูลค่าเพิ่ม" หรือ "ไม่มีภาษีมูลค่าเพิ่ม"
                is_vat_sheet = (sheet_name == "มีภาษีมูลค่าเพิ่ม")
                
                # จัดกลุ่มข้อมูลตามลำดับ
                grouped_by_sequence = {}
                for row in rows_data:
                    sequence = row.get('ลำดับ', '').strip()
                    if sequence:
                        if sequence not in grouped_by_sequence:
                            grouped_by_sequence[sequence] = []
                        grouped_by_sequence[sequence].append(row)
                
                self._log(f"📊 จำนวนลำดับที่แตกต่างกัน: {len(grouped_by_sequence)}")
                
                # ประมวลผลแต่ละลำดับ
                for sequence, rows in sorted(grouped_by_sequence.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 0):
                    # ตรวจสอบสถานะจาก Excel (Resume mode)
                    first_row = rows[0] if rows else None
                    row_index = first_row.get('_row_index') if first_row else None
                    
                    if first_row:
                        excel_status = first_row.get('สถานะ', 'รอประมวลผล')
                        
                        # ข้ามรายการที่เสร็จแล้ว
                        if excel_status == 'สำเร็จ':
                            self._log(f"⏭️ ข้ามลำดับ {sequence} ในชีต '{sheet_name}' (เสร็จแล้ว - สถานะ: {excel_status})")
                            self._notify_progress(success_delta=0)  # ไม่นับซ้ำ แต่ไม่เพิ่ม success
                            continue
                        
                        # ข้ามรายการที่ถูกข้ามแล้ว (ไฟล์ไม่พบ)
                        if excel_status == 'ข้าม (ไฟล์ไม่พบ)':
                            self._log(f"⏭️ ข้ามลำดับ {sequence} ในชีต '{sheet_name}' (ถูกข้ามแล้ว - ไฟล์ไม่พบ)")
                            continue
                        
                        # อัปเดตสถานะเป็น "กำลังประมวลผล"
                        if row_index:
                            self.update_excel_status(excel_path, sheet_name, row_index, "กำลังประมวลผล")
                    
                    # ตรวจสอบไฟล์ PDF ก่อนเริ่มทำงาน
                    self._log(f"\n🔍 กำลังตรวจสอบไฟล์ PDF สำหรับลำดับ: {sequence} (ชีต: {sheet_name})")
                    
                    # ตรวจสอบไฟล์ PDF
                    excel_dir = os.path.dirname(excel_path)
                    missing_files = []
                    found_files = []
                    
                    for row in rows:
                        old_filename = row.get('ชื่อไฟล์เก่า', '').strip()
                        if old_filename:
                            file_path = os.path.join(excel_dir, old_filename)
                            if os.path.exists(file_path):
                                found_files.append(old_filename)
                            else:
                                missing_files.append(old_filename)
                    
                    file_validation = {
                        'valid': len(missing_files) == 0,
                        'missing_files': missing_files,
                        'found_files': found_files
                    }
                    
                    if not file_validation['valid']:
                        missing_files = file_validation['missing_files']
                        self._log(f"❌ ไม่พบไฟล์ PDF สำหรับลำดับ {sequence} (ชีต: {sheet_name}):", level="error")
                        for missing_file in missing_files:
                            self._log(f"   - {missing_file}", level="error")
                        
                        # อัปเดตสถานะใน Excel
                        if first_row and row_index:
                            self.update_excel_status(excel_path, sheet_name, row_index, "ข้าม (ไฟล์ไม่พบ)")
                        
                        self._notify_progress(failure_delta=1)
                        continue
                    
                    self._log(f"✅ ตรวจสอบไฟล์ PDF สำเร็จ: พบ {len(file_validation['found_files'])} ไฟล์")
                    
                    # รีหน้า Express ก่อนเริ่มลำดับใหม่ (เหมือนระบบอ่าน PDF)
                    self.refresh_express_page()
                    
                    self._log(f"\n🔢 ประมวลผลลำดับ: {sequence} ({len(rows)} แถว)")
                    self._status_update(step=f"กำลังประมวลผลลำดับ: {sequence}", file=f"ลำดับ {sequence}")
                    
                    # กรอกข้อมูลแถวแรก
                    if len(rows) > 0:
                        first_row = rows[0]
                        company_name = first_row.get('ชื่อบริษัท', '').strip()
                        self._log(f"📝 กรอกข้อมูลแถวแรก... (บริษัท: {company_name})")
                        self._status_update(step=f"กำลังกรอกข้อมูลลำดับ {sequence} แถวที่ 1", file=company_name or f"ลำดับ {sequence}")
                        
                        # ถ้า folder_group = 'special' ให้ทำงานเหมือน "ไม่มีภาษีมูลค่าเพิ่ม" ทั้งสองชีต
                        effective_is_vat_sheet = is_vat_sheet if self.current_folder_group != 'special' else False
                        if self.fill_form_from_excel_row(first_row, effective_is_vat_sheet, row_index=1, is_first_row=True, folder_group=self.current_folder_group):
                            self._log(f"✅ กรอกข้อมูลแถวแรกสำเร็จ")
                            
                            # ถ้ามีแถวที่ 2, 3, ... (ลำดับเดียวกัน) ให้กรอกต่อทันทีโดยไม่ต้องรอผลบันทึก
                            self._log(f"🔍 ตรวจสอบจำนวนแถว: {len(rows)} แถว")
                            if len(rows) > 1:
                                self._log(f"📊 พบ {len(rows)} แถวในลำดับเดียวกัน - จะกรอกแถวถัดไปทันที")
                                
                                # กรอกแถวที่ 2, 3, ... (ลำดับเดียวกัน)
                                self._log(f"🔄 เริ่มวนลูปกรอกแถวที่ 2-{len(rows)} (range: 1 ถึง {len(rows)-1})")
                                for row_idx in range(1, len(rows)):
                                    self._log(f"🔄 วนลูป: row_idx={row_idx}, len(rows)={len(rows)}")
                                    additional_row = rows[row_idx]
                                    row_number = row_idx + 1  # แถวที่ 2, 3, 4, ...
                                    self._log(f"📝 กรอกข้อมูลแถวที่ {row_number}...")
                                    self._status_update(step=f"กำลังกรอกข้อมูลลำดับ {sequence} แถวที่ {row_number}", file=company_name or f"ลำดับ {sequence}")
                                    
                                    # คลิกปุ่ม "+ เพิ่มรายการ"
                                    try:
                                        # ลองหาปุ่มหลายวิธี
                                        add_item_button = None
                                        button_selectors = [
                                            'div.button-main.button-green.different-border.float-left:has-text("+ เพิ่มรายการ")',
                                            'div.button-main.button-green:has-text("+ เพิ่มรายการ")',
                                            'div.button-green:has-text("+ เพิ่มรายการ")',
                                            'div:has-text("+ เพิ่มรายการ")',
                                            '//div[contains(@class, "button-main") and contains(@class, "button-green") and contains(text(), "+ เพิ่มรายการ")]',
                                            '//div[contains(text(), "+ เพิ่มรายการ")]',
                                            'text=+ เพิ่มรายการ',
                                        ]
                                        
                                        for selector in button_selectors:
                                            try:
                                                add_item_button = self.page.locator(selector)
                                                if add_item_button.count() > 0:
                                                    self._log(f"✅ พบปุ่ม '+ เพิ่มรายการ' ด้วย selector: {selector}")
                                                    break
                                            except:
                                                continue
                                        
                                        if add_item_button and add_item_button.count() > 0:
                                            # ลอง scroll เข้าหา element ก่อน
                                            try:
                                                add_item_button.first.scroll_into_view_if_needed()
                                                time.sleep(0.3)
                                            except:
                                                pass
                                            
                                            add_item_button.first.click()
                                            self._log(f"✅ คลิกปุ่ม '+ เพิ่มรายการ' สำเร็จ")
                                            time.sleep(1.5)  # รอให้แถวใหม่แสดง
                                        else:
                                            self._log(f"⚠️ ไม่พบปุ่ม '+ เพิ่มรายการ' ด้วย selector ใดๆ", level="warning")
                                            # ลองใช้ JavaScript เป็น fallback
                                            try:
                                                self.page.evaluate("""
                                                    const buttons = document.querySelectorAll('div.button-main.button-green');
                                                    for (let btn of buttons) {
                                                        if (btn.textContent.includes('+ เพิ่มรายการ')) {
                                                            btn.click();
                                                            break;
                                                        }
                                                    }
                                                """)
                                                self._log(f"✅ คลิกปุ่ม '+ เพิ่มรายการ' สำเร็จด้วย JavaScript")
                                                time.sleep(1.5)
                                            except Exception as js_e:
                                                self._log(f"⚠️ ไม่สามารถคลิกปุ่ม '+ เพิ่มรายการ' ด้วย JavaScript ได้: {js_e}", level="warning")
                                    except Exception as e:
                                        self._log(f"⚠️ ไม่สามารถคลิกปุ่ม '+ เพิ่มรายการ' ได้: {e}", level="warning")
                                        import traceback
                                        self._log(f"📋 Traceback: {traceback.format_exc()}", level="error")
                                    
                                    # กรอกข้อมูลแถวเพิ่มเติม (ไม่ใช่แถวแรก)
                                    if self.fill_form_from_excel_row(additional_row, effective_is_vat_sheet, row_index=row_number, is_first_row=False, folder_group=self.current_folder_group):
                                        self._log(f"✅ กรอกข้อมูลแถวที่ {row_number} สำเร็จ")
                                    else:
                                        self._log(f"❌ กรอกข้อมูลแถวที่ {row_number} ไม่สำเร็จ", level="error")
                                        break  # หยุดถ้ากรอกไม่สำเร็จ
                                
                                # หลังจากกรอกแถวทั้งหมดเสร็จแล้ว → ทำขั้นตอนสุดท้าย (หมายเหตุ, เลือกรวมภาษี, คลิกปุ่ม)
                                self._log(f"📊 กรอกข้อมูลทั้งหมด {len(rows)} แถวเสร็จแล้ว - กำลังทำขั้นตอนสุดท้าย...")
                                
                                # 5.5. กรอกหมายเหตุ (ใช้ข้อมูลจากแถวแรก)
                                remark = first_row.get('หมายเหตุ', '').strip()
                                if remark:
                                    # ทำความสะอาดข้อความหมายเหตุ (ลบ .pdf, None_vat_, VAT_, WHT_)
                                    remark_cleaned = self._clean_remark_text(remark)
                                    self._log(f"📝 กำลังกรอกหมายเหตุ: {remark_cleaned} (เดิม: {remark})")
                                    try:
                                        remark_field = self.page.locator('#tarremark')
                                        if remark_field.count() > 0:
                                            remark_field.first.fill(remark_cleaned)
                                            self._log(f"✅ กรอกหมายเหตุสำเร็จ")
                                            time.sleep(0.5)
                                        else:
                                            self._log(f"⚠️ ไม่พบฟิลด์หมายเหตุ (#tarremark)", level="warning")
                                    except Exception as e:
                                        self._log(f"⚠️ เกิดข้อผิดพลาดในการกรอกหมายเหตุ: {e}", level="warning")
                                
                                # 7. เลือกดรอปดาวน์ภาษี (เฉพาะ is_vat_sheet)
                                if is_vat_sheet:
                                    # สำหรับ VAT: เลือก "รวมภาษี" ใน ddltaxstatus
                                    self._log(f"🔍 [VAT] กำลังเลือกดรอปดาวน์รวมภาษี (ddltaxstatus)...")
                                    try:
                                        tax_status_dropdown = self.page.locator('//*[@id="ddltaxstatus"]')
                                        if tax_status_dropdown.count() > 0:
                                            tax_status_dropdown.first.click()
                                            time.sleep(1)
                                            
                                            tax_included_option = self.page.locator('text=รวมภาษี, text=Include Tax')
                                            if tax_included_option.count() > 0:
                                                tax_included_option.first.click()
                                                self._log(f"✅ เลือกรวมภาษีสำเร็จ")
                                            else:
                                                # Fallback: เลือกด้วย index
                                                try:
                                                    self.page.select_option('//*[@id="ddltaxstatus"]', index=1)
                                                    self._log(f"✅ เลือกรวมภาษีสำเร็จด้วยวิธี fallback")
                                                except:
                                                    self._log(f"⚠️ ไม่สามารถเลือกรวมภาษีได้", level="warning")
                                    except Exception as e:
                                        self._log(f"⚠️ เกิดข้อผิดพลาดในการเลือกรวมภาษี: {e}", level="warning")
                                
                                # 8. คลิกปุ่ม hidePaymentModal
                                self._log(f"🔍 กำลังคลิกปุ่ม hidePaymentModal...")
                                try:
                                    hide_payment_button = self.page.locator('//*[@id="hidePaymentModal"]')
                                    if hide_payment_button.count() > 0:
                                        hide_payment_button.first.click()
                                        self._log(f"✅ คลิกปุ่ม hidePaymentModal สำเร็จ")
                                    else:
                                        self._log(f"⚠️ ไม่พบปุ่ม hidePaymentModal", level="warning")
                                except Exception as e:
                                    self._log(f"⚠️ เกิดข้อผิดพลาดในการคลิกปุ่ม hidePaymentModal: {e}", level="warning")
                                
                                # 9. คลิกปุ่มสุดท้าย
                                self._log(f"🔍 กำลังกดปุ่มสุดท้าย...")
                                time.sleep(1.5)
                                document_number = None  # ประกาศตัวแปรไว้ก่อน
                                should_continue_processing = True  # ตัวแปรสำหรับควบคุมการทำงานต่อ
                                try:
                                    final_button = self.page.locator('//*[@id="content"]/div[6]/div[6]/div/div[2]/div[2]/div[1]')
                                    if final_button.count() > 0:
                                        final_button.first.click()
                                        self._log(f"✅ คลิกปุ่มสุดท้ายสำเร็จ")
                                        
                                        # 9.5. ตรวจสอบ element #dvredalert ทันทีหลังจากคลิกปุ่มสุดท้าย
                                        time.sleep(1)  # รอให้ alert แสดง
                                        try:
                                            red_alert = self.page.locator('#dvredalert')
                                            if red_alert.count() > 0 and red_alert.first.is_visible():
                                                try:
                                                    alert_label = self.page.locator('#lbredalert')
                                                    alert_text = ''
                                                    if alert_label.count() > 0:
                                                        alert_text = (alert_label.first.text_content() or '').strip()
                                                    else:
                                                        alert_text = (red_alert.first.text_content() or '').strip()
                                                    
                                                    self._log(f"📋 ข้อความใน alert: '{alert_text}'")
                                                    
                                                    # ตรวจสอบว่ามีข้อความ "เลขที่ใบกำกับภาษีอ้างอิงนี้สำหรับผู้ติดต่อนี้เคยถูกใช้งานแล้ว"
                                                    if "เลขที่ใบกำกับภาษีอ้างอิงนี้สำหรับผู้ติดต่อนี้เคยถูกใช้งานแล้ว" in alert_text:
                                                        self._log(f"⚠️ พบเอกสารซ้ำ - กำลังย้ายไฟล์ทันที...")
                                                        self._notify_progress(duplicate_delta=1)
                                                        
                                                        # ย้ายไฟล์ต้นฉบับไปโฟลเดอร์ เอกสารซ้ำรอตรวจ
                                                        old_filename = first_row.get('ชื่อไฟล์เก่า', '').strip()
                                                        if old_filename:
                                                            excel_dir = os.path.dirname(excel_path)
                                                            original_file_path = os.path.join(excel_dir, old_filename)
                                                            if os.path.exists(original_file_path):
                                                                moved = self.file_manager.move_file_to_duplicate_folder(original_file_path)
                                                                if moved:
                                                                    self._log(f"✅ ย้ายไฟล์ต้นฉบับไปยัง 'เอกสารซ้ำรอตรวจ': {moved}")
                                                                else:
                                                                    self._log("❌ ย้ายไฟล์ต้นฉบับไปยังโฟลเดอร์เอกสารซ้ำรอตรวจไม่สำเร็จ", level="warning")
                                                            else:
                                                                self._log(f"⚠️ ไม่พบไฟล์ต้นฉบับ: {original_file_path}", level="warning")
                                                        
                                                        should_continue_processing = False  # หยุดการทำงานต่อ
                                                        # ข้ามไปลำดับถัดไป
                                                        continue
                                                    
                                                    # ตรวจสอบว่ามีข้อความ "โปรดกรอกข้อมูลในช่อง"
                                                    elif "โปรดกรอกข้อมูลในช่อง" in alert_text:
                                                        self._log(f"⚠️ พบข้อความ 'โปรดกรอกข้อมูลในช่อง' - กำลังรีเซ็ตและกรอกข้อมูลใหม่...")
                                                        
                                                        # เก็บข้อมูลสำหรับ retry
                                                        self._current_excel_rows_for_retry = rows
                                                        self._current_excel_sequence_info_for_retry = {
                                                            'sequence': sequence,
                                                            'is_vat_sheet': is_vat_sheet,
                                                            'excel_path': excel_path
                                                        }
                                                        
                                                        # วนลูปสูงสุด 5 ครั้ง
                                                        max_retry = 5
                                                        retry_success = False
                                                        
                                                        for retry_count in range(1, max_retry + 1):
                                                            self._log(f"🔄 เริ่มรีเซ็ตและกรอกข้อมูลใหม่ (ครั้งที่ {retry_count}/{max_retry})...")
                                                            self._refill_attempt_count = retry_count - 1
                                                            
                                                            try:
                                                                # รีหน้าแล้วกรอกใหม่ทั้งชุด
                                                                self.refresh_express_page()
                                                                time.sleep(1)  # รอให้หน้าเว็บโหลดเสร็จ
                                                                
                                                                # กรอกข้อมูลแถวแรก
                                                                if len(rows) > 0:
                                                                    effective_is_vat_sheet = is_vat_sheet if getattr(self, 'current_folder_group', 'regular') != 'special' else False
                                                                    if self.fill_form_from_excel_row(rows[0], effective_is_vat_sheet, row_index=1, is_first_row=True, folder_group=getattr(self, 'current_folder_group', 'regular')):
                                                                        self._log(f"✅ กรอกข้อมูลแถวแรกสำเร็จ")
                                                                        
                                                                        # กรอกแถวที่ 2, 3, ... (ถ้ามี)
                                                                        if len(rows) > 1:
                                                                            for row_idx in range(1, len(rows)):
                                                                                additional_row = rows[row_idx]
                                                                                row_number = row_idx + 1
                                                                                
                                                                                # คลิกปุ่ม "+ เพิ่มรายการ"
                                                                                try:
                                                                                    add_item_button = self.page.locator('div.button-main.button-green.different-border.float-left:has-text("+ เพิ่มรายการ")')
                                                                                    if add_item_button.count() == 0:
                                                                                        add_item_button = self.page.locator('//div[contains(text(), "+ เพิ่มรายการ")]')
                                                                                    if add_item_button.count() > 0:
                                                                                        add_item_button.first.click()
                                                                                        self._log(f"✅ คลิกปุ่ม '+ เพิ่มรายการ' สำหรับแถวที่ {row_number}")
                                                                                        time.sleep(0.5)
                                                                                except Exception as e:
                                                                                    self._log(f"⚠️ ไม่สามารถคลิกปุ่ม '+ เพิ่มรายการ' ได้: {e}")
                                                                                
                                                                                # กรอกข้อมูลแถวที่ row_number
                                                                                self.fill_form_from_excel_row(
                                                                                    additional_row,
                                                                                    is_vat_sheet,
                                                                                    row_index=row_number,
                                                                                    is_first_row=False
                                                                                )
                                                                                self._log(f"✅ กรอกข้อมูลแถวที่ {row_number} สำเร็จ")
                                                                        
                                                                        # ทำขั้นตอนสุดท้าย (หมายเหตุ, เลือกรวมภาษี, คลิกปุ่ม)
                                                                        # กรอกหมายเหตุ
                                                                        remark = rows[0].get('หมายเหตุ', '').strip()
                                                                        if remark:
                                                                            remark_cleaned = self._clean_remark_text(remark)
                                                                            try:
                                                                                remark_field = self.page.locator('#tarremark')
                                                                                if remark_field.count() > 0:
                                                                                    remark_field.first.fill(remark_cleaned)
                                                                                    time.sleep(0.5)
                                                                            except:
                                                                                pass
                                                                        
                                                                        # เลือกรวมภาษี (ถ้า is_vat_sheet)
                                                                        if is_vat_sheet:
                                                                            try:
                                                                                tax_status_dropdown = self.page.locator('//*[@id="ddltaxstatus"]')
                                                                                if tax_status_dropdown.count() > 0:
                                                                                    tax_status_dropdown.first.click()
                                                                                    time.sleep(1)
                                                                                    tax_included_option = self.page.locator('text=รวมภาษี, text=Include Tax')
                                                                                    if tax_included_option.count() > 0:
                                                                                        tax_included_option.first.click()
                                                                                    else:
                                                                                        self.page.select_option('//*[@id="ddltaxstatus"]', index=1)
                                                                            except:
                                                                                pass
                                                                        
                                                                        # คลิกปุ่ม hidePaymentModal
                                                                        try:
                                                                            hide_payment_button = self.page.locator('//*[@id="hidePaymentModal"]')
                                                                            if hide_payment_button.count() > 0:
                                                                                hide_payment_button.first.click()
                                                                        except:
                                                                            pass
                                                                        
                                                                        # คลิกปุ่มสุดท้าย
                                                                        time.sleep(1.5)
                                                                        try:
                                                                            final_button = self.page.locator('//*[@id="content"]/div[6]/div[6]/div/div[2]/div[2]/div[1]')
                                                                            if final_button.count() > 0:
                                                                                final_button.first.click()
                                                                                self._log(f"✅ คลิกปุ่มสุดท้ายสำเร็จ (ครั้งที่ {retry_count})")
                                                                                time.sleep(2)  # รอให้หน้าเว็บโหลด
                                                                                
                                                                                # ตรวจสอบอีกครั้งว่ามี alert หรือไม่
                                                                                red_alert_check = self.page.locator('#dvredalert')
                                                                                if red_alert_check.count() > 0 and red_alert_check.first.is_visible():
                                                                                    alert_label_check = self.page.locator('#lbredalert')
                                                                                    alert_text_check = ''
                                                                                    if alert_label_check.count() > 0:
                                                                                        alert_text_check = (alert_label_check.first.text_content() or '').strip()
                                                                                    else:
                                                                                        alert_text_check = (red_alert_check.first.text_content() or '').strip()
                                                                                    
                                                                                    # ถ้ายังมี "โปรดกรอกข้อมูลในช่อง" ให้ลองใหม่
                                                                                    if "โปรดกรอกข้อมูลในช่อง" in alert_text_check:
                                                                                        self._log(f"⚠️ ยังพบ 'โปรดกรอกข้อมูลในช่อง' (ครั้งที่ {retry_count}) - จะลองใหม่...")
                                                                                        continue  # ลองใหม่
                                                                                    # ถ้ามี "เลขที่ใบกำกับภาษีอ้างอิงนี้สำหรับผู้ติดต่อนี้เคยถูกใช้งานแล้ว" ให้ย้ายไฟล์
                                                                                    elif "เลขที่ใบกำกับภาษีอ้างอิงนี้สำหรับผู้ติดต่อนี้เคยถูกใช้งานแล้ว" in alert_text_check:
                                                                                        self._log(f"⚠️ พบเอกสารซ้ำ (ครั้งที่ {retry_count}) - กำลังย้ายไฟล์...")
                                                                                        old_filename = rows[0].get('ชื่อไฟล์เก่า', '').strip()
                                                                                        if old_filename:
                                                                                            excel_dir = os.path.dirname(excel_path)
                                                                                            original_file_path = os.path.join(excel_dir, old_filename)
                                                                                            if os.path.exists(original_file_path):
                                                                                                moved = self.file_manager.move_file_to_duplicate_folder(original_file_path)
                                                                                                if moved:
                                                                                                    self._log(f"✅ ย้ายไฟล์ต้นฉบับไปยัง 'เอกสารซ้ำรอตรวจ': {moved}")
                                                                                        should_continue_processing = False
                                                                                        break  # ออกจาก retry loop
                                                                                else:
                                                                                    # ไม่มี alert แสดงว่าอาจจะสำเร็จ
                                                                                    self._log(f"✅ ไม่พบ alert - อาจจะบันทึกสำเร็จ (ครั้งที่ {retry_count})")
                                                                                    retry_success = True
                                                                                    break  # ออกจาก retry loop
                                                                        except Exception as e:
                                                                            self._log(f"⚠️ เกิดข้อผิดพลาดในการคลิกปุ่มสุดท้าย (ครั้งที่ {retry_count}): {e}")
                                                                            
                                                            except Exception as e:
                                                                self._log(f"⚠️ เกิดข้อผิดพลาดในการรีเซ็ตและกรอกข้อมูลใหม่ (ครั้งที่ {retry_count}): {e}")
                                                        
                                                        # ล้างข้อมูล retry
                                                        self._current_excel_rows_for_retry = None
                                                        self._current_excel_sequence_info_for_retry = None
                                                        
                                                        if not retry_success:
                                                            self._log(f"⚠️ ไม่สามารถกรอกข้อมูลใหม่ได้สำเร็จหลังจากลอง {max_retry} ครั้ง", level="warning")
                                                            should_continue_processing = False
                                                            continue  # ข้ามไปลำดับถัดไป
                                                        
                                                except Exception as e:
                                                    self._log(f"⚠️ เกิดข้อผิดพลาดในการตรวจสอบ alert: {e}", level="warning")
                                        except Exception as e:
                                            self._log(f"⚠️ เกิดข้อผิดพลาดในการตรวจสอบ #dvredalert: {e}", level="warning")
                                        
                                        # ถ้าไม่ต้องหยุดการทำงานต่อ ให้ตรวจสอบ header element (เพื่ออ่าน document_number สำหรับการเปลี่ยนชื่อไฟล์)
                                        if should_continue_processing:
                                            # 9.6. รอจนกว่าหน้าเว็บจะโหลดเสร็จและ element h3 แสดงข้อมูล
                                            # สำหรับ special → ยังคงต้องอ่าน document_number จาก header เพื่อใช้ในการเปลี่ยนชื่อไฟล์
                                            self._log(f"⏳ กำลังรอให้หน้าเว็บโหลดเสร็จ...")
                                            try:
                                                # รอจนกว่า element h3 จะแสดงข้อมูล
                                                header_element = self.page.locator('h3.ui.header.heading.section-header-doc-left')
                                                header_element.wait_for(state='visible', timeout=0)  # รอไม่จำกัดเวลา
                                                self._log(f"✅ พบ header element แล้ว")
                                                
                                                # อ่านเลขที่เอกสารจาก element (ดึงข้อมูลหลังจาก #)
                                                # สำหรับ special → ใช้สำหรับการเปลี่ยนชื่อไฟล์เท่านั้น (ไม่กรอกในฟอร์ม)
                                                try:
                                                    header_text = header_element.first.text_content() or ''
                                                    self._log(f"📄 ข้อความใน header: {header_text}")
                                                    
                                                    # หาเลขที่เอกสารหลังจาก #
                                                    match = re.search(r'#([A-Z0-9\-]+)', header_text)
                                                    if match:
                                                        document_number = match.group(1)
                                                        self._log(f"✅ อ่านเลขที่เอกสารสำเร็จ: {document_number}")
                                                        if getattr(self, 'current_folder_group', 'regular') == 'special':
                                                            self._log(f"📋 folder_group='special' → ใช้ document_number สำหรับการเปลี่ยนชื่อไฟล์เท่านั้น")
                                                    else:
                                                        self._log(f"⚠️ ไม่พบเลขที่เอกสารใน header", level="warning")
                                                except Exception as e:
                                                    self._log(f"⚠️ เกิดข้อผิดพลาดในการอ่านเลขที่เอกสาร: {e}", level="warning")
                                            except Exception as e:
                                                self._log(f"⚠️ ไม่พบ header element: {e}", level="warning")
                                    else:
                                        self._log(f"⚠️ ไม่พบปุ่มสุดท้าย", level="warning")
                                except Exception as e:
                                    self._log(f"⚠️ เกิดข้อผิดพลาดในการคลิกปุ่มสุดท้าย: {e}", level="warning")
                                
                                # ถ้าไม่ต้องทำงานต่อ ให้ข้ามไปลำดับถัดไป
                                if not should_continue_processing:
                                    continue
                                
                                # 10. ตรวจสอบ VAT (เฉพาะชีต "มีภาษีมูลค่าเพิ่ม" และไม่ใช่ special)
                                if is_vat_sheet and getattr(self, 'current_folder_group', 'regular') != 'special':
                                    vat_amount = first_row.get('ยอดภาษีมูลค่าเพิ่ม', '').strip()
                                    if vat_amount:
                                        try:
                                            vat_amount_clean = re.sub(r'[^\d.]', '', str(vat_amount))
                                            vat_amount_float = float(vat_amount_clean)
                                            if vat_amount_float > 0:  # ตรวจสอบว่ามีค่าจริง
                                                self._log(f"🔍 กำลังตรวจสอบยอดภาษีมูลค่าเพิ่ม: {vat_amount_float}")
                                                
                                                # ใช้ฟังก์ชัน check_and_fix_vat_value ที่มีอยู่แล้ว
                                                if not self.check_and_fix_vat_value(str(vat_amount_float)):
                                                    self._log(f"⚠️ ตรวจสอบ/แก้ไข VAT ไม่สำเร็จ", level="warning")
                                        except Exception as e:
                                            self._log(f"⚠️ เกิดข้อผิดพลาดในการตรวจสอบ VAT: {e}", level="warning")
                                elif getattr(self, 'current_folder_group', 'regular') == 'special':
                                    self._log(f"📋 folder_group='special' → ข้ามการตรวจสอบยอดภาษีมูลค่าเพิ่ม")
                                
                                # หลังจากทำขั้นตอนสุดท้ายเสร็จแล้ว → รอผลบันทึก
                                # เก็บข้อมูล Excel rows (ทั้งหมดของลำดับ) ไว้สำหรับ retry (เหมือนระบบอ่าน PDF)
                                self._current_excel_rows_for_retry = rows  # เก็บ rows ทั้งหมดของลำดับ
                                self._current_excel_sequence_info_for_retry = {
                                    'sequence': sequence,
                                    'is_vat_sheet': is_vat_sheet,
                                    'excel_path': excel_path
                                }
                                self._refill_attempt_count = 0  # รีเซ็ต counter สำหรับ retry
                                
                                self._log(f"📊 ทำขั้นตอนสุดท้ายเสร็จแล้ว - กำลังรอผลบันทึก...")
                                self._status_update(step="กำลังรอผลบันทึก", file=company_name or f"ลำดับ {sequence}")
                                result = self.wait_for_save_result(timeout=20)
                                self._log(f"📋 ผลการบันทึก: {result}")
                                
                                # ล้างข้อมูล retry หลังจากรอผลเสร็จ
                                self._current_excel_rows_for_retry = None
                                self._current_excel_sequence_info_for_retry = None
                                
                                if result == 'approved':
                                    self._log(f"✅ บันทึกสำเร็จ (ลำดับ {sequence}, {len(rows)} แถว)")
                                    self._notify_progress(success_delta=1)
                                    processed_count += 1
                                    
                                    # 11. เปลี่ยนชื่อไฟล์และอัปโหลด (ถ้ามีเลขที่เอกสาร)
                                    # สำหรับ special → ใช้ document_number จาก header (ที่อ่านมาแล้ว) สำหรับการเปลี่ยนชื่อไฟล์
                                    new_filename = None
                                    upload_success = False
                                    if document_number:
                                        # ใช้ effective_is_vat_sheet (ซึ่งจะเป็น False สำหรับ special)
                                        new_filename = self._handle_file_rename_and_upload(
                                            first_row, document_number, effective_is_vat_sheet, excel_path
                                        )
                                        upload_success = (new_filename is not None)
                                        
                                        # เก็บ mapping สำหรับเขียนลง Excel ทีหลัง
                                        if new_filename and first_row and row_index:
                                            key = (sheet_name, row_index)
                                            self.renamed_files_mapping[key] = new_filename
                                            self._log(f"💾 เก็บ mapping: ชีต '{sheet_name}' แถวที่ {row_index} → {new_filename}")
                                        else:
                                            if not new_filename:
                                                self._log(f"⚠️ ไม่มี new_filename → ไม่เก็บ mapping", level="warning")
                                            if not first_row:
                                                self._log(f"⚠️ ไม่มี first_row → ไม่เก็บ mapping", level="warning")
                                            if not row_index:
                                                self._log(f"⚠️ ไม่มี row_index → ไม่เก็บ mapping", level="warning")
                                    else:
                                        self._log(f"⚠️ ไม่พบเลขที่เอกสาร → ข้ามการเปลี่ยนชื่อไฟล์", level="warning")
                                    
                                    # อัปเดตสถานะใน Excel (หลังจากอัปโหลดไฟล์สำเร็จ)
                                    if upload_success or not document_number:  # ถ้าอัปโหลดสำเร็จ หรือไม่มี document_number (ข้ามการอัปโหลด)
                                        if first_row and row_index:
                                            self.update_excel_status(excel_path, sheet_name, row_index, "สำเร็จ")
                                        self._log(f"💾 อัปเดตสถานะใน Excel: ลำดับ {sequence} (ชีต: {sheet_name}) → สำเร็จ")
                                elif result == 'duplicate':
                                    self._log(f"⚠️ เอกสารซ้ำ (ลำดับ {sequence})", level="warning")
                                    self._notify_progress(duplicate_delta=1)
                                    
                                    # อัปเดตสถานะใน Excel
                                    if first_row and row_index:
                                        self.update_excel_status(excel_path, sheet_name, row_index, "เอกสารซ้ำ")
                                    
                                    # ย้ายไฟล์ต้นฉบับไปโฟลเดอร์ เอกสารซ้ำรอตรวจ (เหมือนระบบอ่าน PDF)
                                    old_filename = first_row.get('ชื่อไฟล์เก่า', '').strip()
                                    if old_filename:
                                        excel_dir = os.path.dirname(excel_path)
                                        original_file_path = os.path.join(excel_dir, old_filename)
                                        if os.path.exists(original_file_path):
                                            moved = self.file_manager.move_file_to_duplicate_folder(original_file_path)
                                            if moved:
                                                self._log(f"✅ ย้ายไฟล์ต้นฉบับไปยัง 'เอกสารซ้ำรอตรวจ': {moved}")
                                            else:
                                                self._log("❌ ย้ายไฟล์ต้นฉบับไปยังโฟลเดอร์เอกสารซ้ำรอตรวจไม่สำเร็จ", level="warning")
                                        else:
                                            self._log(f"⚠️ ไม่พบไฟล์ต้นฉบับ: {original_file_path}", level="warning")
                                else:
                                    self._log(f"⚠️ บันทึกไม่สำเร็จ: {result} (ลำดับ {sequence})", level="warning")
                                    self._notify_progress(failure_delta=1)
                                    
                                    # อัปเดตสถานะใน Excel
                                    if first_row and row_index:
                                        self.update_excel_status(excel_path, sheet_name, row_index, f"ล้มเหลว: {result}")
                            else:
                                # ถ้ามีแค่แถวเดียว → ทำขั้นตอนสุดท้าย (หมายเหตุ, เลือกรวมภาษี, คลิกปุ่ม)
                                self._log(f"📊 มีแค่แถวเดียว - กำลังทำขั้นตอนสุดท้าย...")
                                
                                # 5.5. กรอกหมายเหตุ
                                remark = first_row.get('หมายเหตุ', '').strip()
                                if remark:
                                    # ทำความสะอาดข้อความหมายเหตุ (ลบ .pdf, None_vat_, VAT_, WHT_)
                                    remark_cleaned = self._clean_remark_text(remark)
                                    self._log(f"📝 กำลังกรอกหมายเหตุ: {remark_cleaned} (เดิม: {remark})")
                                    try:
                                        remark_field = self.page.locator('#tarremark')
                                        if remark_field.count() > 0:
                                            remark_field.first.fill(remark_cleaned)
                                            self._log(f"✅ กรอกหมายเหตุสำเร็จ")
                                            time.sleep(0.5)
                                        else:
                                            self._log(f"⚠️ ไม่พบฟิลด์หมายเหตุ (#tarremark)", level="warning")
                                    except Exception as e:
                                        self._log(f"⚠️ เกิดข้อผิดพลาดในการกรอกหมายเหตุ: {e}", level="warning")
                                
                                # 7. เลือกดรอปดาวน์ภาษี (เฉพาะ is_vat_sheet)
                                if is_vat_sheet:
                                    # สำหรับ VAT: เลือก "รวมภาษี" ใน ddltaxstatus
                                    self._log(f"🔍 [VAT] กำลังเลือกดรอปดาวน์รวมภาษี (ddltaxstatus)...")
                                    try:
                                        tax_status_dropdown = self.page.locator('//*[@id="ddltaxstatus"]')
                                        if tax_status_dropdown.count() > 0:
                                            tax_status_dropdown.first.click()
                                            time.sleep(1)
                                            
                                            tax_included_option = self.page.locator('text=รวมภาษี, text=Include Tax')
                                            if tax_included_option.count() > 0:
                                                tax_included_option.first.click()
                                                self._log(f"✅ เลือกรวมภาษีสำเร็จ")
                                            else:
                                                # Fallback: เลือกด้วย index
                                                try:
                                                    self.page.select_option('//*[@id="ddltaxstatus"]', index=1)
                                                    self._log(f"✅ เลือกรวมภาษีสำเร็จด้วยวิธี fallback")
                                                except:
                                                    self._log(f"⚠️ ไม่สามารถเลือกรวมภาษีได้", level="warning")
                                    except Exception as e:
                                        self._log(f"⚠️ เกิดข้อผิดพลาดในการเลือกรวมภาษี: {e}", level="warning")
                                
                                # 8. คลิกปุ่ม hidePaymentModal
                                self._log(f"🔍 กำลังคลิกปุ่ม hidePaymentModal...")
                                try:
                                    hide_payment_button = self.page.locator('//*[@id="hidePaymentModal"]')
                                    if hide_payment_button.count() > 0:
                                        hide_payment_button.first.click()
                                        self._log(f"✅ คลิกปุ่ม hidePaymentModal สำเร็จ")
                                    else:
                                        self._log(f"⚠️ ไม่พบปุ่ม hidePaymentModal", level="warning")
                                except Exception as e:
                                    self._log(f"⚠️ เกิดข้อผิดพลาดในการคลิกปุ่ม hidePaymentModal: {e}", level="warning")
                                
                                # 9. คลิกปุ่มสุดท้าย
                                self._log(f"🔍 กำลังกดปุ่มสุดท้าย...")
                                time.sleep(1.5)
                                document_number = None  # ประกาศตัวแปรไว้ก่อน
                                should_continue_processing = True  # ตัวแปรสำหรับควบคุมการทำงานต่อ
                                try:
                                    final_button = self.page.locator('//*[@id="content"]/div[6]/div[6]/div/div[2]/div[2]/div[1]')
                                    if final_button.count() > 0:
                                        final_button.first.click()
                                        self._log(f"✅ คลิกปุ่มสุดท้ายสำเร็จ")
                                        
                                        # 9.5. ตรวจสอบ element #dvredalert ทันทีหลังจากคลิกปุ่มสุดท้าย
                                        time.sleep(1)  # รอให้ alert แสดง
                                        try:
                                            red_alert = self.page.locator('#dvredalert')
                                            if red_alert.count() > 0 and red_alert.first.is_visible():
                                                try:
                                                    alert_label = self.page.locator('#lbredalert')
                                                    alert_text = ''
                                                    if alert_label.count() > 0:
                                                        alert_text = (alert_label.first.text_content() or '').strip()
                                                    else:
                                                        alert_text = (red_alert.first.text_content() or '').strip()
                                                    
                                                    self._log(f"📋 ข้อความใน alert: '{alert_text}'")
                                                    
                                                    # ตรวจสอบว่ามีข้อความ "เลขที่ใบกำกับภาษีอ้างอิงนี้สำหรับผู้ติดต่อนี้เคยถูกใช้งานแล้ว"
                                                    if "เลขที่ใบกำกับภาษีอ้างอิงนี้สำหรับผู้ติดต่อนี้เคยถูกใช้งานแล้ว" in alert_text:
                                                        self._log(f"⚠️ พบเอกสารซ้ำ - กำลังย้ายไฟล์ทันที...")
                                                        self._notify_progress(duplicate_delta=1)
                                                        
                                                        # ย้ายไฟล์ต้นฉบับไปโฟลเดอร์ เอกสารซ้ำรอตรวจ
                                                        old_filename = first_row.get('ชื่อไฟล์เก่า', '').strip()
                                                        if old_filename:
                                                            excel_dir = os.path.dirname(excel_path)
                                                            original_file_path = os.path.join(excel_dir, old_filename)
                                                            if os.path.exists(original_file_path):
                                                                moved = self.file_manager.move_file_to_duplicate_folder(original_file_path)
                                                                if moved:
                                                                    self._log(f"✅ ย้ายไฟล์ต้นฉบับไปยัง 'เอกสารซ้ำรอตรวจ': {moved}")
                                                                else:
                                                                    self._log("❌ ย้ายไฟล์ต้นฉบับไปยังโฟลเดอร์เอกสารซ้ำรอตรวจไม่สำเร็จ", level="warning")
                                                            else:
                                                                self._log(f"⚠️ ไม่พบไฟล์ต้นฉบับ: {original_file_path}", level="warning")
                                                        
                                                        should_continue_processing = False  # หยุดการทำงานต่อ
                                                        # ข้ามไปลำดับถัดไป
                                                        continue
                                                    
                                                    # ตรวจสอบว่ามีข้อความ "โปรดกรอกข้อมูลในช่อง"
                                                    elif "โปรดกรอกข้อมูลในช่อง" in alert_text:
                                                        self._log(f"⚠️ พบข้อความ 'โปรดกรอกข้อมูลในช่อง' - กำลังรีเซ็ตและกรอกข้อมูลใหม่...")
                                                        
                                                        # เก็บข้อมูลสำหรับ retry
                                                        self._current_excel_rows_for_retry = [first_row]  # เก็บเป็น list
                                                        self._current_excel_sequence_info_for_retry = {
                                                            'sequence': sequence,
                                                            'is_vat_sheet': is_vat_sheet,
                                                            'excel_path': excel_path
                                                        }
                                                        
                                                        # วนลูปสูงสุด 5 ครั้ง
                                                        max_retry = 5
                                                        retry_success = False
                                                        
                                                        for retry_count in range(1, max_retry + 1):
                                                            self._log(f"🔄 เริ่มรีเซ็ตและกรอกข้อมูลใหม่ (ครั้งที่ {retry_count}/{max_retry})...")
                                                            self._refill_attempt_count = retry_count - 1
                                                            
                                                            try:
                                                                # รีหน้าแล้วกรอกใหม่ทั้งชุด
                                                                self.refresh_express_page()
                                                                time.sleep(1)  # รอให้หน้าเว็บโหลดเสร็จ
                                                                
                                                                # กรอกข้อมูลแถวแรก
                                                                effective_is_vat_sheet = is_vat_sheet if getattr(self, 'current_folder_group', 'regular') != 'special' else False
                                                                if self.fill_form_from_excel_row(first_row, effective_is_vat_sheet, row_index=1, is_first_row=True, folder_group=getattr(self, 'current_folder_group', 'regular')):
                                                                    self._log(f"✅ กรอกข้อมูลแถวแรกสำเร็จ")
                                                                    
                                                                    # ทำขั้นตอนสุดท้าย (หมายเหตุ, เลือกรวมภาษี, คลิกปุ่ม)
                                                                    # กรอกหมายเหตุ
                                                                    remark = first_row.get('หมายเหตุ', '').strip()
                                                                    if remark:
                                                                        remark_cleaned = self._clean_remark_text(remark)
                                                                        try:
                                                                            remark_field = self.page.locator('#tarremark')
                                                                            if remark_field.count() > 0:
                                                                                remark_field.first.fill(remark_cleaned)
                                                                                time.sleep(0.5)
                                                                        except:
                                                                            pass
                                                                    
                                                                    # เลือกรวมภาษี (ถ้า is_vat_sheet)
                                                                    if is_vat_sheet:
                                                                        try:
                                                                            tax_status_dropdown = self.page.locator('//*[@id="ddltaxstatus"]')
                                                                            if tax_status_dropdown.count() > 0:
                                                                                tax_status_dropdown.first.click()
                                                                                time.sleep(1)
                                                                                tax_included_option = self.page.locator('text=รวมภาษี, text=Include Tax')
                                                                                if tax_included_option.count() > 0:
                                                                                    tax_included_option.first.click()
                                                                                else:
                                                                                    self.page.select_option('//*[@id="ddltaxstatus"]', index=1)
                                                                        except:
                                                                            pass
                                                                    
                                                                    # คลิกปุ่ม hidePaymentModal
                                                                    try:
                                                                        hide_payment_button = self.page.locator('//*[@id="hidePaymentModal"]')
                                                                        if hide_payment_button.count() > 0:
                                                                            hide_payment_button.first.click()
                                                                    except:
                                                                        pass
                                                                    
                                                                    # คลิกปุ่มสุดท้าย
                                                                    time.sleep(1.5)
                                                                    try:
                                                                        final_button = self.page.locator('//*[@id="content"]/div[6]/div[6]/div/div[2]/div[2]/div[1]')
                                                                        if final_button.count() > 0:
                                                                            final_button.first.click()
                                                                            self._log(f"✅ คลิกปุ่มสุดท้ายสำเร็จ (ครั้งที่ {retry_count})")
                                                                            time.sleep(2)  # รอให้หน้าเว็บโหลด
                                                                            
                                                                            # ตรวจสอบอีกครั้งว่ามี alert หรือไม่
                                                                            red_alert_check = self.page.locator('#dvredalert')
                                                                            if red_alert_check.count() > 0 and red_alert_check.first.is_visible():
                                                                                alert_label_check = self.page.locator('#lbredalert')
                                                                                alert_text_check = ''
                                                                                if alert_label_check.count() > 0:
                                                                                    alert_text_check = (alert_label_check.first.text_content() or '').strip()
                                                                                else:
                                                                                    alert_text_check = (red_alert_check.first.text_content() or '').strip()
                                                                                
                                                                                # ถ้ายังมี "โปรดกรอกข้อมูลในช่อง" ให้ลองใหม่
                                                                                if "โปรดกรอกข้อมูลในช่อง" in alert_text_check:
                                                                                    self._log(f"⚠️ ยังพบ 'โปรดกรอกข้อมูลในช่อง' (ครั้งที่ {retry_count}) - จะลองใหม่...")
                                                                                    continue  # ลองใหม่
                                                                                # ถ้ามี "เลขที่ใบกำกับภาษีอ้างอิงนี้สำหรับผู้ติดต่อนี้เคยถูกใช้งานแล้ว" ให้ย้ายไฟล์
                                                                                elif "เลขที่ใบกำกับภาษีอ้างอิงนี้สำหรับผู้ติดต่อนี้เคยถูกใช้งานแล้ว" in alert_text_check:
                                                                                    self._log(f"⚠️ พบเอกสารซ้ำ (ครั้งที่ {retry_count}) - กำลังย้ายไฟล์...")
                                                                                    old_filename = first_row.get('ชื่อไฟล์เก่า', '').strip()
                                                                                    if old_filename:
                                                                                        excel_dir = os.path.dirname(excel_path)
                                                                                        original_file_path = os.path.join(excel_dir, old_filename)
                                                                                        if os.path.exists(original_file_path):
                                                                                            moved = self.file_manager.move_file_to_duplicate_folder(original_file_path)
                                                                                            if moved:
                                                                                                self._log(f"✅ ย้ายไฟล์ต้นฉบับไปยัง 'เอกสารซ้ำรอตรวจ': {moved}")
                                                                                    should_continue_processing = False
                                                                                    break  # ออกจาก retry loop
                                                                            else:
                                                                                # ไม่มี alert แสดงว่าอาจจะสำเร็จ
                                                                                self._log(f"✅ ไม่พบ alert - อาจจะบันทึกสำเร็จ (ครั้งที่ {retry_count})")
                                                                                retry_success = True
                                                                                break  # ออกจาก retry loop
                                                                    except Exception as e:
                                                                        self._log(f"⚠️ เกิดข้อผิดพลาดในการคลิกปุ่มสุดท้าย (ครั้งที่ {retry_count}): {e}")
                                                                        
                                                            except Exception as e:
                                                                self._log(f"⚠️ เกิดข้อผิดพลาดในการรีเซ็ตและกรอกข้อมูลใหม่ (ครั้งที่ {retry_count}): {e}")
                                                        
                                                        # ล้างข้อมูล retry
                                                        self._current_excel_rows_for_retry = None
                                                        self._current_excel_sequence_info_for_retry = None
                                                        
                                                        if not retry_success:
                                                            self._log(f"⚠️ ไม่สามารถกรอกข้อมูลใหม่ได้สำเร็จหลังจากลอง {max_retry} ครั้ง", level="warning")
                                                            should_continue_processing = False
                                                            continue  # ข้ามไปลำดับถัดไป
                                                        
                                                except Exception as e:
                                                    self._log(f"⚠️ เกิดข้อผิดพลาดในการตรวจสอบ alert: {e}", level="warning")
                                        except Exception as e:
                                            self._log(f"⚠️ เกิดข้อผิดพลาดในการตรวจสอบ #dvredalert: {e}", level="warning")
                                        
                                        # ถ้าไม่ต้องหยุดการทำงานต่อ ให้ตรวจสอบ header element (เพื่ออ่าน document_number สำหรับการเปลี่ยนชื่อไฟล์)
                                        if should_continue_processing:
                                            # 9.6. รอจนกว่าหน้าเว็บจะโหลดเสร็จและ element h3 แสดงข้อมูล
                                            # สำหรับ special → ยังคงต้องอ่าน document_number จาก header เพื่อใช้ในการเปลี่ยนชื่อไฟล์
                                            self._log(f"⏳ กำลังรอให้หน้าเว็บโหลดเสร็จ...")
                                            try:
                                                # รอจนกว่า element h3 จะแสดงข้อมูล
                                                header_element = self.page.locator('h3.ui.header.heading.section-header-doc-left')
                                                header_element.wait_for(state='visible', timeout=0)  # รอไม่จำกัดเวลา
                                                self._log(f"✅ พบ header element แล้ว")
                                                
                                                # อ่านเลขที่เอกสารจาก element (ดึงข้อมูลหลังจาก #)
                                                # สำหรับ special → ใช้สำหรับการเปลี่ยนชื่อไฟล์เท่านั้น (ไม่กรอกในฟอร์ม)
                                                try:
                                                    header_text = header_element.first.text_content() or ''
                                                    self._log(f"📄 ข้อความใน header: {header_text}")
                                                    
                                                    # หาเลขที่เอกสารหลังจาก #
                                                    match = re.search(r'#([A-Z0-9\-]+)', header_text)
                                                    if match:
                                                        document_number = match.group(1)
                                                        self._log(f"✅ อ่านเลขที่เอกสารสำเร็จ: {document_number}")
                                                        if getattr(self, 'current_folder_group', 'regular') == 'special':
                                                            self._log(f"📋 folder_group='special' → ใช้ document_number สำหรับการเปลี่ยนชื่อไฟล์เท่านั้น")
                                                    else:
                                                        self._log(f"⚠️ ไม่พบเลขที่เอกสารใน header", level="warning")
                                                except Exception as e:
                                                    self._log(f"⚠️ เกิดข้อผิดพลาดในการอ่านเลขที่เอกสาร: {e}", level="warning")
                                            except Exception as e:
                                                self._log(f"⚠️ ไม่พบ header element: {e}", level="warning")
                                    else:
                                        self._log(f"⚠️ ไม่พบปุ่มสุดท้าย", level="warning")
                                except Exception as e:
                                    self._log(f"⚠️ เกิดข้อผิดพลาดในการคลิกปุ่มสุดท้าย: {e}", level="warning")
                                
                                # ถ้าไม่ต้องทำงานต่อ ให้ข้ามไปลำดับถัดไป
                                if not should_continue_processing:
                                    continue
                                
                                # 10. ตรวจสอบ VAT (เฉพาะชีต "มีภาษีมูลค่าเพิ่ม" และไม่ใช่ special)
                                if is_vat_sheet and getattr(self, 'current_folder_group', 'regular') != 'special':
                                    vat_amount = first_row.get('ยอดภาษีมูลค่าเพิ่ม', '').strip()
                                    if vat_amount:
                                        try:
                                            vat_amount_clean = re.sub(r'[^\d.]', '', str(vat_amount))
                                            vat_amount_float = float(vat_amount_clean)
                                            if vat_amount_float > 0:  # ตรวจสอบว่ามีค่าจริง
                                                self._log(f"🔍 กำลังตรวจสอบยอดภาษีมูลค่าเพิ่ม: {vat_amount_float}")
                                                
                                                # ใช้ฟังก์ชัน check_and_fix_vat_value ที่มีอยู่แล้ว
                                                if not self.check_and_fix_vat_value(str(vat_amount_float)):
                                                    self._log(f"⚠️ ตรวจสอบ/แก้ไข VAT ไม่สำเร็จ", level="warning")
                                        except Exception as e:
                                            self._log(f"⚠️ เกิดข้อผิดพลาดในการตรวจสอบ VAT: {e}", level="warning")
                                elif getattr(self, 'current_folder_group', 'regular') == 'special':
                                    self._log(f"📋 folder_group='special' → ข้ามการตรวจสอบยอดภาษีมูลค่าเพิ่ม")
                                
                                # หลังจากทำขั้นตอนสุดท้ายเสร็จแล้ว → รอผลบันทึก
                                # เก็บข้อมูล Excel row ไว้สำหรับ retry (เหมือนระบบอ่าน PDF)
                                self._current_excel_row_data_for_retry = first_row
                                self._current_excel_sheet_info_for_retry = {
                                    'is_vat_sheet': is_vat_sheet,
                                    'row_index': 1,
                                    'is_first_row': True,
                                    'excel_path': excel_path
                                }
                                self._refill_attempt_count = 0  # รีเซ็ต counter สำหรับ retry
                                
                                self._status_update(step="กำลังรอผลบันทึก", file=company_name or f"ลำดับ {sequence}")
                                result = self.wait_for_save_result(timeout=20)
                                self._log(f"📋 ผลการบันทึก: {result}")
                                
                                # ล้างข้อมูล retry หลังจากรอผลเสร็จ
                                self._current_excel_row_data_for_retry = None
                                self._current_excel_sheet_info_for_retry = None
                                
                                if result == 'approved':
                                    self._log(f"✅ บันทึกแถวแรกสำเร็จ")
                                    self._notify_progress(success_delta=1)
                                    processed_count += 1
                                    
                                    # 11. เปลี่ยนชื่อไฟล์และอัปโหลด (ถ้ามีเลขที่เอกสาร)
                                    # สำหรับ special → ใช้ document_number จาก header (ที่อ่านมาแล้ว) สำหรับการเปลี่ยนชื่อไฟล์
                                    new_filename = None
                                    upload_success = False
                                    if document_number:
                                        # ใช้ effective_is_vat_sheet (ซึ่งจะเป็น False สำหรับ special)
                                        effective_is_vat_sheet = is_vat_sheet if getattr(self, 'current_folder_group', 'regular') != 'special' else False
                                        new_filename = self._handle_file_rename_and_upload(
                                            first_row, document_number, effective_is_vat_sheet, excel_path
                                        )
                                        upload_success = (new_filename is not None)
                                        
                                        # เก็บ mapping สำหรับเขียนลง Excel ทีหลัง
                                        if new_filename and first_row and row_index:
                                            key = (sheet_name, row_index)
                                            self.renamed_files_mapping[key] = new_filename
                                            self._log(f"💾 เก็บ mapping: ชีต '{sheet_name}' แถวที่ {row_index} → {new_filename}")
                                        else:
                                            if not new_filename:
                                                self._log(f"⚠️ ไม่มี new_filename → ไม่เก็บ mapping", level="warning")
                                            if not first_row:
                                                self._log(f"⚠️ ไม่มี first_row → ไม่เก็บ mapping", level="warning")
                                            if not row_index:
                                                self._log(f"⚠️ ไม่มี row_index → ไม่เก็บ mapping", level="warning")
                                    else:
                                        self._log(f"⚠️ ไม่พบเลขที่เอกสาร → ข้ามการเปลี่ยนชื่อไฟล์", level="warning")
                                    
                                    # อัปเดตสถานะใน Excel (หลังจากอัปโหลดไฟล์สำเร็จ)
                                    if upload_success or not document_number:  # ถ้าอัปโหลดสำเร็จ หรือไม่มี document_number (ข้ามการอัปโหลด)
                                        if first_row and row_index:
                                            self.update_excel_status(excel_path, sheet_name, row_index, "สำเร็จ")
                                        self._log(f"💾 อัปเดตสถานะใน Excel: ลำดับ {sequence} (ชีต: {sheet_name}) → สำเร็จ")
                                elif result == 'duplicate':
                                    self._log(f"⚠️ เอกสารซ้ำ (ลำดับ {sequence})", level="warning")
                                    self._notify_progress(duplicate_delta=1)
                                    
                                    # อัปเดตสถานะใน Excel
                                    if first_row and row_index:
                                        self.update_excel_status(excel_path, sheet_name, row_index, "เอกสารซ้ำ")
                                    
                                    # ย้ายไฟล์ต้นฉบับไปโฟลเดอร์ เอกสารซ้ำรอตรวจ (เหมือนระบบอ่าน PDF)
                                    old_filename = first_row.get('ชื่อไฟล์เก่า', '').strip()
                                    if old_filename:
                                        excel_dir = os.path.dirname(excel_path)
                                        original_file_path = os.path.join(excel_dir, old_filename)
                                        if os.path.exists(original_file_path):
                                            moved = self.file_manager.move_file_to_duplicate_folder(original_file_path)
                                            if moved:
                                                self._log(f"✅ ย้ายไฟล์ต้นฉบับไปยัง 'เอกสารซ้ำรอตรวจ': {moved}")
                                            else:
                                                self._log("❌ ย้ายไฟล์ต้นฉบับไปยังโฟลเดอร์เอกสารซ้ำรอตรวจไม่สำเร็จ", level="warning")
                                        else:
                                            self._log(f"⚠️ ไม่พบไฟล์ต้นฉบับ: {original_file_path}", level="warning")
                                else:
                                    self._log(f"⚠️ บันทึกแถวแรกไม่สำเร็จ: {result}", level="warning")
                                    self._notify_progress(failure_delta=1)
                                    
                                    # อัปเดตสถานะใน Excel
                                    if first_row and row_index:
                                        self.update_excel_status(excel_path, sheet_name, row_index, f"ล้มเหลว: {result}")
                        else:
                            self._log(f"❌ กรอกข้อมูลแถวแรกไม่สำเร็จ", level="error")
                            self._notify_progress(failure_delta=1)
                            
                            # อัปเดตสถานะใน Excel
                            if first_row and row_index:
                                self.update_excel_status(excel_path, sheet_name, row_index, "ล้มเหลว: กรอกข้อมูลไม่สำเร็จ")
                    
                    # รอสักครู่ก่อนลำดับถัดไป
                    time.sleep(0.5)
            
            # แสดงสรุปสถานะจาก Excel
            self._log(f"\n📊 สรุปสถานะการทำงาน:")
            self._log(f"   ✅ ประมวลผลสำเร็จ: {processed_count} รายการ")
            self._log(f"   📋 ตรวจสอบสถานะทั้งหมดได้จากคอลัมน์ 'สถานะ' ใน Excel")
            
            # เขียนชื่อไฟล์ที่เปลี่ยนแล้วลง Excel (ถ้ามี)
            if hasattr(self, 'renamed_files_mapping') and self.renamed_files_mapping:
                self._write_renamed_filenames_to_excel(excel_path)
            
            self._log(f"\n✅ ประมวลผลข้อมูลจาก Excel เสร็จสิ้น (ประมวลผล {processed_count} รายการ)")
            self._status_update(step="ประมวลผลเสร็จสิ้น", file='-')
            
            return True
            
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการประมวลผลข้อมูลจาก Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def fill_form_from_excel_row(self, row_data: Dict, is_vat_sheet: bool, row_index: int = 1, is_first_row: bool = True, folder_group: Optional[str] = None) -> bool:
        """กรอกข้อมูลในฟอร์มจากข้อมูลแถวใน Excel
        
        Args:
            row_data: Dictionary ที่มีข้อมูลจากแถวใน Excel
            is_vat_sheet: True ถ้าเป็นชีต "มีภาษีมูลค่าเพิ่ม", False ถ้าเป็น "ไม่มีภาษีมูลค่าเพิ่ม"
            row_index: หมายเลขแถว (1, 2, 3, ...)
            is_first_row: True ถ้าเป็นแถวแรก, False ถ้าเป็นแถวเพิ่มเติม
            folder_group: 'regular' หรือ 'special' (ถ้า None จะใช้ self.current_folder_group)
            
        Returns:
            True ถ้ากรอกสำเร็จ, False ถ้าไม่สำเร็จ
        """
        # ใช้ folder_group จาก parameter หรือจาก self.current_folder_group
        if folder_group is None:
            folder_group = getattr(self, 'current_folder_group', 'regular')
        
        # ถ้า folder_group = 'special' ให้ทำงานเหมือน "ไม่มีภาษีมูลค่าเพิ่ม" ทั้งสองชีต
        if folder_group == 'special':
            is_vat_sheet = False  # บังคับให้ทำงานเหมือน "ไม่มีภาษีมูลค่าเพิ่ม"
            self._log(f"📋 folder_group='special' → ทำงานเหมือน 'ไม่มีภาษีมูลค่าเพิ่ม' (ไม่กรอกเลขที่เอกสาร, ไม่ตรวจสอบยอดภาษี)")
        try:
            self._log(f"📝 กำลังกรอกข้อมูลจาก Excel row (แถวที่ {row_index})...")
            
            # 1. กรอกชื่อบริษัท (เฉพาะแถวแรกเท่านั้น)
            if is_first_row:
                company_name = row_data.get('ชื่อบริษัท', '').strip()
                tax_id = row_data.get('เลขประจำตัวผู้เสียภาษี', '').strip()
                
                if company_name:
                    self._log(f"🏢 กำลังกรอกชื่อบริษัท: {company_name}")
                try:
                    customer_field = self.page.locator('//*[@id="iptcontactname"]')
                    if customer_field.count() > 0:
                        customer_field.first.fill(company_name)
                        self._log(f"✅ กรอกชื่อบริษัทสำเร็จ")
                        time.sleep(1)  # รอให้ dropdown แสดง
                        
                        # ตรวจสอบว่าต้องเพิ่มผู้ติดต่อใหม่หรือไม่
                        if not self.check_and_add_new_contact(company_name, tax_id, customer_field):
                            self._log(f"⚠️ ไม่สามารถจัดการผู้ติดต่อได้", level="warning")
                    else:
                        self._log(f"⚠️ ไม่พบฟิลด์ชื่อบริษัท (iptcontactname)", level="warning")
                except Exception as e:
                    self._log(f"⚠️ เกิดข้อผิดพลาดในการกรอกชื่อบริษัท: {e}", level="warning")
            
            # 2. กรอก Account Code
            account_code = row_data.get('ชื่อบัญชี / โค้ดบัญชี', '').strip()
            if account_code:
                self._log(f"🔢 กำลังกรอก Account Code: {account_code}")
                try:
                    account_field = self.page.locator(f'//*[@id="iptaccountcode{row_index}"]')
                    if account_field.count() > 0:
                        account_field.first.fill(account_code)
                        
                        # รอให้ dropdown แสดง (อย่างน้อย 1 วินาที หรือจนกว่า dropdown จะแสดง)
                        self._log(f"⏳ กำลังรอ dropdown แสดง...")
                        dropdown_appeared = False
                        max_wait_time = 5  # รอสูงสุด 5 วินาที
                        wait_interval = 0.2  # ตรวจสอบทุก 0.2 วินาที
                        elapsed_time = 0
                        
                        while elapsed_time < max_wait_time:
                            try:
                                dropdown_items = self.page.locator('css=ul.ui-autocomplete:visible li')
                                dropdown_count = dropdown_items.count()
                                
                                if dropdown_count > 0:
                                    dropdown_appeared = True
                                    self._log(f"✅ พบ dropdown หลังจากรอ {elapsed_time:.1f} วินาที")
                                    break
                            except:
                                pass
                            
                            time.sleep(wait_interval)
                            elapsed_time += wait_interval
                        
                        # ตรวจสอบว่ามี dropdown แสดงหรือไม่
                        try:
                            dropdown_items = self.page.locator('css=ul.ui-autocomplete:visible li')
                            dropdown_count = dropdown_items.count()
                            
                            if dropdown_count > 0:
                                self._log(f"📋 พบ dropdown {dropdown_count} รายการ")
                                
                                # ตรวจสอบรายการทั้งหมด
                                matched_item = None
                                for i in range(dropdown_count):
                                    try:
                                        item = dropdown_items.nth(i)
                                        item_text = (item.text_content() or '').strip()
                                        self._log(f"📋 รายการที่ {i+1}: '{item_text}'")
                                        
                                        # ตรวจสอบว่ารายการนี้ตรงกับ Account Code ที่กรอกหรือไม่
                                        # รายการอาจเป็น "520103 - ค่าบริการ Grab" และ account_code อาจเป็น "ค่าบริการ Grab" หรือ "520103 - ค่าบริการ Grab"
                                        account_code_lower = account_code.lower().strip()
                                        item_text_lower = item_text.lower().strip()
                                        
                                        # ตรวจสอบหลายแบบ:
                                        # 1. account_code อยู่ใน item_text (เช่น "ค่าบริการ Grab" ใน "520103 - ค่าบริการ Grab")
                                        # 2. item_text อยู่ใน account_code (เช่น "520103 - ค่าบริการ Grab" ใน "520103 - ค่าบริการ Grab")
                                        # 3. แยกส่วน code และชื่อออกมาเปรียบเทียบ (เช่น "520103" หรือ "ค่าบริการ Grab")
                                        is_match = False
                                        
                                        if account_code_lower in item_text_lower or item_text_lower in account_code_lower:
                                            is_match = True
                                        else:
                                            # ลองแยกส่วน code และชื่อออกมา
                                            # รูปแบบ: "520103 - ค่าบริการ Grab"
                                            if ' - ' in item_text:
                                                parts = item_text.split(' - ', 1)
                                                if len(parts) == 2:
                                                    code_part = parts[0].strip().lower()
                                                    name_part = parts[1].strip().lower()
                                                    
                                                    # ตรวจสอบว่า account_code ตรงกับ code หรือ name
                                                    if account_code_lower == code_part or account_code_lower == name_part:
                                                        is_match = True
                                                    elif account_code_lower in name_part or name_part in account_code_lower:
                                                        is_match = True
                                        
                                        if is_match:
                                            # พบรายการที่ตรงกัน - เก็บไว้และคลิก
                                            matched_item = item
                                            self._log(f"✅ พบรายการที่ตรงกัน: '{item_text}' - กำลังคลิก...")
                                            break
                                    except Exception as e:
                                        self._log(f"⚠️ ไม่สามารถอ่านรายการที่ {i+1} ได้: {e}", level="warning")
                                        continue
                                
                                # ถ้าพบรายการที่ตรงกัน ให้คลิก
                                if matched_item:
                                    try:
                                        matched_item.click()
                                        self._log(f"✅ คลิกรายการสำเร็จ")
                                        time.sleep(0.5)
                                    except Exception as e:
                                        self._log(f"⚠️ ไม่สามารถคลิกรายการได้: {e} - จะกด Enter", level="warning")
                                        account_field.first.press('Enter')
                                else:
                                    # ถ้าไม่พบรายการที่ตรงกัน ให้กด Enter
                                    self._log(f"⚠️ ไม่พบรายการที่ตรงกัน - จะกด Enter")
                                    account_field.first.press('Enter')
                            else:
                                # ไม่มี dropdown - ไม่ต้องกด Enter (ตามที่ผู้ใช้ขอ)
                                if not dropdown_appeared:
                                    self._log(f"⚠️ ไม่พบ dropdown หลังจากรอ {max_wait_time} วินาที - ข้ามการเลือก (ไม่กด Enter)")
                                else:
                                    self._log(f"⚠️ ไม่พบ dropdown - ข้ามการเลือก (ไม่กด Enter)")
                        except Exception as e:
                            # ถ้าเกิดข้อผิดพลาดในการตรวจสอบ dropdown - ไม่ต้องกด Enter
                            self._log(f"⚠️ เกิดข้อผิดพลาดในการตรวจสอบ dropdown: {e} - ข้ามการเลือก (ไม่กด Enter)", level="warning")
                        
                        self._log(f"✅ กรอก Account Code สำเร็จ")
                        time.sleep(0.5)
                    else:
                        self._log(f"⚠️ ไม่พบฟิลด์ Account Code (iptaccountcode1)", level="warning")
                except Exception as e:
                    self._log(f"⚠️ เกิดข้อผิดพลาดในการกรอก Account Code: {e}", level="warning")
            
            # 3. กรอกวันที่เอกสาร (เฉพาะแถวแรกเท่านั้น)
            if is_first_row:
                document_date = row_data.get('วันที่', '').strip()
                if document_date:
                    # แปลงวันที่จากรูปแบบ yyyy-mm-dd หรือ datetime object เป็น dd/mm/yyyy
                    document_date = self._convert_date_format(document_date)
                    self._log(f"📅 กำลังกรอกวันที่เอกสาร: {document_date}")
                    try:
                        date_field = self.page.locator('//*[@id="iptdate"]')
                        if date_field.count() > 0:
                            date_field.first.fill(document_date)
                            self._log(f"✅ กรอกวันที่เอกสารสำเร็จ")
                            time.sleep(0.5)
                        else:
                            self._log(f"⚠️ ไม่พบฟิลด์วันที่เอกสาร (iptdate)", level="warning")
                    except Exception as e:
                        self._log(f"⚠️ เกิดข้อผิดพลาดในการกรอกวันที่เอกสาร: {e}", level="warning")
                
                # 3.5. กรอกวันที่ครบกำหนดชำระ (เฉพาะแถวแรกเท่านั้น)
                due_date = row_data.get('ครบกำหนดชำระ', '').strip()
                if due_date:
                    # แปลงวันที่จากรูปแบบ yyyy-mm-dd หรือ datetime object เป็น dd/mm/yyyy
                    due_date = self._convert_date_format(due_date)
                    self._log(f"📅 กำลังกรอกวันที่ครบกำหนดชำระ: {due_date}")
                    try:
                        due_date_field = self.page.locator('//*[@id="iptduedate"]')
                        if due_date_field.count() > 0:
                            due_date_field.first.fill(due_date)
                            self._log(f"✅ กรอกวันที่ครบกำหนดชำระสำเร็จ")
                            time.sleep(0.5)
                        else:
                            self._log(f"⚠️ ไม่พบฟิลด์วันที่ครบกำหนดชำระ (iptduedate)", level="warning")
                    except Exception as e:
                        self._log(f"⚠️ เกิดข้อผิดพลาดในการกรอกวันที่ครบกำหนดชำระ: {e}", level="warning")
                else:
                    self._log(f"📋 ไม่มีข้อมูลในคอลัมน์ 'ครบกำหนดชำระ' - ข้ามขั้นตอนนี้")
                
                # 3.6. กรอกอ้างอิง (เฉพาะแถวแรกเท่านั้น)
                reference = row_data.get('อ้างอิง', '').strip()
                if reference:
                    self._log(f"📝 กำลังกรอกอ้างอิง: {reference}")
                    try:
                        reference_field = self.page.locator('//*[@id="iptrefname"]')
                        if reference_field.count() > 0:
                            reference_field.first.fill(reference)
                            self._log(f"✅ กรอกอ้างอิงสำเร็จ")
                            time.sleep(0.5)
                        else:
                            self._log(f"⚠️ ไม่พบฟิลด์อ้างอิง (iptrefname)", level="warning")
                    except Exception as e:
                        self._log(f"⚠️ เกิดข้อผิดพลาดในการกรอกอ้างอิง: {e}", level="warning")
                else:
                    self._log(f"📋 ไม่มีข้อมูลในคอลัมน์ 'อ้างอิง' - ข้ามขั้นตอนนี้")
                
                # 4. กรอกเลขที่เอกสาร (เฉพาะชีต "มีภาษีมูลค่าเพิ่ม" และแถวแรกเท่านั้น และไม่ใช่ special)
                if is_vat_sheet and folder_group != 'special':
                    document_number = row_data.get('เลขที่เอกสาร', '').strip()
                    if document_number:
                        self._log(f"📄 กำลังกรอกเลขที่เอกสาร: {document_number}")
                        try:
                            invoice_button = self.page.locator('//*[@id="receivedTaxInvoiceAddButton"]')
                            if invoice_button.count() > 0:
                                invoice_button.first.click()
                                self._log(f"✅ คลิกปุ่ม receivedTaxInvoiceAddButton สำเร็จ")
                                time.sleep(1)
                                
                                # กรอกเลขที่เอกสารในป๊อปอัป
                                invoice_input = self.page.locator('//*[@id="inputModalReceivedTaxInvoiceNumber"]')
                                if invoice_input.count() > 0:
                                    invoice_input.first.fill(document_number)
                                    self._log(f"✅ กรอกเลขที่เอกสาร: {document_number}")
                                    
                                    # คลิกปุ่มในป๊อปอัป
                                    popup_button = self.page.locator('//*[@id="receivedTaxInvoiceAddModalSize"]/div/div[3]/div[2]')
                                    if popup_button.count() > 0:
                                        popup_button.first.click()
                                        self._log(f"✅ คลิกปุ่มในป๊อปอัปสำเร็จ")
                                    else:
                                        self._log(f"⚠️ ไม่พบปุ่มในป๊อปอัป", level="warning")
                                else:
                                    self._log(f"⚠️ ไม่พบฟิลด์เลขที่เอกสารในป๊อปอัป", level="warning")
                            else:
                                self._log(f"⚠️ ไม่พบปุ่ม receivedTaxInvoiceAddButton", level="warning")
                        except Exception as e:
                            self._log(f"⚠️ เกิดข้อผิดพลาดในการกรอกเลขที่เอกสาร: {e}", level="warning")
            
            # 5. กรอกยอดหลังบวกภาษีมูลค่าเพิ่ม
            total_amount = row_data.get('ยอดหลังบวกภาษีมูลค่าเพิ่ม', '').strip()
            if total_amount:
                # แปลงเป็นตัวเลข (ลบ comma และช่องว่าง)
                try:
                    total_amount_clean = re.sub(r'[^\d.]', '', str(total_amount))
                    total_amount_float = float(total_amount_clean)
                    self._log(f"💰 กำลังกรอกยอดหลังบวกภาษีมูลค่าเพิ่ม: {total_amount_float}")
                    
                    price_field = self.page.locator(f'//*[@id="iptprice{row_index}"]')
                    if price_field.count() > 0:
                        price_field.first.fill(str(total_amount_float))
                        self._log(f"✅ กรอกยอดหลังบวกภาษีมูลค่าเพิ่มสำเร็จ")
                        time.sleep(0.5)
                    else:
                        self._log(f"⚠️ ไม่พบฟิลด์ยอดหลังบวกภาษีมูลค่าเพิ่ม (iptprice{row_index})", level="warning")
                except Exception as e:
                    self._log(f"⚠️ เกิดข้อผิดพลาดในการแปลงยอดเงิน: {e}", level="warning")
            
            # 6. ตรวจสอบและเลือก ddlvattypeid ตามเงื่อนไข
            # ถ้า folder_group = 'special' → ไม่ตรวจสอบยอดภาษีมูลค่าเพิ่ม (เลือก "ไม่มี" เสมอ)
            if folder_group == 'special':
                # special → ไม่ตรวจสอบยอดภาษีมูลค่าเพิ่ม → เลือก "ไม่มี" เสมอ
                vat_type_value = '1'
                self._log(f"📋 folder_group='special' → ไม่ตรวจสอบยอดภาษีมูลค่าเพิ่ม → เลือก 'ไม่มี' เสมอ")
            else:
                # regular → ตรวจสอบยอดภาษีมูลค่าเพิ่มตามปกติ
                # ตรวจสอบว่ามีข้อมูลใน 3 คอลัมน์หรือไม่
                amount_before_vat = row_data.get('ยอดก่อนภาษีมูลค่าเพิ่ม', '').strip()
                vat_amount = row_data.get('ยอดภาษีมูลค่าเพิ่ม', '').strip()
                total_amount_check = row_data.get('ยอดหลังบวกภาษีมูลค่าเพิ่ม', '').strip()
                
                # ตรวจสอบว่า vat_amount มีค่าจริงหรือไม่ (ไม่ใช่ 0, 0.00, หรือ empty)
                vat_amount_has_value = False
                vat_amount_float = 0.0
                if vat_amount:
                    try:
                        # ลองแปลงเป็นตัวเลข
                        vat_amount_clean = re.sub(r'[^\d.]', '', str(vat_amount))
                        if vat_amount_clean:
                            vat_amount_float = float(vat_amount_clean)
                            if vat_amount_float > 0:
                                vat_amount_has_value = True
                                self._log(f"📊 ยอดภาษีมูลค่าเพิ่ม: {vat_amount_float} (มีค่า)")
                            else:
                                self._log(f"📊 ยอดภาษีมูลค่าเพิ่ม: {vat_amount_float} (ไม่มีค่า)")
                    except:
                        # ถ้าแปลงไม่ได้ ให้ถือว่ามีค่า (เป็น string ที่ไม่ใช่ตัวเลข)
                        vat_amount_has_value = True
                        self._log(f"📊 ยอดภาษีมูลค่าเพิ่ม: {vat_amount} (ไม่สามารถแปลงเป็นตัวเลขได้ - ถือว่ามีค่า)")
                else:
                    self._log(f"📊 ยอดภาษีมูลค่าเพิ่ม: ไม่มีข้อมูล")
                
                # ตรวจสอบว่ามีข้อมูลทั้ง 3 คอลัมน์หรือไม่ (และ vat_amount ต้องมีค่าจริง)
                has_all_three = bool(amount_before_vat and vat_amount_has_value and total_amount_check)
                has_only_two = bool(amount_before_vat and total_amount_check and not vat_amount_has_value)
                
                vat_type_value = None
                if has_all_three:
                    # มีทั้ง 3 คอลัมน์ และยอดภาษีมูลค่าเพิ่ม > 0 → เลือก "7%" (value="3")
                    vat_type_value = '3'
                    self._log(f"📊 พบข้อมูลทั้ง 3 คอลัมน์ และยอดภาษีมูลค่าเพิ่ม > 0 → เลือก '7%'")
                elif has_only_two:
                    # มีแค่ 2 คอลัมน์ (ไม่มียอดภาษีมูลค่าเพิ่ม หรือยอดภาษีมูลค่าเพิ่ม = 0) → เลือก "ไม่มี" (value="1")
                    vat_type_value = '1'
                    self._log(f"📊 พบข้อมูล 2 คอลัมน์ (ยอดภาษีมูลค่าเพิ่ม = 0 หรือไม่มี) → เลือก 'ไม่มี'")
                elif vat_amount_has_value:
                    # มีแค่อย่างเดียวคือยอดภาษีมูลค่าเพิ่ม > 0 → เลือก "7%"
                    vat_type_value = '3'
                    self._log(f"📊 พบยอดภาษีมูลค่าเพิ่ม > 0 → เลือก '7%'")
                else:
                    # ไม่มียอดภาษีมูลค่าเพิ่ม หรือยอดภาษีมูลค่าเพิ่ม = 0 → เลือก "ไม่มี"
                    vat_type_value = '1'
                    self._log(f"📊 ไม่มียอดภาษีมูลค่าเพิ่ม หรือยอดภาษีมูลค่าเพิ่ม = 0 → เลือก 'ไม่มี'")
            
            # เลือก ddlvattypeid ตามที่กำหนด
            if vat_type_value:
                try:
                    vat_type_dropdown = self.page.locator(f'//*[@id="ddlvattypeid{row_index}"]')
                    if vat_type_dropdown.count() > 0:
                        try:
                            self.page.select_option(f'//*[@id="ddlvattypeid{row_index}"]', value=vat_type_value)
                            vat_type_text = '7%' if vat_type_value == '3' else 'ไม่มี'
                            self._log(f"✅ เลือก ddlvattypeid{row_index} เป็น '{vat_type_text}' สำเร็จ")
                            time.sleep(0.3)
                        except:
                            # Fallback: ใช้ JavaScript
                            try:
                                self.page.evaluate(f"""
                                    const dropdown = document.querySelector('#ddlvattypeid{row_index}');
                                    if (dropdown) {{
                                        dropdown.value = '{vat_type_value}';
                                        dropdown.dispatchEvent(new Event('change', {{bubbles: true}}));
                                    }}
                                """)
                                vat_type_text = '7%' if vat_type_value == '3' else 'ไม่มี'
                                self._log(f"✅ เลือก ddlvattypeid{row_index} เป็น '{vat_type_text}' สำเร็จด้วย JavaScript")
                                time.sleep(0.3)
                            except Exception as e:
                                self._log(f"⚠️ ไม่สามารถเลือก ddlvattypeid{row_index} ได้: {e}", level="warning")
                    else:
                        self._log(f"⚠️ ไม่พบดรอปดาวน์ ddlvattypeid{row_index}", level="warning")
                except Exception as e:
                    self._log(f"⚠️ เกิดข้อผิดพลาดในการเลือก ddlvattypeid{row_index}: {e}", level="warning")
                
                # 6.5. ตรวจสอบและเลือก "เปอร์เซ็นต์หัก ณ ที่จ่าย" (ทุกแถว)
                withholding_percent = row_data.get('เปอร์เซ็นต์หัก ณ ที่จ่าย', '').strip()
                if withholding_percent:
                    try:
                        # แปลงเป็นตัวเลข
                        withholding_clean = re.sub(r'[^\d.]', '', str(withholding_percent))
                        if withholding_clean:
                            withholding_float = float(withholding_clean)
                            self._log(f"📊 พบเปอร์เซ็นต์หัก ณ ที่จ่าย (แถวที่ {row_index}): {withholding_float}%")
                            
                            # คลิก dropdown whtDropDown{row_index}
                            wht_dropdown = self.page.locator(f'//*[@id="whtDropDown{row_index}"]')
                            if wht_dropdown.count() > 0:
                                wht_dropdown.first.click()
                                self._log(f"✅ คลิก whtDropDown{row_index} สำเร็จ")
                                time.sleep(0.5)
                                
                                # หา menu item ที่ตรงกับเปอร์เซ็นต์
                                # รองรับ: 0.75, 1, 1.5, 2, 3, 5, 10, 15
                                wht_value = None
                                wht_text = None
                                
                                if withholding_float == 0.75:
                                    wht_value = '0.75'
                                    wht_text = '0.75%'
                                elif withholding_float == 1:
                                    wht_value = '1'
                                    wht_text = '1%'
                                elif withholding_float == 1.5:
                                    wht_value = '1.5'
                                    wht_text = '1.5%'
                                elif withholding_float == 2:
                                    wht_value = '2'
                                    wht_text = '2%'
                                elif withholding_float == 3:
                                    wht_value = '3'
                                    wht_text = '3%'
                                elif withholding_float == 5:
                                    wht_value = '5'
                                    wht_text = '5%'
                                elif withholding_float == 10:
                                    wht_value = '10'
                                    wht_text = '10%'
                                elif withholding_float == 15:
                                    wht_value = '15'
                                    wht_text = '15%'
                                elif withholding_float == 0:
                                    wht_value = '0'
                                    wht_text = 'กำหนดเอง'
                                
                                if wht_value:
                                    # คลิก menu item ที่ตรงกับเปอร์เซ็นต์
                                    try:
                                        # ลองหาด้วย data-value
                                        menu_item = self.page.locator(f'//div[@class="menu transition visible"]//div[@data-value="{wht_value}"]')
                                        if menu_item.count() > 0:
                                            menu_item.first.click()
                                            self._log(f"✅ เลือกเปอร์เซ็นต์หัก ณ ที่จ่าย (แถวที่ {row_index}): {wht_text} สำเร็จ")
                                            time.sleep(0.3)
                                        else:
                                            # Fallback: ใช้ JavaScript
                                            self.page.evaluate(f"""
                                                const menu = document.querySelector('.menu.transition.visible');
                                                if (menu) {{
                                                    const item = menu.querySelector('[data-value="{wht_value}"]');
                                                    if (item) {{
                                                        item.click();
                                                    }}
                                                }}
                                            """)
                                            self._log(f"✅ เลือกเปอร์เซ็นต์หัก ณ ที่จ่าย (แถวที่ {row_index}): {wht_text} สำเร็จด้วย JavaScript")
                                            time.sleep(0.3)
                                    except Exception as e:
                                        self._log(f"⚠️ ไม่สามารถเลือกเปอร์เซ็นต์หัก ณ ที่จ่าย (แถวที่ {row_index}) ได้: {e}", level="warning")
                                else:
                                    self._log(f"⚠️ ไม่รองรับเปอร์เซ็นต์หัก ณ ที่จ่าย (แถวที่ {row_index}): {withholding_float}%", level="warning")
                            else:
                                self._log(f"⚠️ ไม่พบ whtDropDown{row_index}", level="warning")
                    except Exception as e:
                        self._log(f"⚠️ เกิดข้อผิดพลาดในการเลือกเปอร์เซ็นต์หัก ณ ที่จ่าย (แถวที่ {row_index}): {e}", level="warning")
            
            # หมายเหตุ: การกรอกหมายเหตุ, เลือกรวมภาษี, คลิกปุ่ม hidePaymentModal, และคลิกปุ่มสุดท้าย
            # จะทำหลังจากกรอกแถวทั้งหมดเสร็จแล้ว (ดูใน process_excel_data_and_fill_form)
            
            self._log(f"✅ กรอกข้อมูลจาก Excel row เสร็จสิ้น (แถวที่ {row_index})")
            return True
            
        except Exception as e:
            print(f"❌ เกิดข้อผิดพลาดในการกรอกข้อมูลจาก Excel row: {e}")
            import traceback
            traceback.print_exc()
            return False
