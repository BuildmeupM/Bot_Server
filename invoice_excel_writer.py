"""
Excel Writer Service (Version 3.0)
==================================
Service สำหรับเขียนข้อมูลลง Excel file พร้อม 3 Sheets

Sheets:
1. มีภาษีมูลค่าเพิ่ม
2. ไม่มีภาษีมูลค่าเพิ่ม
3. ที่อยู่แต่ละบริษัท

Author: BotV3
Version: 3.0.0
"""

import logging
import shutil
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class InvoiceExcelWriter:
    """Service สำหรับเขียนข้อมูลลง Excel พร้อม 3 Sheets"""
    
    # ชื่อ Sheets
    SHEET_NAMES = {
        1: "มีภาษีมูลค่าเพิ่ม",
        2: "ไม่มีภาษีมูลค่าเพิ่ม",
        3: "ที่อยู่แต่ละบริษัท"
    }
    
    # Header columns สำหรับชีต Invoice
    HEADERS = [
        "ลำดับ",
        "ชื่อบริษัท",
        "เลขประจำตัวผู้เสียภาษี",
        "สาขา",
        "วันที่",
        "ครบกำหนดชำระ",
        "เลขที่เอกสาร",
        "ชื่อบัญชี / โค้ดบัญชี",
        "เปอร์เซ็นต์หัก ณ ที่จ่าย",
        "ยอดก่อนภาษีมูลค่าเพิ่ม",
        "ยอดภาษีมูลค่าเพิ่ม",
        "ยอดหลังบวกภาษีมูลค่าเพิ่ม",
        "อ้างอิง",
        "หมายเหตุ",
        "ชื่อไฟล์ใหม่",
        "ชื่อไฟล์เก่า",
        "ชื่อผู้ซื้อ",
        "เลขประจำตัวผู้เสียภาษี - ผู้ซื้อ",
        "ที่อยู่ผู้ซื้อ",
        "ประเภทเอกสาร",
        "สถานะเอกสาร"
    ]
    
    # Header columns สำหรับชีตที่อยู่
    ADDRESS_HEADERS = [
        "ลำดับ",
        "ชื่อบริษัท",
        "เลขประจำตัวผู้เสียภาษี",
        "ที่อยู่ (รวม)",
        "เลขที่",
        "อื่นๆ",
        "ซอย/ตรอก",
        "ถนน",
        "แขวง/ตำบล",
        "เขต/อำเภอ",
        "จังหวัด",
        "รหัสไปรษณีย์"
    ]
    
    def __init__(self, excel_path: Optional[str] = None):
        """
        Initialize Excel Writer
        
        Args:
            excel_path: Path ของไฟล์ Excel (ถ้าไม่ระบุจะสร้างใหม่)
        """
        self.excel_path = Path(excel_path) if excel_path else None
        self.workbook = None
        self.sheets = {}
    
    def create_new_workbook(self, filename: str = "Invoice_Data.xlsx", target_folder: str = None) -> Path:
        """
        สร้าง Excel workbook ใหม่พร้อม 3 Sheets
        
        Args:
            filename: ชื่อไฟล์ Excel
            target_folder: โฟลเดอร์ที่ต้องการบันทึก (ถ้าไม่ระบุใช้ excel_exports)
        
        Returns:
            Path ของไฟล์ที่สร้าง
        """
        self.workbook = Workbook()
        
        # ลบ sheet default
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # สร้าง 3 Sheets
        for sheet_type, sheet_name in self.SHEET_NAMES.items():
            ws = self.workbook.create_sheet(sheet_name)
            self.sheets[sheet_type] = ws
            
            # เพิ่ม Header ตามประเภทชีต
            if sheet_type == 3:  # ชีตที่อยู่
                ws.append(self.ADDRESS_HEADERS)
                self._format_address_header(ws)
            else:  # ชีต Invoice (มีภาษี/ไม่มีภาษี)
                ws.append(self.HEADERS)
                self._format_header(ws)
        
        # กำหนด path
        if not self.excel_path:
            if target_folder:
                # ใช้โฟลเดอร์ที่ผู้ใช้ระบุ
                output_dir = Path(target_folder)
            else:
                # ใช้โฟลเดอร์ default
                output_dir = Path("excel_exports")
            
            output_dir.mkdir(exist_ok=True)
            self.excel_path = output_dir / filename
        
        return self.excel_path
    
    def load_existing_workbook(self, excel_path: str) -> bool:
        """
        โหลด Excel workbook ที่มีอยู่แล้ว
        
        Args:
            excel_path: Path ของไฟล์ Excel
        
        Returns:
            True ถ้าโหลดสำเร็จ
        """
        try:
            self.excel_path = Path(excel_path)
            if not self.excel_path.exists():
                logger.warning(f"ไม่พบไฟล์ {excel_path} - จะสร้างใหม่")
                return False
            
            # สำคัญ: ต้องปิดไฟล์ก่อนโหลดใหม่ (ถ้ามีการเปิดอยู่) เพื่อให้ได้ข้อมูลล่าสุด
            # และต้องใช้ data_only=False เพื่อให้อ่านค่าที่คำนวณได้ (เช่น max_row) ถูกต้อง
            self.workbook = load_workbook(self.excel_path, data_only=False)
            
            # Map sheets
            for sheet_type, sheet_name in self.SHEET_NAMES.items():
                if sheet_name in self.workbook.sheetnames:
                    self.sheets[sheet_type] = self.workbook[sheet_name]
                    # Log จำนวนแถวปัจจุบันของแต่ละ sheet เพื่อ debug
                    max_row = self.sheets[sheet_type].max_row or 1
                    logger.info(f"📊 โหลด sheet '{sheet_name}': มี {max_row} แถว (รวม header)")
                else:
                    logger.warning(f"ไม่พบ sheet: {sheet_name}")
                    # สร้าง sheet ใหม่ถ้าไม่มี
                    ws = self.workbook.create_sheet(sheet_name)
                    self.sheets[sheet_type] = ws
                    if sheet_type == 3:  # ชีตที่อยู่
                        ws.append(self.ADDRESS_HEADERS)
                        self._format_address_header(ws)
                    else:  # ชีต Invoice
                        ws.append(self.HEADERS)
                        self._format_header(ws)
            
            return True
        except Exception as e:
            logger.error(f"เกิดข้อผิดพลาดในการโหลดไฟล์: {e}")
            return False
    
    def _format_header(self, worksheet):
        """Format header row ให้สวยงาม"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name='TH Sarabun New', size=14, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # ตั้งความกว้างคอลัมน์
        column_widths = {
            1: 8,   # ลำดับ
            2: 40,  # ชื่อบริษัท
            3: 20,  # เลขประจำตัวผู้เสียภาษี
            4: 12,  # สาขา
            5: 15,  # วันที่
            6: 18,  # ครบกำหนดชำระ
            7: 18,  # เลขที่เอกสาร
            8: 30,  # ชื่อบัญชี / โค้ดบัญชี
            9: 18,  # เปอร์เซ็นต์หัก ณ ที่จ่าย
            10: 20, # ยอดก่อนภาษี
            11: 20, # ยอดภาษี
            12: 20, # ยอดหลังบวกภาษี
            13: 20, # อ้างอิง
            14: 25, # หมายเหตุ
            15: 25, # ชื่อไฟล์ใหม่
            16: 30, # ชื่อไฟล์เก่า
            17: 30, # ชื่อผู้ซื้อ
            18: 20, # เลขประจำตัวผู้เสียภาษี - ผู้ซื้อ
            19: 50, # ที่อยู่ผู้ซื้อ
            20: 25, # ประเภทเอกสาร
            21: 15  # สถานะเอกสาร
        }
        
        for col_idx, width in column_widths.items():
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width
        
        # ตั้งความสูงของ header
        worksheet.row_dimensions[1].height = 30
    
    def _format_address_header(self, worksheet):
        """Format header row สำหรับชีตที่อยู่"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name='TH Sarabun New', size=14, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # ตั้งความกว้างคอลัมน์สำหรับชีตที่อยู่
        column_widths = {
            1: 8,   # ลำดับ
            2: 40,  # ชื่อบริษัท
            3: 20,  # เลขประจำตัวผู้เสียภาษี
            4: 80,  # ที่อยู่ (รวม)
            5: 15,  # เลขที่
            6: 20,  # อื่นๆ
            7: 20,  # ซอย/ตรอก
            8: 25,  # ถนน
            9: 20,  # แขวง/ตำบล
            10: 20, # เขต/อำเภอ
            11: 15, # จังหวัด
            12: 15  # รหัสไปรษณีย์
        }
        
        for col_idx, width in column_widths.items():
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width
        
        # ตั้งความสูงของ header
        worksheet.row_dimensions[1].height = 30
    
    def add_row(self, data: Dict[str, Any], override_row_number: Optional[int] = None) -> bool:
        """
        เพิ่มข้อมูลแถวใหม่
        
        Args:
            data: Dictionary ที่มีข้อมูล:
                - document_type: ประเภทเอกสาร (1-4)
                - company_name: ชื่อบริษัท
                - tax_id: เลขประจำตัวผู้เสียภาษี
                - date: วันที่
                - account_name: ชื่อบัญชี
                - account_code: โค้ดบัญชี
                - withholding_tax_percent: เปอร์เซ็นต์หัก ณ ที่จ่าย (%)
                - amount_before_vat: ยอดก่อนภาษี
                - vat_amount: ยอดภาษี
                - total_amount: ยอดหลังบวกภาษี
                - remark: หมายเหตุ
                - new_filename: ชื่อไฟล์ใหม่
                - old_filename: ชื่อไฟล์เก่า
            override_row_number: บังคับใช้ลำดับที่ระบุ (สำหรับแถวที่ใช้ลำดับเดียวกัน)
        
        Returns:
            True ถ้าเพิ่มสำเร็จ
        """
        try:
            # Debug: แสดงข้อมูลที่ได้รับ
            logger.info(f"🔍 [DEBUG Excel] ข้อมูลที่ได้รับ:")
            logger.info(f"   company_name: {data.get('company_name', 'N/A')}")
            logger.info(f"   document_type (เดิม): {data.get('document_type')}")
            logger.info(f"   vat_amount: {data.get('vat_amount')}")
            logger.info(f"   amount_before_vat: {data.get('amount_before_vat')}")
            logger.info(f"   total_amount: {data.get('total_amount')}")
            
            document_type = data.get('document_type', 2)  # Default: ไม่มีภาษี
            if document_type is None:
                document_type = 2  # Default: ไม่มีภาษี
            
            # ตรวจสอบ vat_amount เพื่อกำหนด sheet ที่ถูกต้อง
            vat_amount = data.get('vat_amount')
            if vat_amount is None:
                vat_amount = 0
            elif isinstance(vat_amount, str):
                try:
                    vat_amount = float(vat_amount)
                except (ValueError, TypeError):
                    vat_amount = 0
            else:
                vat_amount = float(vat_amount) if vat_amount else 0
            
            # ตรวจสอบว่าต้องการบังคับให้อยู่ใน sheet "มีภาษีมูลค่าเพิ่ม" หรือไม่ (สำหรับกรณี CUSTOMS_DEPARTMENT ที่มี 2 แถว)
            force_vat_sheet = data.get('force_vat_sheet', False)
            
            # แปลง document_type: 0,3,4 -> 1 หรือ 2 (0 = ไม่มีภาษี -> 2)
            if document_type == 0:
                document_type = 2  # ไม่มีภาษีมูลค่าเพิ่ม
            elif document_type == 3 or document_type == 4:
                # ถ้ามีภาษี ให้ใช้ sheet 1, ถ้าไม่มีภาษี ให้ใช้ sheet 2
                document_type = 1 if vat_amount > 0 else 2
            elif document_type == 1:
                # ถ้า document_type = 1 แต่ vat_amount = 0 หรือ null ให้ส่งไปชีต "ไม่มีภาษีมูลค่าเพิ่ม"
                # ยกเว้นกรณีที่ force_vat_sheet = True (สำหรับ CUSTOMS_DEPARTMENT ที่มี 2 แถว)
                if vat_amount <= 0 and not force_vat_sheet:
                    document_type = 2
                    logger.info(f"⚠️ [DEBUG Excel] document_type=1 แต่ vat_amount={vat_amount} → เปลี่ยนเป็น sheet 'ไม่มีภาษีมูลค่าเพิ่ม'")
                elif vat_amount <= 0 and force_vat_sheet:
                    logger.info(f"✅ [DEBUG Excel] document_type=1, vat_amount={vat_amount} แต่ force_vat_sheet=True → ยังคงอยู่ใน sheet 'มีภาษีมูลค่าเพิ่ม'")
            
            logger.info(f"✅ [DEBUG Excel] document_type (สุดท้าย): {document_type} → Sheet: {self.SHEET_NAMES.get(document_type, 'Unknown')}")
            
            if document_type not in [1, 2]:  # รับแค่ 1 หรือ 2
                logger.error(f"ประเภทเอกสารไม่ถูกต้อง: {document_type} (ต้องเป็น 1 หรือ 2)")
                return False
            
            if document_type not in self.sheets:
                logger.error(f"ไม่พบ sheet สำหรับประเภทเอกสาร: {document_type}")
                return False
            
            worksheet = self.sheets[document_type]
            
            # คำนวณลำดับ (นับแยกกันในแต่ละชีต เริ่มจาก 1, 2, 3...)
            # สำคัญ: ต้องคำนวณลำดับใหม่ทุกครั้งที่เพิ่มแถว เพื่อให้แต่ละรายการมีลำดับที่ถูกต้อง
            if override_row_number is not None:
                row_number = override_row_number
                logger.info(f"📊 [ลำดับ] ใช้ override_row_number: {row_number} (Sheet: {self.SHEET_NAMES.get(document_type, 'Unknown')})")
            else:
                # max_row = 1 ถ้ามีแค่ header, max_row = 2 ถ้ามีข้อมูล 1 แถว, max_row = 3 ถ้ามีข้อมูล 2 แถว
                # จำนวนแถวข้อมูล = max_row - 1 (ลบ header row ออก)
                # ลำดับถัดไป = จำนวนแถวข้อมูล + 1
                # แต่ถ้ายังไม่มีข้อมูลเลย (max_row = 1) จะได้ลำดับ = 1
                # สำคัญ: ต้องอ่าน max_row ใหม่ทุกครั้งเพื่อให้ได้ค่าล่าสุด
                max_row = worksheet.max_row or 1  # ถ้าเป็น None ให้ใช้ 1
                data_row_count = max_row - 1  # จำนวนแถวข้อมูลที่มีอยู่แล้ว
                row_number = max(1, data_row_count + 1)  # ลำดับถัดไป
                logger.info(f"📊 [ลำดับ] Sheet '{self.SHEET_NAMES.get(document_type, 'Unknown')}': max_row={max_row}, data_row_count={data_row_count}, row_number={row_number} (company: {data.get('company_name', 'N/A')[:30]}...)")
            
            # จัดรูปแบบ Account Name / Account Code
            account_name = data.get('account_name', '')
            account_code = data.get('account_code', '')
            if account_name and account_code:
                account_display = f"{account_name} / {account_code}"
            elif account_name:
                account_display = account_name
            elif account_code:
                account_display = account_code
            else:
                account_display = ''
            
            # จัดรูปแบบเปอร์เซ็นต์หัก ณ ที่จ่าย (แสดงเป็นตัวเลขเฉยๆ ไม่มี %)
            withholding_percent = data.get('withholding_tax_percent', 0) or 0
            
            # ถ้าเป็น string ให้ตัด % ออกและแปลงเป็นตัวเลข
            if isinstance(withholding_percent, str):
                # ตัด % และช่องว่างออก
                withholding_percent = withholding_percent.replace('%', '').strip()
                try:
                    withholding_percent = float(withholding_percent)
                except (ValueError, TypeError):
                    withholding_percent = 0
            
            if withholding_percent and withholding_percent > 0:
                # แสดงเป็นตัวเลขเต็ม (เช่น 3 ไม่ใช่ 3.00 หรือ 3%)
                withholding_display = str(int(round(withholding_percent)))
            else:
                withholding_display = ''
            
            # เตรียมข้อมูลแถว
            row_data = [
                row_number,  # ลำดับ
                data.get('company_name', ''),
                data.get('tax_id', ''),
                data.get('branch', ''),  # สาขา
                data.get('date', ''),
                data.get('due_date', ''),  # ครบกำหนดชำระ
                data.get('document_number', ''),  # เลขที่เอกสาร
                account_display,
                withholding_display,  # เปอร์เซ็นต์หัก ณ ที่จ่าย
                data.get('amount_before_vat') or 0,  # ถ้าเป็น None ให้ใช้ 0
                data.get('vat_amount') or 0,  # ถ้าเป็น None ให้ใช้ 0
                data.get('total_amount') or 0,  # ถ้าเป็น None ให้ใช้ 0
                data.get('reference', ''),  # อ้างอิง
                data.get('remark', ''),
                data.get('new_filename', ''),
                data.get('old_filename', ''),
                data.get('buyer_name', ''),  # ชื่อผู้ซื้อ
                data.get('buyer_tax_id', ''),  # เลขประจำตัวผู้เสียภาษี - ผู้ซื้อ
                data.get('buyer_address', ''),  # ที่อยู่ผู้ซื้อ
                data.get('document_type_text', ''),  # ประเภทเอกสาร (ข้อความ)
                data.get('document_status', '')  # สถานะเอกสาร
            ]
            
            worksheet.append(row_data)
            
            # Format แถวที่เพิ่ม
            row_num = worksheet.max_row or 1  # ถ้าเป็น None ให้ใช้ 1
            self._format_data_row(worksheet, row_num)
            
            return True
        except Exception as e:
            logger.error(f"เกิดข้อผิดพลาดในการเพิ่มข้อมูล: {e}", exc_info=True)
            logger.error(f"ข้อมูลที่ส่งมา: company_name={data.get('company_name')}, document_type={data.get('document_type')}, tax_id={data.get('tax_id')}")
            return False
    
    def _format_data_row(self, worksheet, row_num: int):
        """Format data row"""
        data_font = Font(name='TH Sarabun New', size=13)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        center_alignment = Alignment(horizontal='center', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alternate row colors
        if row_num % 2 == 0:
            fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for idx, cell in enumerate(worksheet[row_num], 1):
            cell.font = data_font
            cell.border = border
            cell.fill = fill
            
            # จัด alignment ตามประเภทข้อมูล
            if idx == 1:  # ลำดับ
                cell.alignment = center_alignment
            elif idx == 4:  # สาขา
                cell.alignment = center_alignment
            elif idx == 6:  # ครบกำหนดชำระ
                cell.alignment = center_alignment
            elif idx == 7:  # เลขที่เอกสาร
                cell.alignment = center_alignment
            elif idx == 9:  # เปอร์เซ็นต์หัก ณ ที่จ่าย
                cell.alignment = center_alignment
            elif idx in [10, 11, 12]:  # คอลัมน์ตัวเลข (ยอดเงิน)
                cell.alignment = number_alignment
                # Format เป็นตัวเลข 2 ทศนิยม
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = data_alignment
    
    def save(self) -> bool:
        """
        บันทึกไฟล์ Excel
        
        Returns:
            True ถ้าบันทึกสำเร็จ
        """
        try:
            if not self.workbook:
                logger.error("ไม่มี workbook ให้บันทึก")
                return False
            
            if not self.excel_path:
                logger.error("ไม่ได้ระบุ path ของไฟล์")
                return False
            
            self.workbook.save(self.excel_path)
            logger.info(f"✅ บันทึกไฟล์สำเร็จ: {self.excel_path}")
            return True
        except Exception as e:
            logger.error(f"เกิดข้อผิดพลาดในการบันทึก: {e}")
            return False
    
    def add_address_row(self, data: Dict[str, Any]) -> bool:
        """
        เพิ่มข้อมูลที่อยู่บริษัทในชีตที่อยู่
        
        Args:
            data: Dictionary ที่มีข้อมูล:
                - company_name: ชื่อบริษัท
                - tax_id: เลขประจำตัวผู้เสียภาษี
                - address: ที่อยู่ (รวม) หรือ address_dict (ที่แยกแล้ว)
                - address_full: ที่อยู่รวม (ถ้ามี)
                - building_number: เลขที่
                - other_info: อื่นๆ (เช่น ชื่ออาคาร, Building)
                - soi: ซอย/ตรอก
                - road: ถนน
                - subdistrict: แขวง/ตำบล
                - district: เขต/อำเภอ
                - province: จังหวัด
                - postal_code: รหัสไปรษณีย์
        
        Returns:
            True ถ้าเพิ่มสำเร็จ
        """
        try:
            if 3 not in self.sheets:
                logger.error("ไม่พบ sheet ที่อยู่แต่ละบริษัท")
                return False
            
            worksheet = self.sheets[3]
            
            # ตรวจสอบว่ามีข้อมูลนี้อยู่แล้วหรือไม่ (เช็คจาก company_name และ address)
            company_name = data.get('company_name', '').strip()
            address_text = data.get('address') or data.get('address_full', '')
            address_full = address_text.strip() if address_text else ''
            
            logger.info(f"🔍 ตรวจสอบที่อยู่: Company={company_name}")
            logger.info(f"🔍 Address from data.get('address'): {data.get('address', '(not found)')}")
            logger.info(f"🔍 Address from data.get('address_full'): {data.get('address_full', '(not found)')}")
            logger.info(f"🔍 Final address_full: {address_full[:50] if address_full else '(empty)'}...")
            
            # ตรวจสอบว่ามี company_name และ address_full นี้อยู่แล้วหรือไม่
            if company_name and address_full:
                max_row = worksheet.max_row or 1
                if max_row > 1:  # มีข้อมูลมากกว่า header
                    for row in worksheet.iter_rows(min_row=2, max_row=max_row):
                        existing_company = str(row[1].value or '').strip()  # คอลัมน์ที่ 2 = company_name
                        existing_address = str(row[3].value or '').strip()  # คอลัมน์ที่ 4 = address_full
                        
                        # เปรียบเทียบทั้ง company_name และ address_full
                        if existing_company == company_name and existing_address == address_full:
                            logger.info(f"⚠️ มีข้อมูลที่อยู่ของบริษัทนี้อยู่แล้ว (Company: {company_name}, Address: {address_full[:50]}...)")
                            return True  # ไม่ต้องเพิ่มซ้ำ
            
            # Fallback: ตรวจสอบจาก tax_id (ถ้ามี)
            tax_id = data.get('tax_id', '').strip()
            if tax_id:
                # ตรวจสอบว่ามี tax_id นี้อยู่แล้วหรือไม่
                max_row = worksheet.max_row or 1
                if max_row > 1:  # มีข้อมูลมากกว่า header
                    for row in worksheet.iter_rows(min_row=2, max_row=max_row):
                        existing_tax_id = str(row[2].value or '').strip()  # คอลัมน์ที่ 3 = tax_id
                        if existing_tax_id == tax_id:
                            logger.info(f"⚠️ มีข้อมูลที่อยู่ของบริษัทนี้อยู่แล้ว (Tax ID: {tax_id})")
                            return True  # ไม่ต้องเพิ่มซ้ำ
            
            # ถ้าไม่มีที่อยู่ ให้ log และ return False
            if not address_full:
                logger.warning(f"⚠️ ไม่พบที่อยู่สำหรับบริษัท: {company_name}")
                return False
            
            # คำนวณลำดับ (ถ้ามีแค่ header row = 1, ลำดับจะเป็น 1)
            # max_row = 1 ถ้ามีแค่ header, max_row = 2 ถ้ามีข้อมูล 1 แถว
            # ลำดับ = max_row - 1 (ถ้ามีแค่ header = 0, แต่เราต้องการ 1)
            max_row = worksheet.max_row or 1  # ถ้าเป็น None ให้ใช้ 1
            row_number = max(1, max_row - 1)
            
            # ถ้ามีข้อมูลที่แยกแล้วให้ใช้เลย ไม่เช่นนั้นให้ parse
            if all(key in data for key in ['address_full', 'building_number', 'other_info', 
                                           'soi', 'road', 'subdistrict', 'district', 
                                           'province', 'postal_code']):
                address_dict = data
                logger.info(f"✅ ใช้ข้อมูลที่อยู่ที่แยกแล้ว")
            else:
                # Parse ที่อยู่ออกเป็นส่วนๆ
                address_text = data.get('address') or data.get('address_full', '')
                logger.info(f"🔍 Parse ที่อยู่: {address_text[:100] if address_text else '(empty)'}...")
                address_dict = parse_address(address_text)
                address_dict['address_full'] = address_text  # ใช้ที่อยู่ที่ส่งมา
                logger.info(f"✅ Parse ที่อยู่สำเร็จ: เลขที่={address_dict.get('building_number')}, ถนน={address_dict.get('road')}, เขต={address_dict.get('district')}, แขวง={address_dict.get('subdistrict')}, จังหวัด={address_dict.get('province')}, รหัสไปรษณีย์={address_dict.get('postal_code')}")
            
            # เตรียมข้อมูลแถว
            row_data = [
                row_number,  # ลำดับ
                data.get('company_name', ''),
                data.get('tax_id', ''),
                address_dict.get('address_full', ''),  # ที่อยู่ (รวม)
                address_dict.get('building_number', ''),  # เลขที่
                address_dict.get('other_info', ''),  # อื่นๆ
                address_dict.get('soi', ''),  # ซอย/ตรอก
                address_dict.get('road', ''),  # ถนน
                address_dict.get('subdistrict', ''),  # แขวง/ตำบล
                address_dict.get('district', ''),  # เขต/อำเภอ
                address_dict.get('province', ''),  # จังหวัด
                address_dict.get('postal_code', '')  # รหัสไปรษณีย์
            ]
            
            worksheet.append(row_data)
            
            # Format แถวที่เพิ่ม
            row_num = worksheet.max_row
            self._format_address_data_row(worksheet, row_num)
            
            logger.info(f"✅ เพิ่มข้อมูลที่อยู่สำเร็จ: Company={company_name}, Row={row_num}, Address={address_full[:50] if address_full else '(empty)'}...")
            
            return True
        except Exception as e:
            logger.error(f"เกิดข้อผิดพลาดในการเพิ่มข้อมูลที่อยู่: {e}")
            return False
    
    def _format_address_data_row(self, worksheet, row_num: int):
        """Format data row สำหรับชีตที่อยู่"""
        data_font = Font(name='TH Sarabun New', size=13)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alternate row colors
        if row_num % 2 == 0:
            fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for idx, cell in enumerate(worksheet[row_num], 1):
            cell.font = data_font
            cell.border = border
            cell.fill = fill
            
            # จัด alignment
            if idx == 1:  # ลำดับ
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment
    
    def add_multiple_rows(self, data_list: List[Dict[str, Any]]) -> int:
        """
        เพิ่มข้อมูลหลายแถว
        
        Args:
            data_list: List of dictionaries
        
        Returns:
            จำนวนแถวที่เพิ่มสำเร็จ
        """
        success_count = 0
        for data in data_list:
            if self.add_row(data):
                success_count += 1
        return success_count


# ===== Helper Functions =====

def parse_address(address_text: Optional[str]) -> Dict[str, Optional[str]]:
    """
    แยกที่อยู่เป็นส่วนๆ
    
    Args:
        address_text: ที่อยู่รวม (string)
    
    Returns:
        Dictionary ที่มี:
        - address_full: ที่อยู่รวม
        - building_number: เลขที่
        - other_info: อื่นๆ (เช่น ชื่ออาคาร, Building)
        - soi: ซอย/ตรอก
        - road: ถนน
        - subdistrict: แขวง/ตำบล
        - district: เขต/อำเภอ
        - province: จังหวัด
        - postal_code: รหัสไปรษณีย์
    """
    if not address_text:
        return {
            'address_full': '',
            'building_number': '',
            'other_info': '',
            'soi': '',
            'road': '',
            'subdistrict': '',
            'district': '',
            'province': '',
            'postal_code': ''
        }
    
    import re
    
    # ทำความสะอาดที่อยู่
    address_clean = re.sub(r'\s+', ' ', address_text.strip())
    
    result = {
        'address_full': address_clean,
        'building_number': '',
        'other_info': '',
        'soi': '',
        'road': '',
        'subdistrict': '',
        'district': '',
        'province': '',
        'postal_code': ''
    }
    
    # ลบข้อมูลที่ไม่ใช่ที่อยู่ (เช่น Tel:, Fax:, Phone:)
    address_clean = re.sub(r'\s*(?:Tel|Fax|Phone|โทร|โทรศัพท์):\s*[^\s,]+.*$', '', address_clean, flags=re.IGNORECASE).strip()
    
    # เก็บ address_original สำหรับ parse จากท้ายไปหน้า
    address_original = address_clean
    
    # ===== Parse จากท้ายไปหน้า (รหัสไปรษณีย์ → จังหวัด → เขต/อำเภอ → แขวง/ตำบล) =====
    
    # หารหัสไปรษณีย์ (5 หลัก, มักจะอยู่ท้ายสุด ก่อน Tel:)
    postal_pattern = r'(\d{5})(?:\s|,|$|Tel|Fax|Phone)'
    postal_match = re.search(postal_pattern, address_original)
    if postal_match:
        result['postal_code'] = postal_match.group(1)
        # ตัดรหัสไปรษณีย์ออก
        address_clean = re.sub(rf'{re.escape(postal_match.group(1))}(?:\s|,|$)', '', address_clean).strip()
    
    # ประกาศรายชื่อจังหวัด (ต้องประกาศก่อนใช้งาน)
    provinces_th = ['กรุงเทพมหานคร', 'กระบี่', 'กาญจนบุรี', 'กาฬสินธุ์', 'กำแพงเพชร', 'ขอนแก่น', 
                    'จันทบุรี', 'ฉะเชิงเทรา', 'ชลบุรี', 'ชัยนาท', 'ชัยภูมิ', 'ชุมพร', 'เชียงราย', 
                    'เชียงใหม่', 'ตรัง', 'ตราด', 'ตาก', 'นครนายก', 'นครปฐม', 'นครพนม', 'นครราชสีมา',
                    'นครศรีธรรมราช', 'นครสวรรค์', 'นนทบุรี', 'นราธิวาส', 'น่าน', 'บึงกาฬ', 'บุรีรัมย์',
                    'ปทุมธานี', 'ประจวบคีรีขันธ์', 'ปราจีนบุรี', 'ปัตตานี', 'พระนครศรีอยุธยา', 'พังงา',
                    'พัทลุง', 'พิจิตร', 'พิษณุโลก', 'เพชรบุรี', 'เพชรบูรณ์', 'แพร่', 'ภูเก็ต', 'มหาสารคาม',
                    'มุกดาหาร', 'แม่ฮ่องสอน', 'ยะลา', 'ยโสธร', 'ร้อยเอ็ด', 'ระนอง', 'ระยอง', 'ราชบุรี',
                    'ลพบุรี', 'ลำปาง', 'ลำพูน', 'เลย', 'ศรีสะเกษ', 'สกลนคร', 'สงขลา', 'สตูล', 'สมุทรปราการ',
                    'สมุทรสงคราม', 'สมุทรสาคร', 'สระแก้ว', 'สระบุรี', 'สิงห์บุรี', 'สุโขทัย', 'สุพรรณบุรี',
                    'สุราษฎร์ธานี', 'สุรินทร์', 'หนองคาย', 'หนองบัวลำภู', 'อ่างทอง', 'อำนาจเจริญ',
                    'อุดรธานี', 'อุตรดิตถ์', 'อุทัยธานี', 'อุบลราชธานี']
    
    provinces_en = ['Bangkok', 'Krabi', 'Kanchanaburi', 'Kalasin', 'Kamphaeng Phet', 'Khon Kaen',
                    'Chanthaburi', 'Chachoengsao', 'Chon Buri', 'Chai Nat', 'Chaiyaphum', 'Chumphon', 'Chiang Rai',
                    'Chiang Mai', 'Trang', 'Trat', 'Tak', 'Nakhon Nayok', 'Nakhon Pathom', 'Nakhon Phanom', 'Nakhon Ratchasima',
                    'Nakhon Si Thammarat', 'Nakhon Sawan', 'Nonthaburi', 'Narathiwat', 'Nan', 'Bueng Kan', 'Buriram',
                    'Pathum Thani', 'Prachuap Khiri Khan', 'Prachin Buri', 'Pattani', 'Phra Nakhon Si Ayutthaya', 'Phang Nga',
                    'Phatthalung', 'Phichit', 'Phitsanulok', 'Phetchaburi', 'Phetchabun', 'Phrae', 'Phuket', 'Maha Sarakham',
                    'Mukdahan', 'Mae Hong Son', 'Yala', 'Yasothon', 'Roi Et', 'Ranong', 'Rayong', 'Ratchaburi',
                    'Lop Buri', 'Lampang', 'Lamphun', 'Loei', 'Si Sa Ket', 'Sakon Nakhon', 'Songkhla', 'Satun', 'Samut Prakan',
                    'Samut Songkhram', 'Samut Sakhon', 'Sa Kaeo', 'Saraburi', 'Sing Buri', 'Sukhothai', 'Suphan Buri',
                    'Surat Thani', 'Surin', 'Nong Khai', 'Nong Bua Lamphu', 'Ang Thong', 'Amnat Charoen',
                    'Udon Thani', 'Uttaradit', 'Uthai Thani', 'Ubon Ratchathani']
    
    # หาจังหวัด (รองรับทั้งภาษาไทยและภาษาอังกฤษ)
    # ต้องหา "กทม." (กรุงเทพมหานคร) ก่อน เพราะอาจจะถูกตัดออกไปก่อน
    # Pattern: "กทม." หรือ "กรุงเทพมหานคร"
    province_found = False
    if re.search(r'กทม\.', address_original, re.IGNORECASE):
        result['province'] = 'กรุงเทพมหานคร'
        province_found = True
        # ตัด "กทม." ออก
        address_clean = re.sub(r'กทม\.', '', address_clean, flags=re.IGNORECASE).strip()
    
    # หาจังหวัด (รองรับทั้งภาษาไทยและภาษาอังกฤษ)
    # ต้องหา "จ." ก่อน เพราะอาจจะถูกตัดออกไปก่อน
    # Pattern: "จ.ชลบุรี" หรือ "จังหวัดชลบุรี"
    if not province_found:
        province_pattern_short = r'จ\.\s*([ก-๙A-Za-z\s\-]+?)(?:\s+\d{5}|\s*$)'
        province_match_short = re.search(province_pattern_short, address_original, re.IGNORECASE)
        if province_match_short:
            province_name = province_match_short.group(1).strip()
            # ตรวจสอบว่าเป็นชื่อจังหวัดที่ถูกต้อง (อยู่ใน list)
            for province in provinces_th + provinces_en:
                if province == province_name or province_name in province or province in province_name:
                    result['province'] = province
                    province_found = True
                    # ตัดจังหวัดออก
                    address_clean = re.sub(rf'{re.escape(province_match_short.group(0))}', '', address_clean).strip()
                    break
    
    # ถ้ายังไม่พบจังหวัด ให้ลองหาจาก pattern ปกติ
    if not province_found:
        for province in provinces_th + provinces_en:
            # ใช้ word boundary เพื่อไม่ให้ match ส่วนของคำอื่น
            pattern = rf'\b{re.escape(province)}\b'
            if re.search(pattern, address_clean, re.IGNORECASE):
                result['province'] = province
                # ตัดจังหวัดและส่วนที่เหลือออก (เพราะจังหวัดมักอยู่ท้ายสุด)
                address_clean = re.sub(rf'{re.escape(province)}.*$', '', address_clean).strip()
                province_found = True
                break
    
    # หาอำเภอ (ต้องหาจาก address_original ก่อน เพราะอาจจะถูกตัดออกไป)
    # Pattern: "อ.ศรีราชา" หรือ "อำเภอศรีราชา"
    if not result.get('district'):
        district_pattern_short = r'อ\.\s*([ก-๙A-Za-z\s\-]+?)(?:\s+(?:จ\.|จังหวัด)|\s+\d{5}|\s*$)'
        district_match_short = re.search(district_pattern_short, address_original, re.IGNORECASE)
        if district_match_short:
            district_name = district_match_short.group(1).strip()
            # ตรวจสอบว่าไม่ใช่ส่วนของจังหวัดหรือแขวง/ตำบล
            if district_name and district_name not in [result.get('province', ''), result.get('subdistrict', '')]:
                result['district'] = district_name
                # ตัดอำเภอออกจาก address_clean
                address_clean = re.sub(rf'{re.escape(district_match_short.group(0))}', '', address_clean).strip()
    
    # หาเขต/อำเภอ (รองรับทั้งภาษาไทยและภาษาอังกฤษ)
    # Pattern: ชื่อเขต/อำเภอที่อยู่ก่อนจังหวัด (มักจะเป็นคำสุดท้ายก่อนจังหวัด)
    # ตัวอย่าง: "Vadhana" ก่อน "Bangkok"
    if province_found and not result.get('district'):
        # หาจากส่วนที่เหลือก่อนจังหวัด
        # Pattern: ชื่อเขต/อำเภอ (มักจะเป็นคำภาษาอังกฤษหรือภาษาไทย) ก่อนจังหวัด
        # ใช้ lookahead เพื่อหา pattern ที่อยู่ก่อนจังหวัด
        district_pattern = r'([ก-๙A-Za-z\s\-]+?)(?:\s*,\s*|\s+)(?:' + re.escape(result['province']) + r'|จังหวัด)'
        district_match = re.search(district_pattern, address_original, re.IGNORECASE)
        if district_match:
            district_name = district_match.group(1).strip()
            # ตรวจสอบว่าไม่ใช่ถนนหรือแขวง/ตำบล
            if not re.search(r'(?:ถนน|Road|Rd\.|แขวง|ตำบล)', district_name, re.IGNORECASE):
                result['district'] = district_name
                # ตัดเขต/อำเภอออกจาก address_clean
                address_clean = re.sub(rf'{re.escape(district_name)}\s*,\s*', '', address_clean).strip()
        else:
            # ลองหาแบบมีคำว่า "เขต" หรือ "อำเภอ" หรือ "อ."
            # รองรับทั้งกรณีที่มี comma/space หลังเขต/อำเภอ และกรณีที่ไม่มี (เช่น "เขตลาดกระบัง กรุงเทพมหานคร")
            district_patterns = [
                r'(?:เขต|อำเภอ|อ\.)\s*([ก-๙A-Za-z\s\-]+?)(?:\s*,\s*|\s+|$)',
                r'อ\.\s*([ก-๙A-Za-z\s\-]+?)(?:\s+(?:จ\.|จังหวัด)|$)',
            ]
            for district_pattern in district_patterns:
                district_match = re.search(district_pattern, address_clean, re.IGNORECASE)
                if district_match:
                    district_name = district_match.group(1).strip()
                    # ตรวจสอบว่าไม่ใช่ส่วนของจังหวัด (เช่น "กรุงเทพมหานคร")
                    if district_name and district_name not in [result.get('province', '')]:
                        result['district'] = district_name
                        address_clean = re.sub(rf'{re.escape(district_match.group(0))}', '', address_clean).strip()
                        break
    
    # ถ้ายังไม่พบเขต/อำเภอ ให้ลองหาอีกครั้งจาก address_original (fallback)
    if not result.get('district') and province_found:
        # หาจาก address_original โดยตรง (ก่อนที่จะถูกตัดออก)
        # Pattern: "เขตลาดกระบัง กรุงเทพมหานคร" หรือ "เขต... จังหวัด..."
        district_pattern_fallback = r'(?:เขต|อำเภอ)\s*([ก-๙A-Za-z\s\-]+?)(?:\s+(?:' + re.escape(result['province']) + r'|จังหวัด)|\s*$)'
        district_match_fallback = re.search(district_pattern_fallback, address_original, re.IGNORECASE)
        if district_match_fallback:
            district_name = district_match_fallback.group(1).strip()
            if district_name and district_name not in [result.get('province', '')]:
                result['district'] = district_name
    
    # ถ้ายังไม่พบเขต/อำเภอ ให้ลองหาจาก address_original โดยตรงอีกครั้ง (สำหรับกรณีที่ไม่มีจังหวัดใน pattern)
    if not result.get('district'):
        # หา "เขต..." หรือ "อำเภอ..." หรือ "อ." ที่อยู่ก่อนจังหวัดหรือรหัสไปรษณีย์
        district_pattern_final = r'(?:เขต|อำเภอ|อ\.)\s*([ก-๙A-Za-z\s\-]+?)(?=\s+(?:กรุงเทพ|จังหวัด|\d{5})|\s*$)'
        district_match_final = re.search(district_pattern_final, address_original, re.IGNORECASE)
        if district_match_final:
            district_name = district_match_final.group(1).strip()
            # ตรวจสอบว่าไม่ใช่ส่วนของจังหวัดหรือแขวง/ตำบล
            if district_name and district_name not in [result.get('province', ''), result.get('subdistrict', '')]:
                result['district'] = district_name
    
    # หาแขวง/ตำบล (รองรับทั้งภาษาไทยและภาษาอังกฤษ)
    # Pattern: ชื่อแขวง/ตำบล (มักจะอยู่ก่อนเขต/อำเภอ)
    # ตัวอย่าง: "Klongton-Nua" ก่อน "Vadhana" หรือ "แขวงคลองสามประเวศ" ก่อน "เขตลาดกระบัง"
    if result.get('district'):
        # หาจาก address_original เพื่อหา pattern ที่อยู่ก่อนเขต/อำเภอ
        subdistrict_pattern = r'([ก-๙A-Za-z\s\-]+?)(?:\s*,\s*|\s+)' + re.escape(result['district'])
        subdistrict_match = re.search(subdistrict_pattern, address_original, re.IGNORECASE)
        if subdistrict_match:
            subdistrict_name = subdistrict_match.group(1).strip()
            # ตรวจสอบว่าไม่ใช่ถนน
            if not re.search(r'(?:ถนน|Road|Rd\.)', subdistrict_name, re.IGNORECASE):
                result['subdistrict'] = subdistrict_name
                # ตัดแขวง/ตำบลออกจาก address_clean
                address_clean = re.sub(rf'{re.escape(subdistrict_name)}\s*,\s*', '', address_clean).strip()
        else:
            # ลองหาแบบมีคำว่า "แขวง" หรือ "ตำบล" หรือ "ต."
            subdistrict_patterns = [
                r'(?:แขวง|ตำบล|ต\.)\s*([ก-๙A-Za-z\s\-]+?)(?:\s*,\s*|\s+|$)',
                r'ต\.\s*([ก-๙A-Za-z\s\-]+?)(?:\s+(?:อ\.|เขต|อำเภอ)|$)',
            ]
            for subdistrict_pattern in subdistrict_patterns:
                subdistrict_match = re.search(subdistrict_pattern, address_clean, re.IGNORECASE)
                if subdistrict_match:
                    result['subdistrict'] = subdistrict_match.group(1).strip()
                    address_clean = re.sub(rf'{re.escape(subdistrict_match.group(0))}', '', address_clean).strip()
                    break
    else:
        # ถ้ายังไม่พบเขต/อำเภอ ให้ลองหาแขวง/ตำบลโดยตรง (สำหรับกรณีที่ไม่มีเขต/อำเภอ)
        # Pattern: "แขวง..." หรือ "ตำบล..." หรือ "ต."
        subdistrict_patterns = [
            r'(?:แขวง|ตำบล|ต\.)\s*([ก-๙A-Za-z\s\-]+?)(?:\s*,\s*|\s+|$)',
            r'ต\.\s*([ก-๙A-Za-z\s\-]+?)(?:\s+(?:อ\.|เขต|อำเภอ)|$)',
        ]
        for subdistrict_pattern in subdistrict_patterns:
            subdistrict_match = re.search(subdistrict_pattern, address_clean, re.IGNORECASE)
            if subdistrict_match:
                subdistrict_name = subdistrict_match.group(1).strip()
                # ตรวจสอบว่าไม่ใช่ถนน
                if not re.search(r'(?:ถนน|Road|Rd\.)', subdistrict_name, re.IGNORECASE):
                    result['subdistrict'] = subdistrict_name
                    address_clean = re.sub(rf'{re.escape(subdistrict_match.group(0))}', '', address_clean).strip()
                    break
    
    # ===== Parse จากหน้าไปหลัง (ชื่ออาคาร → เลขที่ → ถนน → ซอย) =====
    
    # หาชื่ออาคาร/Building (เช่น "MSC Building", "อาคาร...", "Building...")
    # ตรวจสอบก่อนเลขที่และถนน
    building_name_patterns = [
        r'^([A-Z][A-Za-z\s]+Building)(?:\s*,\s*|\s+)',  # "MSC Building, " หรือ "MSC Building " (ต้องอยู่ต้น)
        r'^(อาคาร[ก-๙A-Za-z\s]+?)(?:\s*,\s*|\s+)',  # "อาคาร..." (ต้องอยู่ต้น)
        r'^([A-Z][A-Za-z\s]+Tower)(?:\s*,\s*|\s+)',  # "Tower..." (ต้องอยู่ต้น)
        r'^([A-Z][A-Za-z\s]+Plaza)(?:\s*,\s*|\s+)',  # "Plaza..." (ต้องอยู่ต้น)
        r'^([A-Z][A-Za-z\s]+Center)(?:\s*,\s*|\s+)',  # "Center..." (ต้องอยู่ต้น)
    ]
    for pattern in building_name_patterns:
        building_name_match = re.search(pattern, address_clean, re.IGNORECASE)
        if building_name_match:
            result['other_info'] = building_name_match.group(1).strip()
            # ตัดชื่ออาคารออก
            address_clean = re.sub(rf'^{re.escape(building_name_match.group(0))}', '', address_clean).strip()
            # ลบ comma ที่เหลือ
            address_clean = re.sub(r'^,\s*', '', address_clean).strip()
            break
    
    # หาเลขที่ (ตัวเลขที่อยู่ต้นสุด หลังจากตัดชื่ออาคารแล้ว)
    # รองรับรูปแบบ: "เลขที่ 1" หรือ "294/7" หรือ "294" หรือ "571"
    # Pattern 1: หา "เลขที่" ก่อน (สำหรับกรณี "เลขที่ 1 ถ.สุนทรโกษา")
    building_pattern_with_label = r'เลขที่\s+(\d+(?:/\d+)?)(?:\s|,|$)'
    building_match = re.search(building_pattern_with_label, address_clean, re.IGNORECASE)
    if building_match:
        result['building_number'] = building_match.group(1).strip()
        # ตัด "เลขที่ X" ออก
        address_clean = re.sub(rf'{re.escape(building_match.group(0))}', '', address_clean).strip()
        # ลบ comma ที่เหลือ
        address_clean = re.sub(r'^,\s*', '', address_clean).strip()
    else:
        # Pattern 2: หาเลขที่ที่อยู่ต้นสุด (สำหรับกรณี "294/7" หรือ "294")
        building_pattern = r'^(\d+(?:/\d+)?)(?:\s|,|$)'
        building_match = re.search(building_pattern, address_clean)
        if building_match:
            result['building_number'] = building_match.group(1).strip()
            # ตัดเลขที่ออก
            address_clean = re.sub(rf'^{re.escape(building_match.group(1))}\s*', '', address_clean).strip()
            # ลบ comma ที่เหลือ
            address_clean = re.sub(r'^,\s*', '', address_clean).strip()
    
    # หา "หมู่ที่" หรือ "หมู่" (ต้องอยู่หลังเลขที่)
    # Pattern: "หมู่ที่ 1" หรือ "หมู่ 1"
    moo_pattern = r'(?:หมู่ที่|หมู่)\s*(\d+)'
    moo_match = re.search(moo_pattern, address_clean, re.IGNORECASE)
    if moo_match:
        moo_number = moo_match.group(1)
        result['other_info'] = f"หมู่ที่ {moo_number}"
        # ตัดหมู่ที่ออก
        address_clean = re.sub(rf'{re.escape(moo_match.group(0))}\s*', '', address_clean).strip()
    
    # หาถนน (รองรับทั้งภาษาไทยและภาษาอังกฤษ)
    # ใช้ address_clean หลังจากตัดเลขที่แล้ว เพื่อไม่ให้ capture "571"
    # Pattern: "ถนน..." หรือ "ถ." (ย่อของ "ถนน") หรือ "Road" หรือ "Rd." ตามด้วยชื่อถนน
    # ตัวอย่าง: "Sukhumvit 71 Rd." หรือ "ถนนสุขุมวิท" หรือ "ถ.สุนทรโกษา"
    # รองรับทั้ง "Road ..." และ "... Rd." (Rd. อยู่ท้าย)
    # ต้องเริ่มด้วยตัวอักษร (A-Z หรือ ก-๙) ไม่ใช่ตัวเลข
    road_patterns = [
        # Pattern 1: "ถนน..." หรือ "ถ." (ย่อของ "ถนน") หรือ "Road ..." (คำนำหน้าอยู่ต้น)
        r'(?:ถนน|ถ\.|Road)\s+([ก-๙A-Za-z][ก-๙A-Za-z0-9\s\-]*?)(?:\s*,\s*|\s*$)',  
        r'Rd\.\s+([ก-๙A-Za-z][ก-๙A-Za-z0-9\s\-]*?)(?:\s*,\s*|\s*$)',
        # Pattern 2: "... Rd." (Rd. อยู่ท้าย) - เช่น "Sukhumvit 71 Rd."
        # ต้องเริ่มด้วยตัวอักษร ไม่ใช่ตัวเลข
        r'([ก-๙A-Za-z][ก-๙A-Za-z0-9\s\-]*?)\s+Rd\.(?:\s*,\s*|\s*$)',
        # Pattern 3: "... Road" (Road อยู่ท้าย)
        r'([ก-๙A-Za-z][ก-๙A-Za-z0-9\s\-]*?)\s+Road(?:\s*,\s*|\s*$)',
        # Pattern 4: ตามด้วยคำอื่น
        r'(?:ถนน|ถ\.|Road|Rd\.?)\s*([ก-๙A-Za-z][ก-๙A-Za-z0-9\s\-]*?)(?:\s+(?:ซอย|Soi|Soi\.|แขวง|ตำบล|เขต|อำเภอ|จังหวัด|Province))',
    ]
    
    road_found = False
    # ใช้ address_clean หลังจากตัดเลขที่แล้ว เพื่อไม่ให้ capture "571"
    for road_pattern in road_patterns:
        road_match = re.search(road_pattern, address_clean, re.IGNORECASE)
        if road_match:
            road_name = road_match.group(1).strip()
            # ตรวจสอบว่าไม่ใช่ส่วนที่ถูก parse ไปแล้ว (เช่น แขวง/ตำบล, เขต/อำเภอ, ชื่ออาคาร)
            # และไม่ใช่ตัวเลขล้วนๆ
            if road_name and not road_name.isdigit() and road_name not in [result.get('subdistrict', ''), result.get('district', ''), result.get('other_info', '')]:
                # สำหรับ pattern ที่ Rd. อยู่ท้าย ให้รวม Rd. เข้าไปด้วย
                if 'Rd.' in road_match.group(0) and not road_name.endswith('Rd.'):
                    result['road'] = road_name + ' Rd.'
                elif 'Road' in road_match.group(0) and not road_name.endswith('Road'):
                    result['road'] = road_name + ' Road'
                else:
                    # สำหรับภาษาไทย: ถ้าไม่มีคำว่า "ถนน" ให้เพิ่มคำว่า "ถนน" เข้าไป
                    # (รองรับทั้ง "ถ." และ "ถนน")
                    if not road_name.startswith('ถนน'):
                        result['road'] = 'ถนน' + road_name
                    else:
                        result['road'] = road_name
                road_found = True
                # ตัดถนนออกจาก address_clean
                address_clean = re.sub(rf'{re.escape(road_match.group(0))}', '', address_clean).strip()
                # ลบ comma ที่เหลือ
                address_clean = re.sub(r'^,\s*', '', address_clean).strip()
                break
    
    # ถ้ายังหาไม่เจอ ให้ลองหาใน address_original (fallback)
    if not road_found:
        for road_pattern in road_patterns:
            road_match = re.search(road_pattern, address_original, re.IGNORECASE)
            if road_match:
                road_name = road_match.group(1).strip()
                # ตรวจสอบว่าไม่ใช่ตัวเลขล้วนๆ และไม่ใช่ส่วนที่ถูก parse ไปแล้ว
                if road_name and not road_name.isdigit() and road_name not in [result.get('subdistrict', ''), result.get('district', ''), result.get('other_info', ''), result.get('building_number', '')]:
                    # สำหรับ pattern ที่ Rd. อยู่ท้าย ให้รวม Rd. เข้าไปด้วย
                    if 'Rd.' in road_match.group(0) and not road_name.endswith('Rd.'):
                        result['road'] = road_name + ' Rd.'
                    elif 'Road' in road_match.group(0) and not road_name.endswith('Road'):
                        result['road'] = road_name + ' Road'
                    else:
                        # สำหรับภาษาไทย: ถ้าไม่มีคำว่า "ถนน" ให้เพิ่มคำว่า "ถนน" เข้าไป
                        # (รองรับทั้ง "ถ." และ "ถนน")
                        if not road_name.startswith('ถนน'):
                            result['road'] = 'ถนน' + road_name
                        else:
                            result['road'] = road_name
                    # ตัดถนนออกจาก address_clean
                    address_clean = re.sub(rf'{re.escape(road_match.group(0))}', '', address_clean).strip()
                    # ลบ comma ที่เหลือ
                    address_clean = re.sub(r'^,\s*', '', address_clean).strip()
                    break
    
    # หาซอย/ตรอก (มักจะมีคำว่า "ซอย" หรือ "Soi")
    # ถ้ายังมีข้อมูลเหลืออยู่ ให้ลองหาซอย
    soi_pattern = r'(?:ซอย|Soi|Soi\.)\s*([ก-๙A-Za-z0-9\s]+?)(?:\s+(?:ถนน|Road|Rd\.|บ้านเลขที่|เลขที่|แขวง|ตำบล)|\s*$)'
    soi_match = re.search(soi_pattern, address_clean, re.IGNORECASE)
    if soi_match:
        result['soi'] = soi_match.group(1).strip()
        # ตัดซอยออก
        address_clean = re.sub(rf'{re.escape(soi_match.group(0))}', '', address_clean).strip()
    
    # หาบ้านเลขที่ (ถ้ายังไม่มีเลขที่ และมีคำว่า "บ้านเลขที่" หรือ "บ้าน")
    if not result['building_number']:
        house_pattern = r'(?:บ้านเลขที่|บ้าน)\s*(\d+[ก-๙A-Za-z0-9/\-]*)'
        house_match = re.search(house_pattern, address_clean, re.IGNORECASE)
        if house_match:
            result['building_number'] = house_match.group(1).strip()
            # ตัดบ้านเลขที่ออก
            address_clean = re.sub(rf'{re.escape(house_match.group(0))}', '', address_clean).strip()
    
    # ทำความสะอาด address_clean อีกครั้ง (ลบ space หลายตัว)
    address_clean = re.sub(r'\s+', ' ', address_clean).strip()
    
    return result


def write_invoice_data_to_excel(
    data: Dict[str, Any],
    excel_path: Optional[str] = None,
    target_folder: Optional[str] = None,
    create_new: bool = False,
    force_create_new: bool = False
) -> tuple[bool, str]:
    """
    Helper function สำหรับเขียนข้อมูลใบแจ้งหนี้ลง Excel
    
    Args:
        data: Dictionary ที่มีข้อมูลใบแจ้งหนี้
        excel_path: Path ของไฟล์ Excel (ถ้าไม่ระบุจะสร้างใหม่)
        target_folder: โฟลเดอร์ที่ต้องการบันทึก Excel
        create_new: สร้างไฟล์ใหม่หรือไม่ (default: False = append)
        force_create_new: บังคับลบไฟล์เดิมและสร้างใหม่ (default: False)
    
    Returns:
        (success, message)
    """
    try:
        writer = InvoiceExcelWriter(excel_path)
        
        # ถ้าไม่มี excel_path แต่มี target_folder ให้สร้าง path
        if not excel_path and target_folder:
            target_path = Path(target_folder)
            excel_path = str(target_path / "Invoice_Data.xlsx")
            logger.info(f"📂 กำหนด Excel path: {excel_path}")
        
        if create_new or not excel_path or not Path(excel_path).exists():
            # สร้างใหม่
            excel_path = writer.create_new_workbook(target_folder=target_folder)
            logger.info(f"สร้างไฟล์ Excel ใหม่: {excel_path}")
        elif force_create_new and Path(excel_path).exists():
            # ถ้า force_create_new = True ให้ลบไฟล์เดิมและสร้างใหม่
            try:
                Path(excel_path).unlink()
                logger.info(f"ลบไฟล์ Excel เดิม: {excel_path}")
            except Exception as e:
                logger.warning(f"ไม่สามารถลบไฟล์เดิมได้: {e}")
            excel_path = writer.create_new_workbook(target_folder=target_folder)
            logger.info(f"สร้างไฟล์ Excel ใหม่: {excel_path}")
        else:
            # โหลดไฟล์เดิม
            # สำคัญ: ต้องโหลด workbook ใหม่ทุกครั้งเพื่อให้ได้ข้อมูลล่าสุด (รวมถึงลำดับที่ถูกต้อง)
            if not writer.load_existing_workbook(excel_path):
                # ถ้าโหลดไม่ได้ ให้สร้างใหม่
                excel_path = writer.create_new_workbook(target_folder=target_folder)
                logger.info(f"สร้างไฟล์ Excel ใหม่: {excel_path}")
            else:
                logger.info(f"📂 โหลดไฟล์ Excel เดิม: {excel_path} (เพื่อคำนวณลำดับที่ถูกต้อง)")
        
        # ตรวจสอบว่าเป็น MST, CUSTOMS_DEPARTMENT หรือ MYORDER_INTELLIGENCE (มี 2 บรรทัด) หรือไม่
        is_mst_split = (data.get('company') == 'MST' and 
                       data.get('amount_before_vat_line1') is not None and 
                       data.get('amount_before_vat_line2') is not None)
        
        is_customs_split = (data.get('company') == 'CUSTOMS_DEPARTMENT' and 
                           data.get('amount_before_vat_line1') is not None and 
                           data.get('amount_before_vat_line2') is not None)
        
        is_myorder_split = (data.get('company') == 'MYORDER_INTELLIGENCE' and 
                           data.get('amount_before_vat_line1') is not None and 
                           data.get('amount_before_vat_line2') is not None)
        
        is_split = is_mst_split or is_customs_split or is_myorder_split
        
        if is_split:
            # MST หรือ CUSTOMS_DEPARTMENT: สร้าง 2 แถวในชีตเดียวกัน (มีภาษีมูลค่าเพิ่ม) ด้วยลำดับเดียวกันสำหรับแต่ละไฟล์
            company_name = data.get('company', 'UNKNOWN')
            logger.info(f"📋 ตรวจพบ {company_name} ที่มีทั้งมีภาษีและไม่มีภาษี → สร้าง 2 แถวในชีตเดียวกัน")
            
            # คำนวณลำดับปัจจุบัน (ใช้ sheet มีภาษีมูลค่าเพิ่ม)
            vat_sheet = writer.sheets[1]  # Sheet มีภาษีมูลค่าเพิ่ม
            # นับจำนวนแถวข้อมูล (ไม่รวม header)
            # max_row = 1 ถ้ามีแค่ header, max_row = 2 ถ้ามีข้อมูล 1 แถว, max_row = 4 ถ้ามีข้อมูล 2 แถว (2 ไฟล์)
            # ลำดับ = (max_row - 1) / 2 + 1 (เพราะแต่ละไฟล์มี 2 แถว)
            # หรือใช้วิธีง่ายๆ: นับจำนวนลำดับที่มีอยู่แล้ว + 1
            max_row = vat_sheet.max_row or 1  # ถ้าเป็น None ให้ใช้ 1
            data_row_count = max_row - 1  # ลบ header row
            if data_row_count <= 0:
                # ยังไม่มีข้อมูล (มีแค่ header)
                current_row_number = 1
            else:
                # นับจำนวนไฟล์ที่มีอยู่แล้ว (แต่ละไฟล์มี 2 แถว)
                file_count = data_row_count // 2
                current_row_number = file_count + 1
            
            logger.info(f"📊 Sheet มีภาษีมูลค่าเพิ่ม: max_row={max_row}, data_rows={data_row_count}, files={data_row_count // 2}, จะใช้ลำดับ={current_row_number}")
            
            # แถวที่ 1: มีภาษีมูลค่าเพิ่ม (ก่อนภาษี 1,800 + ภาษี 126)
            row1_data = data.copy()
            row1_data['document_type'] = 1  # มีภาษี
            row1_data['amount_before_vat'] = data['amount_before_vat_line1']
            # ใช้ vat_amount_line1 ถ้ามี (สำหรับ MYORDER_INTELLIGENCE) ไม่งั้นใช้ vat_amount เดิม
            row1_data['vat_amount'] = data.get('vat_amount_line1', data.get('vat_amount', 0))
            # ใช้ total_amount_line1 ถ้ามี (สำหรับ MYORDER_INTELLIGENCE) ไม่งั้นคำนวณใหม่
            if 'total_amount_line1' in data:
                row1_data['total_amount'] = data['total_amount_line1']
            elif row1_data['amount_before_vat'] and row1_data['vat_amount'] is not None:
                row1_data['total_amount'] = row1_data['amount_before_vat'] + row1_data['vat_amount']
            else:
                row1_data['total_amount'] = row1_data['amount_before_vat'] or 0
            
            logger.info(f"📝 Row 1: ก่อนภาษี={row1_data['amount_before_vat']}, ภาษี={row1_data['vat_amount']}, รวม={row1_data['total_amount']}")
            
            # แถวที่ 2: ไม่มีภาษีมูลค่าเพิ่ม (ก่อนภาษี 1,500 + ภาษี 0) แต่อยู่ใน Sheet เดียวกัน
            row2_data = data.copy()
            row2_data['document_type'] = 1  # ยังคงอยู่ใน Sheet มีภาษี
            row2_data['amount_before_vat'] = data['amount_before_vat_line2']
            # ใช้ vat_amount_line2 ถ้ามี (สำหรับ MYORDER_INTELLIGENCE) ไม่งั้นใช้ 0
            row2_data['vat_amount'] = data.get('vat_amount_line2', 0)
            # ใช้ total_amount_line2 ถ้ามี (สำหรับ MYORDER_INTELLIGENCE) ไม่งั้นใช้ amount_before_vat_line2
            if 'total_amount_line2' in data:
                row2_data['total_amount'] = data['total_amount_line2']
            else:
                row2_data['total_amount'] = data['amount_before_vat_line2'] or 0
            row2_data['force_vat_sheet'] = True  # บังคับให้อยู่ใน sheet "มีภาษีมูลค่าเพิ่ม" แม้ว่า vat_amount = 0
            
            # สำหรับ CUSTOMS_DEPARTMENT: บรรทัดที่ 2 ไม่ต้องกรอกชื่อบัญชี/โค้ดบัญชี (เว้นว่าง)
            if data.get('company') == 'CUSTOMS_DEPARTMENT':
                row2_data['account_name'] = ''
                row2_data['account_code'] = ''
            
            # สำหรับ MYORDER_INTELLIGENCE: บรรทัดที่ 1 และ 2 ไม่ต้องกรอกชื่อบัญชี/โค้ดบัญชี (เว้นว่าง)
            # แต่ถ้ามีการแก้ไขชื่อบัญชี ให้ใช้ account_name_line1 และ account_name_line2
            if data.get('company') == 'MYORDER_INTELLIGENCE':
                # ใช้ชื่อบัญชีที่แก้ไขแล้วถ้ามี (จากฟอร์มแก้ไข)
                account_name_line1 = data.get('account_name_line1', '')
                account_name_line2 = data.get('account_name_line2', '')
                logger.info(f"📝 MYORDER_INTELLIGENCE - account_name_line1: '{account_name_line1}', account_name_line2: '{account_name_line2}'")
                row1_data['account_name'] = account_name_line1
                row1_data['account_code'] = ''
                row2_data['account_name'] = account_name_line2
                row2_data['account_code'] = ''
                # ตั้งค่า withholding tax percent สำหรับแต่ละบรรทัด
                row1_data['withholding_tax_percent'] = data.get('withholding_tax_percent_line1', 3.0)
                row2_data['withholding_tax_percent'] = data.get('withholding_tax_percent_line2', 1.0)
                logger.info(f"📝 Row 1 account_name: '{row1_data['account_name']}', Row 2 account_name: '{row2_data['account_name']}'")
            
            logger.info(f"📝 Row 2: ก่อนภาษี={row2_data['amount_before_vat']}, ภาษี={row2_data['vat_amount']}, รวม={row2_data['total_amount']}")
            
            # เพิ่มทั้ง 2 แถวด้วยลำดับเดียวกัน (แต่ละไฟล์ใช้ลำดับเดียวกัน)
            success1 = writer.add_row(row1_data, override_row_number=current_row_number)
            logger.info(f"✅ Row 1 added with sequence {current_row_number}: {success1}")
            
            # แถวที่ 2 ใช้ลำดับเดียวกันกับแถวที่ 1
            success2 = writer.add_row(row2_data, override_row_number=current_row_number)
            logger.info(f"✅ Row 2 added with sequence {current_row_number}: {success2}")
            
            if success1 and success2:
                # บันทึกข้อมูลที่อยู่ (ถ้ามี)
                address = data.get('address', '') or data.get('address_full', '')
                if address:
                    logger.info(f"📝 บันทึกข้อมูลที่อยู่: Company={data.get('company_name')}, Address={address[:50]}...")
                    address_data = {
                        'company_name': data.get('company_name', ''),
                        'tax_id': data.get('tax_id', ''),
                        'address': address,
                        'address_full': data.get('address_full', address),
                        'building_number': data.get('building_number', ''),
                        'other_info': data.get('other_info', ''),
                        'soi': data.get('soi', ''),
                        'road': data.get('road', ''),
                        'subdistrict': data.get('subdistrict', ''),
                        'district': data.get('district', ''),
                        'province': data.get('province', ''),
                        'postal_code': data.get('postal_code', '')
                    }
                    address_success = writer.add_address_row(address_data)
                    if address_success:
                        logger.info(f"✅ ผลลัพธ์การบันทึกที่อยู่: สำเร็จ")
                    else:
                        logger.warning(f"⚠️ ผลลัพธ์การบันทึกที่อยู่: ไม่สำเร็จ (อาจมีข้อมูลอยู่แล้วหรือเกิดข้อผิดพลาด)")
                else:
                    logger.info(f"ℹ️ ไม่พบที่อยู่สำหรับบริษัท: {data.get('company_name')} (ไม่จำเป็นต้องบันทึก)")
                
                if writer.save():
                    logger.info(f"💾 บันทึกไฟล์ Excel สำเร็จ: {excel_path}")
                    
                    # สำหรับ MYORDER_INTELLIGENCE: เขียนข้อมูลไปยัง report_myorder.xlsx ด้วย
                    if data.get('company') == 'MYORDER_INTELLIGENCE':
                        logger.info(f"📊 กำลังเขียนข้อมูลไปยัง report_myorder.xlsx...")
                        # ใช้โฟลเดอร์เดียวกับ Invoice_Data.xlsx
                        invoice_excel_path = Path(excel_path)
                        target_folder_for_report = str(invoice_excel_path.parent)
                        report_success, report_message = write_myorder_report_to_excel(
                            data, 
                            target_folder=target_folder_for_report
                        )
                        if report_success:
                            logger.info(f"✅ {report_message}")
                        else:
                            logger.warning(f"⚠️ {report_message}")
                    
                    return True, f"บันทึกข้อมูล {company_name} สำเร็จ (2 แถวในชีตเดียวกัน, ลำดับ {current_row_number}): {excel_path}"
                else:
                    logger.error("❌ เกิดข้อผิดพลาดในการบันทึกไฟล์")
                    return False, "เกิดข้อผิดพลาดในการบันทึกไฟล์"
            else:
                logger.error(f"❌ เกิดข้อผิดพลาดในการเพิ่มข้อมูล: success1={success1}, success2={success2}")
                return False, "เกิดข้อผิดพลาดในการเพิ่มข้อมูล"
        else:
            # บริษัทอื่นๆ: เพิ่ม 1 แถวตามปกติ
            # Log ข้อมูลก่อนส่งไป Excel (สำหรับ debug)
            logger.info(f"📊 ข้อมูลที่จะส่งไป Excel:")
            logger.info(f"   company: {data.get('company')}")
            logger.info(f"   document_type: {data.get('document_type')}")
            logger.info(f"   amount_before_vat: {data.get('amount_before_vat')}")
            logger.info(f"   vat_amount: {data.get('vat_amount')}")
            logger.info(f"   total_amount: {data.get('total_amount')}")
            logger.info(f"   withholding_tax_percent: {data.get('withholding_tax_percent')}")
            
            if writer.add_row(data):
                # บันทึกข้อมูลที่อยู่ (ถ้ามี)
                address = data.get('address', '') or data.get('address_full', '')
                if address:
                    logger.info(f"📝 บันทึกข้อมูลที่อยู่: Company={data.get('company_name')}, Address={address[:50]}...")
                    # ส่งข้อมูลที่แยกแล้วไปด้วย (ถ้ามี) เพื่อให้ระบบใช้ข้อมูลที่แยกแล้ว
                    address_data = {
                        'company_name': data.get('company_name', ''),
                        'tax_id': data.get('tax_id', ''),
                        'address': address,
                        'address_full': data.get('address_full', address),
                        # ข้อมูลที่แยกแล้ว (ถ้ามี)
                        'building_number': data.get('building_number', ''),
                        'other_info': data.get('other_info', ''),
                        'soi': data.get('soi', ''),
                        'road': data.get('road', ''),
                        'subdistrict': data.get('subdistrict', ''),
                        'district': data.get('district', ''),
                        'province': data.get('province', ''),
                        'postal_code': data.get('postal_code', ''),
                    }
                    address_success = writer.add_address_row(address_data)
                    if address_success:
                        logger.info(f"✅ ผลลัพธ์การบันทึกที่อยู่: สำเร็จ")
                    else:
                        logger.warning(f"⚠️ ผลลัพธ์การบันทึกที่อยู่: ไม่สำเร็จ (อาจมีข้อมูลอยู่แล้วหรือเกิดข้อผิดพลาด)")
                else:
                    logger.info(f"ℹ️ ไม่พบที่อยู่สำหรับบริษัท: {data.get('company_name')} (ไม่จำเป็นต้องบันทึก)")
                
                if writer.save():
                    # ตรวจสอบ document_type และ vat_amount อีกครั้งเพื่อแสดงผลลัพธ์
                    document_type = data.get('document_type', 2)
                    if document_type is None:
                        document_type = 2  # Default: ไม่มีภาษี
                    
                    # ตรวจสอบ vat_amount เพื่อกำหนด sheet ที่ถูกต้อง
                    vat_amount = data.get('vat_amount')
                    if vat_amount is None:
                        vat_amount = 0
                    elif isinstance(vat_amount, str):
                        try:
                            vat_amount = float(vat_amount)
                        except (ValueError, TypeError):
                            vat_amount = 0
                    else:
                        vat_amount = float(vat_amount) if vat_amount else 0
                    
                    # แปลง document_type ให้เป็น 1 หรือ 2
                    if document_type == 0:
                        document_type = 2  # ไม่มีภาษีมูลค่าเพิ่ม
                    elif document_type == 3 or document_type == 4:
                        document_type = 1 if vat_amount > 0 else 2
                    elif document_type == 1:
                        # ถ้า document_type = 1 แต่ vat_amount = 0 หรือ null ให้ส่งไปชีต "ไม่มีภาษีมูลค่าเพิ่ม"
                        if vat_amount <= 0:
                            document_type = 2
                    
                    sheet_name = InvoiceExcelWriter.SHEET_NAMES.get(document_type, "Unknown")
                    logger.info(f"✅ [DEBUG Excel] บันทึกข้อมูลสำเร็จ → Sheet: {sheet_name} (vat_amount: {vat_amount})")
                    return True, f"บันทึกข้อมูลลง Excel สำเร็จ (Sheet: {sheet_name}): {excel_path}"
                else:
                    return False, "เกิดข้อผิดพลาดในการบันทึกไฟล์"
            else:
                return False, "เกิดข้อผิดพลาดในการเพิ่มข้อมูล"
    except Exception as e:
        logger.error(f"เกิดข้อผิดพลาด: {e}", exc_info=True)
        return False, f"เกิดข้อผิดพลาด: {e}"


def write_myorder_report_to_excel(
    data: Dict[str, Any],
    report_path: Optional[str] = None,
    target_folder: Optional[str] = None
) -> tuple[bool, str]:
    """
    เขียนข้อมูล MyOrder Intelligence ไปยัง report_myorder.xlsx
    ใน sheet "วางข้อมูล"
    
    Args:
        data: Dictionary ที่มีข้อมูลใบแจ้งหนี้ MyOrder Intelligence
        report_path: Path ของไฟล์ report_myorder.xlsx (ถ้าไม่ระบุจะใช้ค่า default)
        target_folder: โฟลเดอร์ที่ต้องการบันทึก report (ถ้าไม่ระบุจะใช้โฟลเดอร์เดียวกับ Invoice_Data)
    
    Returns:
        (success, message)
    """
    try:
        # หา template จาก temp_excel_report/report_myorder.xlsx
        base_path = Path(__file__).parent
        template_path = base_path / "temp_excel_report" / "report_myorder.xlsx"
        
        # กำหนด path ของไฟล์ report ที่จะบันทึก
        if not report_path:
            # ถ้ามี target_folder ให้ใช้โฟลเดอร์นั้น
            if target_folder:
                target_folder_path = Path(target_folder)
                report_path = str(target_folder_path / "report_myorder.xlsx")
            else:
                # ใช้ path เริ่มต้น: temp_excel_report/report_myorder.xlsx
                report_dir = base_path / "temp_excel_report"
                report_dir.mkdir(exist_ok=True)
                report_path = str(report_dir / "report_myorder.xlsx")
        
        report_path_obj = Path(report_path)
        
        # ถ้า template มีอยู่ ให้คัดลอก template ไปยัง report_path
        if template_path.exists():
            # ถ้าไฟล์ report มีอยู่แล้ว ให้ลบก่อน (เพื่อคัดลอก template ใหม่)
            if report_path_obj.exists():
                try:
                    report_path_obj.unlink()
                    logger.info(f"🗑️ ลบไฟล์ report เดิม: {report_path}")
                except Exception as e:
                    logger.warning(f"⚠️ ไม่สามารถลบไฟล์ report เดิมได้: {e}")
            
            # คัดลอก template
            shutil.copy2(template_path, report_path)
            logger.info(f"📋 คัดลอก template จาก {template_path} ไปยัง {report_path}")
        else:
            logger.warning(f"⚠️ ไม่พบ template: {template_path} - จะสร้างไฟล์ใหม่")
        
        # ตรวจสอบว่าไฟล์ report มีอยู่หรือไม่ (หลังจากคัดลอก template)
        if not report_path_obj.exists():
            logger.warning(f"⚠️ ไม่พบไฟล์ report: {report_path} - จะสร้างไฟล์ใหม่")
            # สร้างไฟล์ใหม่พร้อม sheet "วางข้อมูล"
            wb = Workbook()
            ws = wb.active
            ws.title = "วางข้อมูล"
            
            # เพิ่ม header (เหมือนกับ Invoice_Data.xlsx)
            headers = [
                "ลำดับ",
                "ชื่อบริษัท",
                "เลขประจำตัวผู้เสียภาษี",
                "สาขา",
                "วันที่",
                "ครบกำหนดชำระ",
                "เลขที่เอกสาร",
                "ชื่อบัญชี / โค้ดบัญชี",
                "เปอร์เซ็นต์หัก ณ ที่จ่าย",
                "ยอดก่อนภาษีมูลค่าเพิ่ม",
                "ยอดภาษีมูลค่าเพิ่ม",
                "ยอดหลังบวกภาษีมูลค่าเพิ่ม",
                "อ้างอิง",
                "หมายเหตุ",
                "ชื่อไฟล์ใหม่",
                "ชื่อไฟล์เก่า"
            ]
            ws.append(headers)
            
            # Format header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            wb.save(report_path)
            logger.info(f"✅ สร้างไฟล์ report ใหม่: {report_path}")
        else:
            # โหลดไฟล์ที่มีอยู่ (จาก template ที่คัดลอกมา)
            wb = load_workbook(report_path)
            if "วางข้อมูล" not in wb.sheetnames:
                # สร้าง sheet "วางข้อมูล" ถ้ายังไม่มี
                ws = wb.create_sheet("วางข้อมูล")
                # เพิ่ม header (เหมือนกับ Invoice_Data.xlsx)
                headers = [
                    "ลำดับ",
                    "ชื่อบริษัท",
                    "เลขประจำตัวผู้เสียภาษี",
                    "สาขา",
                    "วันที่",
                    "ครบกำหนดชำระ",
                    "เลขที่เอกสาร",
                    "ชื่อบัญชี / โค้ดบัญชี",
                    "เปอร์เซ็นต์หัก ณ ที่จ่าย",
                    "ยอดก่อนภาษีมูลค่าเพิ่ม",
                    "ยอดภาษีมูลค่าเพิ่ม",
                    "ยอดหลังบวกภาษีมูลค่าเพิ่ม",
                    "อ้างอิง",
                    "หมายเหตุ",
                    "ชื่อไฟล์ใหม่",
                    "ชื่อไฟล์เก่า"
                ]
                ws.append(headers)
                
                # Format header
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                ws = wb["วางข้อมูล"]
        
        # คำนวณลำดับปัจจุบัน (นับจำนวนแถวข้อมูลที่มีอยู่แล้ว)
        max_row = ws.max_row or 1
        data_row_count = max_row - 1  # ลบ header row
        if data_row_count <= 0:
            current_row_number = 1
        else:
            # นับจำนวนไฟล์ที่มีอยู่แล้ว (แต่ละไฟล์มี 2 แถว)
            file_count = data_row_count // 2
            current_row_number = file_count + 1
        
        logger.info(f"📊 Report sheet 'วางข้อมูล': max_row={max_row}, data_rows={data_row_count}, files={data_row_count // 2}, จะใช้ลำดับ={current_row_number}")
        
        # คำนวณยอดรวมสำหรับแต่ละบรรทัด
        amount_before_vat_line1 = data.get('amount_before_vat_line1', 0) or 0
        vat_amount_line1 = data.get('vat_amount_line1', 0) or 0
        total_amount_line1 = amount_before_vat_line1 + vat_amount_line1
        
        amount_before_vat_line2 = data.get('amount_before_vat_line2', 0) or 0
        vat_amount_line2 = data.get('vat_amount_line2', 0) or 0
        total_amount_line2 = amount_before_vat_line2 + vat_amount_line2
        
        # เตรียมข้อมูลชื่อบัญชี/โค้ดบัญชี
        account_name_line1 = data.get('account_name_line1', '') or ''
        account_code_line1 = data.get('account_code_line1', '') or ''
        account_display_line1 = account_name_line1
        if account_code_line1:
            account_display_line1 = f"{account_name_line1} / {account_code_line1}" if account_name_line1 else account_code_line1
        
        account_name_line2 = data.get('account_name_line2', '') or ''
        account_code_line2 = data.get('account_code_line2', '') or ''
        account_display_line2 = account_name_line2
        if account_code_line2:
            account_display_line2 = f"{account_name_line2} / {account_code_line2}" if account_name_line2 else account_code_line2
        
        # เตรียมข้อมูลสำหรับบรรทัดที่ 1 (เหมือนกับ Invoice_Data.xlsx)
        row1_data = [
            current_row_number,  # ลำดับ
            data.get('company_name', ''),
            data.get('tax_id', ''),
            data.get('branch', '') or '',
            data.get('date', ''),
            data.get('due_date', '') or '',  # ครบกำหนดชำระ
            data.get('document_number', ''),
            account_display_line1,  # ชื่อบัญชี / โค้ดบัญชี
            data.get('withholding_tax_percent_line1', 3.0) or 3.0,  # เปอร์เซ็นต์หัก ณ ที่จ่าย 3%
            amount_before_vat_line1,  # ยอดก่อนภาษีบรรทัดที่ 1
            vat_amount_line1,  # ยอดภาษีบรรทัดที่ 1
            total_amount_line1,  # ยอดหลังบวกภาษีมูลค่าเพิ่ม
            data.get('reference', '') or '',  # อ้างอิง
            data.get('remark', '') or '',  # หมายเหตุ
            data.get('new_filename', '') or '',  # ชื่อไฟล์ใหม่
            data.get('old_filename', '') or ''  # ชื่อไฟล์เก่า
        ]
        
        # เตรียมข้อมูลสำหรับบรรทัดที่ 2 (เหมือนกับ Invoice_Data.xlsx)
        row2_data = [
            current_row_number,  # ลำดับเดียวกัน
            data.get('company_name', ''),
            data.get('tax_id', ''),
            data.get('branch', '') or '',
            data.get('date', ''),
            data.get('due_date', '') or '',  # ครบกำหนดชำระ
            data.get('document_number', ''),
            account_display_line2,  # ชื่อบัญชี / โค้ดบัญชี
            data.get('withholding_tax_percent_line2', 1.0) or 1.0,  # เปอร์เซ็นต์หัก ณ ที่จ่าย 1%
            amount_before_vat_line2,  # ยอดก่อนภาษีบรรทัดที่ 2
            vat_amount_line2,  # ยอดภาษีบรรทัดที่ 2
            total_amount_line2,  # ยอดหลังบวกภาษีมูลค่าเพิ่ม
            data.get('reference', '') or '',  # อ้างอิง
            data.get('remark', '') or '',  # หมายเหตุ
            data.get('new_filename', '') or '',  # ชื่อไฟล์ใหม่
            data.get('old_filename', '') or ''  # ชื่อไฟล์เก่า
        ]
        
        # เพิ่มข้อมูลทั้ง 2 แถว
        ws.append(row1_data)
        ws.append(row2_data)
        
        logger.info(f"✅ เพิ่มข้อมูล 2 แถวใน report (ลำดับ {current_row_number})")
        logger.info(f"  บรรทัดที่ 1: ก่อนภาษี={amount_before_vat_line1}, ภาษี={vat_amount_line1}, รวม={total_amount_line1}, หัก ณ ที่จ่าย={row1_data[8]}%")
        logger.info(f"  บรรทัดที่ 2: ก่อนภาษี={amount_before_vat_line2}, ภาษี={vat_amount_line2}, รวม={total_amount_line2}, หัก ณ ที่จ่าย={row2_data[8]}%")
        
        # บันทึกไฟล์
        wb.save(report_path)
        logger.info(f"💾 บันทึกไฟล์ report สำเร็จ: {report_path}")
        
        return True, f"บันทึกข้อมูลลง report_myorder.xlsx สำเร็จ (ลำดับ {current_row_number}): {report_path}"
        
    except Exception as e:
        logger.error(f"เกิดข้อผิดพลาดในการเขียน report: {e}", exc_info=True)
        return False, f"เกิดข้อผิดพลาดในการเขียน report: {e}"


# ===== Usage Example =====
if __name__ == "__main__":
    # ตัวอย่างการใช้งาน
    sample_data = {
        'document_type': 2,  # ไม่มีภาษีมูลค่าเพิ่ม
        'company_name': 'MSC Mediterranean Shipping Company S.A.',
        'tax_id': '0993000003667',
        'date': '03/11/2025',
        'account_name': 'ค่าบริการขนส่ง',
        'account_code': '520310',
        'withholding_tax_percent': 0,  # ไม่มีหัก ณ ที่จ่าย
        'amount_before_vat': 6000.00,
        'vat_amount': 0.00,
        'total_amount': 6000.00,
        'remark': 'EXC-2511-008',
        'new_filename': '2511200301.pdf',
        'old_filename': 'EXC-2511-008_007.pdf'
    }
    
    success, message = write_invoice_data_to_excel(sample_data, create_new=True)
    
    print("=" * 80)
    print("📊 Excel Writer Result")
    print("=" * 80)
    print(f"Success: {success}")
    print(f"Message: {message}")
    print("=" * 80)

