"""
Excel Writer Service
===================
Service สำหรับเขียนข้อมูลลง Excel file

Author: BotV3
Version: 1.0.0
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelWriterService:
    """Service สำหรับเขียนข้อมูลลง Excel"""
    
    def __init__(self, excel_path: Optional[str] = None):
        """
        Initialize Excel Writer Service
        
        Args:
            excel_path: Path ของไฟล์ Excel (ถ้าไม่ระบุจะสร้างใหม่)
        """
        self.excel_path = Path(excel_path) if excel_path else None
        self.workbook = None
        self.worksheet = None
    
    def create_new_workbook(self, filename: str = "MSC_Data.xlsx") -> Path:
        """
        สร้าง Excel workbook ใหม่
        
        Args:
            filename: ชื่อไฟล์ Excel
        
        Returns:
            Path ของไฟล์ที่สร้าง
        """
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "MSC Invoices"
        
        # สร้าง Header
        headers = [
            "ชื่อบริษัท",
            "เลขประจำตัวผู้เสียภาษี",
            "วันที่",
            "ยอดก่อนภาษีมูลค่าเพิ่ม",
            "ยอดรวม",
            "หมายเหตุ",
            "ชื่อไฟล์ใหม่",
            "ชื่อไฟล์เก่า",
            "วันที่บันทึก"
        ]
        
        self.worksheet.append(headers)
        
        # Format header
        self._format_header()
        
        # กำหนด path
        if not self.excel_path:
            output_dir = Path("excel_exports")
            output_dir.mkdir(exist_ok=True)
            self.excel_path = output_dir / filename
        
        return self.excel_path
    
    def load_existing_workbook(self, excel_path: str) -> bool:
        """
        โหลด Excel workbook ที่มีอยู่แล้ว
        
        Args:
            excel_path: Path ของไฟล์ Excel
        
        Returns:
            True ถ้าโหลดสำเร็จ
        """
        try:
            self.excel_path = Path(excel_path)
            if not self.excel_path.exists():
                logger.warning(f"ไม่พบไฟล์ {excel_path} - จะสร้างใหม่")
                return False
            
            self.workbook = load_workbook(self.excel_path)
            self.worksheet = self.workbook.active
            return True
        except Exception as e:
            logger.error(f"เกิดข้อผิดพลาดในการโหลดไฟล์: {e}")
            return False
    
    def _format_header(self):
        """Format header row ให้สวยงาม"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name='TH Sarabun New', size=14, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in self.worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # ตั้งความกว้างคอลัมน์
        column_widths = {
            'A': 40,  # ชื่อบริษัท
            'B': 20,  # เลขประจำตัวผู้เสียภาษี
            'C': 15,  # วันที่
            'D': 20,  # ยอดก่อนภาษี
            'E': 20,  # ยอดรวม
            'F': 20,  # หมายเหตุ
            'G': 25,  # ชื่อไฟล์ใหม่
            'H': 30,  # ชื่อไฟล์เก่า
            'I': 20   # วันที่บันทึก
        }
        
        for col, width in column_widths.items():
            self.worksheet.column_dimensions[col].width = width
        
        # ตั้งความสูงของ header
        self.worksheet.row_dimensions[1].height = 25
    
    def add_row(self, data: Dict[str, Any]) -> bool:
        """
        เพิ่มข้อมูลแถวใหม่
        
        Args:
            data: Dictionary ที่มีข้อมูล:
                - company_name: ชื่อบริษัท
                - tax_id: เลขประจำตัวผู้เสียภาษี
                - date: วันที่ (dd/mm/yyyy)
                - non_taxable_amount: ยอดก่อนภาษี
                - total_amount: ยอดรวม
                - remark: หมายเหตุ
                - new_filename: ชื่อไฟล์ใหม่
                - old_filename: ชื่อไฟล์เก่า
        
        Returns:
            True ถ้าเพิ่มสำเร็จ
        """
        try:
            row_data = [
                data.get('company_name', ''),
                data.get('tax_id', ''),
                data.get('date', ''),
                data.get('non_taxable_amount', 0),
                data.get('total_amount', 0),
                data.get('remark', ''),
                data.get('new_filename', ''),
                data.get('old_filename', ''),
                datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            ]
            
            self.worksheet.append(row_data)
            
            # Format แถวที่เพิ่ม
            row_num = self.worksheet.max_row
            self._format_data_row(row_num)
            
            return True
        except Exception as e:
            logger.error(f"เกิดข้อผิดพลาดในการเพิ่มข้อมูล: {e}")
            return False
    
    def _format_data_row(self, row_num: int):
        """Format data row"""
        data_font = Font(name='TH Sarabun New', size=13)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        number_alignment = Alignment(horizontal='right', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alternate row colors
        if row_num % 2 == 0:
            fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for cell in self.worksheet[row_num]:
            cell.font = data_font
            cell.border = border
            cell.fill = fill
            
            # จัด alignment ตามประเภทข้อมูล
            if cell.column in [4, 5]:  # คอลัมน์ตัวเลข (ยอดเงิน)
                cell.alignment = number_alignment
                # Format เป็นตัวเลข 2 ทศนิยม
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = data_alignment
    
    def save(self) -> bool:
        """
        บันทึกไฟล์ Excel
        
        Returns:
            True ถ้าบันทึกสำเร็จ
        """
        try:
            if not self.workbook:
                logger.error("ไม่มี workbook ให้บันทึก")
                return False
            
            if not self.excel_path:
                logger.error("ไม่ได้ระบุ path ของไฟล์")
                return False
            
            self.workbook.save(self.excel_path)
            logger.info(f"✅ บันทึกไฟล์สำเร็จ: {self.excel_path}")
            return True
        except Exception as e:
            logger.error(f"เกิดข้อผิดพลาดในการบันทึก: {e}")
            return False
    
    def add_multiple_rows(self, data_list: List[Dict[str, Any]]) -> int:
        """
        เพิ่มข้อมูลหลายแถว
        
        Args:
            data_list: List of dictionaries
        
        Returns:
            จำนวนแถวที่เพิ่มสำเร็จ
        """
        success_count = 0
        for data in data_list:
            if self.add_row(data):
                success_count += 1
        return success_count


# ===== Helper Functions =====

def write_msc_data_to_excel(
    data: Dict[str, Any],
    excel_path: Optional[str] = None,
    create_new: bool = False
) -> tuple[bool, str]:
    """
    Helper function สำหรับเขียนข้อมูล MSC ลง Excel
    
    Args:
        data: Dictionary ที่มีข้อมูล MSC
        excel_path: Path ของไฟล์ Excel (ถ้าไม่ระบุจะสร้างใหม่)
        create_new: สร้างไฟล์ใหม่หรือไม่ (default: False = append)
    
    Returns:
        (success, message)
    """
    try:
        writer = ExcelWriterService(excel_path)
        
        if create_new or not excel_path or not Path(excel_path).exists():
            # สร้างใหม่
            excel_path = writer.create_new_workbook()
            logger.info(f"สร้างไฟล์ Excel ใหม่: {excel_path}")
        else:
            # โหลดไฟล์เดิม
            if not writer.load_existing_workbook(excel_path):
                # ถ้าโหลดไม่ได้ ให้สร้างใหม่
                excel_path = writer.create_new_workbook()
                logger.info(f"สร้างไฟล์ Excel ใหม่: {excel_path}")
        
        # เพิ่มข้อมูล
        if writer.add_row(data):
            if writer.save():
                return True, f"บันทึกข้อมูลลง Excel สำเร็จ: {excel_path}"
            else:
                return False, "เกิดข้อผิดพลาดในการบันทึกไฟล์"
        else:
            return False, "เกิดข้อผิดพลาดในการเพิ่มข้อมูล"
    except Exception as e:
        logger.error(f"เกิดข้อผิดพลาด: {e}", exc_info=True)
        return False, f"เกิดข้อผิดพลาด: {e}"


# ===== Usage Example =====
if __name__ == "__main__":
    # ตัวอย่างการใช้งาน
    sample_data = {
        'company_name': 'MSC Mediterranean Shipping Company S.A.',
        'tax_id': '0993000003667',
        'date': '03/11/2025',
        'non_taxable_amount': 6000.00,
        'total_amount': 6000.00,
        'remark': 'EXC-2511-008',
        'new_filename': '2511200301.pdf',
        'old_filename': 'EXC-2511-008_007.pdf'
    }
    
    success, message = write_msc_data_to_excel(sample_data, create_new=True)
    
    print("=" * 80)
    print("📊 Excel Writer Result")
    print("=" * 80)
    print(f"Success: {success}")
    print(f"Message: {message}")
    print("=" * 80)

