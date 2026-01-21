"""
BotV3 - Flask Web Application
Web Application แบบ Flask สำหรับ BotV3
"""

from flask import Flask, render_template, request, jsonify, send_from_directory, session, send_file
from flask_cors import CORS
from pathlib import Path
import sys
import os
import json
import threading
import uuid
from datetime import datetime, timedelta
from typing import Optional, Dict, List, Tuple
import asyncio
import hashlib
import secrets
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
import time
from functools import wraps
import logging
import re

logger = logging.getLogger(__name__)

# เพิ่ม current directory เข้า sys.path
try:
    script_path = Path(__file__)
    if script_path.exists():
        current_dir = script_path.parent.resolve()
    else:
        current_dir = Path(os.path.abspath(__file__)).parent
except Exception:
    try:
        current_dir = Path(os.getcwd())
    except Exception:
        current_dir = Path.cwd()

if current_dir and str(current_dir) not in sys.path:
    sys.path.insert(0, str(current_dir))

# บังคับใช้ Proactor event loop สำหรับ Windows
if sys.platform == 'win32':
    try:
        try:
            # สำหรับ Python 3.10+ ใช้ get_running_loop() หรือ new_event_loop()
            try:
                loop = asyncio.get_running_loop()
                # ถ้ามี loop อยู่แล้วและไม่ใช่ ProactorEventLoop ให้ปิดและสร้างใหม่
                if not isinstance(loop, asyncio.ProactorEventLoop):
                    loop.close()
                    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
                    new_loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(new_loop)
            except RuntimeError:
                # ไม่มี running loop ให้สร้างใหม่
                asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
                asyncio.set_event_loop(asyncio.new_event_loop())
        except RuntimeError:
            asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
            asyncio.set_event_loop(asyncio.new_event_loop())
    except Exception:
        pass

# Import modules
try:
    from config import Config
    from main_system import MainSystemOrchestrator
    from report_manager import get_global_report_manager, set_line_notifications_enabled, get_line_notifications_enabled, line_oa_push
    from email_system.pdf_generator import PDFGenerator
    # Import email_manager จาก email_system
    try:
        from email_system.email_manager import EmailManager, get_global_email_manager, set_global_email_manager
    except ImportError:
        # ถ้ายังไม่มี ให้สร้าง stub
        class EmailManager:
            def __init__(self, *args, **kwargs):
                pass
        def get_global_email_manager():
            return None
        def set_global_email_manager(*args, **kwargs):
            pass
    
    # Import email_service จาก email_system
    from email_system.email_service import EmailService, EmailPattern, get_global_email_service, set_global_email_service
except ImportError as e:
    print(f"⚠️ ไม่สามารถ import โมดูลได้: {e}")
    # สร้าง stub classes
    class Config:
        BASE_FOLDER = "V"
        MAIN_FOLDERS = []
        SKIP_FOLDERS = []
        CUSTOMER_FOLDER = "ลูกค้า"
        AUTOMATION_FOLDER = "ระบบอัตโนมัติ"
        LINE_NOTIFY_ENABLED = True
    
    class MainSystemOrchestrator:
        def __init__(self, *args, **kwargs):
            raise RuntimeError("ไม่สามารถโหลด MainSystemOrchestrator ได้")
    
    def get_global_report_manager():
        return None
    
    def set_line_notifications_enabled(*args, **kwargs):
        pass
    
    def get_line_notifications_enabled():
        return True
    
    class EmailManager:
        def __init__(self, *args, **kwargs):
            pass
    
    def get_global_email_manager():
        return None
    
    def set_global_email_manager(*args, **kwargs):
        pass
    
    class EmailService:
        def __init__(self, *args, **kwargs):
            pass
    
    class EmailPattern:
        def __init__(self, *args, **kwargs):
            pass
    
    def get_global_email_service():
        return None
    
    def set_global_email_service(*args, **kwargs):
        pass

# ฟังก์ชันสำหรับลบไฟล์เก่าใน temp_uploads
def cleanup_temp_files(max_age_minutes: int = 30):
    """
    ลบไฟล์เก่าใน temp_uploads ที่มากกว่า max_age_minutes นาที
    
    Args:
        max_age_minutes: อายุสูงสุดของไฟล์ (นาที) - default 30 นาที
    """
    try:
        temp_dir = Path('temp_uploads')
        if not temp_dir.exists():
            return
        
        deleted_count = 0
        current_time = time.time()
        max_age_seconds = max_age_minutes * 60
        
        for file_path in temp_dir.glob('*'):
            try:
                if file_path.is_file():
                    file_age = current_time - file_path.stat().st_mtime
                    if file_age > max_age_seconds:
                        file_path.unlink()
                        deleted_count += 1
                        age_minutes = file_age / 60
                        logger.info(f"🗑️ [Cleanup] ลบไฟล์เก่า: {file_path.name} (อายุ: {age_minutes:.1f} นาที)")
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถลบไฟล์ได้: {file_path} - {e}")
        
        if deleted_count > 0:
            logger.info(f"✅ [Cleanup] ลบไฟล์เก่า {deleted_count} ไฟล์สำเร็จ")
        elif deleted_count == 0:
            # นับจำนวนไฟล์ที่เหลืออยู่
            remaining_files = len(list(temp_dir.glob('*')))
            if remaining_files > 0:
                logger.debug(f"📊 [Cleanup] ยังมีไฟล์ {remaining_files} ไฟล์ใน temp_uploads (ยังไม่เก่าพอ)")
        
        return deleted_count
    except Exception as e:
        logger.error(f"เกิดข้อผิดพลาดในการลบไฟล์เก่า: {e}", exc_info=True)
        return 0

# สร้าง Flask app
app = Flask(__name__, 
            template_folder='templates',
            static_folder='static')
CORS(app)
app.secret_key = secrets.token_hex(32)  # สำหรับ session

# ตั้งค่า logging สำหรับ tax_form_parser
# ใช้ force=True เพื่อ override การตั้งค่าเดิม (ถ้ามี)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()  # แสดงใน console
    ],
    force=True  # Override การตั้งค่าเดิม
)

# ตั้งค่า logger สำหรับ tax_form_parser ให้แสดง INFO level
tax_form_parser_logger = logging.getLogger('email_system.tax_form_parser')
tax_form_parser_logger.setLevel(logging.INFO)
# ตรวจสอบว่า handler มีอยู่แล้วหรือไม่
if not tax_form_parser_logger.handlers:
    handler = logging.StreamHandler()
    handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
    tax_form_parser_logger.addHandler(handler)
tax_form_parser_logger.propagate = True  # ให้ propagate ไปยัง root logger

# ตั้งค่า logger สำหรับ tax_ocr_processor ให้แสดง INFO level
tax_ocr_processor_logger = logging.getLogger('email_system.tax_ocr_processor')
tax_ocr_processor_logger.setLevel(logging.INFO)
# ตรวจสอบว่า handler มีอยู่แล้วหรือไม่
if not tax_ocr_processor_logger.handlers:
    handler = logging.StreamHandler()
    handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
    tax_ocr_processor_logger.addHandler(handler)
tax_ocr_processor_logger.propagate = True  # ให้ propagate ไปยัง root logger

logger.info("✅ ตั้งค่า logging สำหรับ tax_form_parser และ tax_ocr_processor สำเร็จ")

# Rate limiting configuration (ต้อง define ก่อนใช้งาน)
RATE_LIMIT_WINDOW = 60  # 60 seconds
MAX_REQUESTS_PER_WINDOW = 100  # 100 requests per minute per IP
request_times: Dict[str, List[float]] = {}

# Rate limiting decorator
def rate_limit(max_requests: int = MAX_REQUESTS_PER_WINDOW, window: int = RATE_LIMIT_WINDOW):
    """Rate limiting decorator"""
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            client_ip = request.remote_addr or 'unknown'
            now = time.time()
            
            # Cleanup old entries
            if client_ip in request_times:
                request_times[client_ip] = [t for t in request_times[client_ip] if now - t < window]
            else:
                request_times[client_ip] = []
            
            # ตรวจสอบ rate limit
            if len(request_times[client_ip]) >= max_requests:
                return jsonify({
                    'error': 'เกินจำนวน request ที่อนุญาต กรุณาลองใหม่ในภายหลัง'
                }), 429
            
            # เพิ่ม request time
            request_times[client_ip].append(now)
            
            return f(*args, **kwargs)
        return wrapper
    return decorator

# สำหรับ Flask 2.x+ ที่ไม่มี before_first_request
# ใช้ before_request แทน และเรียก cleanup แค่ครั้งแรก
@app.before_request
def before_request():
    """เรียกใช้ก่อนทุก request - จัดการ cleanup และ rate limiting"""
    # เรียก cleanup แค่ครั้งแรก (ใช้ global variable เพื่อตรวจสอบ)
    if not hasattr(app, '_cleanup_done'):
        try:
            # ลบไฟล์เก่าที่มากกว่า 30 นาที
            cleanup_temp_files(max_age_minutes=30)
            app._cleanup_done = True
            logger.info("✅ เริ่มต้นแอปและลบไฟล์เก่าเรียบร้อย")
        except Exception as e:
            logger.warning(f"⚠️ ไม่สามารถลบไฟล์เก่าได้: {e}")
    
    # Rate limiting สำหรับ API routes (ยกเว้น stats และ jobs GET)
    if request.path.startswith('/api/'):
        if request.path not in ('/api/stats', '/api/jobs') or request.method != 'GET':
            client_ip = request.remote_addr or 'unknown'
            now = time.time()
            
            # Cleanup old entries
            if client_ip in request_times:
                request_times[client_ip] = [t for t in request_times[client_ip] if now - t < RATE_LIMIT_WINDOW]
            else:
                request_times[client_ip] = []
            
            # ตรวจสอบ rate limit
            if len(request_times[client_ip]) >= MAX_REQUESTS_PER_WINDOW:
                return jsonify({
                    'error': 'เกินจำนวน request ที่อนุญาต กรุณาลองใหม่ในภายหลัง'
                }), 429
            
            # เพิ่ม request time
            request_times[client_ip].append(now)

# เพิ่ม after_request เพื่อจัดการ response สำหรับ API routes
@app.after_request
def after_request(response):
    """จัดการ response สำหรับ API routes"""
    if request.path.startswith('/api/'):
        # ตรวจสอบว่า response เป็น HTML error page หรือไม่
        if response.status_code == 404 and 'text/html' in response.content_type:
            print(f"⚠️ [After Request] 404 HTML response detected for {request.path}, converting to JSON")
            response = jsonify({
                'error': 'ไม่พบ endpoint',
                'path': request.path,
                'method': request.method,
                'message': f'Endpoint {request.path} ไม่พบในระบบ'
            })
            response.status_code = 404
            response.headers['Content-Type'] = 'application/json'
        elif request.path.startswith('/api/'):
            # ตรวจสอบว่า API routes return JSON
            if 'application/json' not in response.content_type and response.status_code >= 400:
                print(f"⚠️ [After Request] Non-JSON error response for {request.path}, converting to JSON")
                response = jsonify({
                    'error': 'เกิดข้อผิดพลาด',
                    'path': request.path,
                    'status_code': response.status_code
                })
                response.status_code = response.status_code
                response.headers['Content-Type'] = 'application/json'
    return response

# เพิ่ม error handler เพื่อ return JSON แทน HTML สำหรับ API endpoints
@app.errorhandler(404)
def not_found(error):
    """Return JSON for 404 errors on API routes"""
    try:
        # ใช้ request.path โดยตรง
        path = request.path
        print(f"🔍 [404 Handler] Path: {path}, Method: {request.method}, URL: {request.url}")
        
        if path.startswith('/api/'):
            print(f"✅ [404 Handler] Returning JSON for API route: {path}")
            response = jsonify({
                'error': 'ไม่พบ endpoint',
                'path': path,
                'method': request.method,
                'message': f'Endpoint {path} ไม่พบในระบบ'
            })
            response.headers['Content-Type'] = 'application/json'
            return response, 404
        else:
            print(f"ℹ️ [404 Handler] Non-API route, returning HTML: {path}")
    except Exception as e:
        print(f"❌ [404 Handler] Error: {e}")
        import traceback
        traceback.print_exc()
    # ถ้าไม่ใช่ API route ให้ return error เดิม (HTML)
    return error

@app.errorhandler(500)
def internal_error(error):
    """Return JSON for 500 errors on API routes"""
    try:
        if request.path.startswith('/api/'):
            print(f"🔍 [500 Handler] API route error: {request.path}")
            response = jsonify({'error': 'เกิดข้อผิดพลาดภายในเซิร์ฟเวอร์'})
            response.headers['Content-Type'] = 'application/json'
            return response, 500
    except Exception as e:
        print(f"❌ [500 Handler] Error: {e}")
    return error

@app.errorhandler(403)
def forbidden(error):
    """Return JSON for 403 errors on API routes"""
    try:
        if request.path.startswith('/api/'):
            print(f"🔍 [403 Handler] API route forbidden: {request.path}")
            response = jsonify({'error': 'ไม่มีสิทธิ์เข้าถึง'})
            response.headers['Content-Type'] = 'application/json'
            return response, 403
    except Exception as e:
        print(f"❌ [403 Handler] Error: {e}")
    return error

# Admin configuration (ควรย้ายไป config.py ในอนาคต)
ADMIN_USERNAME = os.getenv('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD_HASH = os.getenv('ADMIN_PASSWORD_HASH', hashlib.sha256('admin123'.encode()).hexdigest())

# Global state
job_store_lock = threading.RLock()  # ใช้ RLock เพื่อป้องกัน deadlock จาก nested locks
job_store: Dict[str, Dict] = {}
folder_locks: Dict[str, str] = {}
JOB_LOG_LIMIT = 200
MAX_JOBS_IN_STORE = 1000  # จำกัดจำนวน jobs ใน store
JOB_CLEANUP_AGE_HOURS = 24  # ลบ jobs ที่เสร็จแล้วเกิน 24 ชั่วโมง
MAX_CONCURRENT_JOBS = 5  # จำกัดจำนวน concurrent jobs
job_executor = ThreadPoolExecutor(max_workers=MAX_CONCURRENT_JOBS, thread_name_prefix="JobWorker")

# ---------- Helper Functions ----------

def _normalize_folder_path(folder_path: str) -> str:
    """Normalize folder path"""
    try:
        folder = Path(folder_path)
        normalized = str(folder.resolve())
        normalized = normalized.replace('\\', '/')
        return normalized.lower()
    except Exception:
        return str(Path(folder_path)).replace('\\', '/').lower()

def _is_folder_locked(folder_path: str) -> Tuple[bool, Optional[str]]:
    """ตรวจสอบว่าโฟลเดอร์กำลังถูกประมวลผลอยู่หรือไม่"""
    normalized_path = _normalize_folder_path(folder_path)
    with job_store_lock:
        if normalized_path in folder_locks:
            job_id = folder_locks[normalized_path]
            job = job_store.get(job_id)
            if job and job.get('status') in ('pending', 'running'):
                return True, job_id
            else:
                folder_locks.pop(normalized_path, None)
        return False, None

def _lock_folder(folder_path: str, job_id: str) -> bool:
    """Lock โฟลเดอร์"""
    normalized_path = _normalize_folder_path(folder_path)
    with job_store_lock:
        if normalized_path in folder_locks:
            existing_job_id = folder_locks[normalized_path]
            existing_job = job_store.get(existing_job_id)
            if existing_job and existing_job.get('status') in ('pending', 'running'):
                return False
            else:
                folder_locks.pop(normalized_path, None)
        folder_locks[normalized_path] = job_id
        return True

def _unlock_folder(folder_path: str, job_id: str):
    """Unlock โฟลเดอร์"""
    normalized_path = _normalize_folder_path(folder_path)
    with job_store_lock:
        if normalized_path in folder_locks and folder_locks[normalized_path] == job_id:
            folder_locks.pop(normalized_path, None)

def _job_add_log(job_id: str, message: str, level: str = "info"):
    """เพิ่ม log ให้ job"""
    try:
        timestamp = datetime.now().strftime("%H:%M:%S")
        with job_store_lock:
            job = job_store.get(job_id)
            if not job:
                return
            job.setdefault('log', []).append({"time": timestamp, "message": message, "level": level})
            if len(job['log']) > JOB_LOG_LIMIT:
                job['log'] = job['log'][-JOB_LOG_LIMIT:]
    except Exception as e:
        print(f"⚠️ [Job Add Log] Error: {e}")

def _job_update_progress(job_id: str, *, total_delta=0, success_delta=0, failure_delta=0, duplicate_delta=0, reset=False, error_details=None):
    """อัพเดต progress ของ job"""
    try:
        with job_store_lock:
            job = job_store.get(job_id)
            if not job:
                return
            if reset or 'progress' not in job:
                job['progress'] = {'total': 0, 'success': 0, 'failed': 0, 'duplicates': 0}
            progress = job['progress']
            if reset:
                progress.update({'total': 0, 'success': 0, 'failed': 0, 'duplicates': 0})
                # รีเซ็ต error details เมื่อ reset
                job['error_details'] = []
            progress['total'] = max(0, progress.get('total', 0) + total_delta)
            progress['success'] = max(0, progress.get('success', 0) + success_delta)
            progress['failed'] = max(0, progress.get('failed', 0) + failure_delta)
            progress['duplicates'] = max(0, progress.get('duplicates', 0) + duplicate_delta)
            
            # เก็บ error details ถ้ามี
            if error_details:
                job.setdefault('error_details', []).append(error_details)
                # จำกัดจำนวน error details (เก็บสูงสุด 100 รายการ)
                if len(job['error_details']) > 100:
                    job['error_details'] = job['error_details'][-100:]
    except Exception as e:
        print(f"⚠️ [Job Update Progress] Error for job {job_id}: {e}")

def _job_update_status(job_id: str, *, folder: Optional[str] = None, file: Optional[str] = None, step: Optional[str] = None):
    """อัพเดต status ของ job"""
    try:
        with job_store_lock:
            job = job_store.get(job_id)
            if not job:
                return
            if folder is not None:
                job['current_folder'] = folder
            if file is not None:
                job['current_file'] = file
            if step is not None:
                job['current_step'] = step
    except Exception as e:
        print(f"⚠️ [Job Update Status] Error for job {job_id}: {e}")

def _job_set_state(job_id: str, status: str):
    """ตั้งค่า status ของ job"""
    try:
        with job_store_lock:
            job = job_store.get(job_id)
            if not job:
                return
            job['status'] = status
            if status in ('success', 'partial_success', 'error'):
                job['end_time'] = datetime.now().isoformat()
                # คำนวณเวลาที่ใช้
                if job.get('start_time'):
                    try:
                        start = datetime.fromisoformat(job['start_time'])
                        end = datetime.fromisoformat(job['end_time'])
                        duration = end - start
                        job['duration_seconds'] = duration.total_seconds()
                        job['duration_formatted'] = str(duration).split('.')[0]  # แสดงเป็น HH:MM:SS
                    except:
                        pass
    except Exception as e:
        print(f"⚠️ [Job Set State] Error for job {job_id}: {e}")

def _get_folder_display_name(folder_path: str) -> str:
    """สร้างชื่อที่แสดงสำหรับโฟลเดอร์"""
    try:
        folder = Path(folder_path)
        folder_name = folder.name
        
        # หา parent folder ที่มี "Build" ในชื่อ หรือ parent ที่เหมาะสม
        current = folder.parent
        parent_name = None
        
        # หา parent folder ที่มี "Build" ในชื่อ
        for _ in range(5):  # ตรวจสอบสูงสุด 5 ระดับ
            if current and current.name:
                if 'build' in current.name.lower() or 'ทดสอบ' in current.name:
                    parent_name = current.name
                    break
            current = current.parent
        
        if parent_name:
            return f"ระบบอัตโนมัติ ({parent_name})"
        else:
            return f"ระบบอัตโนมัติ ({folder_name})"
    except Exception:
        # ถ้าเกิด error ให้ใช้ชื่อโฟลเดอร์โดยตรง
        try:
            return Path(folder_path).name
        except:
            return folder_path

def _check_remaining_pdfs(folder_path: str) -> List[str]:
    """ตรวจสอบไฟล์ PDF ที่ค้างอยู่ในโฟลเดอร์หลัก (ไม่รวมโฟลเดอร์ย่อย)"""
    remaining_pdfs = []
    try:
        folder = Path(folder_path)
        if not folder.exists() or not folder.is_dir():
            return remaining_pdfs
        
        # ตรวจสอบเฉพาะไฟล์ PDF ในโฟลเดอร์หลัก (ไม่รวมโฟลเดอร์ย่อย)
        for item in folder.iterdir():
            if item.is_file() and item.suffix.lower() == '.pdf':
                remaining_pdfs.append(item.name)
        
        # เรียงลำดับชื่อไฟล์
        remaining_pdfs.sort()
    except Exception as e:
        print(f"⚠️ [Check PDFs] Error checking PDFs: {e}")
    
    return remaining_pdfs

def _cleanup_old_jobs():
    """ลบ jobs เก่าที่เสร็จแล้วและเกินอายุที่กำหนด"""
    try:
        now = datetime.now()
        cutoff_time = now - timedelta(hours=JOB_CLEANUP_AGE_HOURS)
        jobs_to_remove = []
        
        with job_store_lock:
            for job_id, job in list(job_store.items()):
                # ลบ jobs ที่เสร็จแล้วและเกินอายุ
                if job.get('status') in ('success', 'error', 'partial_success'):
                    end_time_str = job.get('end_time')
                    if end_time_str:
                        try:
                            end_time = datetime.fromisoformat(end_time_str)
                            if end_time < cutoff_time:
                                jobs_to_remove.append(job_id)
                        except:
                            # ถ้า parse ไม่ได้ ให้ลบถ้า job เก่าเกิน 48 ชั่วโมง
                            start_time_str = job.get('start_time')
                            if start_time_str:
                                try:
                                    start_time = datetime.fromisoformat(start_time_str)
                                    if start_time < cutoff_time:
                                        jobs_to_remove.append(job_id)
                                except:
                                    pass
            
            # ลบ jobs
            for job_id in jobs_to_remove:
                job = job_store.pop(job_id, None)
                if job:
                    folder = job.get('folder', '')
                    if folder:
                        _unlock_folder(folder, job_id)
                    print(f"🧹 [Cleanup] Removed old job {job_id}")
        
        # ถ้ายังมี jobs มากเกินไป ให้ลบ jobs เก่าที่สุด
        if len(job_store) > MAX_JOBS_IN_STORE:
            with job_store_lock:
                # เรียง jobs ตาม end_time หรือ start_time
                sorted_jobs = sorted(
                    job_store.items(),
                    key=lambda x: (
                        x[1].get('end_time') or x[1].get('start_time') or '1970-01-01',
                        x[0]
                    )
                )
                # ลบ jobs เก่าที่สุด
                excess_count = len(job_store) - MAX_JOBS_IN_STORE
                for i in range(excess_count):
                    job_id, job = sorted_jobs[i]
                    job_store.pop(job_id, None)
                    folder = job.get('folder', '')
                    if folder:
                        _unlock_folder(folder, job_id)
                    print(f"🧹 [Cleanup] Removed excess job {job_id}")
    except Exception as e:
        print(f"⚠️ [Cleanup Jobs] Error: {e}")
        import traceback
        traceback.print_exc()

def _get_running_jobs_count() -> int:
    """นับจำนวน jobs ที่กำลังรัน"""
    try:
        with job_store_lock:
            return sum(1 for j in job_store.values() if j.get('status') in ('running', 'pending'))
    except:
        return 0

def _job_worker(job_id: str, folder_path: str):
    """Worker thread สำหรับประมวลผล"""
    with job_store_lock:
        job = job_store.get(job_id)
        if not job:
            return
        job['start_time'] = datetime.now().isoformat()
        job['status'] = 'running'
        # เก็บชื่อที่แสดงไว้ใน job
        job['display_name'] = _get_folder_display_name(folder_path)

    _job_add_log(job_id, f"🚀 เริ่มประมวลผลโฟลเดอร์: {folder_path}", "info")
    _job_update_progress(job_id, reset=True)
    _job_update_status(job_id, folder=folder_path, step="กำลังเตรียมระบบ")

    error_message = None
    remaining_pdfs = []
    error_details = []
    
    try:
        orchestrator = MainSystemOrchestrator(
            str(folder_path),
            progress_callback=lambda **kwargs: _job_update_progress(job_id, **kwargs),
            status_callback=lambda **kwargs: _job_update_status(job_id, **kwargs),
            log_callback=lambda message, level="info": _job_add_log(job_id, message, level)
        )
        
        orchestrator.run_single_folder(folder_path)
        
        # เก็บ error details จาก orchestrator results
        if hasattr(orchestrator, 'results') and orchestrator.results:
            for result in orchestrator.results:
                # เก็บไฟล์ที่อ่านไม่ได้
                if 'pdf_files_read_failed' in result:
                    for failed_file in result['pdf_files_read_failed']:
                        # หา reason จาก pdf_reader.last_batch_report ถ้ามี
                        reason = 'อ่านไม่สำเร็จหรือไม่พบข้อมูลที่ต้องการ'
                        if hasattr(orchestrator, 'pdf_reader') and hasattr(orchestrator.pdf_reader, 'last_batch_report'):
                            skipped_list = orchestrator.pdf_reader.last_batch_report.get('skipped', [])
                            for skipped in skipped_list:
                                if skipped.get('file', '').endswith(failed_file) or failed_file in skipped.get('file', ''):
                                    reason = skipped.get('reason', reason)
                                    break
                        
                        error_details.append({
                            'file': failed_file,
                            'reason': reason,
                            'type': 'read_failed'
                        })
                
                # เก็บไฟล์ที่ skipped จาก pdf_reader
                if hasattr(orchestrator, 'pdf_reader') and hasattr(orchestrator.pdf_reader, 'last_batch_report'):
                    skipped_list = orchestrator.pdf_reader.last_batch_report.get('skipped', [])
                    for skipped in skipped_list:
                        file_path = skipped.get('file', '')
                        file_name = Path(file_path).name if file_path else 'ไม่ทราบชื่อไฟล์'
                        # ตรวจสอบว่าไม่ซ้ำกับไฟล์ที่เก็บไว้แล้ว
                        if not any(err['file'] == file_name or file_name in err.get('file', '') for err in error_details):
                            error_details.append({
                                'file': file_name,
                                'reason': skipped.get('reason', 'ไม่ทราบสาเหตุ'),
                                'type': 'skipped'
                            })
                
                # เก็บ leftover files
                if 'leftover_files' in result:
                    for leftover in result['leftover_files']:
                        leftover_name = Path(leftover).name if leftover else 'ไม่ทราบชื่อไฟล์'
                        error_details.append({
                            'file': leftover_name,
                            'reason': 'ไฟล์ค้างอยู่หลังจากพยายามประมวลผล',
                            'type': 'leftover'
                        })
        
        # เก็บ error details ใน job
        if error_details:
            print(f"📝 [Job Worker] Storing {len(error_details)} error_details for job {job_id}")
            with job_store_lock:
                job = job_store.get(job_id)
                if job:
                    job['error_details'] = error_details
                    print(f"✅ [Job Worker] Stored error_details in job {job_id}")
                else:
                    print(f"⚠️ [Job Worker] Job {job_id} not found in job_store when storing error_details")
        else:
            print(f"ℹ️ [Job Worker] No error_details to store for job {job_id}")
        
        _job_set_state(job_id, 'success')
        _job_add_log(job_id, "✅ งานเสร็จสมบูรณ์", "success")
    except Exception as e:
        error_message = str(e)
        error_type = type(e).__name__
        
        _job_set_state(job_id, 'error')
        _job_add_log(job_id, f"❌ เกิดข้อผิดพลาด: {error_type}: {error_message}", "error")
        
        # ตรวจสอบไฟล์ PDF ที่ค้างอยู่
        remaining_pdfs = _check_remaining_pdfs(folder_path)
        
        if remaining_pdfs:
            pdf_list = ", ".join(remaining_pdfs[:10])  # แสดงสูงสุด 10 ไฟล์
            if len(remaining_pdfs) > 10:
                pdf_list += f" และอีก {len(remaining_pdfs) - 10} ไฟล์"
            
            _job_add_log(job_id, f"📄 พบไฟล์ PDF ค้างอยู่: {len(remaining_pdfs)} ไฟล์", "warning")
            _job_add_log(job_id, f"   ไฟล์: {pdf_list}", "warning")
        
        # เก็บ error information ใน job
        with job_store_lock:
            job = job_store.get(job_id)
            if job:
                job['error_message'] = error_message
                job['error_type'] = error_type
                job['remaining_pdfs'] = remaining_pdfs
                job['remaining_pdf_count'] = len(remaining_pdfs)
    finally:
        try:
            _job_update_status(job_id, step="เสร็จสิ้น", file='-')
            _unlock_folder(folder_path, job_id)
            
            # Cleanup thread reference
            with job_store_lock:
                job = job_store.get(job_id)
                if job:
                    job.pop('future', None)
                    job.pop('thread', None)
        except Exception as e:
            print(f"⚠️ [Job Worker] Error in finally block for job {job_id}: {e}")

# ---------- Routes ----------

@app.route('/')
def index():
    """หน้าแรก"""
    return render_template('index.html')

@app.route('/pdf')
def pdf_page():
    """หน้าประมวลผล PDF"""
    is_admin = session.get('is_admin', False)
    return render_template('pdf_processing.html', is_admin=is_admin)

@app.route('/email')
def email_page():
    """หน้าส่งอีเมลล์"""
    return render_template('email_sending.html')

@app.route('/document-sorting')
def document_sorting_page():
    """หน้าคัดแยกเอกสาร"""
    return render_template('document_sorting.html')

@app.route('/auditcheck')
def auditcheck_page():
    """หน้าตรวจภาษี"""
    return render_template('auditcheck.html')

# ---------- API Routes ----------

@app.route('/api/jobs', methods=['GET'])
def get_jobs():
    """ดึงข้อมูล jobs ทั้งหมด"""
    try:
        # Cleanup old jobs ก่อนดึงข้อมูล
        _cleanup_old_jobs()
        
        jobs = []
        try:
            # ใช้ timeout เพื่อป้องกันการค้าง
            lock_acquired = job_store_lock.acquire(timeout=5.0)
            if not lock_acquired:
                print(f"⚠️ [Get Jobs] Lock timeout")
                return jsonify({'error': 'ไม่สามารถเข้าถึงข้อมูลได้ กรุณาลองใหม่', 'jobs': []}), 503
            
            try:
                for job_id, job in list(job_store.items()):  # ใช้ list() เพื่อป้องกันการเปลี่ยนแปลงระหว่าง iterate
                    try:
                        copy_job = {}
                        for k, v in job.items():
                            if k in ('thread', 'future'):  # ข้าม thread และ future objects
                                continue
                            try:
                                if isinstance(v, dict):
                                    copy_job[k] = v.copy()
                                elif isinstance(v, list):
                                    copy_job[k] = list(v)
                                elif isinstance(v, (str, int, float, bool, type(None))):
                                    copy_job[k] = v
                                else:
                                    # สำหรับ object อื่นๆ ให้แปลงเป็น string
                                    copy_job[k] = str(v)
                            except (AttributeError, TypeError) as e:
                                # ถ้า copy ไม่ได้ ให้ใช้ค่าเดิมหรือแปลงเป็น string
                                try:
                                    copy_job[k] = str(v)
                                except:
                                    copy_job[k] = None
                        
                        if 'progress' not in copy_job:
                            copy_job['progress'] = {'total': 0, 'success': 0, 'failed': 0, 'duplicates': 0}
                        
                        # ตรวจสอบว่ามี id หรือไม่
                        if 'id' not in copy_job:
                            copy_job['id'] = job_id
                        
                        jobs.append(copy_job)
                    except Exception as e:
                        print(f"❌ [Get Jobs] Error copying job {job_id}: {e}")
                        # ข้าม job นี้ไป
                        continue
            finally:
                job_store_lock.release()
        except Exception as lock_error:
            print(f"❌ [Get Jobs] Lock error: {lock_error}")
            import traceback
            traceback.print_exc()
        
        return jsonify({'jobs': jobs}), 200
    except Exception as e:
        print(f"❌ [Get Jobs] Critical error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาดในการดึงข้อมูล jobs: {e}', 'jobs': []}), 500

@app.route('/api/jobs', methods=['POST'])
def create_job():
    """สร้าง job ใหม่"""
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'กรุณาระบุข้อมูล'}), 400
        
        folder_path = data.get('folder_path', '').strip()
        
        if not folder_path:
            return jsonify({'error': 'กรุณาระบุโฟลเดอร์'}), 400
        
        folder = Path(folder_path)
        if not folder.exists():
            return jsonify({'error': f'ไม่พบโฟลเดอร์: {folder_path}'}), 404
        
        # ตรวจสอบจำนวน concurrent jobs
        running_count = _get_running_jobs_count()
        if running_count >= MAX_CONCURRENT_JOBS:
            return jsonify({
                'error': f'มีงานที่กำลังรันอยู่ {running_count} งาน (สูงสุด {MAX_CONCURRENT_JOBS} งาน) กรุณารอให้งานเสร็จก่อน'
            }), 429  # Too Many Requests
        
        is_locked, existing_job_id = _is_folder_locked(folder_path)
        if is_locked:
            return jsonify({'error': f'โฟลเดอร์นี้กำลังถูกประมวลผลอยู่แล้ว (Job ID: {existing_job_id})'}), 400
        
        # Cleanup old jobs ก่อนสร้างใหม่
        _cleanup_old_jobs()
        
        job_id = str(uuid.uuid4())[:8]
        job_data = {
            'id': job_id,
            'folder': str(folder),
            'status': 'pending',
            'progress': {'total': 0, 'success': 0, 'failed': 0, 'duplicates': 0},
            'current_folder': str(folder),
            'current_file': '-',
            'current_step': 'รอเริ่มงาน',
            'log': [],
            'start_time': None,
            'end_time': None,
            'display_name': _get_folder_display_name(str(folder)),  # เพิ่ม display_name ตั้งแต่เริ่มต้น
            'error_details': []  # เพิ่ม error_details สำหรับเก็บรายการไฟล์ที่ผิดพลาด
        }
        
        if not _lock_folder(folder_path, job_id):
            return jsonify({'error': 'ไม่สามารถ lock โฟลเดอร์ได้'}), 400
        
        with job_store_lock:
            job_store[job_id] = job_data
        
        # ใช้ ThreadPoolExecutor แทนการสร้าง thread ใหม่
        future = job_executor.submit(_job_worker, job_id, folder_path)
        
        with job_store_lock:
            if job_id in job_store:
                job_store[job_id]['future'] = future  # เก็บ future แทน thread object
        
        return jsonify({'job_id': job_id, 'message': 'เริ่มงานเรียบร้อย'}), 201
    except Exception as e:
        print(f"❌ [Create Job] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/jobs/<job_id>', methods=['DELETE'])
def delete_job(job_id: str):
    """ลบ job"""
    try:
        with job_store_lock:
            if job_id in job_store:
                job = job_store[job_id]
                folder = job.get('folder', '')
                
                # ลบ future ถ้ามี
                if 'future' in job:
                    future = job.get('future')
                    if future and not future.done():
                        future.cancel()  # ยกเลิก job ที่ยังไม่เสร็จ
                
                job_store.pop(job_id, None)
                if folder:
                    _unlock_folder(folder, job_id)
                return jsonify({'message': 'ลบงานสำเร็จ'}), 200
            return jsonify({'error': 'ไม่พบ job'}), 404
    except Exception as e:
        print(f"❌ [Delete Job] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/jobs/<job_id>', methods=['GET'])
def get_job_detail(job_id: str):
    """ดึงข้อมูล job แบบละเอียด"""
    with job_store_lock:
        job = job_store.get(job_id)
        if not job:
            return jsonify({'error': 'ไม่พบ job'}), 404
        
        copy_job = {
            k: (v.copy() if isinstance(v, dict) else list(v) if isinstance(v, list) else v)
            for k, v in job.items()
            if k != 'thread'
        }
        
        # คำนวณเวลาที่ใช้ถ้ายังไม่มี
        if copy_job.get('start_time') and copy_job.get('end_time'):
            try:
                start = datetime.fromisoformat(copy_job['start_time'])
                end = datetime.fromisoformat(copy_job['end_time'])
                duration = end - start
                copy_job['duration_seconds'] = duration.total_seconds()
                copy_job['duration_formatted'] = str(duration).split('.')[0]
            except:
                pass
        elif copy_job.get('start_time') and copy_job.get('status') == 'running':
            try:
                start = datetime.fromisoformat(copy_job['start_time'])
                now = datetime.now()
                duration = now - start
                copy_job['duration_seconds'] = duration.total_seconds()
                copy_job['duration_formatted'] = str(duration).split('.')[0]
            except:
                pass
        
        return jsonify({'job': copy_job}), 200

# ---------- Admin Routes ----------

@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    """Login สำหรับ admin"""
    data = request.json
    username = data.get('username', '')
    password = data.get('password', '')
    
    password_hash = hashlib.sha256(password.encode()).hexdigest()
    
    if username == ADMIN_USERNAME and password_hash == ADMIN_PASSWORD_HASH:
        session['is_admin'] = True
        session['admin_username'] = username
        return jsonify({'success': True, 'message': 'เข้าสู่ระบบแอดมินสำเร็จ'}), 200
    else:
        return jsonify({'success': False, 'message': 'ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง'}), 401

@app.route('/api/admin/logout', methods=['POST'])
def admin_logout():
    """Logout สำหรับ admin"""
    session.pop('is_admin', None)
    session.pop('admin_username', None)
    return jsonify({'success': True, 'message': 'ออกจากระบบสำเร็จ'}), 200

@app.route('/api/admin/status', methods=['GET'])
def admin_status():
    """ตรวจสอบสถานะ admin"""
    is_admin = session.get('is_admin', False)
    return jsonify({'is_admin': is_admin, 'username': session.get('admin_username')}), 200

@app.route('/api/admin/jobs/clear', methods=['POST'])
def admin_clear_jobs():
    """ลบ jobs ทั้งหมด (admin only)"""
    if not session.get('is_admin', False):
        return jsonify({'error': 'ไม่มีสิทธิ์เข้าถึง'}), 403
    
    try:
        with job_store_lock:
            # Cancel all running jobs
            for job_id, job in list(job_store.items()):
                if 'future' in job:
                    future = job.get('future')
                    if future and not future.done():
                        future.cancel()
            
            count = len(job_store)
            job_store.clear()
            folder_locks.clear()
            return jsonify({'message': f'ลบ jobs ทั้งหมด {count} รายการสำเร็จ'}), 200
    except Exception as e:
        print(f"❌ [Clear Jobs] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/stats', methods=['GET'])
def stats():
    """สถิติระบบ (สำหรับผู้ใช้ทั่วไป)"""
    print(f"✅ [Stats] Route called: {request.path}")
    with job_store_lock:
        total_jobs = len(job_store)
        running_jobs = sum(1 for j in job_store.values() if j.get('status') == 'running')
        pending_jobs = sum(1 for j in job_store.values() if j.get('status') == 'pending')
        success_jobs = sum(1 for j in job_store.values() if j.get('status') == 'success')
        error_jobs = sum(1 for j in job_store.values() if j.get('status') == 'error')
        
        # สถิติไฟล์ทั้งหมด
        total_files = sum(j.get('progress', {}).get('total', 0) for j in job_store.values())
        total_success = sum(j.get('progress', {}).get('success', 0) for j in job_store.values())
        total_failed = sum(j.get('progress', {}).get('failed', 0) for j in job_store.values())
        total_duplicates = sum(j.get('progress', {}).get('duplicates', 0) for j in job_store.values())
        
        return jsonify({
            'jobs': {
                'total': total_jobs,
                'running': running_jobs,
                'pending': pending_jobs,
                'success': success_jobs,
                'error': error_jobs
            },
            'files': {
                'total': total_files,
                'success': total_success,
                'failed': total_failed,
                'duplicates': total_duplicates
            }
        }), 200

@app.route('/api/admin/stats', methods=['GET'])
def admin_stats():
    """สถิติระบบ (admin only) - ใช้ endpoint เดียวกับ /api/stats"""
    return stats()

@app.route('/api/admin/line-notify/status', methods=['GET'])
def line_notify_status():
    """ตรวจสอบสถานะ LINE Notify (admin only)"""
    if not session.get('is_admin', False):
        return jsonify({'error': 'ไม่มีสิทธิ์เข้าถึง'}), 403
    
    try:
        # ตรวจสอบว่า function ถูก import มาแล้วหรือไม่
        if 'get_line_notifications_enabled' not in globals():
            return jsonify({'error': 'Function get_line_notifications_enabled ไม่พบ'}), 500
        enabled = get_line_notifications_enabled()
        print(f"🔍 [LINE Notify Status] สถานะปัจจุบัน: {enabled}")
        return jsonify({'enabled': enabled}), 200
    except NameError as e:
        print(f"❌ [LINE Notify Status] NameError: {e}")
        return jsonify({'error': f'Function ไม่พบ: {e}'}), 500
    except Exception as e:
        print(f"❌ [LINE Notify Status] Exception: {e}")
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/admin/line-notify/toggle', methods=['POST'])
def line_notify_toggle():
    """เปิด/ปิด LINE Notify (admin only)"""
    if not session.get('is_admin', False):
        return jsonify({'error': 'ไม่มีสิทธิ์เข้าถึง'}), 403
    
    try:
        # ตรวจสอบว่า functions ถูก import มาแล้วหรือไม่
        if 'get_line_notifications_enabled' not in globals() or 'set_line_notifications_enabled' not in globals():
            return jsonify({'error': 'Functions สำหรับ LINE notify ไม่พบ'}), 500
        
        # อ่านสถานะปัจจุบัน
        current_status = get_line_notifications_enabled()
        print(f"🔍 [LINE Notify Toggle] สถานะปัจจุบัน: {current_status}")
        
        # เปลี่ยนสถานะ
        new_status = not current_status
        set_line_notifications_enabled(new_status)
        
        # ตรวจสอบว่าตั้งค่าสำเร็จหรือไม่
        verify_status = get_line_notifications_enabled()
        print(f"🔍 [LINE Notify Toggle] สถานะใหม่: {verify_status}")
        
        if verify_status != new_status:
            return jsonify({
                'error': f'ไม่สามารถตั้งค่าสถานะได้ (ตั้งค่า: {new_status}, ตรวจสอบ: {verify_status})'
            }), 500
        
        status_text = 'เปิด' if new_status else 'ปิด'
        status_icon = '🔔' if new_status else '🔕'
        return jsonify({
            'success': True,
            'message': f'{status_icon} {status_text}การแจ้งเตือน LINE สำเร็จ',
            'enabled': new_status
        }), 200
    except NameError as e:
        print(f"❌ [LINE Notify Toggle] NameError: {e}")
        return jsonify({'error': f'Function ไม่พบ: {e}'}), 500
    except Exception as e:
        print(f"❌ [LINE Notify Toggle] Exception: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/config', methods=['GET'])
def get_email_config():
    """ดึงข้อมูลการตั้งค่า SMTP ที่บันทึกไว้"""
    try:
        email_service = get_global_email_service()
        
        if not email_service:
            return jsonify({
                'success': False, 
                'message': 'ยังไม่ได้ตั้งค่า SMTP',
                'config': None
            }), 200
        
        # ดึงข้อมูล config จาก email_service (ไม่ส่ง password กลับไป - เพื่อความปลอดภัย)
        # หมายเหตุ: รหัสผ่านจะไม่ถูกส่งกลับไปให้ frontend และต้องกรอกใหม่ทุกครั้ง
        config = {
            'smtp_server': email_service.smtp_server,
            'smtp_port': email_service.smtp_port,
            'smtp_use_tls': email_service.smtp_use_tls,
            'smtp_username': email_service.smtp_username,
            'email_from': email_service.email_from,
            'is_configured': email_service.is_configured
            # ไม่ส่ง smtp_password เพื่อความปลอดภัย - ต้องกรอกใหม่ทุกครั้ง
        }
        
        return jsonify({
            'success': True,
            'config': config
        }), 200
    except Exception as e:
        return jsonify({
            'success': False, 
            'message': f'เกิดข้อผิดพลาด: {e}',
            'config': None
        }), 500

@app.route('/api/email/config', methods=['POST'])
def save_email_config():
    """บันทึกการตั้งค่า SMTP"""
    data = request.json
    try:
        from pathlib import Path
        # ใช้ path จาก config หรือใช้ path ใน email_system folder
        if hasattr(Config, 'EMAIL_PATTERNS_FILE') and Config.EMAIL_PATTERNS_FILE:
            patterns_file = Path(Config.EMAIL_PATTERNS_FILE)
        else:
            # ถ้าไม่มี config ให้ใช้ path ใน email_system folder
            email_system_dir = Path('email_system')
            patterns_file = email_system_dir / "email_patterns.json"
        
        email_service = EmailService(
            smtp_server=data.get('smtp_server', ''),
            smtp_port=int(data.get('smtp_port', 587)),
            smtp_use_tls=data.get('smtp_use_tls', True),
            smtp_username=data.get('smtp_username', ''),
            smtp_password=data.get('smtp_password', ''),
            email_from=data.get('email_from', '') or data.get('smtp_username', ''),
            patterns_file=patterns_file
        )
        
        # ตรวจสอบว่า email_service มี method test_connection หรือไม่
        if not hasattr(email_service, 'test_connection'):
            return jsonify({
                'success': False, 
                'message': 'EmailService object ไม่มี method test_connection (อาจเป็นปัญหา import หรือ version ไม่ตรงกัน)'
            }), 500
        
        success, message = email_service.test_connection()
        if success:
            set_global_email_service(email_service)
            # ยังคงเก็บ EmailManager สำหรับ backward compatibility
            email_manager = EmailManager(
                smtp_server=data.get('smtp_server', ''),
                smtp_port=int(data.get('smtp_port', 587)),
                smtp_use_tls=data.get('smtp_use_tls', True),
                smtp_username=data.get('smtp_username', ''),
                smtp_password=data.get('smtp_password', ''),
                email_from=data.get('email_from', '') or data.get('smtp_username', '')
            )
            set_global_email_manager(email_manager)
            return jsonify({'success': True, 'message': 'บันทึกการตั้งค่าและทดสอบการเชื่อมต่อสำเร็จ'}), 200
        else:
            return jsonify({'success': False, 'message': f'ทดสอบการเชื่อมต่อล้มเหลว: {message}'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/upload', methods=['POST', 'OPTIONS'])
def upload_email_attachment():
    """อัปโหลดไฟล์สำหรับแนบอีเมลล์ (รองรับการแยกประเภทไฟล์และ OCR)"""
    # Handle CORS preflight
    if request.method == 'OPTIONS':
        response = jsonify({'success': True})
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
        response.headers.add('Access-Control-Allow-Methods', 'POST, OPTIONS')
        return response
    
    try:
        print(f"📎 [Email Upload] Received request: method={request.method}, content_type={request.content_type}")
        print(f"📎 [Email Upload] Files in request: {list(request.files.keys())}")
        
        if 'file' not in request.files:
            print("❌ [Email Upload] No 'file' in request.files")
            return jsonify({'success': False, 'message': 'ไม่มีไฟล์ที่อัปโหลด'}), 400
        
        file = request.files['file']
        if file.filename == '':
            print("❌ [Email Upload] Empty filename")
            return jsonify({'success': False, 'message': 'ไม่ได้เลือกไฟล์'}), 400
        
        # ตรวจสอบประเภทไฟล์
        file_type = request.form.get('file_type', 'tax_attachment')  # default เป็น tax_attachment
        
        print(f"✅ [Email Upload] Processing file: {file.filename}, type: {file_type}")
        
        # สร้างโฟลเดอร์ temp_uploads ถ้ายังไม่มี
        temp_dir = Path('temp_uploads')
        temp_dir.mkdir(exist_ok=True)
        
        # บันทึกไฟล์ชั่วคราว
        import uuid
        file_id = str(uuid.uuid4())
        file_extension = Path(file.filename).suffix
        temp_filename = f"{file_id}{file_extension}"
        temp_file_path = temp_dir / temp_filename
        
        # เก็บชื่อไฟล์เดิม
        original_filename = file.filename
        
        file.save(temp_file_path)
        print(f"✅ [Email Upload] File saved to: {temp_file_path}, original name: {original_filename}")
        
        # ฟังก์ชันสำหรับแสดงผล OCR ใน terminal แบบสวยงาม
        def print_ocr_result(ocr_result, filename, raw_data=None):
            """แสดงผล OCR result ใน terminal แบบสวยงามและอ่านง่าย"""
            if not ocr_result:
                return
            
            print("\n" + "="*70)
            print(f"📋 [OCR] ผลลัพธ์ OCR - {filename}")
            print("="*70)
            
            # แสดงข้อมูลดิบก่อน (ถ้ามี)
            if raw_data:
                print("\n📄 [ข้อมูลดิบ (Raw Data)]:")
                print("-" * 70)
                import json
                try:
                    # แสดงข้อมูลดิบในรูปแบบ JSON ที่อ่านง่าย
                    # แสดง raw_text เต็ม (ทำความสะอาด HTML tags ก่อน)
                    raw_text = raw_data.get('raw_text', '')
                    if raw_text:
                        # ทำความสะอาด HTML tags (ถ้ามี)
                        import re
                        # ลบ HTML tags ทั้งหมด
                        cleaned_text = re.sub(r'<[^>]+>', '', raw_text)
                        # แปลง HTML entities
                        html_entities = {
                            '&nbsp;': ' ',
                            '&amp;': '&',
                            '&lt;': '<',
                            '&gt;': '>',
                            '&quot;': '"',
                            '&#39;': "'",
                        }
                        for entity, char in html_entities.items():
                            cleaned_text = cleaned_text.replace(entity, char)
                        # ลบช่องว่างเกิน
                        cleaned_text = re.sub(r' {3,}', ' ', cleaned_text)
                        # ลบ newline เกิน
                        cleaned_text = re.sub(r'\n{3,}', '\n\n', cleaned_text)
                        cleaned_text = cleaned_text.strip()
                        
                        text_length = len(raw_text)
                        cleaned_length = len(cleaned_text)
                        print(f"📝 Raw Text Length: {text_length} characters (after cleaning: {cleaned_length} characters)")
                        
                        # แสดง raw_text ที่ทำความสะอาดแล้ว
                        print(f"\n📄 Raw Text Content:")
                        print("-" * 70)
                        print(cleaned_text)
                        print("-" * 70)
                    
                    # แสดงข้อมูลอื่นๆ (ไม่รวม raw_text เพื่อไม่ให้ซ้ำ)
                    other_data = {k: v for k, v in raw_data.items() if k != 'raw_text'}
                    if other_data:
                        print(f"\n📊 Other Data:")
                        print("-" * 70)
                        raw_json = json.dumps(other_data, ensure_ascii=False, indent=2)
                        print(raw_json)
                        print("-" * 70)
                    
                    # แสดงข้อมูลดิบของเลขประจำตัวผู้เสียภาษีและชื่อบริษัทจาก raw_text (ถ้ามี)
                    # ใช้ raw_text ที่ดึงมาแล้วข้างบน
                    if raw_text:
                        import re
                        # ค้นหาเลขประจำตัวผู้เสียภาษีใน raw_text
                        tax_id_patterns = [
                            r'เลขประจำตัวผู้เสียภาษีอากร\s+([0-9\s\-]{13,30})',
                            r'เลขประจำตัวผู้เสียภาษีอากร[^\n]*\n([^\n]+)',
                            r'\(ของผู้มีหน้าที่หัก\s*ภาษี\s*ณ\s*ที่จ่าย\)[^\n]*\n([^\n]+)',
                            r'เลขประจำตัวผู้เสียภาษี[^\n]*\n([^\n]+)'
                        ]
                        for pattern in tax_id_patterns:
                            match = re.search(pattern, raw_text, re.IGNORECASE)
                            if match:
                                tax_id_raw = match.group(1).strip()
                                if any(c.isdigit() for c in tax_id_raw):
                                    print(f"🔍 [ข้อมูลดิบ] เลขประจำตัวผู้เสียภาษีที่อ่านได้: {tax_id_raw}")
                                    break
                        
                        # แสดงข้อมูลดิบของชื่อบริษัทจาก raw_text (ถ้ามี)
                        company_patterns = [
                            r'ชื่อผู้นำส่งภาษี[^\n]*?สาขาที่\s*[0-9\s]*\s*(บริษัท[^\n]+?)(?:\s+ที่ตั้ง|\s+อาคาร|\s+ห้อง|$)',
                            r'ชื่อผู้นำส่งภาษี[^\n]*?(บริษัท[^\n]+?)(?:\s+ที่ตั้ง|\s+อาคาร|\s+ห้อง|$)',
                            r'(บริษัท\s+[^\s]+(?:\s+[^\s]+)*?\s+จำกัด(?:\s+มหาชน)?)'
                        ]
                        for pattern in company_patterns:
                            match = re.search(pattern, raw_text, re.IGNORECASE)
                            if match:
                                company_raw = match.group(1).strip()
                                if 'บริษัท' in company_raw or 'ห้างหุ้นส่วน' in company_raw:
                                    print(f"🔍 [ข้อมูลดิบ] ชื่อบริษัทที่อ่านได้: {company_raw}")
                                    break
                except Exception:
                    print(str(raw_data))
                print("-" * 70)
                print()
            
            if ocr_result.get('success', False):
                # แสดงแบบภาษี
                tax_form_type = ocr_result.get('tax_form_type', '')
                if tax_form_type:
                    print(f"📄 แบบภาษี: {tax_form_type}")
                
                # สำหรับแบบภาษี ภ.ง.ด.1, ภ.ง.ด.3, ภ.ง.ด.53, ภ.ง.ด.54 - ต้องมีข้อมูลหลัก
                is_main_tax_form = tax_form_type and any(form in tax_form_type for form in ['ภ.ง.ด.1', 'ภงด.1', 'ภ.ง.ด.3', 'ภงด.3', 'ภ.ง.ด.53', 'ภงด.53', 'ภ.ง.ด.54', 'ภงด.54'])
                
                if is_main_tax_form:
                    print("\n" + "="*70)
                    print("📋 ข้อมูลหลักสำหรับการตรวจสอบ (Required Data):")
                    print("="*70)
                
                # แสดงข้อมูลบริษัท
                company_name = ocr_result.get('company_name', '')
                if company_name:
                    print(f"🏢 ชื่อบริษัท: {company_name}")
                elif is_main_tax_form:
                    print("⚠️  ชื่อบริษัท: ไม่พบข้อมูล (จำเป็นสำหรับการตรวจสอบ)")
                
                # แสดงเลขประจำตัวผู้เสียภาษี
                tax_id = ocr_result.get('tax_id', '')
                if tax_id:
                    # สำหรับ ภ.พ.36 และ ภ.ง.ด.54: แสดงค่าตามที่อ่านได้จาก JSON (ไม่ format)
                    # สำหรับแบบภาษีอื่นๆ: แปลงเลขที่นิติบุคคลเป็นรูปแบบที่มี dash
                    tax_form_type = ocr_result.get('tax_form_type', '')
                    is_pp36 = isinstance(tax_form_type, str) and ('ภ.พ.36' in tax_form_type or 'ภพ.36' in tax_form_type)
                    is_ppnd54 = isinstance(tax_form_type, str) and ('ภ.ง.ด.54' in tax_form_type or 'ภงด.54' in tax_form_type)
                    is_kys = isinstance(tax_form_type, str) and ('กยศ.' in tax_form_type or 'กองทุน' in tax_form_type)
                    is_payin = isinstance(tax_form_type, str) and ('Pay-in' in tax_form_type or 'payin' in tax_form_type.lower())
                    
                    if is_pp36 or is_ppnd54 or is_kys or is_payin:
                        # สำหรับ ภ.พ.36, ภ.ง.ด.54, กยศ. และ Pay-in: แสดงค่าตามที่อ่านได้ (ไม่ format)
                        print(f"🆔 เลขประจำตัวผู้เสียภาษีอากร: {tax_id}")
                    else:
                        # สำหรับแบบภาษีอื่นๆ: แปลงเลขที่นิติบุคคลเป็นรูปแบบที่มี dash
                        tax_id_clean = tax_id.replace(' ', '')
                        if len(tax_id_clean) == 13:
                            tax_id_formatted = f"{tax_id_clean[:5]}-{tax_id_clean[5:10]}-{tax_id_clean[10:12]}-{tax_id_clean[12:]}"
                            print(f"🆔 เลขประจำตัวผู้เสียภาษีอากร: {tax_id_formatted}")
                            print(f"   (รูปแบบดิบ: {tax_id_clean})")
                        else:
                            print(f"🆔 เลขประจำตัวผู้เสียภาษีอากร: {tax_id_clean}")
                elif is_main_tax_form:
                    print("⚠️  เลขประจำตัวผู้เสียภาษีอากร: ไม่พบข้อมูล (จำเป็นสำหรับการตรวจสอบ)")
                
                if is_main_tax_form:
                    print("="*70)
                    print()
                
                # แสดงข้อมูลการยื่น
                # สำหรับ ภ.พ.36 และ ภ.ง.ด.54: ดึงประเภทการยื่นจาก JSON structure ใหม่ (ถ้ามี)
                filing_type = ocr_result.get('filing_type', '')
                tax_form_type = ocr_result.get('tax_form_type', '')
                is_pp36 = isinstance(tax_form_type, str) and ('ภ.พ.36' in tax_form_type or 'ภพ.36' in tax_form_type)
                is_ppnd54 = isinstance(tax_form_type, str) and ('ภ.ง.ด.54' in tax_form_type or 'ภงด.54' in tax_form_type)
                
                if is_pp36 and ocr_result.get('pp36_json'):
                    # ดึงประเภทการยื่นจาก JSON structure ใหม่ (ภ.พ.36)
                    pp36_json = ocr_result.get('pp36_json', {})
                    data_obj = pp36_json.get('data', {})
                    
                    filing_type_normal = data_obj.get('(1) ยื่นปกติ', '')
                    filing_type_supplement = data_obj.get('(2) ยื่นเพิ่มเติมครั้งที่', '')
                    
                    if filing_type_normal and filing_type_normal is not None and filing_type_normal != 'null' and str(filing_type_normal).strip():
                        filing_type_display = str(filing_type_normal).strip()  # เช่น "ยื่นปกติ"
                    elif filing_type_supplement and filing_type_supplement is not None and filing_type_supplement != 'null' and str(filing_type_supplement).strip():
                        filing_type_display = str(filing_type_supplement).strip()  # เช่น "1", "2", หรือข้อความอื่นๆ
                    else:
                        filing_type_display = filing_type
                elif is_ppnd54 and ocr_result.get('ppnd54_json'):
                    # ดึงประเภทการยื่นจาก JSON structure ใหม่ (ภ.ง.ด.54)
                    ppnd54_json = ocr_result.get('ppnd54_json', {})
                    data_obj = ppnd54_json.get('data', {})
                    
                    filing_type_normal = data_obj.get('(1) ยื่นปกติ', '')
                    filing_type_supplement = data_obj.get('(2) ยื่นเพิ่มเติมครั้งที่', '')
                    
                    if filing_type_normal and filing_type_normal is not None and filing_type_normal != 'null' and str(filing_type_normal).strip():
                        filing_type_display = str(filing_type_normal).strip()  # เช่น "ยื่นปกติ"
                    elif filing_type_supplement and filing_type_supplement is not None and filing_type_supplement != 'null' and str(filing_type_supplement).strip():
                        filing_type_display = str(filing_type_supplement).strip()  # เช่น "1", "2", หรือข้อความอื่นๆ
                    else:
                        filing_type_display = filing_type
                else:
                    filing_type_display = filing_type
                
                if filing_type_display:
                    print(f"📝 ประเภทการยื่น: {filing_type_display}")
                
                if ocr_result.get('filing_period'):
                    period = ocr_result['filing_period']
                    if period.get('month') and period.get('year'):
                        print(f"📅 ระยะเวลา: ประจำเดือน {period['month']} {period['year']}")
                
                if ocr_result.get('payment_date'):
                    print(f"💳 วันที่ชำระ: {ocr_result['payment_date']}")
                
                if ocr_result.get('due_date'):
                    print(f"⏰ วันที่ครบกำหนด: {ocr_result['due_date']}")
                
                # สำหรับ กยศ.: แสดงข้อมูลเพิ่มเติม
                tax_form_type = ocr_result.get('tax_form_type', '')
                is_kys = isinstance(tax_form_type, str) and ('กยศ.' in tax_form_type or 'กองทุน' in tax_form_type)
                
                if is_kys:
                    # แสดงข้อมูลเพิ่มเติมจาก กยศ.
                    if ocr_result.get('payment_month'):
                        payment_month = ocr_result.get('payment_month', '')
                        print(f"📅 ชำระเงินของเดือน: {payment_month}")
                    
                    if ocr_result.get('payment_amount'):
                        payment_amount = ocr_result.get('payment_amount', 0)
                        print(f"💰 ยอดชำระ (บาท): {payment_amount:,.2f} บาท")
                
                # สำหรับ Pay-in: แสดงข้อมูลเพิ่มเติม
                is_payin = isinstance(tax_form_type, str) and ('Pay-in' in tax_form_type or 'payin' in tax_form_type.lower())
                
                if is_payin:
                    # แสดงข้อมูลเพิ่มเติมจาก Pay-in
                    if ocr_result.get('payment_amount'):
                        payment_amount = ocr_result.get('payment_amount', 0)
                        print(f"💰 ยอดชำระ (บาท): {payment_amount:,.2f} บาท")
                
                # สำหรับแบบประกันสังคม: แสดงข้อมูลเพิ่มเติม
                is_social_security = isinstance(tax_form_type, str) and ('แบบประกันสังคม' in tax_form_type or 'ประกันสังคม' in tax_form_type)
                
                if is_social_security:
                    # แสดงข้อมูลเพิ่มเติมจากแบบประกันสังคม
                    if ocr_result.get('month'):
                        month_display = ocr_result.get('month')
                        print(f"📅 การนำส่งเงินสมทบสำหรับค่าจ้างเดือน: {month_display}")
                    
                    if ocr_result.get('year'):
                        year_display = ocr_result.get('year')
                        print(f"📅 พ.ศ.: {year_display}")
                    
                    if ocr_result.get('total_wages'):
                        total_wages = ocr_result.get('total_wages', 0)
                        print(f"💰 เงินค่าจ้างทั้งสิ้น: {total_wages:,.2f} บาท")
                    
                    if ocr_result.get('employee_contribution'):
                        employee_contribution = ocr_result.get('employee_contribution', 0)
                        print(f"💰 เงินสมทบผู้ประกันตน: {employee_contribution:,.2f} บาท")
                    
                    if ocr_result.get('employer_contribution'):
                        employer_contribution = ocr_result.get('employer_contribution', 0)
                        print(f"💰 เงินสมทบนายจ้าง: {employer_contribution:,.2f} บาท")
                    
                    if ocr_result.get('total_contribution'):
                        total_contribution = ocr_result.get('total_contribution', 0)
                        print(f"💰 รวมเงินสมทบที่นำส่งทั้งสิ้น: {total_contribution:,.2f} บาท")
                
                # แสดงยอดเงิน
                amounts = ocr_result.get('amounts', {})
                if amounts:
                    print("\n💰 ยอดเงิน:")
                    print("-" * 70)
                    for key, value in amounts.items():
                        if value is not None:
                            if isinstance(value, (int, float)):
                                formatted_value = f"{value:,.2f}"
                            else:
                                formatted_value = str(value)
                            print(f"   • {key}: {formatted_value} บาท")
                    print("-" * 70)
            else:
                error_msg = ocr_result.get('error', 'ไม่สามารถอ่านข้อมูลได้')
                print(f"❌ เกิดข้อผิดพลาด: {error_msg}")
            
            print("="*70 + "\n")
        
        # ตรวจสอบว่าควรทำ OCR หรือไม่ (จากชื่อไฟล์)
        # ตรวจสอบประเภทไฟล์จาก request
        file_type = request.form.get('file_type', 'tax_attachment')
        tax_form_type = request.form.get('tax_form_type', '')
        
        # ถ้าชื่อไฟล์มีคำว่า "แบบ" หรือ "Pay-in" หรือ file_type เป็น tax_filing/pp30 ให้ทำ OCR
        should_do_ocr = False
        specialized_processor_used = False
        
        # รายการ file types ที่ต้องใช้ specialized OCR processors
        specialized_types = ['pp30', 'pp36', 'ppnd1', 'ppnd3', 'ppnd53', 'ppnd54', 'social_security', 'kys', 'payin']
        
        if file_type in specialized_types or tax_form_type in specialized_types:
            should_do_ocr = True
            logger.info(f"📋 [OCR] ผู้ใช้เลือกประเภทไฟล์เป็น {file_type}: {original_filename}")
        elif file_type == 'tax_filing':
            should_do_ocr = True
            logger.info(f"📋 [OCR] ผู้ใช้เลือกประเภทไฟล์เป็นแบบยื่นภาษี: {original_filename}")
        elif file_extension.lower() == '.pdf':
            filename_lower = original_filename.lower()
            if 'แบบ' in filename_lower or 'pay-in' in filename_lower or 'ประกันสังคม' in filename_lower or 'กยศ' in filename_lower:
                should_do_ocr = True
                logger.info(f"📋 [OCR] ตรวจพบคำสำคัญในชื่อไฟล์ - จะทำ OCR: {original_filename}")
            else:
                logger.info(f"⏭️  [OCR] ไม่พบคำสำคัญในชื่อไฟล์ - ข้าม OCR: {original_filename}")
        
        # ถ้าควรทำ OCR ให้เรียกใช้ OCR
        ocr_result = None
        if should_do_ocr:
            
            # ฟังก์ชัน helper สำหรับเรียกใช้ specialized OCR processors
            def process_specialized_ocr(processor_type, file_path, filename):
                """Process OCR using specialized processor"""
                processor_map = {
                    'pp30': {
                        'module': 'testocr.pp30_ocr_processor',
                        'process_func': 'process_pp30_ocr',
                        'parse_func': 'parse_pp30_data',
                        'tax_form_type': 'ภ.พ.30',
                        'method': 'pp30_ocr_processor'
                    },
                    'pp36': {
                        'module': 'testocr.pp36_ocr_processor',
                        'process_func': 'process_pp36_ocr',
                        'parse_func': 'parse_pp36_data',
                        'tax_form_type': 'ภ.พ.36',
                        'method': 'pp36_ocr_processor'
                    },
                    'ppnd1': {
                        'module': 'testocr.ppnd1_ocr_processor',
                        'process_func': 'process_ppnd1_ocr',
                        'parse_func': 'parse_ppnd1_data',
                        'tax_form_type': 'ภ.ง.ด.1',
                        'method': 'ppnd1_ocr_processor'
                    },
                    'ppnd3': {
                        'module': 'testocr.ppnd3_ocr_processor',
                        'process_func': 'process_ppnd3_ocr',
                        'parse_func': 'parse_ppnd3_data',
                        'tax_form_type': 'ภ.ง.ด.3',
                        'method': 'ppnd3_ocr_processor'
                    },
                    'ppnd53': {
                        'module': 'testocr.ppnd53_ocr_processor',
                        'process_func': 'process_ppnd53_ocr',
                        'parse_func': 'parse_ppnd53_data',
                        'tax_form_type': 'ภ.ง.ด.53',
                        'method': 'ppnd53_ocr_processor'
                    },
                    'ppnd54': {
                        'module': 'testocr.ppnd54_ocr_processor',
                        'process_func': 'process_ppnd54_ocr',
                        'parse_func': 'parse_ppnd54_data',
                        'tax_form_type': 'ภ.ง.ด.54',
                        'method': 'ppnd54_ocr_processor'
                    },
                    'social_security': {
                        'module': 'testocr.social_security_ocr_processor',
                        'process_func': 'process_social_security_ocr',
                        'parse_func': 'parse_social_security_data',
                        'tax_form_type': 'แบบประกันสังคม',
                        'method': 'social_security_ocr_processor'
                    },
                    'kys': {
                        'module': 'testocr.kys_ocr_processor',
                        'process_func': 'process_kys_ocr',
                        'parse_func': 'parse_kys_data',
                        'tax_form_type': 'กยศ.',
                        'method': 'kys_ocr_processor'
                    },
                    'payin': {
                        'module': 'testocr.payin_ocr_processor',
                        'process_func': 'process_payin_ocr',
                        'parse_func': 'parse_payin_data',
                        'tax_form_type': 'Pay-in ชำระภาษี',
                        'method': 'payin_ocr_processor'
                    }
                }
                
                if processor_type not in processor_map:
                    return None
                
                config = processor_map[processor_type]
                
                try:
                    # Dynamic import
                    module = __import__(config['module'], fromlist=[config['process_func'], config['parse_func']])
                    process_func = getattr(module, config['process_func'])
                    parse_func = getattr(module, config['parse_func'])
                    
                    logger.info(f"📤 [{config['tax_form_type']} OCR] กำลังประมวลผลไฟล์: {filename}")
                    
                    # Process OCR
                    ocr_result_raw = process_func(str(file_path))
                    
                    if ocr_result_raw.get('success'):
                        # Parse data
                        parsed_data = parse_func(ocr_result_raw)
                        
                        # แปลงเป็นรูปแบบเดียวกับ TaxOCRProcessor
                        ocr_result = {
                            'success': True,
                            'tax_form_type': config['tax_form_type'],
                            'amounts': parsed_data.get('amounts', {}),
                            'method': config['method'],
                            'raw_text': json.dumps(ocr_result_raw.get('raw_response', {}), ensure_ascii=False, indent=2),
                            'formatted_text': json.dumps(parsed_data.get('raw_fields', {}), ensure_ascii=False, indent=2)
                        }
                        
                        # เพิ่มข้อมูลเพิ่มเติมจาก parsed_data (ถ้ามี)
                        for key in ['company_name', 'tax_id', 'address', 'filing_type', 'month', 'year', 
                                   'payment_amount', 'due_date', 'account_number', 'number_of_insured',
                                   'total_wages', 'employee_contribution', 'employer_contribution', 
                                   'total_contribution', 'payment_month', 'pp36_json', 'ppnd54_json', 
                                   'ppnd1_json', 'ppnd3_json', 'ppnd53_json', 'social_security_json', 
                                   'kys_json', 'payin_json']:  # เพิ่ม JSON fields ทั้งหมด
                            if key in parsed_data:
                                ocr_result[key] = parsed_data[key]
                        
                        # ส่งข้อมูลดิบไปแสดงผล
                        full_result_copy = {k: v for k, v in ocr_result.items() if k != 'raw_text'}
                        raw_data = {
                            'raw_text': ocr_result.get('raw_text', ''),
                            'method': config['method'],
                            'full_result': full_result_copy
                        }
                        print_ocr_result(ocr_result, filename, raw_data=raw_data)
                        return ocr_result
                    else:
                        logger.error(f"❌ [{config['tax_form_type']} OCR] ไม่สำเร็จ: {ocr_result_raw.get('error', 'Unknown error')}")
                        ocr_result = {
                            'success': False,
                            'error': f"{config['tax_form_type']} OCR Error: {ocr_result_raw.get('error', 'Unknown error')}"
                        }
                        print_ocr_result(ocr_result, filename)
                        return ocr_result
                        
                except ImportError as import_error:
                    logger.warning(f"⚠️ [{config['tax_form_type']} OCR] ไม่สามารถ import {config['module']} ได้: {import_error}")
                    logger.info("📋 [OCR] จะใช้ TaxOCRProcessor แทน")
                    return None
                except Exception as e:
                    logger.error(f"❌ [{config['tax_form_type']} OCR] เกิดข้อผิดพลาด: {e}", exc_info=True)
                    ocr_result = {
                        'success': False,
                        'error': f'เกิดข้อผิดพลาดในการประมวลผล {config["tax_form_type"]} OCR: {e}'
                    }
                    print_ocr_result(ocr_result, filename)
                    return ocr_result
            
            # ถ้าเป็น specialized processor ให้ใช้ specialized OCR processor
            if file_type in specialized_types:
                ocr_result = process_specialized_ocr(file_type, temp_file_path, original_filename)
                if ocr_result:
                    specialized_processor_used = True
            
            # ถ้าไม่ใช่ specialized processor หรือ specialized OCR ไม่สำเร็จ ให้ใช้ TaxOCRProcessor
            if not specialized_processor_used or (ocr_result and not ocr_result.get('success')):
                try:
                    from email_system.tax_ocr_processor import TaxOCRProcessor
                    processor = TaxOCRProcessor()
                    ocr_result = processor.extract_tax_amounts(temp_file_path)
                    # ส่งข้อมูลดิบไปแสดงผล (ไม่รวม raw_text ใน full_result เพื่อไม่ให้ซ้ำ)
                    full_result_copy = {k: v for k, v in ocr_result.items() if k != 'raw_text'}
                    raw_data = {
                        'raw_text': ocr_result.get('raw_text', ''),
                        'method': ocr_result.get('method', 'unknown'),
                        'full_result': full_result_copy  # ส่งข้อมูลทั้งหมดไปด้วย (ไม่รวม raw_text)
                    }
                    print_ocr_result(ocr_result, original_filename, raw_data=raw_data)
                except ImportError:
                    # ลอง import จาก root directory
                    try:
                        from tax_ocr_processor import TaxOCRProcessor
                        processor = TaxOCRProcessor()
                        ocr_result = processor.extract_tax_amounts(temp_file_path)
                        # ส่งข้อมูลดิบไปแสดงผล (ไม่รวม raw_text ใน full_result เพื่อไม่ให้ซ้ำ)
                        full_result_copy = {k: v for k, v in ocr_result.items() if k != 'raw_text'}
                        raw_data = {
                            'raw_text': ocr_result.get('raw_text', ''),
                            'method': ocr_result.get('method', 'unknown'),
                            'full_result': full_result_copy  # ส่งข้อมูลทั้งหมดไปด้วย (ไม่รวม raw_text)
                        }
                        print_ocr_result(ocr_result, original_filename, raw_data=raw_data)
                    except Exception as ocr_error:
                        logger.error(f"❌ [OCR] เกิดข้อผิดพลาดในการประมวลผล OCR: {ocr_error}", exc_info=True)
                        ocr_result = {
                            'success': False,
                            'error': f'เกิดข้อผิดพลาดในการประมวลผล OCR: {ocr_error}'
                        }
                        print_ocr_result(ocr_result, original_filename)
                except Exception as ocr_error:
                    logger.error(f"❌ [OCR] เกิดข้อผิดพลาดในการประมวลผล OCR: {ocr_error}", exc_info=True)
                    ocr_result = {
                        'success': False,
                        'error': f'เกิดข้อผิดพลาดในการประมวลผล OCR: {ocr_error}'
                    }
                    print_ocr_result(ocr_result, original_filename)
        
        response_data = {
            'success': True,
            'file_id': file_id,
            'filename': original_filename,
            'original_filename': original_filename,
            'temp_path': str(temp_file_path),
            'file_type': file_type,
            'message': f'อัปโหลดไฟล์ "{original_filename}" สำเร็จ'
        }
        
        # เพิ่มผลลัพธ์ OCR ถ้ามี
        if ocr_result:
            response_data['ocr_result'] = ocr_result
        
        return jsonify(response_data), 200
    except Exception as e:
        print(f"❌ [Email Upload] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาดในการอัปโหลด: {e}'}), 500

@app.route('/api/email/upload/<file_id>', methods=['DELETE'])
def delete_uploaded_file(file_id):
    """ลบไฟล์ที่อัปโหลดแล้ว"""
    try:
        temp_dir = Path('temp_uploads')
        if not temp_dir.exists():
            return jsonify({'success': False, 'message': 'ไม่พบโฟลเดอร์ temp_uploads'}), 404
        
        # ค้นหาไฟล์ที่ตรงกับ file_id
        deleted = False
        for file_path in temp_dir.glob(f'{file_id}.*'):
            try:
                file_path.unlink()
                deleted = True
                print(f"🗑️ [Email Upload] ลบไฟล์: {file_path}")
            except Exception as e:
                logger.warning(f"ไม่สามารถลบไฟล์ได้: {file_path} - {e}")
        
        if deleted:
            return jsonify({'success': True, 'message': 'ลบไฟล์สำเร็จ'}), 200
        else:
            return jsonify({'success': False, 'message': 'ไม่พบไฟล์ที่ต้องการลบ'}), 404
    except Exception as e:
        logger.error(f"เกิดข้อผิดพลาดในการลบไฟล์: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/cleanup', methods=['POST'])
def api_cleanup_temp_files():
    """ลบไฟล์เก่าใน temp_uploads ที่ไม่ได้ใช้แล้ว (มากกว่า 30 นาที)"""
    try:
        # ใช้ cleanup_temp_files function
        temp_dir = Path('temp_uploads')
        if not temp_dir.exists():
            return jsonify({'success': True, 'message': 'ไม่มีไฟล์ให้ลบ', 'deleted_count': 0}), 200
        
        # นับไฟล์ก่อนลบ
        files_before = len(list(temp_dir.glob('*')))
        
        # ลบไฟล์เก่ากว่า 30 นาที
        cleanup_temp_files(max_age_minutes=30)
        
        # นับไฟล์หลังลบ
        files_after = len(list(temp_dir.glob('*')))
        deleted_count = files_before - files_after
        
        return jsonify({
            'success': True,
            'message': f'ลบไฟล์เก่า {deleted_count} ไฟล์สำเร็จ',
            'deleted_count': deleted_count,
            'remaining_files': files_after
        }), 200
    except Exception as e:
        logger.error(f"เกิดข้อผิดพลาดในการลบไฟล์เก่า: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/cleanup-all', methods=['POST'])
def api_cleanup_all_temp_files():
    """ลบไฟล์ทั้งหมดใน temp_uploads ทันที (ไม่ต้องรอให้เก่า)"""
    try:
        temp_dir = Path('temp_uploads')
        if not temp_dir.exists():
            return jsonify({'success': True, 'message': 'ไม่มีไฟล์ให้ลบ', 'deleted_count': 0}), 200
        
        deleted_count = 0
        failed_files = []
        
        for file_path in temp_dir.glob('*'):
            try:
                if file_path.is_file():
                    file_path.unlink()
                    deleted_count += 1
                    logger.info(f"🗑️ [Cleanup All] ลบไฟล์: {file_path.name}")
            except Exception as e:
                failed_files.append(file_path.name)
                logger.warning(f"⚠️ ไม่สามารถลบไฟล์ได้: {file_path} - {e}")
        
        message = f'ลบไฟล์ {deleted_count} ไฟล์สำเร็จ'
        if failed_files:
            message += f' (ไม่สามารถลบ {len(failed_files)} ไฟล์: {", ".join(failed_files[:5])})'
        
        return jsonify({
            'success': True,
            'message': message,
            'deleted_count': deleted_count,
            'failed_count': len(failed_files),
            'failed_files': failed_files[:10]  # ส่งรายชื่อไฟล์ที่ลบไม่ได้ (สูงสุด 10 ไฟล์)
        }), 200
    except Exception as e:
        logger.error(f"เกิดข้อผิดพลาดในการลบไฟล์ทั้งหมด: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

def save_email_history(
    pattern_name: Optional[str] = None,
    signature_name: Optional[str] = None,
    sent_time: Optional[datetime] = None,
    to_emails: Optional[List[str]] = None,
    cc_emails: Optional[List[str]] = None,
    subject: Optional[str] = None,
    body: Optional[str] = None,
    line_message: Optional[str] = None
):
    """
    บันทึกประวัติการส่งอีเมลเป็น JSON
    
    Args:
        pattern_name: ชื่อ Pattern ที่ใช้ (ถ้ามี)
        signature_name: ชื่อ Signature ที่ใช้ (ถ้ามี)
        sent_time: เวลาที่ส่ง (ถ้าไม่ระบุจะใช้เวลาปัจจุบัน)
        to_emails: รายการอีเมลล์ผู้รับ
        cc_emails: รายการอีเมลล์ CC
        subject: หัวข้ออีเมลล์
        body: เนื้อหาอีเมลล์
        line_message: ข้อความที่ส่งไปยัง LINE
    """
    try:
        if sent_time is None:
            sent_time = datetime.now()
        
        # สร้างโฟลเดอร์ email_system ถ้ายังไม่มี
        email_system_dir = Path('email_system')
        email_system_dir.mkdir(exist_ok=True)
        
        # ไฟล์ประวัติ
        history_file = email_system_dir / 'email_history.json'
        
        # โหลดประวัติเดิม (ถ้ามี)
        history = []
        if history_file.exists():
            try:
                with open(history_file, 'r', encoding='utf-8') as f:
                    history = json.load(f)
            except Exception as e:
                logger.warning(f"ไม่สามารถโหลด email history ได้: {e}")
                history = []
        
        # เพิ่มประวัติใหม่
        history_entry = {
            'pattern_name': pattern_name or None,
            'signature_name': signature_name or None,
            'sent_time': sent_time.strftime('%Y-%m-%d %H:%M:%S'),
            'to_emails': to_emails or [],
            'cc_emails': cc_emails or [],
            'subject': subject or None,
            'body': body or None,
            'line_message': line_message or None
        }
        
        history.append(history_entry)
        
        # บันทึกลงไฟล์
        with open(history_file, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
        
        logger.info(f"✅ บันทึกประวัติการส่งอีเมล: pattern={pattern_name}, signature={signature_name}, time={sent_time.strftime('%Y-%m-%d %H:%M:%S')}")
    except Exception as e:
        logger.error(f"❌ ไม่สามารถบันทึกประวัติการส่งอีเมลได้: {e}", exc_info=True)

@app.route('/api/email/preview', methods=['POST'])
def preview_email():
    """สร้าง preview ของอีเมลล์ที่จะส่ง"""
    data = request.json
    email_service = get_global_email_service()
    
    if not email_service:
        return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP'}), 400
    
    try:
        # ใช้ default recipients จาก config ถ้าไม่ระบุ
        to_emails = data.get('to_emails', [])
        if not to_emails and hasattr(Config, 'EMAIL_DEFAULT_TO') and Config.EMAIL_DEFAULT_TO:
            to_emails = Config.EMAIL_DEFAULT_TO.copy()
        
        subject = data.get('subject', '')
        body = data.get('body', '')
        
        # รวม CC จาก request และ default
        cc_emails = data.get('cc_emails')
        if cc_emails is None:
            if hasattr(Config, 'EMAIL_DEFAULT_CC') and Config.EMAIL_DEFAULT_CC:
                cc_emails = Config.EMAIL_DEFAULT_CC.copy()
            else:
                cc_emails = []
        elif isinstance(cc_emails, list) and len(cc_emails) == 0:
            cc_emails = []
        elif isinstance(cc_emails, list) and len(cc_emails) > 0:
            if hasattr(Config, 'EMAIL_DEFAULT_CC') and Config.EMAIL_DEFAULT_CC:
                cc_emails = list(set(cc_emails + Config.EMAIL_DEFAULT_CC))
        else:
            cc_emails = []
        
        # ดึงลายเซ็นต์ถ้ามี
        signature_name = data.get('signature_name')
        signature = None
        body_html = None
        
        if signature_name:
            signature = email_service.get_signature(signature_name)
            if signature:
                text_sig, html_sig = signature.generate_signature()
                
                # รวมลายเซ็นต์ใน body
                if body:
                    body_with_sig = body + "\n\n" + text_sig
                else:
                    body_with_sig = text_sig
                
                # สร้าง HTML version
                if html_sig:
                    # แปลง body เป็น HTML
                    body_html = body.replace('\n', '<br>') if body else ''
                    body_html = body_html + "<br><br>" + html_sig
                else:
                    body_html = body_with_sig.replace('\n', '<br>')
        
        # ถ้าไม่มีลายเซ็นต์ แปลง body เป็น HTML
        if not body_html:
            body_html = body.replace('\n', '<br>') if body else ''
        
        # ดึงข้อมูล attachments
        attachments = []
        attachment_data = data.get('attachments', [])
        if attachment_data:
            for att_data in attachment_data:
                if isinstance(att_data, dict):
                    filename = att_data.get('filename') or att_data.get('original_filename')
                    if filename:
                        attachments.append({'filename': filename})
        
        return jsonify({
            'success': True,
            'from_email': email_service.email_from,
            'to_emails': to_emails,
            'cc_emails': cc_emails,
            'subject': subject,
            'body': body,
            'body_html': body_html,
            'signature_name': signature_name,
            'attachments': attachments
        }), 200
    except Exception as e:
        logger.error(f"Error previewing email: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/send', methods=['POST'])
def send_email():
    """ส่งอีเมลล์"""
    data = request.json
    email_service = get_global_email_service()
    
    if not email_service:
        return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP'}), 400
    
    try:
        # ใช้ default recipients จาก config ถ้าไม่ระบุ
        to_emails = data.get('to_emails', [])
        if not to_emails and hasattr(Config, 'EMAIL_DEFAULT_TO') and Config.EMAIL_DEFAULT_TO:
            to_emails = Config.EMAIL_DEFAULT_TO.copy()
        
        subject = data.get('subject', '')
        body = data.get('body', '')
        body_html = data.get('body_html')
        
        # รวม CC จาก request และ default
        # ถ้า cc_emails เป็น None หรือไม่ได้ส่งมา = ใช้ default
        # ถ้า cc_emails เป็น [] (array ว่าง) = ไม่ใช้ CC เลย (ผู้ใช้ลบออกหมด)
        cc_emails = data.get('cc_emails')
        if cc_emails is None:
            # ไม่ได้ส่ง cc_emails มา = ใช้ default
            if hasattr(Config, 'EMAIL_DEFAULT_CC') and Config.EMAIL_DEFAULT_CC:
                cc_emails = Config.EMAIL_DEFAULT_CC.copy()
            else:
                cc_emails = []
        elif isinstance(cc_emails, list) and len(cc_emails) == 0:
            # ส่ง array ว่างมา = ผู้ใช้ต้องการลบ CC ทั้งหมด ไม่รวม default
            cc_emails = []
        elif isinstance(cc_emails, list) and len(cc_emails) > 0:
            # มี CC จากผู้ใช้ = รวมกับ default
            if hasattr(Config, 'EMAIL_DEFAULT_CC') and Config.EMAIL_DEFAULT_CC:
                cc_emails = list(set(cc_emails + Config.EMAIL_DEFAULT_CC))
        else:
            cc_emails = []
        
        # จัดการ BCC เหมือนกับ CC
        bcc_emails = data.get('bcc_emails')
        if bcc_emails is None:
            # ไม่ได้ส่ง bcc_emails มา = ใช้ default
            if hasattr(Config, 'EMAIL_DEFAULT_BCC') and Config.EMAIL_DEFAULT_BCC:
                bcc_emails = Config.EMAIL_DEFAULT_BCC.copy()
            else:
                bcc_emails = []
        elif isinstance(bcc_emails, list) and len(bcc_emails) == 0:
            # ส่ง array ว่างมา = ผู้ใช้ต้องการลบ BCC ทั้งหมด ไม่รวม default
            bcc_emails = []
        elif isinstance(bcc_emails, list) and len(bcc_emails) > 0:
            # มี BCC จากผู้ใช้ = รวมกับ default
            if hasattr(Config, 'EMAIL_DEFAULT_BCC') and Config.EMAIL_DEFAULT_BCC:
                bcc_emails = list(set(bcc_emails + Config.EMAIL_DEFAULT_BCC))
        else:
            bcc_emails = []
        
        # จัดการไฟล์แนบ
        attachments = None
        attachment_paths = []  # ประกาศไว้ก่อนเพื่อป้องกัน UnboundLocalError
        attachment_data = data.get('attachments', [])  # รองรับรูปแบบใหม่ (array of objects)
        attachment_paths_old = data.get('attachment_paths', [])  # รองรับรูปแบบเก่า (backward compatibility)
        
        if attachment_data or attachment_paths_old:
            attachments = []
            
            # รองรับรูปแบบใหม่ (array of objects)
            if attachment_data:
                for att_data in attachment_data:
                    if isinstance(att_data, dict):
                        temp_path = att_data.get('temp_path')
                        original_filename = att_data.get('original_filename') or att_data.get('filename')
                    else:
                        temp_path = att_data
                        original_filename = None
                    
                    if temp_path:
                        file_path = Path(temp_path)
                        if file_path.exists():
                            attachments.append({
                                'path': file_path,
                                'filename': original_filename or file_path.name
                            })
                            attachment_paths.append(temp_path)
            
            # รองรับรูปแบบเก่า (backward compatibility)
            elif attachment_paths_old:
                for temp_path in attachment_paths_old:
                    file_path = Path(temp_path)
                    if file_path.exists():
                        attachments.append(file_path)  # ใช้รูปแบบเก่า
                        attachment_paths.append(temp_path)
        
        # ดึง signature_name ถ้ามี
        signature_name = data.get('signature_name')
        
        # ดึง zip_attachments (default = True เพื่อ backward compatibility)
        zip_attachments = data.get('zip_attachments', True)
        
        # สร้าง PDF สรุป (ถ้ามีข้อมูล)
        summary_data = data.get('summary_data')
        pdf_path = None
        image_path = None
        if summary_data:
            try:
                pdf_generator = PDFGenerator()
                # ใช้ชื่อไฟล์ที่ส่งมาจาก frontend (ถ้ามี)
                pdf_filename = summary_data.get('filename')
                logger.info(f"📄 สร้าง PDF: filename={pdf_filename}")
                pdf_path = pdf_generator.generate_summary_pdf(
                    title=summary_data.get('title', 'สรุปข้อมูล'),
                    company_name=summary_data.get('company_name', ''),
                    period=summary_data.get('period', ''),
                    data=summary_data.get('data', []),
                    filename=pdf_filename,  # ส่งชื่อไฟล์ไปด้วย
                    tax_id=summary_data.get('tax_id', '')  # ส่ง tax_id ไปด้วย
                )
                
                # เพิ่ม PDF เป็น attachment (ไม่ต้องแปลงเป็นรูปภาพแล้ว)
                if pdf_path and pdf_path.exists():
                    if not attachments:
                        attachments = []
                    # ตั้งชื่อไฟล์ตามที่ส่งมาจาก frontend หรือใช้ชื่อ default
                    pdf_attachment_filename = pdf_filename if pdf_filename else 'สรุปภาษี.pdf'
                    logger.info(f"📎 Attach PDF: pdf_filename={pdf_filename}, final_name={pdf_attachment_filename}")
                    attachments.append({
                        'path': pdf_path,
                        'filename': pdf_attachment_filename
                    })
                    
                    # คัดลอก PDF ไปที่ V:\web\line-images ด้วย
                    try:
                        import shutil
                        line_images_dir = Path(r"V:\web\line-images")
                        line_images_dir.mkdir(parents=True, exist_ok=True)
                        line_images_path = line_images_dir / pdf_attachment_filename
                        shutil.copy2(pdf_path, line_images_path)
                        logger.info(f"✅ คัดลอก PDF ไปที่ {line_images_path}")
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถคัดลอก PDF ไปที่ V:\\web\\line-images: {e}")
            except Exception as e:
                logger.error(f"เกิดข้อผิดพลาดในการสร้าง PDF: {e}", exc_info=True)
        
        # กรองโลโก้ออกจาก attachments (ถ้ามี)
        if signature_name and attachments:
            signature = email_service.get_signature(signature_name)
            if signature and signature.logo_path:
                signature_logo_path = Path(signature.logo_path)
                if signature_logo_path.exists():
                    # กรอง signature logo ออกจาก attachments
                    filtered_attachments = []
                    for att in attachments:
                        if isinstance(att, dict):
                            att_path = att.get('path')
                        else:
                            att_path = att
                        
                        if att_path:
                            att_path_obj = Path(att_path) if not isinstance(att_path, Path) else att_path
                            if att_path_obj.resolve() != signature_logo_path.resolve():
                                filtered_attachments.append(att)
                    attachments = filtered_attachments if filtered_attachments else None
        
        success, message = email_service.send_email(
            to_emails=to_emails,
            subject=subject,
            body=body,
            body_html=body_html,
            cc_emails=cc_emails if cc_emails else None,
            bcc_emails=bcc_emails if bcc_emails else None,
            attachments=attachments,
            signature_name=signature_name,
            zip_attachments=zip_attachments
        )
        
        # ลบไฟล์ชั่วคราวหลังจากส่งอีเมลล์เสร็จ (ไม่ว่าจะสำเร็จหรือไม่)
        # เนื่องจากไฟล์ถูกส่งไปแล้ว หรือถ้าไม่สำเร็จก็ไม่จำเป็นต้องเก็บไว้
        if attachment_paths:
            for temp_path in attachment_paths:
                try:
                    file_path = Path(temp_path)
                    if file_path.exists() and 'temp_uploads' in str(file_path):
                        file_path.unlink()
                        logger.debug(f"🗑️ ลบไฟล์ชั่วคราว: {file_path.name}")
                except Exception as e:
                    logger.warning(f"ไม่สามารถลบไฟล์ชั่วคราวได้: {temp_path} - {e}")
        
        if success:
            # บันทึกประวัติการส่งอีเมล (ไม่มี pattern_name เพราะเป็นการส่งปกติ)
            save_email_history(
                pattern_name=None,
                signature_name=signature_name,
                to_emails=to_emails,
                cc_emails=cc_emails,
                subject=subject,
                body=body
            )
            return jsonify({'success': True, 'message': message}), 200
        else:
            return jsonify({'success': False, 'message': message}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/pattern', methods=['POST'])
def create_email_pattern():
    """สร้าง email pattern ใหม่"""
    data = request.json
    email_service = get_global_email_service()
    
    if not email_service:
        return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP'}), 400
    
    try:
        pattern = EmailPattern(
            pattern_name=data.get('pattern_name', ''),
            subject_template=data.get('subject_template', ''),
            body_template=data.get('body_template', ''),
            body_html_template=data.get('body_html_template'),
            default_to=data.get('default_to', []),
            default_cc=data.get('default_cc', []),
            default_bcc=data.get('default_bcc', []),
            description=data.get('description', ''),
            logo_path=data.get('logo_path'),
            company_name=data.get('company_name', ''),
            line_user_id=data.get('line_user_id', ''),
            tax_id=data.get('tax_id', '')
        )
        
        email_service.add_pattern(pattern)
        return jsonify({'success': True, 'message': f'สร้างแพทเทิร์น "{pattern.pattern_name}" สำเร็จ'}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/pattern', methods=['GET'])
def list_email_patterns():
    """รายชื่อ email patterns ทั้งหมด"""
    email_service = get_global_email_service()
    
    if not email_service:
        return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP', 'patterns': []}), 200
    
    try:
        # โหลด patterns ใหม่จากไฟล์ (เพื่อให้แน่ใจว่ามีข้อมูลล่าสุด)
        email_service.load_patterns()
        
        patterns = email_service.list_patterns()
        patterns_data = []
        for pattern_name in patterns:
            pattern = email_service.get_pattern(pattern_name)
            if pattern:
                patterns_data.append(pattern.to_dict())
        
        logger.info(f"✅ โหลด patterns สำเร็จ: {len(patterns_data)} รายการ")
        return jsonify({'success': True, 'patterns': patterns_data}), 200
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการโหลด patterns: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}', 'patterns': []}), 500

@app.route('/api/email/pattern/<pattern_name>', methods=['GET'])
def get_email_pattern(pattern_name):
    """ดึง email pattern"""
    email_service = get_global_email_service()
    
    if not email_service:
        return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP'}), 400
    
    try:
        pattern = email_service.get_pattern(pattern_name)
        if pattern:
            return jsonify({'success': True, 'pattern': pattern.to_dict()}), 200
        else:
            return jsonify({'success': False, 'message': f'ไม่พบแพทเทิร์น: {pattern_name}'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/pattern/<pattern_name>', methods=['PUT', 'PATCH'])
def update_email_pattern(pattern_name):
    """อัปเดต email pattern"""
    data = request.json
    email_service = get_global_email_service()
    
    if not email_service:
        return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP'}), 400
    
    try:
        pattern = EmailPattern(
            pattern_name=data.get('pattern_name', pattern_name),  # ใช้ชื่อใหม่ถ้ามี
            subject_template=data.get('subject_template', ''),
            body_template=data.get('body_template', ''),
            body_html_template=data.get('body_html_template'),
            default_to=data.get('default_to', []),
            default_cc=data.get('default_cc', []),
            default_bcc=data.get('default_bcc', []),
            description=data.get('description', ''),
            logo_path=data.get('logo_path'),
            company_name=data.get('company_name', ''),
            line_user_id=data.get('line_user_id', ''),
            tax_id=data.get('tax_id', '')
        )
        
        if email_service.update_pattern(pattern_name, pattern):
            return jsonify({'success': True, 'message': f'อัปเดตแพทเทิร์น "{pattern.pattern_name}" สำเร็จ'}), 200
        else:
            return jsonify({'success': False, 'message': f'ไม่พบแพทเทิร์น: {pattern_name}'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/pattern/<pattern_name>', methods=['DELETE'])
def delete_email_pattern(pattern_name):
    """ลบ email pattern"""
    email_service = get_global_email_service()
    
    if not email_service:
        return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP'}), 400
    
    try:
        if email_service.delete_pattern(pattern_name):
            return jsonify({'success': True, 'message': f'ลบแพทเทิร์น "{pattern_name}" สำเร็จ'}), 200
        else:
            return jsonify({'success': False, 'message': f'ไม่พบแพทเทิร์น: {pattern_name}'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/pattern/import', methods=['POST'])
def import_email_patterns():
    """นำเข้าข้อมูล Pattern จาก Excel"""
    email_service = get_global_email_service()
    
    if not email_service:
        return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP'}), 400
    
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'ไม่มีไฟล์ที่อัปโหลด'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'ไม่ได้เลือกไฟล์'}), 400
        
        # ตรวจสอบว่าเป็นไฟล์ Excel
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'กรุณาอัปโหลดไฟล์ Excel (.xlsx หรือ .xls)'}), 400
        
        # อ่านไฟล์ Excel
        try:
            from openpyxl import load_workbook
        except ImportError:
            return jsonify({'success': False, 'message': 'ไม่พบไลบรารี openpyxl กรุณาติดตั้งด้วย: pip install openpyxl'}), 500
        
        # บันทึกไฟล์ชั่วคราว
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            tmp_path = tmp_file.name
        
        try:
            wb = load_workbook(tmp_path, data_only=True)
            ws = wb.active
            
            # อ่าน header row
            headers = []
            header_row = None
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                if any(cell and str(cell).strip() for cell in row):
                    # หา header row
                    row_str = [str(cell).strip() if cell else '' for cell in row]
                    if any('ชื่อแพทเทิร์น' in str(cell) or 'pattern' in str(cell).lower() for cell in row if cell):
                        headers = row_str
                        header_row = idx
                        break
            
            if not headers or header_row is None:
                return jsonify({'success': False, 'message': 'ไม่พบ header row ที่ถูกต้อง กรุณาตรวจสอบไฟล์ Excel'}), 400
            
            # แมปชื่อคอลัมน์
            column_map = {}
            for idx, header in enumerate(headers, 1):
                if not header:
                    continue
                header_str = str(header).strip()
                header_lower = header_str.lower()
                
                # ตรวจสอบชื่อแพทเทิร์น
                if 'ชื่อแพทเทิร์น' in header_str or ('pattern' in header_lower and 'name' in header_lower):
                    column_map['pattern_name'] = idx
                # ตรวจสอบคำอธิบาย
                elif 'คำอธิบาย' in header_str or 'description' in header_lower:
                    column_map['description'] = idx
                # ตรวจสอบชื่อบริษัท (รองรับทั้ง "ชื่อบริษัท" และ "ชื่อบริษัทลูกค้า")
                elif 'ชื่อบริษัท' in header_str or ('company' in header_lower and 'name' in header_lower):
                    column_map['company_name'] = idx
                # ตรวจสอบเลขประจำตัวผู้เสียภาษี
                elif 'เลขประจำตัว' in header_str or ('tax' in header_lower and 'id' in header_lower):
                    column_map['tax_id'] = idx
                # ตรวจสอบ LINE User ID
                elif 'line' in header_lower and ('user' in header_lower or 'group' in header_lower or 'room' in header_lower):
                    column_map['line_user_id'] = idx
                # ตรวจสอบ Default To
                elif 'default to' in header_lower or ('to' in header_lower and 'default' in header_lower):
                    column_map['default_to'] = idx
                # ตรวจสอบ Default CC
                elif 'default cc' in header_lower or ('cc' in header_lower and 'default' in header_lower):
                    column_map['default_cc'] = idx
            
            # Debug: แสดง column_map
            logger.info(f"📊 Column mapping: {column_map}")
            
            # ตรวจสอบว่ามีคอลัมน์ที่จำเป็นหรือไม่
            required_columns = ['pattern_name', 'company_name', 'default_cc']
            missing_columns = [col for col in required_columns if col not in column_map]
            if missing_columns:
                return jsonify({
                    'success': False, 
                    'message': f'ไม่พบคอลัมน์ที่จำเป็น: {", ".join(missing_columns)}'
                }), 400
            
            imported_count = 0
            errors = []
            
            # อ่านข้อมูลแต่ละแถว
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
                # ตรวจสอบว่าแถวนี้มีข้อมูลหรือไม่
                if not any(cell and str(cell).strip() for cell in row):
                    continue
                
                try:
                    # ฟังก์ชันช่วยในการอ่านค่า cell
                    def get_cell_value(col_idx):
                        """อ่านค่าจาก cell โดยจัดการ None และ empty"""
                        if col_idx <= 0 or col_idx > len(row):
                            return ''
                        cell_value = row[col_idx - 1]
                        if cell_value is None:
                            return ''
                        # แปลงเป็น string และ strip
                        return str(cell_value).strip()
                    
                    # ดึงข้อมูลจากแต่ละคอลัมน์
                    pattern_name = get_cell_value(column_map['pattern_name'])
                    description = get_cell_value(column_map.get('description', 0))
                    company_name = get_cell_value(column_map['company_name'])
                    tax_id_raw = get_cell_value(column_map.get('tax_id', 0))
                    line_user_id = get_cell_value(column_map.get('line_user_id', 0))
                    default_to_raw = get_cell_value(column_map.get('default_to', 0))
                    default_cc_raw = get_cell_value(column_map['default_cc'])
                    
                    # Validation
                    if not pattern_name:
                        errors.append(f'แถว {row_idx}: ไม่พบชื่อแพทเทิร์น')
                        continue
                    
                    if not company_name:
                        errors.append(f'แถว {row_idx}: ไม่พบชื่อบริษัท')
                        continue
                    
                    if not default_cc_raw:
                        errors.append(f'แถว {row_idx}: ไม่พบ Default CC')
                        continue
                    
                    # Format tax_id เป็น 13 หลัก
                    tax_id = ''
                    if tax_id_raw:
                        # ลบอักขระที่ไม่ใช่ตัวเลข
                        tax_id_clean = ''.join(filter(str.isdigit, tax_id_raw))
                        if len(tax_id_clean) == 12:
                            # เติม 0 นำหน้า
                            tax_id = '0' + tax_id_clean
                        elif len(tax_id_clean) == 13:
                            tax_id = tax_id_clean
                        elif len(tax_id_clean) > 0:
                            errors.append(f'แถว {row_idx}: เลขประจำตัวผู้เสียภาษีอากรไม่ถูกต้อง (ต้องเป็น 12 หรือ 13 หลัก)')
                    
                    # แปลง email lists
                    default_to = [email.strip() for email in default_to_raw.split(',') if email.strip()] if default_to_raw else []
                    default_cc = [email.strip() for email in default_cc_raw.split(',') if email.strip()] if default_cc_raw else []
                    
                    # สร้าง Pattern
                    pattern = EmailPattern(
                        pattern_name=pattern_name,
                        subject_template='',
                        body_template='',
                        body_html_template=None,
                        default_to=default_to,
                        default_cc=default_cc,
                        default_bcc=[],
                        description=description,
                        logo_path=None,
                        company_name=company_name,
                        line_user_id=line_user_id,
                        tax_id=tax_id
                    )
                    
                    # เพิ่มหรืออัปเดต Pattern
                    email_service.add_pattern(pattern)
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f'แถว {row_idx}: {str(e)}')
                    continue
            
            # ลบไฟล์ชั่วคราว
            try:
                os.unlink(tmp_path)
            except:
                pass
            
            return jsonify({
                'success': True,
                'message': f'นำเข้าข้อมูลสำเร็จ {imported_count} รายการ',
                'imported_count': imported_count,
                'errors': errors
            }), 200
            
        except Exception as e:
            # ลบไฟล์ชั่วคราว
            try:
                os.unlink(tmp_path)
            except:
                pass
            return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาดในการอ่านไฟล์ Excel: {str(e)}'}), 500
            
    except Exception as e:
        logger.error(f"Error importing patterns: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {str(e)}'}), 500

@app.route('/api/email/logo/upload', methods=['POST'])
def upload_email_logo():
    """อัปโหลดโลโก้สำหรับใช้ใน email pattern"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'ไม่มีไฟล์ที่อัปโหลด'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'ไม่ได้เลือกไฟล์'}), 400
        
        # ตรวจสอบว่าเป็นไฟล์รูปภาพ
        allowed_extensions = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.svg'}
        file_ext = Path(file.filename).suffix.lower()
        if file_ext not in allowed_extensions:
            return jsonify({'success': False, 'message': f'ไฟล์ต้องเป็นรูปภาพ ({", ".join(allowed_extensions)})'}), 400
        
        # สร้างโฟลเดอร์ email_logos ถ้ายังไม่มี (ใน email_system folder)
        email_system_dir = Path('email_system')
        logos_dir = email_system_dir / 'email_logos'
        logos_dir.mkdir(parents=True, exist_ok=True)
        
        # บันทึกไฟล์โลโก้
        import uuid
        file_id = str(uuid.uuid4())
        logo_filename = f"{file_id}{file_ext}"
        logo_path = logos_dir / logo_filename
        
        file.save(logo_path)
        
        return jsonify({
            'success': True,
            'logo_path': str(logo_path),
            'filename': file.filename,
            'message': f'อัปโหลดโลโก้ "{file.filename}" สำเร็จ'
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาดในการอัปโหลด: {e}'}), 500

@app.route('/api/email/logo/<path:logo_path>', methods=['GET'])
def get_email_logo(logo_path):
    """ดึงโลโก้สำหรับแสดง preview"""
    try:
        logo_file = Path(logo_path)
        # รองรับทั้ง path เดิมและ path ใหม่ใน email_system
        if not logo_file.exists():
            # ลองหาใน email_system folder
            email_system_logo = Path('email_system') / 'email_logos' / logo_file.name
            if email_system_logo.exists():
                logo_file = email_system_logo
        
        if logo_file.exists() and ('email_logos' in str(logo_file) or 'email_system' in str(logo_file)):
            return send_from_directory(logo_file.parent, logo_file.name)
        else:
            return jsonify({'error': 'ไม่พบไฟล์โลโก้'}), 404
    except Exception as e:
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/temp/pdf/<path:pdf_path>', methods=['GET'])
def get_temp_pdf(pdf_path):
    """Serve ไฟล์ PDF ชั่วคราวสำหรับ LINE"""
    try:
        pdf_file = Path(pdf_path)
        # ตรวจสอบว่าไฟล์อยู่ใน temp_uploads หรือไม่
        if not pdf_file.exists():
            # ลองหาใน temp_uploads
            temp_pdf = Path('temp_uploads') / pdf_file.name
            if temp_pdf.exists():
                pdf_file = temp_pdf
            else:
                return jsonify({'error': 'File not found'}), 404
        
        # ตรวจสอบว่าเป็นไฟล์ PDF หรือไม่
        if pdf_file.suffix.lower() != '.pdf':
            return jsonify({'error': 'Invalid file type'}), 400
        
        # ส่งไฟล์ PDF
        return send_from_directory(
            pdf_file.parent,
            pdf_file.name,
            mimetype='application/pdf',
            as_attachment=True,
            download_name='สรุปภาษี.pdf'
        )
    except Exception as e:
        logger.error(f"Error serving PDF: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/temp/image/<path:image_path>', methods=['GET'])
def get_temp_image(image_path):
    """Serve รูปภาพชั่วคราวสำหรับ LINE (ต้องใช้ HTTPS URL)"""
    try:
        image_file = Path(image_path)
        # ตรวจสอบว่าไฟล์อยู่ใน temp_uploads หรือไม่
        if not image_file.exists():
            # ลองหาใน temp_uploads
            temp_image = Path('temp_uploads') / image_file.name
            if temp_image.exists():
                image_file = temp_image
        
        if image_file.exists() and ('temp_uploads' in str(image_file) or image_file.suffix.lower() in ['.png', '.jpg', '.jpeg', '.gif']):
            return send_from_directory(image_file.parent, image_file.name)
        else:
            return jsonify({'error': 'ไม่พบไฟล์รูปภาพ'}), 404
    except Exception as e:
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

# Email Signature Endpoints
@app.route('/api/email/signature', methods=['POST'])
def create_email_signature():
    """สร้าง email signature ใหม่"""
    data = request.json
    email_service = get_global_email_service()
    
    if not email_service:
        return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP'}), 400
    
    try:
        from email_system.email_service import EmailSignature
        signature = EmailSignature(
            signature_name=data.get('signature_name', ''),
            sender_name=data.get('sender_name', ''),
            sender_title=data.get('sender_title', ''),
            sender_email=data.get('sender_email', ''),
            sender_phone=data.get('sender_phone', ''),
            sender_website=data.get('sender_website', ''),
            company_name=data.get('company_name', ''),
            company_address=data.get('company_address', ''),
            signature_html=data.get('signature_html'),
            signature_block=data.get('signature_block'),
            logo_path=data.get('logo_path'),
            description=data.get('description', '')
        )
        
        email_service.add_signature(signature)
        return jsonify({'success': True, 'message': f'สร้างลายเซ็นต์ "{signature.signature_name}" สำเร็จ'}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/signature', methods=['GET'])
def list_email_signatures():
    """รายชื่อ email signatures ทั้งหมด"""
    email_service = get_global_email_service()
    
    if not email_service:
        return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP', 'signatures': []}), 200
    
    try:
        # โหลด signatures จากไฟล์ก่อน
        email_service.load_signatures()
        signatures = email_service.list_signatures()
        signatures_data = []
        for signature_name in signatures:
            signature = email_service.get_signature(signature_name)
            if signature:
                signatures_data.append(signature.to_dict())
        
        return jsonify({'success': True, 'signatures': signatures_data}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}', 'signatures': []}), 500

@app.route('/api/email/signature/<signature_name>', methods=['GET'])
def get_email_signature(signature_name):
    """ดึง email signature"""
    email_service = get_global_email_service()
    
    if not email_service:
        return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP'}), 400
    
    try:
        signature = email_service.get_signature(signature_name)
        if signature:
            return jsonify({'success': True, 'signature': signature.to_dict()}), 200
        else:
            return jsonify({'success': False, 'message': f'ไม่พบลายเซ็นต์: {signature_name}'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/signature/<signature_name>', methods=['PUT', 'PATCH'])
def update_email_signature(signature_name):
    """อัปเดต email signature"""
    data = request.json
    email_service = get_global_email_service()
    
    if not email_service:
        return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP'}), 400
    
    try:
        from email_system.email_service import EmailSignature
        signature = EmailSignature(
            signature_name=data.get('signature_name', signature_name),
            sender_name=data.get('sender_name', ''),
            sender_title=data.get('sender_title', ''),
            sender_email=data.get('sender_email', ''),
            sender_phone=data.get('sender_phone', ''),
            sender_website=data.get('sender_website', ''),
            company_name=data.get('company_name', ''),
            company_address=data.get('company_address', ''),
            signature_html=data.get('signature_html'),
            logo_path=data.get('logo_path'),
            description=data.get('description', '')
        )
        
        if email_service.update_signature(signature_name, signature):
            return jsonify({'success': True, 'message': f'อัปเดตลายเซ็นต์ "{signature.signature_name}" สำเร็จ'}), 200
        else:
            return jsonify({'success': False, 'message': f'ไม่พบลายเซ็นต์: {signature_name}'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/signature/<signature_name>', methods=['DELETE'])
def delete_email_signature(signature_name):
    """ลบ email signature"""
    email_service = get_global_email_service()
    
    if not email_service:
        return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP'}), 400
    
    try:
        if email_service.delete_signature(signature_name):
            return jsonify({'success': True, 'message': f'ลบลายเซ็นต์ "{signature_name}" สำเร็จ'}), 200
        else:
            return jsonify({'success': False, 'message': f'ไม่พบลายเซ็นต์: {signature_name}'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/send/pattern', methods=['POST'])
def send_email_by_pattern():
    """ส่งอีเมลล์ตามแพทเทิร์น"""
    data = request.json
    email_service = get_global_email_service()
    
    if not email_service:
        return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP'}), 400
    
    try:
        pattern_name = data.get('pattern_name', '')
        pattern_data = data.get('data', {})
        to_emails = data.get('to_emails')
        
        # จัดการ CC เหมือนกับ send_email
        cc_emails = data.get('cc_emails')
        if cc_emails is None:
            # ไม่ได้ส่ง cc_emails มา = ใช้ default จาก pattern
            cc_emails = None  # จะใช้ default จาก pattern
        elif isinstance(cc_emails, list) and len(cc_emails) == 0:
            # ส่ง array ว่างมา = ผู้ใช้ต้องการลบ CC ทั้งหมด ไม่รวม default
            cc_emails = []
        elif isinstance(cc_emails, list) and len(cc_emails) > 0:
            # มี CC จากผู้ใช้ = รวมกับ default จาก config (ถ้ามี)
            if hasattr(Config, 'EMAIL_DEFAULT_CC') and Config.EMAIL_DEFAULT_CC:
                cc_emails = list(set(cc_emails + Config.EMAIL_DEFAULT_CC))
        else:
            cc_emails = []
        
        # จัดการ BCC เหมือนกับ send_email
        bcc_emails = data.get('bcc_emails')
        if bcc_emails is None:
            # ไม่ได้ส่ง bcc_emails มา = ใช้ default จาก pattern
            bcc_emails = None  # จะใช้ default จาก pattern
        elif isinstance(bcc_emails, list) and len(bcc_emails) == 0:
            # ส่ง array ว่างมา = ผู้ใช้ต้องการลบ BCC ทั้งหมด ไม่รวม default
            bcc_emails = []
        elif isinstance(bcc_emails, list) and len(bcc_emails) > 0:
            # มี BCC จากผู้ใช้ = รวมกับ default จาก config (ถ้ามี)
            if hasattr(Config, 'EMAIL_DEFAULT_BCC') and Config.EMAIL_DEFAULT_BCC:
                bcc_emails = list(set(bcc_emails + Config.EMAIL_DEFAULT_BCC))
        else:
            bcc_emails = []
        
        # จัดการไฟล์แนบ
        attachments = None
        attachment_paths = []  # ประกาศไว้ก่อนเพื่อป้องกัน UnboundLocalError
        attachment_data = data.get('attachments', [])  # รองรับรูปแบบใหม่ (array of objects)
        attachment_paths_old = data.get('attachment_paths', [])  # รองรับรูปแบบเก่า (backward compatibility)
        
        if attachment_data or attachment_paths_old:
            attachments = []
            
            # รองรับรูปแบบใหม่ (array of objects)
            if attachment_data:
                for att_data in attachment_data:
                    if isinstance(att_data, dict):
                        temp_path = att_data.get('temp_path')
                        original_filename = att_data.get('original_filename') or att_data.get('filename')
                    else:
                        temp_path = att_data
                        original_filename = None
                    
                    if temp_path:
                        file_path = Path(temp_path)
                        if file_path.exists():
                            attachments.append({
                                'path': file_path,
                                'filename': original_filename or file_path.name
                            })
                            attachment_paths.append(temp_path)
            
            # รองรับรูปแบบเก่า (backward compatibility)
            elif attachment_paths_old:
                for temp_path in attachment_paths_old:
                    file_path = Path(temp_path)
                    if file_path.exists():
                        attachments.append(file_path)  # ใช้รูปแบบเก่า
                        attachment_paths.append(temp_path)
        
        # ดึง signature_name ถ้ามี
        signature_name = data.get('signature_name')
        
        # ดึง zip_attachments (default = True เพื่อ backward compatibility)
        zip_attachments = data.get('zip_attachments', True)
        
        # สร้าง PDF สรุป (ถ้ามีข้อมูล)
        summary_data = data.get('summary_data')
        pdf_path = None
        image_path = None
        if summary_data:
            try:
                pdf_generator = PDFGenerator()
                # ใช้ชื่อไฟล์ที่ส่งมาจาก frontend (ถ้ามี)
                pdf_filename = summary_data.get('filename')
                logger.info(f"📄 สร้าง PDF: filename={pdf_filename}")
                pdf_path = pdf_generator.generate_summary_pdf(
                    title=summary_data.get('title', 'สรุปข้อมูล'),
                    company_name=summary_data.get('company_name', ''),
                    period=summary_data.get('period', ''),
                    data=summary_data.get('data', []),
                    filename=pdf_filename,  # ส่งชื่อไฟล์ไปด้วย
                    tax_id=summary_data.get('tax_id', '')  # ส่ง tax_id ไปด้วย
                )
                
                # เพิ่ม PDF เป็น attachment (ไม่ต้องแปลงเป็นรูปภาพแล้ว)
                if pdf_path and pdf_path.exists():
                    if not attachments:
                        attachments = []
                    # ตั้งชื่อไฟล์ตามที่ส่งมาจาก frontend หรือใช้ชื่อ default
                    pdf_attachment_filename = pdf_filename if pdf_filename else 'สรุปภาษี.pdf'
                    logger.info(f"📎 Attach PDF: pdf_filename={pdf_filename}, final_name={pdf_attachment_filename}")
                    attachments.append({
                        'path': pdf_path,
                        'filename': pdf_attachment_filename
                    })
                    
                    # คัดลอก PDF ไปที่ V:\web\line-images ด้วย
                    try:
                        import shutil
                        line_images_dir = Path(r"V:\web\line-images")
                        line_images_dir.mkdir(parents=True, exist_ok=True)
                        line_images_path = line_images_dir / pdf_attachment_filename
                        shutil.copy2(pdf_path, line_images_path)
                        logger.info(f"✅ คัดลอก PDF ไปที่ {line_images_path}")
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถคัดลอก PDF ไปที่ V:\\web\\line-images: {e}")
            except Exception as e:
                logger.error(f"เกิดข้อผิดพลาดในการสร้าง PDF: {e}", exc_info=True)
        
        # กรองโลโก้ออกจาก attachments (ถ้ามี)
        if attachments:
            filtered_attachments = []
            logo_paths_to_skip = []
            
            # เก็บ path ของ signature logo (ถ้ามี)
            if signature_name:
                signature = email_service.get_signature(signature_name)
                if signature and signature.logo_path:
                    sig_logo_path = Path(signature.logo_path)
                    if sig_logo_path.exists():
                        logo_paths_to_skip.append(sig_logo_path.resolve())
            
            # เก็บ path ของ pattern logo (ถ้ามี)
            if pattern_name:
                pattern = email_service.get_pattern(pattern_name)
                if pattern and pattern.logo_path:
                    pattern_logo_path = Path(pattern.logo_path)
                    if pattern_logo_path.exists():
                        logo_paths_to_skip.append(pattern_logo_path.resolve())
            
            # กรอง attachments ที่เป็นโลโก้
            for att in attachments:
                if isinstance(att, dict):
                    att_path = att.get('path')
                else:
                    att_path = att
                
                if att_path:
                    att_path_obj = Path(att_path) if not isinstance(att_path, Path) else att_path
                    if att_path_obj.exists():
                        att_resolved = att_path_obj.resolve()
                        # ตรวจสอบว่าเป็นโลโก้หรือไม่
                        is_logo = False
                        for logo_path_resolved in logo_paths_to_skip:
                            if att_resolved == logo_path_resolved:
                                is_logo = True
                                break
                        
                        # ตรวจสอบชื่อไฟล์ว่าเป็นโลโก้หรือไม่ (fallback)
                        if not is_logo:
                            att_filename_lower = att_path_obj.name.lower()
                            if any(keyword in att_filename_lower for keyword in ['logo', 'signature_logo', 'pattern_logo']):
                                # ตรวจสอบว่าไฟล์อยู่ในโฟลเดอร์ email_logos หรือไม่
                                if 'email_logos' in str(att_resolved) or 'email_system' in str(att_resolved):
                                    is_logo = True
                        
                        if not is_logo:
                            filtered_attachments.append(att)
            
            attachments = filtered_attachments if filtered_attachments else None
        
        # ดึง subject และ body จาก request (ถ้ามี - จะ override pattern)
        subject_override = data.get('subject')
        body_override = data.get('body')
        
        # ดึง subject และ body ที่จะใช้จริง (จาก override หรือจาก pattern)
        final_subject = subject_override
        final_body = body_override
        
        # ถ้าไม่มี override ให้ดึงจาก pattern
        if not final_subject or not final_body:
            pattern = email_service.get_pattern(pattern_name)
            if pattern:
                if not final_subject:
                    final_subject = pattern.subject_template
                if not final_body:
                    final_body = pattern.body_template
        
        success, message = email_service.send_email_by_pattern(
            pattern_name=pattern_name,
            data=pattern_data,
            to_emails=to_emails,
            cc_emails=cc_emails,
            bcc_emails=bcc_emails,
            attachments=attachments,
            signature_name=signature_name,
            subject_override=subject_override,  # เพิ่ม subject override
            body_override=body_override,  # เพิ่ม body override
            zip_attachments=zip_attachments
        )
        
        # ⚠️ ยังไม่ลบไฟล์ชั่วคราว - รอส่ง Pay-in เสร็จก่อน!
        
        if success:
            # บันทึกประวัติการส่งอีเมล (มี pattern_name)
            save_email_history(
                pattern_name=pattern_name,
                signature_name=signature_name,
                to_emails=to_emails if to_emails else None,
                cc_emails=cc_emails if cc_emails else None,
                subject=final_subject,
                body=final_body
            )
            
            response_data = {'success': True, 'message': message}
            # ส่ง pdf_path กลับไปให้ frontend สำหรับส่ง LINE
            if pdf_path and pdf_path.exists():
                response_data['pdf_path'] = pdf_path.name  # ส่งแค่ชื่อไฟล์
                logger.info(f"✅ ส่ง pdf_path กลับไป: {pdf_path.name}")
            
            # ✅ ลบไฟล์ชั่วคราว (ยกเว้น PDF summary)
            if attachment_paths:
                for temp_path in attachment_paths:
                    try:
                        file_path = Path(temp_path)
                        
                        # ไม่ลบ PDF ที่สร้างจาก summary
                        if pdf_path and file_path.exists() and file_path.resolve() == pdf_path.resolve():
                            continue
                        
                        # ลบไฟล์ที่เหลือ
                        if file_path.exists() and 'temp_uploads' in str(file_path):
                            file_path.unlink()
                            logger.debug(f"🗑️ ลบไฟล์ชั่วคราว: {file_path.name}")
                    except Exception as e:
                        logger.warning(f"ไม่สามารถลบไฟล์ชั่วคราวได้: {temp_path} - {e}")
            
            return jsonify(response_data), 200
        else:
            return jsonify({'success': False, 'message': message}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/history', methods=['GET'])
def get_email_history():
    """ดึงประวัติการส่งอีเมล"""
    try:
        email_system_dir = Path('email_system')
        history_file = email_system_dir / 'email_history.json'
        
        if not history_file.exists():
            return jsonify({'success': True, 'history': []}), 200
        
        # โหลดประวัติ
        try:
            with open(history_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
        except Exception as e:
            logger.error(f"ไม่สามารถโหลด email history ได้: {e}")
            return jsonify({'success': True, 'history': []}), 200
        
        # เรียงลำดับจากใหม่ไปเก่า (reverse)
        history.reverse()
        
        return jsonify({'success': True, 'history': history}), 200
    except Exception as e:
        logger.error(f"เกิดข้อผิดพลาดในการดึงประวัติ: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/email/history/export', methods=['GET'])
def export_email_history():
    """ส่งออกประวัติการส่งอีเมลเป็นไฟล์ Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import tempfile
        
        email_service = get_global_email_service()
        if not email_service:
            return jsonify({'success': False, 'message': 'ยังไม่ได้ตั้งค่า SMTP'}), 400
        
        # อ่านประวัติจากไฟล์
        email_system_dir = Path('email_system')
        history_file = email_system_dir / 'email_history.json'
        
        if not history_file.exists():
            return jsonify({'success': False, 'message': 'ไม่พบประวัติการส่งอีเมล'}), 404
        
        # โหลดประวัติ
        try:
            with open(history_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
        except Exception as e:
            logger.error(f"ไม่สามารถโหลด email history ได้: {e}")
            return jsonify({'success': False, 'message': f'ไม่สามารถโหลดประวัติได้: {e}'}), 500
        
        if not history:
            return jsonify({'success': False, 'message': 'ไม่มีประวัติการส่งอีเมล'}), 404
        
        # สร้าง Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "ประวัติการส่งอีเมล"
        
        # สร้าง Header
        headers = [
            "ลำดับ",
            "วันที่ส่ง",
            "เวลาที่ส่ง",
            "ชื่อ Pattern",
            "รายละเอียด Pattern",
            "Company Name (Pattern)",
            "Tax ID (Pattern)",
            "LINE User ID (Pattern)",
            "ชื่อ Signature",
            "ชื่อผู้ส่ง (Signature)",
            "ตำแหน่ง (Signature)",
            "อีเมลล์ผู้ส่ง (Signature)",
            "เบอร์โทรศัพท์ (Signature)",
            "เว็บไซต์ (Signature)",
            "ชื่อบริษัท (Signature)",
            "ที่อยู่บริษัท (Signature)",
            "To Emails",
            "CC Emails",
            "Subject",
            "Body",
            "LINE Message"
        ]
        
        ws.append(headers)
        
        # Format header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # เพิ่มข้อมูล
        for idx, entry in enumerate(history, start=1):
            pattern_name = entry.get('pattern_name')
            signature_name = entry.get('signature_name')
            sent_time = entry.get('sent_time', '')
            
            # แยกวันที่และเวลา
            date_part = ''
            time_part = ''
            if sent_time:
                try:
                    dt = datetime.strptime(sent_time, '%Y-%m-%d %H:%M:%S')
                    date_part = dt.strftime('%Y-%m-%d')
                    time_part = dt.strftime('%H:%M:%S')
                except:
                    date_part = sent_time
                    time_part = ''
            
            # ดึงข้อมูล Pattern
            pattern_description = ''
            pattern_company_name = ''
            pattern_tax_id = ''
            pattern_line_user_id = ''
            
            if pattern_name:
                pattern = email_service.get_pattern(pattern_name)
                if pattern:
                    pattern_description = pattern.description or ''
                    pattern_company_name = pattern.company_name or ''
                    pattern_tax_id = pattern.tax_id or ''
                    pattern_line_user_id = pattern.line_user_id or ''
            
            # ดึงข้อมูล Signature
            sig_sender_name = ''
            sig_sender_title = ''
            sig_sender_email = ''
            sig_sender_phone = ''
            sig_sender_website = ''
            sig_company_name = ''
            sig_company_address = ''
            
            if signature_name:
                signature = email_service.get_signature(signature_name)
                if signature:
                    sig_sender_name = signature.sender_name or ''
                    sig_sender_title = signature.sender_title or ''
                    sig_sender_email = signature.sender_email or ''
                    sig_sender_phone = signature.sender_phone or ''
                    sig_sender_website = signature.sender_website or ''
                    sig_company_name = signature.company_name or ''
                    sig_company_address = signature.company_address or ''
            
            # ดึงข้อมูลเพิ่มเติมจาก entry
            to_emails = entry.get('to_emails', [])
            cc_emails = entry.get('cc_emails', [])
            subject = entry.get('subject', '')
            body = entry.get('body', '')
            line_message = entry.get('line_message', '')
            
            # เพิ่มแถวข้อมูล
            row = [
                idx,
                date_part,
                time_part,
                pattern_name or '',
                pattern_description,
                pattern_company_name,
                pattern_tax_id,
                pattern_line_user_id,
                signature_name or '',
                sig_sender_name,
                sig_sender_title,
                sig_sender_email,
                sig_sender_phone,
                sig_sender_website,
                sig_company_name,
                sig_company_address,
                ', '.join(to_emails) if to_emails else '',
                ', '.join(cc_emails) if cc_emails else '',
                subject,
                body,
                line_message
            ]
            
            ws.append(row)
            
            # Format แถวข้อมูล
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # ปรับความกว้างคอลัมน์
        column_widths = {
            'A': 8,   # ลำดับ
            'Q': 30,  # To Emails
            'R': 30,  # CC Emails
            'S': 40,  # Subject
            'T': 60,  # Body
            'U': 60,  # LINE Message
            'B': 12,  # วันที่ส่ง
            'C': 10,  # เวลาที่ส่ง
            'D': 20,  # ชื่อ Pattern
            'E': 30,  # รายละเอียด Pattern
            'F': 25,  # Company Name (Pattern)
            'G': 15,  # Tax ID (Pattern)
            'H': 20,  # LINE User ID (Pattern)
            'I': 20,  # ชื่อ Signature
            'J': 20,  # ชื่อผู้ส่ง (Signature)
            'K': 20,  # ตำแหน่ง (Signature)
            'L': 25,  # อีเมลล์ผู้ส่ง (Signature)
            'M': 15,  # เบอร์โทรศัพท์ (Signature)
            'N': 25,  # เว็บไซต์ (Signature)
            'O': 25,  # ชื่อบริษัท (Signature)
            'P': 40   # ที่อยู่บริษัท (Signature)
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # สร้างไฟล์ชั่วคราว
        temp_dir = Path(tempfile.gettempdir())
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f'email_history_{timestamp}.xlsx'
        excel_path = temp_dir / excel_filename
        
        wb.save(excel_path)
        
        logger.info(f"✅ สร้างไฟล์ Excel ประวัติการส่งอีเมล: {excel_path}")
        
        # ส่งไฟล์กลับ
        return send_file(
            str(excel_path),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=excel_filename
        )
        
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการส่งออกประวัติ: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

# ---------- Admin Configuration Management Routes ----------

def _extract_folder_code(folder_path: str) -> Optional[str]:
    """แยกรหัสโฟลเดอร์จาก path (เช่น Build495 จาก Build495 บริษัท...)"""
    try:
        path_obj = Path(folder_path)
        for part in path_obj.parts:
            part_str = str(part)
            # กรณี 1: Build495 บริษัท เรบิน่า... -> แยกเป็น Build495
            if part_str.startswith('Build') and ' ' in part_str:
                build_part = part_str.split(' ')[0]  # ได้ "Build495"
                if len(build_part) >= 7:  # Build495 = 7 ตัวอักษรขึ้นไป
                    code_part = build_part[5:]  # ได้ "495"
                    if code_part.isdigit():
                        return f"Build{code_part}"
            # กรณี 2: Build495, Build000 (ไม่มีช่องว่าง)
            elif part_str.startswith('Build') and len(part_str) >= 7:
                code_part = part_str[5:]  # ได้ "495", "000"
                if code_part.isdigit():
                    return f"Build{code_part}"
            # กรณี 3: 495, 000 (3 หลักขึ้นไป)
            elif len(part_str) >= 3 and part_str.isdigit():
                return f"Build{part_str}"
        return None
    except Exception as e:
        print(f"❌ [Extract Folder Code] Error: {e}")
        return None

def _get_login_txt_path(folder_code: str = None):
    """หาพาธไฟล์ txt สำหรับล็อกอิน"""
    if not folder_code:
        folder_code = "Build000"  # default
    
    base_path = Path("V:/A.โฟร์เดอร์หลัก/Build000 ทดสอบระบบ/รหัส")
    txt_path = base_path / f"{folder_code}.txt"
    
    return txt_path

def _get_chart_json_path(folder_code: str = None):
    """หาพาธไฟล์ JSON สำหรับผังบัญชี"""
    if not folder_code:
        folder_code = "Build000"  # default
    
    base_path = Path("V:/A.โฟร์เดอร์หลัก/Build000 ทดสอบระบบ/รหัส")
    json_path = base_path / f"{folder_code}.json"
    
    return json_path

def _get_folder_settings_path():
    """หาพาธไฟล์ folder_settings.json"""
    return Path("V:/A.โฟร์เดอร์หลัก/Build000 ทดสอบระบบ/folder_settings/folder_settings.json")

@app.route('/api/admin/login-info', methods=['GET'])
def get_login_info():
    """ดึงข้อมูลล็อกอินปัจจุบัน (เปิดให้ผู้ใช้ทั่วไป)"""
    
    # รับ folder_path จาก query parameter
    folder_path = request.args.get('folder_path', '').strip()
    folder_code = None
    
    if folder_path:
        folder_code = _extract_folder_code(folder_path)
        if not folder_code:
            return jsonify({'error': f'ไม่สามารถแยกรหัสโฟลเดอร์จาก path: {folder_path}'}), 400
    
    try:
        txt_path = _get_login_txt_path(folder_code)
        if not txt_path.exists():
            return jsonify({'error': f'ไม่พบไฟล์: {txt_path}'}), 404
        
        credentials = {}
        with open(txt_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if ':' in line:
                    # รองรับทั้ง "key : value" และ "key:value"
                    parts = line.split(':', 1)
                    if len(parts) == 2:
                        key = parts[0].strip()
                        value = parts[1].strip()
                        credentials[key] = value
        
        return jsonify({
            'username': credentials.get('Username', ''),
            'password': credentials.get('Password', ''),
            'company_link': credentials.get('Link company', ''),
            'express_link': credentials.get('Link Express', ''),
            'file_path': str(txt_path),
            'folder_code': folder_code or 'Build000'
        }), 200
    except Exception as e:
        print(f"❌ [Get Login Info] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/admin/update-login', methods=['POST'])
def update_login():
    """อัพเดตข้อมูลล็อกอินในไฟล์ txt (เปิดให้ผู้ใช้ทั่วไป)"""
    
    data = request.json
    new_username = data.get('username', '').strip()
    new_password = data.get('password', '').strip()
    folder_path = data.get('folder_path', '').strip()
    
    # แยกรหัสโฟลเดอร์จาก path
    folder_code = None
    if folder_path:
        folder_code = _extract_folder_code(folder_path)
        if not folder_code:
            return jsonify({'error': f'ไม่สามารถแยกรหัสโฟลเดอร์จาก path: {folder_path}'}), 400
    
    if not new_username and not new_password:
        return jsonify({'error': 'กรุณาระบุข้อมูลที่ต้องการแก้ไข'}), 400
    
    try:
        txt_path = _get_login_txt_path(folder_code)
        
        # อ่านไฟล์เดิม
        lines = []
        credentials = {}
        if txt_path.exists():
            with open(txt_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                for line in lines:
                    line_stripped = line.strip()
                    if ':' in line_stripped:
                        parts = line_stripped.split(':', 1)
                        if len(parts) == 2:
                            key = parts[0].strip()
                            value = parts[1].strip()
                            credentials[key] = value
        
        # อัพเดตข้อมูล
        if new_username:
            credentials['Username'] = new_username
        if new_password:
            credentials['Password'] = new_password
        
        # สร้างโฟลเดอร์ถ้ายังไม่มี
        txt_path.parent.mkdir(parents=True, exist_ok=True)
        
        # เขียนไฟล์ใหม่
        with open(txt_path, 'w', encoding='utf-8') as f:
            f.write(f"Username : {credentials.get('Username', '')}\n")
            f.write(f"Password : {credentials.get('Password', '')}\n")
            if 'Link company' in credentials:
                f.write(f"Link company : {credentials.get('Link company', '')}\n")
            if 'Link Express' in credentials:
                f.write(f"Link Express : {credentials.get('Link Express', '')}\n")
        
        messages = []
        if new_username:
            messages.append(f'ชื่อผู้ใช้ถูกเปลี่ยนเป็น: {new_username}')
        if new_password:
            messages.append('รหัสผ่านถูกเปลี่ยน')
        
        return jsonify({
            'success': True,
            'message': ' '.join(messages) + ' สำเร็จ',
            'folder_code': folder_code or 'Build000'
        }), 200
    except Exception as e:
        print(f"❌ [Update Login] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/admin/chart-of-accounts', methods=['GET'])
def get_chart_of_accounts():
    """ดึงข้อมูลผังบัญชีจากไฟล์ JSON (เปิดให้ผู้ใช้ทั่วไป)"""
    
    group = request.args.get('group', 'GROUP1')
    folder_path = request.args.get('folder_path', '').strip()
    
    # แยกรหัสโฟลเดอร์จาก path
    folder_code = None
    if folder_path:
        folder_code = _extract_folder_code(folder_path)
        if not folder_code:
            return jsonify({'error': f'ไม่สามารถแยกรหัสโฟลเดอร์จาก path: {folder_path}'}), 400
    
    try:
        json_path = _get_chart_json_path(folder_code)
        
        if not json_path.exists():
            return jsonify({'error': f'ไม่พบไฟล์: {json_path}'}), 404
        
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        # แปลงข้อมูลจาก JSON เป็นรูปแบบที่ต้องการ
        # JSON format: {"ค่าบริการ Shopee VAT": {"company_name": "...", "customer_id": "...", "account_code": "..."}, ...}
        items = []
        for key, value in json_data.items():
            if isinstance(value, dict):
                items.append({
                    'account_name': key,  # ชื่อผังบัญชีคือ key ของ JSON
                    'company_name': value.get('company_name', ''),
                    'customer_id': value.get('customer_id', ''),
                    'account_code': value.get('account_code', ''),
                    'account_code2': value.get('account_code2', '')
                })
        
        return jsonify({
            'items': items,
            'file_path': str(json_path),
            'folder_code': folder_code or 'Build000'
        }), 200
    except json.JSONDecodeError as e:
        return jsonify({'error': f'ไฟล์ JSON มีปัญหา: {e}'}), 500
    except Exception as e:
        print(f"❌ [Get Chart] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/admin/chart-of-accounts', methods=['POST'])
def add_chart_of_accounts():
    """เพิ่มผังบัญชีในไฟล์ JSON (เปิดให้ผู้ใช้ทั่วไป)"""
    
    data = request.json
    account_name = data.get('account_name', '').strip()
    customer_id = data.get('customer_id', '').strip()
    account_code = data.get('account_code', '').strip()
    group = data.get('group', 'GROUP1')
    folder_path = data.get('folder_path', '').strip()
    
    if not account_name:
        return jsonify({'error': 'กรุณาระบุชื่อผังบัญชี'}), 400
    
    # แยกรหัสโฟลเดอร์จาก path
    folder_code = None
    if folder_path:
        folder_code = _extract_folder_code(folder_path)
        if not folder_code:
            return jsonify({'error': f'ไม่สามารถแยกรหัสโฟลเดอร์จาก path: {folder_path}'}), 400
    
    try:
        json_path = _get_chart_json_path(folder_code)
        
        # อ่านไฟล์เดิม
        json_data = {}
        if json_path.exists():
            with open(json_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
        
        # ตรวจสอบว่ามี account_name นี้อยู่แล้วหรือไม่
        if account_name in json_data:
            return jsonify({'error': f'ผังบัญชี "{account_name}" มีอยู่แล้ว'}), 400
        
        # เพิ่มข้อมูลใหม่
        json_data[account_name] = {
            'company_name': '',  # จะอัพเดตทีหลังถ้าต้องการ
            'customer_id': customer_id,
            'account_code': account_code,
            'account_code2': ''
        }
        
        # สร้างโฟลเดอร์ถ้ายังไม่มี
        json_path.parent.mkdir(parents=True, exist_ok=True)
        
        # เขียนไฟล์ใหม่
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True,
            'message': f'เพิ่มผังบัญชีสำเร็จ',
            'folder_code': folder_code or 'Build000'
        }), 200
    except Exception as e:
        print(f"❌ [Add Chart] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/admin/chart-of-accounts', methods=['PUT'])
def update_chart_of_accounts():
    """อัพเดตผังบัญชีในไฟล์ JSON (เปิดให้ผู้ใช้ทั่วไป)"""
    
    data = request.json
    account_name = data.get('account_name', '').strip()
    customer_id = data.get('customer_id', '').strip()
    account_code = data.get('account_code', '').strip()
    group = data.get('group', 'GROUP1')
    folder_path = data.get('folder_path', '').strip()
    
    if not account_name:
        return jsonify({'error': 'กรุณาระบุชื่อผังบัญชี'}), 400
    
    # แยกรหัสโฟลเดอร์จาก path
    folder_code = None
    if folder_path:
        folder_code = _extract_folder_code(folder_path)
        if not folder_code:
            return jsonify({'error': f'ไม่สามารถแยกรหัสโฟลเดอร์จาก path: {folder_path}'}), 400
    
    try:
        json_path = _get_chart_json_path(folder_code)
        
        if not json_path.exists():
            return jsonify({'error': f'ไม่พบไฟล์: {json_path}'}), 404
        
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        # ตรวจสอบว่ามี account_name นี้อยู่หรือไม่
        if account_name not in json_data:
            return jsonify({'error': f'ไม่พบผังบัญชี: {account_name}'}), 404
        
        # อัพเดตข้อมูล (อัพเดตเฉพาะ customer_id และ account_code)
        if isinstance(json_data[account_name], dict):
            if customer_id:
                json_data[account_name]['customer_id'] = customer_id
            if account_code:
                json_data[account_name]['account_code'] = account_code
        else:
            # ถ้าเป็นรูปแบบเก่า ให้แปลงเป็น dict
            json_data[account_name] = {
                'company_name': '',
                'customer_id': customer_id,
                'account_code': account_code,
                'account_code2': ''
            }
        
        # เขียนไฟล์ใหม่
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True,
            'message': f'อัพเดตผังบัญชี "{account_name}" สำเร็จ',
            'folder_code': folder_code or 'Build000'
        }), 200
    except Exception as e:
        print(f"❌ [Update Chart] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/admin/chart-of-accounts', methods=['DELETE'])
def delete_chart_of_accounts():
    """ลบผังบัญชีจากไฟล์ JSON (เปิดให้ผู้ใช้ทั่วไป)"""
    
    data = request.json
    account_name = data.get('account_name', '').strip()
    group = data.get('group', 'GROUP1')
    folder_path = data.get('folder_path', '').strip()
    
    if not account_name:
        return jsonify({'error': 'กรุณาระบุชื่อผังบัญชี'}), 400
    
    # แยกรหัสโฟลเดอร์จาก path
    folder_code = None
    if folder_path:
        folder_code = _extract_folder_code(folder_path)
        if not folder_code:
            return jsonify({'error': f'ไม่สามารถแยกรหัสโฟลเดอร์จาก path: {folder_path}'}), 400
    
    try:
        json_path = _get_chart_json_path(folder_code)
        
        if not json_path.exists():
            return jsonify({'error': f'ไม่พบไฟล์: {json_path}'}), 404
        
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        # ตรวจสอบว่ามี account_name นี้อยู่หรือไม่
        if account_name not in json_data:
            return jsonify({'error': f'ไม่พบผังบัญชี: {account_name}'}), 404
        
        # ลบข้อมูล
        del json_data[account_name]
        
        # เขียนไฟล์ใหม่
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True,
            'message': f'ลบผังบัญชีสำเร็จ',
            'folder_code': folder_code or 'Build000'
        }), 200
    except Exception as e:
        print(f"❌ [Delete Chart] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

def _update_config_py_company_vat_status(company: str, vat_type: str = None, action: str = 'update'):
    """อัพเดต COMPANY_VAT_STATUS ใน config.py"""
    config_path = current_dir / 'config.py'
    
    if not config_path.exists():
        return False, f'ไม่พบไฟล์: {config_path}'
    
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # หา COMPANY_VAT_STATUS section
        start_idx = None
        end_idx = None
        brace_count = 0
        
        for i, line in enumerate(lines):
            if 'COMPANY_VAT_STATUS' in line and '=' in line:
                start_idx = i
                # นับ opening brace
                brace_count = line.count('{') - line.count('}')
                if brace_count > 0:
                    # ตรวจสอบว่ามี closing brace ในบรรทัดเดียวกันหรือไม่
                    if '}' in line:
                        end_idx = i
                        break
                    continue
            elif start_idx is not None:
                brace_count += line.count('{') - line.count('}')
                if brace_count == 0:
                    end_idx = i
                    break
        
        if start_idx is None:
            return False, 'ไม่พบ COMPANY_VAT_STATUS ใน config.py'
        
        if end_idx is None:
            return False, 'ไม่พบ closing brace ของ COMPANY_VAT_STATUS'
        
        # อ่าน body section
        body_lines = lines[start_idx:end_idx+1]
        body_text = ''.join(body_lines)
        
        # แปลง body เป็น dict
        items = {}
        for line in body_lines[1:-1]:  # ข้ามบรรทัดแรกและสุดท้าย
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if ':' in line:
                # แยก key และ value - รองรับทั้ง "key": "value" และ 'key': 'value'
                import re
                match = re.match(r'["\']?([^"\']+)["\']?\s*:\s*["\']?([^"\']+)["\']?', line)
                if match:
                    key = match.group(1).strip()
                    value = match.group(2).strip().rstrip(',')
                    if key:
                        items[key] = value
        
        # อัพเดต items
        if action == 'add' or action == 'update':
            items[company] = vat_type
        elif action == 'delete':
            if company in items:
                del items[company]
            else:
                return False, f'ไม่พบบริษัท: {company}'
        
        # สร้าง body ใหม่
        new_body_lines = [lines[start_idx]]  # บรรทัดแรก (COMPANY_VAT_STATUS = {)
        for key, value in sorted(items.items()):
            new_body_lines.append(f'        "{key}": "{value}",\n')
        new_body_lines.append('    }\n')  # closing brace
        
        # แทนที่ส่วนเดิม
        new_lines = lines[:start_idx] + new_body_lines + lines[end_idx+1:]
        
        # เขียนไฟล์ใหม่
        with open(config_path, 'w', encoding='utf-8') as f:
            f.writelines(new_lines)
        
        return True, 'อัพเดต config.py สำเร็จ'
    except Exception as e:
        print(f"❌ [Update Config] Error: {e}")
        import traceback
        traceback.print_exc()
        return False, f'เกิดข้อผิดพลาด: {e}'

@app.route('/api/admin/company-types', methods=['GET'])
def get_company_types():
    """ดึงข้อมูลประเภทบริษัทจาก folder_settings.json (เปิดให้ผู้ใช้ทั่วไป)"""
    
    folder_path = request.args.get('folder_path', '').strip()
    
    # แยกรหัสโฟลเดอร์จาก path
    folder_code = None
    if folder_path:
        folder_code = _extract_folder_code(folder_path)
        if not folder_code:
            return jsonify({'error': f'ไม่สามารถแยกรหัสโฟลเดอร์จาก path: {folder_path}'}), 400
    
    try:
        settings_path = _get_folder_settings_path()
        
        if not settings_path.exists():
            return jsonify({'error': f'ไม่พบไฟล์: {settings_path}'}), 404
        
        with open(settings_path, 'r', encoding='utf-8') as f:
            folder_settings = json.load(f)
        
        # ถ้ามี folder_code ให้ดึงข้อมูลเฉพาะ folder_code นั้น
        if folder_code and folder_code in folder_settings:
            folder_info = folder_settings[folder_code]
            group = folder_info.get('group', 'unknown')
            # แปลง group เป็น VAT/NoneVat
            # regular = VAT (จดภาษีมูลค่าเพิ่ม)
            # special = NoneVat (ไม่จดภาษีมูลค่าเพิ่ม)
            vat_status = 'VAT' if group == 'regular' else 'NoneVat' if group == 'special' else 'Unknown'
            return jsonify({
                'items': {folder_code: vat_status},
                'file_path': str(settings_path),
                'folder_code': folder_code,
                'group': group
            }), 200
        elif folder_code:
            # ถ้ามี folder_code แต่ไม่พบใน settings
            return jsonify({
                'items': {},
                'file_path': str(settings_path),
                'folder_code': folder_code,
                'message': f'ไม่พบข้อมูลสำหรับ {folder_code} ใน folder_settings.json'
            }), 200
        else:
            # ถ้าไม่มี folder_code ให้แสดงทั้งหมด
            items = {}
            for code, info in folder_settings.items():
                if isinstance(info, dict):
                    group = info.get('group', 'unknown')
                    vat_status = 'VAT' if group == 'regular' else 'NoneVat' if group == 'special' else 'Unknown'
                    items[code] = vat_status
            
            return jsonify({
                'items': items,
                'file_path': str(settings_path)
            }), 200
    except Exception as e:
        print(f"❌ [Get Company Types] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/admin/company-types', methods=['POST'])
def add_company_type():
    """เพิ่มประเภทบริษัทใน folder_settings.json (เปิดให้ผู้ใช้ทั่วไป)"""
    
    data = request.json
    folder_path = data.get('folder_path', '').strip()
    vat_type = data.get('type', '').strip()
    
    # แยกรหัสโฟลเดอร์จาก path
    folder_code = None
    if folder_path:
        folder_code = _extract_folder_code(folder_path)
        if not folder_code:
            return jsonify({'error': f'ไม่สามารถแยกรหัสโฟลเดอร์จาก path: {folder_path}'}), 400
    else:
        return jsonify({'error': 'กรุณาระบุ folder_path'}), 400
    
    if not vat_type:
        return jsonify({'error': 'กรุณาระบุประเภท'}), 400
    
    if vat_type not in ['VAT', 'NoneVat']:
        return jsonify({'error': 'ประเภทต้องเป็น VAT (จดภาษีมูลค่าเพิ่ม) หรือ NoneVat (ไม่จดภาษีมูลค่าเพิ่ม)'}), 400
    
    try:
        settings_path = _get_folder_settings_path()
        
        # อ่านไฟล์เดิม
        folder_settings = {}
        if settings_path.exists():
            with open(settings_path, 'r', encoding='utf-8') as f:
                folder_settings = json.load(f)
        
        # แปลง VAT/NoneVat เป็น group
        # VAT = regular (จดภาษีมูลค่าเพิ่ม)
        # NoneVat = special (ไม่จดภาษีมูลค่าเพิ่ม)
        group = 'regular' if vat_type == 'VAT' else 'special'
        
        # เพิ่มหรืออัพเดตข้อมูล
        if folder_code not in folder_settings:
            folder_settings[folder_code] = {}
        
        folder_settings[folder_code]['group'] = group
        if 'message' not in folder_settings[folder_code]:
            folder_settings[folder_code]['message'] = f'บริษัท{"จด" if vat_type == "VAT" else "ไม่จด"}ภาษีมูลค่าเพิ่ม'
        
        # สร้างโฟลเดอร์ถ้ายังไม่มี
        settings_path.parent.mkdir(parents=True, exist_ok=True)
        
        # เขียนไฟล์ใหม่
        with open(settings_path, 'w', encoding='utf-8') as f:
            json.dump(folder_settings, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True,
            'message': f'เพิ่มประเภทบริษัทสำเร็จ ({vat_type})',
            'folder_code': folder_code
        }), 200
    except Exception as e:
        print(f"❌ [Add Company Type] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/admin/company-types', methods=['PUT'])
def update_company_type():
    """อัพเดตประเภทบริษัทใน folder_settings.json (เปิดให้ผู้ใช้ทั่วไป)"""
    
    data = request.json
    folder_path = data.get('folder_path', '').strip()
    vat_type = data.get('type', '').strip()
    
    # แยกรหัสโฟลเดอร์จาก path
    folder_code = None
    if folder_path:
        folder_code = _extract_folder_code(folder_path)
        if not folder_code:
            return jsonify({'error': f'ไม่สามารถแยกรหัสโฟลเดอร์จาก path: {folder_path}'}), 400
    else:
        return jsonify({'error': 'กรุณาระบุ folder_path'}), 400
    
    if not vat_type:
        return jsonify({'error': 'กรุณาระบุประเภท'}), 400
    
    if vat_type not in ['VAT', 'NoneVat']:
        return jsonify({'error': 'ประเภทต้องเป็น VAT (จดภาษีมูลค่าเพิ่ม) หรือ NoneVat (ไม่จดภาษีมูลค่าเพิ่ม)'}), 400
    
    try:
        settings_path = _get_folder_settings_path()
        
        if not settings_path.exists():
            return jsonify({'error': f'ไม่พบไฟล์: {settings_path}'}), 404
        
        with open(settings_path, 'r', encoding='utf-8') as f:
            folder_settings = json.load(f)
        
        if folder_code not in folder_settings:
            return jsonify({'error': f'ไม่พบข้อมูลสำหรับ {folder_code}'}), 404
        
        # แปลง VAT/NoneVat เป็น group
        group = 'regular' if vat_type == 'VAT' else 'special'
        
        # อัพเดตข้อมูล
        folder_settings[folder_code]['group'] = group
        folder_settings[folder_code]['message'] = f'บริษัท{"จด" if vat_type == "VAT" else "ไม่จด"}ภาษีมูลค่าเพิ่ม'
        
        # เขียนไฟล์ใหม่
        with open(settings_path, 'w', encoding='utf-8') as f:
            json.dump(folder_settings, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True,
            'message': f'อัพเดตประเภทบริษัทสำเร็จ ({vat_type})',
            'folder_code': folder_code
        }), 200
    except Exception as e:
        print(f"❌ [Update Company Type] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/admin/company-types', methods=['DELETE'])
def delete_company_type():
    """ลบประเภทบริษัทจาก folder_settings.json (เปิดให้ผู้ใช้ทั่วไป)"""
    
    data = request.json
    folder_path = data.get('folder_path', '').strip()
    
    # แยกรหัสโฟลเดอร์จาก path
    folder_code = None
    if folder_path:
        folder_code = _extract_folder_code(folder_path)
        if not folder_code:
            return jsonify({'error': f'ไม่สามารถแยกรหัสโฟลเดอร์จาก path: {folder_path}'}), 400
    else:
        return jsonify({'error': 'กรุณาระบุ folder_path'}), 400
    
    try:
        settings_path = _get_folder_settings_path()
        
        if not settings_path.exists():
            return jsonify({'error': f'ไม่พบไฟล์: {settings_path}'}), 404
        
        with open(settings_path, 'r', encoding='utf-8') as f:
            folder_settings = json.load(f)
        
        if folder_code not in folder_settings:
            return jsonify({'error': f'ไม่พบข้อมูลสำหรับ {folder_code}'}), 404
        
        # ลบข้อมูล
        del folder_settings[folder_code]
        
        # เขียนไฟล์ใหม่
        with open(settings_path, 'w', encoding='utf-8') as f:
            json.dump(folder_settings, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True,
            'message': f'ลบประเภทบริษัทสำเร็จ',
            'folder_code': folder_code
        }), 200
    except Exception as e:
        print(f"❌ [Delete Company Type] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'เกิดข้อผิดพลาด: {e}'}), 500

# ==================== LINE API Endpoints ====================

@app.route('/api/line/settings', methods=['GET'])
def get_line_settings():
    """ดึงการตั้งค่า LINE"""
    try:
        enabled = get_line_notifications_enabled()
        default_user_id = getattr(Config, 'LINE_OA_DEFAULT_TO', '')
        return jsonify({
            'success': True,
            'enabled': enabled,
            'default_user_id': default_user_id
        })
    except Exception as e:
        logger.error(f"Error getting LINE settings: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/line/settings', methods=['POST'])
def update_line_settings():
    """อัปเดตการตั้งค่า LINE"""
    try:
        data = request.json
        enabled = data.get('enabled', True)
        set_line_notifications_enabled(enabled)
        
        # อัปเดต default_user_id ถ้ามี
        if 'default_user_id' in data:
            # เก็บใน session หรือ config (ถ้าต้องการ)
            # สำหรับตอนนี้ใช้ config โดยตรง
            if hasattr(Config, 'LINE_OA_DEFAULT_TO'):
                Config.LINE_OA_DEFAULT_TO = data['default_user_id']
        
        return jsonify({
            'success': True,
            'message': 'อัปเดตการตั้งค่า LINE สำเร็จ'
        })
    except Exception as e:
        logger.error(f"Error updating LINE settings: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/line/send', methods=['POST'])
def send_line_message():
    """ส่งข้อความไปยัง LINE Official Account"""
    try:
        data = request.json
        user_id = data.get('user_id', '')
        message = data.get('message', '')
        
        if not user_id:
            return jsonify({
                'success': False,
                'message': 'กรุณาระบุ LINE User ID / Group ID / Room ID'
            }), 400
        
        if not message:
            return jsonify({
                'success': False,
                'message': 'กรุณาระบุข้อความที่จะส่ง'
            }), 400
        
        # ตรวจสอบสถานะการแจ้งเตือน LINE
        if not get_line_notifications_enabled():
            return jsonify({
                'success': False,
                'message': 'การแจ้งเตือน LINE ถูกปิดอยู่'
            }), 400
        
        # ใช้ line_oa_push จาก report_manager
        success = line_oa_push(message, to=user_id)
        
        if success:
            # บันทึกประวัติ LINE (เพิ่ม line_message)
            save_email_history(
                line_message=message
            )
            return jsonify({
                'success': True,
                'message': f'ส่งข้อความ LINE สำเร็จไปยัง {user_id[:10]}...'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'ส่งข้อความ LINE ไม่สำเร็จ (ตรวจสอบ token และ user_id)'
            }), 500
            
    except Exception as e:
        logger.error(f"Error sending LINE message: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/line/notify', methods=['POST'])
def send_line_notify():
    """ส่งข้อความพร้อมรูปภาพไปยัง LINE Notify"""
    try:
        data = request.json
        message = data.get('message', '')
        image_path = data.get('image_path', '')  # path ของรูปภาพ (optional)
        
        if not message:
            return jsonify({
                'success': False,
                'message': 'กรุณาระบุข้อความที่จะส่ง'
            }), 400
        
        # ตรวจสอบสถานะการแจ้งเตือน LINE
        if not get_line_notifications_enabled():
            return jsonify({
                'success': False,
                'message': 'การแจ้งเตือน LINE ถูกปิดอยู่'
            }), 400
        
        # ตรวจสอบว่ามีรูปภาพหรือไม่
        image_file_path = None
        if image_path:
            full_image_path = Path('temp_uploads') / image_path
            if full_image_path.exists():
                image_file_path = str(full_image_path)
            else:
                logger.warning(f"ไม่พบไฟล์รูปภาพ: {full_image_path}")
        
        # ใช้ line_notify จาก report_manager (รองรับรูปภาพแล้ว)
        from report_manager import line_notify
        success = line_notify(message, image_path=image_file_path)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'ส่งข้อความ LINE Notify สำเร็จ' + (' (พร้อมรูปภาพ)' if image_file_path else '')
            })
        else:
            return jsonify({
                'success': False,
                'message': 'ส่งข้อความ LINE Notify ไม่สำเร็จ (ตรวจสอบ LINE_NOTIFY_TOKEN)'
            }), 500
            
    except Exception as e:
        logger.error(f"Error sending LINE Notify: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

def save_image_to_nas(image_path: str, filename: str = None) -> Optional[str]:
    """
    บันทึกรูปภาพไปยัง NAS และสร้าง HTTPS URL
    
    Args:
        image_path: Path ของไฟล์รูปภาพต้นฉบับ
        filename: ชื่อไฟล์ที่จะบันทึก (ถ้าไม่ระบุจะใช้ชื่อเดิมแต่เปลี่ยนเป็น .jpg)
    
    Returns:
        HTTPS URL ของรูปภาพ หรือ None ถ้าไม่สำเร็จ
    """
    try:
        from pathlib import Path
        from PIL import Image
        import shutil
        
        # โฟลเดอร์ปลายทางบน NAS
        nas_folder = Path('V:/web/line-images')
        
        # สร้างโฟลเดอร์ถ้ายังไม่มี
        nas_folder.mkdir(parents=True, exist_ok=True)
        
        # ตรวจสอบว่าโฟลเดอร์สร้างได้หรือไม่
        if not nas_folder.exists():
            logger.error(f"❌ ไม่สามารถสร้างโฟลเดอร์ได้: {nas_folder}")
            return None
        
        # แปลงเป็น Path object
        source_path = Path(image_path)
        
        if not source_path.exists():
            logger.error(f"❌ ไม่พบไฟล์รูปภาพ: {image_path}")
            return None
        
        # สร้างชื่อไฟล์ (แปลงเป็น .jpg)
        if filename:
            # ถ้ามีการระบุชื่อไฟล์ ให้ใช้ชื่อนั้นแต่เปลี่ยนเป็น .jpg
            output_filename = Path(filename).stem + '.jpg'
        else:
            # ถ้าไม่ระบุ ให้ใช้ชื่อเดิมแต่เปลี่ยนเป็น .jpg
            output_filename = source_path.stem + '.jpg'
        
        # Path ปลายทาง
        destination_path = nas_folder / output_filename
        
        # แปลงรูปภาพเป็น JPG และบันทึก
        try:
            # เปิดรูปภาพ
            img = Image.open(source_path)
            
            # แปลงเป็น RGB ถ้าเป็น RGBA หรือ mode อื่น
            if img.mode in ('RGBA', 'LA', 'P'):
                # สร้างพื้นหลังสีขาว
                rgb_img = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                rgb_img.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = rgb_img
            elif img.mode != 'RGB':
                img = img.convert('RGB')
            
            # บันทึกเป็น JPG
            img.save(destination_path, 'JPEG', quality=95, optimize=True)
            logger.info(f"✅ บันทึกรูปภาพไปยัง NAS สำเร็จ: {destination_path}")
            
        except Exception as e:
            logger.error(f"❌ เกิดข้อผิดพลาดในการแปลงรูปภาพ: {e}", exc_info=True)
            # Fallback: คัดลอกไฟล์ตรงๆ
            try:
                shutil.copy2(source_path, destination_path)
                logger.info(f"✅ คัดลอกรูปภาพไปยัง NAS สำเร็จ (fallback): {destination_path}")
            except Exception as copy_error:
                logger.error(f"❌ เกิดข้อผิดพลาดในการคัดลอกไฟล์: {copy_error}", exc_info=True)
                return None
        
        # สร้าง HTTPS URL (ใช้ urllib.parse.quote สำหรับ URL encoding)
        from urllib.parse import quote
        # ใช้ Port 49335 เพราะ Router Forward Port 443 → 49335
        base_url = "https://buildmeupconsultant.synology.me/line-images"
        # URL encode เฉพาะชื่อไฟล์ (ไม่ encode / ใน path)
        encoded_filename = quote(output_filename, safe='')
        image_url = f"{base_url}/{encoded_filename}"
        
        logger.info(f"✅ สร้าง URL สำหรับรูปภาพ: {image_url}")
        return image_url
        
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการบันทึกรูปภาพไปยัง NAS: {e}", exc_info=True)
        return None

def upload_image_to_google_drive(image_path: str, filename: str) -> Optional[str]:
    """
    [DEPRECATED] อัปโหลดรูปภาพไปยัง Google Drive และสร้าง HTTPS URL
    ใช้ save_image_to_nas แทน
    
    Args:
        image_path: Path ของไฟล์รูปภาพ
        filename: ชื่อไฟล์
    
    Returns:
        HTTPS URL ของรูปภาพ หรือ None ถ้าไม่สำเร็จ
    """
    # ใช้ฟังก์ชันใหม่แทน
    return save_image_to_nas(image_path, filename)

@app.route('/api/line/send/pdf', methods=['POST'])
def send_line_pdf():
    """ส่งไฟล์ PDF ไปยัง LINE Official Account"""
    try:
        data = request.json
        user_id = data.get('user_id', '')
        pdf_filename = data.get('pdf_path', '')
        desired_filename = data.get('pdf_filename')  # ชื่อไฟล์ที่ต้องการ (จาก frontend)
        message = data.get('message', '')  # ข้อความที่จะส่งก่อน PDF
        confirm = data.get('confirm', False)  # ถ้า confirm=True แสดงว่าผู้ใช้ยืนยันแล้ว
        
        if not user_id:
            return jsonify({
                'success': False,
                'message': 'กรุณาระบุ LINE User ID / Group ID / Room ID'
            }), 400
        
        if not pdf_filename:
            return jsonify({
                'success': False,
                'message': 'กรุณาระบุ pdf_path'
            }), 400
        
        # สร้าง full path
        pdf_path = Path('temp_uploads') / pdf_filename
        
        if not pdf_path.exists():
            return jsonify({
                'success': False,
                'message': f'ไม่พบไฟล์ PDF: {pdf_filename}'
            }), 404
        
        # ตรวจสอบสถานะการแจ้งเตือน LINE
        if not get_line_notifications_enabled():
            return jsonify({
                'success': False,
                'message': 'การแจ้งเตือน LINE ถูกปิดอยู่'
            }), 400
        
        # ถ้ายืนยันแล้ว ไม่ส่งข้อความไปยัง LINE
        if confirm:
            # ไม่ส่งข้อความ PDF สรุปไปยัง LINE แล้ว
            return jsonify({
                'success': True,
                'message': f'รับคำขอส่ง PDF แล้ว (ไม่ส่งไปยัง LINE)'
            })
        
        # ถ้ายังไม่ยืนยัน ให้ส่งข้อมูลกลับไป (ไม่แปลง PDF เป็นรูปภาพ)
        pdf_display_name = desired_filename if desired_filename else "สรุปภาษี.pdf"
        return jsonify({
            'success': True,
            'message': f'PDF ({pdf_display_name}) พร้อมส่งไปยัง LINE',
            'pdf_filename': pdf_filename,
            'needs_confirmation': True
        }), 200
    except Exception as e:
        logger.error(f"Error sending LINE PDF: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/line/send/payin', methods=['POST'])
def send_line_payin():
    """ส่งไฟล์ Pay-in (แปลงเป็นรูปภาพ) ไปยัง LINE Official Account"""
    try:
        data = request.json
        user_id = data.get('user_id', '')
        file_id = data.get('file_id', '')
        filename = data.get('filename', 'Pay-in.pdf')
        
        logger.info(f"📄 [Pay-in] รับคำขอส่งไฟล์ Pay-in: file_id={file_id}, filename={filename}")
        
        if not user_id:
            logger.error("❌ [Pay-in] ไม่มี user_id")
            return jsonify({
                'success': False,
                'message': 'กรุณาระบุ LINE User ID / Group ID / Room ID'
            }), 400
        
        if not file_id:
            logger.error("❌ [Pay-in] ไม่มี file_id")
            return jsonify({
                'success': False,
                'message': 'กรุณาระบุ file_id'
            }), 400
        
        # ตรวจสอบสถานะการแจ้งเตือน LINE
        if not get_line_notifications_enabled():
            logger.warning("⚠️ [Pay-in] การแจ้งเตือน LINE ถูกปิดอยู่")
            return jsonify({
                'success': False,
                'message': 'การแจ้งเตือน LINE ถูกปิดอยู่'
            }), 400
        
        # หาไฟล์จาก temp_uploads
        temp_dir = Path('temp_uploads')
        matching_files = list(temp_dir.glob(f"{file_id}*"))
        
        logger.info(f"🔍 [Pay-in] ค้นหาไฟล์: {file_id}* → พบ {len(matching_files)} ไฟล์")
        
        if not matching_files:
            logger.error(f"❌ [Pay-in] ไม่พบไฟล์: {file_id}")
            return jsonify({
                'success': False,
                'message': f'ไม่พบไฟล์: {file_id}'
            }), 404
        
        file_path = matching_files[0]
        logger.info(f"📁 [Pay-in] ใช้ไฟล์: {file_path}")
        
        # ไม่ส่งข้อความไปยัง LINE แล้ว
        return jsonify({
            'success': True,
            'message': f'รับไฟล์ Pay-in ({filename}) แล้ว'
        })
            
    except Exception as e:
        logger.error(f"❌ [Pay-in] Error sending LINE Pay-in: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500

@app.route('/api/line/send/image', methods=['POST'])
def send_line_image():
    """ส่งรูปภาพไปยัง LINE Official Account (ขั้นตอนที่ 1: อัปโหลดและแสดงรูปภาพ)"""
    try:
        data = request.json
        user_id = data.get('user_id', '')
        image_filename = data.get('image_path', '')
        confirm = data.get('confirm', False)  # ถ้า confirm=True แสดงว่าผู้ใช้ยืนยันแล้ว
        
        if not user_id:
            return jsonify({
                'success': False,
                'message': 'กรุณาระบุ LINE User ID / Group ID / Room ID'
            }), 400
        
        if not image_filename:
            return jsonify({
                'success': False,
                'message': 'กรุณาระบุ image_path'
            }), 400
        
        logger.info(f"📄 [LINE Image] รับคำขอส่งรูปภาพ: image_filename={image_filename}, confirm={confirm}")
        
        # ตรวจสอบว่าไฟล์อยู่ที่ NAS หรือ temp_uploads
        nas_path = Path('V:/web/line-images') / image_filename
        temp_path = Path('temp_uploads') / image_filename
        
        if nas_path.exists():
            image_path = nas_path
            logger.info(f"✅ [LINE Image] พบไฟล์ใน NAS: {nas_path}")
        elif temp_path.exists():
            image_path = temp_path
            logger.info(f"✅ [LINE Image] พบไฟล์ใน temp_uploads: {temp_path}")
        else:
            logger.error(f"❌ [LINE Image] ไม่พบไฟล์: {image_filename}")
            logger.error(f"   - ตรวจสอบใน NAS: {nas_path} → {'พบ' if nas_path.exists() else 'ไม่พบ'}")
            logger.error(f"   - ตรวจสอบใน temp_uploads: {temp_path} → {'พบ' if temp_path.exists() else 'ไม่พบ'}")
            return jsonify({
                'success': False,
                'message': f'ไม่พบไฟล์รูปภาพ: {image_filename}'
            }), 404
        
        # ตรวจสอบสถานะการแจ้งเตือน LINE
        if not get_line_notifications_enabled():
            return jsonify({
                'success': False,
                'message': 'การแจ้งเตือน LINE ถูกปิดอยู่'
            }), 400
        
        # ถ้ายืนยันแล้ว ให้ส่งข้อความไปยัง LINE (ไม่ส่งรูปภาพ)
        if confirm:
            from report_manager import line_oa_push
            
            message_text = f"📄 มีรูปภาพ ({image_filename}) แนบในอีเมลล์\n\nกรุณาตรวจสอบอีเมลล์เพื่อดูรูปภาพ"
            success = line_oa_push(message_text, to=user_id)
            
            if success:
                return jsonify({
                    'success': True,
                    'message': f'ส่งข้อความ LINE สำเร็จไปยัง {user_id[:10]}...'
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'ส่งข้อความ LINE ไม่สำเร็จ'
                }), 500
        
        # ถ้ายังไม่ยืนยัน ให้บันทึกรูปภาพไปยัง NAS และส่ง URL กลับไป
        try:
            nas_url = upload_image_to_google_drive(str(image_path), image_filename)  # เรียกใช้ฟังก์ชันที่แก้ไขแล้ว (ใช้ NAS)
            if nas_url:
                # ตรวจสอบว่า URL เข้าถึงได้ (preload เพื่อให้ LINE cache รูปภาพ)
                try:
                    import requests
                    preload_response = requests.head(nas_url, timeout=5, allow_redirects=True)
                    if preload_response.status_code == 200:
                        logger.info(f"✅ URL รูปภาพเข้าถึงได้: {nas_url}")
                    else:
                        logger.warning(f"⚠️ URL รูปภาพอาจมีปัญหา (status: {preload_response.status_code}): {nas_url}")
                except Exception as e:
                    logger.warning(f"⚠️ ไม่สามารถตรวจสอบ URL รูปภาพได้: {e}")
                
                # ส่ง URL กลับไปให้ frontend เพื่อแสดงรูปภาพและยืนยัน
                return jsonify({
                    'success': True,
                    'message': 'บันทึกรูปภาพไปยัง NAS สำเร็จ กรุณาตรวจสอบ preview รูปภาพก่อนยืนยันส่ง',
                    'image_url': nas_url,
                    'image_filename': image_filename,
                    'needs_confirmation': True,  # บอก frontend ว่าต้องยืนยันก่อนส่ง
                    'open_image_first': True  # บอก frontend ให้เปิดลิงค์รูปภาพก่อน
                }), 200
            else:
                # ถ้าบันทึกไปยัง NAS ไม่สำเร็จ ให้ส่งข้อความแทน
                logger.warning("⚠️ บันทึกรูปภาพไปยัง NAS ไม่สำเร็จ - ส่งข้อความแทน")
                from report_manager import line_oa_push
                message_text = f"📄 มีรูปภาพ ({image_filename}) แนบในอีเมลล์\n\nกรุณาตรวจสอบอีเมลล์เพื่อดูรูปภาพ"
                success = line_oa_push(message_text, to=user_id)
                if success:
                    return jsonify({
                        'success': True,
                        'message': f'ส่งข้อความ LINE สำเร็จไปยัง {user_id[:10]}...'
                    })
                else:
                    return jsonify({
                        'success': False,
                        'message': 'ส่งข้อความ LINE ไม่สำเร็จ'
                    }), 500
        except Exception as e:
            logger.error(f"Error saving image to NAS: {e}", exc_info=True)
            # Fallback ไปส่งข้อความแทน
            from report_manager import line_oa_push
            message_text = f"📄 มีรูปภาพ ({image_filename}) แนบในอีเมลล์\n\nกรุณาตรวจสอบอีเมลล์เพื่อดูรูปภาพ"
            success = line_oa_push(message_text, to=user_id)
            if success:
                return jsonify({
                    'success': True,
                    'message': f'ส่งข้อความ LINE สำเร็จไปยัง {user_id[:10]}...'
                })
            else:
                return jsonify({
                    'success': False,
                    'message': f'เกิดข้อผิดพลาด: {e}'
                }), 500
    except Exception as e:
        logger.error(f"Error sending LINE image: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'}), 500


# ==================== OCR Raw Reader API ====================
# Global dict สำหรับเก็บ progress
ocr_progress_store = {}

# Global dict สำหรับเก็บ progress ของ auditcheck OCR
auditcheck_ocr_progress = {}

@app.route('/api/ocr/read-raw', methods=['POST'])
@rate_limit(max_requests=10, window=60)
def api_read_ocr_raw():
    """
    API สำหรับอ่านข้อมูลดิบจากโฟลเดอร์ (Raw OCR Text)
    
    Request body:
        {
            "folder_path": "V:/...",
            "include_subfolders": true,
            "max_files": 50
        }
    
    Response:
        {
            "success": true,
            "session_id": "abc123",  # สำหรับ track progress
            "files": [...],
            "stats": {...}
        }
    """
    try:
        from ocr_reader_service import OCRReaderService
        import uuid
        import threading
        
        data = request.get_json()
        folder_path = data.get('folder_path', '').strip()
        include_subfolders = data.get('include_subfolders', False)  # เปลี่ยนเป็น False (อ่านเฉพาะระดับบนสุด)
        max_files = data.get('max_files', 9999)  # ไม่จำกัดจำนวนไฟล์ (จะพักทุก 50 ไฟล์)
        clear_cache = data.get('clear_cache', False)  # ถ้าเป็น True ให้ลบ cache ก่อนอ่าน
        
        if not folder_path:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุโฟลเดอร์'
            }), 400
        
        # สร้าง session ID สำหรับ track progress
        session_id = str(uuid.uuid4())
        ocr_progress_store[session_id] = {
            'current': 0,
            'total': 0,
            'percent': 0,
            'filename': 'กำลังเริ่มต้น...',
            'status': 'processing',
            'files': [],
            'stats': {'total': 0, 'success': 0, 'error': 0}
        }
        
        # Background worker function
        def read_ocr_worker():
            """Worker function สำหรับอ่าน OCR ใน background"""
            try:
                # Progress callback
                def progress_callback(current, total, percent, filename):
                                            ocr_progress_store[session_id].update({
                        'current': current,
                        'total': total,
                        'percent': percent,
                        'filename': filename,
                        'status': 'processing'
                                            })
                                            logger.debug(f"📊 [OCR Progress] {current}/{total} ({percent}%) - {filename}")
                
                # ถ้าต้องการลบ cache ให้ลบก่อนอ่าน
                if clear_cache:
                    from ocr_cache_manager import OCRCacheManager
                    from pathlib import Path
                    
                    # ดึงชื่อบริษัทจาก folder path
                    company_name = OCRCacheManager._extract_company_name_from_path(folder_path)
                    cache_manager = OCRCacheManager(cache_ttl_hours=720, company_name=company_name)  # 30 วัน
                    folder_path_obj = Path(folder_path)
                    
                    # ลบ cache ที่อยู่ในโฟลเดอร์นี้
                    keys_to_delete = []
                    for cache_key, cache_entry in cache_manager.cache_data.items():
                        cached_filepath = cache_entry.get('filepath', '')
                        if cached_filepath:
                            try:
                                cached_path_obj = Path(cached_filepath)
                                try:
                                    if cached_path_obj.parent.resolve() == folder_path_obj.resolve():
                                        keys_to_delete.append(cache_key)
                                except (OSError, ValueError):
                                    pass
                            except Exception:
                                pass
                    
                    for key in keys_to_delete:
                        del cache_manager.cache_data[key]
                    
                    if keys_to_delete:
                        cache_manager._save_cache()
                        logger.info(f"🗑️ ลบ cache สำหรับโฟลเดอร์ {folder_path}: {len(keys_to_delete)} รายการ (บริษัท: {company_name})")
                
                service = OCRReaderService(max_files=max_files)
                result = service.read_folder_raw(
                    folder_path, 
                    include_subfolders=include_subfolders, 
                    progress_callback=progress_callback
                )
        
                # อัปเดตผลลัพธ์สุดท้าย
                ocr_progress_store[session_id].update({
                    'status': 'completed',
                    'current': result.get('stats', {}).get('total', 0),
                    'total': result.get('stats', {}).get('total', 0),
                    'percent': 100,
                    'filename': 'เสร็จสิ้น',
                    'files': result.get('files', []),
                    'stats': result.get('stats', {'total': 0, 'success': 0, 'error': 0}),
                    'result': result,  # เก็บผลลัพธ์เต็มไว้
                    'completed_at': time.time()  # เก็บเวลาที่เสร็จสิ้นเพื่อ cleanup ภายหลัง
                })
                logger.info(f"✅ [OCR Worker] Session {session_id} completed: {result.get('stats', {})}")
                
            except Exception as e:
                logger.error(f"❌ [OCR Worker] Error in session {session_id}: {e}", exc_info=True)
                ocr_progress_store[session_id].update({
                    'status': 'error',
                    'filename': f'เกิดข้อผิดพลาด: {str(e)}',
                    'error': str(e)
                })
        
        # เริ่ม background thread
        worker_thread = threading.Thread(target=read_ocr_worker, daemon=True)
        worker_thread.start()
        logger.info(f"🚀 [OCR] Started background worker for session {session_id}")
        
        # Return session_id ทันที (ไม่รอให้อ่านเสร็จ)
        return jsonify({
            'success': True,
            'session_id': session_id,
            'message': 'เริ่มอ่านข้อมูล OCR แล้ว กรุณารอสักครู่...',
            'status': 'processing'
        })
        
    except Exception as e:
        logger.error(f"❌ API error: {e}", exc_info=True)
        if 'session_id' in locals() and session_id in ocr_progress_store:
            ocr_progress_store[session_id]['status'] = 'error'
            ocr_progress_store[session_id]['error'] = str(e)
        return jsonify({
            'success': False,
            'error': f'เกิดข้อผิดพลาด: {e}'
        }), 500


@app.route('/api/ocr/cache/check-folder', methods=['POST'])
def api_ocr_cache_check_folder():
    """
    API สำหรับตรวจสอบว่ามี cache สำหรับโฟลเดอร์นี้หรือไม่
    
    Request body:
        {
            "folder_path": "V:/..."
        }
    
    Response:
        {
            "success": true,
            "has_cache": true,
            "cache_count": 10,
            "message": "พบ cache 10 รายการ"
        }
    """
    try:
        from ocr_cache_manager import OCRCacheManager
        from pathlib import Path
        
        data = request.get_json()
        folder_path = data.get('folder_path', '').strip()
        
        if not folder_path:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุโฟลเดอร์'
            }), 400
        
        # ดึงชื่อบริษัทจาก folder path
        company_name = OCRCacheManager._extract_company_name_from_path(folder_path)
        cache_manager = OCRCacheManager(cache_ttl_hours=720, company_name=company_name)  # 30 วัน
        
        # นับจำนวน cache ที่ตรงกับโฟลเดอร์นี้
        folder_path_obj = Path(folder_path)
        cache_count = 0
        
        for cache_entry in cache_manager.cache_data.values():
            cached_filepath = cache_entry.get('filepath', '')
            if cached_filepath:
                try:
                    cached_path_obj = Path(cached_filepath)
                    # ตรวจสอบว่าไฟล์อยู่ในโฟลเดอร์เดียวกันหรือไม่
                    try:
                        if cached_path_obj.parent.resolve() == folder_path_obj.resolve():
                            cache_count += 1
                    except (OSError, ValueError):
                        # ถ้า path ไม่ valid ให้ข้าม
                        pass
                except Exception:
                    pass
        
        return jsonify({
            'success': True,
            'has_cache': cache_count > 0,
            'cache_count': cache_count,
            'message': f'พบ cache {cache_count} รายการ' if cache_count > 0 else 'ไม่พบ cache'
        })
    except Exception as e:
        logger.error(f"❌ [OCR Cache Check Folder] Error: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/ocr/cache/stats', methods=['GET'])
def api_ocr_cache_stats():
    """
    API สำหรับดูสถิติ OCR cache
    
    Response:
        {
            "success": true,
            "stats": {
                "total": 100,
                "valid": 95,
                "expired": 5,
                "ttl_hours": 8
            }
        }
    """
    try:
        from ocr_cache_manager import OCRCacheManager
        cache_manager = OCRCacheManager(cache_ttl_hours=8)
        stats = cache_manager.get_stats()
        
        return jsonify({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        logger.error(f"❌ [OCR Cache Stats] Error: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/ocr/cache/cleanup', methods=['POST'])
def api_ocr_cache_cleanup():
    """
    API สำหรับลบ OCR cache ที่หมดอายุ
    
    Response:
        {
            "success": true,
            "deleted_count": 5
        }
    """
    try:
        from ocr_cache_manager import OCRCacheManager
        cache_manager = OCRCacheManager(cache_ttl_hours=8)
        deleted_count = cache_manager.cleanup_expired()
        
        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'ลบ cache ที่หมดอายุแล้ว: {deleted_count} รายการ'
        })
    except Exception as e:
        logger.error(f"❌ [OCR Cache Cleanup] Error: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/ocr/cache/clear', methods=['POST'])
def api_ocr_cache_clear():
    """
    API สำหรับลบ OCR cache ทั้งหมด
    
    Response:
        {
            "success": true,
            "message": "ลบ cache ทั้งหมดแล้ว"
        }
    """
    try:
        from ocr_cache_manager import OCRCacheManager
        cache_manager = OCRCacheManager(cache_ttl_hours=8)
        cache_manager.clear_all()
        
        return jsonify({
            'success': True,
            'message': 'ลบ cache ทั้งหมดแล้ว'
        })
    except Exception as e:
        logger.error(f"❌ [OCR Cache Clear] Error: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/ocr-progress/<session_id>', methods=['GET'])
def api_get_auditcheck_ocr_progress(session_id: str):
    """API สำหรับดึงข้อมูล progress ของ auditcheck OCR"""
    try:
        if session_id not in auditcheck_ocr_progress:
            return jsonify({
                'error': 'ไม่พบ session',
                'status': 'not_found'
            }), 404
        
        progress_data = auditcheck_ocr_progress[session_id].copy()
        return jsonify(progress_data)
    except Exception as e:
        logger.error(f"❌ Error getting auditcheck OCR progress: {e}")
        return jsonify({
            'error': str(e),
            'status': 'error'
        }), 500


@app.route('/api/ocr/progress/<session_id>', methods=['GET'])
def api_get_ocr_progress(session_id: str):
    """
    API สำหรับดึงข้อมูล progress
    
    Response:
        {
            "current": 5,
            "total": 10,
            "percent": 50,
            "filename": "file.pdf",
            "status": "processing"  # or "completed" or "error"
        }
    """
    try:
        if session_id not in ocr_progress_store:
            return jsonify({
                'error': 'ไม่พบ session',
                'status': 'not_found'
            }), 404
        
        progress_data = ocr_progress_store[session_id].copy()
        
        # ถ้าเสร็จแล้ว ให้ส่งข้อมูล files และ stats กลับไปด้วย
        if progress_data.get('status') == 'completed' and 'result' in progress_data:
            result = progress_data.pop('result')  # เอา result ออกเพื่อไม่ให้ส่งซ้ำ
            progress_data['files'] = result.get('files', [])
            progress_data['stats'] = result.get('stats', {'total': 0, 'success': 0, 'error': 0})
            progress_data['success'] = result.get('success', True)
            if not result.get('success'):
                progress_data['error'] = result.get('error', 'เกิดข้อผิดพลาด')
        
        return jsonify(progress_data)
    except Exception as e:
        logger.error(f"❌ Error getting progress: {e}")
        return jsonify({
            'error': str(e),
            'status': 'error'
        }), 500


@app.route('/api/ocr/read-raw-file', methods=['POST'])
@rate_limit(max_requests=20, window=60)
def api_read_ocr_raw_file():
    """
    API สำหรับอ่านข้อมูลดิบจากไฟล์เดียว
    
    Request body:
        {
            "file_path": "V:/.../file.pdf"
        }
    
    Response:
        {
            "success": true,
            "file": {
                "filename": "file.pdf",
                "status": "success",
                "raw_text": "...",
                "basic_info": {...}
            }
        }
    """
    try:
        from ocr_reader_service import OCRReaderService
        
        data = request.get_json()
        file_path = data.get('file_path', '').strip()
        
        if not file_path:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุไฟล์'
            }), 400
        
        service = OCRReaderService()
        result = service.read_single_file_raw(file_path)
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"❌ API error: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': f'เกิดข้อผิดพลาด: {e}'
        }), 500


# ===== API: ค้นหาไฟล์ Excel ในโฟลเดอร์ =====
@app.route('/api/excel/list-files', methods=['POST'])
def api_list_excel_files():
    """API สำหรับค้นหาไฟล์ Excel ในโฟลเดอร์"""
    try:
        data = request.get_json()
        folder_path = data.get('folder_path', '').strip()
        
        if not folder_path:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุโฟลเดอร์'
            }), 400
        
        folder = Path(folder_path)
        
        # ตรวจสอบว่าโฟลเดอร์มีอยู่จริงหรือไม่ (ใช้ try-except เพื่อจัดการ error)
        try:
            if not folder.exists():
                return jsonify({
                    'success': False,
                    'error': f'ไม่พบโฟลเดอร์: {folder_path}'
                }), 404
        except (OSError, ValueError) as e:
            logger.error(f"❌ ไม่สามารถตรวจสอบโฟลเดอร์ได้: {folder_path} - {e}")
            return jsonify({
                'success': False,
                'error': f'ไม่สามารถเข้าถึงโฟลเดอร์ได้: {folder_path} (อาจเป็น path ยาวเกินไปหรือไม่มีสิทธิ์เข้าถึง)'
            }), 400
        
        # ค้นหาไฟล์ Excel ทั้งหมดในโฟลเดอร์
        excel_files = []
        excel_extensions = ['.xlsx', '.xls']
        
        try:
            for file_path in folder.iterdir():
                try:
                    if file_path.is_file() and file_path.suffix.lower() in excel_extensions:
                        # ข้ามไฟล์ lock ของ Excel
                        if not file_path.name.startswith('~$'):
                            file_info = {
                                'filename': file_path.name,
                                'path': str(file_path),
                                'size': file_path.stat().st_size,
                                'modified': file_path.stat().st_mtime
                            }
                            excel_files.append(file_info)
                except (OSError, PermissionError) as e:
                    # ข้ามไฟล์ที่เข้าถึงไม่ได้
                    logger.warning(f"⚠️ ข้ามไฟล์ที่เข้าถึงไม่ได้: {file_path} - {e}")
                    continue
        except (OSError, PermissionError) as e:
            logger.error(f"❌ ไม่สามารถอ่านโฟลเดอร์ได้: {folder_path} - {e}")
            return jsonify({
                'success': False,
                'error': f'ไม่สามารถอ่านโฟลเดอร์ได้: {folder_path} (อาจไม่มีสิทธิ์เข้าถึง)'
            }), 403
        
        # เรียงตามชื่อไฟล์
        excel_files.sort(key=lambda x: x['filename'])
        
        return jsonify({
            'success': True,
            'files': excel_files,
            'count': len(excel_files)
        }), 200
        
    except Exception as e:
        logger.error(f"❌ [List Excel Files] Error: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': f'เกิดข้อผิดพลาด: {e}'
        }), 500


# ===== API: รันด้วยไฟล์ Excel =====
@app.route('/api/excel/run-with-excel', methods=['POST'])
def api_run_with_excel():
    """API สำหรับรันงานด้วยไฟล์ Excel"""
    try:
        data = request.get_json()
        folder_path = data.get('folder_path', '').strip()
        excel_path = data.get('excel_path', '').strip()  # รับ excel_path จาก request
        
        if not folder_path:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุโฟลเดอร์'
            }), 400
        
        folder = Path(folder_path)
        if not folder.exists():
            return jsonify({
                'success': False,
                'error': f'ไม่พบโฟลเดอร์: {folder_path}'
            }), 404
        
        # ถ้าไม่ระบุ excel_path ให้ใช้ค่า default
        if not excel_path:
            excel_path = str(folder / "Invoice_Data.xlsx")
        
        excel_file = Path(excel_path)
        if not excel_file.exists():
            return jsonify({
                'success': False,
                'error': f'ไม่พบไฟล์ Excel: {excel_path}\n\nกรุณาตรวจสอบว่าได้สร้างไฟล์ Excel แล้ว (ใช้ฟีเจอร์ "อ่านข้อมูล OCR" และ "ส่งทั้งหมดไป Excel")'
            }), 404
        
        # ตรวจสอบว่าไฟล์อยู่ในโฟลเดอร์เดียวกันหรือไม่
        if excel_file.parent != folder:
            return jsonify({
                'success': False,
                'error': f'ไฟล์ Excel ต้องอยู่ในโฟลเดอร์เดียวกัน: {folder_path}'
            }), 400
        
        # ตรวจสอบจำนวน concurrent jobs
        running_count = _get_running_jobs_count()
        if running_count >= MAX_CONCURRENT_JOBS:
            return jsonify({
                'success': False,
                'error': f'มีงานที่กำลังรันอยู่ {running_count} งาน (สูงสุด {MAX_CONCURRENT_JOBS} งาน) กรุณารอให้งานเสร็จก่อน'
            }), 429
        
        is_locked, existing_job_id = _is_folder_locked(folder_path)
        if is_locked:
            return jsonify({
                'success': False,
                'error': f'โฟลเดอร์นี้กำลังถูกประมวลผลอยู่แล้ว (Job ID: {existing_job_id})'
            }), 400
        
        # Cleanup old jobs ก่อนสร้างใหม่
        _cleanup_old_jobs()
        
        # สร้าง job ใหม่
        job_id = str(uuid.uuid4())[:8]
        job_data = {
            'id': job_id,
            'folder': str(folder),
            'status': 'pending',
            'progress': {'total': 0, 'success': 0, 'failed': 0, 'duplicates': 0},
            'current_folder': str(folder),
            'current_file': '-',
            'current_step': 'รอเริ่มงาน',
            'log': [],
            'start_time': None,
            'end_time': None,
            'display_name': _get_folder_display_name(str(folder)),
            'error_details': [],
            'job_type': 'excel_run'  # ระบุว่าเป็น job ประเภท excel_run
        }
        
        if not _lock_folder(folder_path, job_id):
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถ lock โฟลเดอร์ได้'
            }), 400
        
        with job_store_lock:
            job_store[job_id] = job_data
        
        # ใช้ ThreadPoolExecutor สำหรับรัน background job
        future = job_executor.submit(_excel_run_worker, job_id, str(folder), str(excel_path))
        
        with job_store_lock:
            if job_id in job_store:
                job_store[job_id]['future'] = future
        
        return jsonify({
            'success': True,
            'job_id': job_id,
            'message': f'เริ่มงานรันด้วย Excel เรียบร้อย (Job ID: {job_id})',
            'excel_path': str(excel_path)
        }), 201
        
    except Exception as e:
        logger.error(f"❌ [Run with Excel] Error: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': f'เกิดข้อผิดพลาด: {e}'
        }), 500


def _excel_run_worker(job_id: str, folder_path: str, excel_path: str):
    """Worker thread สำหรับรันงานด้วย Excel"""
    with job_store_lock:
        job = job_store.get(job_id)
        if not job:
            return
        job['start_time'] = datetime.now().isoformat()
        job['status'] = 'running'
        job['display_name'] = _get_folder_display_name(folder_path)
    
    _job_add_log(job_id, f"🚀 เริ่มรันงานด้วย Excel: {excel_path}", "info")
    _job_update_status(job_id, folder=folder_path, step="กำลังเตรียมระบบ", file='-')
    
    error_message = None
    error_type = None
    
    try:
        from web_automation_playwright import WebAutomationPlaywright
        
        # สร้าง callback functions
        def progress_callback(**kwargs):
            _job_update_progress(job_id, **kwargs)
        
        def status_callback(**kwargs):
            _job_update_status(job_id, **kwargs)
        
        def log_callback(message, level="info"):
            _job_add_log(job_id, message, level)
        
        # สร้าง WebAutomationPlaywright instance
        automation = WebAutomationPlaywright(
            progress_callback=progress_callback,
            status_callback=status_callback,
            log_callback=log_callback
        )
        
        _job_add_log(job_id, "🔧 กำลังตั้งค่า Playwright...", "info")
        _job_update_status(job_id, folder=folder_path, step="กำลังตั้งค่า Playwright", file='-')
        
        # Setup driver
        if not automation.setup_driver():
            raise Exception("ไม่สามารถตั้งค่า Playwright ได้")
        
        _job_add_log(job_id, "✅ ตั้งค่า Playwright สำเร็จ", "success")
        
        try:
            # อ่านข้อมูลการตั้งค่าจากไฟล์ txt
            main_folder_path = Path(folder_path)
            build_folder = None
            current_path = main_folder_path
            while current_path != current_path.parent:
                if current_path.name.startswith('Build'):
                    build_folder = current_path
                    break
                current_path = current_path.parent
            
            if not build_folder:
                raise Exception(f"ไม่พบโฟลเดอร์ Build* จาก: {folder_path}")
            
            folder_name = build_folder.name
            build_number = folder_name.split()[0]
            local_txt = build_folder / "รหัส" / f"{build_number}.txt"
            
            # หาไฟล์ txt จากหลายตำแหน่ง
            from config import Config
            drive_root = Path(f"{Config.BASE_FOLDER}:/")
            candidates = [local_txt]
            for main_name in getattr(Config, 'MAIN_FOLDERS', ["A.โฟร์เดอร์หลัก", "AA.โฟรเดอร์หลัก", "AAA.โฟรเดอร์หลัก"]):
                candidates.append(
                    drive_root / main_name / f"Build{Config.TEST_SYSTEM_FOLDER}" / "รหัส" / f"{build_number}.txt"
                )
            candidates.append(drive_root / f"Build{Config.TEST_SYSTEM_FOLDER}" / "รหัส" / f"{build_number}.txt")
            
            chosen = None
            for c in candidates:
                if c.exists():
                    chosen = c
                    break
            
            config_file_path = str(chosen) if chosen else str(local_txt)
            
            _job_add_log(job_id, f"📖 กำลังอ่านข้อมูลการตั้งค่าจาก: {config_file_path}", "info")
            
            if not automation.read_config_from_txt(config_file_path):
                raise Exception("ไม่สามารถอ่านข้อมูลการตั้งค่าได้")
            
            _job_add_log(job_id, "✅ อ่านข้อมูลการตั้งค่าสำเร็จ", "success")
            
            # ตรวจสอบว่า credentials ถูกตั้งค่าหรือไม่
            if not automation.credentials or not automation.credentials.get('Username') or not automation.credentials.get('Password'):
                raise Exception("ไม่พบข้อมูล Username หรือ Password ในไฟล์การตั้งค่า")
            
            _job_add_log(job_id, f"🔐 กำลังล็อกอินด้วย Username: {automation.credentials.get('Username')}", "info")
            _job_update_status(job_id, folder=folder_path, step="กำลังล็อกอิน", file='-')
            
            if not automation.login_to_peak_engine(automation.credentials):
                raise Exception("ล็อกอินไม่สำเร็จ - ตรวจสอบ Username และ Password")
            
            _job_add_log(job_id, "✅ ล็อกอินสำเร็จ", "success")
            
            # Navigate to company link
            if automation.company_link:
                _job_add_log(job_id, f"🏢 กำลังไปยัง Company Link...", "info")
                if not automation.navigate_to_company_link(automation.company_link):
                    raise Exception("ไม่สามารถไปยัง Company Link ได้")
                _job_add_log(job_id, "✅ ไปยัง Company Link สำเร็จ", "success")
            
            # Navigate to express link
            if automation.express_link:
                _job_add_log(job_id, f"📝 กำลังไปยัง Express Link...", "info")
                _job_update_status(job_id, folder=folder_path, step="กำลังไปยัง Express Link", file='-')
                if not automation.navigate_to_express_link(automation.express_link):
                    raise Exception("ไม่สามารถไปยัง Express Link ได้")
                _job_add_log(job_id, "✅ ไปยัง Express Link สำเร็จ", "success")
            
            # ประมวลผลข้อมูลจาก Excel
            _job_add_log(job_id, f"📊 กำลังประมวลผลข้อมูลจาก Excel: {excel_path}", "info")
            _job_update_status(job_id, folder=folder_path, step="กำลังประมวลผลข้อมูลจาก Excel", file='-')
            
            # ใช้ process_excel_data_and_fill_form ซึ่งจะทำงานต่อกับระบบบอท
            # ฟังก์ชันนี้จะ:
            # 1. อ่านข้อมูลจาก Excel
            # 2. กดปุ่ม "รันด้วยไฟล์ Excel" บนเว็บ
            # 3. กรอกข้อมูลตามข้อมูลใน Excel
            if not automation.process_excel_data_and_fill_form(excel_path):
                # ใช้ error message ที่ละเอียดจาก automation instance (ถ้ามี)
                detailed_error = getattr(automation, 'last_error_message', None)
                if detailed_error:
                    raise Exception(detailed_error)
                else:
                    raise Exception("การประมวลผลข้อมูลจาก Excel ไม่สำเร็จ")
            
            _job_add_log(job_id, "✅ ประมวลผลข้อมูลจาก Excel สำเร็จ", "success")
            _job_set_state(job_id, 'success')
            _job_add_log(job_id, "✅ งานเสร็จสมบูรณ์", "success")
            
        finally:
            # ปิด Playwright
            _job_add_log(job_id, "🔒 กำลังปิด Playwright...", "info")
            automation.close_driver()
            _job_add_log(job_id, "✅ ปิด Playwright แล้ว", "success")
            
    except Exception as e:
        error_message = str(e)
        error_type = type(e).__name__
        
        _job_set_state(job_id, 'error')
        _job_add_log(job_id, f"❌ เกิดข้อผิดพลาด: {error_type}: {error_message}", "error")
        
        with job_store_lock:
            job = job_store.get(job_id)
            if job:
                job['error_message'] = error_message
                job['error_type'] = error_type
    finally:
        try:
            _job_update_status(job_id, step="เสร็จสิ้น", file='-')
            _unlock_folder(folder_path, job_id)
            
            with job_store_lock:
                job = job_store.get(job_id)
                if job:
                    job.pop('future', None)
                    job.pop('thread', None)
        except Exception as e:
            logger.error(f"⚠️ [Excel Run Worker] Error in finally block for job {job_id}: {e}")


# Background task สำหรับ cleanup
def _cleanup_background_task():
    """Background task สำหรับ cleanup jobs เก่าและไฟล์เก่า"""
    while True:
        try:
            time.sleep(1800)  # รันทุก 30 นาที
            _cleanup_old_jobs()
            # ลบไฟล์เก่าใน temp_uploads (มากกว่า 30 นาที)
            cleanup_temp_files(max_age_minutes=30)
        except Exception as e:
            logger.error(f"⚠️ [Cleanup Background] Error: {e}", exc_info=True)

# Background task สำหรับ cleanup OCR cache
def _cleanup_ocr_cache_task():
    """Background task สำหรับลบ OCR cache ที่หมดอายุ (ทุก 30 วัน) - รองรับ cache แยกตามบริษัท"""
    while True:
        try:
            time.sleep(2592000)  # รันทุก 30 วัน (30 * 24 * 60 * 60 = 2592000 วินาที)
            from ocr_cache_manager import OCRCacheManager
            # ใช้ static method เพื่อลบ cache จากทุกบริษัท
            deleted_count = OCRCacheManager.cleanup_all_expired(cache_dir="cache", cache_ttl_hours=720)
            if deleted_count > 0:
                logger.info(f"🗑️ [Cache Cleanup] ลบ cache ที่หมดอายุแล้ว: {deleted_count} รายการ (จากทุกบริษัท)")
            else:
                logger.debug(f"✅ [Cache Cleanup] ไม่มี cache ที่หมดอายุ")
        except Exception as e:
            logger.error(f"⚠️ [Cache Cleanup Background] Error: {e}", exc_info=True)

# เริ่ม background cleanup task
cleanup_thread = threading.Thread(target=_cleanup_background_task, daemon=True)
cleanup_thread.start()

# เริ่ม background cache cleanup task
cache_cleanup_thread = threading.Thread(target=_cleanup_ocr_cache_task, daemon=True)
cache_cleanup_thread.start()
logger.info("✅ เริ่ม background task สำหรับ cleanup OCR cache (ทุก 8 ชั่วโมง)")


# ===== API: ส่งข้อมูล Invoice ไป Excel (Version 2.0) =====
@app.route('/api/invoice/extract-to-excel', methods=['POST'])
def extract_invoice_to_excel():
    """ดึงข้อมูลจากใบแจ้งหนี้และส่งไป Excel (รองรับหลายบริษัท)"""
    try:
        from invoice_data_extractor import extract_invoice_data
        from invoice_excel_writer import write_invoice_data_to_excel
        from file_renamer_service import rename_pdf_file
        
        data = request.json
        filepath = data.get('filepath')
        raw_text = data.get('raw_text')
        filename = data.get('filename')
        folder_path = data.get('folder_path')  # Path โฟลเดอร์ที่ผู้ใช้กรอก
        excel_path = data.get('excel_path')  # Optional: Excel path เฉพาะ
        auto_rename = data.get('auto_rename', False)
        force_create_new = data.get('force_create_new', False)  # บังคับสร้างไฟล์ใหม่
        edited_data = data.get('edited_data')  # ข้อมูลที่ผู้ใช้แก้ไข
        
        if not filepath or not raw_text or not filename:
            return jsonify({
                'success': False,
                'error': 'ข้อมูลไม่ครบถ้วน (ต้องมี filepath, raw_text, filename)'
            }), 400
        
        # ดึงข้อมูล
        logger.info(f"🔍 กำลังดึงข้อมูลจาก: {filename}")
        extracted_data = extract_invoice_data(raw_text, filename, filepath)
        
        # ถ้าไม่มี extractor match แต่มี edited_data ให้ใช้ข้อมูลจาก edited_data แทน
        if not extracted_data.get('success') and edited_data:
            # ตรวจสอบว่ามีข้อมูลที่จำเป็นหรือไม่ (อย่างน้อยต้องมี company_name หรือ tax_id หรือ total_amount)
            has_essential_data = (
                edited_data.get('company_name') or 
                edited_data.get('tax_id') or 
                edited_data.get('total_amount') is not None
            )
            
            if has_essential_data:
                logger.info(f"⚠️ ไม่พบ extractor ที่ match แต่มีข้อมูลที่แก้ไขแล้ว → ใช้ข้อมูลจาก edited_data")
                # สร้าง extracted_data จาก edited_data
                extracted_data = {
                    'success': True,
                    'company': edited_data.get('company_name') or 'ไม่ระบุ',
                    'company_name': edited_data.get('company_name'),
                    'tax_id': edited_data.get('tax_id'),
                    'document_number': edited_data.get('document_number'),
                    'date': edited_data.get('date'),
                    'amount_before_vat': edited_data.get('amount_before_vat'),
                    'vat_amount': edited_data.get('vat_amount'),
                    'total_amount': edited_data.get('total_amount'),
                    'address': edited_data.get('address'),
                    'address_full': edited_data.get('address_full'),
                    'branch': edited_data.get('branch'),
                    'remark': edited_data.get('remark'),
                    'reference': edited_data.get('reference'),
                    'account_name_line1': edited_data.get('account_name_line1'),
                    'account_name_line2': edited_data.get('account_name_line2'),
                    'account_code': edited_data.get('account_code'),
                    'withholding_tax': edited_data.get('withholding_tax'),
                    'document_type': edited_data.get('document_type', 1),  # Default: มี VAT
                    # เพิ่มข้อมูลที่อยู่แยกส่วน (ถ้ามี)
                    'building_number': edited_data.get('building_number'),
                    'other_info': edited_data.get('other_info'),
                    'soi': edited_data.get('soi'),
                    'road': edited_data.get('road'),
                    'subdistrict': edited_data.get('subdistrict'),
                    'district': edited_data.get('district'),
                    'province': edited_data.get('province'),
                    'postal_code': edited_data.get('postal_code'),
                    'new_filename': edited_data.get('new_filename'),
                    'old_filename': edited_data.get('old_filename') or filename,  # ใช้ old_filename จาก invoice_data หรือ filename
                    'buyer_name': edited_data.get('buyer_name'),
                    'buyer_tax_id': edited_data.get('buyer_tax_id'),
                    'buyer_address': edited_data.get('buyer_address'),
                    'document_type_text': edited_data.get('document_type_text'),
                    'document_status': edited_data.get('document_status'),
                }
                # ตรวจสอบความครบถ้วนของข้อมูล
                from invoice_extractors.manager import InvoiceExtractorManager
                manager = InvoiceExtractorManager()
                completeness_info = manager.check_data_completeness(extracted_data)
                extracted_data['completeness'] = completeness_info
                extracted_data['missing_fields'] = completeness_info['missing_fields']
                extracted_data['completeness_score'] = completeness_info['completeness_score']
                extracted_data['is_complete'] = completeness_info['is_complete']
                extracted_data['has_warnings'] = completeness_info['has_warnings']
            else:
                logger.warning(f"⚠️ ไม่พบ extractor ที่ match และไม่มีข้อมูลที่จำเป็นใน edited_data")
        
        if not extracted_data.get('success'):
            return jsonify({
                'success': False,
                'error': extracted_data.get('error', 'ไม่สามารถดึงข้อมูลได้'),
                'message': extracted_data.get('error', 'ไม่สามารถดึงข้อมูลได้')
            }), 400
        
        # เพิ่ม old_filename ถ้ายังไม่มี (ใช้ชื่อไฟล์ปัจจุบัน)
        # แต่ถ้ามี old_filename ใน edited_data (จาก invoice_data) ให้ใช้ค่านั้นแทน
        if edited_data and edited_data.get('old_filename'):
            extracted_data['old_filename'] = edited_data['old_filename']
        elif 'old_filename' not in extracted_data or not extracted_data.get('old_filename'):
            extracted_data['old_filename'] = filename
        
        # ถ้ามีข้อมูลที่แก้ไข ให้ใช้ข้อมูลที่แก้ไขแทนข้อมูลที่อ่านได้
        if edited_data:
            logger.info(f"✏️ ใช้ข้อมูลที่แก้ไขแล้ว")
            # อัปเดตข้อมูลที่อ่านได้ด้วยข้อมูลที่แก้ไข
            for key, value in edited_data.items():
                # ข้าม old_filename เพราะได้จัดการไว้แล้วข้างบน
                if key == 'old_filename':
                    continue
                # สำหรับ account_name_line1 และ account_name_line2 ให้อัปเดตแม้ว่าจะเป็น empty string
                if key in ['account_name_line1', 'account_name_line2']:
                    extracted_data[key] = value if value is not None else ''
                    logger.info(f"   • {key}: '{value}'")
                elif value is not None and value != '':
                    extracted_data[key] = value
                    logger.info(f"   • {key}: {value}")
        
        # ตรวจสอบว่าดึงข้อมูลได้ครบหรือไม่ (ใช้ข้อมูลจาก completeness check)
        missing_fields = extracted_data.get('missing_fields', [])
        completeness_score = extracted_data.get('completeness_score', 0)
        is_complete = extracted_data.get('is_complete', True)
        has_warnings = extracted_data.get('has_warnings', False)
        
        if missing_fields:
            logger.warning(f"⚠️ ข้อมูลไม่ครบ: {', '.join(missing_fields)} (คะแนน: {completeness_score}%)")
        
        # ถ้าข้อมูลไม่ครบ ให้เพิ่ม warning message
        warning_message = None
        if has_warnings:
            missing_required = extracted_data.get('completeness', {}).get('missing_required_fields', [])
            if missing_required:
                warning_message = f"⚠️ ข้อมูลสำคัญขาดหาย: {', '.join(missing_required)}"
            else:
                missing_recommended = extracted_data.get('completeness', {}).get('missing_recommended_fields', [])
                if missing_recommended:
                    warning_message = f"⚠️ ข้อมูลแนะนำขาดหาย: {', '.join(missing_recommended)}"
        
        # กำหนด target folder สำหรับ Excel
        # ถ้ามี folder_path ให้ใช้โฟลเดอร์เดียวกับไฟล์ PDF
        target_folder = None
        if folder_path:
            from pathlib import Path
            target_folder = str(Path(folder_path))
            logger.info(f"📂 Excel จะถูกบันทึกที่: {target_folder}/Invoice_Data.xlsx")
        
        # เขียนลง Excel
        logger.info(f"📊 กำลังเขียนข้อมูลลง Excel...")
        
        # Debug: แสดงข้อมูลก่อนส่งไป Excel
        logger.info(f"🔍 [DEBUG] ข้อมูลที่จะส่งไป Excel:")
        logger.info(f"   company_name: {extracted_data.get('company_name', 'N/A')}")
        logger.info(f"   document_type: {extracted_data.get('document_type')}")
        logger.info(f"   vat_amount: {extracted_data.get('vat_amount')} (type: {type(extracted_data.get('vat_amount'))})")
        logger.info(f"   amount_before_vat: {extracted_data.get('amount_before_vat')}")
        logger.info(f"   total_amount: {extracted_data.get('total_amount')}")
        
        # ตรวจสอบว่าต้องการสร้างไฟล์ใหม่หรือไม่ (จาก request)
        force_create_new = data.get('force_create_new', False)
        success, message = write_invoice_data_to_excel(
            extracted_data,
            excel_path=excel_path,
            target_folder=target_folder,
            create_new=force_create_new,  # ใช้ค่าจาก request
            force_create_new=force_create_new  # ส่งต่อให้ Excel writer
        )
        
        # เปลี่ยนชื่อไฟล์ (ถ้าต้องการ)
        rename_result = None
        if success and auto_rename and extracted_data.get('new_filename') and filepath:
            logger.info(f"🔄 กำลังเปลี่ยนชื่อไฟล์...")
            rename_success, rename_message = rename_pdf_file(
                filepath,
                extracted_data['new_filename'],
                backup=True
            )
            rename_result = {
                'success': rename_success,
                'message': rename_message
            }
            if rename_success:
                logger.info(f"✅ {rename_message}")
            else:
                logger.warning(f"⚠️ {rename_message}")
        
        if success:
            logger.info(f"✅ {message}")
            response_data = {
                'success': True,
                'message': message,
                'data': extracted_data,
                'missing_fields': missing_fields,
                'completeness_score': completeness_score,
                'is_complete': is_complete,
                'has_warnings': has_warnings
            }
            if warning_message:
                response_data['warning'] = warning_message
            if rename_result:
                response_data['rename_result'] = rename_result
            return jsonify(response_data)
        else:
            logger.error(f"❌ {message}")
            return jsonify({
                'success': False,
                'error': message
            }), 500
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการส่งข้อมูลไป Excel: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ===== API: ส่งข้อมูล MSC ไป Excel (Legacy - เพื่อความเข้ากันได้) =====
@app.route('/api/msc/extract-to-excel', methods=['POST'])
def extract_msc_to_excel():
    """ดึงข้อมูล MSC และส่งไป Excel (Legacy endpoint - redirect to /api/invoice/extract-to-excel)"""
    return extract_invoice_to_excel()


# ===== API: เปลี่ยนชื่อไฟล์ PDF =====
@app.route('/api/file/rename', methods=['POST'])
def rename_file_api():
    """เปลี่ยนชื่อไฟล์ PDF"""
    try:
        from file_renamer_service import rename_pdf_file
        
        data = request.json
        original_path = data.get('original_path')
        new_filename = data.get('new_filename')
        backup = data.get('backup', True)
        
        if not original_path or not new_filename:
            return jsonify({
                'success': False,
                'error': 'ข้อมูลไม่ครบถ้วน (ต้องมี original_path, new_filename)'
            }), 400
        
        success, message = rename_pdf_file(original_path, new_filename, backup)
        
        if success:
            return jsonify({
                'success': True,
                'message': message
            })
        else:
            return jsonify({
                'success': False,
                'error': message
            }), 500
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการเปลี่ยนชื่อไฟล์: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ===== API: ดึงรายชื่อบริษัทที่รองรับ =====
@app.route('/api/invoice/supported-companies', methods=['GET'])
def get_supported_companies():
    """ดึงรายชื่อบริษัทที่ระบบรองรับ"""
    try:
        from invoice_data_extractor import InvoiceExtractorManager
        
        manager = InvoiceExtractorManager()
        companies = manager.get_supported_companies()
        
        return jsonify({
            'success': True,
            'companies': companies,
            'total': len(companies)
        })
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการดึงรายชื่อบริษัท: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ===== API: ระบบคัดแยกเอกสาร =====
@app.route('/api/document-sorting/scan', methods=['POST'])
def scan_duplicate_pdfs():
    """สแกนหาไฟล์ PDF ที่มีชื่อซ้ำกัน"""
    try:
        import shutil
        from collections import defaultdict
        
        data = request.json
        base_path = data.get('base_path', '').strip()
        skip_folders = data.get('skip_folders', [])  # รายการโฟลเดอร์ที่จะข้าม
        include_subfolders = data.get('include_subfolders', True)
        
        if not base_path:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุ base_path'
            }), 400
        
        base_folder = Path(base_path)
        if not base_folder.exists():
            return jsonify({
                'success': False,
                'error': f'ไม่พบโฟลเดอร์: {base_path}'
            }), 404
        
        # แปลง skip_folders เป็น set สำหรับการค้นหาแบบเร็ว (case-insensitive)
        skip_folders_set = set()
        if skip_folders:
            for folder in skip_folders:
                skip_folders_set.add(folder.strip())
        
        # ฟังก์ชันสำหรับดึงคีย์เวิร์ดจากชื่อไฟล์
        def extract_keyword_from_filename(filename):
            """ดึงคีย์เวิร์ดจากด้านหน้าชื่อไฟล์ (None_vat, VAT, WHT ฯลฯ)"""
            import re
            
            # ลบ extension ออก
            name_without_ext = filename.rsplit('.', 1)[0] if '.' in filename else filename
            
            # ตรวจสอบคีย์เวิร์ดแต่ละประเภท (case insensitive)
            
            # 1. NONE_VAT: none_vat, nonevat, nonvat (รองรับทั้งตัวเล็กตัวใหญ่)
            # รูปแบบ: None_vat_*, Nonevat_*, nonvat_*
            if re.match(r'^none[_\s\-]?vat[_\s\-]', name_without_ext, re.IGNORECASE) or \
               re.match(r'^nonevat[_\s\-]', name_without_ext, re.IGNORECASE) or \
               re.match(r'^nonvat[_\s\-]', name_without_ext, re.IGNORECASE):
                return 'NONE_VAT'
            
            # 2. WHT: wht_vat, whtvat, wht (รองรับทั้งตัวเล็กตัวใหญ่)
            # รูปแบบ: Wht_vat_*, WHT_vat_*, wht_vat_*, whtvat_*, wht_*
            if re.match(r'^wht[_\s\-]?vat[_\s\-]', name_without_ext, re.IGNORECASE) or \
               re.match(r'^whtvat[_\s\-]', name_without_ext, re.IGNORECASE) or \
               re.match(r'^wht[_\s\-]', name_without_ext, re.IGNORECASE):
                return 'WHT'
            
            # 3. VAT: vat (รองรับทั้งตัวเล็กตัวใหญ่)
            # รูปแบบ: VAT_*, vat_*
            # ตรวจสอบว่าไม่ใช่ none_vat หรือ wht_vat ก่อน
            if re.match(r'^vat[_\s\-]', name_without_ext, re.IGNORECASE):
                return 'VAT'
            
            return 'OTHER'
        
        # สแกนหาไฟล์ PDF
        pdf_files_by_name = defaultdict(list)  # {filename: [list of file paths]}
        pdf_files_by_keyword = defaultdict(list)  # {keyword: [list of file paths]}
        all_folders = set()
        
        if include_subfolders:
            # สแกนแบบ recursive
            for pdf_file in base_folder.rglob('*.pdf'):
                if pdf_file.is_file():
                    # ตรวจสอบว่าอยู่ในโฟลเดอร์ที่ต้องข้ามหรือไม่
                    should_skip = False
                    for part in pdf_file.parts:
                        # ตรวจสอบว่าโฟลเดอร์ชื่อตรงกับ skip_folders หรือไม่
                        if part in skip_folders_set:
                            should_skip = True
                            break
                    
                    if not should_skip:
                        filename = pdf_file.name
                        pdf_files_by_name[filename].append(str(pdf_file))
                        
                        # จัดกลุ่มตามคีย์เวิร์ด
                        keyword = extract_keyword_from_filename(filename)
                        pdf_files_by_keyword[keyword].append({
                            'filename': filename,
                            'path': str(pdf_file),
                            'folder': str(pdf_file.parent)
                        })
                        
                        all_folders.add(str(pdf_file.parent))
        else:
            # สแกนเฉพาะโฟลเดอร์โดยตรง
            for pdf_file in base_folder.glob('*.pdf'):
                if pdf_file.is_file():
                    filename = pdf_file.name
                    pdf_files_by_name[filename].append(str(pdf_file))
                    
                    # จัดกลุ่มตามคีย์เวิร์ด
                    keyword = extract_keyword_from_filename(filename)
                    pdf_files_by_keyword[keyword].append({
                        'filename': filename,
                        'path': str(pdf_file),
                        'folder': str(pdf_file.parent)
                    })
                    
                    all_folders.add(str(pdf_file.parent))
        
        # หาไฟล์ที่มีชื่อซ้ำกัน (มีมากกว่า 1 ไฟล์)
        duplicate_files = {}
        for filename, file_paths in pdf_files_by_name.items():
            if len(file_paths) > 1:
                duplicate_files[filename] = file_paths
        
        # สร้างสถิติตามคีย์เวิร์ด
        keyword_stats = {}
        for keyword, files in pdf_files_by_keyword.items():
            keyword_stats[keyword] = len(files)
        
        # สร้างสถิติ
        total_pdfs = sum(len(paths) for paths in pdf_files_by_name.values())
        duplicate_count = sum(len(paths) for paths in duplicate_files.values())
        unique_count = total_pdfs - duplicate_count + len(duplicate_files)
        
        return jsonify({
            'success': True,
            'stats': {
                'total_folders': len(all_folders),
                'total_pdf_files': total_pdfs,
                'unique_filenames': len(pdf_files_by_name),
                'duplicate_filenames': len(duplicate_files),
                'duplicate_files_count': duplicate_count,
                'unique_files_count': unique_count,
                'keyword_stats': keyword_stats
            },
            'duplicates': duplicate_files,
            'files_by_keyword': {k: v for k, v in pdf_files_by_keyword.items()},
            'all_folders': sorted(list(all_folders))
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการสแกน PDF: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/document-sorting/extract-duplicate-data', methods=['POST'])
def extract_duplicate_data():
    """อ่านข้อมูลจากไฟล์ PDF ที่มีชื่อซ้ำกัน"""
    try:
        from pdf_reader import PDFReader
        
        data = request.json
        file_paths = data.get('file_paths', [])  # รายการ path ของไฟล์ที่ชื่อซ้ำกัน
        
        if not file_paths or not isinstance(file_paths, list):
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุ file_paths เป็นรายการ'
            }), 400
        
        pdf_reader = PDFReader()
        extracted_data = []
        
        for file_path in file_paths:
            try:
                pdf_file = Path(file_path)
                if not pdf_file.exists():
                    extracted_data.append({
                        'file_path': file_path,
                        'success': False,
                        'error': 'ไม่พบไฟล์'
                    })
                    continue
                
                # อ่านข้อมูลจาก PDF
                pdf_data = pdf_reader.read_pdf(pdf_file)
                
                extracted_data.append({
                    'file_path': file_path,
                    'filename': pdf_file.name,
                    'folder': str(pdf_file.parent),
                    'success': True,
                    'data': pdf_data
                })
            
            except Exception as e:
                extracted_data.append({
                    'file_path': file_path,
                    'success': False,
                    'error': str(e)
                })
        
        return jsonify({
            'success': True,
            'extracted_data': extracted_data
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการอ่านข้อมูล: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/document-sorting/extract-keyword-data', methods=['POST'])
def extract_keyword_data():
    """อ่านข้อมูลจากไฟล์ PDF ที่มีคีย์เวิร์ดเดียวกันและรวมข้อมูล"""
    try:
        import re
        from pdf_reader import PDFReader
        from datetime import datetime
        
        data = request.json
        keyword = data.get('keyword', '').strip()
        file_paths = data.get('file_paths', [])  # รายการ path ของไฟล์ที่มีคีย์เวิร์ดเดียวกัน
        
        if not file_paths or not isinstance(file_paths, list):
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุ file_paths เป็นรายการ'
            }), 400
        
        if not keyword:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุ keyword'
            }), 400
        
        pdf_reader = PDFReader()
        extracted_data = []
        merged_data = {}
        
        # ดึงข้อมูลจากทุกไฟล์
        for file_path in file_paths:
            try:
                pdf_file = Path(file_path)
                if not pdf_file.exists():
                    extracted_data.append({
                        'file_path': file_path,
                        'success': False,
                        'error': 'ไม่พบไฟล์'
                    })
                    continue
                
                # อ่านข้อมูลจาก PDF
                pdf_data = pdf_reader.read_pdf(pdf_file)
                
                extracted_data.append({
                    'file_path': file_path,
                    'filename': pdf_file.name,
                    'folder': str(pdf_file.parent),
                    'success': True,
                    'data': pdf_data
                })
                
                # รวมข้อมูล (ใช้ข้อมูลแรกที่พบ)
                if pdf_data:
                    if pdf_data.get('company_name') and not merged_data.get('company_name'):
                        merged_data['company_name'] = pdf_data.get('company_name')
                    if pdf_data.get('tax_id') and not merged_data.get('tax_id'):
                        merged_data['tax_id'] = pdf_data.get('tax_id')
                    if pdf_data.get('document_number') and not merged_data.get('document_number'):
                        merged_data['document_number'] = pdf_data.get('document_number')
                    if pdf_data.get('document_date') and not merged_data.get('document_date'):
                        merged_data['document_date'] = pdf_data.get('document_date')
            
            except Exception as e:
                extracted_data.append({
                    'file_path': file_path,
                    'success': False,
                    'error': str(e)
                })
        
        # สร้างชื่อไฟล์อัตโนมัติตามคีย์เวิร์ด
        suggested_filename = generate_filename_by_keyword(keyword, extracted_data, merged_data)
        
        return jsonify({
            'success': True,
            'keyword': keyword,
            'extracted_data': extracted_data,
            'merged_data': merged_data,
            'suggested_filename': suggested_filename
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการอ่านข้อมูลตามคีย์เวิร์ด: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def generate_filename_by_keyword(keyword, extracted_data, merged_data):
    """สร้างชื่อไฟล์อัตโนมัติตามคีย์เวิร์ด"""
    try:
        import re
        from datetime import datetime
        
        # ดึงข้อมูลจากไฟล์แรกที่สำเร็จ
        base_filename = None
        document_number = None
        
        for extracted in extracted_data:
            if extracted.get('success') and extracted.get('filename'):
                base_filename = extracted.get('filename')
                # ลองดึง document number จากชื่อไฟล์
                if base_filename:
                    # Pattern 1: EXC-2511-008 หรือ EXC-2511-008_xxx (ดึงส่วนแรกก่อน underscore)
                    match = re.search(r'([A-Z]+-\d{4}-\d{3})', base_filename, re.IGNORECASE)
                    if match:
                        document_number = match.group(1).upper()
                        break
                    
                    # Pattern 2: รูปแบบอื่นๆ เช่น EXC-2511-008
                    match2 = re.search(r'([A-Z]+\d*-\d+-\d+)', base_filename, re.IGNORECASE)
                    if match2:
                        document_number = match2.group(1).upper()
                        break
                
                # ลองดึงจากข้อมูลที่อ่านได้
                if extracted.get('data') and extracted['data'].get('document_number'):
                    document_number = extracted['data'].get('document_number')
                    break
        
        # ถ้ามี merged_data ให้ใช้
        if not document_number and merged_data.get('document_number'):
            document_number = merged_data.get('document_number')
        
        # สร้างชื่อไฟล์
        today = datetime.now().strftime('%Y%m%d')
        
        # ใช้รูปแบบ: {KEYWORD}_{DOCUMENT_NUMBER}_merged_{DATE}.pdf
        if document_number:
            # ลบช่องว่างและอักขระพิเศษออก
            doc_clean = re.sub(r'[^\w\-]', '_', str(document_number))
            suggested_filename = f"{keyword}_{doc_clean}_merged_{today}.pdf"
        else:
            suggested_filename = f"{keyword}_merged_{today}.pdf"
        
        return suggested_filename
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการสร้างชื่อไฟล์: {e}", exc_info=True)
        # คืนค่าแบบง่าย
        from datetime import datetime
        today = datetime.now().strftime('%Y%m%d')
        return f"{keyword}_merged_{today}.pdf"


@app.route('/api/document-sorting/move-file', methods=['POST'])
def move_file_api():
    """ย้ายไฟล์ไปยังโฟลเดอร์ปลายทาง"""
    try:
        import shutil
        
        data = request.json
        source_path = data.get('source_path', '').strip()
        destination_folder = data.get('destination_folder', '').strip()
        new_filename = data.get('new_filename', None)  # ถ้าต้องการเปลี่ยนชื่อไฟล์
        
        if not source_path or not destination_folder:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุ source_path และ destination_folder'
            }), 400
        
        source_file = Path(source_path)
        if not source_file.exists():
            return jsonify({
                'success': False,
                'error': f'ไม่พบไฟล์ต้นฉบับ: {source_path}'
            }), 404
        
        dest_folder = Path(destination_folder)
        dest_folder.mkdir(parents=True, exist_ok=True)
        
        # กำหนดชื่อไฟล์ปลายทาง
        if new_filename:
            dest_file = dest_folder / new_filename
        else:
            dest_file = dest_folder / source_file.name
        
        # ตรวจสอบว่าไฟล์ปลายทางมีอยู่แล้วหรือไม่
        if dest_file.exists():
            return jsonify({
                'success': False,
                'error': f'ไฟล์ปลายทางมีอยู่แล้ว: {dest_file}'
            }), 400
        
        # ย้ายไฟล์
        shutil.move(str(source_file), str(dest_file))
        
        return jsonify({
            'success': True,
            'message': f'ย้ายไฟล์สำเร็จ: {source_file.name} -> {dest_file}',
            'destination_path': str(dest_file)
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการย้ายไฟล์: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/document-sorting/create-folder-and-move', methods=['POST'])
def create_folder_and_move_files():
    """สร้างโฟลเดอร์ตามคีย์เวิร์ดและย้ายไฟล์ทั้งหมดไปยังโฟลเดอร์นั้น"""
    try:
        import shutil
        
        data = request.json
        keyword = data.get('keyword', '').strip()
        base_path = data.get('base_path', '').strip()
        folder_name = data.get('folder_name', keyword).strip()
        file_paths = data.get('file_paths', [])
        
        if not base_path:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุ base_path'
            }), 400
        
        if not folder_name:
            folder_name = keyword or 'UNKNOWN'
        
        if not file_paths or not isinstance(file_paths, list):
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุ file_paths เป็นรายการ'
            }), 400
        
        base_folder = Path(base_path)
        if not base_folder.exists():
            return jsonify({
                'success': False,
                'error': f'ไม่พบโฟลเดอร์: {base_path}'
            }), 404
        
        # สร้างโฟลเดอร์ใหม่ในตำแหน่งเดียวกับ base_path
        new_folder = base_folder / folder_name
        new_folder.mkdir(parents=True, exist_ok=True)
        
        # ย้ายไฟล์ทั้งหมดไปยังโฟลเดอร์ใหม่
        moved_count = 0
        error_count = 0
        errors = []
        
        for file_path in file_paths:
            try:
                source_file = Path(file_path)
                if not source_file.exists():
                    error_count += 1
                    errors.append(f'ไม่พบไฟล์: {file_path}')
                    continue
                
                # ตรวจสอบว่าไฟล์ไม่ได้อยู่ในโฟลเดอร์ปลายทางอยู่แล้ว
                if source_file.parent == new_folder:
                    # ไฟล์อยู่ในโฟลเดอร์ปลายทางแล้ว ไม่ต้องย้าย
                    moved_count += 1
                    continue
                
                # ย้ายไฟล์ไปยังโฟลเดอร์ใหม่
                dest_file = new_folder / source_file.name
                
                # ถ้าไฟล์ปลายทางมีอยู่แล้ว ให้เพิ่มหมายเลขต่อท้าย
                if dest_file.exists():
                    name_part = source_file.stem
                    ext_part = source_file.suffix
                    counter = 1
                    while dest_file.exists():
                        new_name = f"{name_part}_{counter}{ext_part}"
                        dest_file = new_folder / new_name
                        counter += 1
                
                shutil.move(str(source_file), str(dest_file))
                moved_count += 1
            
            except Exception as e:
                error_count += 1
                error_msg = f'ไม่สามารถย้ายไฟล์ {file_path}: {str(e)}'
                errors.append(error_msg)
                logger.error(f"❌ {error_msg}", exc_info=True)
        
        return jsonify({
            'success': True,
            'folder_path': str(new_folder),
            'folder_name': folder_name,
            'moved_count': moved_count,
            'error_count': error_count,
            'total_files': len(file_paths),
            'errors': errors if errors else None
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการสร้างโฟลเดอร์และย้ายไฟล์: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/document-sorting/list-folders', methods=['POST'])
def list_folders_api():
    """ดึงรายการโฟลเดอร์ทั้งหมดใน path"""
    try:
        data = request.json
        base_path = data.get('base_path', '').strip()
        include_subfolders = data.get('include_subfolders', False)
        
        if not base_path:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุ base_path'
            }), 400
        
        base_folder = Path(base_path)
        if not base_folder.exists():
            return jsonify({
                'success': False,
                'error': f'ไม่พบโฟลเดอร์: {base_path}'
            }), 404
        
        folders = set()
        if include_subfolders:
            for item in base_folder.rglob('*'):
                if item.is_dir():
                    folders.add(str(item))
        else:
            for item in base_folder.iterdir():
                if item.is_dir():
                    folders.add(str(item))
        
        return jsonify({
            'success': True,
            'folders': sorted(list(folders))
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการดึงรายการโฟลเดอร์: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ===== API: ระบบตรวจภาษี =====
@app.route('/api/auditcheck/companies/list', methods=['GET'])
@app.route('/api/auditcheck/companies', methods=['GET'])
def get_auditcheck_companies():
    """ดึงรายชื่อบริษัทจากโฟลเดอร์หลัก"""
    # ถ้ามี parameters build, company_name, หรือ customer ให้ redirect ไปที่ get_auditcheck_company
    if request.args.get('build') or request.args.get('company_name') or request.args.get('customer'):
        return get_auditcheck_company()
    try:
        base_paths = [
            Path("V:/A.โฟร์เดอร์หลัก"),
            Path("V:/AA.โฟรเดอร์หลัก"),
            Path("V:/AAA.โฟรเดอร์หลัก")
        ]
        
        # โฟลเดอร์ที่ต้องกรองออก
        skip_folders = [
            "#recycle",
            "#snapshot",
            "A.BMT",
            "A.BMUT",
            "A.Buildmeup",
            "A.งานนอก"
        ]
        
        companies = set()
        
        for base_path in base_paths:
            if not base_path.exists():
                continue
            
            # อ่านชื่อโฟลเดอร์ทั้งหมดในโฟลเดอร์หลัก
            try:
                for item in base_path.iterdir():
                    if item.is_dir():
                        folder_name = item.name
                        
                        # กรองชื่อที่มีคำ "ยกเลิก" ออก
                        if "ยกเลิก" in folder_name:
                            continue
                        
                        # กรองโฟลเดอร์ที่ไม่ต้องการออก
                        should_skip = False
                        for skip_folder in skip_folders:
                            if skip_folder in folder_name or folder_name == skip_folder:
                                should_skip = True
                                break
                        
                        if not should_skip:
                            companies.add(folder_name)
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถอ่านโฟลเดอร์ {base_path}: {e}")
                continue
        
        return jsonify({
            'success': True,
            'companies': sorted(list(companies))
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการดึงรายชื่อบริษัท: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/databases', methods=['GET'])
def get_auditcheck_databases():
    """ดึงรายการฐานข้อมูลทั้งหมดและฐานข้อมูลที่เชื่อมโยงกับบริษัท"""
    try:
        import json
        from pathlib import Path
        
        db_config_file = Path('cache') / 'auditcheck_databases.json'
        company_db_mapping_file = Path('cache') / 'auditcheck_company_database.json'
        
        # โหลดฐานข้อมูลทั้งหมด
        databases = []
        if db_config_file.exists():
            try:
                with open(db_config_file, 'r', encoding='utf-8') as f:
                    databases = json.load(f)
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถอ่านไฟล์ฐานข้อมูล: {e}")
        
        # โหลดการเชื่อมโยงบริษัท-ฐานข้อมูล
        company_database = None
        company = request.args.get('company', '')
        
        if company and company_db_mapping_file.exists():
            try:
                with open(company_db_mapping_file, 'r', encoding='utf-8') as f:
                    mappings = json.load(f)
                    company_database = mappings.get(company)
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถอ่านไฟล์การเชื่อมโยง: {e}")
        
        return jsonify({
            'success': True,
            'databases': databases,
            'companyDatabase': company_database
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการดึงฐานข้อมูล: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/databases', methods=['POST'])
def save_auditcheck_database():
    """บันทึกหรือแก้ไขฐานข้อมูล"""
    try:
        import json
        from pathlib import Path
        import uuid
        from datetime import datetime
        
        data = request.json
        if not data:
            return jsonify({
                'success': False,
                'error': 'ไม่มีข้อมูลส่งมา'
            }), 400
        
        db_config_file = Path('cache') / 'auditcheck_databases.json'
        cache_dir = Path('cache')
        cache_dir.mkdir(exist_ok=True)
        
        # โหลดฐานข้อมูลที่มีอยู่
        databases = []
        if db_config_file.exists():
            try:
                with open(db_config_file, 'r', encoding='utf-8') as f:
                    databases = json.load(f)
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถอ่านไฟล์ฐานข้อมูล: {e}")
        
        db_id = data.get('id')
        db_type = data.get('type', '')
        
        # สร้าง object ฐานข้อมูล
        db_config = {
            'id': db_id or str(uuid.uuid4()),
            'name': data.get('name', ''),
            'type': db_type,
            'host': data.get('host', '') if db_type != 'sqlite' else '',
            'port': data.get('port', '') if db_type != 'sqlite' else '',
            'database': data.get('database', '') if db_type != 'sqlite' else '',
            'username': data.get('username', '') if db_type != 'sqlite' else '',
            'password': data.get('password', ''),  # เก็บ password (ควรเข้ารหัสในอนาคต)
            'sqlite_path': data.get('sqlite_path', '') if db_type == 'sqlite' else '',
            'description': data.get('description', ''),
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat()
        }
        
        # ถ้ามี id แสดงว่าเป็นการแก้ไข
        if db_id:
            index = next((i for i, db in enumerate(databases) if db.get('id') == db_id), -1)
            if index >= 0:
                # เก็บ created_at เดิม
                db_config['created_at'] = databases[index].get('created_at', db_config['created_at'])
                databases[index] = db_config
            else:
                return jsonify({
                    'success': False,
                    'error': 'ไม่พบฐานข้อมูลที่ต้องการแก้ไข'
                }), 404
        else:
            # เพิ่มใหม่
            databases.append(db_config)
        
        # บันทึกไฟล์
        with open(db_config_file, 'w', encoding='utf-8') as f:
            json.dump(databases, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True,
            'database': db_config
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการบันทึกฐานข้อมูล: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/company-database', methods=['POST'])
def save_company_database_mapping():
    """บันทึกการเชื่อมโยงบริษัทกับฐานข้อมูล"""
    try:
        import json
        from pathlib import Path
        
        data = request.json
        if not data:
            return jsonify({
                'success': False,
                'error': 'ไม่มีข้อมูลส่งมา'
            }), 400
        
        company = data.get('company', '')
        database_id = data.get('database_id', '')
        
        if not company:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุบริษัท'
            }), 400
        
        mapping_file = Path('cache') / 'auditcheck_company_database.json'
        cache_dir = Path('cache')
        cache_dir.mkdir(exist_ok=True)
        
        # โหลดการเชื่อมโยงที่มีอยู่
        mappings = {}
        if mapping_file.exists():
            try:
                with open(mapping_file, 'r', encoding='utf-8') as f:
                    mappings = json.load(f)
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถอ่านไฟล์การเชื่อมโยง: {e}")
        
        # อัพเดทการเชื่อมโยง
        if database_id:
            mappings[company] = database_id
        else:
            # ถ้าไม่มี database_id ให้ลบการเชื่อมโยง
            mappings.pop(company, None)
        
        # บันทึกไฟล์
        with open(mapping_file, 'w', encoding='utf-8') as f:
            json.dump(mappings, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการบันทึกการเชื่อมโยง: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/databases/<db_id>', methods=['DELETE'])
def delete_auditcheck_database(db_id):
    """ลบฐานข้อมูล"""
    try:
        import json
        from pathlib import Path
        
        db_config_file = Path('cache') / 'auditcheck_databases.json'
        company_db_mapping_file = Path('cache') / 'auditcheck_company_database.json'
        
        if not db_config_file.exists():
            return jsonify({
                'success': False,
                'error': 'ไม่พบไฟล์ฐานข้อมูล'
            }), 404
        
        # โหลดฐานข้อมูล
        try:
            with open(db_config_file, 'r', encoding='utf-8') as f:
                databases = json.load(f)
        except Exception as e:
            return jsonify({
                'success': False,
                'error': f'ไม่สามารถอ่านไฟล์ฐานข้อมูล: {e}'
            }), 500
        
        # หาและลบฐานข้อมูล
        original_count = len(databases)
        databases = [db for db in databases if db.get('id') != db_id]
        
        if len(databases) == original_count:
            return jsonify({
                'success': False,
                'error': 'ไม่พบฐานข้อมูลที่ต้องการลบ'
            }), 404
        
        # ลบการเชื่อมโยงบริษัท-ฐานข้อมูลที่ใช้ฐานข้อมูลนี้
        if company_db_mapping_file.exists():
            try:
                with open(company_db_mapping_file, 'r', encoding='utf-8') as f:
                    mappings = json.load(f)
                
                # ลบการเชื่อมโยงที่ใช้ฐานข้อมูลนี้
                mappings = {company: mapped_db_id for company, mapped_db_id in mappings.items() if mapped_db_id != db_id}
                
                with open(company_db_mapping_file, 'w', encoding='utf-8') as f:
                    json.dump(mappings, f, ensure_ascii=False, indent=2)
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถอัพเดทการเชื่อมโยง: {e}")
        
        # บันทึกไฟล์ฐานข้อมูล
        with open(db_config_file, 'w', encoding='utf-8') as f:
            json.dump(databases, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            'success': True
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการลบฐานข้อมูล: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/company', methods=['GET'])
def get_auditcheck_company():
    """ดึงข้อมูลบริษัทตาม Build หรือชื่อบริษัท"""
    try:
        import json
        import re
        from pathlib import Path
        
        build = request.args.get('build', '').strip()
        company_name = request.args.get('company_name', '').strip()
        customer = request.args.get('customer', '').strip()
        
        # ถ้าไม่มี customer ให้ใช้ company_name แทน
        if not customer and company_name:
            customer = company_name
        
        logger.info(f"📥 รับ request: build={build}, company_name={company_name}, customer={customer}")
        
        # Sanitize customer name สำหรับใช้เป็นชื่อไฟล์
        def sanitize_filename(name):
            if not name:
                return 'default'
            # ลบอักขระพิเศษและแทนที่ด้วย underscore
            sanitized = re.sub(r'[<>:"/\\|?*]', '_', name)
            sanitized = re.sub(r'\s+', '_', sanitized)
            # จำกัดความยาว
            return sanitized[:100] if len(sanitized) > 100 else sanitized
        
        customer_sanitized = sanitize_filename(customer)
        logger.info(f"🔤 Customer sanitized: {customer_sanitized}")
        
        customer_data_dir = Path('Customer_data')
        customer_data_dir.mkdir(exist_ok=True)
        
        # ลองอ่านไฟล์แยกตาม customer ก่อน
        companies_file = customer_data_dir / f'auditcheck_companies_{customer_sanitized}.json'
        
        logger.info(f"📁 กำลังหาไฟล์: {companies_file}")
        logger.info(f"📁 ไฟล์มีอยู่หรือไม่: {companies_file.exists()}")
        
        # ถ้าไม่มีไฟล์แยกตาม customer ให้ลองอ่านไฟล์เก่า (fallback)
        if not companies_file.exists():
            # ลองอ่านไฟล์เก่าจาก Customer_data (ไฟล์ที่ย้ายมา)
            old_file = customer_data_dir / 'auditcheck_companies.json'
            logger.info(f"📁 กำลังหาไฟล์เก่า: {old_file}")
            logger.info(f"📁 ไฟล์เก่ามีอยู่หรือไม่: {old_file.exists()}")
            
            if old_file.exists():
                companies_file = old_file
                logger.info(f"📂 ใช้ไฟล์เก่า: {companies_file}")
            else:
                logger.warning(f"⚠️ ไม่พบไฟล์ข้อมูลบริษัท: {companies_file} หรือ {old_file}")
                return jsonify({
                    'success': False,
                    'company': None
                }), 200
        else:
            logger.info(f"📂 ใช้ไฟล์แยกตาม customer: {companies_file}")
        
        try:
            with open(companies_file, 'r', encoding='utf-8') as f:
                companies = json.load(f)
        except Exception as e:
            logger.warning(f"⚠️ ไม่สามารถอ่านไฟล์ข้อมูลบริษัท: {e}")
            return jsonify({
                'success': False,
                'company': None
            }), 200
        
        # ค้นหาบริษัทตาม Build หรือชื่อบริษัท
        found_company = None
        
        # ตรวจสอบว่า companies เป็น list หรือไม่
        if not isinstance(companies, list):
            companies = [companies] if companies else []
        
        logger.info(f"🔍 ค้นหาบริษัท: build={build}, company_name={company_name}, จำนวนข้อมูล={len(companies)}")
        
        if build:
            # ค้นหาตาม Build
            found_company = next((comp for comp in companies if comp.get('build', '').strip() == build), None)
            if found_company:
                logger.info(f"✅ พบบริษัทตาม Build: {build}")
        
        if not found_company and company_name:
            # ถ้าไม่พบตาม Build ให้ค้นหาตามชื่อบริษัท (ใช้ partial match)
            company_name_lower = company_name.lower().strip()
            for comp in companies:
                comp_name = comp.get('company_name', '').strip()
                comp_name_lower = comp_name.lower()
                
                # ลองหลายวิธีในการ match
                if (company_name_lower in comp_name_lower or 
                    comp_name_lower in company_name_lower or
                    company_name_lower.replace(' ', '') in comp_name_lower.replace(' ', '') or
                    comp_name_lower.replace(' ', '') in company_name_lower.replace(' ', '')):
                    found_company = comp
                    logger.info(f"✅ พบบริษัทตามชื่อ: {comp_name} (ค้นหา: {company_name})")
                    break
        
        if not found_company:
            logger.warning(f"⚠️ ไม่พบข้อมูลบริษัท: build={build}, company_name={company_name}")
            # แสดงข้อมูลที่มีในไฟล์เพื่อ debug
            if companies:
                logger.info(f"📋 ข้อมูลที่มีในไฟล์: {[comp.get('build', '') + ' - ' + comp.get('company_name', '') for comp in companies]}")
        
        return jsonify({
            'success': True,
            'company': found_company
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการดึงข้อมูลบริษัท: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/companies', methods=['POST'])
def save_auditcheck_company():
    """บันทึกข้อมูลบริษัท"""
    try:
        import json
        import re
        from pathlib import Path
        from datetime import datetime
        
        data = request.json
        if not data:
            return jsonify({
                'success': False,
                'error': 'ไม่มีข้อมูลส่งมา'
            }), 400
        
        # ดึง customer จาก data หรือใช้ company_name แทน
        customer = data.get('customer', '').strip()
        company_name = data.get('company_name', '').strip()
        if not customer and company_name:
            customer = company_name
        
        # Sanitize customer name สำหรับใช้เป็นชื่อไฟล์
        def sanitize_filename(name):
            if not name:
                return 'default'
            # ลบอักขระพิเศษและแทนที่ด้วย underscore
            sanitized = re.sub(r'[<>:"/\\|?*]', '_', name)
            sanitized = re.sub(r'\s+', '_', sanitized)
            # จำกัดความยาว
            return sanitized[:100] if len(sanitized) > 100 else sanitized
        
        customer_sanitized = sanitize_filename(customer)
        customer_data_dir = Path('Customer_data')
        customer_data_dir.mkdir(exist_ok=True)
        companies_file = customer_data_dir / f'auditcheck_companies_{customer_sanitized}.json'
        
        # โหลดข้อมูลบริษัทที่มีอยู่
        companies = []
        if companies_file.exists():
            try:
                with open(companies_file, 'r', encoding='utf-8') as f:
                    companies = json.load(f)
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถอ่านไฟล์ข้อมูลบริษัท: {e}")
        
        # ตรวจสอบว่ามี Build นี้อยู่แล้วหรือไม่
        build = data.get('build', '').strip()
        if not build:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุ Build'
            }), 400
        
        # หาว่ามี Build นี้อยู่แล้วหรือไม่
        existing_index = next((i for i, comp in enumerate(companies) if comp.get('build') == build), -1)
        
        # สร้าง object ข้อมูลบริษัท
        company_data = {
            'build': build,
            'company_name': data.get('company_name', '').strip(),
            'tax_id': data.get('tax_id', '').strip(),
            'vat_status': data.get('vat_status', '').strip(),
            'vat_registration_date': data.get('vat_registration_date', '').strip(),
            'company_address': data.get('company_address', '').strip(),
            'updated_at': datetime.now().isoformat()
        }
        
        # Validation
        if not company_data['company_name']:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุชื่อบริษัท'
            }), 400
        
        if not company_data['vat_status']:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุสถานะบริษัทจดภาษีมูลค่าเพิ่ม'
            }), 400
        
        if not company_data['company_address']:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุที่อยู่บริษัท'
            }), 400
        
        # ถ้ามี Build นี้อยู่แล้ว ให้อัพเดท
        if existing_index >= 0:
            # เก็บ created_at เดิม
            company_data['created_at'] = companies[existing_index].get('created_at', datetime.now().isoformat())
            companies[existing_index] = company_data
            action = 'อัพเดท'
        else:
            # เพิ่มใหม่
            company_data['created_at'] = datetime.now().isoformat()
            companies.append(company_data)
            action = 'เพิ่ม'
        
        # บันทึกไฟล์
        with open(companies_file, 'w', encoding='utf-8') as f:
            json.dump(companies, f, ensure_ascii=False, indent=2)
        
        logger.info(f"✅ {action}ข้อมูลบริษัท: {build} - {company_data['company_name']}")
        
        return jsonify({
            'success': True,
            'message': f'{action}ข้อมูลบริษัทสำเร็จ',
            'company': company_data
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการบันทึกข้อมูลบริษัท: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/check-files', methods=['POST'])
def check_audit_files():
    """ตรวจสอบไฟล์ตามโครงสร้างโฟลเดอร์: บัญชี > 003-ภาษี > ภ.พ.30 > ปี > เดือน-ปี"""
    try:
        data = request.json
        tax_month = data.get('taxMonth', '')  # Format: YYYY-MM
        company = data.get('company', '')
        
        if not tax_month or not company:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเดือนภาษีและชื่อบริษัท'
            }), 400
        
        # แปลงเดือนจาก YYYY-MM
        # ตรวจสอบรูปแบบก่อน split
        if '-' not in tax_month:
            return jsonify({
                'success': False,
                'error': f'รูปแบบเดือนภาษีไม่ถูกต้อง: {tax_month} (คาดหวังรูปแบบ YYYY-MM)'
            }), 400
        
        parts = tax_month.split('-')
        if len(parts) != 2:
            return jsonify({
                'success': False,
                'error': f'รูปแบบเดือนภาษีไม่ถูกต้อง: {tax_month} (คาดหวังรูปแบบ YYYY-MM)'
            }), 400
        
        year, month = parts
        try:
            year_int = int(year)
            month_int = int(month)
        except ValueError as e:
            return jsonify({
                'success': False,
                'error': f'ไม่สามารถแปลงปีหรือเดือนเป็นตัวเลขได้: {tax_month}'
            }), 400
        
        # สร้างรูปแบบเดือน-ปี (รองรับหลายรูปแบบ)
        month_year_patterns = [
            f"{month_int:02d}-{year_int}",  # 01-2026 (เดือน-ปี)
            f"{month_int}-{year_int}",       # 1-2026 (เดือน-ปี)
            f"{year_int}-{month_int:02d}",   # 2026-01 (ปี-เดือน)
            f"{year_int}-{month_int}",       # 2026-1 (ปี-เดือน)
        ]
        
        # ค้นหาในโฟลเดอร์ต่างๆ
        base_paths = [
            Path(f"V:/A.โฟร์เดอร์หลัก/{company}"),
            Path(f"V:/AA.โฟรเดอร์หลัก/{company}"),
            Path(f"V:/AAA.โฟรเดอร์หลัก/{company}")
        ]
        
        found_files = []
        missing_files = []
        folder_structure = []
        
        for base_path in base_paths:
            if not base_path.exists():
                continue
            
            # โครงสร้าง: บัญชี > 003-ภาษี > ภ.พ.30 > ปี > เดือน-ปี
            account_folder = base_path / "บัญชี"
            if not account_folder.exists():
                # แสดงรายการโฟลเดอร์ที่มีอยู่ใน base_path
                existing_items = []
                try:
                    for item in base_path.iterdir():
                        if item.is_dir():
                            existing_items.append(item.name)
                except:
                    pass
                missing_files.append({
                    'type': 'folder',
                    'message': f"ไม่พบโฟลเดอร์: {account_folder}",
                    'expected': 'บัญชี',
                    'existing_items': sorted(existing_items),
                    'parent_path': str(base_path)
                })
                continue
            
            tax_folder = account_folder / "003-ภาษี"
            if not tax_folder.exists():
                existing_items = []
                try:
                    for item in account_folder.iterdir():
                        if item.is_dir():
                            existing_items.append(item.name)
                except:
                    pass
                missing_files.append({
                    'type': 'folder',
                    'message': f"ไม่พบโฟลเดอร์: {tax_folder}",
                    'expected': '003-ภาษี',
                    'existing_items': sorted(existing_items),
                    'parent_path': str(account_folder)
                })
                continue
            
            pph30_folder = tax_folder / "ภ.พ.30"
            if not pph30_folder.exists():
                existing_items = []
                try:
                    for item in tax_folder.iterdir():
                        if item.is_dir():
                            existing_items.append(item.name)
                except:
                    pass
                missing_files.append({
                    'type': 'folder',
                    'message': f"ไม่พบโฟลเดอร์: {pph30_folder}",
                    'expected': 'ภ.พ.30',
                    'existing_items': sorted(existing_items),
                    'parent_path': str(tax_folder)
                })
                continue
            
            year_folder = pph30_folder / year
            if not year_folder.exists():
                existing_items = []
                try:
                    for item in pph30_folder.iterdir():
                        if item.is_dir():
                            existing_items.append(item.name)
                except:
                    pass
                missing_files.append({
                    'type': 'folder',
                    'message': f"ไม่พบโฟลเดอร์: {year_folder}",
                    'expected': year,
                    'existing_items': sorted(existing_items),
                    'parent_path': str(pph30_folder)
                })
                continue
            
            # ค้นหาโฟลเดอร์เดือน-ปี
            month_year_folder = None
            for pattern in month_year_patterns:
                potential_folder = year_folder / pattern
                if potential_folder.exists():
                    month_year_folder = potential_folder
                    break
            
            if not month_year_folder:
                # แสดงรายการโฟลเดอร์ที่มีอยู่ในโฟลเดอร์ปี
                existing_folders = []
                if year_folder.exists():
                    for item in year_folder.iterdir():
                        if item.is_dir():
                            existing_folders.append(item.name)
                    missing_files.append({
                        'type': 'month_year_folder',
                        'message': f"ไม่พบโฟลเดอร์เดือน-ปี: {year}/{month_year_patterns[0]} หรือ {year}/{month_year_patterns[1]}",
                        'expected_patterns': month_year_patterns,
                        'existing_folders': sorted(existing_folders),
                        'year_folder_path': str(year_folder)
                    })
                else:
                    missing_files.append(f"ไม่พบโฟลเดอร์เดือน-ปี: {year}/{month_year_patterns[0]} หรือ {year}/{month_year_patterns[1]}")
                continue
            
            # นับไฟล์ทั้งหมดในโฟลเดอร์เดือน-ปี
            all_files = []
            try:
                for file_path in month_year_folder.iterdir():
                    if file_path.is_file():
                        all_files.append({
                            'name': file_path.name,
                            'path': str(file_path),
                            'size': file_path.stat().st_size,
                            'extension': file_path.suffix.lower()
                        })
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถอ่านไฟล์ในโฟลเดอร์ {month_year_folder}: {e}")
            
            found_files.extend(all_files)
            folder_structure.append({
                'base_path': str(base_path),
                'account_folder': str(account_folder),
                'tax_folder': str(tax_folder),
                'pph30_folder': str(pph30_folder),
                'year_folder': str(year_folder),
                'month_year_folder': str(month_year_folder),
                'month_year_folder_name': month_year_folder.name,
                'files_count': len(all_files),
                'files': all_files
            })
        
        return jsonify({
            'success': True,
            'taxMonth': tax_month,
            'company': company,
            'foundFiles': found_files,
            'missingFiles': missing_files,
            'folderStructure': folder_structure,
            'totalFiles': len(found_files),
            'hasFiles': len(found_files) > 0
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการตรวจสอบไฟล์: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/check-trial-balance', methods=['POST'])
def check_trial_balance():
    """ตรวจสอบว่ามีไฟล์งบทดลองหรือไม่"""
    try:
        data = request.json
        tax_month = data.get('taxMonth', '')  # Format: YYYY-MM
        company = data.get('company', '')
        
        if not tax_month or not company:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเดือนภาษีและชื่อบริษัท'
            }), 400
        
        # ใช้ API ใหม่ที่ตรวจสอบตามโครงสร้างโฟลเดอร์
        # โครงสร้าง: บัญชี > 003-ภาษี > ภ.พ.30 > ปี > เดือน-ปี
        year, month = tax_month.split('-')
        year_int = int(year)
        month_int = int(month)
        
        month_year_patterns = [
            f"{month_int:02d}-{year_int}",  # 01-2026
            f"{month_int}-{year_int}",       # 1-2026
        ]
        
        base_paths = [
            Path(f"V:/A.โฟร์เดอร์หลัก/{company}"),
            Path(f"V:/AA.โฟรเดอร์หลัก/{company}"),
            Path(f"V:/AAA.โฟรเดอร์หลัก/{company}")
        ]
        
        trial_balance_files = []
        
        for base_path in base_paths:
            if not base_path.exists():
                continue
            
            try:
                month_year_folder = None
                year_folder = base_path / "บัญชี" / "003-ภาษี" / "ภ.พ.30" / year
                
                if year_folder.exists():
                    for pattern in month_year_patterns:
                        potential_folder = year_folder / pattern
                        if potential_folder.exists():
                            month_year_folder = potential_folder
                            break
                    
                    if month_year_folder:
                        # ค้นหาไฟล์งบทดลอง
                        for file_path in month_year_folder.rglob("*งบทดลอง*"):
                            if file_path.is_file() and file_path.suffix.lower() in ['.xlsx', '.xls', '.pdf']:
                                trial_balance_files.append(str(file_path))
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์งบทดลองใน {base_path}: {e}")
                continue
        
        return jsonify({
            'success': True,
            'exists': len(trial_balance_files) > 0,
            'fileCount': len(trial_balance_files),
            'files': trial_balance_files
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการตรวจสอบไฟล์งบทดลอง: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/check-ocr-files', methods=['POST'])
def check_ocr_files_legacy():
    """API เก่าสำหรับ backward compatibility - redirect ไป check-excel-files"""
    return check_excel_files()


@app.route('/api/auditcheck/check-excel-files', methods=['POST'])
def check_excel_files():
    """ตรวจสอบว่ามีไฟล์ Excel ที่ส่งออกจากระบบ OCR หรือไม่"""
    try:
        data = request.json
        tax_month = data.get('taxMonth', '')  # Format: YYYY-MM
        company = data.get('company', '')
        
        if not tax_month or not company:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเดือนภาษีและชื่อบริษัท'
            }), 400
        
        # แปลงเดือนจาก YYYY-MM
        if '-' not in tax_month:
            return jsonify({
                'success': False,
                'error': f'รูปแบบเดือนภาษีไม่ถูกต้อง: {tax_month}'
            }), 400
        
        parts = tax_month.split('-')
        if len(parts) != 2:
            return jsonify({
                'success': False,
                'error': f'รูปแบบเดือนภาษีไม่ถูกต้อง: {tax_month}'
            }), 400
        
        year, month = parts
        year_int = int(year)
        month_int = int(month)
        
        # สร้างรูปแบบปี-เดือน (รองรับหลายรูปแบบ)
        month_year_patterns = [
            f"{year_int}-{month_int:02d}",   # 2025-10 (ปี-เดือน)
            f"{year_int}-{month_int}",       # 2025-10 (ปี-เดือน)
            f"{month_int:02d}-{year_int}",   # 10-2025 (เดือน-ปี)
            f"{month_int}-{year_int}",       # 10-2025 (เดือน-ปี)
        ]
        
        # ค้นหาไฟล์ Excel ในโครงสร้าง: บัญชี > 002-รายจ่าย > PV > ปี-เดือน
        excel_file_paths = []
        search_results = []
        
        base_paths = [
            Path(f"V:/A.โฟร์เดอร์หลัก/{company}"),
            Path(f"V:/AA.โฟรเดอร์หลัก/{company}"),
            Path(f"V:/AAA.โฟรเดอร์หลัก/{company}")
        ]
        
        for base_path in base_paths:
            if not base_path.exists():
                continue
            
            try:
                # โครงสร้าง: บัญชี > 002-รายจ่าย > PV
                account_folder = base_path / "บัญชี"
                if not account_folder.exists():
                    continue
                
                expense_folder = account_folder / "002-รายจ่าย"
                if not expense_folder.exists():
                    continue
                
                pv_folder = expense_folder / "PV"
                if not pv_folder.exists():
                    continue
                
                # ขั้นตอนที่ 1: มองหาโฟลเดอร์ ปี-เดือน โดยตรง (เช่น 2025-10)
                month_year_folder = None
                for pattern in month_year_patterns:
                    potential_folder = pv_folder / pattern
                    if potential_folder.exists():
                        month_year_folder = potential_folder
                        search_results.append({
                            'base_path': str(base_path),
                            'search_method': 'direct',
                            'folder_path': str(potential_folder),
                            'pattern_found': pattern
                        })
                        break
                
                # ขั้นตอนที่ 2: ถ้าไม่เจอ ให้มองหาโฟลเดอร์ปีก่อน แล้วเข้าไปมองหาโฟลเดอร์ ปี-เดือน
                if not month_year_folder:
                    # ค้นหาโฟลเดอร์ปี (เช่น 2025, 2568)
                    year_folders = []
                    all_pv_folders = []  # เก็บรายการโฟลเดอร์ทั้งหมดใน PV
                    try:
                        for item in pv_folder.iterdir():
                            if item.is_dir():
                                all_pv_folders.append(item.name)
                                # ตรวจสอบว่าเป็นโฟลเดอร์ปีหรือไม่ (อาจเป็น 2025 หรือ 2568)
                                folder_name = item.name
                                if (year in folder_name or 
                                    str(year_int + 543) in folder_name or  # ปี พ.ศ.
                                    folder_name.isdigit()):
                                    year_folders.append(item)
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถอ่านโฟลเดอร์ PV: {e}")
                    
                    # ลองค้นหาในโฟลเดอร์ปีที่พบ
                    for year_folder in year_folders:
                        year_folder_subfolders = []
                        try:
                            for item in year_folder.iterdir():
                                if item.is_dir():
                                    year_folder_subfolders.append(item.name)
                                    for pattern in month_year_patterns:
                                        if pattern in item.name or item.name == pattern:
                                            potential_folder = year_folder / item.name
                                            if potential_folder.exists():
                                                month_year_folder = potential_folder
                                                search_results.append({
                                                    'base_path': str(base_path),
                                                    'search_method': 'year_then_month',
                                                    'year_folder': str(year_folder),
                                                    'folder_path': str(potential_folder),
                                                    'pattern_found': pattern
                                                })
                                                break
                            if month_year_folder:
                                break
                        except Exception as e:
                            logger.warning(f"⚠️ ไม่สามารถอ่านโฟลเดอร์ปี {year_folder}: {e}")
                    
                    # ถ้ายังไม่พบ ให้เก็บข้อมูลโฟลเดอร์ที่มีอยู่เพื่อแสดงผล
                    if not month_year_folder:
                        search_results.append({
                            'base_path': str(base_path),
                            'search_method': 'not_found',
                            'pv_folder': str(pv_folder),
                            'all_pv_folders': sorted(all_pv_folders),
                            'year_folders': [str(f) for f in year_folders],
                            'year_folder_subfolders': year_folder_subfolders if 'year_folder_subfolders' in locals() else []
                        })
                
                # ค้นหาไฟล์ Excel ในโฟลเดอร์เดือน-ปีที่พบ
                if month_year_folder:
                    # ค้นหาโฟลเดอร์ VAT/vat/Vat ภายในโฟลเดอร์เดือน-ปี
                    vat_folders = []
                    vat_folder_info = {
                        'found': False,
                        'folders': [],
                        'month_year_folder': str(month_year_folder)
                    }
                    try:
                        # ค้นหาโฟลเดอร์ที่มีชื่อว่า VAT, vat, หรือ Vat
                        for item in month_year_folder.iterdir():
                            if item.is_dir() and item.name.lower() == "vat":
                                vat_folders.append(item)
                                vat_folder_info['folders'].append({
                                    'name': item.name,
                                    'path': str(item)
                                })
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์ VAT ใน {month_year_folder}: {e}")
                    
                    vat_folder_info['found'] = len(vat_folders) > 0
                    
                    # ค้นหาไฟล์ Excel ในโฟลเดอร์ VAT/vat/Vat
                    if vat_folders:
                        for vat_folder in vat_folders:
                            try:
                                excel_files_in_vat = []
                                for excel_file in vat_folder.glob("*.xlsx"):
                                    if excel_file.is_file():
                                        filename_lower = excel_file.name.lower()
                                        # อ่านเฉพาะไฟล์ที่มีคำว่า "ocr" ในชื่อไฟล์เท่านั้น
                                        if "ocr" in filename_lower:
                                            if str(excel_file) not in excel_file_paths:
                                                excel_file_paths.append(str(excel_file))
                                                excel_files_in_vat.append(excel_file.name)
                                
                                # เพิ่มข้อมูลเกี่ยวกับไฟล์ Excel ที่พบในโฟลเดอร์ VAT
                                for vat_info in vat_folder_info['folders']:
                                    if vat_info['path'] == str(vat_folder):
                                        vat_info['excel_files'] = excel_files_in_vat
                            except Exception as e:
                                logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์ Excel ใน {vat_folder}: {e}")
                    else:
                        # ถ้าไม่พบโฟลเดอร์ VAT ให้ค้นหาในโฟลเดอร์เดือน-ปีโดยตรง
                        try:
                            for excel_file in month_year_folder.glob("*.xlsx"):
                                if excel_file.is_file():
                                    filename_lower = excel_file.name.lower()
                                    # อ่านเฉพาะไฟล์ที่มีคำว่า "ocr" ในชื่อไฟล์เท่านั้น
                                    if "ocr" in filename_lower:
                                        if str(excel_file) not in excel_file_paths:
                                            excel_file_paths.append(str(excel_file))
                        except Exception as e:
                            logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์ Excel ใน {month_year_folder}: {e}")
                    
                    # เพิ่มข้อมูลเกี่ยวกับโฟลเดอร์ VAT ใน search_results
                    if search_results:
                        # อัปเดต search_result ล่าสุดด้วยข้อมูล VAT
                        last_result = search_results[-1]
                        last_result['vat_folder_info'] = vat_folder_info
                    else:
                        # ถ้าไม่มี search_results ให้สร้างใหม่
                        search_results.append({
                            'base_path': str(base_path),
                            'search_method': 'direct',
                            'folder_path': str(month_year_folder),
                            'vat_folder_info': vat_folder_info
                        })
                        
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์ Excel ใน {base_path}: {e}")
                continue
        
        # ลบไฟล์ซ้ำ
        excel_file_paths = list(set(excel_file_paths))
        
        # อ่านข้อมูลจากไฟล์ Excel (ถ้ามี)
        excel_data = []
        if excel_file_paths:
            try:
                import openpyxl
                for excel_path in excel_file_paths:
                    try:
                        wb = openpyxl.load_workbook(excel_path, data_only=True)
                        ws = wb.active
                        
                        # อ่าน header
                        headers = []
                        if ws.max_row > 0:
                            headers = [cell.value for cell in ws[1]]
                        
                        # นับจำนวนแถวข้อมูล (ไม่รวม header)
                        row_count = ws.max_row - 1 if ws.max_row > 1 else 0
                        
                        excel_data.append({
                            'path': excel_path,
                            'headers': headers,
                            'row_count': row_count,
                            'file_size': Path(excel_path).stat().st_size
                        })
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถอ่านไฟล์ Excel {excel_path}: {e}")
                        excel_data.append({
                            'path': excel_path,
                            'error': str(e)
                        })
            except ImportError:
                logger.warning("⚠️ ไม่พบ openpyxl - ไม่สามารถอ่านข้อมูล Excel ได้")
        
        return jsonify({
            'success': True,
            'exists': len(excel_file_paths) > 0,
            'fileCount': len(excel_file_paths),
            'files': excel_file_paths,
            'excelData': excel_data,
            'searchResults': search_results,
            'monthYearPatterns': month_year_patterns
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการตรวจสอบไฟล์ Excel: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/run-ocr', methods=['POST'])
def run_ocr_for_audit():
    """รัน OCR จากไฟล์ PDF ในโฟลเดอร์ VAT และ return ข้อมูล OCR (ไม่สร้าง Excel)"""
    try:
        data = request.json
        tax_month = data.get('taxMonth', '')  # Format: YYYY-MM
        company = data.get('company', '')
        check_only = data.get('checkOnly', False)  # ถ้าเป็น True จะตรวจสอบเท่านั้น ไม่รัน OCR
        ocr_mode = data.get('ocrMode', 'new')  # 'new' = อ่านใหม่ทั้งหมด, 'continue' = อ่านต่อจากที่ค้าง
        
        if not tax_month or not company:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเดือนภาษีและชื่อบริษัท'
            }), 400
        
        # แปลงเดือนจาก YYYY-MM
        if '-' not in tax_month:
            return jsonify({
                'success': False,
                'error': f'รูปแบบเดือนภาษีไม่ถูกต้อง: {tax_month}'
            }), 400
        
        year, month = tax_month.split('-')
        year_int = int(year)
        month_int = int(month)
        
        # สร้างรูปแบบปี-เดือน
        month_year_patterns = [
            f"{year_int}-{month_int:02d}",
            f"{year_int}-{month_int}",
            f"{month_int:02d}-{year_int}",
            f"{month_int}-{year_int}",
        ]
        
        # ค้นหาโฟลเดอร์ VAT
        base_paths = [
            Path(f"V:/A.โฟร์เดอร์หลัก/{company}"),
            Path(f"V:/AA.โฟรเดอร์หลัก/{company}"),
            Path(f"V:/AAA.โฟรเดอร์หลัก/{company}")
        ]
        
        vat_folder = None
        for base_path in base_paths:
            if not base_path.exists():
                continue
            
            try:
                account_folder = base_path / "บัญชี" / "002-รายจ่าย" / "PV"
                if not account_folder.exists():
                    continue
                
                # ค้นหาโฟลเดอร์เดือน-ปี
                month_year_folder = None
                for pattern in month_year_patterns:
                    potential_folder = account_folder / pattern
                    if potential_folder.exists():
                        month_year_folder = potential_folder
                        break
                
                if not month_year_folder:
                    # ค้นหาในโฟลเดอร์ปี
                    for item in account_folder.iterdir():
                        if item.is_dir() and (year in item.name or str(year_int + 543) in item.name):
                            for pattern in month_year_patterns:
                                potential_folder = item / pattern
                                if potential_folder.exists():
                                    month_year_folder = potential_folder
                                    break
                            if month_year_folder:
                                break
                
                if month_year_folder:
                    # ค้นหาโฟลเดอร์ VAT
                    for item in month_year_folder.iterdir():
                        if item.is_dir() and item.name.lower() == "vat":
                            vat_folder = item
                            break
                    
                    if not vat_folder:
                        vat_folder = month_year_folder
                    
                    break
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์ VAT ใน {base_path}: {e}")
                continue
        
        if not vat_folder:
            return jsonify({
                'success': False,
                'error': f'ไม่พบโฟลเดอร์ VAT สำหรับเดือน {tax_month}'
            }), 404
        
        # ค้นหาไฟล์ PDF ในโฟลเดอร์ VAT (รวมโฟลเดอร์ย่อยด้วย)
        include_subfolders = data.get('include_subfolders', True)  # ค่าเริ่มต้นเป็น True เพื่ออ่านโฟลเดอร์ย่อย
        selected_folders = data.get('selectedFolders', None)  # รายการโฟลเดอร์ที่เลือก (ถ้ามี)
        
        # รายการนามสกุลไฟล์ที่รองรับ OCR
        supported_extensions = ['*.pdf', '*.jpg', '*.jpeg', '*.png', '*.gif', '*.bmp', '*.tiff', '*.tif']
        supported_extensions_upper = ['*.PDF', '*.JPG', '*.JPEG', '*.PNG', '*.GIF', '*.BMP', '*.TIFF', '*.TIF']
        
        # ถ้ามีการเลือกโฟลเดอร์เฉพาะ ให้อ่านเฉพาะโฟลเดอร์ที่เลือก
        if selected_folders and len(selected_folders) > 0:
            pdf_files = []
            for folder_path_str in selected_folders:
                try:
                    folder_path = Path(folder_path_str)
                    if folder_path.exists() and folder_path.is_dir():
                        # ตรวจสอบว่าเป็นโฟลเดอร์หลัก (VAT folder) หรือโฟลเดอร์ย่อย
                        is_main_folder = (str(folder_path) == str(vat_folder))
                        
                        folder_files = []
                        if is_main_folder:
                            # ถ้าเป็นโฟลเดอร์หลัก ให้อ่านเฉพาะไฟล์ในโฟลเดอร์หลักเท่านั้น (ไม่รวมโฟลเดอร์ย่อย)
                            for ext in supported_extensions + supported_extensions_upper:
                                found_files = list(folder_path.glob(ext))
                                # กรองเฉพาะไฟล์จริงๆ (ไม่ใช่โฟลเดอร์)
                                folder_files.extend([f for f in found_files if f.is_file()])
                            logger.info(f"📁 อ่านไฟล์จากโฟลเดอร์หลัก: {folder_path.name} ({len(folder_files)} ไฟล์)")
                        else:
                            # ถ้าเป็นโฟลเดอร์ย่อย ให้อ่านไฟล์ทั้งหมดในโฟลเดอร์ย่อย (รวมโฟลเดอร์ย่อยภายใน)
                            for ext in supported_extensions + supported_extensions_upper:
                                found_files = list(folder_path.rglob(ext))
                                # กรองเฉพาะไฟล์จริงๆ (ไม่ใช่โฟลเดอร์)
                                folder_files.extend([f for f in found_files if f.is_file()])
                            logger.info(f"📁 อ่านไฟล์จากโฟลเดอร์ย่อย: {folder_path.name} ({len(folder_files)} ไฟล์)")
                        
                        # ลบไฟล์ซ้ำ (ถ้ามี) โดยใช้ path เป็น key
                        seen = set()
                        unique_files = []
                        for f in folder_files:
                            file_path_str = str(f)
                            if file_path_str not in seen:
                                seen.add(file_path_str)
                                unique_files.append(f)
                        folder_files = unique_files
                        
                        # Log จำนวนไฟล์ที่ถูกกรองแล้ว
                        if is_main_folder:
                            logger.info(f"📁 อ่านไฟล์จากโฟลเดอร์หลัก: {folder_path.name} ({len(folder_files)} ไฟล์)")
                        else:
                            logger.info(f"📁 อ่านไฟล์จากโฟลเดอร์ย่อย: {folder_path.name} ({len(folder_files)} ไฟล์)")
                        
                        pdf_files.extend(folder_files)
                    else:
                        logger.warning(f"⚠️ โฟลเดอร์ไม่พบหรือไม่ใช่โฟลเดอร์: {folder_path_str}")
                except Exception as e:
                    logger.error(f"❌ ไม่สามารถอ่านโฟลเดอร์ {folder_path_str}: {e}")
            
            # ลบไฟล์ซ้ำทั้งหมด (ถ้ามีการเลือกหลายโฟลเดอร์) โดยใช้ path เป็น key
            seen_all = set()
            unique_all_files = []
            for f in pdf_files:
                file_path_str = str(f)
                if file_path_str not in seen_all:
                    seen_all.add(file_path_str)
                    unique_all_files.append(f)
            pdf_files = unique_all_files
            
            logger.info(f"📊 รวมไฟล์ทั้งหมดหลังลบซ้ำ: {len(pdf_files)} ไฟล์")
        elif include_subfolders:
            # อ่านไฟล์ที่รองรับ OCR ทั้งหมด (รวมโฟลเดอร์ย่อย)
            pdf_files = []
            for ext in supported_extensions + supported_extensions_upper:
                pdf_files.extend(list(vat_folder.rglob(ext)))
        else:
            # อ่านเฉพาะไฟล์ในโฟลเดอร์หลัก
            pdf_files = []
            for ext in supported_extensions + supported_extensions_upper:
                pdf_files.extend(list(vat_folder.glob(ext)))
        
        if not pdf_files:
            return jsonify({
                'success': False,
                'error': f'ไม่พบไฟล์ PDF ในโฟลเดอร์ VAT: {vat_folder}'
            }), 404
        
        # ถ้าเป็น checkOnly ให้ return เฉพาะข้อมูลจำนวนไฟล์และเวลา รวมถึงรายการโฟลเดอร์ย่อย
        if check_only:
            # หาโฟลเดอร์ย่อยทั้งหมดที่มีไฟล์ PDF
            subfolders_info = []
            
            # เพิ่มโฟลเดอร์หลัก (VAT folder) - นับเฉพาะไฟล์ในโฟลเดอร์หลักเท่านั้น (ไม่รวมโฟลเดอร์ย่อย)
            main_folder_files = []
            for ext in supported_extensions + supported_extensions_upper:
                # ใช้ glob() เพื่ออ่านเฉพาะไฟล์ในโฟลเดอร์หลัก (ไม่รวมโฟลเดอร์ย่อย)
                found_files = list(vat_folder.glob(ext))
                # กรองเฉพาะไฟล์จริงๆ (ไม่ใช่โฟลเดอร์)
                main_folder_files.extend([f for f in found_files if f.is_file()])
            
            # ลบไฟล์ซ้ำ (ถ้ามี) โดยใช้ path เป็น key
            seen = set()
            unique_files = []
            for f in main_folder_files:
                if str(f) not in seen:
                    seen.add(str(f))
                    unique_files.append(f)
            main_folder_files = unique_files
            
            # Log สำหรับ debug
            if main_folder_files:
                logger.info(f"📁 โฟลเดอร์หลัก - VAT: พบ {len(main_folder_files)} ไฟล์")
                for f in main_folder_files:
                    logger.debug(f"   - {f.name} ({f.suffix})")
            
            if main_folder_files:
                subfolders_info.append({
                    'path': str(vat_folder),
                    'name': 'โฟลเดอร์หลัก - VAT',
                    'relative_path': '.',
                    'pdf_count': len(main_folder_files)
                })
            
            # หาโฟลเดอร์ย่อยทั้งหมด
            try:
                for item in vat_folder.iterdir():
                    if item.is_dir():
                        subfolder_files = []
                        for ext in supported_extensions + supported_extensions_upper:
                            # ใช้ rglob() เพื่ออ่านไฟล์ในโฟลเดอร์ย่อยแบบ recursive
                            found_files = list(item.rglob(ext))
                            # กรองเฉพาะไฟล์จริงๆ (ไม่ใช่โฟลเดอร์)
                            subfolder_files.extend([f for f in found_files if f.is_file()])
                        
                        # ลบไฟล์ซ้ำ (ถ้ามี) โดยใช้ path เป็น key
                        seen = set()
                        unique_subfolder_files = []
                        for f in subfolder_files:
                            if str(f) not in seen:
                                seen.add(str(f))
                                unique_subfolder_files.append(f)
                        subfolder_files = unique_subfolder_files
                        
                        # Log สำหรับ debug
                        if subfolder_files:
                            logger.info(f"📁 โฟลเดอร์ย่อย '{item.name}': พบ {len(subfolder_files)} ไฟล์")
                            for f in subfolder_files[:5]:  # แสดงแค่ 5 ไฟล์แรก
                                logger.debug(f"   - {f.name} ({f.suffix})")
                        
                        if subfolder_files:
                            # คำนวณ relative path จาก vat_folder
                            try:
                                relative_path = item.relative_to(vat_folder)
                                subfolders_info.append({
                                    'path': str(item),
                                    'name': item.name,
                                    'relative_path': str(relative_path),
                                    'pdf_count': len(subfolder_files)
                                })
                            except Exception as e:
                                logger.warning(f"⚠️ ไม่สามารถคำนวณ relative path สำหรับ {item}: {e}")
                                subfolders_info.append({
                                    'path': str(item),
                                    'name': item.name,
                                    'relative_path': item.name,
                                    'pdf_count': len(subfolder_files)
                                })
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถอ่านโฟลเดอร์ย่อย: {e}")
            
            # คำนวณ total_pdf_files จาก subfolders_info (ไม่ใช่จาก pdf_files ที่อาจมีไฟล์ซ้ำ)
            total_pdf_files_from_folders = sum(folder.get('pdf_count', 0) for folder in subfolders_info)
            
            # คำนวณเวลาที่คาดว่าจะใช้ (1 ไฟล์ใช้เวลา 30 วินาที)
            estimated_seconds = total_pdf_files_from_folders * 30
            estimated_minutes = estimated_seconds / 60
            
            # จัดรูปแบบเวลา
            if estimated_minutes < 1:
                estimated_time_str = f"{estimated_seconds} วินาที"
            elif estimated_minutes < 60:
                estimated_time_str = f"{estimated_minutes:.1f} นาที"
            else:
                hours = int(estimated_minutes // 60)
                minutes = int(estimated_minutes % 60)
                if minutes > 0:
                    estimated_time_str = f"{hours} ชั่วโมง {minutes} นาที"
                else:
                    estimated_time_str = f"{hours} ชั่วโมง"
            
            logger.info(f"📊 Check Only Mode: Found {total_pdf_files_from_folders} files in {len(subfolders_info)} folders, Estimated time: {estimated_time_str}")
            return jsonify({
                'success': True,
                'vatFolder': str(vat_folder),
                'totalPDFFiles': total_pdf_files_from_folders,  # ใช้จำนวนจาก subfolders_info
                'estimatedTime': estimated_time_str,
                'estimatedSeconds': estimated_seconds,
                'subfolders': subfolders_info  # เพิ่มรายการโฟลเดอร์ย่อย
            }), 200
        
        # Import required modules
        try:
            from email_system.tax_ocr_processor import TaxOCRProcessor
            from invoice_data_extractor import extract_invoice_data
            from ocr_cache_manager import OCRCacheManager
        except ImportError as e:
            logger.error(f"❌ ไม่สามารถ import modules: {e}")
            return jsonify({
                'success': False,
                'error': f'ไม่สามารถ import modules ที่จำเป็น: {e}'
            }), 500
        
        # สร้าง Cache Manager สำหรับบริษัทนี้
        cache_manager = OCRCacheManager(cache_ttl_hours=720, company_name="default")  # 30 วัน
        logger.info(f"✅ เริ่มต้น OCR Cache Manager สำหรับบริษัท: {company}")
        
        # สร้าง session ID สำหรับ track progress
        import uuid
        import time
        session_id = str(uuid.uuid4())
        
        # ลบไฟล์ซ้ำ (ถ้ามี) ก่อนนับจำนวน
        seen_all = set()
        unique_pdf_files = []
        for f in pdf_files:
            if str(f) not in seen_all:
                seen_all.add(str(f))
                unique_pdf_files.append(f)
        pdf_files = unique_pdf_files
        
        # ไม่จำกัดจำนวนไฟล์ - จะประมวลผลทั้งหมด (จะพักทุก 50 รายการ)
        total_pdf_count = len(pdf_files)
        logger.info(f"📊 พบ {total_pdf_count} ไฟล์ - จะประมวลผลทั้งหมด (จะพัก 1-2 นาทีทุกๆ 50 รายการ)")
        
        # เริ่มต้น progress store
        auditcheck_ocr_progress[session_id] = {
            'current': 0,
            'total': total_pdf_count,
            'percent': 0,
            'filename': 'กำลังเริ่มต้น...',
            'status': 'processing',
            'success_count': 0,
            'failed_count': 0,
            'current_step': 'กำลังเริ่มต้นระบบ...',
            'step_details': ''
        }
        
        # รัน OCR สำหรับแต่ละไฟล์ PDF
        processor = TaxOCRProcessor()
        processed_files = []
        ocr_data_list = []
        cache_hits = 0  # นับจำนวนไฟล์ที่ใช้ cache
        cache_misses = 0  # นับจำนวนไฟล์ที่เรียก OCR
        
        for idx, pdf_file in enumerate(pdf_files, 1):  # ใช้ไฟล์ที่ถูกกรองแล้ว (ไม่เกิน 50 ไฟล์)
            try:
                # อัปเดต progress
                percent = int((idx / total_pdf_count) * 100)
                auditcheck_ocr_progress[session_id].update({
                    'current': idx,
                    'total': total_pdf_count,
                    'percent': percent,
                    'filename': pdf_file.name,
                    'status': 'processing'
                })
                
                logger.info(f"🔄 กำลังรัน OCR สำหรับ: {pdf_file.name} ({idx}/{total_pdf_count})")
                
                # อัพเดทขั้นตอน: กำลังตรวจสอบ cache
                auditcheck_ocr_progress[session_id].update({
                    'current_step': f'กำลังตรวจสอบ cache สำหรับไฟล์ที่ {idx}/{total_pdf_count}',
                    'step_details': f'ไฟล์: {pdf_file.name}'
                })
                
                # ตรวจสอบ cache ก่อน (ถ้า ocr_mode เป็น 'new' ให้ข้าม cache)
                cached_data = None
                if ocr_mode != 'new':
                    cached_data = cache_manager.get(pdf_file.name, str(pdf_file))
                else:
                    logger.info(f"🔄 OCR Mode: new - ข้าม cache สำหรับ {pdf_file.name}")
                
                if cached_data:
                    cache_hits += 1
                    logger.info(f"✅ พบ cache สำหรับ {pdf_file.name} - ใช้ข้อมูลจาก cache")
                    
                    # อัพเดทขั้นตอน: พบ cache
                    auditcheck_ocr_progress[session_id].update({
                        'current_step': f'✅ พบ cache สำหรับไฟล์ที่ {idx}/{total_pdf_count}',
                        'step_details': f'ใช้ข้อมูลจาก cache: {pdf_file.name}'
                    })
                    
                    # ใช้ข้อมูลจาก cache แทนการเรียก OCR
                    ocr_result = {
                        'success': True,
                        'raw_text': cached_data.get('raw_text_preview', ''),
                        'raw_content': cached_data.get('raw_text', ''),
                        'method': cached_data.get('method', 'cache')
                    }
                    # ใช้ข้อมูลที่ extract แล้วจาก cache
                    extracted_data = {
                        'company_name': cached_data.get('company_name', ''),
                        'tax_id': cached_data.get('tax_id', ''),
                        'branch': cached_data.get('branch', ''),
                        'date': cached_data.get('date', ''),
                        'document_number': cached_data.get('document_number', ''),
                        'amount_before_vat': cached_data.get('amount_before_vat', 0),
                        'vat_amount': cached_data.get('vat_amount', 0),
                        'total_amount': cached_data.get('total_amount', 0),
                        'buyer_name': cached_data.get('buyer_name', ''),
                        'buyer_tax_id': cached_data.get('buyer_tax_id', ''),
                        'buyer_address': cached_data.get('buyer_address', ''),
                        'document_type': cached_data.get('document_type', ''),
                        'document_status': cached_data.get('document_status', ''),
                        'items': cached_data.get('items', []),  # เพิ่มรายการสินค้า
                        'customs_duty': cached_data.get('customs_duty', 0),
                        'has_customs_duty': cached_data.get('has_customs_duty', False)
                    }
                    # ดึง reference number จาก cache หรือจากชื่อไฟล์
                    reference_number = cached_data.get('reference_number')
                    if not reference_number:
                        # ดึง reference number จากชื่อไฟล์
                        import re
                        ref_patterns = [
                            r'\d{2}\.\d{2}\.\d{4}_([A-Z]+-\d+)_',
                            r'^\d{2}\.\d{2}\.\d{4}_([A-Z]+-\d+)_',
                            r'_([A-Z]+-\d+)_',
                            r'([A-Z]{2,}-\d{8,})',
                        ]
                        for pattern in ref_patterns:
                            match = re.search(pattern, pdf_file.name)
                            if match:
                                reference_number = match.group(1)
                                break
                    
                    # สร้าง ocr_data จาก cache
                    ocr_data = {
                        'filename': pdf_file.name,
                        'filepath': str(pdf_file),
                        'success': True,
                        'company_name': extracted_data.get('company_name', ''),
                        'tax_id': extracted_data.get('tax_id', ''),
                        'branch': extracted_data.get('branch', ''),
                        'date': extracted_data.get('date', ''),
                        'document_number': extracted_data.get('document_number', ''),
                        'reference_number': reference_number,
                        'amount_before_vat': extracted_data.get('amount_before_vat', 0),
                        'vat_amount': extracted_data.get('vat_amount', 0),
                        'total_amount': extracted_data.get('total_amount', 0),
                        'buyer_name': extracted_data.get('buyer_name', ''),
                        'buyer_tax_id': extracted_data.get('buyer_tax_id', ''),
                        'buyer_address': extracted_data.get('buyer_address', ''),
                        'document_type': extracted_data.get('document_type', ''),
                        'document_status': extracted_data.get('document_status', ''),
                        'items': extracted_data.get('items', []),  # เพิ่มรายการสินค้าจาก cache
                        'customs_duty': extracted_data.get('customs_duty', 0),
                        'has_customs_duty': extracted_data.get('has_customs_duty', False),
                        'is_customs_department': extracted_data.get('company_name', '').find('กรมศุลกากร') != -1 or extracted_data.get('company_name', '').find('กรมศุล') != -1,
                        'raw_text_preview': cached_data.get('raw_text_preview', ''),
                        'method': 'cache'
                    }
                    
                    # แสดงข้อมูลจาก cache
                    logger.info("=" * 80)
                    logger.info(f"📄 ไฟล์: {pdf_file.name} (จาก cache)")
                    logger.info(f"   ✅ อ่านข้อมูลสำเร็จ (จาก cache)")
                    logger.info(f"   🏢 ชื่อบริษัท: {ocr_data['company_name'] or '(ไม่พบ)'}")
                    logger.info(f"   🆔 เลขประจำตัวผู้เสียภาษี: {ocr_data['tax_id'] or '(ไม่พบ)'}")
                    logger.info(f"   📍 สาขา: {ocr_data['branch'] or '(ไม่พบ)'}")
                    logger.info(f"   📅 วันที่: {ocr_data['date'] or '(ไม่พบ)'}")
                    logger.info(f"   📄 เลขที่เอกสาร: {ocr_data['document_number'] or '(ไม่พบ)'}")
                    logger.info(f"   🔖 เลขที่เอกสารอ้างอิง: {ocr_data['reference_number'] or '(ไม่พบ)'}")
                    logger.info(f"   💰 ยอดก่อนภาษี: {ocr_data['amount_before_vat']:,.2f} บาท")
                    logger.info(f"   💵 ยอดภาษีมูลค่าเพิ่ม: {ocr_data['vat_amount']:,.2f} บาท")
                    logger.info(f"   💵 ยอดรวม: {ocr_data['total_amount']:,.2f} บาท")
                    logger.info("=" * 80)
                    
                    ocr_data_list.append(ocr_data)
                    processed_files.append(pdf_file.name)
                    
                    # อัปเดต progress
                    success_count = sum(1 for item in ocr_data_list if item.get('success'))
                    percent = int((idx / total_pdf_count) * 100) if total_pdf_count > 0 else 0
                    
                    auditcheck_ocr_progress[session_id].update({
                        'current': idx,
                        'total': total_pdf_count,
                        'percent': percent,
                        'filename': pdf_file.name,
                        'status': 'processing',
                        'success_count': success_count,
                        'failed_count': len(ocr_data_list) - success_count
                    })
                    
                    logger.info(f"📊 Progress updated: {idx}/{total_pdf_count} ({percent}%) - {pdf_file.name} (จาก cache)")
                    continue
                
                # ถ้าไม่มี cache ให้เรียก OCR
                cache_misses += 1
                logger.info(f"📝 ไม่พบ cache สำหรับ {pdf_file.name} - เรียก OCR")
                
                # อัพเดทขั้นตอน: กำลังเรียก OCR API
                auditcheck_ocr_progress[session_id].update({
                    'current_step': f'📤 กำลังเรียก OCR API สำหรับไฟล์ที่ {idx}/{total_pdf_count}',
                    'step_details': f'ไฟล์: {pdf_file.name}'
                })
                
                # ใช้ key-extract mode เพื่อดึงข้อมูลแบบ structured
                ocr_result = processor.get_ocr_raw_data(pdf_file)
                
                # อัพเดทขั้นตอน: ได้รับผลลัพธ์จาก OCR API
                if ocr_result.get('success'):
                    auditcheck_ocr_progress[session_id].update({
                        'current_step': f'✅ ได้รับผลลัพธ์จาก OCR API สำหรับไฟล์ที่ {idx}/{total_pdf_count}',
                        'step_details': f'กำลังประมวลผลข้อมูล: {pdf_file.name}'
                    })
                else:
                    auditcheck_ocr_progress[session_id].update({
                        'current_step': f'❌ OCR API ไม่สำเร็จสำหรับไฟล์ที่ {idx}/{total_pdf_count}',
                        'step_details': f'ไฟล์: {pdf_file.name} - {ocr_result.get("error", "Unknown error")}'
                    })
                
                if not ocr_result.get('success'):
                    logger.warning(f"⚠️ OCR ไม่สำเร็จสำหรับ {pdf_file.name}: {ocr_result.get('error')}")
                    ocr_data_list.append({
                        'filename': pdf_file.name,
                        'filepath': str(pdf_file),
                        'success': False,
                        'error': ocr_result.get('error', 'OCR ไม่สำเร็จ')
                    })
                    continue
                
                raw_text = ocr_result.get('raw_text', '') or ocr_result.get('text', '')
                raw_content = ocr_result.get('raw_content', '')
                
                # Log ข้อมูลที่ได้จาก OCR
                logger.info("=" * 80)
                logger.info(f"📊 OCR Result สำหรับ {pdf_file.name}:")
                logger.info(f"   Method: {ocr_result.get('method', 'unknown')}")
                logger.info(f"   Raw Text Length: {len(raw_text) if raw_text else 0} characters")
                logger.info(f"   Raw Content Length: {len(raw_content) if raw_content else 0} characters")
                
                # แสดงตัวอย่าง raw_text (200 ตัวอักษรแรก)
                if raw_text:
                    logger.info(f"   Raw Text Preview: {raw_text[:200]}...")
                else:
                    logger.warning(f"   ⚠️ Raw Text is empty!")
                
                # ถ้า raw_content เป็น JSON (key-extract response) ให้ parse และแสดงข้อมูล
                if raw_content and ('"success"' in raw_content or '"data"' in raw_content):
                    try:
                        import json
                        parsed_json = json.loads(raw_content)
                        if isinstance(parsed_json, dict) and 'data' in parsed_json:
                            data = parsed_json.get('data', {})
                            logger.info(f"   📋 Key-Extract Data Fields: {list(data.keys())}")
                            # แสดงค่าของแต่ละ field
                            for key, value in data.items():
                                if value:  # แสดงเฉพาะค่าที่ไม่ว่าง
                                    logger.info(f"      - {key}: {value}")
                    except Exception as e:
                        logger.warning(f"   ⚠️ Cannot parse raw_content as JSON: {e}")
                
                logger.info("=" * 80)
                
                # ดึง reference number จากชื่อไฟล์เก่า (เช่น EXP-20251000004 จาก 16.10.2025_EXP-20251000004_530117_...)
                reference_number = None
                import re
                # Pattern: วันที่_REF-NUMBER_...
                # ตัวอย่าง: 16.10.2025_EXP-20251000004_530117_...
                ref_patterns = [
                    r'\d{2}\.\d{2}\.\d{4}_([A-Z]+-\d+)_',  # 16.10.2025_EXP-20251000004_
                    r'^\d{2}\.\d{2}\.\d{4}_([A-Z]+-\d+)_',  # เริ่มต้นด้วยวันที่
                    r'_([A-Z]+-\d+)_',  # รูปแบบ _EXP-20251000004_
                    r'([A-Z]{2,}-\d{8,})',  # รูปแบบ EXP-20251000004 (fallback)
                ]
                
                for pattern in ref_patterns:
                    match = re.search(pattern, pdf_file.name)
                    if match:
                        reference_number = match.group(1)
                        logger.info(f"✅ ดึง reference number จากชื่อไฟล์: {reference_number} (จาก {pdf_file.name})")
                        break
                
                # ถ้า raw_content เป็น JSON (key-extract response) ให้ใช้ข้อมูลจาก key-extract โดยตรง
                extracted_data = {}
                if raw_content and ('"success"' in raw_content or '"data"' in raw_content):
                    try:
                        import json
                        parsed_json = json.loads(raw_content)
                        if isinstance(parsed_json, dict) and 'data' in parsed_json:
                            key_extract_data = parsed_json.get('data', {})
                            logger.info(f"✅ ใช้ข้อมูลจาก key-extract API โดยตรง")
                            
                            # แปลงชื่อ fields จาก key-extract เป็นชื่อที่ต้องการ
                            # ชื่อผู้ขาย -> company_name
                            seller_name = key_extract_data.get('ชื่อผู้ขาย', '') or key_extract_data.get('ชื่อบริษัท', '')
                            extracted_data['company_name'] = seller_name
                            
                            # ตรวจสอบว่าชื่อผู้ขายเป็น "กรมศุลกากร" หรือไม่
                            is_customs_department = 'กรมศุลกากร' in seller_name or 'กรมศุล' in seller_name
                            
                            if is_customs_department:
                                # สำหรับกรมศุลกากร: ไม่มีเลขนิติ, ไม่มีที่อยู่, ไม่มีสาขา
                                extracted_data['tax_id'] = ''
                                extracted_data['address'] = ''
                                extracted_data['branch'] = ''
                                logger.info(f"🔍 ตรวจพบชื่อผู้ขายเป็น 'กรมศุลกากร' - เซ็ตค่า tax_id, address, branch เป็นว่าง")
                            else:
                                # เลขประจำตัวผู้เสียภาษี - ผู้ขาย -> tax_id
                                extracted_data['tax_id'] = key_extract_data.get('เลขประจำตัวผู้เสียภาษี - ผู้ขาย', '')
                                # ที่อยู่ผู้ขาย -> address
                                extracted_data['address'] = key_extract_data.get('ที่อยู่ผู้ขาย', '')
                                # สาขา - ผู้ขาย -> branch (ลบ "HQ (" และ ")" ออก)
                                branch_raw = key_extract_data.get('สาขา - ผู้ขาย', '')
                                if branch_raw:
                                    # ลบ "HQ (" และ ")" ออก เช่น "HQ (00000)" -> "00000"
                                    branch_clean = branch_raw.replace('HQ (', '').replace(')', '').strip()
                                    extracted_data['branch'] = branch_clean
                                else:
                                    extracted_data['branch'] = ''
                            # วันที่ -> date (แปลงปี พ.ศ. เป็น ค.ศ.)
                            date_raw = key_extract_data.get('วันที่', '')
                            extracted_data['date'] = convert_buddhist_to_christian_date(date_raw)
                            # เลขที่ใบกำกับภาษี -> document_number
                            extracted_data['document_number'] = key_extract_data.get('เลขที่ใบกำกับภาษี', '')
                            # สำหรับกรมศุลกากร: จัดการรายการภาษีมูลค่าเพิ่มและค่าอากรขาขเข้า
                            if is_customs_department:
                                # ตรวจสอบว่ามีรายการ "ภาษีมูลค่าเพิ่ม" หรือไม่
                                vat_amount_str = key_extract_data.get('ภาษีมูลค่าเพิ่ม', '0') or '0'
                                try:
                                    vat_amount = float(vat_amount_str.replace(',', '').replace('฿', '').strip())
                                except:
                                    vat_amount = 0.0
                                
                                # ถ้ามีรายการภาษีมูลค่าเพิ่ม ให้คำนวณยอดก่อนภาษี
                                if vat_amount > 0:
                                    # คำนวณ: ยอดก่อนภาษี = ยอดภาษี x 100 / 7
                                    amount_before_vat = vat_amount * 100 / 7
                                    # คำนวณ: ยอดหลังบวกภาษี = ยอดก่อนภาษี + ยอดภาษี
                                    total_amount = amount_before_vat + vat_amount
                                    
                                    extracted_data['amount_before_vat'] = amount_before_vat
                                    extracted_data['vat_amount'] = vat_amount
                                    extracted_data['total_amount'] = total_amount
                                    
                                    logger.info(f"💰 กรมศุลกากร - คำนวณยอด: ภาษี={vat_amount:,.2f}, ก่อนภาษี={amount_before_vat:,.2f}, รวม={total_amount:,.2f}")
                                else:
                                    extracted_data['amount_before_vat'] = 0.0
                                    extracted_data['vat_amount'] = 0.0
                                    extracted_data['total_amount'] = 0.0
                                
                                # ตรวจสอบว่ามีรายการ "ค่าอากรขาขเข้า" หรือไม่
                                customs_duty_str = key_extract_data.get('ค่าอากรขาขเข้า', '0') or key_extract_data.get('ค่าอากร', '0') or '0'
                                try:
                                    customs_duty = float(customs_duty_str.replace(',', '').replace('฿', '').strip())
                                except:
                                    customs_duty = 0.0
                                
                                # ถ้ามีค่าอากรขาขเข้า ให้เก็บไว้ในข้อมูลเพิ่มเติม (สำหรับเพิ่มบรรทัดในภายหลัง)
                                if customs_duty > 0:
                                    extracted_data['customs_duty'] = customs_duty
                                    extracted_data['has_customs_duty'] = True
                                    logger.info(f"📦 กรมศุลกากร - พบค่าอากรขาขเข้า: {customs_duty:,.2f}")
                                else:
                                    extracted_data['customs_duty'] = 0.0
                                    extracted_data['has_customs_duty'] = False
                            else:
                                # กรณีปกติ: ใช้ข้อมูลจาก OCR โดยตรง
                                # ยอดรวมก่อนภาษี -> amount_before_vat (แปลง string เป็น float)
                                amount_before_vat_str = key_extract_data.get('ยอดรวมก่อนภาษี', '0') or '0'
                                try:
                                    extracted_data['amount_before_vat'] = float(amount_before_vat_str.replace(',', '').replace('฿', '').strip())
                                except:
                                    extracted_data['amount_before_vat'] = 0.0
                                # ภาษีมูลค่าเพิ่ม -> vat_amount
                                vat_amount_str = key_extract_data.get('ภาษีมูลค่าเพิ่ม', '0') or '0'
                                try:
                                    extracted_data['vat_amount'] = float(vat_amount_str.replace(',', '').replace('฿', '').strip())
                                except:
                                    extracted_data['vat_amount'] = 0.0
                                # ยอดรวมสุทธิ -> total_amount
                                total_amount_str = key_extract_data.get('ยอดรวมสุทธิ', '0') or '0'
                                try:
                                    extracted_data['total_amount'] = float(total_amount_str.replace(',', '').replace('฿', '').strip())
                                except:
                                    extracted_data['total_amount'] = 0.0
                                extracted_data['customs_duty'] = 0.0
                                extracted_data['has_customs_duty'] = False
                            # ชื่อผู้ซื้อ -> buyer_name
                            extracted_data['buyer_name'] = key_extract_data.get('ชื่อผู้ซื้อ', '')
                            # เลขประจำตัวผู้เสียภาษี - ผู้ซื้อ -> buyer_tax_id
                            extracted_data['buyer_tax_id'] = key_extract_data.get('เลขประจำตัวผู้เสียภาษี - ผู้ซื้อ', '')
                            # ที่อยู่ผู้ซื้อ -> buyer_address
                            extracted_data['buyer_address'] = key_extract_data.get('ที่อยู่ผู้ซื้อ', '')
                            # ประเภทเอกสาร -> document_type
                            extracted_data['document_type'] = key_extract_data.get('ประเภทเอกสาร', '')
                            # สถานะเอกสาร -> document_status
                            extracted_data['document_status'] = key_extract_data.get('สถานะเอกสาร', '')
                            
                            # รายการสินค้า -> items
                            items_raw = key_extract_data.get('รายการสินค้า', [])
                            if items_raw and isinstance(items_raw, list):
                                extracted_data['items'] = items_raw
                            else:
                                extracted_data['items'] = []
                            
                            logger.info(f"📊 Extracted from key-extract: company_name={extracted_data.get('company_name')}, tax_id={extracted_data.get('tax_id')}, amount_before_vat={extracted_data.get('amount_before_vat')}")
                    except Exception as e:
                        logger.warning(f"⚠️ Cannot parse key-extract data: {e}, falling back to extract_invoice_data")
                        # Fallback to extract_invoice_data if parsing fails
                        if raw_text:
                            extracted_data = extract_invoice_data(raw_text, pdf_file.name, str(pdf_file))
                        else:
                            extracted_data = {}
                else:
                    # ถ้าไม่ใช่ key-extract response ให้ใช้ extract_invoice_data แบบเดิม
                    if not raw_text:
                        logger.warning(f"⚠️ ไม่พบข้อความจาก OCR สำหรับ {pdf_file.name}")
                        ocr_data_list.append({
                            'filename': pdf_file.name,
                            'filepath': str(pdf_file),
                            'success': False,
                            'error': 'ไม่พบข้อความจาก OCR'
                        })
                        continue
                    
                    logger.info(f"🔄 กำลัง extract invoice data จาก raw_text...")
                    extracted_data = extract_invoice_data(raw_text, pdf_file.name, str(pdf_file))
                    logger.info(f"📊 Extracted Data: {extracted_data}")
                
                # เตรียมข้อมูล OCR
                ocr_data = {
                    'filename': pdf_file.name,
                    'filepath': str(pdf_file),
                    'success': True,
                    'company_name': extracted_data.get('company_name', ''),
                    'tax_id': extracted_data.get('tax_id', ''),
                    'branch': extracted_data.get('branch', ''),
                    'date': extracted_data.get('date', ''),
                    'document_number': extracted_data.get('document_number', ''),
                    'reference_number': reference_number,  # เพิ่ม reference number จากชื่อไฟล์
                    'amount_before_vat': extracted_data.get('amount_before_vat', 0),
                    'vat_amount': extracted_data.get('vat_amount', 0),
                    'total_amount': extracted_data.get('total_amount', 0),
                    'buyer_name': extracted_data.get('buyer_name', ''),
                    'buyer_tax_id': extracted_data.get('buyer_tax_id', ''),
                    'buyer_address': extracted_data.get('buyer_address', ''),
                    'document_type': extracted_data.get('document_type', ''),
                    'document_status': extracted_data.get('document_status', ''),
                    'items': extracted_data.get('items', []),  # เพิ่มรายการสินค้า
                    'customs_duty': extracted_data.get('customs_duty', 0),  # ค่าอากรขาขเข้า (สำหรับกรมศุลกากร)
                    'has_customs_duty': extracted_data.get('has_customs_duty', False),  # มีค่าอากรขาขเข้าหรือไม่
                    'is_customs_department': extracted_data.get('company_name', '').find('กรมศุลกากร') != -1 or extracted_data.get('company_name', '').find('กรมศุล') != -1,
                    'raw_text_preview': raw_text[:500] + '...' if len(raw_text) > 500 else raw_text,
                    'method': ocr_result.get('method', 'unknown')
                }
                
                # แสดงข้อมูลที่อ่านได้ใน terminal
                logger.info("=" * 80)
                logger.info(f"📄 ไฟล์: {pdf_file.name}")
                logger.info(f"   ✅ อ่านข้อมูลสำเร็จ")
                logger.info(f"   🏢 ชื่อบริษัท: {ocr_data['company_name'] or '(ไม่พบ)'}")
                logger.info(f"   🆔 เลขประจำตัวผู้เสียภาษี: {ocr_data['tax_id'] or '(ไม่พบ)'}")
                logger.info(f"   📍 สาขา: {ocr_data['branch'] or '(ไม่พบ)'}")
                logger.info(f"   📅 วันที่: {ocr_data['date'] or '(ไม่พบ)'}")
                logger.info(f"   📄 เลขที่เอกสาร: {ocr_data['document_number'] or '(ไม่พบ)'}")
                logger.info(f"   🔖 เลขที่เอกสารอ้างอิง: {ocr_data['reference_number'] or '(ไม่พบ)'}")
                logger.info(f"   💰 ยอดก่อนภาษี: {ocr_data['amount_before_vat']:,.2f} บาท")
                logger.info(f"   💵 ยอดภาษีมูลค่าเพิ่ม: {ocr_data['vat_amount']:,.2f} บาท")
                logger.info(f"   💵 ยอดรวม: {ocr_data['total_amount']:,.2f} บาท")
                logger.info(f"   👤 ชื่อผู้ซื้อ: {ocr_data['buyer_name'] or '(ไม่พบ)'}")
                logger.info(f"   🆔 เลขประจำตัวผู้เสียภาษี - ผู้ซื้อ: {ocr_data['buyer_tax_id'] or '(ไม่พบ)'}")
                logger.info(f"   📍 ที่อยู่ผู้ซื้อ: {ocr_data['buyer_address'] or '(ไม่พบ)'}")
                logger.info(f"   📋 ประเภทเอกสาร: {ocr_data['document_type'] or '(ไม่พบ)'}")
                logger.info(f"   📋 สถานะเอกสาร: {ocr_data['document_status'] or '(ไม่พบ)'}")
                logger.info("=" * 80)
                
                ocr_data_list.append(ocr_data)
                processed_files.append(pdf_file.name)
                
                # อัพเดทขั้นตอน: กำลังเก็บ cache
                auditcheck_ocr_progress[session_id].update({
                    'current_step': f'💾 กำลังเก็บ cache สำหรับไฟล์ที่ {idx}/{total_pdf_count}',
                    'step_details': f'ไฟล์: {pdf_file.name}'
                })
                
                # เก็บข้อมูล OCR ลง cache
                try:
                    cache_manager.set(pdf_file.name, str(pdf_file), ocr_data)
                    logger.info(f"💾 เก็บ cache สำหรับ {pdf_file.name} สำเร็จ")
                    
                    # อัพเดทขั้นตอน: เก็บ cache สำเร็จ
                    auditcheck_ocr_progress[session_id].update({
                        'current_step': f'✅ เก็บ cache สำเร็จสำหรับไฟล์ที่ {idx}/{total_pdf_count}',
                        'step_details': f'ไฟล์: {pdf_file.name}'
                    })
                except Exception as cache_error:
                    logger.warning(f"⚠️ ไม่สามารถเก็บ cache สำหรับ {pdf_file.name}: {cache_error}")
                    auditcheck_ocr_progress[session_id].update({
                        'current_step': f'⚠️ ไม่สามารถเก็บ cache สำหรับไฟล์ที่ {idx}/{total_pdf_count}',
                        'step_details': f'ไฟล์: {pdf_file.name} - {str(cache_error)}'
                    })
                
                # อัปเดต progress ทันทีหลังจากประมวลผลไฟล์เสร็จ
                success_count = sum(1 for item in ocr_data_list if item.get('success'))
                percent = int((idx / total_pdf_count) * 100) if total_pdf_count > 0 else 0
                
                # อัพเดท progress store (เก็บ current_step และ step_details ไว้ด้วย)
                current_step = auditcheck_ocr_progress[session_id].get('current_step', '')
                step_details = auditcheck_ocr_progress[session_id].get('step_details', '')
                
                auditcheck_ocr_progress[session_id].update({
                    'current': idx,
                    'total': total_pdf_count,
                    'percent': percent,
                    'filename': pdf_file.name,
                    'status': 'processing',
                    'success_count': success_count,
                    'failed_count': len(ocr_data_list) - success_count,
                    'current_step': current_step,  # เก็บไว้
                    'step_details': step_details  # เก็บไว้
                })
                
                logger.info(f"📊 Progress updated: {idx}/{total_pdf_count} ({percent}%) - {pdf_file.name}")
                
                # พักทุก 50 รายการ (หลังจากประมวลผลครบ 50, 100, 150, ...)
                batch_size = 50
                pause_duration = 90  # พัก 90 วินาที (1.5 นาที) - ระหว่าง 1-2 นาทีตามที่ผู้ใช้ต้องการ
                
                if idx > 0 and idx % batch_size == 0 and idx < total_pdf_count:
                    logger.info(f"⏸️  พัก {pause_duration} วินาที (1.5 นาที) หลังจากประมวลผล {idx} ไฟล์...")
                    # อัพเดท progress เพื่อแสดงสถานะการพัก
                    auditcheck_ocr_progress[session_id].update({
                        'current_step': f'⏸️  กำลังพัก {pause_duration} วินาที (1.5 นาที)',
                        'step_details': f'ประมวลผลเสร็จแล้ว {idx}/{total_pdf_count} ไฟล์ - จะเริ่มต่อในอีก {pause_duration} วินาที...'
                    })
                    time.sleep(pause_duration)
                    logger.info(f"✅ พักเสร็จแล้ว เริ่มประมวลผลต่อ...")
                    # อัพเดท progress หลังจากพักเสร็จ
                    auditcheck_ocr_progress[session_id].update({
                        'current_step': f'🔄 เริ่มประมวลผลต่อ...',
                        'step_details': f'ไฟล์ถัดไป: {pdf_files[idx].name if idx < len(pdf_files) else "เสร็จสิ้น"}'
                    })
                
            except Exception as e:
                logger.error(f"❌ เกิดข้อผิดพลาดในการประมวลผล {pdf_file.name}: {e}", exc_info=True)
                ocr_data_list.append({
                    'filename': pdf_file.name,
                    'filepath': str(pdf_file),
                    'success': False,
                    'error': str(e)
                })
                # อัปเดต failed count
                auditcheck_ocr_progress[session_id]['failed_count'] = len(ocr_data_list) - sum(1 for item in ocr_data_list if item.get('success'))
                continue
        
        if not ocr_data_list:
            auditcheck_ocr_progress[session_id]['status'] = 'error'
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถประมวลผลไฟล์ PDF ได้'
            }), 500
        
        # นับจำนวนไฟล์ที่สำเร็จ
        success_count = sum(1 for item in ocr_data_list if item.get('success'))
        
        # อัปเดต progress เป็น completed
        auditcheck_ocr_progress[session_id].update({
            'status': 'completed',
            'current': total_pdf_count,
            'percent': 100,
            'filename': 'เสร็จสิ้น',
            'success_count': success_count,
            'failed_count': len(ocr_data_list) - success_count,
            'current_step': '✅ ประมวลผลเสร็จสิ้น',
            'step_details': f'สำเร็จ: {success_count} ไฟล์, ไม่สำเร็จ: {len(ocr_data_list) - success_count} ไฟล์, Cache Hits: {cache_hits}, Cache Misses: {cache_misses}'
        })
        
        # แสดงสถิติ cache
        logger.info(f"📊 Cache Statistics: Hits={cache_hits}, Misses={cache_misses}, Total={cache_hits + cache_misses}")
        
        # คำนวณเวลาที่คาดว่าจะใช้ (1 ไฟล์ใช้เวลา 30 วินาที) - ใช้จำนวนไฟล์จริงที่ถูกกรองแล้ว
        actual_total_files = len(pdf_files)
        estimated_seconds = actual_total_files * 30
        estimated_minutes = estimated_seconds / 60
        
        # จัดรูปแบบเวลา
        if estimated_minutes < 1:
            estimated_time_str = f"{estimated_seconds} วินาที"
        elif estimated_minutes < 60:
            estimated_time_str = f"{estimated_minutes:.1f} นาที"
        else:
            hours = int(estimated_minutes // 60)
            minutes = int(estimated_minutes % 60)
            if minutes > 0:
                estimated_time_str = f"{hours} ชั่วโมง {minutes} นาที"
            else:
                estimated_time_str = f"{hours} ชั่วโมง"
        
        return jsonify({
            'success': True,
            'sessionId': session_id,  # เพิ่ม session ID สำหรับ track progress
            'vatFolder': str(vat_folder),
            'totalFiles': actual_total_files,
            'totalPDFFiles': actual_total_files,  # จำนวนไฟล์ PDF ทั้งหมด (ใช้จำนวนไฟล์จริงที่ถูกกรองแล้ว)
            'estimatedTime': estimated_time_str,  # เวลาที่คาดว่าจะใช้
            'estimatedSeconds': estimated_seconds,  # เวลาเป็นวินาที
            'processedFiles': processed_files,
            'successCount': success_count,
            'failedCount': len(ocr_data_list) - success_count,
            'cacheHits': cache_hits,  # จำนวนไฟล์ที่ใช้ cache
            'cacheMisses': cache_misses,  # จำนวนไฟล์ที่เรียก OCR
            'ocrData': ocr_data_list
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการรัน OCR: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/upload-excel', methods=['POST'])
def upload_excel_for_audit():
    """อัปโหลดไฟล์ Excel และบันทึกในโฟลเดอร์ VAT"""
    try:
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'ไม่พบไฟล์ที่อัปโหลด'
            }), 400
        
        file = request.files['file']
        tax_month = request.form.get('taxMonth', '')
        company = request.form.get('company', '')
        
        if not tax_month or not company:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเดือนภาษีและชื่อบริษัท'
            }), 400
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'ไม่พบไฟล์ที่เลือก'
            }), 400
        
        # ตรวจสอบว่าเป็นไฟล์ Excel หรือไม่
        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            return jsonify({
                'success': False,
                'error': 'กรุณาอัปโหลดไฟล์ Excel (.xlsx หรือ .xls)'
            }), 400
        
        # แปลงเดือนจาก YYYY-MM
        if '-' not in tax_month:
            return jsonify({
                'success': False,
                'error': f'รูปแบบเดือนภาษีไม่ถูกต้อง: {tax_month}'
            }), 400
        
        year, month = tax_month.split('-')
        year_int = int(year)
        month_int = int(month)
        
        # สร้างรูปแบบปี-เดือน
        month_year_patterns = [
            f"{year_int}-{month_int:02d}",
            f"{year_int}-{month_int}",
            f"{month_int:02d}-{year_int}",
            f"{month_int}-{year_int}",
        ]
        
        # ค้นหาโฟลเดอร์ VAT
        base_paths = [
            Path(f"V:/A.โฟร์เดอร์หลัก/{company}"),
            Path(f"V:/AA.โฟรเดอร์หลัก/{company}"),
            Path(f"V:/AAA.โฟรเดอร์หลัก/{company}")
        ]
        
        vat_folder = None
        for base_path in base_paths:
            if not base_path.exists():
                continue
            
            try:
                account_folder = base_path / "บัญชี" / "002-รายจ่าย" / "PV"
                if not account_folder.exists():
                    continue
                
                # ค้นหาโฟลเดอร์เดือน-ปี
                month_year_folder = None
                for pattern in month_year_patterns:
                    potential_folder = account_folder / pattern
                    if potential_folder.exists():
                        month_year_folder = potential_folder
                        break
                
                if not month_year_folder:
                    # ค้นหาในโฟลเดอร์ปี
                    for item in account_folder.iterdir():
                        if item.is_dir() and (year in item.name or str(year_int + 543) in item.name):
                            for pattern in month_year_patterns:
                                potential_folder = item / pattern
                                if potential_folder.exists():
                                    month_year_folder = potential_folder
                                    break
                            if month_year_folder:
                                break
                
                if month_year_folder:
                    # ค้นหาโฟลเดอร์ VAT
                    for item in month_year_folder.iterdir():
                        if item.is_dir() and item.name.lower() == "vat":
                            vat_folder = item
                            break
                    
                    if not vat_folder:
                        vat_folder = month_year_folder
                    
                    break
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์ VAT ใน {base_path}: {e}")
                continue
        
        if not vat_folder:
            return jsonify({
                'success': False,
                'error': f'ไม่พบโฟลเดอร์ VAT สำหรับเดือน {tax_month}'
            }), 404
        
        # บันทึกไฟล์ Excel
        excel_filename = f"Invoice_Data_OCR_{tax_month.replace('-', '')}.xlsx"
        excel_path = vat_folder / excel_filename
        
        file.save(str(excel_path))
        
        logger.info(f"✅ บันทึกไฟล์ Excel สำเร็จ: {excel_path}")
        
        return jsonify({
            'success': True,
            'excelPath': str(excel_path),
            'filename': excel_filename
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการอัปโหลด Excel: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/check-purchase-tax', methods=['POST'])
def check_purchase_tax():
    """ตรวจสอบว่ามีไฟล์ภาษีซื้อหรือไม่ และเปรียบเทียบจำนวนรายการกับไฟล์ PDF ในโฟลเดอร์ VAT"""
    try:
        data = request.json
        tax_month = data.get('taxMonth', '')  # Format: YYYY-MM
        company = data.get('company', '')
        
        if not tax_month or not company:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเดือนภาษีและชื่อบริษัท'
            }), 400
        
        # แปลงเดือนจาก YYYY-MM เป็นรูปแบบที่ใช้ในระบบ
        if '-' not in tax_month:
            return jsonify({
                'success': False,
                'error': f'รูปแบบเดือนภาษีไม่ถูกต้อง: {tax_month}'
            }), 400
        
        year, month = tax_month.split('-')
        try:
            year_int = int(year)
            month_int = int(month)
        except ValueError:
            return jsonify({
                'success': False,
                'error': f'ไม่สามารถแปลงปีหรือเดือนเป็นตัวเลขได้: {tax_month}'
            }), 400
        
        thai_year = year_int + 543  # แปลงเป็นปี พ.ศ.
        
        # สร้างรูปแบบปี-เดือน (รองรับหลายรูปแบบ)
        month_year_patterns = [
            f"{year_int}-{month_int:02d}",   # 2025-10 (ปี-เดือน)
            f"{year_int}-{month_int}",       # 2025-10 (ปี-เดือน)
            f"{month_int:02d}-{year_int}",   # 10-2025 (เดือน-ปี)
            f"{month_int}-{year_int}",       # 10-2025 (เดือน-ปี)
        ]
        
        # ค้นหาไฟล์ภาษีซื้อในโฟลเดอร์ต่างๆ
        base_paths = [
            Path(f"V:/A.โฟร์เดอร์หลัก/{company}"),
            Path(f"V:/AA.โฟรเดอร์หลัก/{company}"),
            Path(f"V:/AAA.โฟรเดอร์หลัก/{company}")
        ]
        
        purchase_tax_files = []
        purchase_tax_row_count = 0
        
        # 1. ค้นหาไฟล์ภาษีซื้อในโครงสร้างโฟลเดอร์ที่เฉพาะเจาะจง: บัญชี > 003-ภาษี > ภ.พ.30 > [Year] > [Month-Year]
        for base_path in base_paths:
            if not base_path.exists():
                continue
            
            try:
                # โครงสร้าง: บัญชี > 003-ภาษี > ภ.พ.30 > ปี > เดือน-ปี
                account_folder = base_path / "บัญชี"
                if not account_folder.exists():
                    continue
                
                tax_folder = account_folder / "003-ภาษี"
                if not tax_folder.exists():
                    continue
                
                pph30_folder = tax_folder / "ภ.พ.30"
                if not pph30_folder.exists():
                    continue
                
                year_folder = pph30_folder / year
                if not year_folder.exists():
                    continue
                
                # ค้นหาโฟลเดอร์เดือน-ปี
                month_year_folder = None
                for pattern in month_year_patterns:
                    potential_folder = year_folder / pattern
                    if potential_folder.exists():
                        month_year_folder = potential_folder
                        break
                
                # ถ้าไม่เจอโฟลเดอร์เดือน-ปี ให้ลองค้นหาในโฟลเดอร์ปีโดยตรง
                if not month_year_folder:
                    # ลองค้นหาโฟลเดอร์ที่มีรูปแบบเดือน-ปีในชื่อ
                    try:
                        for item in year_folder.iterdir():
                            if item.is_dir():
                                for pattern in month_year_patterns:
                                    if pattern in item.name or item.name == pattern:
                                        month_year_folder = item
                                        break
                                if month_year_folder:
                                    break
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์เดือน-ปีใน {year_folder}: {e}")
                
                # ค้นหาไฟล์ภาษีซื้อในโฟลเดอร์เดือน-ปี
                if month_year_folder:
                    try:
                        # Pattern สำหรับชื่อไฟล์ภาษีซื้อ (ต้องมีคำว่า "ภาษีซื้อ" และไม่ใช่ "ภาษีขาย")
                        purchase_tax_patterns = [
                            f"*ภาษีซื้อ*",
                            f"*Purchase Tax*",
                            f"*purchase tax*",
                            f"*VAT Purchase*",
                        ]
                        
                        # ค้นหาไฟล์ Excel ทั้งหมดในโฟลเดอร์เดือน-ปี
                        all_excel_files = list(month_year_folder.glob("*.xlsx")) + list(month_year_folder.glob("*.xls"))
                        
                        for file_path in all_excel_files:
                            if not file_path.is_file():
                                continue
                            
                            filename = file_path.name
                            filename_lower = filename.lower()
                            
                            # ตรวจสอบว่าเป็นไฟล์ภาษีซื้อหรือไม่ (ต้องมีคำว่า "ภาษีซื้อ" และไม่ใช่ "ภาษีขาย")
                            is_purchase_tax = False
                            
                            # ตรวจสอบว่ามีคำว่า "ภาษีซื้อ" หรือ pattern ที่เกี่ยวข้อง
                            if any(pattern.replace('*', '').lower() in filename_lower for pattern in purchase_tax_patterns):
                                is_purchase_tax = True
                            
                            # ตรวจสอบว่าไม่ใช่ไฟล์ภาษีขาย
                            if 'ภาษีขาย' in filename or 'sales tax' in filename_lower or 'vat sales' in filename_lower:
                                is_purchase_tax = False
                                logger.debug(f"⏭️ ข้ามไฟล์ภาษีขาย: {filename}")
                            
                            if not is_purchase_tax:
                                continue
                            
                            # ตรวจสอบว่าไฟล์ตรงกับเดือนภาษีที่เลือกหรือไม่
                            file_month_match = False
                            
                            # รูปแบบ 1: PP30-YYYYMM หรือ YYYYMM
                            year_month_patterns = [
                                f"{year_int}{month_int:02d}",  # 202510
                                f"{year_int}-{month_int:02d}",  # 2025-10
                                f"{year_int}-{month_int}",      # 2025-10
                                f"{month_int:02d}-{year_int}",  # 10-2025
                                f"{month_int}-{year_int}",      # 10-2025
                            ]
                            
                            for ymp in year_month_patterns:
                                if ymp in filename_lower or ymp in filename:
                                    file_month_match = True
                                    break
                            
                            # รูปแบบ 2: ตรวจสอบในชื่อไฟล์ว่ามีปีและเดือนที่ตรงกัน
                            if not file_month_match:
                                # หา pattern YYYYMM หรือ YYYY-MM ในชื่อไฟล์
                                year_month_matches = re.findall(r'(\d{4})[-]?(\d{1,2})', filename)
                                for y, m in year_month_matches:
                                    try:
                                        y_int = int(y)
                                        m_int = int(m)
                                        if y_int == year_int and m_int == month_int:
                                            file_month_match = True
                                            break
                                    except ValueError:
                                        pass
                            
                            if file_month_match:
                                purchase_tax_files.append(str(file_path))
                                logger.info(f"✅ พบไฟล์ภาษีซื้อที่ตรงกับเดือน {tax_month}: {filename}")
                            else:
                                logger.debug(f"⏭️ ข้ามไฟล์ภาษีซื้อที่ไม่ตรงกับเดือน {tax_month}: {filename}")
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์ภาษีซื้อใน {month_year_folder}: {e}")
            
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์ภาษีซื้อใน {base_path}: {e}")
                continue
        
        # ลบไฟล์ซ้ำ
        purchase_tax_files = list(set(purchase_tax_files))
        
        # 2. อ่านรายการทั้งหมดจากไฟล์ภาษีซื้อ
        purchase_tax_items = []  # เก็บรายการทั้งหมด
        purchase_tax_row_count = 0
        
        if purchase_tax_files:
            try:
                import openpyxl
                for excel_path in purchase_tax_files:
                    try:
                        wb = openpyxl.load_workbook(excel_path, data_only=True)
                        ws = wb.active
                        
                        # อ่าน header (แถวแรก)
                        headers = []
                        if ws.max_row > 0:
                            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
                        
                        # หาคอลัมน์ต่างๆ
                        col_map = {}
                        for idx, header in enumerate(headers, 1):
                            header_lower = header.lower().strip()
                            if 'เลขที่ใบกำกับ' in header_lower or 'เลขที่ใบกำกับภาษี' in header_lower:
                                col_map['invoice_no'] = idx
                            elif 'เลขที่สารอง' in header_lower or 'เลขที่อ้างอิง' in header_lower or 'reference' in header_lower:
                                col_map['reference_no'] = idx
                            elif 'วันที่ใบกำกับ' in header_lower or 'วันที่ใบกำกับภาษี' in header_lower:
                                col_map['invoice_date'] = idx
                            elif 'ผู้ติดต่อ' in header_lower:
                                col_map['contact'] = idx
                            elif 'มูลค่ารวมภาษี' in header_lower or 'รวมภาษี' in header_lower:
                                col_map['total'] = idx
                        
                        # Debug: Log headers และ col_map
                        logger.info(f"📋 Headers จากไฟล์ภาษีซื้อ: {headers}")
                        logger.info(f"📋 Column mapping: {col_map}")
                        
                        # อ่านรายการทั้งหมด (เริ่มจากแถวที่ 2)
                        for row_idx in range(2, ws.max_row + 1):
                            invoice_no = ''
                            reference_no = ''
                            invoice_date = ''
                            contact = ''
                            total = 0
                            
                            # อ่านข้อมูลจากแต่ละคอลัมน์
                            if 'invoice_no' in col_map:
                                cell = ws.cell(row=row_idx, column=col_map['invoice_no'])
                                if cell.value is not None:
                                    invoice_no = str(cell.value).strip()
                            
                            if 'reference_no' in col_map:
                                cell = ws.cell(row=row_idx, column=col_map['reference_no'])
                                if cell.value is not None:
                                    reference_no = str(cell.value).strip()
                            
                            if 'invoice_date' in col_map:
                                cell = ws.cell(row=row_idx, column=col_map['invoice_date'])
                                if cell.value is not None:
                                    invoice_date = str(cell.value).strip()
                            
                            if 'contact' in col_map:
                                cell = ws.cell(row=row_idx, column=col_map['contact'])
                                if cell.value is not None:
                                    contact = str(cell.value).strip()
                            
                            if 'total' in col_map:
                                cell = ws.cell(row=row_idx, column=col_map['total'])
                                if cell.value is not None:
                                    try:
                                        total = float(cell.value) if isinstance(cell.value, (int, float)) else 0
                                    except (ValueError, TypeError):
                                        total = 0
                            
                            # ตรวจสอบว่าเป็นแถวรวมหรือไม่ (มีคำว่า "รวม" ในคอลัมน์ใดๆ หรือไม่มีเลขที่ใบกำกับภาษีและเลขที่สารอง)
                            is_summary_row = False
                            
                            # ตรวจสอบว่ามีคำว่า "รวม" ในแถวนี้หรือไม่
                            for col_idx in range(1, ws.max_column + 1):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                if cell.value is not None:
                                    cell_value = str(cell.value).strip()
                                    if 'รวม' in cell_value or 'total' in cell_value.lower() or 'summary' in cell_value.lower():
                                        is_summary_row = True
                                        break
                            
                            # ถ้าไม่มีเลขที่ใบกำกับภาษีและเลขที่สารอง และมีค่าในคอลัมน์รวมภาษี อาจเป็นแถวรวม
                            if not invoice_no and not reference_no and total > 0:
                                # ตรวจสอบว่ามีข้อมูลในคอลัมน์อื่นๆ หรือไม่ (ถ้าไม่มีเลย อาจเป็นแถวรวม)
                                has_other_data = False
                                for col_idx in range(1, ws.max_column + 1):
                                    if col_idx in col_map.values():
                                        continue  # ข้ามคอลัมน์ที่ map แล้ว
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    if cell.value is not None and str(cell.value).strip():
                                        cell_value = str(cell.value).strip().lower()
                                        # ถ้าไม่ใช่คำว่า "รวม" หรือ "total" หรือตัวเลข ให้ถือว่ามีข้อมูลอื่น
                                        if 'รวม' not in cell_value and 'total' not in cell_value and not cell_value.replace('.', '').replace(',', '').isdigit():
                                            has_other_data = True
                                            break
                                
                                if not has_other_data:
                                    is_summary_row = True
                            
                            # ข้ามแถวรวม
                            if is_summary_row:
                                logger.debug(f"⏭️ ข้ามแถวรวมแถวที่ {row_idx}")
                                continue
                            
                            # เพิ่มเฉพาะแถวที่มีเลขที่ใบกำกับภาษีหรือเลขที่สารอง
                            if invoice_no or reference_no:
                                purchase_tax_items.append({
                                    'invoice_no': invoice_no,
                                    'reference_no': reference_no,  # เลขที่สารอง (เช่น EXP-20251000003)
                                    'invoice_date': invoice_date,
                                    'contact': contact,
                                    'total': total,
                                    'row': row_idx,
                                    'file': excel_path
                                })
                                purchase_tax_row_count += 1
                                logger.debug(f"📝 อ่านรายการแถว {row_idx}: เลขที่={invoice_no}, เลขที่สารอง={reference_no}")
                        
                        logger.info(f"📊 อ่านไฟล์ภาษีซื้อ {excel_path}: {len([item for item in purchase_tax_items if item['file'] == excel_path])} รายการ")
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถอ่านไฟล์ภาษีซื้อ {excel_path}: {e}")
            except ImportError:
                logger.warning("⚠️ ไม่พบ openpyxl - ไม่สามารถอ่านข้อมูล Excel ได้")
        
        # 3. ค้นหาไฟล์ PDF ในโฟลเดอร์ VAT และดึงเลขที่เอกสารจากชื่อไฟล์
        pdf_files_list = []  # เก็บรายการไฟล์ PDF ทั้งหมด
        vat_folder_path = None
        
        for base_path in base_paths:
            if not base_path.exists():
                continue
            
            try:
                # โครงสร้าง: บัญชี > 002-รายจ่าย > PV > [Year-Month] > VAT/vat/Vat
                account_folder = base_path / "บัญชี"
                if not account_folder.exists():
                    continue
                
                expense_folder = account_folder / "002-รายจ่าย"
                if not expense_folder.exists():
                    continue
                
                pv_folder = expense_folder / "PV"
                if not pv_folder.exists():
                    continue
                
                # ค้นหาโฟลเดอร์เดือน-ปี
                month_year_folder = None
                
                # ขั้นตอนที่ 1: มองหาโฟลเดอร์ ปี-เดือน โดยตรง
                for pattern in month_year_patterns:
                    potential_folder = pv_folder / pattern
                    if potential_folder.exists():
                        month_year_folder = potential_folder
                        break
                
                # ขั้นตอนที่ 2: ถ้าไม่เจอ ให้มองหาโฟลเดอร์ปีก่อน แล้วเข้าไปมองหาโฟลเดอร์ ปี-เดือน
                if not month_year_folder:
                    try:
                        for item in pv_folder.iterdir():
                            if item.is_dir():
                                folder_name = item.name
                                if (year in folder_name or 
                                    str(year_int + 543) in folder_name or  # ปี พ.ศ.
                                    folder_name.isdigit()):
                                    # ค้นหาโฟลเดอร์เดือน-ปีภายในโฟลเดอร์ปี
                                    for sub_item in item.iterdir():
                                        if sub_item.is_dir():
                                            for pattern in month_year_patterns:
                                                if pattern in sub_item.name or sub_item.name == pattern:
                                                    month_year_folder = sub_item
                                                    break
                                            if month_year_folder:
                                                break
                                if month_year_folder:
                                    break
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์เดือน-ปี: {e}")
                
                # ค้นหาโฟลเดอร์ VAT/vat/Vat และไฟล์ PDF
                if month_year_folder:
                    vat_folders = []
                    try:
                        for item in month_year_folder.iterdir():
                            if item.is_dir() and item.name.lower() == "vat":
                                vat_folders.append(item)
                                vat_folder_path = str(item)
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์ VAT ใน {month_year_folder}: {e}")
                    
                    # อ่านไฟล์ที่รองรับ OCR ทั้งหมดในโฟลเดอร์ VAT (รวมโฟลเดอร์ย่อยด้วย)
                    # รายการนามสกุลไฟล์ที่รองรับ OCR (ไม่รวม Excel)
                    supported_extensions_check = ['*.pdf', '*.jpg', '*.jpeg', '*.png', '*.gif', '*.bmp', '*.tiff', '*.tif']
                    supported_extensions_upper_check = ['*.PDF', '*.JPG', '*.JPEG', '*.PNG', '*.GIF', '*.BMP', '*.TIFF', '*.TIF']
                    
                    # นามสกุลไฟล์ที่ต้องข้าม (Excel และไฟล์อื่นๆ ที่ไม่ใช่ OCR)
                    excluded_extensions = ['.xlsx', '.xls', '.XLSX', '.XLS', '.doc', '.docx', '.DOC', '.DOCX']
                    
                    if vat_folders:
                        for vat_folder in vat_folders:
                            try:
                                # ใช้ rglob() เพื่อค้นหาไฟล์ที่รองรับ OCR ทั้งหมด รวมโฟลเดอร์ย่อยด้วย
                                pdf_files = []
                                # รายการนามสกุลที่รองรับ OCR เท่านั้น (ไม่รวม Excel)
                                ocr_supported_suffixes_lower = ['.pdf', '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif']
                                ocr_supported_suffixes_upper = ['.PDF', '.JPG', '.JPEG', '.PNG', '.GIF', '.BMP', '.TIFF', '.TIF']
                                ocr_supported_suffixes = ocr_supported_suffixes_lower + ocr_supported_suffixes_upper
                                
                                # ใช้ set เพื่อป้องกันไฟล์ซ้ำ
                                seen_files = set()
                                
                                # ใช้ rglob() เพื่อค้นหาไฟล์ในโฟลเดอร์หลักและโฟลเดอร์ย่อยทั้งหมด (recursive)
                                for ext in supported_extensions_check + supported_extensions_upper_check:
                                    # rglob() จะค้นหาไฟล์ในโฟลเดอร์หลักและโฟลเดอร์ย่อยทั้งหมดแบบ recursive
                                    found_files = list(vat_folder.rglob(ext))
                                    logger.debug(f"🔍 ค้นหาไฟล์ด้วย pattern '{ext}' ใน {vat_folder}: พบ {len(found_files)} รายการ")
                                    
                                    # กรองเฉพาะไฟล์จริงๆ ที่มีนามสกุลรองรับ OCR เท่านั้น
                                    for f in found_files:
                                        # ตรวจสอบว่าเป็นไฟล์จริงๆ (ไม่ใช่โฟลเดอร์)
                                        if not f.is_file():
                                            logger.debug(f"⏭️ ข้าม (ไม่ใช่ไฟล์): {f}")
                                            continue
                                        
                                        # ตรวจสอบว่านามสกุลไฟล์อยู่ในรายการที่รองรับ OCR
                                        file_suffix = f.suffix
                                        if file_suffix not in ocr_supported_suffixes:
                                            logger.debug(f"⏭️ ข้ามไฟล์ที่ไม่รองรับ OCR: {f.name} (นามสกุล: {file_suffix})")
                                            continue
                                        
                                        # ตรวจสอบอีกครั้งว่าไม่ใช่ไฟล์ Excel (เพื่อความปลอดภัย)
                                        if file_suffix.lower() in ['.xlsx', '.xls']:
                                            logger.debug(f"⏭️ ข้ามไฟล์ Excel: {f.name}")
                                            continue
                                        
                                        # ใช้ path แบบ absolute string เพื่อตรวจสอบซ้ำ
                                        file_path_str = str(f.resolve())
                                        if file_path_str not in seen_files:
                                            seen_files.add(file_path_str)
                                            pdf_files.append(f)
                                            # Log เพื่อดูว่าไฟล์มาจากโฟลเดอร์ไหน
                                            relative_path = f.relative_to(vat_folder)
                                            logger.debug(f"✅ เพิ่มไฟล์: {f.name} (จาก: {relative_path.parent if relative_path.parent != Path('.') else 'โฟลเดอร์หลัก'})")
                                
                                logger.info(f"📁 ค้นหาไฟล์ที่รองรับ OCR ในโฟลเดอร์ VAT (รวมโฟลเดอร์ย่อย): {vat_folder}")
                                logger.info(f"   ✅ พบ {len(pdf_files)} ไฟล์ที่รองรับ OCR (ไม่รวม Excel และโฟลเดอร์)")
                                
                                # Log รายชื่อไฟล์ที่พบเพื่อ debug (แสดง path เต็ม)
                                if pdf_files:
                                    logger.info(f"📋 รายการไฟล์ที่รองรับ OCR ({len(pdf_files)} ไฟล์):")
                                    for pf in pdf_files:
                                        relative_path = pf.relative_to(vat_folder)
                                        folder_info = f" (ในโฟลเดอร์ย่อย: {relative_path.parent})" if relative_path.parent != Path('.') else " (ในโฟลเดอร์หลัก)"
                                        logger.info(f"   - {pf.name} ({pf.suffix}){folder_info}")
                                else:
                                    logger.warning(f"⚠️ ไม่พบไฟล์ที่รองรับ OCR ในโฟลเดอร์ VAT: {vat_folder}")
                                
                                for pdf_file in pdf_files:
                                    # ตรวจสอบอีกครั้งว่าเป็นไฟล์จริงๆ (ไม่ใช่ directory)
                                    if not pdf_file.is_file():
                                        logger.debug(f"⏭️ ข้าม (ไม่ใช่ไฟล์): {pdf_file.name}")
                                        continue
                                    
                                    # ตรวจสอบอีกครั้งว่านามสกุลรองรับ OCR
                                    if pdf_file.suffix not in ocr_supported_suffixes:
                                        logger.debug(f"⏭️ ข้ามไฟล์ที่ไม่รองรับ OCR: {pdf_file.name} (นามสกุล: {pdf_file.suffix})")
                                        continue
                                    
                                    # ดึงเลขที่เอกสารจากชื่อไฟล์ (ลองหลายรูปแบบ)
                                    filename = pdf_file.name
                                    filename_no_ext = pdf_file.stem  # ไม่มีนามสกุลไฟล์
                                    
                                    # รูปแบบ 1: ชื่อไฟล์อาจมีเลขที่เอกสารอยู่ต้นๆ เช่น "34815_xxx.pdf"
                                    # รูปแบบ 2: ชื่อไฟล์อาจมีเลขที่เอกสารอยู่กลาง เช่น "xxx_34815_xxx.pdf"
                                    # รูปแบบ 3: ชื่อไฟล์อาจมีเลขที่เอกสารอยู่ท้าย เช่น "xxx_34815.pdf"
                                    
                                    # ลองหาเลขที่เอกสารจากชื่อไฟล์ (อาจเป็นตัวเลขหรือตัวอักษร+ตัวเลข)
                                    # Pattern: หาตัวเลขที่อาจเป็นเลขที่เอกสาร (อย่างน้อย 4 หลัก)
                                    doc_number_candidates = re.findall(r'(\d{4,})', filename_no_ext)
                                    
                                    # หรือหาตัวอักษร+ตัวเลข เช่น "EXP-20251100058"
                                    doc_number_candidates.extend(re.findall(r'([A-Z]+[-]?\d{4,})', filename_no_ext))
                                    
                                    pdf_files_list.append({
                                        'filename': filename,
                                        'path': str(pdf_file),
                                        'doc_numbers': doc_number_candidates,  # รายการเลขที่เอกสารที่อาจเป็น
                                        'matched': False  # ใช้สำหรับ tracking ว่า match กับ Excel หรือยัง
                                    })
                                    
                                    logger.debug(f"📄 พบไฟล์ที่รองรับ OCR: {filename} (เลขที่ที่พบ: {doc_number_candidates})")
                            except Exception as e:
                                logger.warning(f"⚠️ ไม่สามารถอ่านไฟล์ PDF ใน {vat_folder}: {e}")
            
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์ PDF ใน {base_path}: {e}")
                continue
        
        # 4. เปรียบเทียบจำนวนรายการกับจำนวนไฟล์ที่รองรับ OCR (ไม่รวม Excel)
        pdf_files_count = len(pdf_files_list)
        
        # Log เพื่อ debug
        logger.info(f"📊 สรุปการเปรียบเทียบ:")
        logger.info(f"   - จำนวนรายการในไฟล์ภาษีซื้อ (Excel): {purchase_tax_row_count} รายการ")
        logger.info(f"   - จำนวนไฟล์ที่รองรับ OCR (PDF/JPG/PNG) ในโฟลเดอร์ VAT: {pdf_files_count} ไฟล์ (ไม่รวม Excel)")
        
        # 5. สรุปผลการเปรียบเทียบ (เปรียบเทียบแค่จำนวนเท่านั้น)
        count_match = purchase_tax_row_count == pdf_files_count
        all_match = count_match
        
        return jsonify({
            'success': True,
            'exists': len(purchase_tax_files) > 0,
            'filePath': purchase_tax_files[0] if purchase_tax_files else None,
            'purchaseTaxFiles': purchase_tax_files,
            'purchaseTaxRowCount': purchase_tax_row_count,
            'pdfFilesCount': pdf_files_count,
            'vatFolderPath': vat_folder_path,
            'countMatch': count_match,
            'allMatch': all_match,
            'message': f'จำนวนรายการในไฟล์ภาษีซื้อ ({purchase_tax_row_count}) {"ตรงกับ" if count_match else "ไม่ตรงกับ"} จำนวนไฟล์ที่รองรับ OCR (PDF/JPG/PNG) ในโฟลเดอร์ VAT ({pdf_files_count}) - ไม่รวมไฟล์ Excel'
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการตรวจสอบไฟล์ภาษีซื้อ: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/compare-trial-balance-files', methods=['POST'])
def compare_trial_balance_files():
    """เปรียบเทียบข้อมูลรายการในงบทดลองและจำนวนไฟล์ OCR"""
    try:
        data = request.json
        tax_month = data.get('taxMonth', '')  # Format: YYYY-MM
        company = data.get('company', '')
        
        if not tax_month or not company:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเดือนภาษีและชื่อบริษัท'
            }), 400
        
        # แปลงเดือนจาก YYYY-MM
        if '-' not in tax_month:
            return jsonify({
                'success': False,
                'error': f'รูปแบบเดือนภาษีไม่ถูกต้อง: {tax_month}'
            }), 400
        
        parts = tax_month.split('-')
        if len(parts) != 2:
            return jsonify({
                'success': False,
                'error': f'รูปแบบเดือนภาษีไม่ถูกต้อง: {tax_month}'
            }), 400
        
        year, month = parts
        year_int = int(year)
        month_int = int(month)
        
        # สร้างรูปแบบปี-เดือน (รองรับหลายรูปแบบ)
        month_year_patterns = [
            f"{year_int}-{month_int:02d}",   # 2025-10 (ปี-เดือน)
            f"{year_int}-{month_int}",       # 2025-10 (ปี-เดือน)
            f"{month_int:02d}-{year_int}",   # 10-2025 (เดือน-ปี)
            f"{month_int}-{year_int}",       # 10-2025 (เดือน-ปี)
        ]
        
        # 1. ค้นหาไฟล์ Excel ที่มีคำว่า "งบทดลอง" ในโฟลเดอร์ VAT/vat/Vat
        trial_balance_files = []
        trial_balance_count = 0
        
        base_paths = [
            Path(f"V:/A.โฟร์เดอร์หลัก/{company}"),
            Path(f"V:/AA.โฟรเดอร์หลัก/{company}"),
            Path(f"V:/AAA.โฟรเดอร์หลัก/{company}")
        ]
        
        for base_path in base_paths:
            if not base_path.exists():
                continue
            
            try:
                # โครงสร้าง: บัญชี > 002-รายจ่าย > PV
                account_folder = base_path / "บัญชี"
                if not account_folder.exists():
                    continue
                
                expense_folder = account_folder / "002-รายจ่าย"
                if not expense_folder.exists():
                    continue
                
                pv_folder = expense_folder / "PV"
                if not pv_folder.exists():
                    continue
                
                # ค้นหาโฟลเดอร์เดือน-ปี
                month_year_folder = None
                
                # ขั้นตอนที่ 1: มองหาโฟลเดอร์ ปี-เดือน โดยตรง
                for pattern in month_year_patterns:
                    potential_folder = pv_folder / pattern
                    if potential_folder.exists():
                        month_year_folder = potential_folder
                        break
                
                # ขั้นตอนที่ 2: ถ้าไม่เจอ ให้มองหาโฟลเดอร์ปีก่อน แล้วเข้าไปมองหาโฟลเดอร์ ปี-เดือน
                if not month_year_folder:
                    try:
                        for item in pv_folder.iterdir():
                            if item.is_dir():
                                folder_name = item.name
                                if (year in folder_name or 
                                    str(year_int + 543) in folder_name or  # ปี พ.ศ.
                                    folder_name.isdigit()):
                                    # ค้นหาโฟลเดอร์เดือน-ปีภายในโฟลเดอร์ปี
                                    for sub_item in item.iterdir():
                                        if sub_item.is_dir():
                                            for pattern in month_year_patterns:
                                                if pattern in sub_item.name or sub_item.name == pattern:
                                                    month_year_folder = sub_item
                                                    break
                                            if month_year_folder:
                                                break
                                if month_year_folder:
                                    break
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์เดือน-ปี: {e}")
                
                # ค้นหาโฟลเดอร์ VAT/vat/Vat และไฟล์ Excel ที่มีคำว่า "งบทดลอง"
                if month_year_folder:
                    vat_folders = []
                    try:
                        for item in month_year_folder.iterdir():
                            if item.is_dir() and item.name.lower() == "vat":
                                vat_folders.append(item)
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์ VAT ใน {month_year_folder}: {e}")
                    
                    # ค้นหาไฟล์ Excel ที่มีคำว่า "งบทดลอง" ในโฟลเดอร์ VAT
                    if vat_folders:
                        for vat_folder in vat_folders:
                            try:
                                for excel_file in vat_folder.glob("*.xlsx"):
                                    if excel_file.is_file():
                                        filename_lower = excel_file.name.lower()
                                        # ตรวจสอบว่าชื่อไฟล์มีคำว่า "งบทดลอง"
                                        if "งบทดลอง" in excel_file.name:
                                            trial_balance_files.append(str(excel_file))
                            except Exception as e:
                                logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์งบทดลองใน {vat_folder}: {e}")
                    else:
                        # ถ้าไม่พบโฟลเดอร์ VAT ให้ค้นหาในโฟลเดอร์เดือน-ปีโดยตรง
                        try:
                            for excel_file in month_year_folder.glob("*.xlsx"):
                                if excel_file.is_file():
                                    # ตรวจสอบว่าชื่อไฟล์มีคำว่า "งบทดลอง"
                                    if "งบทดลอง" in excel_file.name:
                                        trial_balance_files.append(str(excel_file))
                        except Exception as e:
                            logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์งบทดลองใน {month_year_folder}: {e}")
            
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์งบทดลองใน {base_path}: {e}")
                continue
        
        # ลบไฟล์ซ้ำ
        trial_balance_files = list(set(trial_balance_files))
        
        # อ่านไฟล์ Excel และอ่านข้อมูลงบทดลอง
        trial_balance_data = {
            'company_name': None,
            'report_date': None,
            'period': None,
            'purchase_tax': {'debit': 0, 'credit': 0},
            'purchase_tax_not_due': {'debit': 0, 'credit': 0},
            'sales_tax_vat30': {'debit': 0, 'credit': 0}
        }
        
        if trial_balance_files:
            try:
                import openpyxl
                for excel_path in trial_balance_files:
                    try:
                        wb = openpyxl.load_workbook(excel_path, data_only=True)
                        ws = wb.active
                        
                        # 1. อ่านชื่อกิจการจาก K3
                        company_name_cell = ws['K3']
                        if company_name_cell.value:
                            trial_balance_data['company_name'] = str(company_name_cell.value).strip()
                        
                        # 2. อ่านวันที่ออกรายงานจาก K4
                        report_date_cell = ws['K4']
                        if report_date_cell.value:
                            trial_balance_data['report_date'] = str(report_date_cell.value).strip()
                        
                        # 3. อ่านช่วงเวลาจาก K5
                        period_cell = ws['K5']
                        if period_cell.value:
                            trial_balance_data['period'] = str(period_cell.value).strip()
                        
                        # 4. ค้นหาข้อมูลโดยใช้โค้ดบัญชี (Column A) และชื่อบัญชี (Column B)
                        # กำหนดโค้ดบัญชีที่ต้องการ:
                        # - 115401 = ภาษีซื้อ
                        # - 115402 = ภาษีซื้อยังไม่ถึงกำหนด
                        # - 215101 = ภาษีขาย ภ.พ.30
                        for row in range(3, ws.max_row + 1):
                            cell_a = ws[f'A{row}']  # โค้ดบัญชี
                            cell_b = ws[f'B{row}']  # ชื่อบัญชี
                            
                            # อ่านโค้ดบัญชีจาก Column A
                            account_code = None
                            if cell_a.value:
                                try:
                                    # แปลงเป็น string แล้วลบช่องว่าง
                                    account_code_str = str(cell_a.value).strip()
                                    # ถ้าเป็นตัวเลข ให้แปลงเป็น string เพื่อเปรียบเทียบ
                                    if account_code_str.replace('.', '').isdigit():
                                        account_code = account_code_str.split('.')[0]  # เอาเฉพาะส่วนที่เป็นจำนวนเต็ม
                                    else:
                                        account_code = account_code_str
                                except (ValueError, TypeError):
                                    account_code = None
                            
                            # อ่านชื่อบัญชีจาก Column B
                            account_name = None
                            if cell_b.value:
                                account_name = str(cell_b.value).strip()
                            
                            # ถ้าไม่มีโค้ดบัญชีหรือชื่อบัญชี ให้ข้าม
                            if not account_code and not account_name:
                                continue
                            
                            # อ่านยอดคงเหลือจากคอลัมน์ G (เดบิต) และ H (เครดิต)
                            cell_g = ws[f'G{row}']
                            cell_h = ws[f'H{row}']
                            
                            debit_value = 0
                            credit_value = 0
                            
                            # อ่านค่าเดบิตจาก G
                            if cell_g.value is not None:
                                try:
                                    debit_value = float(cell_g.value) if isinstance(cell_g.value, (int, float)) else 0
                                except (ValueError, TypeError):
                                    debit_value = 0
                            
                            # อ่านค่าเครดิตจาก H
                            if cell_h.value is not None:
                                try:
                                    credit_value = float(cell_h.value) if isinstance(cell_h.value, (int, float)) else 0
                                except (ValueError, TypeError):
                                    credit_value = 0
                            
                            # ตรวจสอบจากโค้ดบัญชี (Column A) และชื่อบัญชี (Column B)
                            # 1. ภาษีซื้อ: โค้ด = 115401 และชื่อ = ภาษีซื้อ
                            if account_code == '115401' and account_name and 'ภาษีซื้อ' in account_name and 'ยังไม่ถึงกำหนด' not in account_name:
                                trial_balance_data['purchase_tax']['debit'] += debit_value
                                trial_balance_data['purchase_tax']['credit'] += credit_value
                                logger.info(f"✅ พบภาษีซื้อ: โค้ด={account_code}, ชื่อ={account_name}, เดบิต={debit_value}, เครดิต={credit_value}")
                            
                            # 2. ภาษีซื้อยังไม่ถึงกำหนด: โค้ด = 115402 และชื่อ = ภาษีซื้อยังไม่ถึงกำหนด
                            elif account_code == '115402' and account_name and 'ภาษีซื้อยังไม่ถึงกำหนด' in account_name:
                                trial_balance_data['purchase_tax_not_due']['debit'] += debit_value
                                trial_balance_data['purchase_tax_not_due']['credit'] += credit_value
                                logger.info(f"✅ พบภาษีซื้อยังไม่ถึงกำหนด: โค้ด={account_code}, ชื่อ={account_name}, เดบิต={debit_value}, เครดิต={credit_value}")
                            
                            # 3. ภาษีขาย ภ.พ.30: โค้ด = 215101 และชื่อ = ภาษีขาย ภ.พ.30
                            elif account_code == '215101' and account_name and 'ภาษีขาย' in account_name and 'ภ.พ.30' in account_name:
                                trial_balance_data['sales_tax_vat30']['debit'] += debit_value
                                trial_balance_data['sales_tax_vat30']['credit'] += credit_value
                                logger.info(f"✅ พบภาษีขาย ภ.พ.30: โค้ด={account_code}, ชื่อ={account_name}, เดบิต={debit_value}, เครดิต={credit_value}")
                            
                            # Fallback: ถ้าไม่มีโค้ดบัญชี ให้ใช้การตรวจสอบจากชื่อบัญชีแบบเดิม (เพื่อรองรับกรณีที่ไม่มีโค้ด)
                            elif not account_code:
                                if account_name:
                                    if "ภาษีซื้อ" in account_name and "ยังไม่ถึงกำหนด" not in account_name:
                                        trial_balance_data['purchase_tax']['debit'] += debit_value
                                        trial_balance_data['purchase_tax']['credit'] += credit_value
                                        logger.info(f"⚠️ พบภาษีซื้อ (ไม่มีโค้ด): ชื่อ={account_name}, เดบิต={debit_value}, เครดิต={credit_value}")
                                    elif "ภาษีซื้อยังไม่ถึงกำหนด" in account_name:
                                        trial_balance_data['purchase_tax_not_due']['debit'] += debit_value
                                        trial_balance_data['purchase_tax_not_due']['credit'] += credit_value
                                        logger.info(f"⚠️ พบภาษีซื้อยังไม่ถึงกำหนด (ไม่มีโค้ด): ชื่อ={account_name}, เดบิต={debit_value}, เครดิต={credit_value}")
                                    elif "ภาษีขาย" in account_name and "ภ.พ.30" in account_name:
                                        trial_balance_data['sales_tax_vat30']['debit'] += debit_value
                                        trial_balance_data['sales_tax_vat30']['credit'] += credit_value
                                        logger.info(f"⚠️ พบภาษีขาย ภ.พ.30 (ไม่มีโค้ด): ชื่อ={account_name}, เดบิต={debit_value}, เครดิต={credit_value}")
                        
                        # นับจำนวนรายการ (ใช้จำนวนแถวที่มีข้อมูล)
                        row_count = ws.max_row - 2 if ws.max_row > 2 else 0
                        trial_balance_count += row_count
                        
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถอ่านไฟล์ Excel {excel_path}: {e}")
            except ImportError:
                logger.warning("⚠️ ไม่พบ openpyxl - ไม่สามารถอ่านข้อมูล Excel ได้")
        
        # Step 4 ไม่ต้องเปรียบเทียบกับไฟล์ OCR - แสดงเฉพาะข้อมูลงบทดลองเท่านั้น
        return jsonify({
            'success': True,
            'trialBalanceCount': trial_balance_count,
            'trialBalanceFiles': trial_balance_files,
            'trialBalanceData': trial_balance_data,
            'message': f'พบข้อมูลงบทดลอง: {trial_balance_count} รายการ'
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการเปรียบเทียบข้อมูล: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def convert_buddhist_to_christian_date(date_str):
    """
    แปลงวันที่จากปี พ.ศ. เป็น ค.ศ.
    รองรับรูปแบบ: dd/mm/yyyy (พ.ศ.) -> dd/mm/yyyy (ค.ศ.)
    ตัวอย่าง: 16/10/2568 -> 16/10/2025
    """
    if not date_str:
        return date_str
    
    import re
    # Pattern: dd/mm/yyyy (พ.ศ.)
    pattern = r'(\d{1,2})/(\d{1,2})/(\d{4})'
    match = re.match(pattern, date_str.strip())
    
    if match:
        day = match.group(1)
        month = match.group(2)
        year_buddhist = int(match.group(3))
        
        # ถ้าปีมากกว่า 2500 แสดงว่าเป็นปี พ.ศ. ให้แปลงเป็น ค.ศ.
        if year_buddhist > 2500:
            year_christian = year_buddhist - 543
            return f"{day}/{month}/{year_christian}"
        # ถ้าปีน้อยกว่า 2500 แสดงว่าเป็นปี ค.ศ. อยู่แล้ว
        else:
            return date_str
    
    return date_str


@app.route('/api/auditcheck/compare-purchase-tax-ocr', methods=['POST'])
def compare_purchase_tax_ocr():
    """เปรียบเทียบข้อมูลภาษีซื้อและไฟล์ OCR"""
    try:
        data = request.json
        tax_month = data.get('taxMonth', '')  # Format: YYYY-MM
        company = data.get('company', '')
        
        if not tax_month or not company:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเดือนภาษีและชื่อบริษัท'
            }), 400
        
        # แปลงเดือนจาก YYYY-MM
        year, month = tax_month.split('-')
        year_int = int(year)
        month_int = int(month)
        thai_year = year_int + 543  # แปลงเป็นปี พ.ศ.
        
        base_paths = [
            Path(f"V:/A.โฟร์เดอร์หลัก/{company}"),
            Path(f"V:/AA.โฟรเดอร์หลัก/{company}"),
            Path(f"V:/AAA.โฟรเดอร์หลัก/{company}")
        ]
        
        # 1. ค้นหาไฟล์ภาษีซื้อเฉพาะเดือนภาษีที่เลือก
        purchase_tax_files = []
        purchase_tax_patterns = [
            f"*ภาษีซื้อ*{year}*{month}*",
            f"*ภาษีซื้อ*{thai_year}*{month}*",
            f"*Purchase Tax*{year}*{month}*",
            f"*purchase tax*{year}*{month}*",
            f"*VAT Purchase*{year}*{month}*",
        ]
        
        # รูปแบบเดือน-ปีที่ใช้ในการกรอง
        month_year_formats = [
            f"{year}-{month_int:02d}",  # 2025-10
            f"{year}-{month_int}",      # 2025-10
            f"{month_int:02d}-{year}",  # 10-2025
            f"{month_int}-{year}",      # 10-2025
            f"{year}{month_int:02d}",   # 202510
            f"{year}{month_int}",       # 202510
            f"{month_int:02d}{year}",   # 102025
            f"{month_int}{year}",       # 102025
        ]
        
        for base_path in base_paths:
            if not base_path.exists():
                continue
            for pattern in purchase_tax_patterns:
                for file_path in base_path.rglob(pattern):
                    if file_path.is_file() and file_path.suffix.lower() in ['.xlsx', '.xls']:
                        # ตรวจสอบว่าไฟล์ตรงกับเดือนภาษีที่เลือกหรือไม่
                        file_name = file_path.name
                        file_path_str = str(file_path)
                        
                        # ตรวจสอบว่า path หรือชื่อไฟล์มีเดือน-ปีที่เลือกหรือไม่
                        matches_month = False
                        
                        # แยก path เป็นส่วนๆ เพื่อตรวจสอบ
                        path_parts = file_path_str.split('\\')
                        
                        # ตรวจสอบรูปแบบเดือน-ปีใน path
                        for part in path_parts:
                            # ตรวจสอบรูปแบบ 2025-10 หรือ 10-2025
                            if (f"{year_int}-{month_int:02d}" in part or 
                                f"{month_int:02d}-{year_int}" in part or
                                f"{year_int}-{month_int}" in part or
                                f"{month_int}-{year_int}" in part):
                                matches_month = True
                                break
                            
                            # ตรวจสอบรูปแบบ PP30-YYYYMM หรือ PP30-YYYY-MM
                            if (f"PP30-{year_int}{month_int:02d}" in part or
                                f"PP30-{year_int}-{month_int:02d}" in part or
                                f"PP30-{year_int}-{month_int}" in part):
                                matches_month = True
                                break
                            
                            # ตรวจสอบรูปแบบเดือน.ปี พ.ศ. (เช่น 10.2568)
                            if (f"{month_int:02d}.{thai_year}" in part or
                                f"{month_int}.{thai_year}" in part):
                                matches_month = True
                                break
                            
                            # ตรวจสอบรูปแบบ YYYYMM (เช่น 202510)
                            if f"{year_int}{month_int:02d}" in part:
                                # ตรวจสอบว่าไม่ใช่ส่วนของตัวเลขอื่น (เช่น 2025100 ไม่ควร match)
                                if part == f"{year_int}{month_int:02d}" or f"-{year_int}{month_int:02d}" in part or f"{year_int}{month_int:02d}-" in part:
                                    matches_month = True
                                    break
                        
                        # ตรวจสอบในชื่อไฟล์
                        if not matches_month:
                            # ตรวจสอบรูปแบบ PP30-YYYYMM หรือ PP30-YYYY-MM ในชื่อไฟล์
                            if (f"PP30-{year_int}{month_int:02d}" in file_name or
                                f"PP30-{year_int}-{month_int:02d}" in file_name or
                                f"PP30-{year_int}-{month_int}" in file_name):
                                matches_month = True
                        
                        # ตรวจสอบสุดท้าย: ตรวจสอบว่าไม่มีเดือนอื่นที่ชัดเจนใน path
                        if matches_month:
                            # ตรวจสอบว่าไม่มีเดือนอื่น (เช่น 09, 11, 12) ในรูปแบบที่ชัดเจน
                            for other_month in range(1, 13):
                                if other_month == month_int:
                                    continue
                                
                                other_month_str = f"{other_month:02d}"
                                
                                # ตรวจสอบรูปแบบเดือน-ปี (เช่น 09-2025, 2025-09)
                                if (f"{other_month_str}-{year_int}" in file_path_str or
                                    f"{year_int}-{other_month_str}" in file_path_str or
                                    f"{other_month_str}.{thai_year}" in file_path_str or
                                    f"PP30-{year_int}{other_month_str}" in file_path_str):
                                    # ถ้าพบเดือนอื่นที่ชัดเจน ให้ไม่เพิ่มไฟล์นี้
                                    matches_month = False
                                    break
                        
                        if matches_month:
                            purchase_tax_files.append(str(file_path))
        
        purchase_tax_files = list(set(purchase_tax_files))
        
        # 2. ค้นหาไฟล์ Excel OCR (จาก Step 2)
        month_year_patterns = [
            f"{year_int}-{month_int:02d}",
            f"{year_int}-{month_int}",
            f"{month_int:02d}-{year_int}",
            f"{month_int}-{year_int}",
        ]
        
        ocr_excel_files = []
        vat_folder_path = None  # เก็บ path ของโฟลเดอร์ VAT
        
        for base_path in base_paths:
            if not base_path.exists():
                continue
            try:
                account_folder = base_path / "บัญชี"
                if not account_folder.exists():
                    continue
                expense_folder = account_folder / "002-รายจ่าย"
                if not expense_folder.exists():
                    continue
                pv_folder = expense_folder / "PV"
                if not pv_folder.exists():
                    continue
                
                # ค้นหาโฟลเดอร์เดือน-ปี
                month_year_folder = None
                for pattern in month_year_patterns:
                    potential_folder = pv_folder / pattern
                    if potential_folder.exists():
                        month_year_folder = potential_folder
                        break
                
                if not month_year_folder:
                    try:
                        for item in pv_folder.iterdir():
                            if item.is_dir():
                                folder_name = item.name
                                if (year in folder_name or str(year_int + 543) in folder_name or folder_name.isdigit()):
                                    for sub_item in item.iterdir():
                                        if sub_item.is_dir():
                                            for pattern in month_year_patterns:
                                                if pattern in sub_item.name or sub_item.name == pattern:
                                                    month_year_folder = sub_item
                                                    break
                                            if month_year_folder:
                                                break
                                if month_year_folder:
                                    break
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์เดือน-ปี: {e}")
                
                # ค้นหาไฟล์ Excel OCR ในโฟลเดอร์ VAT
                if month_year_folder:
                    vat_folders = []
                    try:
                        for item in month_year_folder.iterdir():
                            if item.is_dir() and item.name.lower() == "vat":
                                vat_folders.append(item)
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์ VAT: {e}")
                    
                    if vat_folders:
                        for vat_folder in vat_folders:
                            try:
                                # เก็บ path ของโฟลเดอร์ VAT (ใช้ตัวแรกที่เจอ)
                                if not vat_folder_path:
                                    vat_folder_path = str(vat_folder)
                                    logger.info(f"📁 พบโฟลเดอร์ VAT: {vat_folder_path}")
                                
                                for excel_file in vat_folder.glob("*.xlsx"):
                                    if excel_file.is_file():
                                        filename_lower = excel_file.name.lower()
                                        # อ่านเฉพาะไฟล์ที่มีคำว่า "ocr" ในชื่อไฟล์เท่านั้น
                                        if "ocr" in filename_lower:
                                            ocr_excel_files.append(str(excel_file))
                            except Exception as e:
                                logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์ OCR: {e}")
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์ OCR ใน {base_path}: {e}")
                continue
        
        ocr_excel_files = list(set(ocr_excel_files))
        
        # 3. อ่านข้อมูลจากไฟล์ภาษีซื้อ
        purchase_tax_data = []
        if purchase_tax_files:
            try:
                import openpyxl
                for excel_path in purchase_tax_files:
                    try:
                        wb = openpyxl.load_workbook(excel_path, data_only=True)
                        ws = wb.active
                        
                        # อ่าน header (แถวแรก)
                        headers = []
                        if ws.max_row > 0:
                            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
                        
                        # หาคอลัมน์ตามชื่อ header
                        col_map = {}
                        for idx, header in enumerate(headers, 1):
                            header_lower = header.lower().strip()
                            
                            # เลขที่ใบกำกับภาษี
                            if 'เลขที่ใบกำกับ' in header_lower or 'เลขที่ใบกำกับภาษี' in header_lower:
                                col_map['invoice_no'] = idx
                            # วันที่ใบกำกับภาษี
                            elif 'วันที่ใบกำกับ' in header_lower or 'วันที่ใบกำกับภาษี' in header_lower:
                                col_map['invoice_date'] = idx
                            # ผู้ติดต่อ
                            elif 'ผู้ติดต่อ' in header_lower:
                                col_map['contact'] = idx
                            # เลขทะเบียนผู้เสียภาษี
                            elif 'เลขทะเบียนผู้เสียภาษี' in header_lower:
                                col_map['tax_id'] = idx
                            # สาขา/สำนักงานใหญ่
                            elif 'สาขา' in header_lower and ('สำนักงาน' in header_lower or 'ใหญ่' in header_lower):
                                col_map['branch'] = idx
                            # รายการยกเว้นภาษี
                            elif 'รายการยกเว้นภาษี' in header_lower:
                                col_map['exempt'] = idx
                            # รายการภาษี 0%
                            elif 'รายการภาษี 0%' in header_lower or 'ภาษี 0%' in header_lower:
                                col_map['tax_0'] = idx
                            # รายการภาษี 7%
                            elif 'รายการภาษี 7%' in header_lower or 'ภาษี 7%' in header_lower:
                                col_map['tax_7'] = idx
                            # ภาษีมูลค่าเพิ่ม
                            elif 'ภาษีมูลค่าเพิ่ม' in header_lower or ('vat' in header_lower and 'ภาษี' in header_lower):
                                col_map['vat'] = idx
                            # มูลค่ารวมภาษี
                            elif 'มูลค่ารวมภาษี' in header_lower or 'รวมภาษี' in header_lower:
                                col_map['total_with_vat'] = idx
                            # ประเภทใบกำกับ
                            elif 'ประเภทใบกำกับ' in header_lower:
                                col_map['invoice_type'] = idx
                            # เลขที่เอกสารอ้างอิง
                            elif 'เลขที่เอกสารอ้างอิง' in header_lower or 'เลขที่อ้างอิง' in header_lower:
                                col_map['reference_no'] = idx
                        
                        # Debug: Log headers และ col_map
                        logger.info(f"📋 Headers จากไฟล์ภาษีซื้อ: {headers[:10]}...")  # แสดง 10 headers แรก
                        logger.info(f"📋 Column mapping: {col_map}")
                        
                        # อ่านข้อมูล (เริ่มจากแถวที่ 2)
                        for row_idx in range(2, ws.max_row + 1):
                            row_data = {
                                'row': row_idx,
                                'invoice_no': '',
                                'invoice_date': '',
                                'contact': '',
                                'tax_id': '',
                                'branch': '',
                                'reference_no': '',  # เพิ่ม reference_no
                                'exempt': 0,
                                'tax_0': 0,
                                'tax_7': 0,
                                'vat': 0,
                                'total_with_vat': 0,
                                'invoice_type': ''
                            }
                            
                            # อ่านข้อมูลจากแต่ละคอลัมน์
                            for key, col_idx in col_map.items():
                                cell = ws.cell(row=row_idx, column=col_idx)
                                if cell.value is not None:
                                    if key in ['exempt', 'tax_0', 'tax_7', 'vat', 'total_with_vat']:
                                        try:
                                            row_data[key] = float(cell.value) if isinstance(cell.value, (int, float)) else 0
                                        except (ValueError, TypeError):
                                            row_data[key] = 0
                                    else:
                                        row_data[key] = str(cell.value).strip()
                            
                            # ตรวจสอบว่าเป็นรายการ "สรุปรวม" หรือไม่
                            # ถ้าคอลัมน์ "สาขา/สำนักงานใหญ่" เป็น "รวม" หรือ "summary" หรือ "total" ให้ข้ามไป
                            branch_value = row_data.get('branch', '').lower().strip()
                            is_summary_row = branch_value in ['รวม', 'summary', 'total', 'รวมทั้งหมด', 'รวมทั้งสิ้น']
                            
                            # เพิ่มเฉพาะแถวที่มีข้อมูลและไม่ใช่รายการสรุปรวม
                            if not is_summary_row and (row_data['invoice_no'] or row_data['total_with_vat'] > 0 or row_data['tax_7'] > 0):
                                purchase_tax_data.append(row_data)
                            elif is_summary_row:
                                logger.info(f"🔍 ข้ามรายการสรุปรวมแถวที่ {row_idx}: สาขา = '{row_data.get('branch', '')}'")
                        
                        logger.info(f"📊 อ่านข้อมูลภาษีซื้อได้ {len(purchase_tax_data)} รายการ")
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถอ่านไฟล์ภาษีซื้อ {excel_path}: {e}")
            except ImportError:
                logger.warning("⚠️ ไม่พบ openpyxl - ไม่สามารถอ่านข้อมูล Excel ได้")
        
        # 4. อ่านข้อมูล OCR (จาก Step 2 หรือจากไฟล์ Excel OCR)
        ocr_data = []
        
        # ตรวจสอบว่ามีข้อมูล OCR จาก Step 2 หรือไม่
        ocr_data_from_step2 = data.get('ocrDataFromStep2', [])
        if ocr_data_from_step2 and len(ocr_data_from_step2) > 0:
            logger.info(f"📊 ใช้ข้อมูล OCR จาก Step 2: {len(ocr_data_from_step2)} รายการ")
            # แปลงข้อมูล OCR จาก Step 2 เป็นรูปแบบเดียวกับข้อมูลจาก Excel
            for ocr_item in ocr_data_from_step2:
                if ocr_item.get('success'):
                    # เพิ่มรายการหลัก (ภาษีมูลค่าเพิ่ม)
                    ocr_data.append({
                        'row': len(ocr_data) + 2,  # เริ่มจากแถวที่ 2
                        'company_name': ocr_item.get('company_name', ''),
                        'tax_id': ocr_item.get('tax_id', ''),
                        'branch': ocr_item.get('branch', ''),
                        'document_no': ocr_item.get('document_number', ''),
                        'date': convert_buddhist_to_christian_date(ocr_item.get('date', '')),
                        'amount_before_vat': ocr_item.get('amount_before_vat', 0),
                        'vat_amount': ocr_item.get('vat_amount', 0),
                        'total_amount': ocr_item.get('total_amount', 0),
                        'buyer_name': ocr_item.get('buyer_name', ''),
                        'buyer_tax_id': ocr_item.get('buyer_tax_id', ''),
                        'buyer_address': ocr_item.get('buyer_address', ''),
                        'document_type': ocr_item.get('document_type', ''),
                        'document_status': ocr_item.get('document_status', ''),
                        'old_filename': ocr_item.get('filename', ''),
                        'filename': ocr_item.get('filename', ''),  # เพิ่ม filename สำหรับแสดงในหน้าเว็บ
                        'reference_number': ocr_item.get('reference_number', ''),  # ดึง reference number จากชื่อไฟล์
                        'items': ocr_item.get('items', []),  # เพิ่มรายการสินค้า
                        'is_customs_duty': False  # ไม่ใช่ค่าอากรขาขเข้า
                    })
                    
                    # สำหรับกรมศุลกากร: ถ้ามีค่าอากรขาขเข้า ให้เพิ่มอีก 1 บรรทัด
                    if ocr_item.get('has_customs_duty', False) and ocr_item.get('customs_duty', 0) > 0:
                        customs_duty_amount = ocr_item.get('customs_duty', 0)
                        logger.info(f"📦 เพิ่มรายการค่าอากรขาขเข้า: {customs_duty_amount:,.2f} บาท")
                        ocr_data.append({
                            'row': len(ocr_data) + 2,
                            'company_name': ocr_item.get('company_name', ''),
                            'tax_id': ocr_item.get('tax_id', ''),
                            'branch': ocr_item.get('branch', ''),
                            'document_no': ocr_item.get('document_number', ''),
                            'date': convert_buddhist_to_christian_date(ocr_item.get('date', '')),
                            'amount_before_vat': customs_duty_amount,  # ค่าอากรขาขเข้า (ไม่มีภาษีมูลค่าเพิ่ม)
                            'vat_amount': 0,  # ไม่มีภาษีมูลค่าเพิ่ม
                            'total_amount': customs_duty_amount,  # ยอดรวม = ค่าอากรขาขเข้า
                            'buyer_name': ocr_item.get('buyer_name', ''),
                            'buyer_tax_id': ocr_item.get('buyer_tax_id', ''),
                            'buyer_address': ocr_item.get('buyer_address', ''),
                            'document_type': ocr_item.get('document_type', ''),
                            'document_status': ocr_item.get('document_status', ''),
                            'old_filename': ocr_item.get('filename', ''),
                            'reference_number': ocr_item.get('reference_number', ''),
                            'is_customs_duty': True  # เป็นค่าอากรขาขเข้า
                        })
        
        # ถ้าไม่มีข้อมูล OCR จาก Step 2 ให้อ่านจากไฟล์ Excel OCR
        if not ocr_data and ocr_excel_files:
            try:
                import openpyxl
                for excel_path in ocr_excel_files:
                    try:
                        wb = openpyxl.load_workbook(excel_path, data_only=True)
                        ws = wb.active
                        
                        # อ่าน header (แถวแรก)
                        headers = []
                        if ws.max_row > 0:
                            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
                        
                        # หาคอลัมน์ตามชื่อ header
                        col_map = {}
                        for idx, header in enumerate(headers, 1):
                            header_lower = header.lower().strip()
                            
                            # ชื่อบริษัท
                            if 'ชื่อบริษัท' in header_lower:
                                col_map['company_name'] = idx
                            # เลขประจำตัวผู้เสียภาษี
                            elif 'เลขประจำตัวผู้เสียภาษี' in header_lower and 'ผู้ซื้อ' not in header_lower:
                                col_map['tax_id'] = idx
                            # สาขา
                            elif 'สาขา' in header_lower and 'ผู้ซื้อ' not in header_lower:
                                col_map['branch'] = idx
                            # เลขที่เอกสาร
                            elif 'เลขที่เอกสาร' in header_lower or 'เลขที่ใบกำกับ' in header_lower:
                                col_map['document_no'] = idx
                            # วันที่
                            elif 'วันที่' in header_lower and 'ใบกำกับ' not in header_lower:
                                col_map['date'] = idx
                            # ยอดก่อนภาษีมูลค่าเพิ่ม
                            elif 'ยอดก่อนภาษี' in header_lower or 'ก่อนภาษี' in header_lower:
                                col_map['amount_before_vat'] = idx
                            # ยอดภาษีมูลค่าเพิ่ม (ต้องระวังไม่ให้ match กับยอดหลังบวก)
                            elif 'ยอดภาษีมูลค่าเพิ่ม' in header_lower and 'หลังบวก' not in header_lower:
                                if 'vat_amount' not in col_map:  # ใช้เฉพาะถ้ายังไม่มี
                                    col_map['vat_amount'] = idx
                            # ยอดหลังบวกภาษีมูลค่าเพิ่ม (ต้องมีคำว่า "หลังบวก")
                            elif 'ยอดหลังบวก' in header_lower and 'ภาษีมูลค่าเพิ่ม' in header_lower:
                                col_map['total_amount'] = idx
                            # ยอดรวม (fallback - ต้องตรวจสอบว่าไม่ใช่ยอดภาษี)
                            elif 'ยอดรวม' in header_lower and 'ภาษี' not in header_lower:
                                if 'total_amount' not in col_map:  # ใช้เฉพาะถ้ายังไม่มี
                                    col_map['total_amount'] = idx
                            # TotalAmount (fallback สำหรับภาษาอังกฤษ)
                            elif 'totalamount' in header_lower.replace(' ', ''):
                                if 'total_amount' not in col_map:  # ใช้เฉพาะถ้ายังไม่มี
                                    col_map['total_amount'] = idx
                            # ชื่อผู้ซื้อ
                            elif 'ชื่อผู้ซื้อ' in header_lower:
                                col_map['buyer_name'] = idx
                            # เลขประจำตัวผู้เสียภาษี - ผู้ซื้อ
                            elif 'เลขประจำตัวผู้เสียภาษี' in header_lower and 'ผู้ซื้อ' in header_lower:
                                col_map['buyer_tax_id'] = idx
                            # ที่อยู่ผู้ซื้อ
                            elif 'ที่อยู่ผู้ซื้อ' in header_lower:
                                col_map['buyer_address'] = idx
                            # ประเภทเอกสาร
                            elif 'ประเภทเอกสาร' in header_lower:
                                col_map['document_type'] = idx
                            # สถานะเอกสาร
                            elif 'สถานะเอกสาร' in header_lower:
                                col_map['document_status'] = idx
                            # รายการสินค้า (items) - อาจเป็น JSON string
                            elif 'รายการสินค้า' in header_lower or 'items' in header_lower or 'รายการ' in header_lower:
                                col_map['items'] = idx
                            # ชื่อไฟล์เก่า (old_filename)
                            elif 'ชื่อไฟล์เก่า' in header_lower or 'old_filename' in header_lower or 'old filename' in header_lower:
                                col_map['old_filename'] = idx
                        
                        # Debug: Log headers และ col_map
                        logger.info(f"📋 Headers จากไฟล์ OCR: {headers[:15]}...")  # แสดง 15 headers แรก
                        logger.info(f"📋 Column mapping: {col_map}")
                        
                        # อ่านข้อมูล (เริ่มจากแถวที่ 2)
                        for row_idx in range(2, ws.max_row + 1):
                            row_data = {
                                'row': row_idx,
                                'company_name': '',
                                'tax_id': '',
                                'branch': '',
                                'document_no': '',
                                'date': '',
                                'amount_before_vat': 0,
                                'vat_amount': 0,
                                'total_amount': 0,
                                'buyer_name': '',
                                'buyer_tax_id': '',
                                'buyer_address': '',
                                'document_type': '',
                                'document_status': '',
                                'old_filename': '',  # เพิ่ม old_filename
                                'filename': '',  # เพิ่ม filename สำหรับแสดงในหน้าเว็บ
                                'reference_number': None,  # เพิ่ม reference_number
                                'items': []  # เพิ่ม items
                            }
                            
                            # อ่านข้อมูลจากแต่ละคอลัมน์
                            for key, col_idx in col_map.items():
                                cell = ws.cell(row=row_idx, column=col_idx)
                                if cell.value is not None:
                                    if key in ['amount_before_vat', 'vat_amount', 'total_amount']:
                                        try:
                                            row_data[key] = float(cell.value) if isinstance(cell.value, (int, float)) else 0
                                        except (ValueError, TypeError):
                                            row_data[key] = 0
                                    elif key == 'items':
                                        # ถ้าเป็น items ให้พยายาม parse เป็น JSON
                                        try:
                                            import json
                                            items_str = str(cell.value).strip()
                                            if items_str:
                                                # ลอง parse เป็น JSON
                                                items_data = json.loads(items_str)
                                                if isinstance(items_data, list):
                                                    row_data['items'] = items_data
                                                else:
                                                    row_data['items'] = []
                                        except (json.JSONDecodeError, ValueError, TypeError):
                                            # ถ้า parse ไม่ได้ ให้เก็บเป็น empty list
                                            row_data['items'] = []
                                    elif key == 'old_filename':
                                        row_data[key] = str(cell.value).strip()
                                        # ถ้ายังไม่มี filename ให้ใช้ old_filename
                                        if not row_data.get('filename'):
                                            row_data['filename'] = row_data[key]
                                    else:
                                        row_data[key] = str(cell.value).strip()
                            
                            # ดึง reference number จากชื่อไฟล์เก่า (ถ้ามี)
                            if row_data.get('old_filename'):
                                import re
                                old_filename = row_data['old_filename']
                                # Pattern: วันที่_REF-NUMBER_...
                                # ตัวอย่าง: 16.10.2025_EXP-20251000004_530117_...
                                ref_patterns = [
                                    r'\d{2}\.\d{2}\.\d{4}_([A-Z]+-\d+)_',  # 16.10.2025_EXP-20251000004_
                                    r'^\d{2}\.\d{2}\.\d{4}_([A-Z]+-\d+)_',  # เริ่มต้นด้วยวันที่
                                    r'_([A-Z]+-\d+)_',  # รูปแบบ _EXP-20251000004_
                                    r'([A-Z]{2,}-\d{8,})',  # รูปแบบ EXP-20251000004 (fallback)
                                ]
                                
                                for pattern in ref_patterns:
                                    match = re.search(pattern, old_filename)
                                    if match:
                                        row_data['reference_number'] = match.group(1)
                                        logger.info(f"✅ ดึง reference number จากชื่อไฟล์เก่า: {row_data['reference_number']} (จาก {old_filename})")
                                        break
                            
                            # เพิ่มเฉพาะแถวที่มีข้อมูล
                            if row_data['document_no'] or row_data['total_amount'] > 0 or row_data['amount_before_vat'] > 0:
                                ocr_data.append(row_data)
                        
                        logger.info(f"📊 อ่านข้อมูล OCR ได้ {len(ocr_data)} รายการ")
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถอ่านไฟล์ OCR {excel_path}: {e}")
            except ImportError:
                logger.warning("⚠️ ไม่พบ openpyxl - ไม่สามารถอ่านข้อมูล Excel ได้")
        
        # 5. เปรียบเทียบข้อมูลตามคู่คอลัมน์ที่ระบุ
        comparisons = []
        
        # สร้าง dictionary สำหรับค้นหาข้อมูล OCR โดยใช้หลายวิธี
        ocr_dict_by_doc_no = {}  # ใช้เลขที่เอกสารเป็น key
        ocr_dict_by_reference = {}  # ใช้ reference number เป็น key (สำหรับหน้า auditcheck)
        ocr_list = []  # เก็บรายการทั้งหมดเพื่อค้นหาแบบ fuzzy match
        
        for ocr_item in ocr_data:
            doc_no = ocr_item.get('document_no', '').strip()
            if doc_no:
                # เก็บแบบปกติ
                ocr_dict_by_doc_no[doc_no] = ocr_item
                # เก็บแบบ normalize (ลบ space, -)
                doc_no_normalized = doc_no.replace(' ', '').replace('-', '').replace('_', '')
                if doc_no_normalized != doc_no:
                    ocr_dict_by_doc_no[doc_no_normalized] = ocr_item
            
            # เก็บ reference number สำหรับการเปรียบเทียบ (สำหรับหน้า auditcheck)
            reference_no = ocr_item.get('reference_number', '').strip() if ocr_item.get('reference_number') else None
            if reference_no:
                ocr_dict_by_reference[reference_no] = ocr_item
                logger.info(f"📋 เก็บ reference number: {reference_no} สำหรับการเปรียบเทียบ")
            
            ocr_list.append(ocr_item)
        
        # Debug: Log ข้อมูลที่อ่านได้
        logger.info(f"📊 จำนวนข้อมูลภาษีซื้อ: {len(purchase_tax_data)}")
        logger.info(f"📊 จำนวนข้อมูล OCR: {len(ocr_data)}")
        if purchase_tax_data:
            logger.info(f"📊 ตัวอย่างข้อมูลภาษีซื้อ: {purchase_tax_data[0]}")
        if ocr_data:
            logger.info(f"📊 ตัวอย่างข้อมูล OCR: {ocr_data[0]}")
        
        # เปรียบเทียบข้อมูลภาษีซื้อกับ OCR - ใช้เลขที่เอกสารอ้างอิงเป็นหลัก
        for idx, purchase_item in enumerate(purchase_tax_data, 1):
            invoice_no = purchase_item.get('invoice_no', '').strip()
            reference_no = purchase_item.get('reference_no', '').strip()  # เลขที่เอกสารอ้างอิง
            
            # ใช้ reference number เป็นหลักในการเทียบเท่านั้น
            ocr_item = None
            if reference_no:
                # ค้นหาโดยใช้ reference number จากชื่อไฟล์ OCR
                ocr_item = ocr_dict_by_reference.get(reference_no, None)
                if ocr_item:
                    logger.info(f"✅ พบ OCR โดยใช้ reference number: {reference_no} -> {ocr_item.get('document_no', '')}")
                else:
                    logger.info(f"⚠️ ไม่พบ OCR ที่มี reference number: {reference_no}")
            else:
                logger.info(f"⚠️ ไม่พบ reference number ในข้อมูลภาษีซื้อ: invoice_no={invoice_no}")
            
            if ocr_item:
                # เปรียบเทียบข้อมูลตามคู่คอลัมน์ที่ระบุ
                match_results = {
                    'reference_no_match': False,  # OCR reference number เทียบกับ ภาษีซื้อ เลขที่เอกสารอ้างอิง (สำหรับหน้า auditcheck)
                    'company_name_match': False,  # OCR ชื่อบริษัท เทียบกับ ภาษีซื้อ ผู้ติดต่อ
                    'tax_id_match': False,  # OCR เลขประจำตัวผู้เสียภาษี เทียบกับ ภาษีซื้อ เลขทะเบียนผู้เสียภาษี
                    'branch_match': False,  # OCR สาขา เทียบกับ ภาษีซื้อ สาขา/สำนักงานใหญ่
                    'document_no_match': False,  # OCR เลขที่เอกสาร เทียบกับ ภาษีซื้อ เลขที่ใบกำกับภาษี
                    'date_match': False,  # OCR วันที่ เทียบกับ ภาษีซื้อ วันที่ใบกำกับภาษี
                    'amount_before_vat_match': False,  # OCR ยอดก่อนภาษี เทียบกับ ภาษีซื้อ รายการภาษี 7%
                    'vat_amount_match': False,  # OCR ยอดภาษี เทียบกับ ภาษีซื้อ ภาษีมูลค่าเพิ่ม
                    'total_amount_match': False,  # OCR ยอดหลังบวกภาษี เทียบกับ ภาษีซื้อ มูลค่ารวมภาษี
                    'document_type_match': False  # OCR ประเภทเอกสาร เทียบกับ ภาษีซื้อ ประเภทใบกำกับ
                }
                
                # 0. เปรียบเทียบ reference number (สำหรับหน้า auditcheck)
                ocr_reference = ocr_item.get('reference_number', '').strip() if ocr_item.get('reference_number') else ''
                purchase_reference = reference_no
                if ocr_reference and purchase_reference:
                    match_results['reference_no_match'] = ocr_reference == purchase_reference
                    logger.info(f"🔍 เปรียบเทียบ reference number: OCR={ocr_reference} vs Purchase={purchase_reference} -> {match_results['reference_no_match']}")
                elif ocr_reference or purchase_reference:
                    # ถ้ามีแค่ฝั่งเดียว ให้ไม่ match
                    match_results['reference_no_match'] = False
                
                # 1. เปรียบเทียบชื่อบริษัท (OCR) กับ ผู้ติดต่อ (ภาษีซื้อ)
                ocr_company = ocr_item.get('company_name', '').strip()
                purchase_contact = purchase_item.get('contact', '').strip()
                
                # Debug logic สำหรับบริษัทที่ชื่อภาษาอังกฤษและไทยต่างกัน
                def normalize_company_names(ocr_comp, purchase_comp):
                    """Normalize company names สำหรับบริษัทที่ชื่อภาษาอังกฤษและไทยต่างกัน - ถือว่าเป็นบริษัทเดียวกัน"""
                    ocr_lower = ocr_comp.lower().strip()
                    purchase_lower = purchase_comp.lower().strip()
                    
                    # 1. Shopee: Shopee (Thailand) Co., Ltd. = บริษัท ช้อปปี้ (ประเทศไทย) จำกัด
                    shopee_names_ocr = ['shopee (thailand) co., ltd.', 'shopee (thailand) co.,ltd.', 'shopee (thailand) co ltd', 'shopee (thailand) ltd']
                    shopee_names_thai = ['บริษัท ช้อปปี้ (ประเทศไทย) จำกัด', 'บริษัทช้อปปี้(ประเทศไทย)จำกัด', 'ช้อปปี้ (ประเทศไทย) จำกัด']
                    
                    is_shopee_ocr = any(name in ocr_lower for name in shopee_names_ocr)
                    is_shopee_thai = any(name in purchase_lower for name in shopee_names_thai)
                    
                    if is_shopee_ocr and is_shopee_thai:
                        logger.info(f'🔧 [Company Debug] พบ Shopee: OCR="{ocr_comp}" = Purchase="{purchase_comp}" (ถือว่าเป็นบริษัทเดียวกัน)')
                        return True
                    
                    # 2. TikTok Shop: TikTok Shop (Thailand) Ltd. = บริษัท ติ๊กต๊อก ช็อป (ประเทศไทย) จำกัด
                    tiktok_names_ocr = ['tiktok shop (thailand) ltd.', 'tiktok shop (thailand) ltd', 'tiktok shop (thailand) co., ltd.', 'tiktok shop (thailand) co.,ltd.']
                    tiktok_names_thai = ['บริษัท ติ๊กต๊อก ช็อป (ประเทศไทย) จำกัด', 'บริษัทติ๊กต๊อกช็อป(ประเทศไทย)จำกัด', 'ติ๊กต๊อก ช็อป (ประเทศไทย) จำกัด', 'บริษัท ติ๊กต๊อก ช็อป (ประเทศไทย) จํากัด']
                    
                    is_tiktok_ocr = any(name in ocr_lower for name in tiktok_names_ocr)
                    is_tiktok_thai = any(name in purchase_lower for name in tiktok_names_thai)
                    
                    if is_tiktok_ocr and is_tiktok_thai:
                        logger.info(f'🔧 [Company Debug] พบ TikTok Shop: OCR="{ocr_comp}" = Purchase="{purchase_comp}" (ถือว่าเป็นบริษัทเดียวกัน)')
                        return True
                    
                    return False
                
                # ตรวจสอบบริษัทที่ชื่อภาษาอังกฤษและไทยต่างกันก่อน
                if ocr_company and purchase_contact:
                    company_match = normalize_company_names(ocr_company, purchase_contact)
                    if company_match:
                        match_results['company_name_match'] = True
                    else:
                        match_results['company_name_match'] = ocr_company.lower() == purchase_contact.lower()
                else:
                    match_results['company_name_match'] = False
                
                # 2. เปรียบเทียบเลขประจำตัวผู้เสียภาษี
                ocr_tax_id = ocr_item.get('tax_id', '').strip().replace('-', '').replace(' ', '')
                purchase_tax_id = purchase_item.get('tax_id', '').strip().replace('-', '').replace(' ', '')
                match_results['tax_id_match'] = ocr_tax_id == purchase_tax_id if ocr_tax_id and purchase_tax_id else False
                
                # 3. เปรียบเทียบสาขา (normalize: แยกตัวเลขออกจากข้อความ)
                ocr_branch = ocr_item.get('branch', '').strip()
                purchase_branch = purchase_item.get('branch', '').strip()
                
                def normalize_branch(branch_str):
                    """Normalize branch: แยกตัวเลขออกจากข้อความ เช่น 'HQ (00000)' -> '00000'"""
                    if not branch_str:
                        return ''
                    # ลบวงเล็บและข้อความรอบๆ แล้วหาเฉพาะตัวเลข
                    import re
                    # หาตัวเลขในสาขา (อาจมีหลายตัวเลข ให้เอาเฉพาะตัวเลขที่ยาวที่สุด)
                    numbers = re.findall(r'\d+', branch_str)
                    if numbers:
                        # เอาเฉพาะตัวเลขที่ยาวที่สุด (มักจะเป็นรหัสสาขา)
                        return max(numbers, key=len)
                    # ถ้าไม่มีตัวเลข ให้ใช้ค่าต้นฉบับ
                    return branch_str.strip()
                
                ocr_branch_normalized = normalize_branch(ocr_branch)
                purchase_branch_normalized = normalize_branch(purchase_branch)
                match_results['branch_match'] = ocr_branch_normalized == purchase_branch_normalized if ocr_branch_normalized and purchase_branch_normalized else False
                
                # Debug logging สำหรับสาขา
                if not match_results['branch_match'] and ocr_branch and purchase_branch:
                    logger.debug(f"🔍 [Branch Compare] OCR: '{ocr_branch}' -> '{ocr_branch_normalized}' | Purchase: '{purchase_branch}' -> '{purchase_branch_normalized}'")
                
                # 4. เปรียบเทียบเลขที่เอกสาร
                ocr_doc_no = ocr_item.get('document_no', '').strip()
                
                # Debug logic สำหรับ Shopee: ลบช่องว่างตรงกลางออกจากเลขที่เอกสาร
                # เช่น: "TRSPEMKP00-00000-25 1201-0017434" -> "TRSPEMKP00-00000-251201-0017434"
                def normalize_shopee_document_number(doc_no):
                    """Normalize เลขที่เอกสาร Shopee - ลบช่องว่างตรงกลาง"""
                    if not doc_no:
                        return doc_no
                    
                    # ตรวจสอบว่าเป็นรูปแบบ Shopee หรือไม่ (มี TRSPEMKP และมีช่องว่างในรูปแบบ -XX XXXX-)
                    if 'trspemkp' in doc_no.lower() and ' ' in doc_no:
                        # ลบช่องว่างตรงกลาง เช่น "25 1201" -> "251201"
                        # Pattern: -XX XXXX- -> -XXXXXX-
                        import re
                        normalized = re.sub(r'(\d+)\s+(\d+)', r'\1\2', doc_no)
                        if normalized != doc_no:
                            logger.info(f'🔧 [Shopee Debug] แปลงเลขที่เอกสาร: "{doc_no}" -> "{normalized}"')
                        return normalized
                    return doc_no
                
                # Normalize เลขที่เอกสาร OCR สำหรับ Shopee
                ocr_doc_no_normalized = normalize_shopee_document_number(ocr_doc_no)
                
                # เปรียบเทียบ (ใช้ normalized version)
                if ocr_doc_no_normalized and invoice_no:
                    match_results['document_no_match'] = ocr_doc_no_normalized == invoice_no
                    
                    # ถ้าไม่ match ให้ลองเปรียบเทียบกับ version เดิมด้วย (fallback)
                    if not match_results['document_no_match'] and ocr_doc_no:
                        match_results['document_no_match'] = ocr_doc_no == invoice_no
                else:
                    match_results['document_no_match'] = False
                
                # อัพเดท ocr_doc_no ใน ocr_item เพื่อให้ใช้ normalized version ต่อไป
                if ocr_doc_no_normalized != ocr_doc_no:
                    ocr_item['document_no'] = ocr_doc_no_normalized
                    logger.info(f'🔧 [Shopee Debug] อัพเดท document_no ใน ocr_item: "{ocr_doc_no}" -> "{ocr_doc_no_normalized}"')
                
                # 5. เปรียบเทียบวันที่ (แปลงปี พ.ศ. เป็น ค.ศ. ก่อนเปรียบเทียบ)
                ocr_date = ocr_item.get('date', '').strip()
                purchase_date = purchase_item.get('invoice_date', '').strip()
                # แปลงวันที่เป็นรูปแบบเดียวกันเพื่อเปรียบเทียบ (แปลงปี พ.ศ. เป็น ค.ศ.)
                ocr_date_converted = convert_buddhist_to_christian_date(ocr_date)
                purchase_date_converted = convert_buddhist_to_christian_date(purchase_date)
                match_results['date_match'] = ocr_date_converted == purchase_date_converted if ocr_date_converted and purchase_date_converted else False
                
                # 6. เปรียบเทียบยอดก่อนภาษี (OCR) กับ รายการภาษี 7% (ภาษีซื้อ)
                ocr_amount_before_vat = ocr_item.get('amount_before_vat', 0)
                purchase_tax_7 = purchase_item.get('tax_7', 0)
                match_results['amount_before_vat_match'] = abs(ocr_amount_before_vat - purchase_tax_7) < 0.01
                
                # 7. เปรียบเทียบยอดภาษี (OCR) กับ ภาษีมูลค่าเพิ่ม (ภาษีซื้อ)
                ocr_vat = ocr_item.get('vat_amount', 0)
                purchase_vat = purchase_item.get('vat', 0)
                match_results['vat_amount_match'] = abs(ocr_vat - purchase_vat) < 0.01
                
                # 8. เปรียบเทียบยอดหลังบวกภาษี (OCR) กับ มูลค่ารวมภาษี (ภาษีซื้อ)
                ocr_total = ocr_item.get('total_amount', 0)
                purchase_total = purchase_item.get('total_with_vat', 0)
                # ถ้าไม่มีมูลค่ารวมภาษี ให้คำนวณจากรายการภาษี 7% + ภาษีมูลค่าเพิ่ม
                if purchase_total == 0:
                    purchase_total = purchase_tax_7 + purchase_vat
                match_results['total_amount_match'] = abs(ocr_total - purchase_total) < 0.01
                
                # 9. เปรียบเทียบประเภทเอกสาร (OCR) กับ ประเภทใบกำกับ (ภาษีซื้อ)
                ocr_doc_type = ocr_item.get('document_type', '').strip()
                purchase_invoice_type = purchase_item.get('invoice_type', '').strip()
                
                def normalize_document_type(doc_type_str):
                    """Normalize document type: แปลงเป็นรูปแบบมาตรฐานสำหรับเปรียบเทียบ"""
                    if not doc_type_str:
                        return ''
                    
                    doc_type_lower = doc_type_str.lower().strip()
                    
                    # Mapping ระหว่างภาษาไทยและอังกฤษ
                    doc_type_mapping = {
                        # ภาษาไทย -> รูปแบบมาตรฐาน
                        'ใบกำกับภาษี': 'tax_invoice',
                        'ใบกำกับ': 'tax_invoice',
                        'ใบกำกับภาษีมูลค่าเพิ่ม': 'tax_invoice',
                        'ใบกำกับภาษีมูลค่าเพิ่ม (vat)': 'tax_invoice',
                        
                        # ภาษาอังกฤษ -> รูปแบบมาตรฐาน
                        'tax invoice': 'tax_invoice',
                        'receipt / tax invoice': 'tax_invoice',
                        'receipt/tax invoice': 'tax_invoice',
                        'invoice': 'tax_invoice',
                        'vat invoice': 'tax_invoice',
                        
                        # ใบเสร็จ
                        'ใบเสร็จ': 'receipt',
                        'receipt': 'receipt',
                        'ใบเสร็จรับเงิน': 'receipt',
                        
                        # ใบกำกับภาษี/ใบเสร็จ
                        'ใบกำกับภาษี/ใบเสร็จ': 'tax_invoice_receipt',
                        'tax invoice/receipt': 'tax_invoice_receipt',
                        'receipt / tax invoice': 'tax_invoice_receipt',  # อาจเป็นทั้งสอง
                    }
                    
                    # ตรวจสอบ mapping
                    for key, value in doc_type_mapping.items():
                        if key in doc_type_lower:
                            return value
                    
                    # ถ้าไม่เจอ mapping ให้ใช้ค่าต้นฉบับแปลงเป็นตัวพิมพ์เล็ก
                    return doc_type_lower
                
                ocr_doc_type_normalized = normalize_document_type(ocr_doc_type)
                purchase_invoice_type_normalized = normalize_document_type(purchase_invoice_type)
                
                # เปรียบเทียบ: ถ้าเป็น tax_invoice ทั้งคู่ หรือมี tax_invoice ใน OCR และ purchase_invoice_type เป็น "ใบกำกับภาษี" ก็ถือว่าตรงกัน
                if ocr_doc_type_normalized and purchase_invoice_type_normalized:
                    # กรณีที่ตรงกันตรงๆ
                    if ocr_doc_type_normalized == purchase_invoice_type_normalized:
                        match_results['document_type_match'] = True
                    # กรณีพิเศษ: "receipt / tax invoice" หรือ "tax invoice/receipt" ควรตรงกับ "ใบกำกับภาษี"
                    elif (ocr_doc_type_normalized == 'tax_invoice_receipt' and purchase_invoice_type_normalized == 'tax_invoice') or \
                         (ocr_doc_type_normalized == 'tax_invoice' and purchase_invoice_type_normalized == 'tax_invoice_receipt'):
                        match_results['document_type_match'] = True
                    # กรณีที่ OCR เป็น "receipt / tax invoice" และภาษีซื้อเป็น "ใบกำกับภาษี" (ซึ่ง normalize เป็น tax_invoice)
                    elif 'tax_invoice' in ocr_doc_type_normalized and purchase_invoice_type_normalized == 'tax_invoice':
                        match_results['document_type_match'] = True
                    elif ocr_doc_type_normalized == 'tax_invoice' and 'tax_invoice' in purchase_invoice_type_normalized:
                        match_results['document_type_match'] = True
                    else:
                        match_results['document_type_match'] = False
                else:
                    match_results['document_type_match'] = False
                
                # Debug logging สำหรับประเภทเอกสาร
                if not match_results['document_type_match'] and ocr_doc_type and purchase_invoice_type:
                    logger.debug(f"🔍 [DocType Compare] OCR: '{ocr_doc_type}' -> '{ocr_doc_type_normalized}' | Purchase: '{purchase_invoice_type}' -> '{purchase_invoice_type_normalized}'")
                elif match_results['document_type_match']:
                    logger.debug(f"✅ [DocType Match] OCR: '{ocr_doc_type}' -> '{ocr_doc_type_normalized}' | Purchase: '{purchase_invoice_type}' -> '{purchase_invoice_type_normalized}'")
                
                # นับจำนวน fields ที่ตรงกัน (ไม่รวม reference_no_match)
                match_fields_without_ref = {k: v for k, v in match_results.items() if k != 'reference_no_match'}
                matched_count = sum(1 for v in match_fields_without_ref.values() if v)
                total_count = len(match_fields_without_ref)
                
                # ตรวจสอบว่าข้อมูลทั้งหมดตรงกันหรือไม่ (ไม่รวม reference_no_match)
                all_fields_match = all(match_fields_without_ref.values())
                
                # สำหรับหน้า auditcheck: ถ้า reference number match แล้ว ให้ถือว่าตรงกัน (แม้ว่าฟิลด์อื่นจะไม่ match)
                # แต่สำหรับการแสดงผล debug ยังคงแสดงสถานะจริงตามฟิลด์ที่ตรงกัน
                reference_no_match = match_results.get('reference_no_match', False)
                if reference_no_match:
                    logger.info(f"✅ พบการ match โดยใช้ reference number: {reference_no}")
                    # ถ้า reference number match แต่ฟิลด์อื่นๆ ไม่ตรงกันทั้งหมด ให้ใช้ all_fields_match สำหรับการแสดงผล
                    # แต่ match_status จะถูกกำหนดตาม matched_count เพื่อแสดงสถานะจริง
                
                # กำหนดสถานะการ match:
                # - all_fields_match = True และ matched_count == total_count: ตรงกันทั้งหมด (สีเขียว)
                # - matched_count > 0 และ matched_count < total_count: ตรงกันบางส่วน (สีเหลือง)
                # - matched_count == 0: ไม่ตรงเลยสักอัน (สีแดง)
                # หมายเหตุ: แม้ว่า reference_no_match จะเป็น True แต่ถ้าฟิลด์อื่นๆ ไม่ตรงกันทั้งหมด ก็จะแสดงเป็น partial_match
                if all_fields_match and matched_count == total_count:
                    match_status = 'full_match'
                elif matched_count > 0:
                    match_status = 'partial_match'
                else:
                    match_status = 'no_match'
                
                # สร้างหมายเหตุอัตโนมัติที่อธิบายว่าไม่ตรงกันตรงไหน เพราะอะไร
                def generate_initial_note(match_results, purchase_data, ocr_data):
                    """สร้างหมายเหตุอัตโนมัติที่อธิบายว่าไม่ตรงกันตรงไหน เพราะอะไร"""
                    notes = []
                    
                    # ตรวจสอบแต่ละ field ที่ไม่ตรงกัน
                    if not match_results.get('company_name_match', True):
                        purchase_company = purchase_data.get('contact', '-')
                        ocr_company = ocr_data.get('company_name', '-')
                        notes.append(f"ชื่อบริษัทไม่ตรงกัน: ภาษีซื้อ='{purchase_company}' แต่ OCR='{ocr_company}'")
                    
                    if not match_results.get('tax_id_match', True):
                        purchase_tax_id = purchase_data.get('tax_id', '-')
                        ocr_tax_id = ocr_data.get('tax_id', '-')
                        notes.append(f"เลขประจำตัวผู้เสียภาษีไม่ตรงกัน: ภาษีซื้อ='{purchase_tax_id}' แต่ OCR='{ocr_tax_id}'")
                    
                    if not match_results.get('branch_match', True):
                        purchase_branch = purchase_data.get('branch', '-')
                        ocr_branch = ocr_data.get('branch', '-')
                        notes.append(f"สาขาไม่ตรงกัน: ภาษีซื้อ='{purchase_branch}' แต่ OCR='{ocr_branch}'")
                    
                    if not match_results.get('document_no_match', True):
                        purchase_doc_no = purchase_data.get('invoice_no', '-')
                        ocr_doc_no = ocr_data.get('document_no', '-')
                        notes.append(f"เลขที่ใบกำกับไม่ตรงกัน: ภาษีซื้อ='{purchase_doc_no}' แต่ OCR='{ocr_doc_no}'")
                    
                    if not match_results.get('date_match', True):
                        purchase_date = purchase_data.get('invoice_date', '-')
                        ocr_date = ocr_data.get('date', '-')
                        notes.append(f"วันที่ไม่ตรงกัน: ภาษีซื้อ='{purchase_date}' แต่ OCR='{ocr_date}'")
                    
                    if not match_results.get('amount_before_vat_match', True):
                        purchase_amount = purchase_data.get('tax_7', 0)
                        ocr_amount = ocr_data.get('amount_before_vat', 0)
                        notes.append(f"รายการ 7% ไม่ตรงกัน: ภาษีซื้อ={purchase_amount:,.2f} แต่ OCR={ocr_amount:,.2f} (ต่างกัน {abs(purchase_amount - ocr_amount):,.2f})")
                    
                    if not match_results.get('vat_amount_match', True):
                        purchase_vat = purchase_data.get('vat', 0)
                        ocr_vat = ocr_data.get('vat_amount', 0)
                        notes.append(f"ภาษีมูลค่าเพิ่มไม่ตรงกัน: ภาษีซื้อ={purchase_vat:,.2f} แต่ OCR={ocr_vat:,.2f} (ต่างกัน {abs(purchase_vat - ocr_vat):,.2f})")
                    
                    if not match_results.get('total_amount_match', True):
                        purchase_total = purchase_data.get('total_with_vat', 0)
                        ocr_total = ocr_data.get('total_amount', 0)
                        notes.append(f"มูลค่ารวมไม่ตรงกัน: ภาษีซื้อ={purchase_total:,.2f} แต่ OCR={ocr_total:,.2f} (ต่างกัน {abs(purchase_total - ocr_total):,.2f})")
                    
                    if not match_results.get('document_type_match', True):
                        purchase_doc_type = purchase_data.get('invoice_type', '-')
                        ocr_doc_type = ocr_data.get('document_type', '-')
                        notes.append(f"ประเภทเอกสารไม่ตรงกัน: ภาษีซื้อ='{purchase_doc_type}' แต่ OCR='{ocr_doc_type}'")
                    
                    # รวมหมายเหตุทั้งหมดด้วยเครื่องหมายขึ้นบรรทัดใหม่
                    if notes:
                        return '\n'.join(notes)
                    return ''
                
                initial_note = ''
                if match_status != 'full_match':
                    initial_note = generate_initial_note(match_results, purchase_item, ocr_item)
                
                comparisons.append({
                    'item': f'รายการที่ {idx}',
                    'purchaseTax': f'{purchase_total:,.2f}',
                    'ocrFile': f'{ocr_total:,.2f}',
                    'match': all_fields_match,
                    'match_status': match_status,  # เพิ่ม match_status สำหรับ debug
                    'matched_count': matched_count,  # จำนวน fields ที่ตรงกัน
                    'total_count': total_count,  # จำนวน fields ทั้งหมด
                    'invoice_no': invoice_no,
                    'match_details': match_results,
                    'purchase_data': purchase_item,
                    'ocr_data': ocr_item,
                    'initial_note': initial_note  # เพิ่มหมายเหตุเบื้องต้น
                })
            else:
                # ไม่พบข้อมูล OCR ที่ตรงกัน
                purchase_total = purchase_item.get('total_with_vat', 0)
                if purchase_total == 0:
                    purchase_total = purchase_item.get('tax_7', 0) + purchase_item.get('vat', 0)
                
                initial_note = f"ไม่พบข้อมูล OCR ที่ตรงกับเลขที่ใบกำกับ: {invoice_no}"
                
                comparisons.append({
                    'item': f'รายการที่ {idx}',
                    'purchaseTax': f'{purchase_total:,.2f}',
                    'ocrFile': '0.00',
                    'match': False,
                    'match_status': 'no_match',  # ไม่พบข้อมูล OCR
                    'matched_count': 0,
                    'total_count': 0,
                    'invoice_no': invoice_no,
                    'match_details': None,
                    'purchase_data': purchase_item,
                    'ocr_data': None,
                    'initial_note': initial_note  # เพิ่มหมายเหตุเบื้องต้น
                })
        
        # เพิ่มรายการ OCR ที่ไม่พบในภาษีซื้อ (ใช้ reference_no เป็นหลัก)
        purchase_reference_nos = {item.get('reference_no', '').strip() for item in purchase_tax_data if item.get('reference_no', '').strip()}
        matched_ocr_reference_nos = set()  # เก็บ reference_no ของ OCR ที่ถูก match แล้ว
        
        # เก็บ reference_no ของ OCR ที่ถูก match แล้วจาก comparisons
        for comp in comparisons:
            ocr_data_comp = comp.get('ocr_data')
            if ocr_data_comp:
                ocr_ref = ocr_data_comp.get('reference_number', '').strip() if ocr_data_comp.get('reference_number') else ''
                if ocr_ref:
                    matched_ocr_reference_nos.add(ocr_ref)
        
        for ocr_item in ocr_data:
            ocr_reference_no = ocr_item.get('reference_number', '').strip() if ocr_item.get('reference_number') else ''
            doc_no = ocr_item.get('document_no', '').strip()
            
            # ตรวจสอบว่า OCR item นี้ถูก match ไปแล้วหรือยัง (ใช้ reference_no เป็นหลัก)
            already_matched = False
            if ocr_reference_no:
                already_matched = ocr_reference_no in matched_ocr_reference_nos
            
            # ถ้ายังไม่ถูก match และมี reference_no ให้ตรวจสอบว่าไม่พบในภาษีซื้อ
            if not already_matched:
                if ocr_reference_no:
                    # ใช้ reference_no เป็นหลักในการตรวจสอบ
                    if ocr_reference_no not in purchase_reference_nos:
                        ocr_total = ocr_item.get('total_amount', 0)
                        initial_note = f"ไม่พบข้อมูลภาษีซื้อที่ตรงกับเลขที่เอกสารอ้างอิง: {ocr_reference_no}"
                        
                        comparisons.append({
                            'item': f'รายการที่ {len(comparisons) + 1}',
                            'purchaseTax': '0.00',
                            'ocrFile': f'{ocr_total:,.2f}',
                            'match': False,
                            'match_status': 'no_match',  # ไม่พบข้อมูลภาษีซื้อ
                            'matched_count': 0,
                            'total_count': 0,
                            'invoice_no': doc_no,
                            'match_details': None,
                            'purchase_data': None,
                            'ocr_data': ocr_item,
                            'initial_note': initial_note  # เพิ่มหมายเหตุเบื้องต้น
                        })
                elif doc_no:
                    # ถ้าไม่มี reference_no แต่มี doc_no ให้แสดง (กรณี fallback)
                    ocr_total = ocr_item.get('total_amount', 0)
                    initial_note = f"ไม่พบข้อมูลภาษีซื้อที่ตรงกับเลขที่ใบกำกับ: {doc_no} (ไม่มีเลขที่เอกสารอ้างอิง)"
                    
                    comparisons.append({
                        'item': f'รายการที่ {len(comparisons) + 1}',
                        'purchaseTax': '0.00',
                        'ocrFile': f'{ocr_total:,.2f}',
                        'match': False,
                        'match_status': 'no_match',  # ไม่พบข้อมูลภาษีซื้อ
                        'matched_count': 0,
                        'total_count': 0,
                        'invoice_no': doc_no,
                        'match_details': None,
                        'purchase_data': None,
                        'ocr_data': ocr_item,
                        'initial_note': initial_note  # เพิ่มหมายเหตุเบื้องต้น
                    })
        
        all_match = all(comp['match'] for comp in comparisons) if comparisons else False
        
        return jsonify({
            'success': True,
            'allMatch': all_match,
            'comparisons': comparisons,
            'purchaseTaxFileCount': len(purchase_tax_files),
            'ocrFileCount': len(ocr_excel_files),
            'purchaseTaxFiles': purchase_tax_files,
            'ocrFiles': ocr_excel_files,
            'purchaseTaxDataCount': len(purchase_tax_data),
            'ocrDataCount': len(ocr_data),
            'vatFolderPath': vat_folder_path,  # เพิ่ม path ของโฟลเดอร์ VAT
            'debug': {
                'purchase_tax_sample': purchase_tax_data[0] if purchase_tax_data else None,
                'ocr_sample': ocr_data[0] if ocr_data else None,
                'purchase_invoice_nos': [item.get('invoice_no', '') for item in purchase_tax_data[:5]],
                'ocr_document_nos': [item.get('document_no', '') for item in ocr_data[:5]]
            }
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการเปรียบเทียบข้อมูลภาษีซื้อและ OCR: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/find-pdf-by-reference', methods=['POST'])
def find_pdf_by_reference():
    """ค้นหาไฟล์ PDF ในโฟลเดอร์ VAT โดยใช้เลขที่เอกสารอ้างอิง"""
    try:
        data = request.json
        reference_no = data.get('referenceNo', '')
        tax_month = data.get('taxMonth', '')  # Format: YYYY-MM
        company = data.get('company', '')
        
        if not reference_no:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเลขที่เอกสารอ้างอิง'
            }), 400
        
        if not tax_month or not company:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเดือนภาษีและชื่อบริษัท'
            }), 400
        
        # แปลงเดือนจาก YYYY-MM
        year, month = tax_month.split('-')
        year_int = int(year)
        month_int = int(month)
        
        base_paths = [
            Path(f"V:/A.โฟร์เดอร์หลัก/{company}"),
            Path(f"V:/AA.โฟรเดอร์หลัก/{company}"),
            Path(f"V:/AAA.โฟรเดอร์หลัก/{company}")
        ]
        
        month_year_patterns = [
            f"{year_int}-{month_int:02d}",
            f"{year_int}-{month_int}",
            f"{month_int:02d}-{year_int}",
            f"{month_int}-{year_int}",
        ]
        
        # ค้นหาไฟล์ PDF ในโฟลเดอร์ VAT
        pdf_files = []
        for base_path in base_paths:
            if not base_path.exists():
                continue
            try:
                account_folder = base_path / "บัญชี"
                if not account_folder.exists():
                    continue
                expense_folder = account_folder / "002-รายจ่าย"
                if not expense_folder.exists():
                    continue
                pv_folder = expense_folder / "PV"
                if not pv_folder.exists():
                    continue
                
                # ค้นหาโฟลเดอร์เดือน-ปี
                month_year_folder = None
                for pattern in month_year_patterns:
                    potential_folder = pv_folder / pattern
                    if potential_folder.exists():
                        month_year_folder = potential_folder
                        break
                
                if not month_year_folder:
                    try:
                        for item in pv_folder.iterdir():
                            if item.is_dir():
                                folder_name = item.name
                                if (str(year_int) in folder_name or str(year_int + 543) in folder_name):
                                    for sub_item in item.iterdir():
                                        if sub_item.is_dir():
                                            for pattern in month_year_patterns:
                                                if pattern in sub_item.name or sub_item.name == pattern:
                                                    month_year_folder = sub_item
                                                    break
                                            if month_year_folder:
                                                break
                                if month_year_folder:
                                    break
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์เดือน-ปี: {e}")
                
                # ค้นหาไฟล์ PDF ในโฟลเดอร์ VAT
                if month_year_folder:
                    vat_folders = []
                    try:
                        for item in month_year_folder.iterdir():
                            if item.is_dir() and item.name.lower() == "vat":
                                vat_folders.append(item)
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์ VAT: {e}")
                    
                    if vat_folders:
                        for vat_folder in vat_folders:
                            try:
                                logger.info(f"🔍 กำลังค้นหาไฟล์ที่รองรับ OCR (PDF/JPG/PNG) ในโฟลเดอร์: {vat_folder} (รวมโฟลเดอร์ย่อย)")
                                logger.info(f"🔍 เลขอ้างอิงที่ค้นหา (original): '{reference_no}'")
                                
                                # Normalize เลขอ้างอิงสำหรับการค้นหา (ลบ space, dash, underscore)
                                reference_no_normalized = reference_no.replace(' ', '').replace('-', '').replace('_', '').upper()
                                logger.info(f"🔍 เลขอ้างอิงที่ normalize แล้ว: '{reference_no_normalized}'")
                                
                                # สร้างรูปแบบการค้นหาที่หลากหลาย
                                search_patterns = [
                                    reference_no,  # รูปแบบเดิม (เช่น EXP-20251000003)
                                    reference_no.replace('-', ''),  # ลบ dash (เช่น EXP20251000003)
                                    reference_no.replace('EXP-', ''),  # ลบ EXP- (เช่น 20251000003)
                                    reference_no_normalized,  # normalize เต็มรูปแบบ
                                ]
                                
                                # แสดง patterns ทั้งหมดที่จะค้นหา
                                logger.info(f"🔍 Search patterns: {search_patterns}")
                                
                                # รายการนามสกุลไฟล์ที่รองรับ (PDF, JPG, PNG)
                                supported_extensions = ['*.pdf', '*.jpg', '*.jpeg', '*.png', '*.PDF', '*.JPG', '*.JPEG', '*.PNG']
                                
                                # ค้นหาไฟล์ที่รองรับ OCR (รวมโฟลเดอร์ย่อยด้วย)
                                file_count = 0
                                all_files_in_folder = []
                                
                                for ext_pattern in supported_extensions:
                                    # ใช้ rglob() เพื่อค้นหาไฟล์ในโฟลเดอร์หลักและโฟลเดอร์ย่อยทั้งหมด
                                    for file_path in vat_folder.rglob(ext_pattern):
                                        if file_path.is_file():
                                            filename = file_path.name
                                            all_files_in_folder.append(filename)
                                            filename_normalized = filename.replace(' ', '').replace('-', '').replace('_', '').upper()
                                            
                                            file_count += 1
                                            
                                            # ตรวจสอบว่าชื่อไฟล์มีเลขที่เอกสารอ้างอิงหรือไม่ (ใช้หลายรูปแบบ)
                                            is_match = False
                                            matched_pattern = None
                                            
                                            for pattern in search_patterns:
                                                pattern_normalized = pattern.replace(' ', '').replace('-', '').replace('_', '').upper()
                                                
                                                # ลองหาทั้งใน filename เดิมและ normalized
                                                if pattern in filename:
                                                    is_match = True
                                                    matched_pattern = f"{pattern} (exact in filename)"
                                                    logger.info(f"✅ Match found (exact): '{pattern}' in '{filename}'")
                                                    break
                                                elif pattern_normalized in filename_normalized:
                                                    is_match = True
                                                    matched_pattern = f"{pattern} (normalized)"
                                                    logger.info(f"✅ Match found (normalized): '{pattern_normalized}' in normalized '{filename_normalized}'")
                                                    break
                                            
                                            if is_match:
                                                pdf_files.append({
                                                    'path': str(file_path),
                                                    'filename': filename,
                                                    'size': file_path.stat().st_size if file_path.exists() else 0,
                                                    'extension': file_path.suffix.lower()
                                                })
                                                logger.info(f"✅ พบไฟล์: {filename} ({file_path.suffix})")
                                                logger.info(f"✅ Match pattern: {matched_pattern}")
                                                logger.info(f"✅ Path: {file_path}")
                                
                                # แสดงรายชื่อไฟล์ทั้งหมดในโฟลเดอร์ (สำหรับ debug)
                                logger.info(f"📊 จำนวนไฟล์ที่รองรับ OCR ในโฟลเดอร์ (รวมโฟลเดอร์ย่อย): {file_count} ไฟล์")
                                if file_count > 0:
                                    logger.info(f"📋 รายชื่อไฟล์ทั้งหมด (5 ไฟล์แรก):")
                                    for i, fname in enumerate(all_files_in_folder[:5]):
                                        logger.info(f"   {i+1}. {fname}")
                                    if len(all_files_in_folder) > 5:
                                        logger.info(f"   ... และอีก {len(all_files_in_folder) - 5} ไฟล์")
                                
                                logger.info(f"📊 จำนวนไฟล์ที่ตรงกัน: {len(pdf_files)} ไฟล์")
                            except Exception as e:
                                logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์ PDF: {e}")
                                import traceback
                                logger.warning(f"⚠️ Traceback: {traceback.format_exc()}")
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์ PDF ใน {base_path}: {e}")
                continue
        
        if pdf_files:
            # ถ้าพบหลายไฟล์ ให้เลือกไฟล์แรก
            found_file = pdf_files[0]
            return jsonify({
                'success': True,
                'found': True,
                'pdfPath': found_file['path'],
                'filename': found_file['filename'],
                'fileExtension': found_file.get('extension', '.pdf'),
                'totalFound': len(pdf_files)
            }), 200
        else:
            return jsonify({
                'success': True,
                'found': False,
                'message': f'ไม่พบไฟล์ (PDF/JPG/PNG) สำหรับเลขที่อ้างอิง: {reference_no}'
            }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการค้นหาไฟล์ PDF: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/view-pdf/<path:filepath>')
def view_pdf(filepath):
    """แสดงไฟล์ PDF, JPG, PNG หรือไฟล์รูปภาพอื่นๆ"""
    try:
        # ตรวจสอบว่าไฟล์มีอยู่จริง
        file_path = Path(filepath)
        if not file_path.exists() or not file_path.is_file():
            return jsonify({
                'success': False,
                'error': 'ไม่พบไฟล์'
            }), 404
        
        # ตรวจสอบนามสกุลไฟล์เพื่อกำหนด mimetype
        file_extension = file_path.suffix.lower()
        mimetype_map = {
            '.pdf': 'application/pdf',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.gif': 'image/gif',
            '.bmp': 'image/bmp',
            '.tiff': 'image/tiff',
            '.tif': 'image/tiff'
        }
        
        mimetype = mimetype_map.get(file_extension, 'application/octet-stream')
        
        logger.info(f"📄 แสดงไฟล์: {file_path.name} (ประเภท: {mimetype})")
        
        # ส่งไฟล์
        return send_file(
            file_path,
            mimetype=mimetype,
            as_attachment=False,
            download_name=file_path.name
        )
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการแสดงไฟล์: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/move-document-to-review', methods=['POST'])
def move_document_to_review():
    """ย้ายเอกสาร PDF ไปยังโฟลเดอร์ 'ไฟล์ที่ต้องตรวจสอบ'"""
    try:
        data = request.json
        reference_no = data.get('referenceNo', '')
        tax_month = data.get('taxMonth', '')  # Format: YYYY-MM
        company = data.get('company', '')
        
        if not reference_no:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเลขที่เอกสารอ้างอิง'
            }), 400
        
        if not tax_month or not company:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเดือนภาษีและชื่อบริษัท'
            }), 400
        
        # แปลงเดือนจาก YYYY-MM
        year, month = tax_month.split('-')
        year_int = int(year)
        month_int = int(month)
        
        base_paths = [
            Path(f"V:/A.โฟร์เดอร์หลัก/{company}"),
            Path(f"V:/AA.โฟรเดอร์หลัก/{company}"),
            Path(f"V:/AAA.โฟรเดอร์หลัก/{company}")
        ]
        
        month_year_patterns = [
            f"{year_int}-{month_int:02d}",
            f"{year_int}-{month_int}",
            f"{month_int:02d}-{year_int}",
            f"{month_int}-{year_int}",
        ]
        
        # ค้นหาไฟล์ PDF ที่จะย้าย
        source_pdf_path = None
        source_vat_folder = None
        
        for base_path in base_paths:
            if not base_path.exists():
                continue
            try:
                account_folder = base_path / "บัญชี"
                if not account_folder.exists():
                    continue
                expense_folder = account_folder / "002-รายจ่าย"
                if not expense_folder.exists():
                    continue
                pv_folder = expense_folder / "PV"
                if not pv_folder.exists():
                    continue
                
                # ค้นหาโฟลเดอร์เดือน-ปี
                month_year_folder = None
                for pattern in month_year_patterns:
                    potential_folder = pv_folder / pattern
                    if potential_folder.exists():
                        month_year_folder = potential_folder
                        break
                
                if not month_year_folder:
                    try:
                        for item in pv_folder.iterdir():
                            if item.is_dir():
                                folder_name = item.name
                                if (str(year_int) in folder_name or str(year_int + 543) in folder_name):
                                    for sub_item in item.iterdir():
                                        if sub_item.is_dir():
                                            for pattern in month_year_patterns:
                                                if pattern in sub_item.name or sub_item.name == pattern:
                                                    month_year_folder = sub_item
                                                    break
                                            if month_year_folder:
                                                break
                                if month_year_folder:
                                    break
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์เดือน-ปี: {e}")
                
                # ค้นหาไฟล์ PDF ในโฟลเดอร์ VAT
                if month_year_folder:
                    vat_folders = []
                    try:
                        for item in month_year_folder.iterdir():
                            if item.is_dir() and item.name.lower() == "vat":
                                vat_folders.append(item)
                    except Exception as e:
                        logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์ VAT: {e}")
                    
                    if vat_folders:
                        for vat_folder in vat_folders:
                            try:
                                # Normalize เลขอ้างอิงสำหรับการค้นหา
                                search_patterns = [
                                    reference_no,
                                    reference_no.replace('-', ''),
                                    reference_no.replace('EXP-', ''),
                                    reference_no.replace(' ', '').replace('-', '').replace('_', '').upper(),
                                ]
                                
                                for pdf_file in vat_folder.glob("*.pdf"):
                                    if pdf_file.is_file():
                                        filename = pdf_file.name
                                        filename_normalized = filename.replace(' ', '').replace('-', '').replace('_', '').upper()
                                        
                                        # ตรวจสอบว่าชื่อไฟล์มีเลขอ้างอิงหรือไม่
                                        for pattern in search_patterns:
                                            pattern_normalized = pattern.replace(' ', '').replace('-', '').replace('_', '').upper()
                                            if pattern in filename or pattern_normalized in filename_normalized:
                                                source_pdf_path = pdf_file
                                                source_vat_folder = vat_folder
                                                logger.info(f"✅ พบไฟล์ที่จะย้าย: {filename}")
                                                break
                                        
                                        if source_pdf_path:
                                            break
                                
                                if source_pdf_path:
                                    break
                            except Exception as e:
                                logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์ PDF: {e}")
                
                if source_pdf_path:
                    break
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์ PDF ใน {base_path}: {e}")
                continue
        
        if not source_pdf_path or not source_vat_folder:
            return jsonify({
                'success': False,
                'message': f'ไม่พบไฟล์ PDF สำหรับเลขที่อ้างอิง: {reference_no}'
            }), 200
        
        # สร้างโฟลเดอร์ "ไฟล์ที่ต้องตรวจสอบ" ภายใต้โฟลเดอร์ VAT
        review_folder = source_vat_folder / "ไฟล์ที่ต้องตรวจสอบ"
        try:
            review_folder.mkdir(parents=True, exist_ok=True)
            logger.info(f"📁 สร้างโฟลเดอร์: {review_folder}")
        except Exception as e:
            logger.error(f"❌ ไม่สามารถสร้างโฟลเดอร์ได้: {e}")
            return jsonify({
                'success': False,
                'error': f'ไม่สามารถสร้างโฟลเดอร์ได้: {str(e)}'
            }), 500
        
        # ย้ายไฟล์
        import shutil
        destination_path = review_folder / source_pdf_path.name
        
        try:
            # ตรวจสอบว่าไฟล์ปลายทางมีอยู่แล้วหรือไม่
            if destination_path.exists():
                logger.warning(f"⚠️ ไฟล์ปลายทางมีอยู่แล้ว: {destination_path}")
                return jsonify({
                    'success': False,
                    'error': f'ไฟล์ {source_pdf_path.name} มีอยู่ในโฟลเดอร์ "ไฟล์ที่ต้องตรวจสอบ" แล้ว'
                }), 200
            
            # ย้ายไฟล์
            shutil.move(str(source_pdf_path), str(destination_path))
            logger.info(f"✅ ย้ายไฟล์สำเร็จ: {source_pdf_path.name}")
            logger.info(f"✅ จาก: {source_pdf_path}")
            logger.info(f"✅ ไปยัง: {destination_path}")
            
            return jsonify({
                'success': True,
                'message': f'ย้ายเอกสารสำเร็จ',
                'filename': source_pdf_path.name,
                'source': str(source_pdf_path),
                'destination': str(destination_path)
            }), 200
            
        except Exception as e:
            logger.error(f"❌ ไม่สามารถย้ายไฟล์ได้: {e}", exc_info=True)
            return jsonify({
                'success': False,
                'error': f'ไม่สามารถย้ายไฟล์ได้: {str(e)}'
            }), 500
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการย้ายเอกสาร: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/move-all-mismatched-documents', methods=['POST'])
def move_all_mismatched_documents():
    """ย้ายเอกสาร PDF ทั้งหมดที่ไม่ตรงกันไปยังโฟลเดอร์ 'ไฟล์ที่ต้องตรวจสอบ'"""
    try:
        data = request.json
        reference_nos = data.get('referenceNos', [])  # Array ของ reference numbers
        tax_month = data.get('taxMonth', '')  # Format: YYYY-MM
        company = data.get('company', '')
        vat_folder_path = data.get('vatFolderPath', '')  # path ของโฟลเดอร์ VAT
        
        if not reference_nos or len(reference_nos) == 0:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเลขที่เอกสารอ้างอิง'
            }), 400
        
        if not tax_month or not company:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเดือนภาษีและชื่อบริษัท'
            }), 400
        
        logger.info(f"📦 กำลังย้ายเอกสาร {len(reference_nos)} รายการที่ไม่ตรงกัน...")
        
        # ถ้ามี vat_folder_path ให้ใช้เลย ไม่ต้องค้นหา
        if vat_folder_path:
            source_vat_folder = Path(vat_folder_path)
            if not source_vat_folder.exists():
                return jsonify({
                    'success': False,
                    'error': f'โฟลเดอร์ VAT ไม่มีอยู่: {vat_folder_path}'
                }), 400
        else:
            # ค้นหาโฟลเดอร์ VAT (ใช้โค้ดเดียวกับ move-document-to-review)
            year, month = tax_month.split('-')
            year_int = int(year)
            month_int = int(month)
            thai_year = year_int + 543
            thai_months = ['', 'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน',
                          'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม']
            month_name = thai_months[month_int]
            
            month_year_patterns = [
                f"{year_int}-{month_int:02d}",
                f"{month_name} {thai_year}",
                f"{month_name}{thai_year}",
                f"{month_int:02d}-{year_int}"
            ]
            
            base_paths = [
                Path(f"V:/A.โฟร์เดอร์หลัก/{company}"),
                Path(f"V:/AA.โฟรเดอร์หลัก/{company}"),
                Path(f"V:/AAA.โฟรเดอร์หลัก/{company}")
            ]
            
            source_vat_folder = None
            
            for base_path in base_paths:
                if not base_path.exists():
                    continue
                try:
                    account_folder = base_path / "บัญชี"
                    if not account_folder.exists():
                        continue
                    expense_folder = account_folder / "002-รายจ่าย"
                    if not expense_folder.exists():
                        continue
                    pv_folder = expense_folder / "PV"
                    if not pv_folder.exists():
                        continue
                    
                    # ค้นหาโฟลเดอร์เดือน-ปี
                    month_year_folder = None
                    for pattern in month_year_patterns:
                        potential_folder = pv_folder / pattern
                        if potential_folder.exists():
                            month_year_folder = potential_folder
                            break
                    
                    if not month_year_folder:
                        try:
                            for item in pv_folder.iterdir():
                                if item.is_dir():
                                    folder_name = item.name
                                    if (str(year_int) in folder_name or str(thai_year) in folder_name):
                                        for sub_item in item.iterdir():
                                            if sub_item.is_dir():
                                                for pattern in month_year_patterns:
                                                    if pattern in sub_item.name or sub_item.name == pattern:
                                                        month_year_folder = sub_item
                                                        break
                                                if month_year_folder:
                                                    break
                                    if month_year_folder:
                                        break
                        except Exception as e:
                            logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์เดือน-ปี: {e}")
                    
                    # ค้นหาโฟลเดอร์ VAT
                    if month_year_folder:
                        try:
                            for item in month_year_folder.iterdir():
                                if item.is_dir() and item.name.lower() == "vat":
                                    source_vat_folder = item
                                    logger.info(f"✅ พบโฟลเดอร์ VAT: {source_vat_folder}")
                                    break
                        except Exception as e:
                            logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์ VAT: {e}")
                    
                    if source_vat_folder:
                        break
                except Exception as e:
                    logger.warning(f"⚠️ ไม่สามารถค้นหาโฟลเดอร์ VAT ใน {base_path}: {e}")
                    continue
            
            if not source_vat_folder:
                return jsonify({
                    'success': False,
                    'error': 'ไม่พบโฟลเดอร์ VAT'
                }), 400
        
        # สร้างโฟลเดอร์ "ไฟล์ที่ต้องตรวจสอบ" ภายใต้โฟลเดอร์ VAT
        review_folder = source_vat_folder / "ไฟล์ที่ต้องตรวจสอบ"
        try:
            review_folder.mkdir(parents=True, exist_ok=True)
            logger.info(f"📁 สร้างโฟลเดอร์: {review_folder}")
        except Exception as e:
            logger.error(f"❌ ไม่สามารถสร้างโฟลเดอร์ได้: {e}")
            return jsonify({
                'success': False,
                'error': f'ไม่สามารถสร้างโฟลเดอร์ได้: {str(e)}'
            }), 500
        
        # ย้ายไฟล์ทั้งหมด
        import shutil
        results = {
            'success': [],
            'failed': [],
            'not_found': []
        }
        
        for reference_no in reference_nos:
            if not reference_no:
                continue
            
            source_pdf_path = None
            
            # Normalize เลขอ้างอิงสำหรับการค้นหา
            reference_no_normalized = reference_no.replace(' ', '').replace('-', '').replace('_', '').upper()
            search_patterns = [
                reference_no,
                reference_no.replace('-', ''),
                reference_no.replace('EXP-', ''),
                reference_no_normalized,
            ]
            
            # ค้นหาไฟล์ PDF
            try:
                for pdf_file in source_vat_folder.glob("*.pdf"):
                    if pdf_file.is_file():
                        filename = pdf_file.name
                        filename_normalized = filename.replace(' ', '').replace('-', '').replace('_', '').upper()
                        
                        # ตรวจสอบว่าชื่อไฟล์มีเลขอ้างอิงหรือไม่
                        for pattern in search_patterns:
                            pattern_normalized = pattern.replace(' ', '').replace('-', '').replace('_', '').upper()
                            if pattern in filename or pattern_normalized in filename_normalized:
                                source_pdf_path = pdf_file
                                break
                        
                        if source_pdf_path:
                            break
            except Exception as e:
                logger.warning(f"⚠️ ไม่สามารถค้นหาไฟล์ PDF สำหรับ {reference_no}: {e}")
            
            if not source_pdf_path:
                results['not_found'].append({
                    'reference_no': reference_no,
                    'message': 'ไม่พบไฟล์ PDF'
                })
                continue
            
            # ย้ายไฟล์
            destination_path = review_folder / source_pdf_path.name
            
            try:
                # ตรวจสอบว่าไฟล์ปลายทางมีอยู่แล้วหรือไม่
                if destination_path.exists():
                    results['failed'].append({
                        'reference_no': reference_no,
                        'filename': source_pdf_path.name,
                        'message': f'ไฟล์ {source_pdf_path.name} มีอยู่ในโฟลเดอร์ "ไฟล์ที่ต้องตรวจสอบ" แล้ว'
                    })
                    continue
                
                # ย้ายไฟล์
                shutil.move(str(source_pdf_path), str(destination_path))
                logger.info(f"✅ ย้ายไฟล์สำเร็จ: {source_pdf_path.name}")
                
                results['success'].append({
                    'reference_no': reference_no,
                    'filename': source_pdf_path.name,
                    'source': str(source_pdf_path),
                    'destination': str(destination_path)
                })
            except Exception as e:
                logger.error(f"❌ ไม่สามารถย้ายไฟล์ {source_pdf_path.name} ได้: {e}")
                results['failed'].append({
                    'reference_no': reference_no,
                    'filename': source_pdf_path.name,
                    'message': str(e)
                })
        
        # สรุปผล
        total_count = len(reference_nos)
        success_count = len(results['success'])
        failed_count = len(results['failed'])
        not_found_count = len(results['not_found'])
        
        logger.info(f"📊 สรุปผลการย้าย: สำเร็จ {success_count}/{total_count}, ล้มเหลว {failed_count}, ไม่พบ {not_found_count}")
        
        return jsonify({
            'success': True,
            'message': f'ย้ายเอกสารเสร็จสิ้น: สำเร็จ {success_count}/{total_count} รายการ',
            'total': total_count,
            'success_count': success_count,
            'failed_count': failed_count,
            'not_found_count': not_found_count,
            'results': results
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการย้ายเอกสารทั้งหมด: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/export-excel', methods=['POST'])
def export_audit_report_to_excel():
    """ส่งออกรายงานตรวจภาษีเป็น Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from datetime import datetime
        
        data = request.json
        
        # ตรวจสอบว่า data ไม่เป็น None
        if data is None:
            logger.error(f"❌ request.json คืนค่า None")
            return jsonify({
                'success': False,
                'error': 'ไม่ได้รับข้อมูลจากหน้าเว็บ (request.json is None)'
            }), 400
        
        # ตรวจสอบว่า data เป็น dict หรือไม่
        if not isinstance(data, dict):
            logger.error(f"❌ request.json ไม่ใช่ dict (type: {type(data)})")
            return jsonify({
                'success': False,
                'error': 'รูปแบบข้อมูลไม่ถูกต้อง (request.json is not a dict)'
            }), 400
        
        tax_month = data.get('taxMonth', '')  # Format: YYYY-MM
        company = data.get('company', '')
        notes = data.get('notes', {})  # หมายเหตุจากหน้าเว็บ
        vat_folder_path = data.get('vatFolderPath', '')  # path ของโฟลเดอร์ VAT จาก Step 5
        ocr_data_from_step2 = data.get('ocrDataFromStep2', [])  # ข้อมูล OCR จาก Step 2
        invalid_documents = data.get('invalidDocuments', {})  # สถานะเอกสารใช้ไม่ได้ (key = index, value = true/false)
        approvals = data.get('approvals', {})  # ข้อมูลการอนุมัติฟิลด์ (key = "index-fieldKey", value = true/false)
        
        # ตรวจสอบว่า notes, invalid_documents, approvals เป็น dict หรือไม่
        if not isinstance(notes, dict):
            notes = {}
        if not isinstance(invalid_documents, dict):
            invalid_documents = {}
        if not isinstance(approvals, dict):
            approvals = {}
        if not isinstance(ocr_data_from_step2, list):
            ocr_data_from_step2 = []
        
        if not tax_month or not company:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุเดือนภาษีและชื่อบริษัท'
            }), 400
        
        logger.info(f"📊 กำลังสร้างรายงาน Excel สำหรับ {company} เดือน {tax_month}")
        if vat_folder_path:
            logger.info(f"📁 โฟลเดอร์ VAT ที่ได้รับ: {vat_folder_path}")
        if ocr_data_from_step2:
            logger.info(f"📊 ข้อมูล OCR จาก Step 2: {len(ocr_data_from_step2)} รายการ")
        
        # แปลงเดือนจาก YYYY-MM
        year, month = tax_month.split('-')
        year_int = int(year)
        month_int = int(month)
        
        # ====== ดึงข้อมูลจริงจากระบบ (ใช้โค้ดเดียวกับ compare-purchase-tax-ocr) ======
        # เรียก internal function เพื่อดึงข้อมูล
        request_body = {'taxMonth': tax_month, 'company': company}
        if ocr_data_from_step2:
            request_body['ocrDataFromStep2'] = ocr_data_from_step2
        
        with app.test_client() as client:
            response = client.post('/api/auditcheck/compare-purchase-tax-ocr',
                                 json=request_body,
                                 content_type='application/json')
            
            if response.status_code != 200:
                logger.error(f"❌ ไม่สามารถดึงข้อมูลได้: {response.status_code}")
                return jsonify({
                    'success': False,
                    'error': 'ไม่สามารถดึงข้อมูลจากระบบได้'
                }), 500
            
            comparison_data = response.get_json()
            
            # ตรวจสอบว่า comparison_data ไม่เป็น None หรือไม่ใช่ dict
            if comparison_data is None:
                logger.error(f"❌ ไม่สามารถดึงข้อมูลได้: response.get_json() คืนค่า None")
                return jsonify({
                    'success': False,
                    'error': 'ไม่สามารถดึงข้อมูลจากระบบได้ (response is None)'
                }), 500
            
            if not isinstance(comparison_data, dict):
                logger.error(f"❌ ข้อมูลไม่ถูกต้อง: comparison_data ไม่ใช่ dict (type: {type(comparison_data)})")
                return jsonify({
                    'success': False,
                    'error': 'ไม่สามารถดึงข้อมูลจากระบบได้ (response format is invalid)'
                }), 500
            
            if not comparison_data.get('success'):
                logger.error(f"❌ ข้อมูลไม่ถูกต้อง: comparison_data.get('success') = {comparison_data.get('success')}")
                return jsonify({
                    'success': False,
                    'error': 'ไม่พบข้อมูลสำหรับสร้างรายงาน'
                }), 500
            
            comparisons = comparison_data.get('comparisons', [])
            
            if not comparisons:
                logger.warning(f"⚠️ ไม่มีข้อมูลเปรียบเทียบ")
                return jsonify({
                    'success': False,
                    'error': 'ไม่มีข้อมูลสำหรับสร้างรายงาน'
                }), 400
        
        logger.info(f"✅ ดึงข้อมูลได้ {len(comparisons)} รายการ")
        
        # สร้าง Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "รายงานสรุปการตรวจสอบ"
        
        # กำหนดสไตล์
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(name='TH Sarabun New', size=12, bold=True, color="FFFFFF")
        
        subheader_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        subheader_font = Font(name='TH Sarabun New', size=12, bold=True, color="FFFFFF")
        
        # ฟอนต์สำหรับข้อมูลในตาราง (ใช้ขนาดเดียวกันทุกเซลล์)
        data_font = Font(name='TH Sarabun New', size=11, bold=False, italic=False)
        
        match_fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")  # เขียว
        partial_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")  # เหลือง
        mismatch_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # แดง
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ส่วนหัวรายงาน
        ws.merge_cells('A1:O1')
        ws['A1'] = f"รายงานสรุปการตรวจสอบภาษีซื้อ"
        ws['A1'].font = Font(name='TH Sarabun New', size=18, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:O2')
        ws['A2'] = f"บริษัท: {company}"
        ws['A2'].font = Font(name='TH Sarabun New', size=14, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:O3')
        month_name_th = ['มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน',
                         'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม']
        ws['A3'] = f"เดือนภาษี: {month_name_th[month_int-1]} {year_int+543}"
        ws['A3'].font = Font(name='TH Sarabun New', size=12)
        ws['A3'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A4:O4')
        ws['A4'] = f"วันที่สร้างรายงาน: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A4'].font = Font(name='TH Sarabun New', size=11)
        ws['A4'].alignment = Alignment(horizontal='center')
        
        # Header ตาราง (เพิ่มคอลัมน์ชื่อไฟล์ OCR, รายการยกเว้นภาษี, รายการภาษี 0%)
        headers = [
            'ลำดับ', 'เลขที่อ้างอิง', 'ชื่อบริษัท', 'เลขที่ใบกำกับ', 'วันที่',
            'เลขทะเบียนผู้เสียภาษี', 'สาขา', 'รายการ 7%', 'ภาษี 7%',
            'มูลค่ารวม', 'รายการยกเว้นภาษี', 'รายการภาษี 0%', 'ชื่อไฟล์ OCR', 'สถานะ', 'หมายเหตุ'
        ]
        
        row = 6
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # ปรับความกว้างคอลัมน์ (เพิ่มคอลัมน์ชื่อไฟล์ OCR, รายการยกเว้นภาษี, รายการภาษี 0%)
        column_widths = [8, 20, 30, 15, 12, 18, 10, 15, 12, 15, 18, 18, 35, 15, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # เพิ่มข้อมูล
        row = 7
        item_no = 1
        
        # สีสำหรับไฮไลท์
        red_fill = PatternFill(start_color="FB1919", end_color="FB1919", fill_type="solid")  # สีแดง (#FB1919)
        green_fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")  # สีเขียวอ่อน
        
        for idx, comp in enumerate(comparisons):
            # ตรวจสอบว่า comp ไม่เป็น None และเป็น dict
            if not comp or not isinstance(comp, dict):
                logger.warning(f"⚠️ ข้าม comparison ที่ index {idx}: comp ไม่ถูกต้อง (type: {type(comp)})")
                continue
            
            purchase_data = comp.get('purchase_data', {})
            ocr_data = comp.get('ocr_data', {})
            match_status = comp.get('match_status', 'no_match')
            match_details = comp.get('match_details', {})  # ใช้ match_details แทน field_matches
            
            # ตรวจสอบว่า purchase_data และ ocr_data เป็น dict หรือไม่
            if not isinstance(purchase_data, dict):
                purchase_data = {}
            if not isinstance(ocr_data, dict):
                ocr_data = {}
            if not isinstance(match_details, dict):
                match_details = {}
            
            # ตรวจสอบว่าเอกสารนี้ใช้ไม่ได้หรือไม่
            is_invalid = invalid_documents.get(str(idx), False) if isinstance(invalid_documents, dict) else False
            
            # สร้าง match_details ที่ปรับตามการอนุมัติ (ถ้าฟิลด์ถูกอนุมัติแล้ว ให้ถือว่า match)
            adjusted_match_details = match_details.copy() if match_details else {}
            for field_key in ['company_name_match', 'tax_id_match', 'branch_match', 'document_no_match', 
                            'date_match', 'amount_before_vat_match', 'vat_amount_match', 'total_amount_match', 
                            'document_type_match', 'reference_no_match']:
                approval_key = f"{idx}-{field_key}"
                if isinstance(approvals, dict) and approvals.get(approval_key, False):
                    # ถ้าฟิลด์นี้ถูกอนุมัติแล้ว ให้ถือว่า match
                    adjusted_match_details[field_key] = True
            
            # อัปเดต match_status ตาม adjusted_match_details
            if adjusted_match_details:
                matched_count = sum(1 for v in adjusted_match_details.values() if v)
                total_count = len(adjusted_match_details)
                if matched_count == total_count:
                    match_status = 'full_match'
                elif matched_count > 0:
                    match_status = 'partial_match'
                else:
                    match_status = 'no_match'
            
            # ใช้ adjusted_match_details แทน match_details เดิม
            match_details = adjusted_match_details
            
            # ข้อมูลจากภาษีซื้อ
            if purchase_data:
                # กำหนดค่าแต่ละเซลล์และตั้งค่า font ให้เท่ากันทุกเซลล์
                ws.cell(row=row, column=1, value=item_no).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=1).font = data_font
                ws.cell(row=row, column=2, value=purchase_data.get('reference_no', '-')).font = data_font
                ws.cell(row=row, column=3, value=purchase_data.get('contact', '-')).font = data_font
                ws.cell(row=row, column=4, value=purchase_data.get('invoice_no', '-')).font = data_font
                ws.cell(row=row, column=5, value=purchase_data.get('invoice_date', '-')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=5).font = data_font
                ws.cell(row=row, column=6, value=purchase_data.get('tax_id', '-')).font = data_font
                ws.cell(row=row, column=7, value=purchase_data.get('branch', '-')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=7).font = data_font
                ws.cell(row=row, column=8, value=purchase_data.get('tax_7', 0)).number_format = '#,##0.00'
                ws.cell(row=row, column=8).font = data_font
                ws.cell(row=row, column=9, value=purchase_data.get('vat', 0)).number_format = '#,##0.00'
                ws.cell(row=row, column=9).font = data_font
                ws.cell(row=row, column=10, value=purchase_data.get('total_with_vat', 0)).number_format = '#,##0.00'
                ws.cell(row=row, column=10).font = data_font
                
                # รายการยกเว้นภาษี (คอลัมน์ 11)
                exempt_amount = purchase_data.get('exempt', 0) or purchase_data.get('tax_exempt', 0) or 0
                ws.cell(row=row, column=11, value=exempt_amount).number_format = '#,##0.00'
                ws.cell(row=row, column=11).font = data_font
                
                # รายการภาษี 0% (คอลัมน์ 12)
                tax_0_amount = purchase_data.get('tax_0', 0) or purchase_data.get('tax_zero', 0) or 0
                ws.cell(row=row, column=12, value=tax_0_amount).number_format = '#,##0.00'
                ws.cell(row=row, column=12).font = data_font
                
                # ชื่อไฟล์ OCR (คอลัมน์ 13)
                ocr_filename = ocr_data.get('filename') or ocr_data.get('old_filename') or '-'
                ws.cell(row=row, column=13, value=ocr_filename).font = data_font
                
                # สถานะ (คอลัมน์ 14)
                status_text = ''
                status_fill = None
                if match_status == 'full_match':
                    status_text = 'ตรงกันทั้งหมด'
                    status_fill = match_fill
                elif match_status == 'partial_match':
                    status_text = 'ตรงกันบางส่วน'
                    status_fill = partial_fill
                else:
                    status_text = 'ไม่ตรงกัน'
                    status_fill = mismatch_fill
                
                status_cell = ws.cell(row=row, column=14, value=status_text)
                status_cell.alignment = Alignment(horizontal='center')
                status_cell.font = data_font
                if status_fill:
                    status_cell.fill = status_fill
                
                # หมายเหตุ (คอลัมน์ 15) - ใช้ initial_note จาก backend ถ้า notes จากหน้าเว็บว่างเปล่า
                # ใช้ item_no - 1 เป็น key เพื่อให้ตรงกับ index ที่เริ่มจาก 0 ในหน้าเว็บ
                # เพราะหน้าเว็บใช้ index (0, 1, 2, ...) เป็น key ใน comparisonNotes
                item_key = str(item_no - 1)
                note_text = ''
                if isinstance(notes, dict):
                    note_text = notes.get(item_key, '')
                # ถ้าไม่มี notes จากหน้าเว็บ ให้ใช้ initial_note จาก backend
                if not note_text:
                    note_text = comp.get('initial_note', '') if isinstance(comp, dict) else ''
                ws.cell(row=row, column=15, value=note_text).font = data_font
                
                # ตรวจสอบว่าเอกสารใช้ไม่ได้หรือไม่ - ถ้าใช่ให้ไฮไลท์สีแดงทั้งแถบ (ยกเว้นช่องหมายเหตุ)
                if is_invalid:
                    for col in range(1, 16):
                        if col != 15:  # ไม่ใส่สีในช่องหมายเหตุ (column 15)
                            cell = ws.cell(row=row, column=col)
                            cell.fill = red_fill
                            cell.border = border
                            cell.font = data_font
                        else:
                            # ช่องหมายเหตุไม่ใส่สี
                            cell = ws.cell(row=row, column=col)
                            cell.border = border
                            cell.font = data_font
                # ถ้า full_match และไม่ใช่เอกสารใช้ไม่ได้ → ไฮไลท์เขียวทั้งแถว (ยกเว้นช่องหมายเหตุ)
                elif match_status == 'full_match':
                    for col in range(1, 16):
                        if col != 15:  # ไม่ใส่สีในช่องหมายเหตุ (column 15)
                            cell = ws.cell(row=row, column=col)
                            cell.fill = green_fill
                            cell.border = border
                            cell.font = data_font
                        else:
                            # ช่องหมายเหตุไม่ใส่สี
                            cell = ws.cell(row=row, column=col)
                            cell.border = border
                            cell.font = data_font
                else:
                    # ขีดสีแดงที่ช่องข้อมูลที่ไม่ตรงกัน (เฉพาะช่องที่ไม่ตรงกัน) - แต่ถ้าเอกสารใช้ไม่ได้แล้วจะไม่ทำเพราะไฮไลท์ทั้งแถบแล้ว
                    if not is_invalid and ocr_data:
                        if match_details:
                            # Column 3: ชื่อบริษัท
                            if not match_details.get('company_name_match', True):
                                ws.cell(row=row, column=3).fill = red_fill
                            
                            # Column 4: เลขที่ใบกำกับ
                            if not match_details.get('document_no_match', True):
                                ws.cell(row=row, column=4).fill = red_fill
                            
                            # Column 5: วันที่
                            if not match_details.get('date_match', True):
                                ws.cell(row=row, column=5).fill = red_fill
                            
                            # Column 6: เลขทะเบียนผู้เสียภาษี
                            if not match_details.get('tax_id_match', True):
                                ws.cell(row=row, column=6).fill = red_fill
                            
                            # Column 7: สาขา
                            if not match_details.get('branch_match', True):
                                ws.cell(row=row, column=7).fill = red_fill
                            
                            # Column 8: รายการ 7%
                            if not match_details.get('amount_before_vat_match', True):
                                ws.cell(row=row, column=8).fill = red_fill
                            
                            # Column 9: ภาษี 7%
                            if not match_details.get('vat_amount_match', True):
                                ws.cell(row=row, column=9).fill = red_fill
                            
                            # Column 10: มูลค่ารวม
                            if not match_details.get('total_amount_match', True):
                                ws.cell(row=row, column=10).fill = red_fill
                        else:
                            # ถ้าไม่มี match_details ให้ไฮไลท์ทุกช่องเป็นสีแดง (กรณี no_match)
                            logger.warning(f"⚠️ ไม่พบ match_details สำหรับรายการที่ {item_no}")
                            for col in range(3, 13):  # ไฮไลท์เฉพาะคอลัมน์ข้อมูล (ไม่รวมลำดับ, เลขที่อ้างอิง, ชื่อไฟล์ OCR, สถานะ, หมายเหตุ)
                                ws.cell(row=row, column=col).fill = red_fill
                    
                    # ใส่ border และ font ทุกเซลล์
                    for col in range(1, 16):
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        if not cell.font:  # ถ้ายังไม่ได้ตั้งค่า font ให้ตั้งค่า
                            cell.font = data_font
                
                row += 1
            
            # ข้อมูลจาก OCR (แสดงต่อด้านล่างด้วยลำดับเดียวกัน) - แสดงเสมอแม้ไม่มีข้อมูล OCR
            # ใช้ data_font เดียวกันกับแถวภาษีซื้อเพื่อให้ขนาดเท่ากัน
            
            # ตั้งค่าข้อมูลและ font พร้อมกัน
            ws.cell(row=row, column=1, value=item_no).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=1).font = data_font
            
            if ocr_data:
                ws.cell(row=row, column=2, value=ocr_data.get('reference_number', ocr_data.get('document_no', '-')))
                ws.cell(row=row, column=2).font = data_font
                ws.cell(row=row, column=3, value=ocr_data.get('company_name', '-'))
                ws.cell(row=row, column=3).font = data_font
                ws.cell(row=row, column=4, value=ocr_data.get('document_no', '-'))
                ws.cell(row=row, column=4).font = data_font
                ws.cell(row=row, column=5, value=ocr_data.get('date', '-')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=5).font = data_font
                ws.cell(row=row, column=6, value=ocr_data.get('tax_id', '-'))
                ws.cell(row=row, column=6).font = data_font
                ws.cell(row=row, column=7, value=ocr_data.get('branch', '-')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=7).font = data_font
                ws.cell(row=row, column=8, value=ocr_data.get('amount_before_vat', 0)).number_format = '#,##0.00'
                ws.cell(row=row, column=8).font = data_font
                ws.cell(row=row, column=9, value=ocr_data.get('vat_amount', 0)).number_format = '#,##0.00'
                ws.cell(row=row, column=9).font = data_font
                ws.cell(row=row, column=10, value=ocr_data.get('total_amount', 0)).number_format = '#,##0.00'
                ws.cell(row=row, column=10).font = data_font
                # รายการยกเว้นภาษี (คอลัมน์ 11) - สำหรับ OCR ไม่มีข้อมูลนี้
                ws.cell(row=row, column=11, value='-').font = data_font
                # รายการภาษี 0% (คอลัมน์ 12) - สำหรับ OCR ไม่มีข้อมูลนี้
                ws.cell(row=row, column=12, value='-').font = data_font
                # ชื่อไฟล์ OCR (คอลัมน์ 13)
                ocr_filename = ocr_data.get('filename') or ocr_data.get('old_filename') or '-'
                ws.cell(row=row, column=13, value=ocr_filename).font = data_font
            else:
                # ถ้าไม่มีข้อมูล OCR ให้แสดง "-" ในทุกช่อง
                for col in range(2, 14):
                    ws.cell(row=row, column=col, value='-').font = data_font
                    if col == 5 or col == 7:  # วันที่และสาขา
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
                    elif col >= 8 and col <= 12:  # ตัวเลข (รายการ 7%, ภาษี 7%, มูลค่ารวม, รายการยกเว้นภาษี, รายการภาษี 0%)
                        ws.cell(row=row, column=col).number_format = '#,##0.00'
            
            ws.cell(row=row, column=14, value='OCR Data').alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=14).fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
            ws.cell(row=row, column=14).font = data_font
            ws.cell(row=row, column=15, value='').font = data_font
            
            # ตรวจสอบว่าเอกสารใช้ไม่ได้หรือไม่ - ถ้าใช่ให้ไฮไลท์สีแดงทั้งแถบ (ยกเว้นช่องหมายเหตุ)
            if is_invalid:
                for col in range(1, 16):
                    if col != 15:  # ไม่ใส่สีในช่องหมายเหตุ (column 15)
                        cell = ws.cell(row=row, column=col)
                        cell.fill = red_fill
                        cell.border = border
                        cell.font = data_font
                    else:
                        # ช่องหมายเหตุไม่ใส่สี
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        cell.font = data_font
            # ถ้า full_match และไม่ใช่เอกสารใช้ไม่ได้ → ไฮไลท์เขียวทั้งแถว OCR ด้วย (ยกเว้นช่องหมายเหตุ)
            elif match_status == 'full_match':
                for col in range(1, 16):
                    if col != 15:  # ไม่ใส่สีในช่องหมายเหตุ (column 15)
                        cell = ws.cell(row=row, column=col)
                        cell.fill = green_fill
                        cell.border = border
                        cell.font = data_font
                    else:
                        # ช่องหมายเหตุไม่ใส่สี
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        cell.font = data_font
            else:
                # ขีดสีแดงที่ช่องข้อมูลที่ไม่ตรงกัน (เฉพาะช่องที่ไม่ตรงกัน) สำหรับแถว OCR
                if purchase_data and ocr_data:
                    if match_details:
                        # Column 3: ชื่อบริษัท
                        if not match_details.get('company_name_match', True):
                            cell = ws.cell(row=row, column=3)
                            cell.fill = red_fill
                            cell.font = data_font
                        
                        # Column 4: เลขที่ใบกำกับ
                        if not match_details.get('document_no_match', True):
                            cell = ws.cell(row=row, column=4)
                            cell.fill = red_fill
                            cell.font = data_font
                        
                        # Column 5: วันที่
                        if not match_details.get('date_match', True):
                            cell = ws.cell(row=row, column=5)
                            cell.fill = red_fill
                            cell.font = data_font
                        
                        # Column 6: เลขทะเบียนผู้เสียภาษี
                        if not match_details.get('tax_id_match', True):
                            cell = ws.cell(row=row, column=6)
                            cell.fill = red_fill
                            cell.font = data_font
                        
                        # Column 7: สาขา
                        if not match_details.get('branch_match', True):
                            cell = ws.cell(row=row, column=7)
                            cell.fill = red_fill
                            cell.font = data_font
                        
                        # Column 8: รายการ 7%
                        if not match_details.get('amount_before_vat_match', True):
                            cell = ws.cell(row=row, column=8)
                            cell.fill = red_fill
                            cell.font = data_font
                        
                        # Column 9: ภาษี 7%
                        if not match_details.get('vat_amount_match', True):
                            cell = ws.cell(row=row, column=9)
                            cell.fill = red_fill
                            cell.font = data_font
                        
                        # Column 10: มูลค่ารวม
                        if not match_details.get('total_amount_match', True):
                            cell = ws.cell(row=row, column=10)
                            cell.fill = red_fill
                            cell.font = data_font
                    else:
                        # ถ้าไม่มี match_details ให้ไฮไลท์ทุกช่องเป็นสีแดง (กรณี no_match)
                        for col in range(3, 11):  # ไฮไลท์เฉพาะคอลัมน์ข้อมูล
                            cell = ws.cell(row=row, column=col)
                            cell.fill = red_fill
                            cell.font = data_font
                
                    # ใส่ border และ font ทุกเซลล์
                    for col in range(1, 14):
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        if not cell.font:  # ถ้ายังไม่ได้ตั้งค่า font ให้ตั้งค่า
                            cell.font = data_font
            
            row += 1
            
            item_no += 1
        
        # ส่วนสรุป
        summary_row = row + 1
        ws.merge_cells(f'A{summary_row}:G{summary_row}')
        ws.cell(row=summary_row, column=1, value='รวมทั้งหมด').font = Font(name='TH Sarabun New', size=12, bold=True)
        ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal='right')
        ws.cell(row=summary_row, column=1).fill = subheader_fill
        ws.cell(row=summary_row, column=1).font = subheader_font
        
        # คำนวณยอดรวมจากข้อมูลจริง
        total_tax_7 = sum(comp.get('purchase_data', {}).get('tax_7', 0) for comp in comparisons if comp.get('purchase_data'))
        total_vat = sum(comp.get('purchase_data', {}).get('vat', 0) for comp in comparisons if comp.get('purchase_data'))
        total_amount = sum(comp.get('purchase_data', {}).get('total_with_vat', 0) for comp in comparisons if comp.get('purchase_data'))
        
        ws.cell(row=summary_row, column=8, value=total_tax_7).number_format = '#,##0.00'
        ws.cell(row=summary_row, column=8).font = Font(name='TH Sarabun New', size=12, bold=True)
        ws.cell(row=summary_row, column=8).fill = subheader_fill
        ws.cell(row=summary_row, column=8).font = subheader_font
        
        ws.cell(row=summary_row, column=9, value=total_vat).number_format = '#,##0.00'
        ws.cell(row=summary_row, column=9).font = Font(name='TH Sarabun New', size=12, bold=True)
        ws.cell(row=summary_row, column=9).fill = subheader_fill
        ws.cell(row=summary_row, column=9).font = subheader_font
        
        ws.cell(row=summary_row, column=10, value=total_amount).number_format = '#,##0.00'
        ws.cell(row=summary_row, column=10).font = Font(name='TH Sarabun New', size=12, bold=True)
        ws.cell(row=summary_row, column=10).fill = subheader_fill
        ws.cell(row=summary_row, column=10).font = subheader_font
        
        # ใส่ border
        for col in range(1, 13):
            ws.cell(row=summary_row, column=col).border = border
        
        # สร้างชื่อไฟล์
        filename = f"รายงานตรวจภาษี_{company}_{tax_month}.xlsx"
        
        # บันทึกไฟล์ Excel
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # ====== บันทึกไฟล์ลงโฟลเดอร์ VAT ======
        saved_file_path = None
        if vat_folder_path:
            try:
                vat_folder = Path(vat_folder_path)
                if vat_folder.exists() and vat_folder.is_dir():
                    file_path = vat_folder / filename
                    with open(file_path, 'wb') as f:
                        f.write(excel_file.getvalue())
                    saved_file_path = str(file_path)
                    logger.info(f"💾 บันทึกไฟล์ Excel ลงโฟลเดอร์ VAT: {file_path}")
                else:
                    logger.warning(f"⚠️ โฟลเดอร์ VAT ไม่มีอยู่: {vat_folder_path}")
                    return jsonify({
                        'success': False,
                        'error': f'ไม่พบโฟลเดอร์ VAT: {vat_folder_path}'
                    }), 400
            except Exception as e:
                logger.error(f"❌ ไม่สามารถบันทึกไฟล์ลงโฟลเดอร์ VAT: {e}")
                return jsonify({
                    'success': False,
                    'error': f'ไม่สามารถบันทึกไฟล์ได้: {str(e)}'
                }), 500
        else:
            logger.warning(f"⚠️ ไม่มี path ของโฟลเดอร์ VAT")
            return jsonify({
                'success': False,
                'error': 'ไม่พบ path ของโฟลเดอร์ VAT'
            }), 400
        
        # ส่งข้อมูลกลับไปบอกว่าบันทึกสำเร็จ (ไม่ส่งไฟล์ให้ดาวน์โหลด)
        return jsonify({
            'success': True,
            'message': 'บันทึกรายงานสำเร็จ',
            'filePath': saved_file_path,
            'fileName': filename,
            'vatFolderPath': vat_folder_path
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการส่งออก Excel: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/save-state', methods=['POST'])
def save_auditcheck_state():
    """บันทึกสถานะการตรวจสอบลงไฟล์"""
    try:
        import json
        from pathlib import Path
        from datetime import datetime
        
        data = request.json
        if not data:
            return jsonify({
                'success': False,
                'error': 'ไม่มีข้อมูลส่งมา'
            }), 400
        
        company = data.get('company', '')
        tax_month = data.get('taxMonth', '')
        
        if not company or not tax_month:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุบริษัทและเดือนภาษี'
            }), 400
        
        # สร้างโฟลเดอร์ cache สำหรับ auditcheck states
        cache_dir = Path('cache')
        cache_dir.mkdir(exist_ok=True)
        
        # สร้างชื่อไฟล์ที่ปลอดภัย
        safe_company = company.replace('/', '_').replace('\\', '_').replace(':', '_')
        safe_tax_month = tax_month.replace('/', '_').replace('\\', '_').replace(':', '_')
        state_filename = f"auditcheck_state_{safe_company}_{safe_tax_month}.json"
        state_file_path = cache_dir / state_filename
        
        # อัพเดท timestamp
        data['last_saved'] = datetime.now().isoformat()
        
        # บันทึกไฟล์
        try:
            with open(state_file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"✅ บันทึกสถานะการตรวจสอบ: {state_filename}")
            
            return jsonify({
                'success': True,
                'message': 'บันทึกสถานะสำเร็จ',
                'filename': state_filename
            }), 200
        except Exception as e:
            logger.error(f"❌ ไม่สามารถบันทึกไฟล์สถานะ: {e}")
            return jsonify({
                'success': False,
                'error': f'ไม่สามารถบันทึกไฟล์ได้: {str(e)}'
            }), 500
            
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการบันทึกสถานะ: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/load-state', methods=['GET'])
def load_auditcheck_state():
    """โหลดสถานะการตรวจสอบจากไฟล์"""
    try:
        import json
        from pathlib import Path
        
        company = request.args.get('company', '')
        tax_month = request.args.get('taxMonth', '')
        
        if not company or not tax_month:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุบริษัทและเดือนภาษี'
            }), 400
        
        # สร้างชื่อไฟล์ที่ปลอดภัย
        safe_company = company.replace('/', '_').replace('\\', '_').replace(':', '_')
        safe_tax_month = tax_month.replace('/', '_').replace('\\', '_').replace(':', '_')
        state_filename = f"auditcheck_state_{safe_company}_{safe_tax_month}.json"
        state_file_path = Path('cache') / state_filename
        
        # ตรวจสอบว่าไฟล์มีอยู่หรือไม่
        if not state_file_path.exists():
            return jsonify({
                'success': False,
                'error': 'ไม่พบข้อมูลสถานะที่บันทึกไว้'
            }), 404
        
        # อ่านไฟล์
        try:
            with open(state_file_path, 'r', encoding='utf-8') as f:
                state_data = json.load(f)
            
            logger.info(f"✅ โหลดสถานะการตรวจสอบ: {state_filename}")
            
            return jsonify({
                'success': True,
                'state': state_data,
                'filename': state_filename
            }), 200
        except json.JSONDecodeError as e:
            logger.error(f"❌ ไม่สามารถอ่านไฟล์สถานะ (JSON error): {e}")
            return jsonify({
                'success': False,
                'error': f'ไฟล์สถานะเสียหาย: {str(e)}'
            }), 500
        except Exception as e:
            logger.error(f"❌ ไม่สามารถอ่านไฟล์สถานะ: {e}")
            return jsonify({
                'success': False,
                'error': f'ไม่สามารถอ่านไฟล์ได้: {str(e)}'
            }), 500
            
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการโหลดสถานะ: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/auditcheck/mark-documents-invalid', methods=['POST'])
def mark_documents_invalid():
    """ย้ายไฟล์ไปยังโฟลเดอร์เอกสารใช้งานไม่ได้และทำเครื่องหมาย"""
    try:
        import shutil
        from pathlib import Path
        
        data = request.json
        tax_month = data.get('taxMonth', '')
        company = data.get('company', '')
        vat_folder_path = data.get('vatFolderPath', '')
        documents = data.get('documents', [])
        
        if not vat_folder_path:
            return jsonify({
                'success': False,
                'error': 'ไม่พบ path ของโฟลเดอร์ VAT'
            }), 400
        
        vat_folder = Path(vat_folder_path)
        if not vat_folder.exists():
            return jsonify({
                'success': False,
                'error': f'ไม่พบโฟลเดอร์ VAT: {vat_folder_path}'
            }), 400
        
        # สร้างโฟลเดอร์ "เอกสารใช้งานไม่ได้" ในโฟลเดอร์ VAT
        invalid_folder = vat_folder / 'เอกสารใช้งานไม่ได้'
        invalid_folder.mkdir(exist_ok=True)
        
        moved_count = 0
        moved_files = []
        
        # ย้ายไฟล์ PDF ที่เกี่ยวข้อง
        for doc in documents:
            ocr_filename = doc.get('ocrFilename', '')
            reference_no = doc.get('referenceNo', '')
            
            # ค้นหาไฟล์ PDF ที่เกี่ยวข้อง
            if ocr_filename:
                # ลองหาไฟล์จากชื่อไฟล์ OCR
                pdf_file = vat_folder / ocr_filename
                if pdf_file.exists() and pdf_file.is_file():
                    try:
                        dest_file = invalid_folder / pdf_file.name
                        shutil.move(str(pdf_file), str(dest_file))
                        moved_files.append(pdf_file.name)
                        moved_count += 1
                        logger.info(f"✅ ย้ายไฟล์: {pdf_file.name} → {dest_file}")
                    except Exception as e:
                        logger.error(f"❌ ไม่สามารถย้ายไฟล์ {pdf_file.name}: {e}")
            
            # ลองหาไฟล์จาก reference number
            if reference_no:
                for pdf_file in vat_folder.glob(f'*{reference_no}*'):
                    if pdf_file.is_file() and pdf_file.name not in moved_files:
                        try:
                            dest_file = invalid_folder / pdf_file.name
                            shutil.move(str(pdf_file), str(dest_file))
                            moved_files.append(pdf_file.name)
                            moved_count += 1
                            logger.info(f"✅ ย้ายไฟล์: {pdf_file.name} → {dest_file}")
                        except Exception as e:
                            logger.error(f"❌ ไม่สามารถย้ายไฟล์ {pdf_file.name}: {e}")
        
        return jsonify({
            'success': True,
            'message': f'ย้ายไฟล์สำเร็จ {moved_count} ไฟล์',
            'movedCount': moved_count,
            'movedFiles': moved_files,
            'invalidFolderPath': str(invalid_folder)
        }), 200
    
    except Exception as e:
        logger.error(f"❌ เกิดข้อผิดพลาดในการย้ายไฟล์: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


if __name__ == '__main__':
    # สร้างโฟลเดอร์ templates และ static ถ้ายังไม่มี
    Path('templates').mkdir(exist_ok=True)
    Path('static').mkdir(exist_ok=True)
    
    # ตรวจสอบ routes ที่สำคัญ
    print("🔍 ตรวจสอบ routes...")
    routes = []
    for rule in app.url_map.iter_rules():
        methods = sorted(rule.methods - {'HEAD', 'OPTIONS'})
        routes.append({
            'path': str(rule),
            'methods': methods,
            'endpoint': rule.endpoint
        })
    
    route_paths = [r['path'] for r in routes]
    important_routes = ['/api/stats', '/api/admin/line-notify/status', '/api/admin/line-notify/toggle']
    for route in important_routes:
        if route in route_paths:
            route_info = next((r for r in routes if r['path'] == route), None)
            methods = ', '.join(route_info['methods']) if route_info else ''
            endpoint = route_info['endpoint'] if route_info else ''
            print(f"  ✅ {route} [{methods}] -> {endpoint}")
        else:
            print(f"  ❌ {route} - ไม่พบ!")
            # แสดง routes ที่ใกล้เคียง
            similar = [r['path'] for r in routes if 'stats' in r['path'].lower() or 'line' in r['path'].lower()]
            if similar:
                print(f"     Routes ที่ใกล้เคียง: {', '.join(similar[:5])}")
    
    # แสดง API routes ทั้งหมด
    print(f"\n📋 API Routes ทั้งหมด:")
    api_routes = [r for r in routes if r['path'].startswith('/api/')]
    for route in sorted(api_routes, key=lambda x: x['path']):
        methods = ', '.join(route['methods'])
        print(f"  {route['path']:50} [{methods}] -> {route['endpoint']}")
    
    print(f"\n📊 จำนวน routes ทั้งหมด: {len(routes)}")
    print(f"📊 จำนวน API routes: {len(api_routes)}")
    print("🚀 เริ่มต้น BotV3 Web Application...")
    print("🌐 Web App จะรันที่: http://localhost:5000")
    print("📖 หน้าแรก: http://localhost:5000/")
    print("📄 หน้าประมวลผล PDF: http://localhost:5000/pdf")
    print("📧 หน้าส่งอีเมลล์: http://localhost:5000/email")
    print("📋 หน้าคัดแยกเอกสาร: http://localhost:5000/document-sorting")
    print("🛑 กด Ctrl+C เพื่อหยุดการทำงาน")
    print("=" * 50)
    
    app.run(host='0.0.0.0', port=5000, debug=True, use_reloader=False)

